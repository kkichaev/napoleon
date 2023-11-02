# -*- coding: cp1251 -*-

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import sys

reload(sys)
sys.setdefaultencoding("cp1251")

class DailyPlan:
    __slots__ = ['folder', 'plan', 'fact']
    
    def __init__(self, item):
        self.fact = 0
        self.plan = item.weight
        self.folder = item.id

class OrgPlanData:
    __slots__ = ['items', 'id', 'created']
    
    def __init__(self, doc):
        self.items = list()
        self.id = doc.id
        self.addPlan(doc)

    def addPlan(self, doc):
        self.created = doc.created
        self.items = list()
        for di in doc.items:
            item = DailyPlan(di)
            self.items.append(item)
            
    def addFact(self, doc, price):
        for di in doc.items:
            if not di.id in price: continue
            
            prc = price[di.id]
            for pi in self.items:
                if pi.folder == prc.fid:
                    pi.fact += di.qty * prc.weight
                    break;
        
        
        
class AgentPlanData:
    __slots__ = ['items', 'agent', 'orgs']
    
    def __init__(self, agent, server):
        self.agent = agent
        self.items = list()
        server.ChangeUser("'" + agent.id + "'")
        self.orgs = server.Get('Org','', 'id')
        server.RestoreUser()
        print len(self.orgs)
        
    def addPlan(self, doc):
        for oi in self.items:
            if oi.id == doc.id:
                if oi.created < doc.created:
                    oi.addPlan(doc)
                return
        oi = OrgPlanData(doc)
        self.items.append(oi)
        
    def addFact(self, doc, price):
        for oi in self.items:
            if oi.id == doc.id:
                oi.addFact(doc, price)
                break
    
class DivisionPlanData:
    __slots__ = ['items', 'division']
    
    def __init__(self, division):
        self.items = list()
        self.division = division
        
    def addPlan(self, agent, doc, server):
        for ai in self.items:
            if ai.agent.id == agent.id:
                ai.addPlan(doc)
                return
        ai = AgentPlanData(agent, server)
        self.items.append(ai)
        ai.addPlan(doc)


def findDivision(divisions, data, agent):
    for d in divisions:
        for da in d.agents:
            if da.id == agent.id:
                for di in data:
                    if di.division.id == d.id:
                        return di
                    
                di = DivisionPlanData(d)
                data.append(di)
                return di
            
    return None

def findPlan(data, userid):
    for di in data:
        for ai in di.items:
            if ai.agent.id == userid:
                return ai
    return None
        

def loadData(server):
    params = server.Params[0]

    divisions = server.Get('Division', '')
    agents = server.Get('Agents','','id')
    folders = server.Get('ManagerFolder','','id')
    price = server.Get('ManagerPrice','','id')
    
    uids = '"userid" in('
    for ai in params.items:
        uids += "'" + ai.id + "',"
        
    uids = uids[:-1] + ")"
    
    data = list()
    
    where = uids + ' and "date">=ToDate("{0}") and "date"<=ToDate("{0} 23:59:59")' . format(params.date.strftime("%d/%m/%Y"))
    plans = server.Get('DailyPlan', where)
    
    for doc in plans:
        if not doc.userid in agents: continue
        
        agent = agents[doc.userid]
        dd = findDivision(divisions, data, agent)
        if dd == None: continue
        
        dd.addPlan(agent, doc, server)
        
    where = uids + ' and "created" >= ToDate("{0}") and "created" <= ToDate("{0} 23:59:59")' . format(params.date.strftime("%d/%m/%Y"))
    docs = server.Get('Order', where)
    for doc in docs:
        pd = findPlan(data, doc.userid)
        if pd == None: continue
        pd.addFact(doc, price)
        
    return data, folders

def printOut(data, folders):
    xlb = XLBuilder()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    cr = 1
    cc = 1
    
    xlb.makeHead(sheet, cr, ['Отдел', 'Фамилия ТП','Клиент/Группа','План','Факт','Отклонение'], True, cc)
    cr += 1
    
    for divPlan in data:
        dp = 0
        df = 0
        for agtPlan in divPlan.items:
            ap = 0
            af = 0
            for orgPlan in agtPlan.items:
                op = 0
                of = 0
                for iPlan in orgPlan.items:
                    f = folders[iPlan.folder].name if iPlan.folder in folders else iPlan.folder
                    values = ['','', f, iPlan.plan, iPlan.fact, "" if iPlan.plan == 0 else str(float(iPlan.fact) / iPlan.plan * 100) + '%' ]
                    xlb.makeCells(sheet, cr, values, cc)
                    sheet.row_dimensions[cr+1].outline_level = 3 
                    cr += 1
                    op += iPlan.plan
                    of += iPlan.fact
                 
                org = agtPlan.orgs[orgPlan.id].name if orgPlan.id in agtPlan.orgs else orgPlan.id 
                values = ['','', org, op, of, "" if op == 0 else str(float(of) / op * 100) + '%']
                xlb.makeCells(sheet, cr, values, cc)
                sheet.row_dimensions[cr+1].outline_level = 2 
                cr += 1
                ap += op
                af += of
            values = ['', agtPlan.agent.name, '', ap, af, "" if ap == 0 else str(float(af)/ ap * 100) + '%']
            xlb.makeCells(sheet, cr, values, cc)
            sheet.row_dimensions[cr+1].outline_level = 1 
            cr += 1
            dp += ap
            df += af
        values = [divPlan.division.name, '', '', dp, df, "" if dp == 0 else str(float(df)/ dp * 100) + '%']
        xlb.makeCells(sheet, cr, values, cc)
        cr += 1
    

    ccl = cc + 1
    wdh = [30,30,30,12,12,12]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(ccl)].width = w
        ccl += 1
    
    
    return wb
    
def run(server):
   data, folders = loadData(server)
   wb = printOut(data, folders)
   
   XLBuilder().workbookToObject(wb, "d_rpt.xlsx", server) 
   
