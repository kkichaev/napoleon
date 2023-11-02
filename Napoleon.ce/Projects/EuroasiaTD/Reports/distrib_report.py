# -*- coding: cp1251 -*-
from importlib import reload
import sys
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

class DocData:
    def __init__(self):  
      pass
    
    def values(self, row):
      result = [self.agent, "{0} {1}".format(self.orgname, self.orgaddress), self.created.strftime('%d.%m.%Y %H:%M:%S')]
      result.extend(self.pricevalues)
      sz = len(self.pricevalues)
      sc = 3
      rn = get_column_letter(sz + sc)
      f = "=SUM(D{0}:{2}{0})".format(row + 1, sc, rn) if sz > 0 else ' '
      result.append(f)
      result.append(self.remark)
      
      return result
    
    def values2(self, row):
      result = [self.agent, "{0} {1}".format(self.orgname, self.orgaddress), self.created.strftime('%d.%m.%Y %H:%M:%S')]
      return result    
                    
class ReportData:
    def __init__(self):
        self.data = list()
        self.pcidx = dict()
        self.agents = dict()
        self.orgs = dict()
        self.start = ''
        self.finish = ''
        self.division = ''
        
    def add(self, doc):
        data = DocData()
        data.created = doc.created
        
        on = '';
        oa = '';
        oc = ''
        
        mtx = []
        
        if doc.id in self.orgs:
          o = self.orgs[doc.id]
          on = o.name
          oa = o.address
          
          for m in o.matrix:
            mtx.append(m.id)
          
          
        data.orgname = on
        data.orgaddress = oa
        data.partval = 0
        data.orgcateg = oc
        data.agent = self.agents[doc.userid] if doc.userid in self.agents else doc.userid
        data.remark = doc.remark
        
        data.pricevalues = []
        data.priceremarks = []
        data.pricematrix = []
        
        for i in range(0, len(self.price)):
          data.pricevalues.append(' ')
          data.priceremarks.append(' ')
          data.pricematrix.append(1 if self.price[i].id in mtx else 0)
        
        for i in doc.items:
          if i.id in self.pcidx:
            data.pricevalues[self.pcidx[i.id]] = i.qty
            data.priceremarks[self.pcidx[i.id]] = i.remark
        
        self.data.append(data)
        
    def setPrice(self, pi):
      self.price = []
      
      for p in pi.values():
        self.price.append(p)
      
      self.price = sorted(self.price, key=lambda p: p.name)
      
      for i in range(0, len(self.price)):
        p = self.price[i]
        self.pcidx[p.id] = i
      
    def priceCells(self):
      result = []
      
      for p in self.price:
        result.append(p.name)
      
      return result

def loadData(params, server):
    orgs = server.Get('CommonOrgs', '', 'id')
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    price = server.Get('ManagerPrice',  "", 'id')
    folders = server.Get('ManagerFolder', '')
    fids = []
    plv = -1
    
    for f in folders:
      if len(fids) == 0 and f.id == params.folderid:
        fids.append(f.id)
        plv = f.level
      elif len(fids) > 0:
        if f.level > plv:
          fids.append(f.id)
        else:
          break
        
    fprice = dict()
    
    for p in price.values():
      if p.fid in fids:
        fprice[p.id] = p
        
    data = ReportData();
    data.start = params.start
    data.finish = params.finish
    data.division = params.division
    data.setPrice(fprice)
    
    for id in params.users.split(','):
      server.ChangeUser("'" + id + "'")
      orgs.update(server.Get("Org", "", "id"))
      data.orgs = orgs
      a = server.CurrentUser();
      data.agents[a.id] = a.name
      server.RestoreUser()
        
      q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}") and "userid" = "{2}"'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"),
        id)
      
      docs = server.Get('OrgDistrib', q)
      
      if docs != None:
        for d in docs:
          data.add(d)
            
    return data

class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    #cell.style.alignment.text_rotation = 90
    return column
  
  def makeHead(self, sheet, row, titles, wrap_text = False, startColumn = 0):
    XLBuilder.makeHead(self, sheet, row, titles, wrap_text, startColumn)
    
def rangeBorders(range, border):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = border
            cell.style.borders.right.border_style = border
            cell.style.borders.top.border_style = border
            cell.style.borders.bottom.border_style = border

def printTop(sh, data):
    sh.cell(row=0, column=0).value = 'Период: {0} - {1}'.format(data.start.strftime('%d.%m.%Y'), data.finish.strftime('%d.%m.%Y'))
    sh.cell(row=1, column=0).value = 'Подразделение: {0}'.format(data.division)

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Форма1"
    printTop(sheet, data)
    xlb = XLBuilderEx()
    START_ROW = 2
    
    row = START_ROW
    
    head = ['Мерчендайзер', 'Наименование магазина', 'Дата / Время отправки отчета']
    st = len(head)         
    pc = data.priceCells()
    
    if len(pc) > 0:
      head.extend(pc)
      
    head.extend(['Итого SKU', 'Примечание'])  
      
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      xlb.makeCells(sheet, row, item.values(row))
      row += 1        
    
    sheet.cell(row=row, column=2).value = "Итого"
    rangeBorders(sheet.range("A{1}:{0}{1}".format(get_column_letter(len(head)), row+1)), Border.BORDER_THIN)
    
    cc = st
    for i in range(0,len(pc) + 1):
      c = sheet.cell(row=row, column=cc)
      cn = get_column_letter(cc+1)
      dc = len(data.data)
      dsz = dc + 3
      c.value = "=SUM({0}4:{0}{1})".format(cn, dsz) if dc > 0 else ' ' 
      cc += 1
    
    cc = 1
    for w in [30,30,30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    #-------------------------------------------------------------------------------------------------------------
    
    sheet = wb.create_sheet()
    sheet.title = "Форма2"
    printTop(sheet, data)
    row = START_ROW
    head = ['Мерчендайзер', 'Наименование магазина', 'Дата / Время отправки отчета', 'Примечание']
    xlb.makeHead(sheet, row, head, True)
    
    cc = 1
    for w in [30,30,30, 30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    row += 1
    for item in data.data:
      d = item.values2(row)
      d.append(' ')
      xlb.makeCells(sheet, row, d)
      sheet.cell(row=row, column=1).style.font.bold = True
      sheet.cell(row=row, column=2).style.font.bold = True
        
      row += 1  
      
      prices = data.priceCells()
      for idx in range(0, len(prices)):
        d[len(d)-3] = prices[idx]
        d[len(d)-2] = item.pricevalues[idx]
        d[len(d)-1] = item.priceremarks[idx]
        xlb.makeCells(sheet, row, d)
        
        if item.pricematrix[idx] == 0:
          xlb.setBackColor(sheet.cell(row=row, column=2),"c0c0c0")
          sheet.cell(row=row, column=2).value = ' '
        
        row += 1
    
    return wb
    
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info('end')
    