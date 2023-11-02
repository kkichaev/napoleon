# -*- coding: cp1251 -*-
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3
        
class AgentData:
  __slots__ = ['data', 'name']
  
  def __init__(self):
    self.data = []
    self.name = ""
  
  def getData(self, row):
    return ['','','Итого',"=SUM(D{0}:D{1})".format(row - len(self.data)+1, row) if len(self.data) > 0 else 0]
    
class ReportData(AgentData):
  pass

class RouteItem:
  __slots__ = ['id', 'date', 'org']
  
  def __init__(self):
    self.id = ""
    self.date = None
    self.org = ""
  
  def __repr__(self):
    return "{0}, {1}".format(self.id, self.date)
    
class RoutePair:
  __slots__ = ['first', 'second', 'distance', 'date']
  
  def __init__(self):
    self.first = RouteItem()
    self.second = RouteItem()
    self.distance = 0
    self.date = None
    
  def __repr__(self):
    return "[{0} - {1} : {2}]".format(self.first, self.second, self.distance)
  
  def getData(self, row):
    return [self.date.strftime("%d.%m.%Y") if self.date != None else "X", self.first.org, self.second.org, self.distance / 1000]

def filterGPS(st, fn, gps):
  ret = []
    
  for g in gps:
    tm = g.date.time()
    
    if tm >= st and tm <= fn:
      ret.append(g)
    
  return ret

def calcDistance(gps):
  ret = 0
  lastpos = None
        
  for g in gps:
    if lastpos == None:
        lastpos = g
        continue

    ret += coordutils.distance(lastpos.latitude, lastpos.longitude, g.latitude, g.longitude)
    lastpos = g

  return ret
  
def loadData(params, server):
  agents = server.Get("Agents", "", "id")
  
  data = ReportData()  
  orgData = {}
  
  st = params.start.time()
  ft = params.finish.time()
  
  for aid in params.userid.split(','):
    item = AgentData()
    item.aid = aid
    data.data.append(item)
    
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    item.name = server.CurrentUser().name
    server.RestoreUser()
    
    date = params.start
    
    while date <= params.finish:
      where = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}") and "userid"=\'{2}\''.format(
        date.strftime("%d/%m/%Y %H:%M:%S"), date.strftime("%d/%m/%Y 23:59:29"), aid) 
      
      gps = server.Get("GPSPos", where)
      
      if ft.hour * ft.minute * ft.second * ft.microsecond == 0:
          ft = ft.replace(23,59,59,999999)
      
      gpsList = []
      
      for g in gps:
        if g.isGSM == 1 and params.gms == 0:
          continue
          
        gt = g.date.time()
        
        if st <= gt and ft >= gt:
            gpsList.append(g)
      
      where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
          aid,
          date.strftime("%d/%m/%Y 0:0:0"),
          date.strftime("%d/%m/%Y 23:59:59"))
      
      docNames = ["Order", "VisitInfo", "OrgRemnants", "Answer", "Incass", "TaskDone", "Sales"]

      allDocList = []
      
      for name in docNames:
        docList = server.Get(name, where)
        
        if docList == None:
          continue
          
        if docList != None:
          for d in docList:
            allDocList.append(d)
      
      allDocList = sorted(allDocList, cmp=lambda lhs, rhs: cmp(lhs.created, rhs.created))

      if len(allDocList) > 0:
        pairs = [RoutePair()]
        
        for d in allDocList:
          p = pairs[len(pairs)-1]
          
          if len(p.first.id) == 0:
            p.first.id = d.id
            p.first.date = d.created
            p.date = date
            
          if d.id == p.first.id:
            continue
          
          if len(p.second.id) > 0 and p.second.id != d.id:
            prev = p
            p = RoutePair()
            p.date = date
            p.first.id = prev.second.id
            p.first.date = prev.second.date
            pairs.append(p)
          
          p.second.id = d.id
          p.second.date = d.created
          
        for p in pairs:
          if p.first.date != None and p.second.date != None:
            p.distance = calcDistance(filterGPS(p.first.date.time(), p.second.date.time(), gps))
            
          item.data.append(p)
          p.first.org = "{0} ({1})".format(orgs[p.first.id].name, orgs[p.first.id].address) if p.first.id in orgs else "Контрагент с кодом <{0}>".format(p.first.id)
          p.second.org = "{0} ({1})".format(orgs[p.second.id].name, orgs[p.second.id].address) if p.second.id in orgs else "Контрагент с кодом <{0}>".format(p.second.id)
          
      date = date + timedelta(days=1)
  
  return data
    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 3:
      cell.style.number_format.format_code = NumberFormat.FORMAT_NUMBER
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    for ad in data.data:
      c = sheet.cell(row=0,column=0)
      c.value = "Отчет по пробегу"
      c.style.font.bold = True
      c.style.font.size = 18
      
      c = sheet.cell(row=1,column=0)
      c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
      
      c = sheet.cell(row=2, column=0)
      c.value = "ФИО Менеджера {0}".format(ad.name)
      
      xlb = XLBuilderEx()
      row = 3
      
      head = ['Дата', 'Адрес', 'Адрес', 'Расстояние, км']
      date = params.start
        
      xlb.makeHead(sheet, row, head, True)
      row += 1

      for item in ad.data:
        xlb.makeCells(sheet, row, item.getData(row))
        row += 1
      
      xlb.makeCells(sheet, row, ad.getData(row))
      
      cc = 1
      for w in [15,50,50,15]:
          sheet.column_dimensions[get_column_letter(cc)].width = w
          cc += 1
          
      sheet = wb.create_sheet()
    
    
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    