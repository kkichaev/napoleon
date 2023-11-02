# -*- coding: cp1251 -*-
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat

import datetime


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

class Item:
    __slots__ = ['org', 'created', 'sended', 'sum', 'number']
    
    def __init__(self, doc, orgs):
        org = orgs[doc.id].name if doc.id in orgs else "Контрагент с кодом <{0}>".format(doc.id)
        
        self.org = org
        self.created = doc.created
        self.sended = doc.sended
        self.sum = 0
        self.number = ""

class AgentPage:
    __slots__= ['id', 'name', 'items']
    
    def __init__(self, agent):
        self.id = agent[0]
        self.name = agent[1]
        self.items = list()
        
    def data(self, docs, orgs):
        for i in docs:
            if i.autoMode == 0 and len(i.items) == 0:
                item = Item(i, orgs)
                item.sum = i.sum
                self.items.append(item)
            else:
                for ii in i.items:
                    item = Item(i, orgs)
                    item.sum = ii.sum
                    item.number = ii.number
                    self.items.append(item)
    
    def sum(self):
        sum = 0
        
        for i in self.items:
            sum += i.sum
        
        return sum
        
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = dict()

def agentName(id, arr):
    name = id
    
    if id in arr:
        name = arr[id].name
        
    return id, name
    
def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    data = ReportData()
    
    for a in params.agents:
        userid = a.id
        
        if not userid in data.data:
            data.data[userid] = AgentPage(agentName(userid, agents))
            
        ap = data.data[userid]

        server.ChangeUser("'" + userid + "'")
        orgs = server.Get("Org", "", "id")
        server.RestoreUser()
        
        where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
            userid,
            params.start.strftime("%d/%m/%Y 0:0:0"),
            params.finish.strftime("%d/%m/%Y 23:59:59"))

        incass = server.Get("Incass", where)
        
        ap.data(incass, orgs)
            
    return data
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    
    sheet = None

    for ap in data.data.values():
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()

        sheet.title = ap.name[:31]
        
        xlb = XLBuilder()
        sheet.cell(row=0, column=0).value = ap.name
        
        row = 1
        
        head = ['Контрагент', 'Тип документа', 'Дата', 'Время создания', 'Дата передачи', 'Сумма', 'Номер накладной']
        
        xlb.makeHead(sheet, row, head, True)
        row += 1

        for item in ap.items:
            values = [item.org, "ПКО", item.created.strftime("%d.%m.%Y"), item.created.strftime("%H:%M"), item.sended.strftime("%d.%m.%Y"), item.sum, item.number]
            xlb.makeCells(sheet, row, values)
            row += 1
        
        row += 1
        sheet.cell(row=row, column=0).value = "Итоговая сумма: {0} руб.".format(ap.sum())
        
        row += 2
        sheet.cell(row=row, column=0).value = "Сдал ___________________"
        
        row += 2
        sheet.cell(row=row, column=0).value = "Принял _________________"
        
        cc = 1
        for w in [70,20,20,20,20,20,20]:
            sheet.column_dimensions[get_column_letter(cc)].width = w
            cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    