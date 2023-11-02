# -*- coding: cp1251 -*-
from util import UnknownOrg
from openpyxl import Workbook
from openpyxl.style import Border, Alignment, Color, Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
from datetime import datetime
from datetime import timedelta
import tempfile
import io
from util import scope
from __builtin__ import str

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

class Group:
    id = ""
    name = ""
    items = set()
    
    def __init__(self, src):
        self.items = set()
        self.id = src.id
        self.name = src.name
        for item in src.items:
            self.items.add(item.id)
    
    def haveItem(self, id):
        return id in self.items
    
class Groups:
    groups = set()
    
    def load(self, server):
        self.groups = set()
        
        userid = "'" + server.Params[0].agent+"'"
        server.ChangeUser(userid)
        items = server.Get("ItemGroups", '')
        server.RestoreUser()
        
        print "Items " + str(len(items))
        
#         where = '"userid"='+"'"+server.Params[0].agent+"'"
#         items = server.Get("ItemGroups", where)
#         where = '"userid" is null or "userid"=""'        
#         if items == None or len(items) == 0:
#             items = server.Get("ItemGroups",  where)
        
        if items == None:
            return
        
        for i in items:
            item = Group(i)
            self.groups.add(item)


class GroupVisitData:
    """ Наличие """
    rest = set()
    
    """ Заказано """
    orders = set()
    
    """ Стало """
    current = set()
    
    def __init__(self):
        self.rest = set()
        self.orders = set()
        self.current = set()

    def __repr__(self):
        return "rest " + str(len(self.rest)) + " ord " + str(len(self.orders)) + " cur " + str(len(self.current))

class VisitData:
    start = 0
    finish = 0
    
    """ group->GroupVisitData """
    groups = {}
    
    def __init__(self):
        self.groups = dict()
    
    def __repr__(self):
        str = ""
        for g, d in self.groups.iteritems() :
            str = str + " group " + g.name + " data " + d.__repr__();
        return str
    
    def haveOrders(self):        
        for k, v in self.groups.iteritems() :
            if len(v.orders) > 0: return True
        return False
    
    def getGroupData(self, group):
        groupVisitData = None
        if group in self.groups:
            groupVisitData = self.groups[group]
        else:
            groupVisitData = GroupVisitData()
            self.groups[group] = groupVisitData
            
        return groupVisitData
    
    def refreshOrders(self, item, groups):
        for group in groups:
            if not group.haveItem(item) : continue
            groupVisitData = self.getGroupData(group)                
            groupVisitData.orders.add(item)
            groupVisitData.current.add(item)
            

    def refreshRest(self, item, groups):
        for group in groups:
            if not group.haveItem(item) : continue
            
            groupVisitData = self.getGroupData(group)                
            groupVisitData.rest.add(item)
            groupVisitData.current.add(item)

class AgentDailyData:
    orgs = dict()
    agentOrgs = dict()
    
    rest = 0.0
    orders = 0.0
    current = 0.0
    prcCurrent = 0.0
    
    def __init__(self, server, id):
        self.orgs = dict()
        
        where = '"userid"=' + scope(id)
        self.agentOrgs = server.Get("Org", where, "id")
        
    def __repr__(self):
        str = ""
        for o,d in self.orgs.iteritems() :
            str = str + "org " + o.name + " data " + d.__repr__() + " ";
        return str
        
    def Visited(self):
        return len(self.orgs)
    
    def Orders(self):
        count = 0
        
        for k, v in self.orgs.iteritems() :            
            if v.haveOrders() :
                count = count + 1
        
        return count
    
    def TotalRest(self, group):
        rest = 0
        for od in self.orgs.itervalues() :
            if group in od.groups:
                rest = rest + len(od.groups[group].rest)
                
        return rest

    def TotalOrder(self, group):
        count = 0
        for od in self.orgs.itervalues() :
            if group in od.groups:
                count = count + len(od.groups[group].orders)
                
        return count

    def TotalCurrent(self, group):
        count = 0
        for od in self.orgs.itervalues() :
            if group in od.groups:
                count = count + len(od.groups[group].current)
                
        return count

    def getOrgData(self, org):
        orgData = None
        if org in self.orgs:
            orgData = self.orgs[org]
        else:
            orgData = VisitData()
            self.orgs[org] = orgData
            
        return orgData
    
    def getOrg(self, orgId):
        org = None
        if orgId not in self.agentOrgs:
            org = UnknownOrg.get(orgId, self.agentOrgs)
        else :
            org = self.agentOrgs[orgId]
        
        return org
    
    def loadOrder(self, groups, doc, mgrPrice, folders):
        org = self.getOrg(doc.id)
        orgData = self.getOrgData(org)
        for item in doc.items:
            if not (item.id in mgrPrice):
                continue
            pid = "," + mgrPrice[item.id].fid + ","
            if folders.find(pid) < 0 :
                continue
            orgData.refreshOrders(item.id, groups)
            
    def loadRest(self, groups, doc, mgrPrice, folders):
        org = self.getOrg(doc.id)
        orgData = self.getOrgData(org)
                
        for item in doc.items:
            if not (item.id in mgrPrice):
                continue
            pid = "," + mgrPrice[item.id].fid + ","
            if folders.find(pid) < 0 :
                continue
            orgData.refreshRest(item.id, groups)
            
    """ Вызывать после загрузки всех заявок и остатков """
    def loadWorkTime(self, worktime):
        org = self.getOrg(worktime.id)
        if org not in self.orgs: 
            return
        
        orgData = self.orgs[org]                
        if orgData.start == 0 or orgData.start > worktime.start:
            orgData.start = worktime.start
        if orgData.finish == 0 or orgData.finish < worktime.stop:
            orgData.finish = worktime.stop                    
 
                  
def setVal(cell, value, vrt = Alignment.VERTICAL_CENTER, hrz= Alignment.HORIZONTAL_CENTER, wrap=False):
    cell.value = value
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    makeBorder(cell)
            
def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN
                        

def makeReportHeader(sheet, groups, agent, date, agentData):
    dealsCount = agentData.Orders()
    visitsCount = agentData.Visited()
    
    sheet.cell(row=0, column=0).value = ""
    sheet.cell(row=0, column=1).value = "ДНЕВНОЙ ОТЧЕТ"
    sheet.cell(row=1, column=1).value = "ДАТА"
    sheet.cell(row=1, column=2).value = date
    sheet.cell(row=1, column=3).value = "АГЕНТ"
    sheet.cell(row=1, column=4).value = ""
    
    sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=10)
    sheet.cell(row=1, column=5).value = agent.name
        
    sheet.cell(row=2, column=1).value = "РАЙОН"
    sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    sheet.cell(row=2, column=3).value = "Из офиса"
    sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=10)
    sheet.merge_cells(start_row=2, start_column=11, end_row=2, end_column=13)
    
    sheet.cell(row=2, column=11).value = "Посещено"
    sheet.merge_cells(start_row=2, start_column=14, end_row=2, end_column=16)
    sheet.cell(row=2, column=14).value = visitsCount
     
    sheet.merge_cells(start_row=2, start_column=17, end_row=3, end_column=19)
    cell = sheet.cell(row=2, column=17)
    cell.value = "КПД посещений Сделок/Посещений"        
    cell.style.alignment.wrap_text = True
    sheet.merge_cells(start_row=2, start_column=20, end_row=3, end_column=21)
    prc = 0
    if visitsCount != 0 : prc = float(dealsCount) / visitsCount * 100
    setVal(sheet.cell(row=2, column=20), str(prc) + "%")
    
    sheet.cell(row=3, column=1).value = "МАРШРУТ"
    sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
    sheet.cell(row=3, column=3).value = "В офис"
    sheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=10)
    sheet.merge_cells(start_row=3, start_column=11, end_row=3, end_column=13)
    
    sheet.cell(row=3, column=11).value = "Сделок"
    sheet.merge_cells(start_row=3, start_column=14, end_row=3, end_column=16)
    sheet.cell(row=3, column=14).value = dealsCount
      
    r = sheet.range("B2:K2")
    rangeBorders(r)
    r = sheet.range("B3:V4")
    rangeBorders(r)
    
    sheet.merge_cells(start_row=5, start_column=0, end_row=7, end_column=0)
    setVal(sheet.cell(row=5, column=0), "№")
    
    sheet.merge_cells(start_row=5, start_column=1, end_row=7, end_column=1)
    setVal(sheet.cell(row=5, column=1), "Название магазина")
    
    sheet.merge_cells(start_row=5, start_column=2, end_row=7, end_column=2)
    setVal(sheet.cell(row=5, column=2), "Адрес")
    
    sheet.merge_cells(start_row=5, start_column=3, end_row=7, end_column=3)
    setVal(sheet.cell(row=5, column=3), "Приб")
    
    sheet.merge_cells(start_row=5, start_column=4, end_row=7, end_column=4)
    setVal(sheet.cell(row=5, column=4), "Убыл")
        
    sc = 5
    scc = sc

    for f in groups:
        sheet.merge_cells(start_row=5, start_column=sc, end_row=5, end_column=sc + 2)
        cell = sheet.cell(row=5, column=sc)
        cell.value = f.name
        makeBorder(cell)
        
        cell = sheet.cell(row=6, column=sc)
        cell.value = "Сток"
        makeBorder(cell)
        sheet.merge_cells(start_row=6, start_column=sc + 1, end_row=6, end_column=sc + 2)
        cell = sheet.cell(row=6, column=sc + 1)
        cell.value = len(f.items)
        makeBorder(cell)

        cell = sheet.cell(row=7, column=sc)
        cell.value = "Нал"
        makeBorder(cell)
        
        cell = sheet.cell(row=7, column=sc + 1)
        cell.value = "Зак"
        makeBorder(cell)
        
        cell = sheet.cell(row=7, column=sc + 2)
        cell.value = "Стало"
        makeBorder(cell)
        
        sc = sc + 3
        
    sheet.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=sc + 2)
    setVal(sheet.cell(row=4, column=sc),"ПОЗИЦИЙ")
    sheet.merge_cells(start_row=5, start_column=sc, end_row=6, end_column=sc)
    setVal(sheet.cell(row=5, column=sc), "Всего сток",wrap=True)
    setVal(sheet.cell(row=7, column=sc), "Маг")
    sc = sc + 1
    
    sheet.merge_cells(start_row=5, start_column=sc, end_row=6, end_column=sc)
    setVal(sheet.cell(row=7, column=sc), "Зак")
    sc = sc + 1
    
    sheet.merge_cells(start_row=5, start_column=sc, end_row=7, end_column=sc)
    setVal(sheet.cell(row=5, column=sc),"Стало магаз с зак",wrap=True)
    sc = sc + 1
    
    sheet.merge_cells(start_row=4, start_column=sc, end_row=6, end_column=sc)
    setVal(sheet.cell(row=4, column=sc),"% Охват сток",wrap=True)
    setVal(sheet.cell(row=7, column=sc), "Стало/Ст")
    
    sheet.column_dimensions[get_column_letter(1)].width = 3
    sheet.column_dimensions[get_column_letter(2)].width = 30
    sheet.column_dimensions[get_column_letter(3)].width = 30
    sheet.column_dimensions[get_column_letter(4)].width = 7
    sheet.column_dimensions[get_column_letter(5)].width = 7
     
    for x in range(scc+1, sc+1):
        sheet.column_dimensions[get_column_letter(x)].width = 6

def makeBorder(cell):
    cell.style.borders.left.border_style = Border.BORDER_THIN
    cell.style.borders.right.border_style = Border.BORDER_THIN
    cell.style.borders.top.border_style = Border.BORDER_THIN
    cell.style.borders.bottom.border_style = Border.BORDER_THIN

def makeReportData(sheet, groups, agentData):
    cr = 8
    count = 1
    
    data = sorted(agentData.orgs.iteritems(), key = lambda x: x[1].start)
    for org, orgData in data:
        cc = 0

        cell = sheet.cell(row=cr, column=cc) 
        cell.value = count
        makeBorder(cell)
        
        cc = cc + 1        
        cell = sheet.cell(row=cr, column=cc);
        cell.value = org.name
        makeBorder(cell)

        cc = cc + 1
        cell = sheet.cell(row=cr, column=cc);
        cell.value = org.address
        makeBorder(cell)
        
        cc = cc + 1
        cell = sheet.cell(row=cr, column=cc);
        cell.value = orgData.start.strftime("%H:%M") if orgData.start != 0 else ""
        makeBorder(cell)
 
        cc = cc + 1
        cell = sheet.cell(row=cr, column=cc);
        cell.value = orgData.finish.strftime("%H:%M") if orgData.finish != 0 else ""
        makeBorder(cell) 
        cc = cc + 1
        
        stock = 0
        rest = 0
        orders = 0
        current = 0
        
        for group in groups:
            for ctr in range(0, 4):        
                makeBorder(sheet.cell(row=cr+ctr, column=cc))
                makeBorder(sheet.cell(row=cr+ctr, column=cc+1))
                makeBorder(sheet.cell(row=cr+ctr, column=cc+2))
                makeBorder(sheet.cell(row=cr+ctr, column=cc+3))

            stock = stock + len(group.items)
            if group not in orgData.groups :
                cc = cc + 3
                continue
                        
            groupData = orgData.groups[group]
            setVal(sheet.cell(row=cr, column=cc), len(groupData.rest))
            rest = rest + len(groupData.rest)
            cc = cc + 1
            setVal(sheet.cell(row=cr, column=cc), len(groupData.orders))
            orders = orders + len(groupData.orders)
            cc = cc + 1
            setVal(sheet.cell(row=cr, column=cc), len(groupData.current))
            current = current + len(groupData.current)
            cc = cc + 1
        
        
        setVal(sheet.cell(row=cr, column=cc), rest)
        cc = cc + 1
        setVal(sheet.cell(row=cr, column=cc), orders)
        cc = cc + 1
        setVal(sheet.cell(row=cr, column=cc), current)
        cc = cc + 1
        prc = float(current) / stock * 100
        setVal(sheet.cell(row=cr, column=cc), "{0:.1f}".format(prc) + "%")
        cc = cc + 1
        
        agentData.rest = agentData.rest + rest
        agentData.orders = agentData.orders + orders
        agentData.current = agentData.current + current
        agentData.prcCurrent = agentData.prcCurrent + prc
        
        count = count + 1
        cr = cr + 1
    return cr   

def makeReportTail(sheet, cr, groups, agentData):
    cc = 2
    sheet.merge_cells(start_row=cr, start_column=cc, end_row=cr, end_column=cc + 2)
    setVal(sheet.cell(row=cr, column=cc), "Наличие в магазинах", hrz= Alignment.HORIZONTAL_LEFT)
    
    sheet.merge_cells(start_row=cr+1, start_column=cc, end_row=cr+1, end_column=cc + 2)
    setVal(sheet.cell(row=cr+1, column=cc), "Заказано в магазине", hrz= Alignment.HORIZONTAL_LEFT)
    
    sheet.merge_cells(start_row=cr+2, start_column=cc, end_row=cr+2, end_column=cc + 2)
    setVal(sheet.cell(row=cr+2, column=cc), "Всего будет с заказом", hrz= Alignment.HORIZONTAL_LEFT)
    
    sheet.merge_cells(start_row=cr+3, start_column=cc, end_row=cr+3, end_column=cc + 2)
    setVal(sheet.cell(row=cr+3, column=cc), "Процент охвата группы", hrz= Alignment.HORIZONTAL_LEFT)
    
    orgs = len(agentData.orgs)
    if orgs == 0: 
        orgs = 1
    
    cc = 5
    for group in groups:
        items = len(group.items)

        for ctr in range(0, 4):        
            makeBorder(sheet.cell(row=cr+ctr, column=cc))
            makeBorder(sheet.cell(row=cr+ctr, column=cc+1))
            makeBorder(sheet.cell(row=cr+ctr, column=cc+2))
            makeBorder(sheet.cell(row=cr+ctr, column=cc+3))

        rest = agentData.TotalRest(group)
        setVal(sheet.cell(row=cr, column=cc), "{0:.1f}".format(float(rest) / orgs))
        cc = cc + 1
    
        orders = agentData.TotalOrder(group)
        setVal(sheet.cell(row=cr+1, column=cc), "{0:.1f}".format(float(orders) / orgs))
        cc = cc + 1
        
        current = agentData.TotalCurrent(group)
        value = float(current) / orgs
        setVal(sheet.cell(row=cr+2, column=cc), "{0:.1f}".format(value))
        setVal(sheet.cell(row=cr+3, column=cc), "{0:.1f}".format(value / items * 100))
        cc = cc + 1

    setVal(sheet.cell(row=cr, column=cc), "{0:.1f}".format(agentData.rest / orgs))
    makeBorder(sheet.cell(row=cr+1, column=cc))
    makeBorder(sheet.cell(row=cr+2, column=cc))
    makeBorder(sheet.cell(row=cr+3, column=cc))
    cc = cc + 1

    setVal(sheet.cell(row=cr+1, column=cc), "{0:.1f}".format(agentData.orders / orgs))
    makeBorder(sheet.cell(row=cr, column=cc))
    makeBorder(sheet.cell(row=cr+2, column=cc))
    makeBorder(sheet.cell(row=cr+3, column=cc))
    cc = cc + 1
    
    setVal(sheet.cell(row=cr+2, column=cc), "{0:.1f}".format(agentData.current / orgs))
    makeBorder(sheet.cell(row=cr+1, column=cc))
    makeBorder(sheet.cell(row=cr, column=cc))
    makeBorder(sheet.cell(row=cr+3, column=cc))

    sheet.merge_cells(start_row=cr+1, start_column=cc+1, end_row=cr+3, end_column=cc+1)
    setVal(sheet.cell(row=cr+1, column=cc+1), "{0:.1f}".format(agentData.prcCurrent / orgs))
    makeBorder(sheet.cell(row=cr, column=cc+1))
    makeBorder(sheet.cell(row=cr+3, column=cc+1))
    
    sheet.cell(row=cr+4, column=0).value="КОММЕНТАРИИ:(максимальная дополнительная информация о клиента, конкурентных поставщиках, конкурирующих товарах, замечания и т.п., влияющее на изменение оборота товара)."

    return

def makeReport(sheet, groups, agent, date, agentData):
    makeReportHeader(sheet, groups, agent, date, agentData)
    cr = makeReportData(sheet, groups, agentData)
    makeReportTail(sheet, cr, groups, agentData)
    return
                
def doReport(server, param, sheet):
    
    agents = server.Get("Agents","","id")
    grp = Groups()
    grp.load(server)
    groups = grp.groups

    mgrPrice = server.Get("ManagerPrice", "", "id")
    
    agent = agents[param.agent]    
    agentData = AgentDailyData(server, param.agent)        

    start = param.date.strftime('%d/%m/%Y')
    tomorrow = param.date + timedelta(days=1) 
    finish = tomorrow.strftime('%d/%m/%Y')
    where = '"created" >= ToDate("' + start + '") and "created" < ToDate("' + finish + '") and "userid"=' + scope(param.agent)
    
    orders = server.Get("Order", where)
    
    if orders != None :
        for doc in orders:
            agentData.loadOrder(groups, doc, mgrPrice, param.folders)
    
    remnants = server.Get("OrgRemnants", where)
    if remnants != None :
        for doc in remnants:
            agentData.loadRest(groups, doc, mgrPrice, param.folders)
    
    where = '"start" >= ToDate(' + scope(start) + ') and "start" < ToDate(' + scope(finish) + ') and "userid"=' + scope(param.agent)
    worktimes = server.Get("WorkTime", where)
    if worktimes != None :
        for wt in worktimes:
            agentData.loadWorkTime(wt)
    
    
    makeReport(sheet, groups, agent, param.date, agentData)
#     updateGlobalData(server, param)        
#     makeHeader(server, param, sheet)
#     makeData(server, param, sheet)
#     makeFooter(server, param, sheet)

def doExcel(server, param, outObj):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()

    doReport(server, param, sheet)
    
    repName = "oneday.xlsx" 
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)
     
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
     
    obj = outObj.New()
    obj.name = repName
    obj.file = bytes



def run(server):
    print __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    doExcel(server, server.Params[0], outObj)
    server.Put(outObj)
    
    print __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
