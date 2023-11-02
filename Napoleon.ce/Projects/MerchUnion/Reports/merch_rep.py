# -*- coding: cp1251 -*-

from importlib import reload
import sys;
import logging

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys)

class Report:
  def __init__(self):
    self.start = datetime.now()
    self.finish = datetime.now()
    self.items = list()
        
class Item:
  def getData(self, row):
    return [self.orgName, self.orgAddress, self.codeTT, self.location, self.date, self.start, self.finish, self.agent,
      self.price, self.format, self.qty, self.face, self.cost, self.promo, self.oos, self.remark]

def loadData(server):
  start = server.Params[0].start
  finish = server.Params[0].finish
  title = server.Params[0].title
  
  WHERE_STR = '"created" >= ToDate("{0}") and "created" < ToDate("{1} 23:59:59") and "userid"=\'{2}\'';         
  
  res = Report()
  res.title = title
  res.start = start
  res.finish = finish
  
  commonPrice = server.Get("Price", "", "id")
  
  for a in server.Params[0].agents:
    server.ChangeUser(a.id)
    agent = server.CurrentUser().name
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    
    price = None
    
    if len(commonPrice) == 0:
      price = server.Get("Price", "", "id")
      
    orgs.update(porg)
    server.RestoreUser()
    
    where = WHERE_STR.format(start.strftime("%d/%m/%Y"), finish.strftime("%d/%m/%Y"),a.id)
    scripts = server.Get("ScriptDoc", where)
    remnants = server.Get("OrgRemnants", where)
    
    rmnMap = {}
    
    for r in remnants:
      rmnMap[r.created] = r
    
    for s in scripts:
       if s.items != None:
        for i in s.items:
          if i.type == 'OrgRemnants' and i.date in rmnMap:
            rmn = rmnMap[i.date]
            
            orgName = orgs[s.id].name if s.id in orgs else "Контрагент с кодом <{0}>".format(s.id)
            orgAddress = orgs[s.id].address if s.id in orgs else ""
            codeTT = orgs[s.id].codeTT if s.id in orgs else ""
            loaction = "{0} {1}".format(rmn.latitude, rmn.longitude)
            strdate = s.created.strftime("%d.%m.%Y")
            strstart = s.items[0].date.strftime("%d.%m.%Y %H:%M")
            strfinish = s.items[len(s.items)-1].date.strftime("%d.%m.%Y %H:%M")
            
            for ri in rmn.items:
              item = Item()
              item.orgName = orgName
              item.orgAddress = orgAddress
              item.codeTT = codeTT
              item.location = loaction
              item.date = strdate
              item.start = strstart
              item.finish = strfinish
              item.agent = agent
              item.price = getPriceName(commonPrice, price, ri.id)
              item.format = ri.format
              item.qty = ri.qty
              item.face = ri.face
              item.cost = ri.cost
              item.promo = "Да" if ri.promo == 1 else "Нет"
              item.oos = ri.oos
              item.remark = ri.remark
              
              res.items.append(item)
  
  return res

def getPriceName(all, price, id):
  res = "Товар с кодом <{0}>".format(id)
  
  if price != None and id in price:
    res = price[id].name
  elif id in all:
    res = all[id].name
   
  return res
  
def item_cmp(x, y):
  res = cmp(x.org, y.org)

  return res
  
def setCellWidth(sh, wa):
  cc = 1
  for w in wa:
    sh.column_dimensions[get_column_letter(cc)].width = w
    cc += 1
        
class XLBuilderEx(XLBuilder):
  HEAD_COLOR = "FFB6DDE8"

  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    self.setBackColor(cell, self.HEAD_COLOR)
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 3:
      cell.style.number_format._set_format_code('0%')
    
def printOut(rep, xlb):
  wb = Workbook(False, 'cp1251')
  sh = wb.get_active_sheet()
  
  cell = sh.cell(row=0, column=0)
  cell.value = "Отчет по мерчендайзингу"
  cell.style.font.bold = True
  cell.style.font.size = 18

  
  sh.cell(row=1, column=0).value = 'Интервал с {0} по {1}'.format(
      rep.start.strftime("%d.%m.%Y"), rep.finish.strftime("%d.%m.%Y"))
  
  head = ["Наименование", "Адрес", "Код ТТ", "Координаты", "Дата визита", "Начало визита", "Конец визита", "Торговый представитель", "Номенклатура(согласно матрице)",
    "Формат ТТ", "Остатки(кол-во)", "Фейсинг(кол-во шт)", "Цена полка", "Отметка промо", "Причина OOS", "Комментарий"]
  
  r = 2
  xlb.makeHead(sh, r, head)
  
  for d in rep.items:
    r += 1
    xlb.makeCells(sh, r, d.getData(r))
  
  setCellWidth(sh, [31,29,12,12,12,12,12,17,30,12,12,12,12,12,27,27])
              
  return wb

def doReport(server):
  data = loadData(server)
  xlb = XLBuilderEx()
  wb = printOut(data,xlb)
  xlb.workbookToObject(wb, "merch_rep.xlsx", server)
        
def run(server):
  print ("start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
  doReport(server)
  print ("finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    