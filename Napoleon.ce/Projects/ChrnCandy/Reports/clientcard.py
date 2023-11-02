# -*- coding: cp1251 -*-

import sys;
import locale

from datetime import datetime, timedelta
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment

reload(sys);
sys.setdefaultencoding("cp1251")

columns = list()

class Data:
    org = None 
    items = None
    
    def __init__(self, org):
        self.org = org
        self.items = dict()
        
    def addOrder(self, order, item, price):
        if not item.id in self.items:
            self.items[item.id] = Item(getItemName(price, item.id))
            
        self.items[item.id].addQty(order.created.strftime("%d.%m.%Y"), item.qty)    
        
    def addRemnants(self, remn, item, price):
        if not item.id in self.items:
            self.items[item.id] = Item(getItemName(price, item.id))
            
        self.items[item.id].addRemn(remn.created.strftime("%d.%m.%Y"), item.qty)  
        
class Item:
    name = None
    items = None
    
    def __init__(self, name):
        self.name = name
        self.items = dict()
        
    def addQty(self, date, qty):
        if not date in self.items:
            self.items[date] = Day(date)
        
        self.items[date].qty += qty
        
    def addRemn(self, date, rem):
        if not date in self.items:
            self.items[date] = Day(date)
        
        self.items[date].rem += rem
        
class Day:
    date = None
    qty = 0
    rem = 0    
    
    def __init__(self, date):
        self.date = date
        self.qty = 0
        self.rem = 0
        
def loadData(server):
    global columns
    
    userid = server.Params[0].userid
    start = server.Params[0].start
    finish = server.Params[0].finish
    finish = finish + timedelta(days=1)
    
    where = '"userid" in ({0}) and "date" >= ToDate("{1}") and "date" < ToDate("{2}")'.format(userid, start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"))
    orders = server.Get("Order", where)
    remnants = server.Get("OrgRemnants", where)
    price = server.Get("ManagerPrice", "", "id")
    
    server.ChangeUser("'{0}'".format(userid))
    orgs = server.Get("Org", "", "id")        
    server.RestoreUser()

    data = dict()
    
    c = start
    
    while c < finish:
        columns.append(c.strftime("%d.%m.%Y"))
        c += timedelta(days=1)
        
    for o in orders:
        if not o.id in data:
            data[o.id] = Data(getOrgName(orgs, o.id))
        
        for i in o.items:
            data[o.id].addOrder(o, i, price)
    
    for r in remnants:
        if not r.id in data:
            data[r.id] = Data(getOrgName(orgs, r.id))
        
        for i in r.items:
            data[r.id].addRemnants(r, i, price)
    
    return data        
            
def getOrgName(orgs, id):
    return orgs[id].name if id in orgs else "Контрагент с кодом<{0}>".format(id)
    
def getItemName(price, id):
    return price[id].name if id in price else "Товар с кодом<{0}>".format(id)
    
def printOut(data):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    printData(xlb, sh, data)
    
    return wb

def printData(xlb, sh, data):
    items = sorted(data.values(), lambda x,y: cmp(x.org, y.org))
    row = 0
    for i in items:
        row = printHead(row, xlb, sh, i)
        
        ii = sorted(i.items.values(), lambda x,y: cmp(x.name, y.name))

        pos = 1
        for t in ii:
            row = printRow(xlb, sh, t, pos, row)
            pos += 1
        
        row += 1
            
    setCellWidth(sh, [10,50])
            
def printRow(xlb, sh, data, pos, row):
    arr = [pos, data.name]
    
    for c in columns:
        if c in data.items:
            arr.append(data.items[c].rem)
            arr.append(data.items[c].qty)
        else:
            arr.append(0)
            arr.append(0)
    
    xlb.makeCells(sh, row, arr)
    
    return row + 1
    
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def printHead(row, xlb, sh, data):
    global columns
    bkgColor = "ffC0C0C0"
    
    cell = sh.cell(row=row, column=0);
    cell.value = "№ п/п"
    sh.merge_cells(start_row=row, start_column = 0, end_row = row+2, end_column = 0)
    xlb.paintHeadCell(cell)
    xlb.setBackColor(cell, bkgColor)
    xlb.makeBorder(sh.cell(row=row+1, column=0), XLBuilder.HEAD_BORDER_STYLE)
    xlb.makeBorder(sh.cell(row=row+2, column=0), XLBuilder.HEAD_BORDER_STYLE)

    
    cell = sh.cell(row=row, column=1);
    cell.value = "Продукты"
    sh.merge_cells(start_row=row, start_column = 1, end_row = row+2, end_column = 1)
    xlb.paintHeadCell(cell)
    xlb.setBackColor(cell, bkgColor)
    xlb.makeBorder(sh.cell(row=row+1, column=1), XLBuilder.HEAD_BORDER_STYLE)
    xlb.makeBorder(sh.cell(row=row+2, column=1), XLBuilder.HEAD_BORDER_STYLE)
    
    s = 2
    sh.merge_cells(start_row=row, start_column = s, end_row = row, end_column = s + len(columns) * 2 - 1)
    cell = sh.cell(row=row, column=s)
    cell.value = data.org
    xlb.paintHeadCell(cell)
    xlb.setBackColor(cell, bkgColor)
    
    row += 1
    
    for cx in range(0,len(columns)):
        cidx = s + cx * 2
        xlb.makeBorder(sh.cell(row=row-1, column=cidx), XLBuilder.HEAD_BORDER_STYLE)
        xlb.makeBorder(sh.cell(row=row-1, column=cidx+1), XLBuilder.HEAD_BORDER_STYLE)
        
        sh.merge_cells(start_row=row, start_column = cidx, end_row = row, end_column = cidx + 1)
        cell = sh.cell(row=row, column=cidx)
        cell.value = columns[cx]
        xlb.paintHeadCell(cell)
        xlb.setBackColor(cell, bkgColor)
        xlb.makeBorder(sh.cell(row=row, column=cidx), XLBuilder.HEAD_BORDER_STYLE)
        xlb.makeBorder(sh.cell(row=row, column=cidx+1), XLBuilder.HEAD_BORDER_STYLE)
        
        cell = sh.cell(row=row+1, column=cidx)
        cell.value="ост."
        xlb.paintHeadCell(cell)
        xlb.setBackColor(cell, bkgColor)
        
        cell = sh.cell(row=row+1, column=cidx + 1)
        cell.value="зак."
        xlb.paintHeadCell(cell)
        xlb.setBackColor(cell, bkgColor)
        
    return row + 2
    
class XLBuilderEx(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        
        if column == 0:
            cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
            cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    
def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    XLBuilder().workbookToObject(wb, "time.xlsx", server)
          
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
   
