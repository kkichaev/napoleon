# -*- coding: cp1251 -*-

import sys;
import locale

from datetime import datetime, timedelta
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter
from grsoft.xl_base import XLBuilder
from imp import reload
import logging
from openpyxl.style import Alignment, NumberFormat
from manager.summary import plannedOrgs, makeQuery
from manager.document import docTypes, Order

reload(sys)

bkgColor = "ffC0C0C0"

class RItem:
  name = None
  data = None
  
class Report:
  start = None
  finish = None
  items = None
  agents = None
  
  def __init__(self):
    self.items = list()
  
  def setData(self, data):
    for d in data.values():
      i = RItem()
      i.name = self.agents[d.id].name if d.id in self.agents else "Сотрудник с кодом <{0}>".format(d.id)
      i.data = d
       
      self.items.append(i)
      
    self.items = sorted(self.items, key=lambda p: p.name)  
     
class Data:
  id = None
  route = None
  orders = None
  summa = 0.0
  visitedWithOrder = None
  orders_count = 0
  orders_in_route = 0
  visits_in_route = 0
  visits_out_route = 0
  route_count = 0
  visited = None

  def __init__(self, id):
    self.id = id
    self.visited = set()
    self.visitedWithOrder = set()
    self.orders = set()
    
  def addOrder(self, order):
    self.orders.add(order.id)
    self.add(order)
    self.visitedWithOrder.add(order.id)
    
  def add(self, doc):
    self.summa += doc.sum()
    self.visited.add(doc.id)
        
  def updatePlan(self):
    for orgId in self.visited:
      if orgId in self.route:
        self.visits_in_route += 1
      else:
        self.visits_out_route += 1
                
    for orgId in self.visitedWithOrder:
      if orgId in self.route:
        self.orders_in_route += 1   
   
    self.route_count = len(self.route)
    self.orders_count = len(self.orders)
    
  def append(self, d):
    self.summa += d.summa
    self.orders_in_route += d.orders_in_route
    self.visits_in_route += d.visits_in_route
    self.route_count += d.route_count
    self.visits_out_route += d.visits_out_route
    self.orders_count += d.orders_count
    
def staffList(staff, mid):
  result = list()
  
  for id in staff:
    if id != mid:
      result.append(id)
  
  return result

def sumup(s, d):
  for i in d.values():
    if i.id in s:
      s[i.id].append(i)
    else:
      s[i.id] = i

def loadData(server):
  s = server.Params[0].start
  f = server.Params[0].finish
  
  cu = server.CurrentUser()
  staff = staffList(cu._SubUsers, cu.id)
      
  r = Report()
  r.start = s
  r.finish = f + timedelta(-1)
  r.agents = server.Get("Agents", "", "id")
  
  summary = dict()
  
  while s < f:
    dd = dayData(server, staff, s)
    sumup(summary, dd)
    s = s +  timedelta(1)
    
  r.setData(summary)  
  
  return r    
 
def useridin(users):
  str = ""
  
  for id in users:
    if len(str) > 0:
        str += ','
        
    str += "'{0}'".format(id)
  
  return '"userid" in({0})'.format(str)
      
def dayData(server, staff, date ):
    agentQuery = useridin(staff)
    result = dict()   
    
    for id in staff:
      d = Data(id)
      d.route = plannedOrgs(server, id, date)
      result[id] = d
        
    where = makeQuery(agentQuery,date)
  
    for dt in docTypes:
      docs = dt.docList(server, where)
      
      if dt.docWraper == Order or issubclass(dt.docWraper, Order)  :
        for doc in docs:
          result[doc.userid].addOrder(doc)
          
      else:
        for doc in docs:
          result[doc.userid].add(doc)
                
    for data in result.values():
      data.updatePlan()

    return result  
   
def printOut(data):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    printData(xlb, sh, data)
    
    return wb

def printData(xlb, sh, data):
  sh.cell(row=0, column=0).value = "Отчет о результативности визитов."
  sh.cell(row=1, column=0).value = "Период с {0} по {1}".format(data.start.strftime("%d.%m.%Y"), data.finish.strftime("%d.%m.%Y"))
  head = ["ФИО ТП", "Кол-во ТТ в маршруте (общее)", "Кол-во визитов общее", "Кол- во визитов по маршруту", "Кол- во визитов вне маршрута", "Сумма заказов", "Кол-во заявок", "% исполнения маршрута", "% заявок"]
  
  row = 3
  xlb.makeHead(sh, row, head, True)
  row += 2
  
  for i in data.items:
    xlb.makeCells(sh, row, [i.name, i.data.route_count, "=D{0} + E{0}".format(row+1), i.data.visits_in_route, 
      i.data.visits_out_route, i.data.summa, i.data.orders_count,
      "=IFERROR(D{0}/B{0}, 0)".format(row+1),"=IFERROR(G{0}/C{0}, 0)".format(row+1)])
    row += 1
  
  row = 4
  r1 = 6
  r2 = len(data.items) + r1 - 1
  
  if r2 > r1:
    xlb.makeCells(sh, row, ["ИТОГО", "=SUM(B{0}:B{1})".format(r1, r2), "=SUM(C{0}:C{1})".format(r1, r2),
      "=SUM(D{0}:D{1})".format(r1, r2), "=SUM(E{0}:E{1})".format(r1, r2), "=SUM(F{0}:F{1})".format(r1, r2),
      "=SUM(G{0}:G{1})".format(r1, r2), "=IFERROR(D{0}/B{0}, 0)".format(r1 - 1, r2), "=IFERROR(G{0}/C{0}, 0)".format(r1 - 1, r2)])  
    
  setCellWidth(sh, [30,20,20,20,20,20,20,20,20])
            
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
   
class XLBuilderEx(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment.wrap_text = True
        self.setBackColor(cell, bkgColor)
        return column  
    
    def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        
        if row == 4:
          self.setBackColor(cell,'ffE5E4E2')

        if column == 7 or column == 8 :
            cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE_00
    
def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    XLBuilder().workbookToObject(wb, "summary_report.xlsx", server)
          
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')
    
    doReport(server)
    
    logging.info('end')   
