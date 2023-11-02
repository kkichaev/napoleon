# -*- coding: cp1251 -*-
from importlib import reload
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

import datetime
from datetime import timedelta


import sys;
reload(sys)
#sys.setdefaultencoding("cp1251")


class QuestHelper:
  LIST_TYPE = 2
  DATASET_TYPE = 5
  PHOTO_TYPE = 7
  ORG_DATASET_TYPE = "Организация"

class Item:
  def __init__(self):
    self.address = ""
    self.org = ""
    self.created = ""
    self.user = ""
    self.orgcode = ""
    self.orgid = ""
    self.userid = ""
    self.div = ""
    self.remark = ""
  
  def values(self):
    res = [self.org, self.address, self.created.strftime('%d.%m.%Y'), self.div, "", "", self.user]
    res.extend(self.items)
    res[len(res) - 1] = self.remark
    res.append(self.sumItems())
    
    return res
    
  def sumItems(self):
    sum = 0.0
    
    for i in self.items:
      try:
        sum += float(i)
      except:
        sum += 0
    
    return sum
    
class ItemAgentSum:
  def __init__(self):
    pass
    
class ItemSlsSum:
  def __init__(self):
    pass
    
class ItemResultSum:
  def __init__(self):
    pass    
    
class ReportData:
  KEY_FMT =  "{0}\t{1}"
  ITEM_KEY_FMT = "{0}\t{1}\t{2}"

  def __init__(self):
    self.items = list()
    self.quests = list()
    self.cellidx = dict()
    self.agentsumitems = dict()
    self.slsnetitems = list()
    self.resultsum = list()
    
  def prepare(self, docs, quests):
    qu = list()
    
    for d in docs:
      if not d.question in qu:
        qu.append(d.question)
    
    idx = 0
    
    ql = quests.values()
    ql = sorted(ql, key= lambda x: x.name)
    
    for q in ql:
      if q.idquest in qu:
        quest = quests[q.idquest]
        self.quests.append(quest)
        
        quest.items = sorted(quest.items, key=lambda x: x.number)
          
        for i in quest.items:
          if i.type == QuestHelper.LIST_TYPE:
            for v in i.values:
              key = self.ITEM_KEY_FMT.format(quest.idquest, i.iditem, v.value)
              self.cellidx[key] = idx     
              idx += 1          
          else:
            key = self.KEY_FMT.format(quest.idquest, i.iditem)
            self.cellidx[key] = idx      
            idx += 1    

  def loadDocs(self, docs, pics, href):
  
    for d in docs:
      i = Item()
      i.created = d.created
      i.address = ''
      i.org = 'Контрагент с кодом <{0}>'.format(d.id)
      
      if d.id in self.orgs:
        i.address = self.orgs[d.id].address
        i.org = self.orgs[d.id].name
        i.orgcode = d.id
      
      i.orgid = d.id
      i.userid = d.userid
      
      if d.userid in self.agents:
        i.user = self.agents[d.userid].name
      
      if d.userid in self.agentDivision:
        i.div = self.agentDivision[d.userid].name
      
      i.items = [""]*len(self.cellidx)
      
      for n in d.items:
        if n.type ==QuestHelper.LIST_TYPE:
          self.putItemValue(i.items, self.ITEM_KEY_FMT.format(d.question, n.iditem, n.answer), "X")
          
        elif n.type == QuestHelper.PHOTO_TYPE:
          val = '=HYPERLINK("{0}{1}", "Фото")'.format(href, pics[n.answer].name) if n.answer in pics else "Фото не найдено!"
          self.putItemValue(i.items, self.KEY_FMT.format(d.question, n.iditem), val)
            
        elif n.type == QuestHelper.DATASET_TYPE:
          val = ''
          if n.remark == QuestHelper.ORG_DATASET_TYPE:
            val = self.orgs[n.answer].name if n.answer in self.orgs else n.answer
          
          self.putItemValue(i.items, self.KEY_FMT.format(d.question, n.iditem), val)
        else:
          val = ''
          
          if n.answer == '1':
            for q in self.quests:
              if q.idquest == d.question:
                for qi in q.items:
                  if qi.iditem == n.iditem and len(qi.values) > 1:
                    val = float(qi.values[0].value2) / 100
                    break
                    
              if val != '':
                break
          elif n.id == 'Комментарий':
            i.remark = n.answer.encode('utf-8','ignore').decode('utf-8')

          self.putItemValue(i.items, self.KEY_FMT.format(d.question, n.iditem), val)
      
      if i.sumItems() > 0:
        self.items.append(i)   
      
      
    #data for agentsummary
      
    agentsum = dict()
    
    for i in self.items:
      if not i.userid in agentsum:
        agentsum[i.userid] = dict()
        
      orgval = agentsum[i.userid]

      if not i.orgid in orgval:
        orgval[i.orgid] = list()
        
      orgval[i.orgid].append(i.sumItems())  
    
    for a in agentsum:
      orgavg = agentsum[a]
      
      for oid in orgavg:
        s = 0.0
        c = 0.0
        
        for v in orgavg[oid]:
          s += v
          c += 1
          
        if not a in self.agentsumitems:
          self.agentsumitems[a] = list()
        
        ias = ItemAgentSum()          
        self.agentsumitems[a].append(ias)
        ias.name = oid
        ias.address = ''
        ias.org = 'Контрагент с кодом <{0}>'.format(oid)
        
        if oid in self.orgs:
          ias.address = self.orgs[oid].address
          ias.org = self.orgs[oid].name
          
        ias.user = a
        
        if a in self.agents:
          ias.user = self.agents[a].name
        
        ias.avg = s / c if c > 0 else 0
    
    for a in self.agentsumitems:
      sorted(self.agentsumitems[a], key= lambda x: x.org+x.address)
      
    #data for slssum
    
    slssum = dict()
    
    for i in self.items:
      if i.orgid in self.orgs:
        sid = self.orgs[i.orgid].sid
      
        if not sid in slssum:
          slssum[sid] = list()
        
        slssum[sid].append(i.sumItems())  
        
    
    for ss in slssum:
      s = 0.0
      c = 0.0
      
      for v in slssum[ss]:
        s += v
        c += 1
        
      iss = ItemSlsSum()
      self.slsnetitems.append(iss)
      iss.slsnet = ss
      
      if ss in self.slsnet:
        iss.slsnet = self.slsnet[ss].name
        
      iss.avg = s / c if c > 0 else 0  
    
    sorted(self.slsnetitems, key=lambda x: x.slsnet)
    
    #data for resultsum
    
    ressum = dict()
    
    for i in self.items:
      if not i.userid in ressum:
        ressum[i.userid] = list()
      ressum[i.userid].append(i.sumItems())  
    
    for rs in ressum:
      s = 0.0
      c = 0.0
      
      for v in ressum[rs]:
        s += v
        c += 1
        
      irs = ItemResultSum()
      self.resultsum.append(irs)
      irs.user = rs
      
      if rs in self.agents:
        irs.user = self.agents[rs].name
      
      if rs in self.agentDivision:
        irs.div = self.agentDivision[rs].name
        
      irs.avg = s / c if c > 0 else 0  
    
    sorted(self.resultsum, key = lambda x: x.user)
    
  def putItemValue(self, items, key, value):
    if key in self.cellidx:
      idx = self.cellidx[key]
      items[idx] = value

class XLB(XLBuilder):
  YELLOW = Color("FFFFFF00")
  FIXED_CELL_COLOR = Color("FFB6DDE8")
  QUESTS_COLORS = [Color("FFD8D8D8"), Color("FFC2D69A")]
  
  def makeCell(self, sheet, row, column, cell, value):
    XLBuilder.makeCell(self, sheet, row, column, cell, value)        
    
    if column < 6:
      fill = cell.style.fill
      fill.fill_type = Fill.FILL_SOLID
      fill.start_color = XLB.FIXED_CELL_COLOR
      
    if column > 5:
      cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE
  
  def printHead(self, row, titles, sheet, report):
    cix = len(titles)
    
    sheet.merge_cells(start_row=row, start_column=0, end_row=row, end_column=cix - 1)
    self.printCell(sheet, row, 0, "Анкеты", XLB.YELLOW)
    
    for x in range(0, cix):
      sheet.merge_cells(start_row=row + 1, start_column=x, end_row=row + 2, end_column=x)
      self.printCell(sheet, row + 1, x, titles[x], XLB.FIXED_CELL_COLOR)
      sheet.column_dimensions[get_column_letter(x + 1)].width = 25
    
    i = 1
    for q in report.quests:
      color = Color("FFD8D8D8") #XLB.QUESTS_COLORS[len(XLB.QUESTS_COLORS) % i]
      i += 1
      cix = self.printGroup(row, cix, q, sheet, report, color)
    
    sheet.merge_cells(start_row=row, start_column=cix, end_row=row+2, end_column=cix)
    self.printCell(sheet, row, cix, "Итого", XLB.YELLOW)
    
    for x in range(0,3):
      self.makeBorder(sheet.cell(row=row+x, column=cix), XLBuilder.HEAD_BORDER_STYLE)
    
    row += 1
    
    return row
    
  def printGroup(self, row, column, quest, sheet, report, color):
    self.printCell(sheet, row, column, quest.name, XLB.YELLOW, True)
    cv = column

    for i in quest.items:
      self.printCell(sheet, row + 1, cv, i.id, color)
      
      if i.type == QuestHelper.LIST_TYPE:
        s = cv
        
        for v in i.values:
          self.printHeadColumn(sheet, row + 2, cv, v.value, color, 4)
          cv += 1
          
        sheet.merge_cells(start_row=row + 1, start_column=s, end_row=row + 1, end_column=cv-1)
          
      elif i.type == QuestHelper.DATASET_TYPE:
        self.printHeadColumn(sheet, row + 2, cv, i.values[0].value, color, 13)
        cv += 1
        
      else:
        sheet.column_dimensions[get_column_letter(cv + 1)].width = 13
        sheet.merge_cells(start_row=row + 1, start_column=cv, end_row=row + 2, end_column=cv)
        cv += 1
      
    sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=cv-1)
    
    for c in range(0, cv):
      for r in range(2, 6):
        self.makeBorder(sheet.cell(row=r, column=c), XLBuilder.HEAD_BORDER_STYLE)
    
    return cv  

  def printCell(self, sheet, row, column, value, color, bold=False):
    cell = sheet.cell(row=row, column=column)
    cell.value = value
    self.styleCell(cell, color, bold)
   
  def printHeadColumn(self, sheet, row, column, value, color, width):
    self.printCell(sheet, row, column, value, color)
    sheet.column_dimensions[get_column_letter(column + 1)].width = width
    
  def styleCell(self, cell, color, bold=False):
    style = cell.style
    style.font.bold = bold
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    fill = cell.style.fill;
    fill.fill_type = Fill.FILL_SOLID
    fill.start_color = color

class XLBW(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.fill.fill_type = Fill.FILL_SOLID
    cell.style.fill.start_color = Color("FFB6DDE8")
    return column
    
class XLB2(XLBW):    
  def makeCell(self, sheet, row, column, cell, value):
    XLBuilder.makeCell(self, sheet, row, column, cell, value)        
    
    if column == 3:
      cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE
      
class XLB3(XLBW):    
  def makeCell(self, sheet, row, column, cell, value):
    XLBuilder.makeCell(self, sheet, row, column, cell, value)        
    
    if column == 3:
      cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE   

class XLB4(XLBW):    
  def makeCell(self, sheet, row, column, cell, value):
    XLBuilder.makeCell(self, sheet, row, column, cell, value)        
    
    if column == 4:
      cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE      
      
def loadData(data, params, server):
    orgs = dict()
    porg = server.Get("PotenzialOrg", "", "id")
    data.agents = server.Get("Agents", "", "id")
    quests = server.Get("Question", '"idquest" is null or "idquest" is not null', "idquest")
    pictures = server.Get("PicStoreSrc", "", "id")
    orgs.update(porg)
    data.orgs = orgs
    data.agentDivision = server.Get("AgentDivision", "", "userid")
    
    if params.param == 0:
      for item in params.userids:
        server.ChangeUser(item.id)
        aorgs = server.Get("Org", '', 'id')
        server.RestoreUser()
        
        if aorgs != None:
            data.orgs.update(aorgs)

    where = '"created" >= ToDate("{0} 0:0:0") and "created" <= ToDate("{1} 0:0:0")'.format(params.start.strftime('%d.%m.%Y'), (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'))
    
    if len(params.userids) > 0:
      arr = ''
      
      for i in params.userids:
        if len(arr) > 0:
          arr += ','
        
        arr += "'" + i.id + "'"
      
      where = '{0} and "userid" in ({1})'.format(where, arr)
    
    if len(params.quests) > 0:
      arr = ''
      
      for i in params.quests:
        if len(arr) > 0:
          arr += ','
        
        arr += i.id
      
      where = '{0} and "question" in ({1})'.format(where, arr)
    
    print ("where: ", where)
    answers = server.Get("Answer" if params.param == 0 else "MAnswer", where)
    
    print("answers: ", len(answers))

    if params.param == 1:
      data.agents = server.Get("DivisionManager", "", "login")
      ids = []
      for a in answers:
        if len(a.id) > 0 and not a.id in ids: 
          ids.append(a.id)
          aorgs = server.Get("Org", '"id"=\'{0}\''.format(a.id), 'id')
          
          if aorgs != None:
            data.orgs.update(aorgs)
    
    data.prepare(answers, quests)
    data.orgs = orgs
    data.slsnet = server.Get("Slsnet", "", "id")
    
    href = params.hrefBase

    data.loadDocs(answers, pictures, href)
    
    return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLB()
    
    c = sheet.cell(row=0, column=0)
    c.value = "Отчет по анкетам"  
    c.style.font.bold = True
    c.style.font.size = 18

    c = sheet.cell(row=1, column=0)
    c.value = "Интервал: c {0} по {1}".format(params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y'))  
    
    r = 2
    arr = ["Наименование", "Адрес", "Дата", "Подразделение", "ТС", "СВ", "Мерчендайзер"]
    r += xlb.printHead(r, arr, sheet, data);
    
    for i in data.items:
      xlb.makeCells(sheet, r, i.values())
      r += 1

    #
    xlb = XLB2()
    
    for i in data.agentsumitems.values():
      titleName = "{0}".format(i[0].user)[:30]
      sheet = wb.create_sheet()
      sheet.title = titleName
      
      r = 2
      xlb.makeHead(sheet, r, ["Наименование", "Адрес", "Мерчендайзер", "Итого среднее по торговой точке"], True)
      
      r += 1
      for y in i:
        xlb.makeCells(sheet, r, [y.org, y.address, y.user, y.avg])
        r += 1
      
      cc = 1
      for w in [30, 30, 30, 30]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    #
    xlb = XLB3()
    sheet = wb.create_sheet()
    sheet.title = "Итого по сетям"
    
    r = 2
    xlb.makeHead(sheet, r, ["ТС", "СВ", "Сеть", "Итого среднее значение"], True)
    r+=1
    
    for i in data.slsnetitems:
      xlb.makeCells(sheet, r, ["", "", i.slsnet, i.avg])
      r += 1
    
    cc = 1
    for w in [30, 30, 30, 30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
      
    #
    xlb = XLB4()
    sheet = wb.create_sheet()
    sheet.title = "Итоговый отчет"
    
    r = 2
    xlb.makeHead(sheet, r, ["Подразделение", "ТС", "СВ", "Мерчендайзер", "Итого среднее значение"], True)
    r+=1
    
    for i in data.resultsum:
      xlb.makeCells(sheet, r, [i.div, "", "", i.user, i.avg])
      r += 1
    
    cc = 1
    for w in [30, 30, 30, 30, 30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1  
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "quest_rep.xlsx", server)                
    logging.info('end')
    
