# -*- coding: cp1251 -*-

from datetime import datetime, date, timedelta
from openpyxl import Workbook
from imp import reload
import logging
from openpyxl.style import Border
from openpyxl.style import Alignment, NumberFormat
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *
from grsoft.xl_base import XLBuilder
from calendar import monthrange

import sys;
import tempfile
import io
import time

reload(sys)

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN

class ItemCell:
    face = None
    qty = None
    
    def __init__(self):
        self.face = 0.0
        self.qty = 0.0

class PageItem:
    city = None
    user = None
    slsnet = None
    org = None
    data = None
    created = None
    partshelf = 0
    fact = None
    visit = None
    planogram = 0
    date = None
    address = None
    facemy = 0
    faceall = 0
    faceplan = 0
    facefact = 0
    remark = None
    usrid = None
    
    id = None
    avgItems = 0.0
    avgGroup = 0.0
    sid = None
    
    def __init__(self):
        self.data = dict()
        self.visit = list()
        self.avgItems = 0.0
        self.avgGroup = 0.0
        
    def process(self, c, report):
        self.city = report.getCity(c.id)
        self.user = report.getUser(c.userid)
        self.userid = c.userid
        self.slsnet = report.getSls(c.id)
        self.org = report.getOrgName(c.id)
        self.address = report.getOrgAddress(c.id)
        self.created = c.created.strftime("%d.%m.%Y")
        self.post_process(c, report)
        self.date = datetime(c.created.year, c.created.month, c.created.day)
        self.remark = c.remark
        self.id = c.id
        self.sid = report.getSlsID(c.id)

    def countAvg(self, avgItems, avgGroup):
        self.avgItems = 0
        self.avgGroup = 0
        
        if self.id in avgItems:
            cdata = avgItems[self.id]
            for val in cdata.values():
                self.avgItems += float(val)
                
            self.avgItems /= len(cdata)

        if self.id in avgGroup:
            cdata = avgGroup[self.id]
            for val in cdata.values():
                self.avgGroup += float(val)
                
            self.avgGroup = int(self.avgGroup / len(cdata) + 0.5)
        
    def post_process(self, c, report):
        ps = report.getPartShelf(c.id)
        self.partshelf = ps
        self.faceplan = report.getFacePlan(c.id);
        self.planogram = report.getPlanogram(c)
        
        other = 0
        my = 0
        all = 0
        
        for i in c.items:
            if i.id in report.mprice:
                self.data[i.id] = i.qty
                
                if report.mprice[i.id].my > 0:
                    my = my + i.qty
                else:
                    other = other + i.qty
                     
        all = my + other
        onepercent  = all / 100
        f = my / onepercent if onepercent != 0 else 0
        self.fact = "{:0.0f}".format(f)
        self.faceall = all
        self.facemy = my
        
        p = self.facemy / self.faceplan * 100 if self.faceplan != 0 else 0
        self.facefact = "{:0.0f}".format(p)
        report.getVisit(c, self.visit)
        
    def avgMePrc(self):
        res = 0 if self.avgGroup == 0 else int(self.avgItems * 100 / self.avgGroup + 0.5)
        return res
                   
class MPageItem(PageItem):
    def post_process(self, c, report):
       for i in c.items:
            if i.id in report.mprice:
                self.data[i.id] = i.cost

class AgentResultPageItem:
    userid = None
    agent = None
    plan = None
    fact = None
    akbm = None
    
    def __init__(self):
        self.userid = ""
        self.agent = ""
        self.plan = 0.0
        self.fact = 0.0
        self.akbm = 0.0 
        self.fitogo = 0.0
        
    def getData(self):
        return [self.agent, self.akbm, "='{0}'!E2".format(self.agent), "='{0}'!F2".format(self.agent)]    
        
class AgentResultPage:
    items = None
    
    def __init__(self):
        self.items = list()
                             
class Page:
    report = None
    items = None;
    title = None
    width = 0
    price = None
    factSlsIdx = 0;
    
    def __init__(self, report):
        self.report = report
        self.items = list()
        self.title = "Отчет по доли полки"
        self.width = 14
        self.price = list()
        self.factSlsIdx = -3;
        
    def collumns(self, sheet, r, c):
        c -= 1
        
        titles = ["Кол-во усл.ед. по нашей продукции", "Среднее по нашей продукции", "Итого доля полки в усл. ед.","План по доле полки в усл. ед.",
                  "Факт по доле полке, %", "План по доли полки %", "Факт по доли полки  на дату", "СРЕДНЕЕ по группе усл. ед.", 
                  "Факт по доле полки по торговой точке", 'Факт по доле полки по все сети']
        
        for str in titles:
            c += 1
            sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
            setVal(sheet.cell(row=r, column=c), str, Alignment.VERTICAL_BOTTOM, rotation=90)
            col_letter = get_column_letter(c + 1) 
            sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Отметка о выполнении планограммы", Alignment.VERTICAL_BOTTOM, rotation=90)    
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Комментарий", Alignment.VERTICAL_BOTTOM, rotation=0)   
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 35

    def itemsdata(self, sheet, r, c, i):
        prcInt = 0 if i.faceall == 0 else int(i.facemy * 100/i.faceall + 0.5) 
        prcMe = str(prcInt) + '%'
        planPrc = str(0 if i.avgGroup == 0 else int(i.faceplan * 100 / i.avgGroup + 0.5)) + '%'
        avgMePrc = str(i.avgMePrc()) + '%'
        
        data = [i.facemy, i.avgItems, i.faceall, i.faceplan, i.facefact, planPrc, prcMe, i.avgGroup, avgMePrc, 0]
        
        for d in data:
            setVal(sheet.cell(row=r, column=c), d)
            c = c + 1

        setVal(sheet.cell(row=r, column=c), i.planogram if i.planogram > 0 else "")      
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.remark)      
        c = c + 1
        
        for v in i.visit:
            idx = 0
            if v.items != None:
                for vi in v.items:
                    cell = sheet.cell(row=r, column=c) 
                    cell.hyperlink = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                    cell.value = "Фото{0}".format(idx + 1)
                    idx = idx + 1
                    c = c + 1        
                
        return c
    
    def process(self, report):
        avgItems = dict()
        avgGroup = dict()
        
        for c in report.contract:
            if not (report.allowedOrgs == None or c.id in report.allowedOrgs) :
                continue
            pi = PageItem()
            pi.process(c, report)
            self.items.append(pi)
            
            ctDate = None
            if not (c.id in avgItems):
                ctDate = dict()
                avgItems[c.id] = ctDate
            else:
                ctDate = avgItems[c.id]

            gtData = None
            if not (c.id in avgGroup):
                gtData = dict()
                avgGroup[c.id] = gtData
            else:
                gtData = avgGroup[c.id]
            
            cdata = c.created.date()
            val = 0
            if (cdata in ctDate):
                val = ctDate[cdata]
            ctDate[cdata] = pi.facemy + val  

            val = 0
            if (cdata in gtData):
                val = gtData[cdata]
            gtData[cdata] = pi.faceall + val  
            
            if not c.userid in report.agentPages:
                ap = self.createAgentPage()
                ap.process(c.userid, report)
                report.agentPages[c.userid] = ap  
            
        self.items = sorted(self.items, key=lambda x:x.created)
        
        for item in self.items:
            item.countAvg(avgItems, avgGroup)
    
    def createAgentPage(self):
        return AgentPage()  
      
class PageV2(Page):
    def __init__(self, report):
        Page.__init__(self, report)
        self.width = 6
        self.factSlsIdx = -2;
        
    def collumns(self, sheet, r, c):
        c -= 1
        
        titles = ["Кол-во усл.ед. по нашей продукции", "Среднее по нашей продукции", "Итого доля полки в усл. ед.","План по доле полки в усл. ед.",
                  "План по сети %", "Выполнение плана по сети %", "СРЕДНЕЕ по группе усл. ед.", 
                  "Факт по доле полки по торговой точке", 'Факт по доле полки по все сети']
        
        for str in titles:
            c += 1
            sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
            setVal(sheet.cell(row=r, column=c), str, Alignment.VERTICAL_BOTTOM, rotation=90)
            col_letter = get_column_letter(c + 1) 
            sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Комментарий", Alignment.VERTICAL_BOTTOM, rotation=0)   
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 35    
        
    def itemsdata(self, sheet, r, c, i):
        prcInt = 0 if i.faceall == 0 else int(i.facemy * 100/i.faceall + 0.5) 
        prcMe = str(prcInt) + '%'
        planPrc = str(0 if i.avgGroup == 0 else int(i.faceplan * 100 / i.avgGroup + 0.5)) + '%'
        avgMePrc = str(i.avgMePrc()) + '%'
        
        slsPlan = 0
        
        if i.sid in self.report.slsnet:
            slsPlan = self.report.slsnet[i.sid].plan
        
        data = [i.facemy, i.avgItems, i.faceall, "=ROUND({1}{0}*{2}{0}/100, 0)".format(r+1, get_column_letter(c+3), get_column_letter(c+5)), 
                slsPlan, "=IFERROR(ROUND({1}{0}/{2}{0}*100, 0), 0)".format(r+1, get_column_letter(c+1), get_column_letter(c+4)), i.avgGroup, avgMePrc, 0]
        
        for d in data:
            setVal(sheet.cell(row=r, column=c), d)
            c = c + 1

        setVal(sheet.cell(row=r, column=c), i.remark)      
        c = c + 1
        
        for v in i.visit:
            idx = 0
            if v.items != None:
                for vi in v.items:
                    cell = sheet.cell(row=r, column=c) 
                    cell.hyperlink = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                    cell.value = "Фото{0}".format(idx + 1)
                    idx = idx + 1
                    c = c + 1        
                
        return c

class PageV4(Page):
    def createAgentPage(self):
        return AgentPageV4()      
        
class MonitoringPage(Page):
    def __init__(self, report):
        Page.__init__(self, report)
        self.title = "Ценовой мониторинг"
        self.width = 6
        
    def collumns(self, sheet, r, c):
        pass
    
    def itemsdata(self, sheet, r, c, i):
        pass
    
    def process(self, report):
        for m in report.monitoring:
            pi = MPageItem()
            pi.process(m, report)
            self.items.append(pi)
                         
# def price_cmp(lhs, rhs):
#     result = (int)(lhs.my - rhs.my)
    
#     if result == 0:
#         result = cmp(lhs.group, rhs.group)
    
#     if result == 0:
#         result = cmp(lhs.name, rhs.name)
            
#     return result

class AgentPageItem:
    slsnet = None
    sid = None
    akbm = 0
    plan = 0
    
    def __init__(self):
        self.slsnet = ""
        self.sid = ""
        self.akbm = 0
        self.plan = 0

    def getData(self, row):
        return [self.slsnet, self.akbm, self.plan / 100, 0, "=B{0}*C{0}".format(row+1), "=B{0}*D{0}".format(row+1)]     
    
class AgentPage:
    agent = None
    items = None
    userid = None
    
    def __init__(self):
        self.agent = ""
        self.userid = ""
        self.items = dict()
        
    def process(self, userid, report):
        self.agent = report.getUser(userid)
        self.userid = userid
        server = report.server
        
        server.ChangeUser("'" + userid + "'")
        orgs = server.Get("Org", "", "id")
        server.RestoreUser()
        
        self.initItems(orgs, report)
            
    def initItems(self, orgs, report):
        for o in orgs.values():
            if not o.sid in self.items:
                api = AgentPageItem()
                api.slsnet = o.sid
                api.sid = o.sid
                
                if o.sid in report.slsnet:
                    sls = report.slsnet[o.sid]
                    api.slsnet = sls.name
                    api.koef = sls.koef
                    api.plan = sls.plan
                
                self.items[o.sid] = api
                
            api = self.items[o.sid]
            api.akbm += 1

class AgentPageItemV4:
    __slots__ = ["id", "org", "address", "plan", "responsible"]    
    
    def __init__(self):
        self.org = ""
        self.id = ""
        self.address = ""
        self.responsible = ""
    
    def getData(self, r):
        return [self.org, self.address, "", "", "","", ""]    
               
class AgentPageV4(AgentPage): 
    def initItems(self, orgs, report):
        for o in orgs.values():
            if not o.id in report.orgIds: continue
            
            if not o.id in self.items:
                api = AgentPageItemV4()
                
                api.id = o.id
                api.org = o.name
                api.address = o.address
                api.responsible = o.responsible
                
                if o.sid in report.slsnet:
                    sls = report.slsnet[o.sid]
                    api.plan = sls.plan
                    
                self.items[o.id] = api
                    
class Report:
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье" ]

    pages = None
    
    """ Список прайс """
    price = None
    """ Позиции в списке по ID """
    pidx = None
    
    slsnet = None
    city = None
    org = None
    agents = None
    partshelf = None
    mprice = None
    visit = None
    planogram = None
    contract = None
    
    matrix = None
    allowedOrgs = None

    visitdata = None
    planogramdata = None
    monitoringdata = None
    
    reportItemID = None
    agentPages = None
    server = None
    resultAgentPage = None
    
    orgIds = None
    
    def loadMatrix(self, server):
        param = server.Params[0]
        cid = param.cid
        matrixName = param.matrix

        if matrixName == "" or matrixName == None :
            return
        omtx = server.Get("OrgMatrix", '"cdef"=' + "'" + cid + "'" + ' and "name"=' + "'" + param.matrix + "'")
        if omtx == None:
            return

        self.allowedOrgs = list()
        for omi in omtx:
            self.allowedOrgs.append(omi.id)

        mtx = server.Get("CommonMatrix", '"name"=' +"'" + omtx[0].name + "'")
        if mtx == None:
            return
        self.matrix = list()
        for oi in mtx[0].items:
            if oi.id in self.mprice:
               self.matrix.append(oi.id)

    def __init__(self, server, start, finish, pageType):
        param = server.Params[0]
        cid = param.cid
        usePhoto = param.photo
        userid = ','.join(["'%s'" % x.id for x in param.userids])
        
        self.server = server
        self.reportItemID = None
        self.agentPages = dict()
        self.resultAgentPage = AgentResultPage()
        
        #self.weeks = Weeks(start, finish)
        
        pids = ""
        
        if cid != None:
          conDef = server.Get("ContractDef", '"id"=' + "'" + cid + "'")
          
          for cd in conDef:
              for i in cd.items:
                  iid = "'" + i.id +"'"
                  if iid in pids: 
                      continue
                  
                  if len(pids) > 0:
                      pids += ","
                      
                  pids += iid
        
        self.mprice = server.Get("ManagerPrice", '"id" in (' + pids + ')', "id")

        if param.item != '' and param.item != None:
            removed = list()
            item = param.item.upper()
            for p in self.mprice.values():
                if p.name.upper() != item:
                    removed.append(p.id)
                elif p.my == 1: 
                    self.reportItemID = p.id
                    
            for id in removed:
                self.mprice.pop(id, None)
                
        self.price = list();
        self.price.extend(self.mprice.values())
        self.price = sorted(self.price, key=lambda x: (x.my,x.group,x.name))
        
        self.loadMatrix(server)
        
        self.pidx = dict()
        for i in range(0, len(self.price)):
            self.pidx[self.price[i].id] = i    
        
        self.org = server.Get("Org", '"id" is not null', "id")
        self.agents = server.Get("Agents","","id")
        
        endRange = finish + timedelta(days=1) 
        
        where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + \
          endRange.strftime('%d/%m/%Y') + '") and "def"='+"'" + cid + "'" + ' and "userid" in (' + userid + ')'
          
        self.contract = server.Get("Contract", where)
        vobj = "Visit" if usePhoto == 1 else "VisitInfo"
        self.visit = server.Get(vobj, where)
        
        #where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + endRange.strftime('%d/%m/%Y') + '")'
        self.planogram = server.Get("Planogram", where)
        self.monitoring = server.Get("CMonitoring", where)
        self.partshelf = server.Get("PartShelf", '"cid"= ' + "'" + cid + "'", "sid") 
        self.btlplan = server.Get("BtlPlan", '"cid"= ' + "'" + cid + "'", "id")
        
        self.orgIds = dict()
        self.addActiveOrgs(self.contract)
        self.addActiveOrgs(self.visit)
        self.addActiveOrgs(self.planogram)
        self.addActiveOrgs(self.monitoring)
        
        class_ = getattr(__import__(__name__), pageType)
        p = class_(self)
        
        self.pages = [MonitoringPage(self), p]
        
        self.slsnet = server.Get("Slsnet","","id")
        self.city = server.Get("City", "", "id")
        
        for p in self.pages:
            p.process(self)
            hasval = list()
            
            for r in p.items:
                for d in r.data:
                    if not (self.matrix == None or d in self.matrix):
                        continue
                    if not d in hasval and r.data[d] > 0 and d in self.mprice:
                        p.price.append(self.mprice[d])
                        hasval.append(d)
                        
            if not self.matrix == None:
                for d in self.matrix:
                    if not d in hasval:
                        p.price.append(self.mprice[d])
                        hasval.append(d)

            p.price = sorted(p.price, key=lambda x: (x.my,x.group,x.name))
        
        psd = self.pages[1]
        avgFactAgentOrg = dict()
        
        for pi in psd.items:
            if not pi.userid in avgFactAgentOrg:
                avgFactAgentOrg[pi.userid] = dict()
                
            if not pi.id in avgFactAgentOrg[pi.userid]:
                avgFactAgentOrg[pi.userid][pi.id] = pi.avgMePrc()
        
        avgFactAgent = dict()
        
        for userid in avgFactAgentOrg:
            sum = 0.0
            count = 0
            
            for v in avgFactAgentOrg[userid].values():
                sum += v
                count += 1
            
            avgFactAgent[userid] = sum / count if count != 0 else 0    
             
        ap = server.Get("AgentPlan", "", "id")
        self.doResultPage(ap, avgFactAgent)
    
    def addActiveOrgs(self, docs):
        for d in docs:
            self.orgIds[d.id] = True
        
    def doResultPage(self, ap, avgFactAgent):
        for k,v in self.agentPages.items():
            pi = AgentResultPageItem()
            pi.agent = v.agent
            pi.userid = k
            
            if k in ap:
                pi.plan = ap[k].plan
                    
            for i in v.items.values():
                pi.akbm += i.akbm
            
            if v.userid in avgFactAgent:
                pi.fact = avgFactAgent[v.userid]
            
            self.resultAgentPage.items.append(pi)
                        
    def getCity(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].cid
        
        return result
    
    def getVisitFacePlan(self, id):
        return self.btlplan[id].face if id in self.btlplan else 0.0
    
    def getVisitInRoute(self, userid, orgid, route):
        count = 0
        usedDates = list()
        for v in self.visit:
            if v.userid != userid or v.id == orgid or v.date.date() in usedDates:
                continue

            usedDates.append(v.date.date())
            curDay = Report.days[v.date.weekday()]
            if not curDay in route:
                continue

            for ofi in route[curDay].items:
                if ofi.name == orgid:
                    count += 1
                    break
        return count
                
        
    def getFacePlan(self, id):
        result = 0.0;
        
        if id in self.btlplan:
            plan = self.btlplan[id]
            
            if self.reportItemID != None:
                for item in plan.items:
                    if item.id == self.reportItemID:
                        result = item.face
                        break
            else:
                result = plan.face
        
        return result;    
        
    def getOrgName(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].name
        
        return result
    
    def getOrgAddress(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].address
        
        return result
            
    def getSls(self, id):
        result = id
        
        if id in self.org and self.org[id].sid in self.slsnet:
            result = self.slsnet[self.org[id].sid].name
        
        return result      
    
    def getSlsID(self, id):
        result = id
        
        if id in self.org and self.org[id].sid in self.slsnet:
            result = self.slsnet[self.org[id].sid].id
        
        return result 
    
    def getUser(self, id):
        result = id
        
        if id in self.agents:
            result = self.agents[id].name
            
        return result
      
    def getAddress(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].address 
            
        return result 
    
    def sameDay(self, d1, d2):
        return d1.day == d2.day and d1.month == d2.month and d1.year == d2.year
     
    def getPartShelf(self, id):
        result = 0
        
        if id in self.org and self.org[id].sid in self.partshelf:
            result = self.partshelf[self.org[id].sid].part
            
        return result;
    
    def getVisit(self, c, val):
        if self.visitdata == None and len(self.visit) > 0:
            self.visitdata = dict()
            
            for v in self.visit:
                if not v.id in self.visitdata:
                    self.visitdata[v.id] = dict()
                 
                v1 = self.visitdata[v.id]
                
                if not v.userid in v1:
                    v1[v.userid] = dict()
                
                v2 = v1[v.userid]
                dt = datetime(v.created.year, v.created.month, v.created.day)
                if not dt in v2:
                    v2[dt] = list()
                    
                v2[dt].append(v)         
        
        if self.visitdata != None:
            cdt = datetime(c.created.year, c.created.month, c.created.day)
            
            if c.id in self.visitdata and c.userid in self.visitdata[c.id] and cdt in self.visitdata[c.id][c.userid]:              
                val.extend(self.visitdata[c.id][c.userid][cdt])
    
    def getPlanogram(self, c):
        if self.planogramdata == None:
            self.planogramdata = dict()
            
            for p in self.planogram:
                if not p.id in self.planogramdata:
                    self.planogramdata[p.id] = dict()
                    
                p1 = self.planogramdata[p.id]
                
                if not p.uesrid in p1:
                    p1[p.userid] = dict()
                
                p2 = p1[p.userid]  
                dt = datetime(p.created.year, p.created.month, p.created.day)
                p2[dt] = p
        
        cdt = datetime(c.created.year, c.created.month, c.created.day)
        if c.id in self.planogramdata and c.userid in self.planogramdata[c.id] and cdt in self.planogramdata[c.id][c.userid]:              
            return self.planogramdata[c.id][c.userid][cdt].approved
         
        return 0    
              
def setVal(cell, value, vrt = Alignment.VERTICAL_CENTER, hrz= Alignment.HORIZONTAL_CENTER, wrap=True, bold=False, rotation=0):
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    cell.style.alignment.text_rotation = rotation
    cell.style.font.bold = bold
    cell.value = value
                      

def countAvgSls(items):
    avgCount = dict()
    avgVal = dict()
    
    orgs = list()
    
    for i in items:
        if i.id in orgs:
            continue
        
        avg = 0
        avgctr = 0
        key = i.slsnet + i.user
        
        if key in avgVal:
            avg = avgVal[key]
            avgctr = avgCount[key]
            
        avg += i.avgMePrc()
        avgctr += 1
        avgVal[key] = avg
        avgCount[key] = avgctr
        
        avg = 0
        avgctr = 0
        key = i.slsnet
        
        if key in avgVal:
            avg = avgVal[key]
            avgctr = avgCount[key]
            
        avg += i.avgMePrc()
        avgctr += 1
        avgVal[key] = avg
        avgCount[key] = avgctr
        
        orgs.append(i.id)

    avgSls = dict()
    
    for k, v in avgVal.items():
        avg = int(float(v) / avgCount[k] + 0.5)
        avgSls[k] = avg
        
    return avgSls                      

class XLBuilderEx(XLBuilder):
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 2 or column == 3 or column == 4 or column == 5:
      cell.style.number_format._set_format_code('0%')
      
def doReportV12(server):
    server.RegisterType("Result[name:s,file:b,items[name:s,photo:b]]")
    outObj = server.New("Result")
    p = server.Params[0]
    pt = "Page" if p.version == 1 else "PageV2"
    report = Report(server, p.start, p.finish, pt)
    wb = Workbook(False, 'cp1251')
    
    sheet = None
    slsAvg = []
    for page in report.pages:
        if len(page.price) == 0:
            continue
        
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()
        sheet.title = page.title
        
        i = 1
        r = 0

        sheet.merge_cells(start_row=r, start_column=0, end_row=r+1, end_column=0)    
        setVal(sheet.cell(row=r, column=0), "Город")
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
        setVal(sheet.cell(row=r, column=1), "Мерчендайзер")
        col_letter = get_column_letter(2) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        setVal(sheet.cell(row=r, column=2), "Наименование торговой сети")
        col_letter = get_column_letter(3) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        setVal(sheet.cell(row=r, column=3), "Наименование торговой точки")
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)
        setVal(sheet.cell(row=r, column=4), "Адрес")
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        setVal(sheet.cell(row=r, column=5), "Дата заполнения отчета")
        col_letter = get_column_letter(6) 
        sheet.column_dimensions[col_letter].width = 20

        group = None
        START_COL = 6
        c = START_COL
        pc = c
        isMy = False
        
        for p in page.price:
            if group != None and group != p.group: 
                sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
                setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
                pc = c
                
            group = p.group
            isMy = p.my > 0
            
            setVal(sheet.cell(row=r + 1, column=c), p.name, vrt = Alignment.VERTICAL_BOTTOM, rotation=90)
            col_letter = get_column_letter(c + 1) 
            sheet.column_dimensions[col_letter].width = 5
            
            c = c + 1
            
        sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
        setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
        sheet.row_dimensions[2].height = 95
        
        page.collumns(sheet, r, c)
        
        r = r + 2
        dt = None
        sr = r + 1
        dsr = sr
        
        sum_group = dict()
        sum_result = dict()
        last_date = None
        
        slsAvg = countAvgSls(page.items)
        for i in page.items:
            if dt == None:
                dt = i.date
                   
            setVal(sheet.cell(row=r, column=0), i.city, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=1), i.user, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=2), i.slsnet, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=3), i.org, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=4), i.address, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=5), i.created, hrz= Alignment.HORIZONTAL_RIGHT)
            
            c = START_COL;
            last_date = i.created
            
            for p in page.price:
                if p.id in i.data:
                    setVal(sheet.cell(row=r, column=c), i.data[p.id])
                    
                    if not p.id in sum_group:
                        sum_group[p.id] = i.data[p.id]
                    else:     
                        sum_group[p.id] = sum_group[p.id] + i.data[p.id]  
                    
                c = c + 1
            
            lastCol = page.itemsdata(sheet, r, c, i)
            avgCount = slsAvg[i.slsnet] if i.slsnet in slsAvg else 0
            
            if avgCount > 0:
                avgVal = str(avgCount) + '%'
                setVal(sheet.cell(row=r, column=lastCol + page.factSlsIdx), avgVal, hrz= Alignment.HORIZONTAL_LEFT)
            
            r = r + 1
        
        rangeBorders(sheet.range("A1:"+get_column_letter(lastCol)+str(r)) )
        
    apl = list()
    apl.extend(report.agentPages.values())
    apl = sorted(apl, key=lambda x:x.agent)
    
    xlb = XLBuilderEx()
    head = ["Торговая сеть", "АКБ (м)", "Плановая доля полки", "Фактическая доля полки по сети", "Расчет плана ИТОГО по МЧ, %", "Расчет факта ИТОГО по МЧ, %"]
    resitems = dict()
    
    for ap in apl:
        sheet = wb.create_sheet()
        sheet.title = ap.agent
        xlb.makeHead(sheet,0,head,True)
        cc = 1
        for w in [20,20,20,20,20,20]:
            sheet.column_dimensions[get_column_letter(cc)].width = w
            cc += 1
        
        items = ap.items.values()
        items = sorted(items, key=lambda x:x.slsnet)
        
        r = 1
        sz = len(items)+2
        itogo = ["Итого","=SUM(B3:B{0})".format(sz),"","","=IFERROR(SUM(E3:E{0})/B2, 0)".format(sz),"=IFERROR(SUM(F3:F{0})/B2, 0)".format(sz)]
        xlb.makeCells(sheet, r, itogo)
        r += 1
        
        for d in items:
            xlb.makeCells(sheet, r, d.getData(r))
            key = d.slsnet+ap.agent
            avgCount = slsAvg[key] if key in slsAvg else 0
            avgVal = str(avgCount) + '%'
            setVal(sheet.cell(row=r, column=3), avgVal, hrz= Alignment.HORIZONTAL_RIGHT)
                
            r += 1
            
            if not d.sid in resitems:
                api = AgentPageItem()
                api.slsnet = d.slsnet
                api.plan = d.plan
                resitems[d.sid] = api
            
            api = resitems[d.sid]
            api.akbm += d.akbm
            
# ResultPage    
    sheet = wb.create_sheet()
    sheet.title = "ИТОГО по всем"
    xlb.makeHead(sheet,0,head,True)
    
    cc = 1
    for w in [20,20,20,20,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
            
    items = resitems.values()  
    items = sorted(items, key=lambda x:x.slsnet)
    r = 1
    
    sz = len(items)+2
    itogo = ["Итого","=SUM(B3:B{0})".format(sz),"","","=IFERROR(SUM(E3:E{0})/B2, 0)".format(sz),"=IFERROR(SUM(F3:F{0})/B2, 0)".format(sz)]
    xlb.makeCells(sheet, r, itogo)
    r += 1
    
    for d in items:
        xlb.makeCells(sheet, r, d.getData(r))
        
        avgCount = slsAvg[d.slsnet] if d.slsnet in slsAvg else 0
        avgVal = str(avgCount) + '%'
        setVal(sheet.cell(row=r, column=3), avgVal, hrz= Alignment.HORIZONTAL_RIGHT)
            
        r += 1  
    
# AgentResultPage
    sheet = wb.create_sheet()
    sheet.title = "ИТОГО по МЧ"
    
    head = ["Торговая сеть", "АКБ (м)", "Плановая доля полки", "Фактическая доля полки по сети"]
    xlb.makeHead(sheet,0,head,True)
    
    cc = 1
    for w in [20,20,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
            
    items = report.resultAgentPage.items  
    items = sorted(items, key=lambda x: x.agent)
    r = 1
    
    for d in items:
        xlb.makeCells(sheet, r, d.getData())
        r += 1  
            
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    for v in report.visit:
        idx = 0
        if v.items != None:
            for vi in v.items:
                item = obj.items.New()
                item.name = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                item.photo = vi.id
                idx = idx + 1
    
    server.Put(outObj)

class Pair:
    p1 = None
    p2 = None
    
    def __init__(self, a, b):
        self.p1 = a;
        self.p2 = b;
      
def getTimeRanges(server):
    result = list();
    s = server.Params[0].start
    f = server.Params[0].finish
    
    while s < f:
        s = datetime(s.year, s.month, 1, 0, 0, 0)
        d1, d2 = monthrange(s.year, s.month)
        ss = s + timedelta(days=d2-1)
        result.append(Pair(s, ss))
        s = ss + timedelta(days=1)
        
    return result;

class XLBuilder3Ex(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        if column > 0:
            cell.style.number_format._set_format_code('[$-F419]yyyy\,\ mmmm;@')
            
        return column
          
def doReportV3(server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    arr = getTimeRanges(server);
    
    reports = dict()
    agents = dict()
    slsnets = dict()
    orgs = dict()
    
    for a in arr:
        params = server.Params[0]
        params.start = a.p1
        params.finish = a.p2
        reports[a] = Report(server, params.start, params.finish, "Page")
        
    for r in reports.values():
        for ri in r.resultAgentPage.items:
            if not ri.userid in agents:
                agents[ri.userid] = ri;
                
        apl = list()
        apl.extend(r.agentPages.values())
        
        for ap in apl:
            for d in ap.items.values():
                if not d.sid in slsnets:
                    slsnets[d.sid] = d
        
        if len(r.pages) >= 1:
            p = r.pages[1];
            for i in p.items:
                if not i.id in orgs:
                    orgs[i.id] = i
                        
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    xlb = XLBuilder3Ex()
    head = ["Параметр/период"]
    
    for a in arr:
        head.append(a.p1);
    
    xlb.makeHead(sheet,0,head)
    
    # пїЅпїЅпїЅпїЅпїЅпїЅ
    r = 1
    for k,v in agents.items():
        d = []
        d.append(v.agent + ' (мерч)')
        
        for a in arr:
            added = False
            if a in reports:
                rp = reports[a]
                
                if len(rp.pages) >= 1:
                    p = rp.pages[1]
                    slsAvg = countAvgSls(p.items)
                    akbm = 0.0
                    sumv = 0.0
                    
                    if k in rp.agentPages:
                        ap = rp.agentPages[k]
                        for ki in ap.items.values():
                            akbm += ki.akbm
                            key = ki.slsnet + ap.agent
                            avg = slsAvg[key] if key in slsAvg else 0;
                            sumv += ki.akbm * avg  

                    val = 0 if akbm == 0 else round(sumv / akbm)
                    d.append(str(val) + '%') 
                    added = True        
            
            if not added:
                d.append('0%')            
                
        xlb.makeCells(sheet, r, d)
        r += 1
    # пїЅпїЅпїЅпїЅ
    for k, v in slsnets.items():
        d = []
        d.append(v.slsnet + ' (сеть)')
        
        for a in arr:
            added = False;
            if a in reports:
                rp = reports[a];
                
                if len(rp.pages) >= 1:
                    p = rp.pages[1]
                    slsAvg = countAvgSls(p.items)
                    avgCount = slsAvg[v.slsnet] if v.slsnet in slsAvg else '0'
                    avgVal = str(avgCount) + '%'
                    d.append(avgVal)
                    added = True
                
            if not added:
                d.append('0%')     
        
        xlb.makeCells(sheet, r, d)
        r += 1

    # пїЅпїЅпїЅпїЅпїЅпїЅпїЅпїЅ пїЅпїЅпїЅпїЅпїЅ
    for k, v in orgs.items():
        d = []
        d.append(v.org + ' ('+ v.address+ ')')
        
        for a in arr:
            added = False;
            
            if a in reports:
                rp = reports[a];
                
                if len(rp.pages) >= 1:
                    p = rp.pages[1]
                    
                    for i in p.items:
                        if i.id == k:
                            a = str(i.avgMePrc()) + '%'
                            d.append(a);
                            added = True
                            break;
            
            if not added:
                d.append('0%') 
                
        xlb.makeCells(sheet, r, d)
        r += 1
    
    for c in range(len(head)):
        col_letter = get_column_letter(c+1)
        v = 35 if c == 0 else 14 
        sheet.column_dimensions[col_letter].width = v
    
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    server.Put(outObj)
    
class RRPV4:
    # __slots__ = ["agvItems", "avgMePrc", "planFace", "avgGroup"]
    
    def __init__(self):
        self.avgItems = 0.0
        self.avgMePrc = 0.0
        self.planFace = 0.0
        self.avgGroup = 0.0
         
def printReportV4ReportPage(sheet, title, page):
    sheet.title = title
    
    i = 1
    r = 0

    sheet.merge_cells(start_row=r, start_column=0, end_row=r+1, end_column=0)    
    setVal(sheet.cell(row=r, column=0), "Город")
    col_letter = get_column_letter(1) 
    sheet.column_dimensions[col_letter].width = 20
    sheet.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
    setVal(sheet.cell(row=r, column=1), "Мерчендайзер")
    col_letter = get_column_letter(2) 
    sheet.column_dimensions[col_letter].width = 20
    sheet.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
    setVal(sheet.cell(row=r, column=2), "Наименование торговой сети")
    col_letter = get_column_letter(3) 
    sheet.column_dimensions[col_letter].width = 20
    sheet.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
    setVal(sheet.cell(row=r, column=3), "Наименование торговой точки")
    col_letter = get_column_letter(4) 
    sheet.column_dimensions[col_letter].width = 20
    sheet.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)
    setVal(sheet.cell(row=r, column=4), "Адрес")
    col_letter = get_column_letter(5) 
    sheet.column_dimensions[col_letter].width = 20
    sheet.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
    setVal(sheet.cell(row=r, column=5), "Дата заполнения отчета")
    col_letter = get_column_letter(6) 
    sheet.column_dimensions[col_letter].width = 20

    group = None
    START_COL = 6
    c = START_COL
    pc = c
    isMy = False
    
    for p in page.price:
        if group != None and group != p.group: 
            sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
            setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
            pc = c
            
        group = p.group
        isMy = p.my > 0
        
        setVal(sheet.cell(row=r + 1, column=c), p.name, vrt = Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        
    if c - 1 >= pc:    
        sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
        setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
        
    sheet.row_dimensions[2].height = 95
    
    page.collumns(sheet, r, c)
    
    r = r + 2
    dt = None
    sr = r + 1
    dsr = sr
    
    sum_group = dict()
    sum_result = dict()
    last_date = None
    
    slsAvg = countAvgSls(page.items)
    
    lastCol = 16
    
    res = dict() # userid-orgid-data
    for i in page.items:
        if dt == None:
            dt = i.date
               
        setVal(sheet.cell(row=r, column=0), i.city, hrz= Alignment.HORIZONTAL_LEFT)
        setVal(sheet.cell(row=r, column=1), i.user, hrz= Alignment.HORIZONTAL_LEFT)
        setVal(sheet.cell(row=r, column=2), i.slsnet, hrz= Alignment.HORIZONTAL_LEFT)
        setVal(sheet.cell(row=r, column=3), i.org, hrz= Alignment.HORIZONTAL_LEFT)
        setVal(sheet.cell(row=r, column=4), i.address, hrz= Alignment.HORIZONTAL_LEFT)
        setVal(sheet.cell(row=r, column=5), i.created, hrz= Alignment.HORIZONTAL_RIGHT)
        
        c = START_COL;
        last_date = i.created
        
        for p in page.price:
            if p.id in i.data:
                setVal(sheet.cell(row=r, column=c), i.data[p.id])
                
                if not p.id in sum_group:
                    sum_group[p.id] = i.data[p.id]
                else:     
                    sum_group[p.id] = sum_group[p.id] + i.data[p.id]  
                
            c = c + 1

        if not i.userid in res:
            res[i.userid] = dict()
            
        resuser = res[i.userid]
        
        if not i.id in resuser:
            resuser[i.id] = RRPV4()
             
        resuser[i.id].avgItems = i.avgItems
        resuser[i.id].avgMePrc = i.avgMePrc()
        resuser[i.id].planFace = i.faceplan
        resuser[i.id].avgGroup = i.avgGroup
            
        lastCol = page.itemsdata(sheet, r, c, i)
        avgCount = slsAvg[i.slsnet] if i.slsnet in slsAvg else 0
        
        if avgCount > 0:
            avgVal = str(avgCount) + '%'
            setVal(sheet.cell(row=r, column=lastCol + page.factSlsIdx), avgVal, hrz= Alignment.HORIZONTAL_LEFT)
        
        r = r + 1
    
    rangeBorders(sheet.range("A1:"+get_column_letter(lastCol)+str(r)) )
    
    return res

class ReportV4(Report):
    def __init__(self, server, start, finish, pageType):
      Report.__init__(self, server, start, finish, pageType)
      
    def doResultPage(self, ap, avgFactAgent):
        pass
    
class XLBuilder4Ex(XLBuilder):
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 2 or column == 3 or column == 4 or column == 11:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
      
class XLBuilder4ResEx(XLBuilder):
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 5 or column == 6 or column == 7 or column == 14:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
      
def doReportV4(server):
    server.RegisterType("Result[name:s,file:b,items[name:s,photo:b]]")
    outObj = server.New("Result")
    params = server.Params[0]
    f = params.finish
    s = f.replace(day=1)
    
    reportCurMonth = ReportV4(server, s, f, "PageV4")
    
    where = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}") and "userid" in ({2})'.format(
               s.strftime("%d/%m/%Y 0:0:0"), 
               s.strftime("%d/%m/%Y 23:59:59"),
               ','.join(["'%s'" % x.id for x in params.userids]))

    planFact = server.Get('VisitPlanFact', where, 'userid')
    
    f = s - timedelta(days=1)
    s = f.replace(day=1)
    
    reportPrevMonth = ReportV4(server, s, f, "PageV4")
    
    wb = Workbook(False, 'cp1251')
    
    d1 = printReportV4ReportPage(wb.get_active_sheet(), "Отчет по ДП(Прошлый месяц)", reportPrevMonth.pages[1])
    d2 = printReportV4ReportPage(wb.create_sheet(), "Отчет по ДП(Текущий месяц)", reportCurMonth.pages[1])
    
    asf = dict(reportPrevMonth.agentPages)
    asf.update(reportCurMonth.agentPages)
    
    apl = list()
    apl.extend(asf.values())
    apl = sorted(apl, key=lambda x: x.agent)
    
    xlb = XLBuilder4Ex()
    head = ["Торговая точка", "Адрес", "Факт в усл. ед. прошлый месяц", "План по доле полки в усл. ед текущий месяц", 
            "Фактическая доля полки в усл. ед. текущий месяц", "Факт по торговой точке в % текущий месяц", "СРЕДНЕЕ по группе усл. ед.",
            "Количество визитов в месяц План", "Количество визтов в месяц Факт", 
            "План по фейсам в месяц", "Факт по фейсам в месяц", "Эффективность"]
    resitems = dict()
    
    res = []
    
    for ap in apl:
        sheet = wb.create_sheet()
        sheet.title = ap.agent
        xlb.makeHead(sheet,0,head,True)
        cc = 1
        for w in [20,20,20,20,20,20]:
            sheet.column_dimensions[get_column_letter(cc)].width = w
            cc += 1
        
        items = ap.items.values()
        items = sorted(items, key=lambda x: x.org)
        
        uid = "'" + ap.userid + "'"
        server.ChangeUser(uid)
        route = server.Get("OrgFolder", "", "name")
        server.RestoreUser()

        agentPlanFact = planFact[ap.userid] if ap.userid in planFact else None
        
        r = 1
        sz = len(items)+2
        
        avgVisitEff = 0.0
        visitEffCount = 0
        
        for d in items:
            resRow = ["",d.responsible,ap.agent]
            sz = len(resRow)
            data = d.getData(r)
            resRow.extend(data)
            xlb.makeCells(sheet, r, data)
            
            avg = 0
            
            facePlan = reportCurMonth.getVisitFacePlan(d.id)
            factVisits = reportCurMonth.getVisitInRoute(ap.userid, d.id, route)
            planVisits = 0
            if agentPlanFact != None:
                for apfi in agentPlanFact.items:
                    if apfi.id == d.id:
                        planVisits = apfi.plan - apfi.miss
                        break 
            
            if ap.userid in d1 and d.id in d1[ap.userid]:
                avg = d1[ap.userid][d.id].avgItems
                
            c = 2    
            setVal(sheet.cell(row=r, column=c), avg, hrz= Alignment.HORIZONTAL_RIGHT)
            resRow[c+sz] = avg
            
            planFace = 0
            
            if ap.userid in d2 and d.id in d2[ap.userid]:
                planFace = d2[ap.userid][d.id].planFace
            
            #если в текущем месяце не было данных берем данные о плане из прошлого месяца
            if planFace == 0 and ap.userid in d1 and d.id in d1[ap.userid]:
                planFace = d1[ap.userid][d.id].planFace    
            
            c = 3
            setVal(sheet.cell(row=r, column=c), planFace, hrz= Alignment.HORIZONTAL_RIGHT)
            resRow[c+sz] = planFace
            
            avg = 0
            
            if ap.userid in d2 and d.id in d2[ap.userid]:
                avg = d2[ap.userid][d.id].avgItems
            
            c = 4
            setVal(sheet.cell(row=r, column=c), avg, hrz= Alignment.HORIZONTAL_RIGHT)    
            resRow[c+sz] = avg
            
            avgMePrc = 0
            
            if ap.userid in d2 and d.id in d2[ap.userid]:
                avgMePrc = d2[ap.userid][d.id].avgMePrc
            
            c = 5            
            strAvgMePrc = str(avgMePrc) + '%'
            setVal(sheet.cell(row=r, column=c), strAvgMePrc, hrz= Alignment.HORIZONTAL_RIGHT)
            resRow[c+sz] = strAvgMePrc
            
            avgGroup = 0
            
            if ap.userid in d2 and d.id in d2[ap.userid]:
                avgGroup = d2[ap.userid][d.id].avgGroup
            
            c = 6            
            setVal(sheet.cell(row=r, column=c), avgGroup, hrz= Alignment.HORIZONTAL_RIGHT)
            resRow[c+sz] = avgGroup

            c += 1            
            ccell = sheet.cell(row=r, column=c)
            setVal(ccell, planVisits, hrz= Alignment.HORIZONTAL_RIGHT)
            xlb.makeBorder(ccell, Border.BORDER_THIN)
            resRow.append(planVisits)

            c += 1            
            ccell = sheet.cell(row=r, column=c)
            setVal(ccell, factVisits, hrz= Alignment.HORIZONTAL_RIGHT)
            xlb.makeBorder(ccell, Border.BORDER_THIN)
            resRow.append(factVisits)
            
            c += 1            
            ccell = sheet.cell(row=r, column=c)
            fp = facePlan * planVisits
            setVal(ccell, fp, hrz= Alignment.HORIZONTAL_RIGHT)
            xlb.makeBorder(ccell, Border.BORDER_THIN)
            resRow.append(fp)

            c += 1            
            ccell = sheet.cell(row=r, column=c)
            fc = facePlan * factVisits
            setVal(ccell, fc, hrz= Alignment.HORIZONTAL_RIGHT)
            xlb.makeBorder(ccell, Border.BORDER_THIN)
            resRow.append(fc)
            
            c += 1            
            eff = '' if planVisits == 0 else float(factVisits) / planVisits * 100.0
            
            if eff != '':
                avgVisitEff += eff
                visitEffCount += 1
                
            resRow.append(eff)
            
            ccell = sheet.cell(row=r, column=c)
            setVal(ccell, eff, hrz= Alignment.HORIZONTAL_RIGHT)
            ccell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            xlb.makeBorder(ccell, Border.BORDER_THIN)
            
            res.append(resRow)
            r += 1
       
        avgValue = '' if visitEffCount == 0 else avgVisitEff / visitEffCount
        ccell = sheet.cell(row=r, column=c)
        setVal(ccell, avgValue, hrz= Alignment.HORIZONTAL_RIGHT)
        ccell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
        xlb.makeBorder(ccell, Border.BORDER_THIN)
    
    xlb = XLBuilder4ResEx()
    sheet = wb.create_sheet()
    sheet.title = "ИТОГО"
    resHead = ["ТС","СВ","МЧ"]
    resHead.extend(head)
    xlb.makeHead(sheet,0,resHead,True)      
    cc = 1
    
    for w in [20,20,20,20,20,20,20,20,20,20,15,15,15,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1    
    
    r = 1    
    
    for d in res:
      xlb.makeCells(sheet, r, d)
      r += 1
    
    ccell = sheet.cell(row=r, column=14)
    setVal(ccell, "=IFERROR(AVERAGE(O2:O{0}),0)".format(r), hrz= Alignment.HORIZONTAL_RIGHT)
    ccell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
    xlb.makeBorder(ccell, Border.BORDER_THIN)
    
    
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    vv = list()
    vv.extend(reportPrevMonth.visit)
    vv.extend(reportCurMonth.visit)
    
    for v in vv:
        idx = 0
        if v.items != None:
            for vi in v.items:
                item = obj.items.New()
                item.name = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                item.photo = vi.id
                idx = idx + 1
    
    server.Put(outObj)
                                      
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    version = params.version
    logging.info("params " + str(params))
     
    if version == 1 or version == 2 :
        doReportV12(server)
    elif version ==3:   
        doReportV3(server)
    elif version == 4:
        doReportV4(server)    
    
    logging.info('end')
    