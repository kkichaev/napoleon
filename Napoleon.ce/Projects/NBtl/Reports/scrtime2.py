# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
import logging
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from imp import reload
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from manager import summary

import tempfile
import sys;

reload(sys)

bkgColor = "ff90ffff"

class Report:
    items = None
    start = None
    finish = None
    
    def __init__(self, server):
        self.start, self.finish, dummy = inflateParams(server)
        self.items = list()
        self.finish -= timedelta(days=1)
        
class Item:
    divid = None
    data = None
    finish = None
    inside = None
    outside = None
    
    def __init__(self):
      self.data = datetime.now()
      self.finish = datetime.now()
      self.inside = timedelta()
      self.outside = timedelta()
      self.divid = ""

class DivData:
    div = None
    inside = None
    outside = None
    agents = None
    title = None
    
    def __init__(self):
        self.div = None
        self.inside = timedelta()
        self.outside = timedelta()
        self.agents = list()
        self.title = ""
        
    def getData(self, row):
        return [self.title, timeDeltaToStr(self.inside), timeDeltaToStr(self.outside)]
        
def timeDeltaToStr(val):
    hours, remainder = divmod(val.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
        
    return "{:02d}:{:02d}:{:02d}".format(int(hours), int(minutes), int(seconds))  
        
def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].divid

def insideParent(val, id, m):
    if id != -1 and id in m:
        insideParent(val, m[id].div.parent, m)
    
    if id in m:
        m[id].inside += val 

def outsideParent(val, id, m):
    if id != -1 and id in m:
        outsideParent(val, m[id].div.parent, m)
    
    if id in m:
        m[id].outside += val
                
def loadData(server):
    start, finish, div = inflateParams(server)
    
    rd = server.Get("Division", '"id"={0}'.format(div))
    dvs = list()
    da = summary.loadAgents(server, rd, dvs)
    dvsmap = dict()
    
    for d in dvs:
        dd = DivData()
        dd.div = d
        dd.title = d.name
        dvsmap[d.id] = dd
        
    for d in dvs:
        for a in d.agents:
            WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
            WHERE_STR = WHERE_STR_ALL + ' and "userid" in (\'{2}\')'; 
        
            where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), a)
            
            scripts = server.Get("ScriptDoc", where)
            agent = ""
            
            map = dict()    
            
            for s in scripts:
                k = s.id + s.created.strftime("%d/%m/%Y")
              
                if not k in map:
                    i = Item()
                    i.divid = d.id
                    i.data = s.created
                    map[k] = i
              
                    m = map[k]
              
                    dt = m.data
              
                if s.items != None:
                    for s in s.items:
                        if s.state == 1 and s.date > dt:
                            dt = s.date
                    
                m.finish = dt

                if m.finish.date() == m.data.date():
                    m.inside = m.finish - m.data
              
                if d.id in dvsmap:
                    dvsmap[d.id].inside += m.inside
                    insideParent(m.inside, dvsmap[d.id].div.parent, dvsmap)
      
            data = list()
    
            for v in map.values():
                data.append(v)
      
            data = sorted(data, key=lambda x:x.data)

            for i in range(0, len(data)):
                if i > 0:
                    t = timedelta()
                    
                    if data[i].data.date() == data[i-1].finish.date():
                        t = data[i].data - data[i-1].finish
                        
                    data[i].outside = t
                    
                    if data[i].divid in dvsmap:
                        dvsmap[d.id].outside += t
                        outsideParent(t, dvsmap[d.id].div.parent, dvsmap)
            
            if d.id in dvsmap:
                adv = DivData() 
                adv.title = da[a].name if a in da else a
                             
                for dt in data:
                    adv.inside += dt.inside
                    adv.outside += dt.outside 
                
                dvsmap[d.id].agents.append(adv)    
                
    ret = Report(server)
    
    for d in dvs:
        dd = dvsmap[d.id]
        ret.items.append(dd)
        
        for a in sorted(dd.agents, key=lambda x:x.title):
            ret.items.append(a)
                
    return ret
    
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    sh.cell(row=1, column=0).value = "Период: c {0} по {1}".format(data.start.strftime("%d.%m.%Y"), 
        data.finish.strftime("%d.%m.%Y"))
    
    head = ["Название", "Итого время в ТТ", "Итого время передвижения от ТТ до ТТ"]
    
    r = 2
    xlb.makeHead(sh, r, head)
    r += 1
    
    for d in data.items:
        xlb.makeCells(sh, r, d.getData(r))
        r += 1
            
    setCellWidth(sh, [30,30,30])
    
class XLBuilderEx(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment.wrap_text = True
        return column
        
    def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
        XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
        if column == 1 or column == 2:
            cell.style.number_format._set_format_code('[h]:mm:ss')
            cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
    
def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')
    
    doReport(server)

    logging.info('end')        