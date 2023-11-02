# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import datetime

import sys

from imp import reload;
reload(sys)

def loadData(params, server):
    agents = server.Get('Agents','','id')
    orgs = server.Get('CommonOrgs','','id')
    
    where = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}") and "userid"={2}'.format(
               params.date.strftime("%d/%m/%Y 0:0:0"), 
               params.date.strftime("%d/%m/%Y 23:59:59"),
               "'" + params.userid + "'")
    data = server.Get('VisitPlanFact', where)

    agent = agents[params.userid].name if params.userid in agents else params.userid             
    return (agent,orgs,data)
    
def printOut(agent, orgs, data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 1
    
    cell = sheet.cell(row=row, column=0)
    cell.value = 'Сотрудник: ' + agent

    row += 1
    cell = sheet.cell(row=row, column=0)
    cell.value = 'Период {0}'.format( params.date.strftime("%m/%Y"))
    
    row += 2
    head = ['Сеть', 'Адрес', 'План', 'Отсутствие по уважительной причине', 'План итого', 'Комментарий']    
    xlb.makeHead(sheet, row, head, True)
    
    row += 1
    for doc in data:
        for item in doc.items:
            orgName = item.id
            orgAdr = ''
            if item.id in orgs:
                org = orgs[item.id]
                orgName = org.name
                orgAdr = org.address
            
            values = [orgName,orgAdr,item.plan,item.miss,item.plan - item.miss,item.comment]

            xlb.makeCells(sheet, row, values)
            row += 1        
    
    cc = 1
    for w in [30,40,20,20,20,50]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    (agent,orgs,data) = loadData(params, server)
    wb = printOut(agent, orgs, data, params)

    XLBuilder().workbookToObject(wb, "plan_fact.xlsx", server)                
    logging.info('end')
    