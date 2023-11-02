# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import datetime


import sys

from imp import reload
reload(sys)


class ItemData :
    def __init__(self, item, groups, price, plu, docdata):
        self.categ = ''
        self.exists = 'Да' if item.exist == 1 else 'Нет'
        self.itemStatus = item.itemStatus
        self.mfr = ''
        self.plu = ''
        
        key = item.id + docdata.sid

        if key in plu:
            self.plu = plu[key]

        if item.id in price:
            prc = price[item.id]
            self.name = prc.name
            self.categ = prc.categ
            
            if prc.fid in groups:
                self.mfr = groups[prc.fid].name
        else:
            self.categ = ''
            self.name = 'Товар с кодом <' + item.id + '>'
                
class ItemData2(ItemData):
    def __init__(self, item, groups, price, plu, docdata):
        self.categ = ''
        self.exists = '0'
        self.name = '0'
        self.mfr = ''
        self.itemStatus = ''
        self.plu = ''
                
class DocData:
    def __init__(self, server, doc, agents, salesNet, groups, orgs, price, plu):        
        self.agent = agents[doc.userid].name if doc.userid in agents else 'Агент с кодом <' + doc.userid + '>'
        self.date = doc.created.date().strftime("%d.%m.%Y")
        self.slsnet = ''
        self.sid = ''
        
        if doc.id in orgs:
            org = orgs[doc.id]
            self.city = org.cid
            self.orgName = org.name
            self.address = org.address
            self.code = org.code
            self.sid = org.sid

            if org.sid in salesNet:
                self.slsnet = salesNet[org.sid].name
        else:
            self.city = ''
            self.orgName = ''
            self.address =''
            self.code='' 
        
        self.items = list()
        self.initItems(server, doc, groups, price, orgs, plu)
            
    def initItems(self, server, doc, groups, price, orgs, plu):
        for item in doc.items:
            i = ItemData(item, groups, price, plu, self)
            self.items.append(i)
                    
class DocData2(DocData):
    def initItems(self, server, doc, groups, price, orgs, plu):
        i = ItemData2(None, groups, price, plu, self)
        ex = 0
        for t in doc.items:
            if t.exist == 1:
                ex += 1 
        
        i.exists = str(ex)
        
        if doc.id in orgs:
            mn = orgs[doc.id].goodsMatrix
            gm = server.Get("GoodsMatrix",'"name"=\'{0}\''.format(mn),'name')
            
            if gm != None and mn in gm:
                i.name = str(len(gm[mn].items))
             
        self.items.append(i)

class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()
        
    def add(self, server, doc, agents, salesNet, groups, orgs, price, plu):
        data = self.createDocData(server, doc, agents, salesNet, groups, orgs, price, plu)
        self.data.append(data)
        
    def createDocData(self, server, doc, agents, salesNet, groups, orgs, price, plu):    
        return DocData(server, doc, agents, salesNet, groups, orgs, price, plu)
    
class ReportData2(ReportData):
    def createDocData(self, server, doc, agents, salesNet, groups, orgs, price, plu):    
        return DocData2(server, doc, agents, salesNet, groups, orgs, price, plu)

def loadData(params, server):
    
    agents = server.Get("Agents", "", "id")
    salesNet = server.Get('Slsnet','"id"="{0}"'.format(params.slsnet), 'id')
    groups = server.Get('GroupGoods','','id')
    orgs = server.Get('CommonOrgs', '', 'id')
    price = server.Get('Price', '"isGoods"=1', 'id')
    plu = server.Get('PLU', "")

    pluHash = {}

    for p in plu:
        key = p.item + p.sls
        pluHash[key] = p.code
    
    agentQuery = '"userid" in('
    for agent in params.userids:
        agentQuery += "'" + agent.id + "',"
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    data = ReportData2() if params.type == 1 else ReportData()
    docs = server.Get('Distrib', q)
    if docs != None:
        for d in docs:
            if d.id in orgs and orgs[d.id].sid == params.slsnet:
                data.add(server, d, agents, salesNet, groups, orgs, price, pluHash)
            
    return data
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 1
    
    head = ['Город', 'Мерчендайзер', 'Наименование торговой сети', 'Наименование торговой точки']

    if params.type == 0 or params.type == 2: head.append('Код торговой точки')

    head.extend(['Адрес', 'Категория', 'Производитель', 'Дата заполнения отчета'])
    if params.type == 2: head.append('PLU')
    head.extend(['SKU','Наличие на полке'])

    if params.type == 2: head.append('Статус товара')
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
        docValues = [item.city, item.agent, item.slsnet, item.orgName]
        if params.type == 0 or params.type == 2: docValues.append(item.code)
        docValues.append(item.address)
        for oi in item.items:
            values = []
            values.extend(docValues)
            
            values.extend([oi.categ, oi.mfr, item.date])
            
            if params.type == 2:
                values.append(oi.plu)

            values.extend([oi.name, oi.exists])
            
            if params.type == 2:
                values.append(oi.itemStatus if oi.itemStatus >= 0 else '')

            xlb.makeCells(sheet, row, values)
            row += 1        
    
    cc = 1
    collumns = [20] * len(head)
    for w in collumns:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info('end')
    