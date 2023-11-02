# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
import logging
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from imp import reload
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);

class Report:
  pass

class Data:
  def getData(self):
      res = [self.agent]
      
      for d in self.days:
        res.append(d.start.strftime("%H:%M") + " " if d.start != None else " ")
        res.append(d.finish.strftime("%H:%M") + " " if d.finish != None else " ")
      return res
      
class DayData():
    def __init__(self):
      self.start = None
      self.finish = None
      
    def __str__(self):
      return 'start: {0}, finish: {1}'.format(self.start.strftime("%H:%M") if self.start != None else " ",
        self.finish.strftime("%H:%M") if self.finish != None else " ")
    
class Item:
    def __init__(self):
      self.data = datetime.now()
      self.finish = datetime.now()
      
    def getData(self):
      return [self.agent]
      
def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, ','.join(["'%s'" % x.id for x in server.Params[0].userids])
    
def loadData(server):
    start, finish, userid = inflateParams(server)
    
    report = Report()
    report.data = {}
    report.start = start
    report.finish = finish
    report.days = []
    
    s = start
    
    while s < finish:
      d = DayData()
      d.date = s.date()
      report.days.append(d)
      s += timedelta(days=1)
    
    WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
    WHERE_STR = WHERE_STR_ALL + ' and "userid" in ({2})'; 

    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
    scripts = server.Get("ScriptDoc", where)
    
    agent = ""
    
    usersOrg = {}
    agents = {}
    
    for uid in userid.split(','):
      server.ChangeUser(uid)
      agent = server.CurrentUser().name
      agentid = server.CurrentUser().id
      agents[agentid] = agent
      usersOrg[agentid] = server.Get("Org", "", "id")
      porg = server.Get("PotenzialOrg", "", "id")
      usersOrg[agentid].update(porg)
      server.RestoreUser()
      report.data[agentid] = Data()
      report.data[agentid].agent = agent
      report.data[agentid].days = []

      for d in range(0,len(report.days)):
        report.data[agentid].days.append(DayData())
        
      report.data[agentid].items = {}

    data = [] 
    
    orgmap = {}
    
    for s in scripts:
      d = Item()
      
      if not s.userid in orgmap:
        orgmap[s.userid] = {}
        
      if not s.created.date() in orgmap[s.userid]:
        orgmap[s.userid][s.created.date()] = 0
      
      orgmap[s.userid][s.created.date()] += 1
      
      d.org = orgmap[s.userid][s.created.date()]
      d.address = usersOrg[s.userid][s.id].address if s.id in usersOrg[s.userid] else ""
      d.data = s.created
      d.userid = s.userid
      d.agent = agents[s.userid] if s.userid in agents else ""
      dt = d.data
      
      if s.items != None:
        for s in s.items:
          if s.state == 1 and s.date > dt and s.date.date() == dt.date():
            dt = s.date
            
      d.finish = dt
      
      data.append(d)
      
    data = sorted(data, key=lambda x:(x.agent,x.data))
    
    for d in data:
      idx = (d.data - start).days
      dd = report.data[d.userid].days[idx]
      
      if dd.start == None or d.data < dd.start:
        dd.start = d.data
      
      if dd.finish == None or d.finish > dd.finish:
        dd.finish = d.finish
      
      if not d.org in report.data[d.userid].items:
        report.data[d.userid].items[d.org] = []
        
        for dx in range(0,len(report.days)):
          report.data[d.userid].items[d.org].append(DayData())
      
      orgitem = report.data[d.userid].items[d.org][idx]
      
      if orgitem.start == None or d.data < orgitem.start:
        orgitem.start = d.data
      
      if orgitem.finish == None or d.finish > orgitem.finish:
        orgitem.finish = d.finish
        
    return report, agent
      
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintFisrtSheet(xlb, sh, report, agent):
    head = ["Мерчендайзер"]
    r = 0
    xlb.makeHead(sh, r, head)
    
    col = len(head)
    dn = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']
    
    for d in report.days:
      sh.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
      cell = sh.cell(row=r, column=col)
      cell.value = "{0} {1}".format(dn[d.date.weekday()], d.date.strftime("%d.%m"))
      cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
      sh.cell(row=r+1, column=col).value = 'начало р.д'
      sh.cell(row=r+1, column=col+1).value = 'конец р.д'
      col += 2
      
    for c in range(0, len(report.days)*2 + 1):
      for r in range(0,2):
        cel = sh.cell(row=r, column=c)
        cel.style.font.bold = True
        xlb.makeBorder(cel, xlb.HEAD_BORDER_STYLE)
    
    r = 1
    
    for d in sorted(report.data.values(), key=lambda x: x.agent):
      r+=1
      xlb.makeCells(sh, r, d.getData())
    
    arr = [30]
    
    for a in range(0, len(report.days) * 2):
      arr.append(12)
      
    setCellWidth(sh, arr)
    
def printSecondSheet(xlb, sh, report, agent):
    head = ["Мерчендайзер"]
    r = 0
    xlb.makeHead(sh, r, head)
    
    col = len(head)
    dn = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']
    
    for d in report.days:
      sh.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
      cell = sh.cell(row=r, column=col)
      cell.value = "{0} {1}".format(dn[d.date.weekday()], d.date.strftime("%d.%m"))
      cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
      sh.cell(row=r+1, column=col).value = 'начало р.д'
      sh.cell(row=r+1, column=col+1).value = 'конец р.д'
      col += 2
      
    for c in range(0, len(report.days)*2 + 1):
      for r in range(0,2):
        cel = sh.cell(row=r, column=c)
        cel.style.font.bold = True
        xlb.makeBorder(cel, xlb.HEAD_BORDER_STYLE)
    
    r = 1
    
    for d in sorted(report.data.values(), key = lambda x: x.agent):
      r+=1
      xlb.makeCells(sh, r, d.getData())
      
      for org in sorted(d.items, key = lambda x:x):
        list = d.items[org]
        r += 1
        arr = ["Точка {0}".format(org)]
        
        for dd in list:
          arr.append(dd.start.strftime("%H:%M") + " " if dd.start != None else " ")
          arr.append(dd.finish.strftime("%H:%M") + " " if dd.finish != None else " ")
        
        xlb.makeCells(sh, r, arr)
        cel = sh.cell(row=r, column=0)
        cel.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
    
    arr = [30]
    
    for a in range(0, len(report.days) * 2):
      arr.append(12)
      
    setCellWidth(sh, arr)    
    
def printOut(d, a):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilder()
    ptintFisrtSheet(xlb, sh, d, a)
    
    sh = wb.create_sheet()
    printSecondSheet(xlb, sh, d, a)
    
    return wb

def doReport(server):
    data, agent = loadData(server)
    wb = printOut(data, agent)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    doReport(server)

    logging.info('end')                