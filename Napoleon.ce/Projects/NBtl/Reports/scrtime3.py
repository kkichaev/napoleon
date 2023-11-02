# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
import logging
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from imp import reload
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from contract import ReportV4, countAvgSls, RRPV4, Report

import tempfile
import sys;

reload(sys)

bkgColor = "ff90ffff"

class Item:
    data = None
    org = None
    address = None
    finish = None
    orgid = None
    userid = None
    
    def __init__(self):
      self.org = ""
      self.address = ""
      self.data = datetime.now()
      self.finish = datetime.now()
      self.orgid = ""
      self.userid = ""
      
    def getData(self, row, isLastInDay):
      return [self.data.strftime("%d.%m.%Y"), self.org, self.address, 
              self.data.strftime("%H:%M:%S") if self.data != None else "" , 
              self.finish.strftime("%H:%M:%S") if self.finish != None else "",
              "=E{0}-D{0}".format(row+1), 
              "=D{0}-E{1}".format(row+2, row+1) if not isLastInDay else "" ]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, "'" + server.Params[0].userids[0].id + "'"
    
def loadData(server):
    start, finish, userid = inflateParams(server)
    
    WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
    WHERE_STR = WHERE_STR_ALL + ' and "userid" in ({2})'; 

    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
    
    scripts = server.Get("ScriptDoc", where)
    agent = ""
    
    server.ChangeUser(userid)
    agent = server.CurrentUser().name
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    server.RestoreUser()

    map = dict()    
    
    for s in scripts:
      k = s.id + s.created.strftime("%d/%m/%Y")
      
      if not k in map:
        d = Item()
        d.org = orgs[s.id].name if s.id in orgs else "Контрагент с кодом <{0}>".format(s.id)
        d.address = orgs[s.id].address if s.id in orgs else ""
        d.data = s.created
        d.orgid = s.id
        d.userid = s.userid
        map[k] = d
      
      m = map[k]
      
      dt = m.data
      
      if s.items != None:
        for s in s.items:
          if s.state == 1 and s.date > dt:
            dt = s.date
            
      m.finish = dt
      
    data = list()
    
    for v in map.values():
      data.append(v)
      
    data = sorted(data, key=lambda x: x.data)
    
    
    f = server.Params[0].finish
    s = f.replace(day=1)
    
    reportCurMonth = ReportV4(server, s, f, "PageV4")
    d2 = calcReportV4ReportPage(reportCurMonth.pages[1])
    
    return data, agent, d2
    
def calcReportV4ReportPage(page):
    res = dict() # userid-orgid-data
    dt = None
    sum_group = dict()
    sum_result = dict()
    last_date = None
    slsAvg = countAvgSls(page.items)
    
    for i in page.items:
        if dt == None:
            dt = i.date

            last_date = i.created
        
        for p in page.price:
            if p.id in i.data:
                if not p.id in sum_group:
                    sum_group[p.id] = i.data[p.id]
                else:     
                    sum_group[p.id] = sum_group[p.id] + i.data[p.id]  
                
        if not i.userid in res:
            res[i.userid] = dict()
            
        resuser = res[i.userid]
        
        if not i.id in resuser:
            resuser[i.id] = RRPV4()
             
        resuser[i.id].avgItems = i.avgItems
        resuser[i.id].avgMePrc = i.avgMePrc()
        resuser[i.id].planFace = i.faceplan
        resuser[i.id].avgGroup = i.avgGroup
            
        avgCount = slsAvg[i.slsnet] if i.slsnet in slsAvg else 0
        
        if avgCount > 0:
            avgVal = str(avgCount) + '%'
        
    return res
    
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def addExtend(row, d, exdata):
    planFace = 0
                  
    if d.userid in exdata and d.orgid in exdata[d.userid]:
      planFace = exdata[d.userid][d.orgid].planFace
        
    row.append(planFace)
    
    avg = 0
    
    if d.userid in exdata and d.orgid in exdata[d.userid]:
      avg = exdata[d.userid][d.orgid].avgItems
      
    row.append(avg)

    avgMePrc = 0
        
    if d.userid in exdata and d.orgid in exdata[d.userid]:
      avgMePrc = exdata[d.userid][d.orgid].avgMePrc
    
    row.append(avgMePrc)
    
    avgGroup = 0
        
    if d.userid in exdata and d.orgid in exdata[d.userid]:
      avgGroup = exdata[d.userid][d.orgid].avgGroup
    
    row.append(avgGroup)
    
def ptintSheet(xlb, sh, data, agent, exdata):
    sh.cell(row=0, column=0).value = agent
    sh.cell(row=1, column=0).value = "Дата: " + datetime.now().strftime("%d.%m.%Y")
    
    head = ["Дата", "Название ТТ", "Адрес ТТ", "Время начала визита в ТТ", "Время окончания визита в ТТ", "Итого время в ТТ", "Итого время передвижения от ТТ до ТТ",
      "План по доле полки в усл. ед текущий месяц", "Фактическая доля полки в усл. ед. текущий месяц", "Факт по торговой точке в % текущий месяц", "СРЕДНЕЕ по группе усл. ед."]
    
    r = 2
    xlb.makeHead(sh, r, head)
    
    svData = None
    for d in data:
        if svData != None:
            r += 1
            row = svData.getData(r, svData.data.date() != d.data.date())
            addExtend(row,d,exdata)
            xlb.makeCells(sh, r, row)
        svData = d
            
    if svData != None:
        r += 1
        row = svData.getData(r, False)
        addExtend(row,d,exdata)
        xlb.makeCells(sh, r, row)
    
    sh.cell(row=r, column=6).value = ""
    
    if r > 2:
      r += 1
      sh.cell(row=r, column=5).value = "=SUM(F2:F{0})".format(r)
      sh.cell(row=r, column=5).style.number_format._set_format_code('[h]:mm:ss')
      sh.cell(row=r, column=6).value = "=SUM(G2:G{0})".format(r)
      sh.cell(row=r, column=6).style.number_format._set_format_code('[h]:mm:ss')
    
    setCellWidth(sh, [30,30,30,30,30,30,30,30,30,30,30])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 5 or column == 6:
      cell.style.number_format._set_format_code('[h]:mm:ss')
      
    
def printOut(d, a, d2):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d, a, d2)
                
    return wb

def doReport(server):
    data, agent, data2 = loadData(server)
    wb = printOut(data, agent, data2)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    doReport(server)

    logging.info('end')            