from importlib import reload
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

import datetime
from datetime import timedelta


import sys;
reload(sys)

class ReportData:
  pass

class ReportItem:
  def __init__(self) -> None:
    self.addresses = {}
    self.heads = set()
    self.items = [] 

  def append(self, d):
    key = "{0}#{1}".format(d.city,d.address)
    if not key in self.addresses:
      self.addresses[key] = {}

    date = d.created.strftime('%d.%m.%Y')
    self.heads.add(date)
    self.addresses[key][date] = d.items

    if len(self.items) == 0:
      for i in d.items:
        self.items.append(i.id)

    self.items = sorted(self.items)

  def toData(self):
    ret = []

    def getVal(list, key):
      for i in list:
        if i.id == key:
          return i.answer
        
      return ''  
    
    for a in self.addresses:
      for i in self.items:
        row = []
        adr = a.split('#')
        row.append(adr[0])
        row.append(adr[1])
        row.append(i)

        for h in self.heads:
          val = ''
          for d in self.addresses[a].keys():
            if d == h:
              val = getVal(self.addresses[a][d], i)
              break

          row.append(val)
        ret.append(row)
      ret.append([])
    return ret  

def upackQuests(quests):
  res = ''
  for i in quests:
      res += "{},".format(i.id)
  return res[:-1]

def loadData(data,p,s):
  stmt = '''
    select a.id as orgid, a.created, a.userid, i.answer, i.id,  o.cid as city, o.address, o.name as org, s.name 
      from answer as a
      join answer$items as i on a.created=i.[Answer$created] and a.userid=i.[Answer$userid]
      left join org as o on a.id=o.id
      left join slsnet as s on s.id=o.sid 
      where question in ({0}) and i.type=4
      and a.created >= ToDate("{1}") and a.created < ToDate("{2}")

      order by a.id
    '''.format(
        upackQuests(p.quests),
        p.start.strftime('%d.%m.%Y'),
        (p.finish + datetime.timedelta(days=1)).strftime('%d.%m.%Y'))
  
  data.sls = {}

  docs = s.Query(
    stmt, 'Docs[orgid:s, userid:s,id:s,created:dt,name:s,city:s,address:s,org:s,items(userid,created)[id:s,answer:s]]')
  
  for d in docs:
    if not d.name in data.sls:
      data.sls[d.name] = ReportItem()

    data.sls[d.name].append(d)

  return data

def printOut(d,p):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()

  for s in d.sls:
    sheet.title = s
    head = ['город', 'адрес', 'наименование']
    head.extend(d.sls[s].heads)
    xlb = XLBuilder()
    xlb.makeHead(sheet, 0, head, True)

    row = 1
    for r in d.sls[s].toData():
      xlb.makeCells(sheet, row, r)
      row += 1

    cc = 1
    for w in range(0, len(head)):
      sheet.column_dimensions[get_column_letter(cc)].width = 40 if w < 2 else 10
      cc += 1

    sheet = wb.create_sheet()

  return wb

def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(ReportData(), params, server)
  wb = printOut(data, params)

  XLBuilder().workbookToObject(wb, "quest_rep_3.xlsx", server)                
  logging.info('end')