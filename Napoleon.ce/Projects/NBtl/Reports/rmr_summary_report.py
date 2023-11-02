# -*- coding: cp1251 -*-

import datetime
from importlib import reload
import sys
import logging
from grsoft.route import AgentRoute
from grsoft.xl_base import XLBuilder

from manager import summary
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from rmr_report_style import XLBuilderCommon
from datetime import timedelta
from openpyxl.style import Border

class AgentData:
    __slots__ = ['route', 'orgs_id_inroute', 'orgs_days', 'agent_name', 'orgs_days_count', 'orgs_days_count_fact']
    days_short = {"Понедельник" : "пн", "Вторник" : "вт", "Среда" : "ср", "Четверг" : "чт", "Пятница" : "пт", "Суббота" : "сб", "Воскресенье" : "вс"}
    days_short_idx = {"пн" : 0, "вт" : 1, "ср" : 2, "чт" : 3, "пт" : 4, "сб" : 5, "вс" : 6}
    
    def __init__(self, route):
      self.orgs_id_inroute = []
      self.route = route
      self.orgs_days = {}
      self.orgs_days_count = {}
      
      for f in self.route.route.values():
        for o in f.items:
          if not o.name in self.orgs_id_inroute:
            self.orgs_id_inroute.append(o.name)
            
          if not o.name in self.orgs_days:
            self.orgs_days[o.name] = []
          
          if f.name[0].isdigit():
            f.name = f.name[1:]
          
          if not self.days_short[f.name] in self.orgs_days[o.name]:  
            self.orgs_days[o.name].append(self.days_short[f.name])
        
    def orgRouteDays(self, id):
      ret = ""
      
      if id in self.orgs_days:
        self.orgs_days[id].sort(key=lambda x: self.days_short_idx[x])
        for d in self.orgs_days[id]:
          if len(ret) > 0:
            ret += ', '
            
          ret += d
          
      return ret
      
    def orgRouteDaysCount(self, id, start, finish):
      ret = [0,0,0,0,0,0,0]
      
      if not id in self.orgs_days_count:
        while start <= finish:
          idx = start.weekday()
          sh =  self.days_short[self.route.days[idx]]
          
          if id in self.orgs_days and sh in self.orgs_days[id]:
            ret[idx] += 1
          
          start += datetime.timedelta(days=1)
          
      return ret    
              
    def inRoute(self, id):
      return id in self.orgs_id_inroute
        
class Item:
  def __init__(self):
    self.fact_visit = []
  
  def getData(self):
    ret = [self.sls, self.org, self.address, self.orgroutedays, self.agent_name, self.city]
    
    cc = 0
    
    for c in self.org_days_count:
      if c == 0:
        ret.append('')
      else:
        ret.append(c)
        cc += c
        
    ret.append(cc)
    
    cc = 0
    fact = [0,0,0,0,0,0,0]
    
    for d in self.fact_visit:
      fact[d.weekday()] += 1
      
    for c in fact:
      if c == 0:
        ret.append('')
      else:
        ret.append(c)
        cc += c  
        
    ret.append(cc)    
    
    return ret    
    
  def setVisist(self, doc, dayRoute):
    list = []
    
    for o in dayRoute:
      list.append(o.id)
      
    if doc.id in list and not doc.created.date() in self.fact_visit:
      self.fact_visit.append(doc.created.date())
      
  
class ReportData:
    __slots__ = ['data', 'server', 'agents', 'items', 'slsnet', 'cites', 'start', 'finish']
    
    def __init__(self, server):
        self.data = dict()
        self.server = server
        self.agents = server.Get('Agents', '', 'id')
        self.items = {}
        self.slsnet = server.Get('Slsnet', '', 'id')
        self.start = server.Params[0].start
        self.finish = server.Params[0].finish
    
    def init(self, userid):
        self.data[userid] = AgentData(AgentRoute(self.server, userid))
        self.server.ChangeUser("'" + userid + "'")
        self.data[userid].agent_name = self.server.CurrentUser().name
        self.server.RestoreUser()

      
    def add(self, doc):
        userid = doc.userid
        
        if self.data[userid].inRoute(doc.id) and doc.id in self.data[userid].route.orgs:
          agentData = self.data[userid]
          
          if not doc.id in self.items:
            item = Item()
            item.sls = self.slsnet[self.data[userid].route.orgs[doc.id].sid].name if self.data[userid].route.orgs[doc.id].sid in self.slsnet else ''
            item.org = agentData.route.orgs[doc.id].name
            item.address = agentData.route.orgs[doc.id].address
            item.orgroutedays = agentData.orgRouteDays(doc.id)
            item.agent_name = agentData.agent_name
            item.city = agentData.route.orgs[doc.id].cid
            item.org_days_count = agentData.orgRouteDaysCount(doc.id, self.start, self.finish)
            self.items[doc.id] = item
          
          item = self.items[doc.id]
          item.setVisist(doc, agentData.route.getDayRoute(doc.created))

def loadData(params, server):
    data = ReportData(server)
    
    agentQuery = '"userid" in('
    
    for agent in params.userids:
        agentQuery += "'" + agent.id + "',";
        data.init(agent.id)
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    docs = server.Get('ScriptDoc', q)
    
    if docs != None:
        for d in docs:
            data.add(d)
            
    return data
  
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 1
    
    cell = sheet.cell(row=row, column=0)
    cell.value = 'Дата с {0} по {1}'.format( params.start.strftime("%d/%m/%Y"), params.finish.strftime("%d/%m/%Y"))
    row += 2
    
    head = ['Торговая сеть', 'Название', 'Адрес', 'Дни посещений', 'Фио мерчендайзера', 'Город', 'пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс', 'Итого посещений в неделю',
      'пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс', 'По маршруту']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1
    
    for i in data.items.values():
      xlb.makeCells(sheet, row, i.getData())
      row += 1
      
    cc = 1
    for w in [20,20,20,20,20,20,3,3,3,3,3,3,3,15,3,3,3,3,3,3,3,15]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "vist_rep.xlsx", server)                
    logging.info('end')
    