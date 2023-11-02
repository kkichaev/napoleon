# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from imp import reload
from openpyxl.style import Color, Fill, Border, Alignment

import datetime


import sys;
reload(sys)


class ItemData :
    __slots__ = ['categ', 'mfr', 'name', 'qty', 'qtyFlag', 'qtyFact', 'factFlag', 'qtyInput', 'qtyUnload', 'code']
    
    def __init__(self, item, groups, goodValues, price):
        self.qty = item.qty
        self.mfr = ''
        self.qtyFlag = 0
        self.factFlag = 0
        self.qtyFact = item.qtyFact
        self.qtyInput = item.qtyInput
        self.qtyUnload = item.qtyUnload
        self.code = ''
        
        for qvi in goodValues:
            if qvi.id == item.id:
                if self.qty <= qvi.redLine: self.qtyFlag = -1
                elif self.qty >= qvi.greenLine: self.qtyFlag = 1

                if self.qtyFact <= qvi.redLine: self.factFlag = -1
                elif self.qtyFact >= qvi.greenLine: self.factFlag = 1
                break 
        
        
        if item.id in price:
            prc = price[item.id]
            self.name = prc.name
            self.categ = prc.categ
            self.code = prc.code
            
            if prc.fid in groups:
                self.mfr = groups[prc.fid].name
        else:
            self.categ = ''
            self.name = 'Товар с кодом <' + item.id + '>'
                
        
class DocData:
    __slots__ = ['city', 'agent', 'slsnet', 'orgName', 'address', 'date', 'items', 'code']
    
    def __init__(self, doc, agents, salesNet, groups, goodValues, orgs, price):        
        self.agent = agents[doc.userid].name if doc.userid in agents else 'Агент с кодом <' + doc.userid + '>'
        self.date = doc.created.date().strftime("%d.%m.%Y")
        self.slsnet = ''
        self.code = ''
        
        if doc.id in orgs:
            org = orgs[doc.id]
            self.city = org.cid
            self.orgName = org.name
            self.address = org.address
            self.code = org.code
            
            if org.sid in salesNet:
                self.slsnet = salesNet[org.sid].name
        else:
            self.city = ''
            self.orgName = ''
            self.address ='' 
        
        self.items = list()
        
        gvals = goodValues[doc.id].items if doc.id in goodValues else list() 
        for item in doc.items:
            i = ItemData(item, groups, gvals, price)
            self.items.append(i)

class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()
        
    def add(self, doc, agents, salesNet, groups, goodValues, orgs, price):
        data = DocData(doc, agents, salesNet, groups, goodValues, orgs, price)
        self.data.append(data)

def loadData(params, server):
    
    agents = server.Get("Agents", "", "id")
    salesNet = server.Get('Slsnet','', 'id')
    groups = server.Get('GroupGoods','','id')
    goodValues = server.Get('GoodsValues','','id')
    orgs = server.Get('CommonOrgs', '', 'id')
    price = server.Get('Price', '"isGoods"=1', 'id')
    
    agentQuery = '"userid" in('
    for agent in params.userids:
        agentQuery += "'" + agent.id + "',";
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    data = ReportData()
    docs = server.Get('OrgRemnants', q)
    if docs != None:
        for d in docs:
            data.add(d, agents, salesNet, groups, goodValues, orgs, price)
            
    return data
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 0
    
    head = ['Дата аудита', 'Номер магазина', 'Адрес', 'Наименование товара', 'Номенклатурный номер товара', 'По учету', 'По факту']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    date = None
    
    for item in data.data:
        docValues = [item.date, item.code, "{0}, {1}".format(item.city, item.address)]
        
        if date != None and date != item.date:
          for i in range(0,7):
            c = sheet.cell(row=row, column=i)
            xlb.setBackColor(c,"ffFFFF00")
            xlb.makeBorder(c, Border.BORDER_THIN)
          row += 1
          
        date = item.date  
        
        for oi in item.items:
            values = []
            values.extend(docValues)
            qtyExit = oi.qtyUnload + oi.qtyInput
            values.extend([oi.name, oi.code, oi.qty, oi.qtyFact])
            
            xlb.makeCells(sheet, row, values)
            row += 1  
    
    cc = 1
    for w in [12,23,60,60,32,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "rmnt_rep.xlsx", server)                
    logging.info('end')
    