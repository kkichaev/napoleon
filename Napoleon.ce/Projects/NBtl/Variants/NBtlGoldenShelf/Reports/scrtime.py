# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class Item:
    data = None
    org = None
    address = None
    finish = None
    
    def __init__(self):
      self.org = ""
      self.address = ""
      self.data = datetime.now()
      self.finish = datetime.now()
      
    def getData(self, row, isLastInDay):
      return [self.data.strftime("%d.%m.%Y"), self.org, self.address, 
              self.data.strftime("%H:%M") if self.data != None else "" , 
              self.finish.strftime("%H:%M") if self.finish != None else "",
              "=E{0}-D{0}".format(row+1), 
              "=D{0}-E{1}".format(row+2, row+1) if not isLastInDay else "" ]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].userids
    
def loadData(server):
    start, finish, userid = inflateParams(server)
    
    WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
    WHERE_STR = WHERE_STR_ALL + ' and "userid" in ({2})'; 

    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
    
    scripts = server.Get("ScriptDoc", where)
    agent = ""
    
    server.ChangeUser(userid)
    agent = server.CurrentUser().name
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    server.RestoreUser()

    map = dict()    
    
    for s in scripts:
      k = s.id + s.created.strftime("%d/%m/%Y")
      
      if not k in map:
        d = Item()
        d.org = orgs[s.id].name if s.id in orgs else "Контрагент с кодом <{0}>".format(s.id)
        d.address = orgs[s.id].address if s.id in orgs else ""
        d.data = s.created
        map[k] = d
      
      m = map[k]
      
      dt = m.data
      
      if s.items != None:
        for s in s.items:
          if s.state == 1 and s.date > dt:
            dt = s.date
            
      m.finish = dt
      
    data = list()
    
    for v in map.values():
      data.append(v)
      
    data = sorted(data, cmp=item_cmp)
    return data, agent
    
def item_cmp(x, y):
  res = cmp(x.data, y.data)

  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data, agent):
    sh.cell(row=0, column=0).value = agent
    sh.cell(row=1, column=0).value = "Дата: " + datetime.now().strftime("%d.%m.%Y")
    
    head = ["Дата", "Название ТТ", "Адрес ТТ", "Время начала визита в ТТ", "Время окончания визита в ТТ", "Итого время в ТТ", "Итого время передвижения от ТТ до ТТ"]
    
    r = 2
    xlb.makeHead(sh, r, head)
    
    svData = None
    for d in data:
        if svData != None:
            r += 1
            xlb.makeCells(sh, r, svData.getData(r, svData.data.date() != d.data.date()))
        svData = d
            
    if svData != None:
        r += 1
        xlb.makeCells(sh, r, svData.getData(r, False))
    
    sh.cell(row=r, column=6).value = ""
    
    if r > 2:
      r += 1
      sh.cell(row=r, column=5).value = "=SUM(F2:F{0})".format(r)
      sh.cell(row=r, column=5).style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME4)
      sh.cell(row=r, column=6).value = "=SUM(G2:G{0})".format(r)
      sh.cell(row=r, column=6).style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME4)
    
    setCellWidth(sh, [30,30,30,30,30,30,30])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 5 or column == 6:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME4)
      
    
def printOut(d, a):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d, a)
                
    return wb

def doReport(server):
    data, agent = loadData(server)
    wb = printOut(data, agent)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    