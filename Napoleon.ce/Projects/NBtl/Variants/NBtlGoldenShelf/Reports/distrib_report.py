# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import datetime


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")


class ItemData :
    __slots__ = ['categ', 'mfr', 'name', 'exists']
    
    def __init__(self, item, groups, price):
        self.exists = 'Да' if item.exist == 1 else 'Нет'
        self.mfr = ''
        
        if item.id in price:
            prc = price[item.id]
            self.name = prc.name
            self.categ = prc.categ
            
            if prc.fid in groups:
                self.mfr = groups[prc.fid].name
        else:
            self.categ = ''
            self.name = 'Товар с кодом <' + item.id + '>'
                
        
class DocData:
    __slots__ = ['city', 'agent', 'slsnet', 'orgName', 'address', 'date', 'items']
    
    def __init__(self, doc, agents, salesNet, groups, orgs, price):        
        self.agent = agents[doc.userid].name if doc.userid in agents else 'Агент с кодом <' + doc.userid + '>'
        self.date = doc.created.date().strftime("%d.%m.%Y")
        self.slsnet = ''
        
        if doc.id in orgs:
            org = orgs[doc.id]
            self.city = org.cid
            self.orgName = org.name
            self.address = org.address
            
            if org.sid in salesNet:
                self.slsnet = salesNet[org.sid].name
        else:
            self.city = ''
            self.orgName = ''
            self.address ='' 
        
        self.items = list()
        
        for item in doc.items:
            i = ItemData(item, groups, price)
            self.items.append(i)

class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()
        
    def add(self, doc, agents, salesNet, groups, orgs, price):
        data = DocData(doc, agents, salesNet, groups, orgs, price)
        self.data.append(data)

def loadData(params, server):
    
    agents = server.Get("Agents", "", "id")
    salesNet = server.Get('Slsnet','', 'id')
    groups = server.Get('GroupGoods','','id')
    orgs = server.Get('CommonOrgs', '', 'id')
    price = server.Get('Price', '"isGoods"=1', 'id')
    
    agentQuery = '"userid" in('
    for agent in params.agents:
        agentQuery += "'" + agent.id + "',";
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    data = ReportData()
    docs = server.Get('Distrib', q)
    if docs != None:
        for d in docs:
            data.add(d, agents, salesNet, groups, orgs, price)
            
    return data
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 1
    
    head = ['Город', 'Мерчендайзер', 'Наименование торговой сети', 'Наименование торговой точки', 
            'Адрес', 'Категория', 'Производитель', 'Дата заполнения отчета', 'SKU','Наличие на полке']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
            
        docValues = [item.city, item.agent, item.slsnet, item.orgName, item.address]
        for oi in item.items:
            values = []
            values.extend(docValues)
            values.extend([oi.categ, oi.mfr, item.date, oi.name, oi.exists])
            
            xlb.makeCells(sheet, row, values)
            row += 1        
    
    cc = 1
    for w in [20,20,20,20,20,20,20,20,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info('end')
    