# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import datetime


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

class AgentDailyData:
    __slots__ = ['route', 'visited', 'fullVisit', 'outRoute', 'badAddress', 'orgs']
    
    MIN_DISTANCE = 300
    DOC_INITED = 1
    
    def __init__(self, route, orgs):
        self.route = list()
        self.visited = list()
        self.fullVisit = list()
        self.badAddress = list()
        self.outRoute = list()
        self.orgs = orgs
        
        for oi in route: 
            if not oi.id in self.route:
                self.route.append(oi.id)

    def isScriptCompleete(self, doc):
        for i in doc.items:
            if i.state != AgentDailyData.DOC_INITED:
                return False
            
        return True
        
    def add(self, doc, locationService):
        id = doc.id
        if not id in self.visited:
            self.visited.append(id)
            
        if not id in self.route and not id in self.outRoute:
            self.outRoute.append(id)
        
        if id in self.orgs:
            org = self.orgs[id]
            loc = locationService.getLocation(org)
            lp = LocationPoint(doc.latitude, doc.longitude)
            
            if lp.distance(loc) > AgentDailyData.MIN_DISTANCE and (not id in self.badAddress):
                self.badAddress.append(id)
                
            if self.isScriptCompleete(doc) and (not id in self.fullVisit):
                self.fullVisit.append(id)
                

class AgentData:
    __slots__ = ['dailyData', 'route']
    
    def __init__(self, route):
        self.dailyData = dict()
        self.route = route
        
    def add(self, doc, locationService):
        docDate = doc.created.date();
        if not docDate in self.dailyData:
            agentDailyDocs = AgentDailyData(self.route.getDayRoute(docDate), self.route.orgs)
            self.dailyData[docDate] = agentDailyDocs
        self.dailyData[docDate].add(doc, locationService)
        
    def count(self, total, dayList):
        workDays = 0
        for cd in dayList:
            if not cd in self.dailyData:
                notVisited = len(self.route.getDayRoute(cd))
                if notVisited > 0:
                    total.notVisited += notVisited
                    workDays += 1
            else:
                workDays += 1
                data = self.dailyData[cd]
                route = len(data.route)
                visited = len(data.visited)
                outRoute = len(data.outRoute)
                onRoute = visited - outRoute
                
                total.plan += route
                total.visits += visited
                total.fact += onRoute
                total.notVisited += route - onRoute
                total.outRoute += outRoute
                total.outAddress += len(data.badAddress)
                total.fullVisit += len(data.fullVisit)
                
                if route > 0:
                    cp = float(onRoute) / route
                    #print cp
                    total.progress += cp
                
        if workDays > 0:
            total.progress /= workDays
        

class Total:
    __slots__ = ['visits', 'plan', 'fact', 'notVisited', 'outRoute', 'outAddress', 'fullVisit', 'progress']
    
    def __init__(self):
        self.visits = 0
        self.plan = 0
        self.fact = 0
        self.notVisited = 0
        self.outRoute = 0
        self.outAddress = 0
        self.fullVisit = 0
        self.progress = 0

    def add(self, total):
        self.visits += total.visits
        self.plan += total.plan
        self.fact += total.fact
        self.notVisited += total.notVisited
        self.outRoute += total.outRoute
        self.outAddress += total.outAddress
        self.fullVisit += total.fullVisit
        self.progress += total.progress

class ReportData:
    __slots__ = ['data', 'locationService', 'server', 'agents']
    
    def __init__(self, server):
        self.data = dict()
        self.locationService = OrgLocation(server)
        self.server = server
        self.agents = server.Get('Agents', '', 'id')
        
    def add(self, doc):
        userid = doc.userid
        if not userid in self.data:
            agentData = AgentData(AgentRoute(self.server, userid))
            agentData.add(doc, self.locationService)
            self.data[userid] = agentData
        else:
            self.data[userid].add(doc, self.locationService)
        
    def getTotalData(self, userid, dayList):
        total = Total()
        
        if userid in self.data:
            self.data[userid].count(total, dayList)
        
        return total


def loadData(params, server):
    data = ReportData(server)
    
    agentQuery = '"userid" in('
    for agent in params.agents:
        agentQuery += "'" + agent.id + "',";
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    docs = server.Get('ScriptDoc', q)
    if docs != None:
         
        for d in docs:
            data.add(d)
            
    return data
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 1
    
    cell = sheet.cell(row=row, column=0)
    cell.value = 'Дата с {0} по {1}'.format( params.start.strftime("%d/%m/%Y"), params.finish.strftime("%d/%m/%Y"))
    row += 2
    
    head = ['Подразделение / агент', 'визиты', 'план по маршруту', 'факт по маршруту', 'не посетил', 'не по маршруту', 
            'адрес ТТ не соответствует адресу GPS', 'кол-во завершенных визитов (где пройдены все шаги сценария)', 'прогресс']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1
    
    dayList = list()
    cd = params.start;
    while cd <= params.finish:
        dayList.append(cd.date())
        cd = cd + datetime.timedelta(days=1)
    
    agentCount = 0
    total = Total()
    for agent in params.agents:

        name = data.agents[agent.id].name if agent.id in data.agents else 'Агент с кодом <' + agent.id + '>'
        ct = data.getTotalData(agent.id, dayList)
        values = [name, ct.visits, ct.plan, ct.fact, ct.notVisited, ct.outRoute, ct.outAddress, ct.fullVisit, int(ct.progress * 100)]        
        xlb.makeCells(sheet, row, values)
        
        row += 1        
        agentCount += 1
    
        total.add(ct)
    
    if agentCount > 0: total.progress /= agentCount
    
    values = ['Итого', total.visits, total.plan, total.fact, total.notVisited, total.outRoute, total.outAddress, total.fullVisit, int(total.progress * 100)]        
    xlb.makeCells(sheet, row, values)
    
    
    cc = 1
    for w in [20,20,20,20,20,20,20,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "vist_rep.xlsx", server)                
    logging.info('end')
    