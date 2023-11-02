# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *
from grsoft.xl_base import XLBuilder

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN

class ItemCell:
    face = None
    qty = None
    
    def __init__(self):
        self.face = 0.0
        self.qty = 0.0

class PageItem:
    city = None
    user = None
    slsnet = None
    org = None
    data = None
    created = None
    partshelf = 0
    fact = None
    visit = None
    planogram = 0
    date = None
    address = None
    facemy = 0
    faceall = 0
    faceplan = 0
    facefact = 0
    remark = None
    
    id = None
    avgItems = 0.0
    avgGroup = 0.0
    
    def __init__(self):
        self.data = dict()
        self.visit = list()
        self.avgItems = 0.0
        self.avgGroup = 0.0
        
    def process(self, c, report):
        self.city = report.getCity(c.id)
        self.user = report.getUser(c.userid)
        self.slsnet = report.getSls(c.id)
        self.org = report.getOrgName(c.id)
        self.address = report.getOrgAddress(c.id)
        self.created = c.created.strftime("%d.%m.%Y")
        self.post_process(c, report)
        self.date = datetime(c.created.year, c.created.month, c.created.day)
        self.remark = c.remark
        self.id = c.id

    def countAvg(self, avgItems, avgGroup):
        self.avgItems = 0
        self.avgGroup = 0
        
        if self.id in avgItems:
            cdata = avgItems[self.id]
            for val in cdata.itervalues():
                self.avgItems += float(val)
                
            self.avgItems /= len(cdata)

        if self.id in avgGroup:
            cdata = avgGroup[self.id]
            for val in cdata.itervalues():
                self.avgGroup += float(val)
                
            self.avgGroup = int(self.avgGroup / len(cdata) + 0.5)
        
    def post_process(self, c, report):
        ps = report.getPartShelf(c.id)
        self.partshelf = ps
        self.faceplan = report.getFacePlan(c.id);
        self.planogram = report.getPlanogram(c)
        
        other = 0
        my = 0
        all = 0
        
        for i in c.items:
            if i.id in report.mprice:
                self.data[i.id] = i.qty
                
                if report.mprice[i.id].my > 0:
                    my = my + i.qty
                else:
                    other = other + i.qty
                     
        all = my + other
        onepercent  = all / 100
        f = my / onepercent if onepercent != 0 else 0
        self.fact = "{:0.0f}".format(f)
        self.faceall = all
        self.facemy = my
        
        p = self.facemy / self.faceplan * 100 if self.faceplan != 0 else 0
        self.facefact = "{:0.0f}".format(p)
        report.getVisit(c, self.visit)
        
    def avgMePrc(self):
        return 0 if self.avgGroup == 0 else int(self.avgItems * 100 / self.avgGroup + 0.5)

    def sumQty(self):
        r = 0;
        
        for v in self.data.values():
            r += v
            
        return r
        
class MPageItem(PageItem):
    def post_process(self, c, report):
       for i in c.items:
            if i.id in report.mprice:
                self.data[i.id] = i.cost
                     
class Page:
    report = None
    items = None;
    title = None
    width = 0
    price = None
    
    def __init__(self, report):
        self.report = report
        self.items = list()
        self.title = "Отчет по доли полки"
        self.width = 14
        self.price = list()
        
    def collumns(self, sheet, r, c):
        c -= 1
        
        titles = ["Кол-во усл.ед. по нашей продукции", "Среднее по нашей продукции", "Итого доля полки в усл. ед.","План по доле полки в усл. ед.",
                  "Факт по доле полке, %", "План по доли полки %", "Факт по доли полки  на дату", "СРЕДНЕЕ по группе усл. ед.", 
                  "Факт по доли полки по торговой точке", 'Факт по доле полки по все сети']
        
        for str in titles:
            c += 1
            sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
            setVal(sheet.cell(row=r, column=c), str, Alignment.VERTICAL_BOTTOM, rotation=90)
            col_letter = get_column_letter(c + 1) 
            sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Отметка о выполнении планограммы", Alignment.VERTICAL_BOTTOM, rotation=90)    
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Комментарий", Alignment.VERTICAL_BOTTOM, rotation=0)   
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 35

    def itemsdata(self, sheet, r, c, i):
        prcInt = 0 if i.faceall == 0 else int(i.facemy * 100/i.faceall + 0.5) 
        prcMe = str(prcInt) + '%'
        planPrc = str(0 if i.avgGroup == 0 else int(i.faceplan * 100 / i.avgGroup + 0.5)) + '%'
        avgMePrc = str(i.avgMePrc()) + '%'
        
        data = [i.facemy, i.avgItems, i.faceall, i.faceplan, i.facefact, planPrc, prcMe, i.avgGroup, avgMePrc, 0]
        
        for d in data:
            setVal(sheet.cell(row=r, column=c), d)
            c = c + 1

        setVal(sheet.cell(row=r, column=c), i.planogram if i.planogram > 0 else "")      
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.remark)      
        c = c + 1
        
        for v in i.visit:
            idx = 0
            if v.items != None:
                for vi in v.items:
                    cell = sheet.cell(row=r, column=c) 
                    cell.hyperlink = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                    cell.value = "Фото{0}".format(idx + 1)
                    idx = idx + 1
                    c = c + 1        
                
        return c
    
    def process(self, report):
        avgItems = dict()
        avgGroup = dict()
        
        for c in report.contract:
            if not (report.allowedOrgs == None or c.id in report.allowedOrgs) :
                continue
            pi = PageItem()
            pi.process(c, report)
            self.items.append(pi)
            
            ctDate = None
            if not (c.id in avgItems):
                ctDate = dict()
                avgItems[c.id] = ctDate
            else:
                ctDate = avgItems[c.id]

            gtData = None
            if not (c.id in avgGroup):
                gtData = dict()
                avgGroup[c.id] = gtData
            else:
                gtData = avgGroup[c.id]
            
            cdata = c.created.date()
            val = 0
            if (cdata in ctDate):
                val = ctDate[cdata]
            ctDate[cdata] = pi.facemy + val  

            val = 0
            if (cdata in gtData):
                val = gtData[cdata]
                
            gtData[cdata] = pi.faceall + val
            
            if not c.userid in report.agentPages:
                ap = AgentPage()
                ap.process(c.userid, report)
                report.agentPages[c.userid] = ap  
            
            
        self.items = sorted(self.items, cmp=lambda lhs, rhs: cmp(lhs.created, rhs.created))
        for item in self.items:
            item.countAvg(avgItems, avgGroup)
        
class MonitoringPage(Page):
    def __init__(self, report):
        Page.__init__(self, report)
        self.title = "Ценовой мониторинг"
        self.width = 6
        
    def collumns(self, sheet, r, c):
        pass
    
    def itemsdata(self, sheet, r, c, i):
        pass
    
    def process(self, report):
        for m in report.monitoring:
            pi = MPageItem()
            pi.process(m, report)
            self.items.append(pi)
                         
def price_cmp(lhs, rhs):
    result = (int)(lhs.my - rhs.my)
    
    if result == 0:
        result = cmp(lhs.group, rhs.group)
    
    if result == 0:
        result = cmp(lhs.name, rhs.name)
            
    return result

class AgentPageItem:
    slsnet = None
    sid = None
    akbm = 0
    koef = 0
    akbr = 0
    plan = 0
    
    def __init__(self):
        self.slsnet = ""
        self.sid = ""
        self.akbm = 0
        self.koef = 0
        self.akbr = 0
        self.plan = 0
        self.routeOrg = list()
        self.orgVisited = list()
    
    def markVisited(self, orgid):
        if orgid in self.routeOrg and not orgid in self.orgVisited:
            self.orgVisited.append(orgid)


    def getData(self, row):
        return [self.slsnet, len(self.routeOrg), self.koef, len(self.orgVisited), "=IFERROR(D{0}/D2,0)".format(row+1), self.plan / 100, 3]     
        # return [self.slsnet, self.akbm, self.koef, "=B{0}*C{0}".format(row+1), "=IFERROR(D{0}/D2,0)".format(row+1), self.plan / 100, 3]     
    
class AgentPage:
    agent = None
    items = None
    
    def __init__(self):
        self.agent = ""
        self.items = dict()

    def markVisited(self, orgid):
        for ap in self.items.itervalues():
            ap.markVisited(orgid) 

    def getRouteOrgs(self, route, orgs):
        ret = dict()

        for r in route:
            for ri in r.items:
                if ri.name in orgs:
                    org = orgs[ri.name]
                    if not org.sid in ret:
                        ret[org.sid] = list()
                    ret[org.sid].append(org.id)

        return ret

    def process(self, userid, report):
        self.agent = report.getUser(userid)
        server = report.server
        
        server.ChangeUser("'" + userid + "'")
        orgs = server.Get("Org", "", "id")
        route = server.Get('OrgFolder', '')
        server.RestoreUser()
        
        routeOrgs = self.getRouteOrgs(route, orgs)

        for o in orgs.values():
            if not o.sid in self.items:
                api = AgentPageItem()
                api.slsnet = o.sid
                api.sid = o.sid
                
                if o.sid in routeOrgs:
                    api.routeOrg = routeOrgs[o.sid]

                if o.sid in report.slsnet:
                    sls = report.slsnet[o.sid]
                    api.slsnet = sls.name
                    api.koef = sls.koef
                    api.plan = sls.plan
                
                self.items[o.sid] = api
                
            api = self.items[o.sid]
            api.akbm += 1    
                 
class Report:
    pages = None
    
    """ Список прайс """
    price = None
    """ Позиции в списке по ID """
    pidx = None
    
    slsnet = None
    city = None
    org = None
    agents = None
    partshelf = None
    mprice = None
    visit = None
    planogram = None
    contract = None
    
    matrix = None
    allowedOrgs = None

    visitdata = None
    planogramdata = None
    monitoringdata = None
    
    reportItemID = None
    agentPages = None
    server = None
    
    def loadMatrix(self, server):
        param = server.Params[0]
        cid = param.cid
        matrixName = param.matrix

        if matrixName == "" :
            return
        omtx = server.Get("OrgMatrix", '"cdef"=' + "'" + cid + "'" + ' and "name"=' + "'" + param.matrix + "'")
        if omtx == None:
            return

        self.allowedOrgs = list()
        for omi in omtx:
            self.allowedOrgs.append(omi.id)

        mtx = server.Get("CommonMatrix", '"name"=' +"'" + omtx[0].name + "'")
        if mtx == None:
            return
        self.matrix = list()
        for oi in mtx[0].items:
            if oi.id in self.mprice:
               self.matrix.append(oi.id)

    def __init__(self, server):
        param = server.Params[0]
        start = param.start
        finish = param.finish
        cid = param.cid
        usePhoto = param.photo
        userid = param.userid
        
        self.server = server
        self.reportItemID = None
        self.agentPages = dict()
        
        #self.weeks = Weeks(start, finish)
        
        conDef = server.Get("ContractDef", '"id"=' + "'" + cid + "'")
        
        pids = ""
        
        for cd in conDef:
            for i in cd.items:
                iid = "'" + i.id +"'"
                if iid in pids: 
                    continue
                
                if len(pids) > 0:
                    pids += ","
                    
                pids += iid
        
        self.mprice = server.Get("ManagerPrice", '"id" in (' + pids + ')', "id")
        if param.item != '':
            removed = list()
            item = param.item.upper()
            for p in self.mprice.itervalues():
                if p.name.upper() != item:
                    removed.append(p.id)
                elif p.my == 1: 
                    self.reportItemID = p.id
                    
            for id in removed:
                self.mprice.pop(id, None)
                
        self.price = list();
        self.price.extend(self.mprice.values())
        self.price = sorted(self.price, cmp=price_cmp)
        
        self.loadMatrix(server)
        
        self.pidx = dict()
        for i in range(0, len(self.price)):
            self.pidx[self.price[i].id] = i    
        
        self.org = server.Get("Org", '"id" is not null', "id")
        self.agents = server.Get("Agents","","id")
        
        endRange = finish + timedelta(days=1) 
        
        where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + \
          endRange.strftime('%d/%m/%Y') + '") and "def"='+"'" + cid + "'" + ' and "userid" in (' + userid + ')'
          
        self.contract = server.Get("Contract", where)
        
        vobj = "Visit" if usePhoto == 1 else "VisitInfo"
        self.visit = server.Get(vobj, where)
        
        #where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + endRange.strftime('%d/%m/%Y') + '")'
        self.planogram = server.Get("Planogram", where)
        self.monitoring = server.Get("CMonitoring", where)
        self.partshelf = server.Get("PartShelf", '"cid"= ' + "'" + cid + "'", "sid") 
        self.btlplan = server.Get("BtlPlan", '"cid"= ' + "'" + cid + "'", "id")
        
        self.pages = {MonitoringPage(self), Page(self)}
        
        self.slsnet = server.Get("Slsnet","","id")
        self.city = server.Get("City", "", "id")
        
        for p in self.pages:
            p.process(self)
            hasval = list()
            
            for r in p.items:
                for d in r.data:
                    if not (self.matrix == None or d in self.matrix):
                        continue
                    if not d in hasval and r.data[d] > 0 and d in self.mprice:
                        p.price.append(self.mprice[d])
                        hasval.append(d)
                        
            if not self.matrix == None:
                for d in self.matrix:
                    if not d in hasval:
                        p.price.append(self.mprice[d])
                        hasval.append(d)

            p.price = sorted(p.price, cmp=price_cmp)    
            #print p.price        
        self.loadVisited()

    def loadVisited(self):
        docsList = [self.visit, self.planogram, self.monitoring]
        for docs in docsList:
            for d in docs:
                try: 
                    if d.userid in self.agentPages:
                        self.agentPages[d.userid].markVisited(d.id)
                except:
                    print d

    def getCity(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].cid
        
        return result
    
    def getFacePlan(self, id):
        result = 0.0;
        
        if id in self.btlplan:
            plan = self.btlplan[id]
            if self.reportItemID != None:
                for item in plan.items:
                    if item.id == self.reportItemID:
                        result = item.face
                        break
            else:
                result = plan.face
        
        #print id, result    
        return result;    
        
    def getOrgName(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].name
        
        return result
    
    def getOrgAddress(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].address
        
        return result
            
    def getSls(self, id):
        result = id
        
        if id in self.org and self.org[id].sid in self.slsnet:
            result = self.slsnet[self.org[id].sid].name
        
        return result      
    
    def getUser(self, id):
        result = id
        
        if id in self.agents:
            result = self.agents[id].name
            
        return result
      
    def getAddress(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].address 
            
        return result 
    
    def sameDay(self, d1, d2):
        return d1.day == d2.day and d1.month == d2.month and d1.year == d2.year
     
    def getPartShelf(self, id):
        result = 0
        
        if id in self.org and self.org[id].sid in self.partshelf:
            result = self.partshelf[self.org[id].sid].part
            
        return result;
    
    def getVisit(self, c, val):
        if self.visitdata == None and len(self.visit) > 0:
            self.visitdata = dict()
            
            for v in self.visit:
                if not v.id in self.visitdata:
                    self.visitdata[v.id] = dict()
                 
                v1 = self.visitdata[v.id]
                
                if not v.userid in v1:
                    v1[v.userid] = dict()
                
                v2 = v1[v.userid]
                dt = datetime(v.created.year, v.created.month, v.created.day)
                if not dt in v2:
                    v2[dt] = list()
                    
                v2[dt].append(v)         
        
        if self.visitdata != None:
            cdt = datetime(c.created.year, c.created.month, c.created.day)
            
            if c.id in self.visitdata and c.userid in self.visitdata[c.id] and cdt in self.visitdata[c.id][c.userid]:              
                val.extend(self.visitdata[c.id][c.userid][cdt])
    
    def getPlanogram(self, c):
        if self.planogramdata == None:
            self.planogramdata = dict()
            
            for p in self.planogram:
                if not p.id in self.planogramdata:
                    self.planogramdata[p.id] = dict()
                    
                p1 = self.planogramdata[p.id]
                
                if not p.uesrid in p1:
                    p1[p.userid] = dict()
                
                p2 = p1[p.userid]  
                dt = datetime(p.created.year, p.created.month, p.created.day)
                p2[dt] = p
        
        cdt = datetime(c.created.year, c.created.month, c.created.day)
        if c.id in self.planogramdata and c.userid in self.planogramdata[c.id] and cdt in self.planogramdata[c.id][c.userid]:              
            return self.planogramdata[c.id][c.userid][cdt].approved
         
        return 0    
              
def setVal(cell, value, vrt = Alignment.VERTICAL_CENTER, hrz= Alignment.HORIZONTAL_CENTER, wrap=True, bold=False, rotation=0):
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    cell.style.alignment.text_rotation = rotation
    cell.style.font.bold = bold
    cell.value = value
                      

def countAvgSls(items):
    avgCount = dict()
    avgVal = dict()
    
    orgs = list()
    
    for i in items:
        if i.id in orgs:
            continue
        
        avg = 0
        avgctr = 0
        key = i.slsnet + i.user
        
        if key in avgVal:
            avg = avgVal[key]
            avgctr = avgCount[key]
            
        avg += i.avgMePrc()
        avgctr += 1
        avgVal[key] = avg
        avgCount[key] = avgctr
        
        avg = 0
        avgctr = 0
        key = i.slsnet
        
        if key in avgVal:
            avg = avgVal[key]
            avgctr = avgCount[key]
            
        avg += i.avgMePrc()
        avgctr += 1
        avgVal[key] = avg
        avgCount[key] = avgctr

    avgSls = dict()
    
    for k, v in avgVal.iteritems():
        avg = float(v) / avgCount[k] + 0.5
        avgSls[k] = avg

    return avgSls                      

class XLBuilderEx(XLBuilder):
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 4 :
      cell.style.number_format._set_format_code('0%')
    elif column == 5:
      cell.style.number_format._set_format_code('0.00%') 
    elif column == 6:
      cell.style.number_format._set_format_code('0.0%') 
                            
def run(server):
    print "contract start"
    server.RegisterType("Result[name:s,file:b,items[name:s,photo:b]]")
    outObj = server.New("Result")

    report = Report(server)
    wb = Workbook(False, 'cp1251')
    
    sheet = None
    for page in report.pages:
        if len(page.price) == 0:
            continue
        
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()
        sheet.title = page.title
        
        i = 1
        r = 0

        sheet.merge_cells(start_row=r, start_column=0, end_row=r+1, end_column=0)    
        setVal(sheet.cell(row=r, column=0), "Город")
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
        setVal(sheet.cell(row=r, column=1), "Мерчендайзер")
        col_letter = get_column_letter(2) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        setVal(sheet.cell(row=r, column=2), "Наименование торговой сети")
        col_letter = get_column_letter(3) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        setVal(sheet.cell(row=r, column=3), "Наименование торговой точки")
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)
        setVal(sheet.cell(row=r, column=4), "Адрес")
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        setVal(sheet.cell(row=r, column=5), "Дата заполнения отчета")
        col_letter = get_column_letter(6) 
        sheet.column_dimensions[col_letter].width = 20

        setVal(sheet.cell(row=r, column=6), "Итого фейсов")
        col_letter = get_column_letter(7) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r+1, column=6), "Товар", vrt = Alignment.VERTICAL_BOTTOM, rotation=90)
       
        
        group = None
        START_COL = 7
        c = START_COL
        pc = c
        isMy = False
        
        for p in page.price:
            isMy = p.my > 0
            
            if isMy:
                if group != None and group != p.group: 
                    sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
                    setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
                    pc = c
                    
                group = p.group
                setVal(sheet.cell(row=r + 1, column=c), p.name, vrt = Alignment.VERTICAL_BOTTOM, rotation=90)
                col_letter = get_column_letter(c + 1) 
                sheet.column_dimensions[col_letter].width = 5
                
                c = c + 1
            
        sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
        setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
        sheet.row_dimensions[2].height = 95
        
        page.collumns(sheet, r, c)
        
        r = r + 2
        dt = None
        sr = r + 1
        dsr = sr
        
        sum_group = dict()
        sum_result = dict()
        last_date = None
        
        slsAvg = countAvgSls(page.items)
        for i in page.items:
            if dt == None:
                dt = i.date
                   
            setVal(sheet.cell(row=r, column=0), i.city, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=1), i.user, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=2), i.slsnet, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=3), i.org, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=4), i.address, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=5), i.created, hrz= Alignment.HORIZONTAL_RIGHT)
            setVal(sheet.cell(row=r, column=6), i.sumQty(), hrz= Alignment.HORIZONTAL_RIGHT)
            
            c = START_COL;
            last_date = i.created
            
            for p in page.price:
              qty = i.data[p.id] if p.id in i.data else 0
              
              if p.my != 0:
                  setVal(sheet.cell(row=r, column=c), qty)
                  c = c + 1
              
              if not p.id in sum_group:
                  sum_group[p.id] = qty
              else:     
                  sum_group[p.id] = sum_group[p.id] + qty
                    
            lastCol = page.itemsdata(sheet, r, c, i)
            key = i.slsnet + i.user
            avgCount = slsAvg[key] if key in slsAvg else 0
            if avgCount > 0:
                avgVal = str(avgCount) + '%'
                setVal(sheet.cell(row=r, column=lastCol - 3), avgVal, hrz= Alignment.HORIZONTAL_LEFT)
            
            r = r + 1
        
        rangeBorders(sheet.range("A1:"+get_column_letter(lastCol)+str(r)) )
    
    apl = list()
    apl.extend(report.agentPages.values())
    apl = sorted(apl, cmp=lambda lhs, rhs: cmp(lhs.agent, rhs.agent))
    
    xlb = XLBuilderEx()
    head = ["Торговая сеть", "Кол-во ТТ план (по маршруту)", "Коэффициент", "Кол-во ТТ факт", "Доли сети", "Плановая доля полки", "Фактическая доля полки по сети"]
    resitems = dict()
    
    for ap in apl:
        sheet = wb.create_sheet()
        sheet.title = ap.agent
        xlb.makeHead(sheet,0,head,True)
        cc = 1
        for w in [20,20,20,20,20,20,20]:
            sheet.column_dimensions[get_column_letter(cc)].width = w
            cc += 1
        
        items = ap.items.values()
        items = sorted(items, cmp=lambda lhs, rhs: cmp(lhs.slsnet, rhs.slsnet))
        
        r = 1
        sz = len(items)+2
        itogo = ["Итого","=SUM(B3:B{0})".format(sz),"","=SUM(D3:D{0})".format(sz),"100%","=SUMPRODUCT(E3:E{0},F3:F{0})".format(sz),
                 "=SUMPRODUCT(E3:E{0},G3:G{0})".format(sz)]
        xlb.makeCells(sheet, r, itogo)
        r += 1
        
        for d in items:
            xlb.makeCells(sheet, r, d.getData(r))
            key = d.slsnet+ap.agent
            avgCount = slsAvg[key] if key in slsAvg else 0
            cell = sheet.cell(row=r, column=6)
            setVal(cell, avgCount / 100, hrz= Alignment.HORIZONTAL_RIGHT)
            cell.style.number_format._set_format_code('0.0%') 
            
            r += 1
            
            if not d.sid in resitems:
                api = AgentPageItem()
                api.slsnet = d.slsnet
                api.koef = d.koef
                api.plan = d.plan
                resitems[d.sid] = api
            
            api = resitems[d.sid]
            api.akbm += d.akbm
            api.routeOrg.extend(d.routeOrg)
            api.orgVisited.extend(d.orgVisited)
    
    sheet = wb.create_sheet()
    sheet.title = "ИТОГО по всем"
    xlb.makeHead(sheet,0,head,True)
    
    cc = 1
    for w in [20,20,20,20,20,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
            
    items = resitems.values()  
    items = sorted(items, cmp=lambda lhs, rhs: cmp(lhs.slsnet, rhs.slsnet))              
    r = 1
    sz = len(items)+2
    itogo = ["Итого","=SUM(B3:B{0})".format(sz),"","=SUM(D3:D{0})".format(sz),"100%","=SUMPRODUCT(E3:E{0},F3:F{0})".format(sz),
             "=SUMPRODUCT(E3:E{0},G3:G{0})".format(sz)]
    xlb.makeCells(sheet, r, itogo)
    r += 1
    
    for d in items:
        xlb.makeCells(sheet, r, d.getData(r))
        
        avgCount = slsAvg[d.slsnet] if d.slsnet in slsAvg else 0
        avgVal = str(avgCount) + '%'
        setVal(sheet.cell(row=r, column=6), avgVal, hrz= Alignment.HORIZONTAL_RIGHT)
            
        r += 1                
                    
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    for v in report.visit:
        idx = 0
        if v.items != None:
            for vi in v.items:
                item = obj.items.New()
                item.name = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                item.photo = vi.id
                idx = idx + 1
    
    server.Put(outObj)
    
    print "contract finish"
    