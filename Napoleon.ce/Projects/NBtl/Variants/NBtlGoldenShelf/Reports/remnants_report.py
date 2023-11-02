# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import datetime


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")


class ItemData :
    __slots__ = ['categ', 'mfr', 'name', 'qty', 'qtyFlag', 'qtyFact', 'factFlag']
    
    def __init__(self, item, groups, goodValues, price):
        self.qty = item.qty
        self.mfr = ''
        self.qtyFlag = 0
        self.factFlag = 0
        self.qtyFact = item.qtyFact
        
        for qvi in goodValues:
            if qvi.id == item.id:
                if self.qty <= qvi.redLine: self.qtyFlag = -1
                elif self.qty >= qvi.greenLine: self.qtyFlag = 1

                if self.qtyFact <= qvi.redLine: self.factFlag = -1
                elif self.qtyFact >= qvi.greenLine: self.factFlag = 1
                break 
        
        
        if item.id in price:
            prc = price[item.id]
            self.name = prc.name
            self.categ = prc.categ
            
            if prc.fid in groups:
                self.mfr = groups[prc.fid].name
        else:
            self.categ = ''
            self.name = 'Товар с кодом <' + item.id + '>'
                
        
class DocData:
    __slots__ = ['city', 'agent', 'slsnet', 'orgName', 'address', 'date', 'items']
    
    def __init__(self, doc, agents, salesNet, groups, goodValues, orgs, price):        
        self.agent = agents[doc.userid].name if doc.userid in agents else 'Агент с кодом <' + doc.userid + '>'
        self.date = doc.created.date().strftime("%d.%m.%Y")
        self.slsnet = ''
        
        if doc.id in orgs:
            org = orgs[doc.id]
            self.city = org.cid
            self.orgName = org.name
            self.address = org.address
            
            if org.sid in salesNet:
                self.slsnet = salesNet[org.sid].name
        else:
            self.city = ''
            self.orgName = ''
            self.address ='' 
        
        self.items = list()
        
        gvals = goodValues[doc.id].items if doc.id in goodValues else list() 
        for item in doc.items:
            i = ItemData(item, groups, gvals, price)
            self.items.append(i)

class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()
        
    def add(self, doc, agents, salesNet, groups, goodValues, orgs, price):
        data = DocData(doc, agents, salesNet, groups, goodValues, orgs, price)
        self.data.append(data)

def loadData(params, server):
    
    agents = server.Get("Agents", "", "id")
    salesNet = server.Get('Slsnet','', 'id')
    groups = server.Get('GroupGoods','','id')
    goodValues = server.Get('GoodsValues','','id')
    orgs = server.Get('CommonOrgs', '', 'id')
    price = server.Get('Price', '"isGoods"=1', 'id')
    
    agentQuery = '"userid" in('
    for agent in params.agents:
        agentQuery += "'" + agent.id + "',";
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    data = ReportData()
    docs = server.Get('OrgRemnants', q)
    if docs != None:
        for d in docs:
            data.add(d, agents, salesNet, groups, goodValues, orgs, price)
            
    return data
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    row = 1
    
    head = ['Город', 'Мерчендайзер', 'Наименование торговой сети', 'Наименование торговой точки', 'Адрес', 'Категория', 
            'Производитель', 'Дата заполнения отчета', 'SKU','Остаток по учету','Остаток по факту']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
            
        docValues = [item.city, item.agent, item.slsnet, item.orgName, item.address]
        for oi in item.items:
            values = []
            values.extend(docValues)
            values.extend([oi.categ, oi.mfr, item.date, oi.name, oi.qty, oi.qtyFact])
            
            xlb.makeCells(sheet, row, values)
            
            if oi.qtyFlag != 0:
                cell = sheet.cell(row=row, column=9)
                color = 'FFFF0000' if oi.qtyFlag < 0 else 'FF0000FF'
                xlb.setBackColor(cell, color)

            if oi.factFlag != 0:
                cell = sheet.cell(row=row, column=10)
                color = 'FFFF0000' if oi.factFlag < 0 else 'FF0000FF'
                xlb.setBackColor(cell, color)
            
            row += 1        
    
    cc = 1
    for w in [20,20,20,20,20,20,20,20,20,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "rmnt_rep.xlsx", server)                
    logging.info('end')
    