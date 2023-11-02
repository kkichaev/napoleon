# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN

class PageItem:
    org = None
    num = None
    date = None
    payd = None
    sum = None
    pdz = None
    
    def __init__(self, dlv, rpt):
        self.org = rpt.getOrgName(dlv.id)
        self.num = dlv.number
        self.date = dlv.date
        self.payd = dlv.payDate
        
        self.sum = 0
        
        for i in dlv.items:
            self.sum = self.sum + i.sum
        
        self.pdz = 0
        
        if dlv.payDate < datetime.now():
            self.pdz = dlv.sumD    
                   
class Page:
    report = None
    items = None;
    title = None
    
    def __init__(self, title):
        self.items = list()
        self.title = title
        
    def process(self, report, d):
        self.items.append(PageItem(d, report))                
             
class Report:
    pages = None
    dlv = None
    org = None
    
    def __init__(self, server):
        where = server.Params[0].uid;
        self.agents = server.Get("Agents","", "id")
        self.dlv = server.Get("Delivery", where)
        self.org = server.Get("Org", where, "id")
        
        self.pages = dict()
        
        for d in self.dlv:
            if d.userid in self.agents:
                if not d.userid in self.pages:
                    self.pages[d.userid] = Page(self.agents[d.userid].name)
                 
                self.pages[d.userid].process(self, d)  
                
    def getOrgName(self, id):
        result = id
        
        if id in self.org:
            o = self.org[id]
            result = o.name + "(" + o.address + ")"
            
        return result                  
    
def setVal(cell, value, vrt = Alignment.VERTICAL_CENTER, hrz= Alignment.HORIZONTAL_CENTER, wrap=True, bold=False, rotation=0):
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    cell.style.alignment.text_rotation = rotation
    cell.style.font.bold = bold
    cell.value = value
                      
def run(server):
    print "dlvrpt start"
    server.RegisterType("Result[name:s,file:b,items[name:s,photo:b]]")
    outObj = server.New("Result")

    report = Report(server)
    wb = Workbook(False, 'cp1251')
     
    plist = report.pages.values()
    plist = sorted(plist, cmp=lambda lhs, rhs: cmp(lhs.title, rhs.title))
 
    sheet = None
    
    for page in plist:
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()
         
        sheet.title = page.title
           
        r = 0
   
        setVal(sheet.cell(row=r, column=0), "Контрагент (адрес)", bold=True)
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 50
        setVal(sheet.cell(row=r, column=1), "Номер документа", bold=True)
        col_letter = get_column_letter(2) 
        sheet.column_dimensions[col_letter].width = 30
        setVal(sheet.cell(row=r, column=2), "Дата отгрузки", bold=True)
        col_letter = get_column_letter(3) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r, column=3), "Дни отсрочки", bold=True)
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r, column=4), "Дата оплаты", bold=True)
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r, column=5), "Общая сумма долга", bold=True)
        col_letter = get_column_letter(6) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r, column=6), "ПДЗ", bold=True)
        col_letter = get_column_letter(7) 
        sheet.column_dimensions[col_letter].width = 20 
        r = r + 1
         
        for i in page.items:
            setVal(sheet.cell(row=r, column=0), i.org, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=1), i.num, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=2), i.date, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=3), (i.payd - i.date).days, hrz= Alignment.HORIZONTAL_RIGHT)
            setVal(sheet.cell(row=r, column=4), i.payd, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=5), i.sum, hrz= Alignment.HORIZONTAL_RIGHT)
            setVal(sheet.cell(row=r, column=6), i.pdz, hrz= Alignment.HORIZONTAL_RIGHT) 
            r = r + 1

        rangeBorders(sheet.range("A1:G{0}".format(r)))
                 
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)
 
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
 
    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
     
    server.Put(outObj)
    
    print "dlvrpt finish"
    