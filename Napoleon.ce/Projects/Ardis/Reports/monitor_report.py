# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
import datetime
import tempfile
import io


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def getDivisionAgents(server, division, agents):
   if division == None or len(division) == 0:
      return

   for d in division :
      for a in d.agents :
        agents.append(a.id)

      getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)

def loadAgents(server, division):
   ret = dict()
   agents = server.Get("Agents", "")

   divagents = []
   getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
   divagents = set(divagents)

   for a in agents:
      if a.id in divagents:
         ret[a.id] = a

   return ret

def loadAgent(server, agentID):
   ret = dict()
   agents = server.Get("Agents", "")

   for a in agents:
      if a.id == agentID:
         ret[a.id] = a

   return ret

def makeIDStr(server, agents):
   res = '"userid" in ('

   for id in agents.iterkeys():
      res += "'" + id + "',"

   res = res[:-1] + ")"
   return res

def setTitle(sheet, sr1, cc, esr, ecc, title):
   cell = sheet.cell(row=sr1, column=cc)
   cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
   cell.style.alignment.vertical = Alignment.VERTICAL_CENTER

   cell.value = title
   if esr != sr1 or ecc != cc :
      sheet.merge_cells(start_row=sr1, start_column=cc, end_row=esr, end_column=ecc)
      if esr != sr1 :
         c = cell
         for sr in range(sr1+1, esr+1) :
           c = sheet.cell(row=sr, column=cc)
           c.style.borders.left.border_style = Border.BORDER_THIN
           c.style.borders.right.border_style = Border.BORDER_THIN
         c.style.borders.bottom.border_style = Border.BORDER_THIN
      else :
         c = cell
         for cl in range(cc+1, ecc+1) :
           c = sheet.cell(row=sr1, column=cl)
           c.style.borders.top.border_style = Border.BORDER_THIN
           c.style.borders.bottom.border_style = Border.BORDER_THIN
         c.style.borders.right.border_style = Border.BORDER_THIN

   cell.style.borders.left.border_style = Border.BORDER_THIN
   cell.style.borders.right.border_style = Border.BORDER_THIN
   cell.style.borders.top.border_style = Border.BORDER_THIN
   cell.style.borders.bottom.border_style = Border.BORDER_THIN

def drawHead(sheet, items, sr1, cc, titles):
   setTitle(sheet, sr1, cc, sr1+1, cc, "№")
   cc += 1

   setTitle(sheet, sr1, cc, sr1+1, cc, "Время")
   cc += 1

   setTitle(sheet, sr1, cc, sr1+1, cc, "Наименование ТТ")
   cc += 1

   setTitle(sheet, sr1, cc, sr1+1, cc, "Адрес")
   cc += 1

   for i in items :
      setTitle(sheet, sr1, cc, sr1, cc + len(titles)-1, i.name)

      nr = sr1+1
      for title in titles :
         setTitle(sheet, nr, cc, nr, cc, title)
         cc += 1

   setTitle(sheet, sr1, cc, sr1+1, cc, "Комментарии")

   return cc

def findItem(doc, itemID) :
   for i in doc.items :
     if i.id == itemID :
        return i

   return None
     
def drawData(sheet, docs, items, agentID, orgs, dateS, dateE, sr1, sc, isOur, getData):
   cr = sr1 + 2
   idx = 1
   for doc in docs :
      if doc.userid != agentID or doc.created < dateS or doc.created > dateE :
         continue

      cc = sc

      cell = sheet.cell(row=cr, column=cc)
      cell.value = idx
      cc += 1

      cell = sheet.cell(row=cr, column=cc)
      cell.value = doc.created.strftime("%d/%m/%Y %H:%M")
      cc += 1

      name = "объект с кодом '" + doc.id + "'"
      address = ""
      if orgs.has_key(doc.id) :
         o = orgs[doc.id]
         name = o.name
         address = o.address

      cell = sheet.cell(row=cr, column=cc)
      cell.value = name
      cc += 1

      cell = sheet.cell(row=cr, column=cc)
      cell.value = address
      cc += 1

      shift = 4 if isOur else 6
      for i in items :
         item = findItem(doc, i.id)
         if item != None and len(item.items) >= 4 :
            if not isOur :
               sheet.cell(row=cr, column=cc).value = item.face
               cc += 1
               sheet.cell(row=cr, column=cc).value = item.sku
               cc += 1
            for ci in range(4) :
               value = getData(item.items[ci])
               if value != 0 :
                 cell = sheet.cell(row=cr, column=cc)
                 cell.value = value
               cc += 1
         else :
            cc += shift

      cell = sheet.cell(row=cr, column=cc)
      cell.value = doc.remark
      cc += 1
      idx += 1
      cr += 1

   return cr

def doReport(orgs, price, docs, agents, dateS, dateE, isOur, isConcurents, repName, outObj):
   wb = Workbook(False, 'cp1251')

   items = []
   concItems = []
   for p in price.itervalues() :
      if (p.flags & 1) > 0 :
         items.append(p)
      else :
         concItems.append(p)
   items.sort(key=attrgetter('name'))
   concItems.sort(key=attrgetter('name'))

   sheet = None
   sheet2 = None
   sheet3 = None

   if isOur > 0 :
      sheet = wb.get_active_sheet()
      sheet.title = "Суммы"

      sheet2 = wb.create_sheet()
      sheet2.title = "Количества"

   if isConcurents > 0 :
      if isOur > 0 :
         sheet3 = wb.create_sheet()
         sheet3.title = "Конкуренты"
      else :
         sheet3 = wb.get_active_sheet()
         sheet3.title = "Конкуренты"

   sr1 = 0
   sr2 = 0
   sr3 = 0
   for agent in agents.values() :
      if isOur > 0 :
         sheet.cell(row=sr1, column=0).value = "Ф.И.О."
         sheet.cell(row=sr1, column=1).value = agent.name
         sr1 += 1
         sheet.cell(row=sr1, column=0).value = "Ежедневный отчет"
         sr1 += 2

         lastCell = drawHead(sheet, items, sr1, 0, ["0.25", "0.5", "0.7", "1"])
         lastRow = drawData(sheet, docs, items, agent.id, orgs, dateS, dateE, sr1, 0, True, lambda x: x.cost )
         sr1 = lastRow + 2

         sheet2.cell(row=sr2, column=0).value = "Ф.И.О."
         sheet2.cell(row=sr2, column=1).value = agent.name
         sr2 += 1
         sheet2.cell(row=sr2, column=0).value = "Ежедневный отчет"
         sr2 += 2

         lastCell = drawHead(sheet2, items, sr2, 0, ["0.25", "0.5", "0.7", "1"])
         lastRow = drawData(sheet2, docs, items, agent.id, orgs, dateS, dateE, sr2, 0, True, lambda x: x.qty )
         sr2 = lastRow + 2

      if isConcurents > 0 :
         sheet3.cell(row=sr3, column=0).value = "Ф.И.О."
         sheet3.cell(row=sr3, column=1).value = agent.name
         sr3 += 1

         lastCell = drawHead(sheet3, concItems, sr3, 0, ["Фейс (кол-во)", "SKU (0.25-1.0)", "0.25", "0.5", "0.7", "1"])
         lastRow = drawData(sheet3, docs, concItems, agent.id, orgs, dateS, dateE, sr3, 0, False, lambda x: x.cost )
         sr3 = lastRow + 2

   fileName = tempfile.gettempdir() + '/' + repName
   wb.save(fileName)

   file = io.open(fileName, 'rb')
   bytes = file.read(-1)
   file.close()

   obj = outObj.New()
   obj.name = repName
   obj.file = bytes


def run(server):

   print "start"

   # get list of all params
   params = server.Params

   param = params[0]

   uidFilter = ""
   if param.divisionID != 0 :
      agents = loadAgents(server, param.divisionID)
      uidFilter = makeIDStr(server, agents)
   else:
      agents = loadAgent(server, param.agentID)
      uidFilter = '"userid" in (' +"'" + param.agentID + "')"

   orgs = server.Get("Org", uidFilter, "id")
   orgs.update(server.Get("PotenzialOrg", uidFilter, "id"))

   price = server.Get("MonitoringItem", "", "id")

   edate = param.dateEnd + datetime.timedelta(days=1)
   dateFilter = '"date" >= ToDate(\'' + param.date.strftime("%d/%m/%Y") + '\') and "date" < ToDate(\'' + edate.strftime("%d/%m/%Y") +"')"
   docs = server.Get("Monitoring", uidFilter + " and " + dateFilter)

   server.RegisterType("Result[name:s,file:b]")
   outObj = server.New("Result")

   doReport(orgs, price, docs, agents, param.date, edate, param.isOur, param.isConcurents, "res.xlsx", outObj)

   server.Put(outObj)

   print "done " +  str(len(docs))

   
