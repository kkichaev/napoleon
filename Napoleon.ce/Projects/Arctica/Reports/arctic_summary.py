# -*- coding: cp1251 -*-
import logging
import sys
from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment


def makeWB(params):
    HEADER_COLOR = Color("FFA8A8A8")
    DIV_COLOR = Color("FFD8D8D8")
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilder()
    
    c = sheet.cell(row=0, column=0)
    c.value = "Итоговый отчёт подразделения"  
    c.style.font.bold = True
    c.style.font.size = 18

    c = sheet.cell(row=2, column=0)
    c.value = "Интервал: c {0} по {1}".format(params.start.strftime('%d.%m.%Y'), params.end.strftime('%d.%m.%Y'))  
    
    r = 4
    arr = ["Подразделение / агент", "визиты", "заявки", "сумма", "результат.", 
           "звонки", "по телефону", "сумма", "результат.", "прогресс"]
    xlb.makeHead(sheet, r, arr, True);
    for ic in range(0, len(arr)):
        c = sheet.cell(row=r, column=ic)
        fill = c.style.fill
        fill.fill_type = Fill.FILL_SOLID
        fill.start_color = HEADER_COLOR
        
    r += 1
    
    total = None
    for i in params.items:
        if total == None:
            total = i
        orders = i.orders - i.byphone
        visits = i.visits - i.calls
        values = [i.name, visits, orders, i.sum - i.phoneSum, "0%" if orders == 0 else str((orders/visits) * 100) + "%",
                  i.calls, i.byphone, i.phoneSum, "0%" if i.calls == 0 else str(i.byphone / i.calls * 100) + "%",
                  i.progress]
        xlb.makeCells(sheet, r, values)
        if i.isDivision > 0 :
            for ic in range(0, len(values)):
                c = sheet.cell(row=r, column=ic)
                fill = c.style.fill
                fill.fill_type = Fill.FILL_SOLID
                fill.start_color = DIV_COLOR
                
        r += 1
    
    if total != None:
        orders = total.orders - total.byphone
        visits = total.visits - total.calls
        values = ["Итого", visits, orders, total.sum - total.phoneSum, "0%" if orders == 0 else str((orders/visits) * 100) + "%", 
                  total.calls, total.byphone, total.phoneSum, "0%" if i.calls == 0 else str(i.byphone / i.calls * 100) + "%",
                  total.progress]
        xlb.makeCells(sheet, r, values)
        for ic in range(0, len(values)):
            c = sheet.cell(row=r, column=ic)
            c.style.font.bold = True
        
    x = 1;
    for w in [40, 15, 15, 15, 15, 15, 15, 15, 15]:
        sheet.column_dimensions[get_column_letter(x)].width = w
        x += 1

    return wb

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    reload(sys)
    sys.setdefaultencoding("cp1251")

    params = server.Params[0]

    wb = makeWB(params)
    
    XLBuilder().workbookToObject(wb, "quest_rep.xlsx", server)                
    logging.info('end')