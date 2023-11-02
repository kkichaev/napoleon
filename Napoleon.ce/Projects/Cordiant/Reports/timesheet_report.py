# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging
import calendar

import datetime

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat
from calendar import monthrange

reload(sys);

class Item:
  pass
  
  def __init__(self):
    self.k = 0
    self.b = 0
    self.a = 0
    self.o = 0
    self.h = 0
  
  def values(self, idx, date):
    res = [idx, self.name, self.divname]
    walkThroughMonth(date, lambda f, x: res.append(getattr(self, f)))
    res.extend([self.notWorkCount, self.k, self.b, self.a, self.o])
    return res
  
def walkThroughMonth(date, func):
  mr = monthrange(date.year, date.month)
  
  for x in range(1, mr[1]+1):
    fn = "day{0}".format(x)
    func(fn, x)
      
class ReportData:
  def __init__(self):
    self.items = []
    pass

class XLB(XLBuilder):
  FIXED_CELL_COLOR = Color("FFB6DDE8")
  
  def styleHolidayCell(self, cell, column):
    d = column - 2
    if d > 0 and d <= monthrange(self.year, self.month)[1]:
      if dateIsHolyiday(self.year, self.month, d):
        fill = cell.style.fill;
        fill.fill_type = Fill.FILL_SOLID
        fill.start_color = self.FIXED_CELL_COLOR
  
  def adjustHeadCell(self, sheet, cell, row, column):
    XLBuilder.adjustHeadCell(self, sheet, cell, row, column)
    self.styleHolidayCell(cell, column)
        
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)

    if (isinstance(value, float) or isinstance(value, int)) and value < 0:
      if value == -1:
        cell.value = "К"
      if value == -2:
        cell.value = "в"
      if value == -3:
        cell.value = "a"
      if value == -4:
        cell.value = "o"        
      if value == -5:
        cell.value = "Б"  
        
    if column > 2:
      cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER 

    if (isinstance(value, float) or isinstance(value, int)) and value == 0:
      cell.value = ""
    
    self.styleHolidayCell(cell, column)

def loadData(data, params, server):
  timeSheet = server.Get("TimeSheetQuery", "ToDate('{0}.{1}.{2}')".format(params.start.day, params.start.month, params.start.year));
  
  for ts in timeSheet:
    item = Item()
    item.divid = ts.divid
    item.id = ts.id
    item.name = ts.name
    item.divname = ts.divname
    item.notWorkCount = 0
    
    if ts.start.year > 1970:
      loadDays(item, params.start, ts)
      item.notWorkCount = ts.notWorkCount
    else: 
      loadDays(item, params.start, None)  
      
    data.items.append(item)  
  
  return data
  
def defaultValue(d, s):
  defHour = 8
  holyday = -2
  
  if dateIsHolyiday(d.year, d.month, s):
    return holyday
  else:
    return defHour
    
def dateIsHolyiday(year, month, day):
  dt = datetime.date(year, month, day).weekday()
  return dt == 5 or dt == 6
  
def loadDays(d, t, s):
  walkThroughMonth(t, lambda fn, x: initAttr(s, fn, x, t, d))
    
def initAttr(src, fn, x, t, dst):
  value = getattr(src, fn) if src != None else defaultValue(t, x)
  setattr(dst, fn, value)  
  
  if value == -1:
    dst.k += 1
    
  if value == -2:
    dst.h += 1  
    
  if value == -3:
    dst.a += 1  
    
  if value == -4:
    dst.o += 1  
    
  if value == -5:
    dst.b += 1  
  
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Лист 1"
    
    xlb = XLB()
    xlb.year = params.start.year
    xlb.month = params.start.month
    
    r = 0
    cell = sheet.cell(row=r, column=0)
    cell.value = "Табель учета рабочего времени за {0} 2021 год".format(calendar.month_name[params.start.month])
    
    r+= 1
    
    arr = ["№","Ф.И.О.", "Подразделение"]
    walkThroughMonth(params.start, lambda fn, x: arr.append(x))
    arr.extend(["вых.","К", "Б", "А", "О"])
    
    xlb.makeHead(sheet, r, arr);
    
    r += 1
    
    idx = 1
    for i in data.items:
      xlb.makeCells(sheet, r, i.values(idx, params.start))
      r += 1
      idx += 1
    
    cc = 0
    for w in [10, 40, 40]:
      cc += 1
      setClmnWidth(sheet, cc, w)
    
    walkThroughMonth(params.start, lambda fn, x: setClmnWidth(sheet, x+cc, 5))    
    
    return wb
    
def setClmnWidth(sheet, let, width):
  sheet.column_dimensions[get_column_letter(let)].width = width
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "timesheet.xlsx", server)                
    logging.info('end')
    
