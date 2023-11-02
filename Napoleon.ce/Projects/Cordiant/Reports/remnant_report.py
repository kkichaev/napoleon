# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

reload(sys);

class Item:
  def values(self):
    return [self.id, self.date.strftime('%d.%m.%Y'), self.yur_name, self.fact_name, self.region, self.city, self.address, self.typePTT, self.specPTT, 
      self.width, self.wall, self.diameter, self.brand, self.subbrand, self.model, self.type_auto, self.season, self.face, self.qty, self.mrcChanged]

class AgentItem:
  def __init__(self):
    self.items = []
    pass
    
class ReportData:
  def __init__(self):
    self.items = []
    pass

class XLB(XLBuilder):
  HEAD_COLOR = "FFD8D8D8"

  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True    
    self.setBackColor(cell, self.HEAD_COLOR)
    
    return column


def loadData(data, params, server):
  for a in params.userids:
    agentItem = AgentItem()
    data.items.append(agentItem)
    aid = a.id
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    price = server.Get("Price", "", "id")
    agentItem.name = server.CurrentUser().name
    server.RestoreUser()
    
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))
    
    remnant = server.Get('OrgRemnants', where)
    
    region = server.Get('OrgRegion', '', 'id')
    city = server.Get('City', '', 'id')
    typePTT = server.Get('TypePTT', '', 'id')
    specPTT = server.Get('SpecPTT', '', 'id')
    
    for d in remnant:
      org = orgs[d.id] if d.id in orgs else None
      
      for di in d.items:
        item = Item()
        agentItem.items.append(item)
        
        item.id = d.id
        item.date = d.created
        item.yur_name = org.nameYur if org != None else 'Контрагент с кодом<{0}>'.format(d.id)
        item.fact_name = org.nameFakt if org != None else 'Контрагент с кодом<{0}>'.format(d.id)
        item.region = ''
        item.city = ''
        item.typePTT = ''
        item.specPTT = ''
        item.address = ''
        item.face = di.face
        item.qty = di.qty
        item.mrcChanged = "да" if di.mrcChanged == 0 else "нет"
        
        if org != None:
          item.region = region[org.regionID].name if org.regionID in region else 'Регион с кодом <{0}>'.format(org.regionID)
          item.city = city[org.cityID].name if org.cityID in city else 'Город с кодом <{0}>'.format(org.cityID)
          item.specPTT = specPTT[org.specpttID].name if org.specpttID in specPTT else 'Принадлежность РТТ с кодом <{0}>'.format(org.specpttID)
          item.typePTT = typePTT[org.typepttID].name if org.typepttID in typePTT else 'Тип РТТ с кодом <{0}>'.format(org.typepttID)
          item.address = org.address
          
          
        item.width = 0
        item.wall = 0
        item.diameter = 0
        item.brand = ''
        item.subbrand = ''
        item.model = ''
        item.type_auto = ''
        item.season = ''
        item.cost = 0
        
        if di.id in price:
          p = price[di.id]
          item.width = p.width
          item.wall = p.wall
          item.diameter = p.diameter
          item.brand = p.brand
          item.subbrand = p.subbrand
          item.model = p.model
          item.type_auto = p.autoType
          item.season = 'Зимняя' if p.season == 1 else 'Летняя' if p.season == 2 else 'Всесезонка'
          item.cost = di.cost
      
  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Ценовой мониторинг"
    
    xlb = XLB()
    
    for ai in data.items:
      sheet.title = ai.name
      
      r = 0
      
      arr = ["Номер клиента в базе","Дата", "Юр. названиеРТТ", "Факт. название РТТ", "Область/край", "Город/поселок", "Улица, дом", "Тип РТТ", "Принадлежность РТТ", "Ширина(мм)", "Высота(%)", "Диаметр(дюйм)", "Бренд", "Суббренд", "Модель", "Тип авто", "Сезонность", "Фейсинг", "Остаток", "Выполнение МРЦ"]
      
      xlb.makeHead(sheet, r, arr);
      
      r += 1
      for i in ai.items:
        xlb.makeCells(sheet, r, i.values())
        r += 1
      
      cc = 1
      for w in [10,12,13,15,15,15,15,15,15,15,10,10,10,15,15,15,15,10,10]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
      sheet = wb.create_sheet()
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "remnant_report.xlsx", server)                
    logging.info('end')
    
