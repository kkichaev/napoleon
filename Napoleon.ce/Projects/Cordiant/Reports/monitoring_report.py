# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

reload(sys);

class Item:
  def values(self):
    return [self.id, self.date.strftime('%d.%m.%Y'), self.visit_type, self.org_name, self.fact_name, self.region, self.city, self.address, self.typePTT, self.specPTT, 
      self.width, self.wall, self.diameter, self.brand, self.subbrand, self.model, self.type_auto, self.season, self.cost]

class ReportData:
  def __init__(self):
    self.items = []
    pass

class XLB(XLBuilder):
  HEAD_COLOR = "FFD8D8D8"

  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True    
    self.setBackColor(cell, self.HEAD_COLOR)
    
    return column


def loadData(data, params, server):
  if len(params.userids) > 0:
    aid = params.userids[0].id
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    price = server.Get("Price", "", "id")
    server.RestoreUser()
    
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))
    
    monitoring = server.Get('CMonitoring', where)
    
    region = server.Get('OrgRegion', '', 'id')
    city = server.Get('City', '', 'id')
    typePTT = server.Get('TypePTT', '', 'id')
    specPTT = server.Get('SpecPTT', '', 'id')
    
    visits = []
    
    for d in monitoring:
      key = d.created.strftime('%Y.%m.%d') + d.id
      vt = 'вторичный'
      
      if not key in visits:
        visits.append(key)
        vt = 'первичный'
      
      org = orgs[d.id] if d.id in orgs else None
      
      for di in d.items:
        item = Item()
        data.items.append(item)
        
        item.id = di.id
        item.date = d.created
        item.visit_type = vt
        item.org_name = org.name if org != None else 'Контрагент с кодом<{0}>'.format(di.id)
        item.fact_name = org.nameFakt if org != None else ''
        item.region = ''
        item.city = ''
        item.typePTT = ''
        item.specPTT = ''
        
        if org != None:
          item.region = region[org.regionID].name if org.regionID in region else 'Регион с кодом <{0}>'.format(org.regionID)
          item.city = city[org.cityID].name if org.cityID in city else 'Город с кодом <{0}>'.format(org.cityID)
          item.specPTT = specPTT[org.specpttID].name if org.specpttID in specPTT else 'Принадлежность РТТ с кодом <{0}>'.format(org.specpttID)
          item.typePTT = typePTT[org.typepttID].name if org.typepttID in typePTT else 'Тип РТТ с кодом <{0}>'.format(org.typepttID)
          item.address = org.address
          
          
        item.width = 0
        item.wall = 0
        item.diameter = 0
        item.brand = ''
        item.subbrand = ''
        item.model = ''
        item.type_auto = ''
        item.season = ''
        item.cost = 0
        
        if di.id in price:
          p = price[di.id]
          item.width = p.width
          item.wall = p.wall
          item.diameter = p.diameter
          item.brand = p.brand
          item.subbrand = p.subbrand
          item.model = p.model
          item.type_auto = p.autoType
          item.season = 'Зимняя' if p.season == 1 else 'Летняя' if p.season == 2 else 'Всесезонка'
          item.cost = di.cost
      
  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Ценовой мониторинг"
    
    xlb = XLB()
    
    r = 0
    arr = ["ЕНС","дата визита", "первичный визит/вторичный визит", "Юр. название РТТ", "Факт. название РТТ", "Область/край", "Город/поселок", "Улица, дом", "Тип РТТ", "Принадлежность РТТ", "Ширина(мм)", "Высота(%)", "Диаметр(дюйм)", "Бренд", "Суббренд", "Модель", "Тип авто", "Сезонность", "цена"]
    xlb.makeHead(sheet, r, arr);
    
    r += 1
    for i in data.items:
      xlb.makeCells(sheet, r, i.values())
      r += 1
    
    cc = 1
    for w in [10,12,13,15,15,15,15,15,15,15,10,10,10,15,15,15,15,10,10]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "monitoring_report.xlsx", server)                
    logging.info('end')
    
