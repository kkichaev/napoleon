# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat
from calendar import monthrange
from datetime import timedelta, datetime
from grsoft.route import AgentRoute

reload(sys);

class MonCell:
  pass
  
class ReportData:
  def __init__(self):
    self.items = []
    self.months = {}
    
  def compileMonCells(self, s, f):
    idx = 0
    while s < f:
      mr = monthrange(s.year, s.month)
      mc = MonCell()
      mc.idx = idx
      mc.title = " {0}.{1} ".format(s.strftime("%m"), s.strftime("%y"))
      self.months[s] = mc
      s = s + timedelta(days=mr[1])
      idx += 1
      
  def monthTitles(self):
    res = [None] * len(self.months)
    
    for mc in self.months.values():
      res[mc.idx] = mc.title
      
    return res  

class Item:
  def __init__(self, s, f):
    self.cells = {}
    
    while s < f:
      self.cells[s] = [0, 0]
      mr = monthrange(s.year, s.month)
      s = s + timedelta(days=mr[1])
  
  def values(self):
    res = [self.username]
    res.extend(self.cellVales())
    return res
    
  def cellVales(self):
    res = []
    keys = sorted(self.cells.keys())
    
    for k in keys:
      res.append(self.cells[k][0])
      res.append(self.cells[k][1])
    
    return res
    
class XLB(XLBuilder):
  FIXED_CELL_COLOR = Color("FFB6DDE8")

  def adjustHeadCell(self, sheet, cell, row, column):
    XLBuilder.adjustHeadCell(self, sheet, cell, row, column)
    fill = cell.style.fill;
    fill.fill_type = Fill.FILL_SOLID
    fill.start_color = self.FIXED_CELL_COLOR
    
    if row == 0 and (column % 2) == 1:
      sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)

    return column

def loadData(data, params, server):
  data.compileMonCells(params.start, params.finish)
  
  for u in params.userids:
    item = Item(params.start, params.finish)
    data.items.append(item)
    
    aid = u.id
    
    agentRoute = AgentRoute(server, aid)
    
    server.ChangeUser("'" + aid + "'")
    item.userid = aid
    item.username = server.CurrentUser().name
    server.RestoreUser()
    
    where = '"userid"={0} and "start" >= ToDate("{1}") and "start" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))
      
    ts = server.Get("TimeSheet", where)  
    
    timeSheet = {}
    
    for t in ts:
      d = t.start
      
      mr = monthrange(d.year, d.month)
      
      for x in range(1, mr[1]+1):
        d = datetime(d.year, d.month, x)
        c = getattr(t, "day{0}".format(x))
        timeSheet[d] = c
    
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))

    docNames = ["VisitInfo", "OrgRemnants", "CMonitoring"]
    
    s = params.start
    
    plans = {}
    facts = {}
    
    while s <= params.finish:
      plans[s] = []
      facts[s] = []
      
      if not s in timeSheet or timeSheet[s] != -2:
        rt = agentRoute.getDayRoute(s)
        d = datetime(s.year, s.month, 1)
        item.cells[d][0] = item.cells[d][0] + len(rt)
        plans[s] = []
        
        for o in rt:
          plans[s].append(o.id)
        
      s = s + timedelta(days=1)
      
      
    for name in docNames:
      docList = server.Get(name, where)
    
      if docList == None:
        continue
      
      for d in docList:
        s = datetime(d.created.year, d.created.month, d.created.day)
        
        if d.id in plans[s] and not d.id in facts[s]:
          facts[s].append(d.id)

    for f in facts:
      ff = facts[f]
      s = datetime(f.year, f.month, 1)
      item.cells[s][1] = item.cells[s][1] + len(ff)
      
  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Визиты"
    
    xlb = XLB()
    
    r = 0
    arr = ["Территория"]
    
    for m in data.monthTitles():
      arr.append(m)
      arr.append("")
      
    xlb.makeHead(sheet, r, arr);
    
    r += 1
    arr = ["Визиты"]
    
    for m in data.monthTitles():
      arr.append("План")
      arr.append("Факт")
      
    xlb.makeHead(sheet, r, arr);  
    
    r += 1
    
    for i in data.items:
      xlb.makeCells(sheet, r, i.values())
      r += 1
    
    cc = 2
    
    arr = ["Факт"]
    for m in data.monthTitles():
      let = get_column_letter(cc)
      sheet.column_dimensions[let].width = 10
      arr.append("=SUM({0}2:{0}{1})".format(let, r))
      
      let = get_column_letter(cc+1)
      sheet.column_dimensions[let].width = 10
      arr.append("=SUM({0}2:{0}{1})".format(let, r))

      cc += 2
      
    xlb.makeCells(sheet, r, arr);
    
    cc = 1
    for w in [30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "visit_report.xlsx", server)                
    logging.info('end')
    
