# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

reload(sys);

class PTypeItem:
  def __init__(self):
    self.items = {}
    
  def values(self):
    res =  [self.name]
    return res
    
  def getItems(self):
    return self.items.values()  
  
class AgentItem:
  def __init__(self):
    self.summer = 0
    self.winter = 0
    self.d17_18_summer = 0
    self.d17_18_winter = 0
    self.lgsh = 0
    
  def values(self):
    return [self.name, self.summer, self.winter, self.d17_18_summer, self.d17_18_winter, self.lgsh]
  
class ReportData:
  def __init__(self):
    self.items = {}
    self.agents = {}
    
  def getItems(self):
    return self.items.values()

class XLB(XLBuilder):
  FIXED_CELL_COLOR = Color("FFB6DDE8")
    
  def adjustHeadCell(self, sheet, cell, row, column):
    XLBuilder.adjustHeadCell(self, sheet, cell, row, column)
    fill = cell.style.fill;
    fill.fill_type = Fill.FILL_SOLID
    fill.start_color = self.FIXED_CELL_COLOR

    return column  

  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border) 

    if column in [1,2,3,4,5]:
      cell.style.number_format._set_format_code('0.0')
      
  def makeFoot(self, sheet, row, titles):
    XLBuilder.makeHead(self, sheet, row, titles)
    
    for x in range(0, len(titles)):
      if x > 0:
        c = sheet.cell(row=row, column=x)
        c.style.number_format._set_format_code('0.0')

    
def putRemnant(doc, data, idx, item):
  if not doc.id in data[idx]:
    data[idx][doc.id] = {}
    
  om = data[idx][doc.id]

  if not item.id in om:
    om[item.id] = []

  om[item.id].append(item.qty)
  
def vals(dict):
    res = 0
    cn = len(dict)
    
    for oid in dict:
      org_dict = dict[oid]
      
      for pid in org_dict:
        price_avg = 0
        
        vals = org_dict[pid]
        
        if len(vals) > 0:
          for v in vals:
            price_avg += 1
            
          res += price_avg / len(vals)  
          
    return res / cn if cn > 0 else 0  

def loadData(data, params, server):
  ptypes = server.Get('TypePTT')
  
  for pt in ptypes:
    pi = PTypeItem()
    pi.id = pt.id
    pi.name = pt.name
    
    data.items[pi.id] = pi
    
    for u in params.userids:
      ai = AgentItem()
      ai.id = u.id
      
      pi.items[ai.id] = ai
      
      if ai.id in data.agents:
        ai.name = data.agents[ai.id]
      else:
        server.ChangeUser("'" + ai.id + "'")
        ai.name = server.CurrentUser().name
        server.RestoreUser()
        data.agents[ai.id] = ai.name
        
      pi.items[ai.id] = ai  
      
      
  for u in params.userids:
    aid = u.id
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))
    remnants = server.Get('OrgRemnants', where)
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    price = server.Get("Price", "", "id")
    server.RestoreUser()
    
    summer = {}
    winter = {}
    d17_18_summer = {}
    d17_18_winter = {}
    lgsh = {}
    
    for r in remnants:
    
      if r.id in orgs:
        org = orgs[r.id]
        
        if not org.typepttID in summer:
          summer[org.typepttID] = {}
          
        if not org.typepttID in d17_18_summer:
          d17_18_summer[org.typepttID] = {}   
          
        if not org.typepttID in winter:
          winter[org.typepttID] = {}
          
        if not org.typepttID in d17_18_winter:
          d17_18_winter[org.typepttID] = {}   

        if not org.typepttID in lgsh:
          lgsh[org.typepttID] = {}
          
        for ri in r.items:
          if ri.id in price:
            p = price[ri.id]
            
            if (ri.qty > 0 or ri.face > 0) and p.docFilter == 1 and p.season == 2:
              putRemnant(r, summer, org.typepttID, ri)

            if (ri.qty > 0 or ri.face > 0)and p.docFilter == 1 and p.season == 2 and p.diameter in [17,18]:
              putRemnant(r, d17_18_summer, org.typepttID, ri)
  
            if (ri.qty > 0 or ri.face > 0) and p.docFilter == 1 and p.season == 0:
              putRemnant(r, winter, org.typepttID, ri)              
              
            if (ri.qty > 0 or ri.face > 0) and p.docFilter == 1 and p.season == 0 and p.diameter in [17,18]:
              putRemnant(r, d17_18_winter, org.typepttID, ri)
              
            if (ri.qty > 0 or ri.face > 0) and p.docFilter == 1 and p.autoType == 'ЛГШ':
              putRemnant(r, lgsh, org.typepttID, ri)
  
    for pt in data.items:
      pti = data.items[pt]
      
      if aid in pti.items:
        ai = pti.items[aid]
        ai.summer = vals(summer[pt]) if pt in summer else 0
        ai.winter = vals(winter[pt]) if pt in winter else 0
        ai.d17_18_summer = vals(d17_18_summer[pt]) if pt in d17_18_summer else 0
        ai.d17_18_winter = vals(d17_18_winter[pt]) if pt in d17_18_winter else 0
        ai.lgsh = vals(lgsh[pt]) if pt in lgsh else 0
        
  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "КБ offline"
    
    xlb = XLB()
    
    r = 0
    arr = ["SKU Лето", "SKU Зима", "17-18 Лето", "17-18 Зима", "ЛГШ"]
    
    for i in data.getItems():
      head = i.values()
      head.extend(arr)
      xlb.makeHead(sheet, r, head);
      r+=1
      
      for i2 in i.getItems():
        xlb.makeCells(sheet, r, i2.values())
        r += 1
    
    pos = 2
    
    for i in data.getItems():
      head = i.values()
      s1 = pos
      s2 = pos+len(i.items)-1
      arr = ["=AVERAGE(B{0}:B{1})".format(s1, s2),
             "=AVERAGE(C{0}:C{1})".format(s1, s2),
             "=AVERAGE(D{0}:D{1})".format(s1, s2),
             "=AVERAGE(E{0}:E{1})".format(s1, s2),
             "=AVERAGE(F{0}:F{1})".format(s1, s2)]
      
      head.extend(arr)
      xlb.makeFoot(sheet, r, head);
      r+=1
      pos += 1 + len(i.items)
    
    cc = 1
    
    for w in [20, 20, 20, 20, 20, 20]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "kboffline_report.xlsx", server)                
    logging.info('end')
    
