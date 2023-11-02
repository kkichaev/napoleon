# -*- coding: cp1251 -*-
from importlib import reload
from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
#sys.setdefaultencoding("cp1251")

def borders(sheet, sr, c1, c2):
    for cl in range(c1, c2) :
       c = sheet.cell(row=sr, column=cl)
       c.style.borders.top.border_style = Border.BORDER_THIN
       c.style.borders.bottom.border_style = Border.BORDER_THIN
       c.style.borders.right.border_style = Border.BORDER_THIN 

class DataItem:
    org = None
    agent = None
    created = None
    item = None
    start = None
    finish = None
    
class Data:
    items = None
    
    def __init__(self):
        self.items = list()
        
    def load(self, server):
        begin = server.Params[0].begin
        end = server.Params[0].end
        userids = server.Params[0].userids
        
        price = server.Get("Price", "", 'id')
        where = '"created" > ToDate("{0}") and "created" <= ToDate("{1}")  and "userid" in ({2})'.format(
            begin.strftime("%d/%m/%Y 0:0:0"), end.strftime("%d/%m/%Y 23:59:59"), userids)
   
        merch = server.Get("MerchEnd", where)
        where = '"userid" in ({0})'.format(userids)
        orgs = server.Get("Org", where, "id")
        users = server.Get("Agents", "", "id")
        
        if merch != None:
            for m in merch:
                for i in m.items:
                    item = DataItem()
                    item.created = m.created
                    
                    if m.id in orgs and orgs != None:
                        item.org = orgs[m.id].name
                    else:
                        item.org = "контрагент <{0}>".format(m.id);
                     
                    if m.userid in users and users != None:
                        item.agent = users[m.userid].name 
                    else:
                        item.agent =  "торговый <{0}>".format(m.userid);
                    
                    if i.id in price and price != None:
                        item.item = price[i.id].name
                    else:
                        item.item = "товар <{0}>".format(i.id)
                        
                    item.start = i.start
                    item.finish = i.finish
                    self.items.append(item)         
        
def doReport(data):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.cell(row=0, column=0).value = "дата"
    sheet.cell(row=0, column=1).value = "агент"
    sheet.cell(row=0, column=2).value = "организация"
    sheet.cell(row=0, column=3).value = "товар"
    sheet.cell(row=0, column=4).value = "начало"
    sheet.cell(row=0, column=5).value = "конец"
    
    r = 1
    for i in data.items:
        sheet.cell(row=r, column=0).value = i.created
        sheet.cell(row=r, column=1).value = i.agent
        sheet.cell(row=r, column=2).value = i.org
        sheet.cell(row=r, column=3).value = i.item
        sheet.cell(row=r, column=4).value = i.start
        sheet.cell(row=r, column=5).value = i.finish
        r = r + 1
             
    repName = "merch.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName) 
    return fileName
               
def run(server):
    print (__name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    print ("")
    
    #getcontext().prec = 2 
    #getcontext().rounding = ROUND_05UP
    
    params = server.Params
    param = params[0]
    
    data = Data()
    data.load(server)
    fileName= doReport(data)
    
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
 
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
 
    obj = outObj.New()
    obj.name ="merchreport.xlsx" 
    obj.file = bytes
    
    server.Put(outObj)
    
    print (__name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))