# -*- coding: cp1251 -*-
from importlib import reload
from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
#sys.setdefaultencoding("cp1251")

def borders(sheet, sr, c1, c2):
    for cl in range(c1, c2) :
       c = sheet.cell(row=sr, column=cl)
       c.style.borders.top.border_style = Border.BORDER_THIN
       c.style.borders.bottom.border_style = Border.BORDER_THIN
       c.style.borders.right.border_style = Border.BORDER_THIN 

class DataItem:
    org = None
    done = None
    create = None
    
    def __init__(self):
        self.org = ""
        self.done = 0
        self.create = 0
    
class Data:
    items = None
    begin = None
    end = None
    agent = None
    
    def __init__(self):
        self.items = list()
        
    def load(self, server):
        self.begin = server.Params[0].start
        self.end = server.Params[0].finish
        userid = server.Params[0].userid
        
        price = server.Get("Price", "", 'id')
        where = '"created" > ToDate("{0}") and "created" <= ToDate("{1}")  and "userid" = \'{2}\''.format(
            self.begin.strftime("%d/%m/%Y 0:0:0"), self.end.strftime("%d/%m/%Y 23:59:59"), userid)
   
        atask = server.Get("ATask", where)
        mtask = server.Get("MTask", where)
        answer = server.Get("TaskAnswer", where)
        
        where = '"userid" in (\'{0}\')'.format(userid)
        orgs = server.Get("Org", where, "id")
        
        agents = server.Get("Agents", "", "id")
        
        if userid in agents:
            self.agent = agents[userid].name
        else:
            self.agent = userid    
        
        collect = dict();
        
        if atask != None:
            for a in atask:
                if not a.id in collect:
                    collect[a.id] = DataItem()
                
                i = collect[a.id]
                if orgs != None and a.id in orgs:
                    i.org = orgs[a.id].name
                else:
                    i.org = "контрагент <{0}>".format(a.id);
                
                i.create = i.create + 1
                
        if mtask != None:
            for m in mtask:
                if not m.id in collect:
                    collect[m.id] = DataItem()     
                    
                i = collect[m.id]
                if orgs != None and m.id in orgs:
                    i.org = orgs[m.id].name
                else:
                    i.org = "контрагент <{0}>".format(m.id);
                
                i.create = i.create + 1    
                
        if answer != None:
            for a in answer:
                if not a.id in collect:
                    collect[a.id] = DataItem()
                
                i = collect[a.id]
                if orgs != None and a.id in orgs:
                    i.org = orgs[a.id].name
                else:
                    i.org = "контрагент <{0}>".format(a.id);
                
                i.done = i.done + 1
                
        self.items.extend(collect.values())  
        self.items = sorted(self.items, key = lambda x: x.org)      
                        
def doReport(data):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    r = 0
    sheet.cell(row=r, column=0).value = "{0} период {1} - {2}".format(data.agent, data.begin.strftime("%d/%m/%Y"), data.end.strftime("%d/%m/%Y"))

    color = 'C0C0C0'
    
    r = r + 2
    sheet.cell(row=r, column=0).value = "Организация"
    sheet.cell(row=r, column=0).style.fill.fill_type = Fill.FILL_SOLID
    sheet.cell(row=r, column=0).style.fill.start_color.index = color
    sheet.cell(row=r, column=1).value = "Поставлено"
    sheet.cell(row=r, column=1).style.fill.fill_type = Fill.FILL_SOLID
    sheet.cell(row=r, column=1).style.fill.start_color.index = color
    sheet.cell(row=r, column=2).value = "Выполнено"
    sheet.cell(row=r, column=2).style.fill.fill_type = Fill.FILL_SOLID
    sheet.cell(row=r, column=2).style.fill.start_color.index = color
    borders(sheet, r, 0, 3)
    
    r = r + 1
    for i in data.items:
        sheet.cell(row=r, column=0).value = i.org
        sheet.cell(row=r, column=1).value = i.create
        sheet.cell(row=r, column=2).value = i.done
        borders(sheet, r, 0, 3)
        r = r + 1

    sheet.column_dimensions[get_column_letter(1)].width = 25
    sheet.column_dimensions[get_column_letter(2)].width = 15
    sheet.column_dimensions[get_column_letter(3)].width = 15
             
    repName = "task.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName) 
    return fileName
               
def run(server):
    print (__name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    print ("")
    
    #getcontext().prec = 2 
    #getcontext().rounding = ROUND_05UP
    
    params = server.Params
    param = params[0]
    
    data = Data()
    data.load(server)
    fileName= doReport(data)
    
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
 
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
 
    obj = outObj.New()
    obj.name ="merchreport.xlsx" 
    obj.file = bytes
    
    server.Put(outObj)
    
    print (__name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))