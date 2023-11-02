# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border, NumberFormat
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *
from grsoft.xl_base import XLBuilder
from rmr_report_style import XLBuilderCommon

reload(sys);
#sys.setdefaultencoding("cp1251")

def borders(sheet, sr, c1, c2):
    for cl in range(c1, c2) :
       c = sheet.cell(row=sr, column=cl)
       c.style.borders.top.border_style = Border.BORDER_THIN
       c.style.borders.bottom.border_style = Border.BORDER_THIN
       c.style.borders.right.border_style = Border.BORDER_THIN 

class DataItem:
    name = None
    id = None
    times = None
    
    def __init__(self, a):
      self.name = a.name
      self.id = a.id
      self.times = dict()
    
    def process(self, visit):
      for vi in visit.items:
        if not vi.date.date() in self.times:
          self.times[vi.date.date()] = vi.date
        
        if self.times[vi.date.date()] > vi.date:
          self.times[vi.date.date()] = vi.date

def idsToString(list):
  res = ""

  for i in list:
    if len(res) > 0:
      res += ","
    res += i.id  

  return res  

class Data:
    items = None
    
    def __init__(self):
        self.items = list()
        
    def load(self, server):
        begin = server.Params[0].start
        end = server.Params[0].finish
        userids = idsToString(server.Params[0].userids)
        
        where = '"created" > ToDate("{0}") and "created" <= ToDate("{1}")  and "userid" in ({2})'.format(
            begin.strftime("%d/%m/%Y 0:0:0"), end.strftime("%d/%m/%Y 23:59:59"), userids)
   
        visit = server.Get("VisitNoPhoto", where)
        agents = server.Get("Agents", "", "id")
        
        data = dict()
        
        if visit != None:
            for v in visit:
              if not v.userid in agents:
                continue
              
              if not v.userid in data:
                data[v.userid] = DataItem(agents[v.userid])
                
              di = data[v.userid]
              di.process(v)
        
        for u in userids.split(","):
          u = u[1:-1]
          
          if u in agents and not u in data:
            data[u] = DataItem(agents[u])
        
        self.items.extend(data.values()) 
        self.items = sorted(self.items, key = lambda x: x.name)
        
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)     
    
    if column > 0:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME3)        
    
def doReport(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.cell(row=0, column=0).value = "Время начала работы"
    sheet.cell(row=1, column=0).value = "Подразделение: {0}".format(data.divname) 
    
    head = []
    head.append("Агент")
    
    s = params.start.date()
    
    while s <= params.finish.date():
      head.append(s.strftime("%d.%m.%Y"))
      s = s + timedelta(days=1)
      
    head.append("Итого")  
    
    xlb = XLBuilderEx()
    xlb.makeHead(sheet, 2, head)
    
    r = 3
    for i in data.items:
      s = params.start.date()
    
      cells = [i.name]
      
      while s <= params.finish.date():
        c = ""
        if s in i.times:
          c = i.times[s].time()
        
        cells.append(c)
        s = s + timedelta(days=1)
      
      c = "=IFERROR(AVERAGEIF({1}{0}:{2}{0},\">0\"), \" \")".format(r+1, get_column_letter(2), get_column_letter((params.finish.date() - params.start.date()).days + 2))
      cells.append(c)
      
      xlb.makeCells(sheet, r, cells)
      
      r = r + 1
    
    sheet.column_dimensions[get_column_letter(1)].width = 25
    
    for x in range(2, (params.finish.date() - params.start.date()).days + 4):
      sheet.column_dimensions[get_column_letter(x)].width = 12

    return wb  

def loadData(params, server):
  data = Data()
  data.load(server)    
  rootDivision = server.Get('Division', '"id"=' + str(params.divid))
  data.divname = rootDivision[0].name
  return data

def printOut(data, params):
  return doReport(data, params)  
               
def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilderCommon().workbookToObject(wb, "start_work_report.xlsx", server)                
  logging.info('end')

    # print __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    # data = Data()
    # data.load(server)
    # fileName= doReport(data, server.Params[0])
    
    # server.RegisterType("Result[name:s,file:b]")
    # outObj = server.New("Result")
 
    # file = io.open(fileName, 'rb')
    # bytes = file.read(-1)
    # file.close()
 
    # obj = outObj.New()
    # obj.name ="merchreport.xlsx" 
    # obj.file = bytes
    
    # server.Put(outObj)
    
    # print __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')