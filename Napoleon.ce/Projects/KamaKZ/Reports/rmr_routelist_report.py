# -*- coding: cp1251 -*-
from importlib import reload
import logging
import sys;
from openpyxl import Workbook
from rmr_report_style import XLBuilderCommon
from rmr_visit_report import Data, AgentSheet
from datetime import datetime
from openpyxl.cell import get_column_letter
from openpyxl.style import Alignment, Border

reload(sys);
#sys.setdefaultencoding("cp1251")

class Month:
  __slots__ = ["data"]
  
  def __init__(self):
    self.data = [Week(), Week(), Week(), Week()]

class Week:
  __slots_ = ["data"]

  Monday = 'Понедельник'
  Thuesday = 'Вторник'
  Wednesday = 'Среда'
  Thursday = 'Четверг'
  Friday = 'Пятница'
  Saturday = 'Суббота'
  Sunday = 'Воскресенье'
  
  @staticmethod
  def toArray():
    return [Week.Monday, Week.Thuesday, Week.Wednesday, Week.Thursday, Week.Friday, Week.Saturday, Week.Sunday]
    
  def __init__(self):
    self.data = {Week.Monday : [], Week.Thuesday : [], Week.Wednesday : [], Week.Thursday : [], Week.Friday : [], Week.Saturday : [], Week.Sunday : []}
    
class RepData(Data):
  __slots__ = ["allorgs", "agents"]

  def __init__(self):
    Data.__init__(self)
    self.allorgs = dict()
    self.agents = None

def loadData(params, server):
  data = RepData()
  data.agents = server.Get('Agents', '', 'id')
  allorgs = {}

  for ai in params.userids:
    route_month = Month()
    
    page = AgentSheet()
    page.id = id
    data.items.append(page)
    
    server.ChangeUser("'" + ai.id + "'")
    orgs = server.Get("Org", "", "id")
    page.name = server.CurrentUser().name
    folders = server.Get("OrgFolder", "")
    server.RestoreUser()
      
    for k in orgs.keys():
      data.allorgs[k] = orgs[k]
      
    cnt_rows = 0;
    
    for f in folders:
      if cnt_rows < len(f.items):
        cnt_rows = len(f.items)
      
      idx = -1
      
      if f.name[0].isdigit():
        name = f.name[1:]
        idx = int(f.name[0]) - 1
      else:
        name = f.name
      
      for mid in range(0,len(route_month.data)):
        if idx == mid or idx == -1:
          rw = route_month.data[mid]
          
          if name in rw.data:
            list = []
            
            for on in f.items:
              list.append( "{0} ({1})". format(orgs[on.name].name, orgs[on.name].address) if on.name in orgs else on.name)
              
            rw.data[name].extend(sorted(list))
      
    page.items.append(route_month)
      
  data.items = sorted(data.items, key=lambda lhs: lhs.name)  
  
  return data

def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  xlb = XLBuilderCommon()
  
  for page in data.items:
    sheet.title = page.name
    cell = sheet.cell(row=0, column=0)
    cell.value = "Маршуртный лист: {0}".format(page.name)
    cell.style.font.bold = True
    
    titles = ['№']
    titles.extend(Week.toArray())
    
    row = 2
    idx = 1
    
    for m in page.items:
      for w in m.data:
        sheet.cell(row=row, column=0).value = 'Неделя: {0}'.format(idx)
        idx += 1
        row += 1
        xlb.makeHead(sheet, row, titles, True)
        row +=1
        column = 0
        start_row = row
        
        max_row = row
        
        for k in Week.toArray():
          row = start_row
          rn = 1
          
          if k in w.data:
            for c in w.data[k]:
              cell = sheet.cell(row=row, column=0)
              cell.value = rn
              cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
              cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
              rn += 1 
            
              cell = sheet.cell(row=row, column=column + 1)
              cell.value = c
              cell.style.alignment.horizontal = Alignment.HORIZONTAL_LEFT
              cell.style.alignment.vertical = Alignment.VERTICAL_TOP

              cell.style.alignment.wrap_text = True   
              row += 1
              
          column += 1
          
          if max_row < row:
            max_row = row
        
        for r in range(start_row, max_row):
          for c in range(0, len(titles)):
            cell = sheet.cell(row=r, column=c)
            xlb.makeBorder(cell, Border.BORDER_THIN)

        row = max_row + 1
        
      CLMN_SZ = 35  
      cc = 1
      for w in [3, CLMN_SZ,CLMN_SZ,CLMN_SZ,CLMN_SZ,CLMN_SZ,CLMN_SZ,CLMN_SZ]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    sheet = wb.create_sheet()
    
  return wb;
  
def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilderCommon().workbookToObject(wb, "rmr_routelist_report.xlsx", server)                
  logging.info('end')