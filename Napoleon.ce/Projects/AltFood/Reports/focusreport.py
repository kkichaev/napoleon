# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class Item:
    data = None
    org = None
    folder = None
    item = None
    remark = None
    
    def __init__(self):
      self.org = ""
      self.address = ""
      self.data = datetime.now()
      self.finish = datetime.now()
      self.order = 0
      self.incass = 0
      
    def getData(self, row):
      return [self.data.strftime("%d.%m.%Y"), self.org, self.folder, self.item, self.remark]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].userid
    
def loadData(server):
    start, finish, userid = inflateParams(server)
    
    WHERE_STR = '"created" >= ToDate("{0}") and "created" < ToDate("{1}") and "userid"=\'{2}\'';         
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
    orders = server.Get("Order", where)
    agent = ""
    
    server.ChangeUser(userid)
    agent = server.CurrentUser().name
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    price = server.Get("Price", "", "id")
    folders = server.Get("Folder", "", "fid")
    orgs.update(porg)
    server.RestoreUser()

    data = list()
    
    for o in orders:
        if o.focusedFolders != None:
            for fi in o.focusedFolders:
                i = Item()
                i.data = o.created
                i.org = orgs[o.id].name if o.id in orgs else "Контрагент с кодом <{0}>".format(o.id)
                i.folder = folders[fi.fid].name if fi.fid in folders else "Папка с кодом <{0}>".format(fi.fid)
                i.item = ""
                i.remark = fi.remark
                data.append(i)
        if o.focusedItems != None:
            for fi in o.focusedItems:
                i = Item()
                i.data = o.created
                i.org = orgs[o.id].name if o.id in orgs else "Контрагент с кодом <{0}>".format(o.id)
                i.folder = ""
                i.item = price[fi.id].name if fi.id in price else "Товар с кодом <{0}>".format(fi.id)
                i.remark = fi.remark
                data.append(i)
      
    data = sorted(data, cmp=item_cmp)
    
    return data, agent
    
def item_cmp(x, y):
  res = cmp(x.data, y.data)
  
  if res == 0:
      res = cmp(x.org, y.org)

  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data, agent):
    sh.cell(row=0, column=0).value = agent
    sh.cell(row=1, column=0).value = "Непроданный фокусный товар, дата: " + datetime.now().strftime("%d.%m.%Y")
    
    head = ["Дата", "Контрагент", "Папка", "Товар", "Примечание"]
    
    r = 2
    xlb.makeHead(sh, r, head)
    
    for d in data:
      r += 1
      xlb.makeCells(sh, r, d.getData(r))
    
    setCellWidth(sh, [11,45,30,30,30])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
def printOut(d, a):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d, a)
                
    return wb

def doReport(server):
    data, agent = loadData(server)
    wb = printOut(data, agent)
    workbookToObject(wb, "focusreport.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    