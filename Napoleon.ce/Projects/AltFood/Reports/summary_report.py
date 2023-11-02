# -*- coding: cp1251 -*-
import sys;
import logging
import locale

from grsoft import xl_base
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Fill
from openpyxl.style import Color
from openpyxl.style import NumberFormat

reload(sys);
sys.setdefaultencoding("cp1251")

class XLB(xl_base.XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        fill = cell.style.fill;
        fill.fill_type = Fill.FILL_SOLID
        fill.start_color = Color('FFAFAFAF')
        
        return column
 
    def makeTotal(self, sheet, row, values, startColumn=0):
        cc = startColumn
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                self.makeCell(sheet, row, cc, c, value)
                c.style.font.bold = True
            cc += 1
    
    def makeCell(self, sheet, row, column, cell, value):
        xl_base.XLBuilder.makeCell(self, sheet, row, column, cell, value)        
        
        if column == 4:
            cell.style.number_format.format_code = NumberFormat.FORMAT_NUMBER
        
def calcOrderPercent(v, o):
    return o * 100.0 / v if v != 0 else 0

def printOut(params):
    xlb = XLB()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"

    row = 0
    cc = 0
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Итоговый отчёт подразделения ' + params.division
    style = cell.style
    style.font.bold = True
    style.font.size = 14
    
    row += 2
    cell =  sheet.cell(row=row, column=cc)
    cell.value = 'Период: ' + params.start.strftime("%d.%m.%Y") + " - " + params.finish.strftime("%d.%m.%Y")

    row += 2
    xlb.makeHead(sheet, row, ['Подразделение / агент', 'визиты','заявки','сумма', 'процент заявок', 'прогресс'], True)
    
    for item in params.items:
        row += 1
        values = [item.name, item.visit, item.orders, item.sum, calcOrderPercent(item.visit, item.orders), item.progress]
        xlb.makeCells(sheet, row, values)

    if len(params.items) > 0:
        row += 1
        item = params.items[0]
        values = ["Итого", item.visit, item.orders, item.sum, calcOrderPercent(item.visit, item.orders), item.progress] 
        xlb.makeTotal(sheet, row, values)
    
    cc = 1
    wdh = [45,11,11,11,11]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    return wb

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting")

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))

    wb = printOut(params)
    xl_base.XLBuilder().workbookToObject(wb, "summary_report.xlsx", server)                

    logging.info("ended")
