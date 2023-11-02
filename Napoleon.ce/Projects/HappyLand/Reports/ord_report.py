# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

import datetime
from datetime import datetime
from datetime import timedelta

import io
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
from openpyxl.cell import Cell
import sys
import tempfile
import time

reload(sys);
sys.setdefaultencoding("cp1251")

scheduleStart = dict()

def getWeekIndex(server, data, agentid):  
    scStart = None
     
    if agentid in scheduleStart: 
        scStart = scheduleStart[agentid]
    else :
        where = '"userid"' + " in ('" + agentid + "')"
        cfg = server.Get("ServerConfig", where)
        
        if cfg != None:
            for c in cfg:
                if c.key == 'SheduleStart':
                    strptime = lambda date_string, format: datetime(*(time.strptime(date_string, format)[0:6])) 
                    scStart = strptime (c.value, '%Y-%m-%d')
                    break
                
        scheduleStart[scStart] = scStart;
    
    result = -1
    
    if scStart != None:
        d = data - scStart
        result = ((d.days / 7) % 4) + 1;
    
    return result 

def getRoutePerDay(server, orgFolder, a, b1):
    days = {"Monday" : "Понедельник", 
                "Tuesday" : "Вторник",
                "Wednesday" : "Среда",
                "Thursday" : "Четверг",
                "Friday" : "Пятница",
                "Saturday" : "Суббота",
                "Sunday" : "Воскресенье",
        };
    d = days[b1.strftime("%A")]
    widx = getWeekIndex(server, b1, a.id)
    plans = list();
    
    if orgFolder != None:
        for of in orgFolder:
            if of.name == d or of.name == str(widx) + d:
                for i in of.items:
                    if not i.name in plans:
                        plans.append(i.name) 
    return plans 

class ItemItem:
    id = None
    name = None
    sum = None
    
    def __init__(self):
        id = ""
        name = ""
        sum = 0

class Item:
    manager = None
    agent = None
    org = None
    status = None
    result = None
    start = None
    stop = None
    route_time = None
    ord_count = None
    items = None
    number = 0
    visit_time = None
    color = None
        
    def __init__(self):
        self.manager = None
        self.agent = None
        self.org = None
        self.status = None
        self.result = None
        self.start = None
        self.stop = None
        self.route_time = 0
        self.ord_count = 0
        self.number = 0
        self.items = list()
        self.visit_time = 0
        self.color = None
        
class WTData:
    start = None
    stop = None
    prev_finish = None
    visit_time = None
    
    def __init__(self, orgid, wt):
        self.visit_time = 0
        for w in wt:
            if self.start == None and orgid != w.id: 
                self.prev_finish = w.finish
            elif orgid == w.id:
                if self.start == None:
                    self.start = w.start
                 
                self.stop = w.stop
                self.visit_time = self.visit_time + (w.stop - w.start).seconds        
        
        if self.prev_finish == None:
            self.prev_finish = self.start        
            
        self.visit_time = self.visit_time
        
    def time_in_route(self):
        return (self.start - self.prev_finish).seconds   
        
class ReportData:
    start = None
    finish = None
    items = None
        
    def __init__(self, server):
        param = server.Params[0];
        self.start = param.start
        self.finish = param.finish 
        agentMap = server.Get("Agents", "", "id")
        divMap = server.Get("Division","", "id")
        manager = server.Get("DivisionManager")
        orgs = server.Get("Org", "", "id")
        price = server.Get("Price", "", "id")
        
        divDict = dict()
                
        for d in divMap:
            for di in divMap[d].agents:
                if not di.id in divDict:
                    divDict[di.id] = d
        
        manDict = dict()
        
        for m in manager:
            if not m.division in manDict: 
                manDict[m.division] = m.name
                
        agents = list()
                
        for s in param.userid.split(","):
            if s in agentMap:
                agents.append(agentMap[s])
            
        agents = sorted(agents, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
        
        self.items = list()  
        
        manager = None
        manager_order = 0
        
        for a in agents:
            num_doc = 1
            agentIdFilter = '"userid"' + " in ('" + a.id + "')"
            orgFolder = server.Get("OrgFolder", agentIdFilter)  
           
            man = ""
            d = divDict[a.id] if a.id in divDict else None
            
            if d != None:
                man = manDict[d] if d in manDict else ""
            
            start = param.start
            finish = param.finish 
            total_route_time = 0
            total_visit_time = 0
            total_ord_count = 0
            has_data = False
        
            while start < finish:
                wtSql = '"start"' + " >= ToDate('" + start.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                        '"start"' + " <= ToDate('" + start.strftime("%d/%m/%Y 23:59:59") + "') and "+\
                        '"userid"' + " in ('" + a.id + "') order by " + '"start"'
                worktimes = server.Get("WorkTime", wtSql)
                workitems = sorted(worktimes, cmp=lambda lhs, rhs: cmp(lhs.start, rhs.start))
                
                where = '"created" > ToDate("{0}") and "created" <= ToDate("{1}") and "userid"=\'{2}\''.format(start.strftime("%d/%m/%Y 0:0:0"), start.strftime("%d/%m/%Y 23:59:59"), a.id)
                orders = server.Get("Order", where)
                
                docs = dict()
                ord_items = dict()
                ord_cnt = dict() 
                photo_cnt = dict()
                
                for o in orders:
                    if o.id in ord_cnt:
                        ord_cnt[o.id] = ord_cnt[o.id] + 1
                    else:
                        ord_cnt[o.id] = 1
                        
                    if not o.id in docs:
                        docs[o.id] = list()
                        
                    docs[o.id].append(o)
                    
                    if not o.id in ord_items:
                        ord_items[o.id] = list()
                        
                    for i in o.items:
                        ord_items[o.id].append(i)
                   
                visits = server.Get("VisitRpt", where)
                
                for v in visits:
                    if not v.id in docs:
                        docs[v.id] = list()
                        
                    docs[v.id].append(v)
                    
                    pc = len(v.items)
                    
                    if pc > 0:
                        if v.id in photo_cnt:
                            photo_cnt[v.id] = photo_cnt[v.id] + pc
                        else:
                            photo_cnt[v.id] = pc
                    
                plan = getRoutePerDay(server, orgFolder, a, start)
                check_plan = list()
                check_plan.extend(plan)
                day_items = list()
                
                if not has_data:
                    has_data = len(docs) > 0
                    
                for d in docs:
                    if d in check_plan:
                        check_plan.remove(d)
                    
                    rdi = Item()
                    rdi.manager = man
                    rdi.agent = a.name
                    rdi.org = orgs[d].name if d in orgs else d
                    rdi.status = "Плановый" if d in plan else "Внеплановый"
                    
                    result_txt = ""
                    
                    if d in ord_cnt:
                        result_txt = "Зак: " + str(ord_cnt[d])
                    
                    if d in photo_cnt:
                        if len(result_txt) > 0:
                            result_txt = result_txt + " "
                            
                        result_txt = result_txt + "М: " + str(photo_cnt[d]) + " Фото есть"
                    
                    rdi.result = result_txt
                    
                    wtdata = WTData(d, worktimes)
                    rdi.start = wtdata.start
                    rdi.stop = wtdata.stop
                    rdi.route_time = wtdata.time_in_route()
                    total_route_time = total_route_time + rdi.route_time
                    rdi.visit_time = wtdata.visit_time
                    total_visit_time = total_visit_time + rdi.visit_time
                    rdi.ord_count = ord_cnt[d] if d in ord_cnt else 0
                    total_ord_count = total_ord_count + rdi.ord_count
                    rdi.number = num_doc
                    num_doc = num_doc + 1
                    
                    item_list = list()
                    item_doc = dict()
                    
                    if d in ord_items:
                        item_list = ord_items[d]

                        for i in item_list:
                            ii = None
                             
                            if i in item_doc:
                                ii = item_doc[i.id]
                            else:    
                                ii = ItemItem()
                                ii.id = i.id
                                ii.name = price[i.id].name if i.id in price else i.id
                                ii.sum = 0
                                item_doc[i.id] = ii
                                 
                            ii.sum = ii.sum + i.cost * i .qty    
                    
                    item_list = item_doc.values()
                    item_list = sorted(item_list, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
                    rdi.items.extend(item_list)
                            
                    day_items.append(rdi)
                
                for u in check_plan:
                    rdi = Item()
                    rdi.manager = man
                    rdi.agent = a.name
                    rdi.org = orgs[u].name if u in orgs else u
                    rdi.status = "Плановый"
                    rdi.result = "не посетил"
                    rdi.color = 'FFCCFFCC' 
                    
                    self.items.append(rdi)
                    
                self.items.extend(day_items)
               
                start = start + timedelta(days=1);
            
            if has_data:   
                result_item = Item()
                result_item.manager = man
                result_item.agent = a.name + " Итог"
                result_item.route_time = total_route_time
                result_item.visit_time = total_visit_time
                result_item.ord_count = total_ord_count
                result_item.color = Color.YELLOW
            
                self.items.append(result_item) 
                
                if manager == None:
                    manager = man
                
                manager_order = manager_order + total_ord_count
                
                if manager != man:
                    result_item = Item()
                    result_item.manager = man
                    result_item.agent = ""
                    result_item.ord_count = manager_order
                    result_item.color = 'FF00FFFF'
                    manager_order = 0
                    
                    self.items.append(result_item)
        
        if manager != None:
            result_item = Item()
            result_item.manager = manager
            result_item.agent = ""
            result_item.ord_count = manager_order
            result_item.color = 'FF00FFFF'
            
            self.items.append(result_item)
            
def cellCaption(sheet, r, c, value):
    cell = sheet.cell(row=r, column=c)
    cell.value = value
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    
def convertTimeToHuman(second):  
    h = second / 3600
    m = second % 3600 / 60
    s = second % 60
    
    return "{0:02d}:{1:02d}:{2:02d}".format(h,m,s)  

def setBorder(sheet, r, col_cnt, color):
    c = 0
    
    while c < col_cnt:
        cell = sheet.cell(row=r, column=c) 
         
        if color != None:
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = color
             
        cell.style.borders.left.border_style = Border.BORDER_THIN
        cell.style.borders.right.border_style = Border.BORDER_THIN
        cell.style.borders.top.border_style = Border.BORDER_THIN
        cell.style.borders.bottom.border_style = Border.BORDER_THIN
        
        c = c + 1
    
def doReport(data, out):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()

    sheet.cell(row=0, column=0).value = "Отчет по результатам посещений за период: {0} - {1} ".format(data.start.strftime("%d.%m.%Y 00:00:00"), data.finish.strftime("%d.%m.%Y 23:59:59"))
    
    cellCaption(sheet, 1, 0, "Супервайзер")
    cellCaption(sheet, 1, 1, "ТП")
    cellCaption(sheet, 1, 2, "Дата визита")
    cellCaption(sheet, 1, 3, "№")
    cellCaption(sheet, 1, 4, "Контрагент")
    cellCaption(sheet, 1, 5, "Статус визита")
    cellCaption(sheet, 1, 6, "Результат визита")
    cellCaption(sheet, 1, 7, "Время начала визита")
    cellCaption(sheet, 1, 8, "Время конца визита")
    cellCaption(sheet, 1, 9, "Время в пути")
    cellCaption(sheet, 1, 10, "Время в точке")
    cellCaption(sheet, 1, 11, "Номенклатура")
    cellCaption(sheet, 1, 12, "Сумма")
    cellCaption(sheet, 1, 13, "Число заказов")
    cellCaption(sheet, 1, 14, "Комментарий")
    cellCaption(sheet, 1, 15, "Действия СВ")
    
    sheet.row_dimensions[2].height = 45
     
    sheet.column_dimensions[get_column_letter(1)].width = 21
    sheet.column_dimensions[get_column_letter(2)].width = 21
    sheet.column_dimensions[get_column_letter(3)].width = 15
    sheet.column_dimensions[get_column_letter(4)].width = 10
    sheet.column_dimensions[get_column_letter(5)].width = 35
    sheet.column_dimensions[get_column_letter(6)].width = 20
    sheet.column_dimensions[get_column_letter(7)].width = 20
    sheet.column_dimensions[get_column_letter(12)].width = 25
 
    COLUMN_CNT = 16
    setBorder(sheet, 1, COLUMN_CNT, None)
    
    row = 2;
     
    for r in data.items:
        setBorder(sheet, row, COLUMN_CNT, r.color)
        
        sheet.cell(row=row, column=0).value = r.manager
        sheet.cell(row=row, column=1).value = r.agent
        sheet.cell(row=row, column=2).value = "" if r.start == None else r.start.strftime("%d.%m.%Y")
        sheet.cell(row=row, column=3).value = "" if r.number == 0 else r.number
        sheet.cell(row=row, column=4).value = r.org
        sheet.cell(row=row, column=5).value = r.status
        sheet.cell(row=row, column=6).value = r.result
        sheet.cell(row=row, column=7).set_value_explicit("" if r.start == None else r.start.strftime("%H:%M:%S"), data_type=Cell.TYPE_STRING)
        sheet.cell(row=row, column=8).set_value_explicit("" if r.stop == None else r.stop.strftime("%H:%M:%S"), data_type=Cell.TYPE_STRING)
        sheet.cell(row=row, column=9).set_value_explicit(convertTimeToHuman(r.route_time), data_type=Cell.TYPE_STRING)
        sheet.cell(row=row, column=10).set_value_explicit(convertTimeToHuman(r.visit_time), data_type=Cell.TYPE_STRING)
        sheet.cell(row=row, column=13).value = r.ord_count
        
        for i in r.items:
            sheet.cell(row=row, column=11).value = i.name
            sheet.cell(row=row, column=12).value = i.sum
            setBorder(sheet, row, COLUMN_CNT, r.color)
            row = row + 1
             
        if len(r.items) == 0:
            row = row + 1

    repName = "ord_report.xlsx" 
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = out.New()
    obj.name = repName
    obj.file = bytes    
    
def run(server):

    print "start", __name__, datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    type = "Result[name:s,file:b]"
    server.RegisterType(type)
    outObj = server.New("Result")
    
    data = ReportData(server)
    doReport(data, outObj)
    
    server.Put(outObj)
   
    print "done", __name__, datetime.now().strftime('%d/%m/%Y %H:%M:%S')
