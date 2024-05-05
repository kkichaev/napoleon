# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging
import locale

from grsoft import xl_base
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Fill
from openpyxl.style import Color
from openpyxl.style import NumberFormat

reload(sys)
#sys.setdefaultencoding("cp1251")

class XLB(xl_base.XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        fill = cell.style.fill;
        fill.fill_type = Fill.FILL_SOLID
        fill.start_color = Color('FFAFAFAF')
        
        return column
 
    def makeTotal(self, sheet, row, values, startColumn=0):
        cc = startColumn
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                self.makeCell(sheet, row, cc, c, value)
                c.style.font.bold = True
            cc += 1
        

class Totals:
    __slots__ = ['visit','route','outRoute','docs','sum','qty','weight','day','notVisit']
    
    def __init__(self, item):
        self.route = item.routeCount        
        self.visit = list()
        self.outRoute = list()
        self.docs = 0
        self.sum = 0
        self.qty = 0
        self.weight = 0
        self.day = item.workDate
        self.notVisit = 0
        
    def add(self, item):
        visitDoc = item.type != 'Не посетил'
        if item.outRoute != 0:
            if not item.id in self.outRoute: self.outRoute.append(item.id)
        else:
            if visitDoc and not item.id in self.visit: self.visit.append(item.id)
            
        if visitDoc: 
            self.docs += 1
            self.sum += item.sum
            self.qty += item.qty
            self.weight += item.weight
        else: self.notVisit += 1


    def output(self, sheet, row):
        cell = sheet.cell(row=row, column=0)
        cell.value = 'Итого по дню: посетил: {0}, по маршруту {1}, вне маршрута {2}, не посетил {3} документов: {4}, сумма: {5:.2f}р., штук {6}, вес {7} кг'.format(
                   len(self.visit) + len(self.outRoute),
                   len(self.visit),
                   len(self.outRoute),
                   self.notVisit,
                   self.docs,
                   self.sum,
                   self.qty,
                   self.weight
                   )

def printOut(params):
    xlb = XLB()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"

    row = 0
    cc = 0
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Отчет агента ' + params.name
    style = cell.style
    style.font.bold = True
    style.font.size = 14
    
    row += 2
    cell =  sheet.cell(row=row, column=cc)
    cell.value = 'Период: ' + params.start.strftime("%d.%m.%Y") + " - " + params.finish.strftime("%d.%m.%Y")

    row += 2
    total = None
    for item in params.items:
        if total == None or total.day != item.workDate: 
            if total != None:
                row += 1        
                total.output(sheet,row)
                row += 1        
            
            cell = sheet.cell(row=row, column=cc)
            cell.value = 'Дата {0} ({1})'.format(item.workDate.strftime("%d.%m"), item.workDateName)
            row += 2
            
            xlb.makeHead(sheet, row, ['контрагенты', 'тип посещения','по маршруту','дата','время создания','дата передачи','сумма','штук','позиций','вес, кг','комментарий', 'Доверенность, №'], True)
            row += 1
            total = Totals(item)
        else:
            visitDoc = item.type != 'Не посетил'
            values = [item.name, item.type, 'да' if item.outRoute == 0 else 'нет', item.docDate, 
                      item.created.strftime("%H:%M") if visitDoc else '',
                      item.sended.strftime("%d.%m.%Y %H:%M") if visitDoc else '',
                      item.sum  if visitDoc else '',
                        item.qty  if visitDoc else '', item.items if visitDoc else '',
                        item.weight if visitDoc else '',
                        item.remark if visitDoc else '',
                        item.dover ]
            xlb.makeCells(sheet, row, values)
            cell = sheet.cell(row=row,column=6)
            cell.style.number_format.format_code = NumberFormat.FORMAT_NUMBER_00
            row += 1
            total.add(item)

    if total != None:
        row += 1        
        total.output(sheet,row)
    
    cc = 1
    wdh = [45,11,11,11,11,20]
    for w in wdh:
        try:
            sheet.column_dimensions[get_column_letter(cc)].width = w
        except:
            pass
        cc += 1
    return wb

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting")

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))

    wb = printOut(params)
    xl_base.XLBuilder().workbookToObject(wb, "summary_report.xlsx", server)                

    logging.info("ended")
