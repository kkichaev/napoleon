# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Fill
from openpyxl.style import Color
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

class Data:
    name = None
    address  = None
    color = None

class Day:
    name = None
    data = None
    
    def __init__(self):
        self.name = ""
        self.data = list()
    
class Item:
    idx = 0
    week = None
    
    def __init__(self):
        self.week = list()
    
class Report:
    month = None
    org = None
    corg = None
    agent = None 
    items = None
    
    def __init__(self, server):
        w1 = list()
        w2 = list()
        w3 = list()
        w4 = list()
        
        userid = server.Params[0].userid
        where = '"userid" in ({0})'.format("'{0}'".format(userid))
        of = server.Get("OrgFolder", where)
        self.org = server.Get("Org", where, 'id')
        self.corg = server.Get("Org", '', 'id')
        
        users = server.Get("Agents", '', 'id')
        
        if userid in users:
            self.agent = users[userid].name
        else:
            self.agent = userid
        
        self.month = list();
        self.month.append(w1)
        self.month.append(w2)
        self.month.append(w3)
        self.month.append(w4)
        
        for f in of:
            f.items.sort(cmp=lambda x,y: cmp(x.pos, y.pos))
            
            name = f.name
            
            if len(name) > 0:
                idx = name[0]
                
                if idx.isdigit():
                    w = self.month[int(idx)-1]
                    w.append(f)
                    f.name = f.name[1:]
                else:
                    w1.append(f)
                    w2.append(f)
                    w3.append(f)
                    w4.append(f)
                    
        items = list()        
        
        days = {"Понедельник" : 0, "Вторник" : 1, "Среда" : 2, "Четверг" : 3, "Пятница" : 4, "Суббота" : 5, "Воскресенье" : 6}
        self.items = list()
        
        for idx in range(len(self.month)):
            if (idx == 0) or (not weekEquals(self.month[idx-1], self.month[idx])):
                item = Item()
                self.items.append(item)
                item.idx = idx + 1
                week = [None, None, None, None, None, None, None]
                w = self.month[idx]
                for f in w:
                    name = f.name
    
                    if name in days:
                        i = days[name]
                        
                        if i < len(week):
                            ls = Day()
                            item.week.insert(i,ls)
                            ls.name = name
                            
                            for fi in f.items:
                                o = self.getOrg(fi.name)
                                 
                                d = Data()
                                ls.data.append(d)
                                if o != None:
                                    d.name = o.name
                                    d.address = o.address
                                    d.color = hex(int(o.color))
                                else: 
                                    d.name = fi.name
                                    d.color = Color.BLACK
                    
    def getOrg(self, id):
        result = None
        
        if id in self.org:
            result = self.org[id]
        elif id in self.corg:
            result = self.corg[id]
                
        return result                 
                    
def repName():
    return "res.xlsx"

def weekItemsEquals(w1, w2):
    result = len(w1) == len(w2)
                
    if result:
        for ii in range(0, len(w1)):
            result = w1[ii].name == w2[ii].name
            
            if not result:
                break

    return result
        
def weekEquals(w1, w2):
    result = len(w1) == len(w2)
    
    if result:
        wa = list()
        wb = list()

        wa.extend(w1)
        wb.extend(w2)
        
        wa.sort(cmp=lambda x,y: cmp(x.name,y.name))
        wb.sort(cmp=lambda x,y: cmp(x.name,y.name))
        
        for i in range(0,len(wa)):
            result = wa[i].name == wb[i].name
            
            if result:
                result = weekItemsEquals(wa[i].items, wb[i].items)
                
                if not result:
                    break
                            
    return result

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN
               
def createXML(report):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.cell(row=0, column=0).value = report.agent
    sheet.cell(row=0, column=0).style.font.bold = True
    sheet.cell(row=2, column=0).value = "Маршрут"
    sheet.cell(row=2, column=0).style.font.bold = True
    
    r = 3
    for i in report.items:
        sheet.cell(row=r, column=0).value = "Неделя {0}".format(i.idx)
        r = r + 1
        sheet.cell(row=r, column=0).value = "№"
        sheet.cell(row=r, column=0).style.fill.fill_type = Fill.FILL_SOLID
        sheet.cell(row=r, column=0).style.fill.start_color.index = 'C0C0C0'
        
        c = 1
        mr = 0
        
        for w in i.week:
            sheet.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
            sheet.cell(row=r, column=c).value = w.name
            sheet.cell(row=r, column=c).style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
            sheet.cell(row=r, column=c).style.fill.fill_type = Fill.FILL_SOLID
            sheet.cell(row=r, column=c).style.fill.start_color.index = 'C0C0C0'
            
            sheet.column_dimensions[get_column_letter(c+1)].width = 30
            sheet.column_dimensions[get_column_letter(c+2)].width = 30
            
            tr = r + 1
            
            for d in w.data:
                sheet.cell(row=tr, column=0).value = tr - r
                sheet.cell(row=tr, column=c).value = d.name
                sheet.cell(row=tr, column=c).style.font.color.index = d.color
                sheet.cell(row=tr, column=c+1).value = d.address
                sheet.cell(row=tr, column=c+1).style.font.color.index = d.color
                tr = tr + 1
                
                if mr < tr: mr = tr
            
            c = c + 2
            
        rangeBorders(sheet.range("A"+str(r+1)+":"+get_column_letter(c)+str(mr)))    
        r = mr + 1
    
    result = tempfile.gettempdir() + '/' + repName()
    wb.save(result)
    
    return result 

def doReport(server):
    report = Report(server)
    xml = createXML(report)
    
    file = io.open(xml, 'rb')
    bytes = file.read(-1)
    file.close()
 
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    obj = outObj.New()
    obj.name = repName()
    obj.file = bytes
     
    server.Put(outObj)
                   
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')