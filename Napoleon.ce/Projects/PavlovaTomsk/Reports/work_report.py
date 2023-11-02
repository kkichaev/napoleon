# -*- coding: cp1251 -*-
from importlib import reload
import sys
import logging
import locale
import time

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

reload(sys);

class Item:
    def __init__(self):
        self.userid = ''
        self.username = ''
        self.org_in_route = 0.0
        self.order_out_route = 0.0
        self.visit_in_route = 0.0
        self.visit_out_route = 0.0
        self.order_in_route = 0.0
        self.order_sum = 0.0
        now = datetime.now()
        self.start = None
        self.finish = None
        self.aid = ''
        self.incass_count = 0
        self.visit_doc_count = 0
        self.incass_sum = 0.0
        self.min_time_count = 0.0
        self.max_time_count = 0.0
        self.distance = 0.0
        
    def getData(self, row):
      row += 1
      return [self.aid, self.username, self.org_in_route, self.visit_in_route, self.visit_out_route, "=G{0}+H{0}".format(row), 
        self.order_in_route, self.order_out_route,  "=J{0}+K{0}".format(row), self.order_sum, self.incass_count, self.visit_doc_count, self.incass_sum,
         "=IFERROR((G{0})/F{0},0)".format(row), 
         "=IFERROR((J{0} + K{0})/F{0},0)".format(row), 
         "=IFERROR((O{0})/F{0},0)".format(row), 
        "=IFERROR(J{0}/F{0},0)".format(row),
        self.start.strftime("%H:%M") if self.start != None else "", 
        self.finish.strftime("%H:%M") if self.finish != None else "", 
        self.min_time_count, self.max_time_count, self.distance]
      
class ReportData:
    def __init__(self):
        self.data = list()

def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    data = ReportData()

    div = server.Get('Division', '"id"=' + str(params.divid))
    
    if len(div) > 0:
      data.divname = div[0].name    
      
    for aid in params.userids:
      item = Item()
      day_start = list()
      day_finish = list()
      item.aid = aid.id
      
      ar = AgentRoute(server, item.aid)
      
      date = params.start
      
      while date <= params.finish:
        route = ar.getDayRoute(date)  
        
        route_ids = []
        start = None
        finish = None
        
        for i in route:
          route_ids.append(i.id)
        
        item.org_in_route += len(route)

        server.ChangeUser("'" + item.aid + "'")
        item.userid = item.aid
        item.username = server.CurrentUser().name
        server.RestoreUser()
        
        where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
            item.aid,
            date.strftime("%d/%m/%Y 0:0:0"),
            date.strftime("%d/%m/%Y 23:59:59"))
        
        orgs_order_in_route = []
        orgs_order_out_route = []
        visit_in_route = []
        visit_out_route = []

        docNames = ["Order", "VisitInfo", "OrgRemnants", "Answer", "Incass", "TaskDone", "Sales"]
        allDocsAgendPerDay = []
        
        for name in docNames:
          docList = server.Get(name, where)
          
          if docList == None:
            continue

          allDocsAgendPerDay.extend(docList)
          
          if docList != None:
            sz = len(docList)
            
            if name == "Incass":
              item.incass_count += sz
              
            if name == 'VisitInfo':
              vi = []
              for d in docList:
                if d.id in route_ids and not d.id in vi:
                  vi.append(d.id)
                  
              item.visit_doc_count += len(vi)
          
            for d in docList:
              if d.id in route_ids and not d.id in visit_in_route:
                if not d.id in visit_in_route: visit_in_route.append(d.id)
              elif not d.id in route_ids and not d.id in visit_out_route:
                if not d.id in visit_out_route: visit_out_route.append(d.id)
                
              if name == "Order" or name == "Sales":
                if d.id in visit_in_route:
                  if not d.id in orgs_order_in_route: orgs_order_in_route.append(d.id)
                else:
                  if not d.id in orgs_order_out_route: orgs_order_out_route.append(d.id)
                  
                for i in d.items:
                  item.order_sum += i.cost * i.qty  
              
              if name == "Incass":
                item.incass_sum += d.sum
                
              if start == None or start > d.created:
                start = d.created
                
              if finish == None or finish < d.created:
                finish = d.created
        
        allDocsAgendPerDay = sorted(allDocsAgendPerDay, key=lambda doc: doc.created)

        if start != None:
          day_start.append(start)
          day_finish.append(finish)
        
        id = ''
        cr = None
          
        for d in allDocsAgendPerDay:
          if id != d.id and cr != None:   
            td = d.created - cr
            
            elapsed = td.total_seconds();
            
            if elapsed > 60 * 60:
              item.max_time_count += 1
            elif elapsed < 10 * 60:
              item.min_time_count += 1
            
          id = d.id
          cr = d.created
          
      
        item.visit_out_route += len(visit_out_route)
        item.visit_in_route += len(visit_in_route)
        item.order_in_route += len(orgs_order_in_route)
        item.order_out_route += len(orgs_order_out_route)
        date += timedelta(days=1)

      item.distance = computeDist(server, params, item.aid)
      if len(day_start) != 0:
         sumH = 0
         for val in day_start:
           sumH += val.hour * 3600 + val.minute * 60 + val.second
         item.start = datetime(2019, 1, 1, int((sumH / 3600) / len(day_start)), int((sumH % 3600 / 60)/ len(day_start)), int((sumH % 60)/len(day_start)))

         sumH = 0
         for val in day_finish:
           sumH += val.hour * 3600 + val.minute * 60 + val.second
         item.finish = datetime(2019, 1, 1, int((sumH / 3600) / len(day_finish)), int((sumH % 3600 / 60)/ len(day_finish)), int((sumH % 60)/len(day_finish)))

      data.data.append(item)
      
    data.data = sorted(data.data, key=lambda lhs: lhs.username)
    
    return data
    
def computeDist(server, params, userid):

    q = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}") and "userid"=\'{2}\' and "isGSM" = 0'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"), userid)
    
    gpspos = server.Get("GPSPos",q)
    
    lastpos = None
    distance = 0
    td1 = timedelta(0, 0)
    td2 = timedelta(23, 59)
    ctr = 0
    
    if params.time != None and len(params.time) > 0:
      arr = params.time.split("|")
      
      if len(arr) == 2:
        t = arr[0].split(":");
        td1 = timedelta(hours=int(t[0]), minutes=int(t[1]))
        t = arr[1].split(":");
        td2 = timedelta(hours=int(t[0]), minutes=int(t[1]))
        
      
    for pos in gpspos:
      date = pos.date.replace(hour=0, minute=0, second=0, microsecond=0)

      if pos.date < (date + td1) or pos.date > (date + td2):
        continue
        
      if lastpos == None:
          lastpos = pos
          continue
      distance += coordutils.distance(lastpos.latitude, lastpos.longitude, pos.latitude, pos.longitude)
      lastpos = pos

    return distance

    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 13 or column == 15:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
    
    if column >= 16 and column <= 19:
      cell.style.number_format._set_format_code('0%')
      
    if column == 24 and value != 0:
      cell.value = '{0}км {1}м'.format(int(value/1000), int(value%1000))
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 3
    
    head = ['Дата начала', 'Дата конца', 'филиал', 'ИД', 'ТП', 'Итого точек в маршруте', 'По маршруту посещений', 'Не по маршруту посещений', 'Итого посещений', 
      'Заявок по маршруту', 'Заявок не по маршруту', 'Итого Заявок', "Сумма заявок", 'ПКО', 'Фотоотчеты', "Сумма ПКО", "% Посещений", "% заявок", "% Фотоотчетов", 
      "Прогресс", "Начало", "Окончание", "< 10 мин", "> 1 час", "Пробег"]
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      values = [params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"), data.divname]
      values.extend(item.getData(row))
      
      xlb.makeCells(sheet, row, values)
      row += 1
        
        
    cc = 1
    SZ = 12
    for w in [SZ,SZ,SZ,SZ,26,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    