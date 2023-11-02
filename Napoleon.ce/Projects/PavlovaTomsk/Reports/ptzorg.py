# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class Item:
    __slots__ = ['agent', 'org', 'addr', 'fio', 'phone']
    
    def __init__(self):
      self.agent = ""
      self.org = ""
      self.addr = ""
      self.fio = ""
      self.phone = 0
      
    def getData(self):
      return [self.agent, self.org, self.addr, self.fio, self.phone]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].userids
    
def loadData(server):
    start, finish, userid = inflateParams(server)
    
    WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
    WHERE_STR = WHERE_STR_ALL + ' and "userid" in ({2})'; 

    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
    
    orgs = server.Get("PotenzialOrg", where, "id")
    agents = server.Get("Agents", "", "id");
    
    data = list()
    
    for o in orgs.values():
        i = Item()
        i.agent = agents[o.userid].name if o.userid in agents else o.userid
        i.org = o.name
        i.addr = o.address
        
        if len(o.contacts) > 0:
            i.fio = o.contacts[0].name
            i.phone = o.contacts[0].phone 
        
        data.append(i)
    
    data = sorted(data, cmp=item_cmp)
    return data
    
def item_cmp(x, y):
  res = cmp(x.agent, y.agent)

  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    head = ["Торговый представитель", "Название торговой точки", "Адрес", "ВКонтактное лицо", "Номер телефона"]
    
    r = 0
    xlb.makeHead(sh, r, head)
    
    for d in data:
      r += 1
      xlb.makeCells(sh, r, d.getData())
    
    setCellWidth(sh, [30,30,30,30,30])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 5 or column == 6:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME4)
      
    
def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    