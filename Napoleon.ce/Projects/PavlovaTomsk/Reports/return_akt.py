# -*- coding: cp1251 -*-
from importlib import reload
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS, Alignment
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys
reload(sys)
#sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
  __slots__ = ['id', 'pos', 'name', 'item', 'qty', 'sum', 'number', 'date', 'cause']
  
  def __init__(self):
    self.id = ""
    self.pos = 0
    self.name = ""
    self.item = ""
    self.qty = 0.0
    self.sum = 0.0
    self.number = ""
    self.date = ""
    self.cause = ""
      
  def getData(self):
    return [self.pos, self.name, '', '', '', '', self.item, self.qty, self.sum, self.number, self.date, self.cause]
  
class ReportData():
  __slots__ = ['items', 'name', 'fullname', 'address', 'factAddress', 'inn', 'rsh', 'ks', 'bik', 'orgname', 'orgfullname', 'orgaddress', 'orgfactAddress', 'orginn', 'orgrsh', 'orgks', 'orgbik', 'price', 'returns']

  def __init__(self):
    self.items = []
    self.name = ""
    self.fullname = ""
    self.address = ""
    self.factAddress = ""
    self.inn = ""
    self.rsh = ""
    self.ks = ""
    self.bik = ""
    self.orgname = ""
    self.orgfullname = ""
    self.orgaddress = ""
    self.orgfactAddress = ""
    self.orginn = ""
    self.orgrsh = ""
    self.orgks = ""
    self.orgbik = ""
    self.price = None
    self.returns = []
    
def loadData(params, server):
  data = ReportData()
  
  firms = server.Get('Firm', '')
  
  if len(firms) > 0:
    f = firms[0]
    data.name = f.name
    data.fullname = f.fullname
    data.address = f.address
    data.factAddress = f.factAddress
    data.inn = f.inn 
    data.rsh = f.rsh
    data.ks = f.ks
    data.bik = f.bik
  
  where = '"created"= ToDate("{0}") and "userid"="{1}"'.format(params.created.strftime('%d.%m.%Y %H:%M:%S'), params.userid)
  ret = server.Get("Returns", where) 
  price = server.Get("ManagerPrice", '', "id")
  
  server.ChangeUser("'" + params.userid + "'")
  orgs = server.Get("Org", "", "id")
  server.RestoreUser()
  
  data.price = price
  p = 1
  if ret != None and len(ret) > 0:
    r = ret[0]
    
    if r.id in orgs:
      o = orgs[r.id]
      data.orgname = o.name
      data.orgfullname = o.fullName
      data.orgaddress = o.address
      data.orgfactAddress = o.factAddress
      data.orginn = o.inn
      data.orgrsh = o.rsh
      data.orgks = o.ks
      data.orgbik = o.bik
  
    for i in r.items:
      for ii in i.items:
        t = Item()
        t.id = i.id
        t.pos = p
        t.name = price[i.id].name if i.id in price else "Това с кодом <{0}>".format(i.id)
        t.item = "шт"
        t.qty = ii.qty
        t.sum = ii.cost * ii.qty
        t.number = ii.number
        t.date = ii.date.strftime('%d.%m.%Y')
        t.cause = i.cause
        p += 1
      
        data.items.append(t)
  
  rts = server.Get('ReturnCause', '')
  
  for r in rts:
    data.returns.append(r.text)
    
  return data
    
class XLBuilderEx(XLBuilder):
  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 1:
      sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    

def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 8

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Возврат товара разрешаю"
    
    c = sheet.cell(row=0, column=2)
    c.value = "_________________________"
    
    c = sheet.cell(row=0, column=4)
    c.value = "(_________________________)"
    
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
    c = sheet.cell(row=1, column=2)
    c.value = "(подпись)"
    
    sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    c = sheet.cell(row=1, column=4)
    c.value = "(расшифровка подписи)"
    
    c = sheet.cell(row=4, column=0)
    c.value = " "
    sheet.row_dimensions[c.row].height = 50
    
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=10)
    c = sheet.cell(row=4, column=1)
    c.style.alignment.wrap_text = True    
    c.style.font.bold = True
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.value = data.fullname + ', ' + data.factAddress + ',ИНН ' + data.inn + ',р/с ' + data.rsh + ',БИК ' + data.bik + ',к/с ' + data.ks
    
    c = sheet.cell(row=5, column=1)
    c.value = ""

    sheet.merge_cells(start_row=7, start_column=0, end_row=7, end_column=10)
    c = sheet.cell(row=7, column=0)
    c.value = "АКТ"
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.font.bold = True
    
    sheet.merge_cells(start_row=8, start_column=0, end_row=8, end_column=10)
    c = sheet.cell(row=8, column=0)
    c.value = "ВОЗВРАТ ТОВАРА"
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.font.bold = True
    
    c = sheet.cell(row=9, column=0)
    c.value = "Справочно: номер счета-фактуры в ЭДО _________________"
    c.style.font.bold = True
   
    c = sheet.cell(row=10, column=1)
    c.value = "Место приемки товара:"
    c.style.font.bold = True
    sheet.row_dimensions[c.row].height = 20
    
    sheet.merge_cells(start_row=10, start_column=3, end_row=10, end_column=10)
    c = sheet.cell(row=10, column=3)
    c.style.alignment.wrap_text = True
    c.value = data.orgfullname + ', ' + data.orgaddress
    
    c = sheet.cell(row=11, column=1)
    c.value = "Настоящий акт составлен:"
    c.style.alignment.wrap_text = True 
    c.style.font.bold = True
    sheet.row_dimensions[c.row].height = 20
    
    sheet.merge_cells(start_row=11, start_column=2, end_row=11, end_column=10)
    
    c = sheet.cell(row=12, column=1)
    c.value = "Покупатель (Грузоотправитель):"
    c.style.alignment.wrap_text = True 
    c.style.font.bold = True
    sheet.row_dimensions[c.row].height = 40
    
    sheet.merge_cells(start_row=12, start_column=2, end_row=12, end_column=10)
    c = sheet.cell(row=12, column=2)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.wrap_text = True
    c.value = data.orgname + ', ' + data.orgfactAddress +  ',ИНН ' + data.orginn + ',р/с ' + data.orgrsh + ',БИК ' + data.orgbik + ',к/с ' + data.orgks
    
    c = sheet.cell(row=13, column=1)
    c.value = "Поставщик (Грузополучатель):"
    c.style.alignment.wrap_text = True 
    c.style.font.bold = True
    sheet.row_dimensions[c.row].height = 40
    
    sheet.merge_cells(start_row=13, start_column=2, end_row=13, end_column=10)
    c = sheet.cell(row=13, column=2)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.wrap_text = True
    c.value = data.fullname + ', ИНН ' + data.inn + ', ' + data.address + ',р/с ' + data.rsh + ',БИК ' + data.bik + ',к/с ' + data.ks
    
    c = sheet.cell(row=16, column=1)
    c.value = "Наименование,количество осмотренной продукции (товара) и характеристика выявленных дефектов"
    
    c = sheet.cell(row=18, column=0)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.value = "№"
    
    sheet.merge_cells(start_row=18, start_column=1, end_row=18, end_column=5)
    c = sheet.cell(row=18, column=1)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.value = "Товар"
    
    c = sheet.cell(row=18, column=6)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.value = "Ед. из"
    
    sheet.merge_cells(start_row=18, start_column=7, end_row=18, end_column=8)
    c = sheet.cell(row=18, column=7)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "Возврат товара"
    
    sheet.merge_cells(start_row=18, start_column=9, end_row=18, end_column=10)
    c = sheet.cell(row=18, column=9)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "Документ-основание"
    
    c = sheet.cell(row=18, column=11)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.wrap_text = True
    c.value = "Причина возврата"
    
    sheet.row_dimensions[c.row].height = 22
    
    sheet.merge_cells(start_row=19, start_column=1, end_row=19, end_column=5)
    c = sheet.cell(row=19, column=1)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.value = "наименование, характеристика, сорт, артикул товара"
    
    c = sheet.cell(row=19, column=7)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "Кол-во"
    
    c = sheet.cell(row=19, column=8)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.wrap_text = True
    c.value = "сумма, в т.ч. НДС"
    
    c = sheet.cell(row=19, column=9)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.wrap_text = True
    c.value = "Номер"
    
    c = sheet.cell(row=19, column=10)
    c.style.alignment.vertical = Alignment.VERTICAL_TOP
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.wrap_text = True
    c.value = "Дата"
    
    sheet.row_dimensions[c.row].height = 22
    
    c = sheet.cell(row=20, column=0)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "1"
    
    sheet.merge_cells(start_row=20, start_column=1, end_row=20, end_column=5)
    c = sheet.cell(row=20, column=1)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "2"
    
    c = sheet.cell(row=20, column=6)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "3"
    
    c = sheet.cell(row=20, column=7)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "4"
    
    c = sheet.cell(row=20, column=8)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "5"
    
    c = sheet.cell(row=20, column=9)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "6"
    
    c = sheet.cell(row=20, column=10)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "7"
    
    c = sheet.cell(row=20, column=11)
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.value = "8"
    
    xlb = XLBuilderEx()
    row = 21
    
    for i in data.items:
      xlb.makeCells(sheet, row, i.getData())
      row += 1
    
    sheet.merge_cells(start_row=row, start_column=0, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=0)
    c.value = "Итого по накладной"
    
    for r in range(18, row+1):
      for c in range(0, 12):
        cell = sheet.cell(row=r, column=c)
        xlb.makeBorder(cell, Border.BORDER_THIN)

    if len(data.items) > 0:
      c = sheet.cell(row=row, column=7)
      c.value="=SUM(H{0}:H{1})".format(row-len(data.items)+1, row)
      c = sheet.cell(row=row, column=8)
      c.value="=SUM(I{0}:I{1})".format(row-len(data.items)+1, row)
    
    row += 2
    c = sheet.cell(row=row, column=0)
    c.value = "Причины:"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "1- отказ клиента от заказа"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "2- отказ от товара по срокам годности"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "3- потеря товарного вида"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "4- брак в процессе доставки"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "5- просрочка"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "6- брак производственный"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "7- закрытие ТТ"
    
    row += 1
    c = sheet.cell(row=row, column=1)
    c.value = "8 - не продается"
    
    
    # idx = 4
    
    # for r in data.returns:
      # row += 1
      # c = sheet.cell(row=row, column=1)
      # c.value = '{0} - {1}'.format(idx, r)
      # idx += 1
    
    row += 3
    c = sheet.cell(row=row, column=0)
    c.value = "Покупатель:"
    c.style.font.bold = True
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.value = data.orgfullname
    c.style.font.bold = True
    
    c = sheet.cell(row=row, column=7)
    c.value = "Поставщик:"
    c.style.font.bold = True
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.value = data.name
    c.style.font.bold = True
    
    sheet.row_dimensions[c.row].height = 22
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "Адрес юрид.:"
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.value = data.orgfactAddress
    
    c = sheet.cell(row=row, column=7)
    c.value = "Адрес юрид.:"
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.value = data.address
    
    sheet.row_dimensions[c.row].height = 22
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "Адрес факт.:"
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.value = data.orgaddress
    
    c = sheet.cell(row=row, column=7)
    c.value = "Адрес факт.:"
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.value = data.factAddress
    
    sheet.row_dimensions[c.row].height = 22
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "ИНН:"
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.orginn + ' '
    
    c = sheet.cell(row=row, column=7)
    c.value = "ИНН:"
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.inn + ' '
    
    sheet.row_dimensions[c.row].height = 22
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "р/сч:"
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.orgrsh + ' '
    
    c = sheet.cell(row=row, column=7)
    c.value = "р/сч:"
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.rsh + ' '
    
    sheet.row_dimensions[c.row].height = 22
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "к/сч:"
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.orgks + ' '
    
    c = sheet.cell(row=row, column=7)
    c.value = "к/сч:"
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.ks + ' '
    
    sheet.row_dimensions[c.row].height = 22
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "БИК:"
    
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    c = sheet.cell(row=row, column=2)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.orgbik + ' '
    
    c = sheet.cell(row=row, column=7)
    c.value = "БИК:"
    
    sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    c = sheet.cell(row=row, column=8)
    c.style.alignment.wrap_text = True
    c.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)
    c.value = data.bik + ' '
    
    sheet.row_dimensions[c.row].height = 22
    
    c = sheet.cell(row=row + 2, column=9)
    c.value = "Юрьева Г.А."
    
    row += 3
    c = sheet.cell(row=row, column=0)
    c.value = "______________________"
    c = sheet.cell(row=row, column=2)
    c.value = "(___________________________________)"
    c = sheet.cell(row=row, column=7)
    c.value = "______________________"
    c = sheet.cell(row=row, column=9)
    c.value = "(___________________________________)"
    
    row += 1
    c = sheet.cell(row=row, column=0)
    c.value = "подпись"
    c = sheet.cell(row=row, column=2)
    c.value = "расшифровка"
    c = sheet.cell(row=row, column=7)
    c.value = "подпись"
    c = sheet.cell(row=row, column=9)
    c.value = "расшифровка"
    
    row += 2
    c = sheet.cell(row=row, column=0)
    c.style.font.bold = True
    c.value = "МП"
    c = sheet.cell(row=row, column=7)
    c.style.font.bold = True
    c.value = "МП"
    
    cc = 1
    for w in [6, 19, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    