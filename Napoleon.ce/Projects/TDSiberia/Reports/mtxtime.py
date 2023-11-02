# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class Item:
    org = None
    address = None
    agent = None
    data = None
    matrix = None
    typedoc = None
    time = None
    remark = ""
    
    def __init__(self):
      self.org = ""
      self.address = ""
      self.agent = ""
      self.data = ""
      self.matrix = ""
      self.typedoc = "Заявка"
      self.time = 0
      self.remark = ""
      
    def getData(self):
      return [self.org, self.address, self.agent, self.data.strftime("%d.%m.%Y"), self.matrix, self.typedoc, self.time, self.remark]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].userids
    
def loadData(server):
    start, finish, userids = inflateParams(server)
    
    WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
    WHERE_STR = WHERE_STR_ALL + ' and "userid" in ({2})'; 

    if len(userids) == 0 :
      where = WHERE_STR_ALL.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"))
    else:  
      where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userids)
    
    orders = server.Get("Order", where)
    agents = server.Get("Agents", '', 'id')
    
    allorgs = dict()
    
    for u in userids.split(','):    
      print "agentid:", u
      server.ChangeUser(u)
      orgs = server.Get("Org", "", "id")
      porg = server.Get("PotenzialOrg", "", "id")
      orgs.update(porg)
      server.RestoreUser()
      
      for k in orgs.keys():
        if not k in allorgs:
          allorgs[k] = orgs[k]
    
    time = dict()    
    
    for o in orders:
      k = o.id + o.created.strftime("%d/%m/%Y") + o.userid
      
      for m in o.mtxstage:
        ky = k + m.name 
        p = (m.finish - m.start).seconds
        if not ky in time:
          d = Item()
          d.org = allorgs[o.id].name if o.id in allorgs else "Контрагент с кодом <{0}>".format(o.id)
          d.address = allorgs[o.id].address if o.id in allorgs else ""
          d.agent = agents[o.userid].name if o.userid in agents else "Агент с кодом <{0}>".format(o.userid)
          d.data = o.created
          d.matrix = m.name
          d.remark = m.remark
          time[ky] = d
    
        time[ky].time += p
    
    data = list()
    
    for v in time.values():
      data.append(v)
      
    data = sorted(data, cmp=item_cmp)
    
    return data 
    
def item_cmp(x, y):
  res = cmp(x.data, y.data)

  if res == 0:  
    res = cmp(x.agent, y.agent)
    
  if res == 0:
    res = cmp(x.org, y.org)
    
  if res == 0:
    res = cmp(x.matrix, y.matrix)
    
  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    head = ["Название ТТ", "Адрес", "Торговый представитель", "Дата", "Категория", "Тип документа", "Время", "Комментарий"]
    
    r = 0
    xlb.makeHead(sh, r, head)
    
    for d in data:
      r += 1
      xlb.makeCells(sh, r, d.getData())
    
    if r > 1:
      r += 1
      sh.cell(row=r, column=6).value = "=SUM(G2:G{0})".format(r)
    
    setCellWidth(sh, [30,30,30,30,30,30,30,30])

def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilder()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    