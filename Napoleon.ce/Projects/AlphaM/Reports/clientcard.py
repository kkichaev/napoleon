# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#


from importlib import reload
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
import tempfile
import io


import sys;
reload(sys);
# sys.setdefaultencoding("cp1251")

class CardVal:
    orders = None
    remains = None
    returns = None
    
    def __init__(self):
        self.orders = 0
        self.remains = 0
        self.returns = 0
        
def update(objects, values, priceAxe, dateAxe, orgs, i):
    for o in objects:
        if not o.id in orgs:
            orgs.append(o.id)
            
        if not o.id in values:
            values[o.id] = dict()
            priceAxe[o.id] = dict()
            dateAxe[o.id] = dict()
         
        val = values[o.id]
        pAxe = priceAxe[o.id]
        dAxe = dateAxe[o.id]
        
        dt = o.created.strftime("%d/%m/%Y")
        dAxe[dt] = True
        for oi in o.items:
            key = str(dt) + oi.id
            
            if not key in val:
                val[key] = CardVal()
                
            cv = val[key]     
            if  i == 0:
                cv.orders += oi.qty
            elif i == 1:
                cv.remains += oi.qty
            elif i == 2:
                cv.returns += oi.qty    
                
            pAxe[oi.id] = True
            
def makeHead(sheet, o, dates, r):
    sheet.cell(row=r, column=0).value = "N п/п"
    sheet.cell(row=r, column=1).value = o
    sheet.cell(row=r + 1, column=1).value = o
    sheet.cell(row=r + 2, column=1).value = o
    sheet.cell(row=r, column=2).value = "Продукты"
    sheet.merge_cells(start_row=r, start_column=3, end_row=r, end_column= 3 + len(dates) * 3 - 1)
    
    idx = 0
    for d in dates:
        shift = idx * 3 + 3
        idx = idx + 1 
        sheet.cell(row=r + 1, column=shift).value = d
        #sheet.merge_cells(start_row=r + 1, start_column=shift, end_row=r + 1, end_column= shift + len(dates) * 3 - 1)
        
        sheet.cell(row=r + 2, column=shift).value = "взв."
        sheet.cell(row=r + 2, column=shift + 1).value = "ост."
        sheet.cell(row=r + 2, column=shift + 2).value = "зак."
        
        
def doReport(server, params, outObj):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()

    start = params.start
    finish = params.finish
    userid = params.userid

    where = '"created" > ToDate("{0}") and "created" <= ToDate("{1}") and "userid"="{2}"'.format(
         start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 23:59:59"), userid)

    orders = server.Get("Order", where)
    remnants = server.Get("OrgRemnants", where)
    returns = server.Get("Returns", where)
    price = server.Get("Price","","id")
    srvorg = server.Get("Org","","id")
    
    values = dict()
    priceAxe = dict()
    dateAxe = dict()
    orgs = list()
    
    update(orders, values, priceAxe, dateAxe, orgs, 0)
    update(remnants, values, priceAxe, dateAxe, orgs, 1)
    update(returns, values, priceAxe, dateAxe, orgs, 2)
    
    row = 0
   
    for o in orgs:
        if o in priceAxe:
            num = 1
            goods = priceAxe[o]
            dates = dateAxe[o]
            datelist = dates.keys()
            datelist.sort()
            
            if o in srvorg.keys():
                orgname = srvorg[o].name
            else:
                orgname = o  
                
            makeHead(sheet, orgname, datelist, row)
            row = row + 3
            pl = list()
            bun = 1
            
            for g in goods.keys():
                if g in price:
                    pl.append(price[g])
        
            sorted(pl, key=lambda x: x.name)
           
            for p in pl:
                sheet.cell(row=row, column=0).value = num
                sheet.cell(row=row, column=1).value = orgname
                sheet.cell(row=row, column=2).value = p.name + " " + p.id
                
                idx = 0
                for d in datelist:
                    shift = idx * 3 + 3
                    idx = idx + 1
                    
                    if o in values:
                        v = values[o]
                        key = str(d) + p.id
                        
                        if key in v:
                            ret = ""
                            if v[key].returns != 0:
                                ret = v[key].returns
                            
                            rmn = ""
                            if v[key].remains != 0:
                                rmn = v[key].remains
                            
                            ord = ""
                            
                            if v[key].orders != 0:
                                ord = v[key].orders
                                
                            sheet.cell(row=row, column=shift).value = ret
                            sheet.cell(row=row, column=shift + 1).value = rmn
                            sheet.cell(row=row, column=shift + 2).value = ord 
                
                
                row = row + 1
                num = num + 1    
            row = row + 1
    
    try:
        sheet.column_dimensions[get_column_letter(3)].width = 90
    except:   
        print("error:", sys.exc_info()[0])

    repName = "clientcard.xlsx" 
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)
    
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
    
    obj = outObj.New()
    obj.name = repName
    obj.file = bytes

def run(server):

    print("start clientcard")

    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    doReport(server, server.Params[0], outObj)
    server.Put(outObj)

    print("finish org_report")


   
