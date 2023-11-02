# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment

import datetime
from datetime import timedelta


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")


class QuestHelper:
  LIST_TYPE = 2
  DATASET_TYPE = 5
  PHOTO_TYPE = 7
  ORG_DATASET_TYPE = "Организация"

class Item:
  __slots__ = ["address", "org", "created", "user", "items", "orgcode", "orgData"]
  
  def __init__(self):
    self.address = ""
    self.org = ""
    self.created = ""
    self.user = ""
    self.orgcode = ""
    self.orgData = None
  
  def values(self):
    coord = ""
    if self.orgData != None :
        val = str(self.orgData.longitude) + ", " + str(self.orgData.latitude)
        coord = '=HYPERLINK("https://maps.yandex.ru/?ll={0},{1}&z=18&pt={0},{1},comma", "{2}")'.format(self.orgData.longitude, self.orgData.latitude, val)
        
    res = [self.org, self.orgcode, coord, self.address, self.created.strftime('%d.%m.%Y'), self.user]
    res.extend(self.items)
    
    return res
    
class ReportData:
  __slots__ = ['items', 'quests', 'orgs', 'agents', "cellidx", 'photos']
  KEY_FMT =  "{0}\t{1}"
  ITEM_KEY_FMT = "{0}\t{1}\t{2}"

  def __init__(self):
    self.items = list()
    self.quests = list()
    self.cellidx = dict()
    self.photos = dict()
  
  def countPhotos(self, docList, questItem):
    if questItem in self.photos: 
        return self.photos[questItem]
    
    countPhoto = 0
    
    for d in docList:
        count = 0
        for i in d.items:
            if i.iditem == questItem:
                count+=1
        if count > countPhoto:
            countPhoto = count
    
    self.photos[questItem] = countPhoto
    
    return countPhoto  
        
  def prepare(self, docs, quests):
    qu = list()
    
    ansByQ = dict()
    
    for d in docs:
      if not d.question in qu:
        qu.append(d.question)
        alist = list()
        ansByQ[d.question] = alist
      ansByQ[d.question].append(d)
    
    idx = 0
    
    ql = quests.values()
    ql.sort(cmp= lambda x, y: cmp(x.name, y.name))
    
    for q in ql:
      if q.idquest in qu:
        quest = quests[q.idquest]
        self.quests.append(quest)

        quest.items.sort(cmp= lambda x, y: cmp(x.number, y.number))
          
        for i in quest.items:
          if i.type == QuestHelper.LIST_TYPE:
            for v in i.values:
              key = self.ITEM_KEY_FMT.format(quest.idquest, i.iditem, v.value)
              self.cellidx[key] = idx     
              idx += 1
          elif i.type == QuestHelper.PHOTO_TYPE:
            cf = self.countPhotos(ansByQ[q.idquest], i.iditem)
            
            for ctr in range(0, cf):  
              key = self.KEY_FMT.format(quest.idquest, i.iditem + str(ctr))
              self.cellidx[key] = idx      
              idx += 1    
          else:
            key = self.KEY_FMT.format(quest.idquest, i.iditem)
            self.cellidx[key] = idx      
            idx += 1    

  def loadDocs(self, docs, pics, href):
    for d in docs:
      i = Item()
      i.created = d.created
      
      if d.id in self.orgs:
        i.address = self.orgs[d.id].address
        i.org = self.orgs[d.id].name
        i.orgcode = d.id
        i.orgData = self.orgs[d.id]
      
      if d.userid in self.agents:
        i.user = self.agents[d.userid].name
      
      i.items = [""]*len(self.cellidx)
      
      photoCount = dict()
      for n in d.items:
        if n.type ==QuestHelper.LIST_TYPE:
          self.putItemValue(i.items, self.ITEM_KEY_FMT.format(d.question, n.iditem, n.answer), "X")
          
        elif n.type == QuestHelper.PHOTO_TYPE:
          ctr = photoCount[n.iditem] if n.iditem in photoCount else 0
          val = '=HYPERLINK("{0}{1}", "Фото")'.format(href, pics[n.answer].name) if n.answer in pics else "Фото не найдено!"
          key = self.KEY_FMT.format(d.question, n.iditem + str(ctr))
          self.putItemValue(i.items, key, val)
          ctr += 1
          photoCount[n.iditem] = ctr
            
        elif n.type == QuestHelper.DATASET_TYPE:
          val = ''
          if n.remark == QuestHelper.ORG_DATASET_TYPE:
            val = self.orgs[n.answer].name if n.answer in self.orgs else n.answer
          
          self.putItemValue(i.items, self.KEY_FMT.format(d.question, n.iditem), val)
        else: 
          self.putItemValue(i.items, self.KEY_FMT.format(d.question, n.iditem), n.answer)
        
      self.items.append(i)   
      
  def putItemValue(self, items, key, value):
    if key in self.cellidx:
        idx = self.cellidx[key]
        items[idx] = value

class XLB(XLBuilder):
  __slots__ = ['staticTitles']  
  
  def __init__(self):
      self.staticTitles = 5
    
  YELLOW = Color("FFFFFF00")
  FIXED_CELL_COLOR = Color("FFB6DDE8")
  QUESTS_COLORS = [Color("FFD8D8D8"), Color("FFC2D69A")]
  
  def makeCell(self, sheet, row, column, cell, value):
    XLBuilder.makeCell(self, sheet, row, column, cell, value)        
    
    if column < self.staticTitles:
      fill = cell.style.fill
      fill.fill_type = Fill.FILL_SOLID
      fill.start_color = XLB.FIXED_CELL_COLOR
  
  def printHead(self, row, titles, sheet, report):
    cix = len(titles)
    self.staticTitles = cix
    
    sheet.merge_cells(start_row=row, start_column=0, end_row=row, end_column=cix - 1)
    self.printCell(sheet, row, 0, "Анкеты", XLB.YELLOW)
    
    for x in range(0, cix):
      sheet.merge_cells(start_row=row + 1, start_column=x, end_row=row + 2, end_column=x)
      self.printCell(sheet, row + 1, x, titles[x], XLB.FIXED_CELL_COLOR)
      sheet.column_dimensions[get_column_letter(x + 1)].width = 25
    
    i = 1
    for q in report.quests:
      color = Color("FFD8D8D8") #XLB.QUESTS_COLORS[len(XLB.QUESTS_COLORS) % i]
      i += 1
      cix = self.printGroup(row, cix, q, sheet, report, color)
    
    row += 1
    
    return row
    
  def printGroup(self, row, column, quest, sheet, report, color):
    self.printCell(sheet, row, column, quest.name, XLB.YELLOW, True)
    cv = column

    for i in quest.items:
      self.printCell(sheet, row + 1, cv, i.id, color)
      
      if i.type == QuestHelper.LIST_TYPE:
        s = cv
        
        for v in i.values:
          self.printHeadColumn(sheet, row + 2, cv, v.value, color, 4)
          cv += 1
          
        sheet.merge_cells(start_row=row + 1, start_column=s, end_row=row + 1, end_column=cv-1)

      elif i.type == QuestHelper.PHOTO_TYPE:
        phCount = report.photos[i.iditem]
        stColumn = cv
        for ctr in range(0, phCount):
            cell = sheet.cell(row=row+1, column=cv)
            
            cell.value = i.id if ctr == 0 else ""
            sheet.column_dimensions[get_column_letter(cv + 1)].width = 13
            cv += 1
        sheet.merge_cells(start_row=row + 1, start_column=stColumn, end_row=row + 2, end_column=cv - 1)
          
      elif i.type == QuestHelper.DATASET_TYPE:
        self.printHeadColumn(sheet, row + 2, cv, i.values[0].value, color, 13)
        cv += 1
        
      else:
        sheet.column_dimensions[get_column_letter(cv + 1)].width = 13
        sheet.merge_cells(start_row=row + 1, start_column=cv, end_row=row + 2, end_column=cv)
        cv += 1
      
    sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=cv-1)
    
    for c in range(0, cv):
      for r in range(2, 6):
        self.makeBorder(sheet.cell(row=r, column=c), XLBuilder.HEAD_BORDER_STYLE)
    
    return cv  

  def printCell(self, sheet, row, column, value, color, bold=False):
    cell = sheet.cell(row=row, column=column)
    cell.value = value
    self.styleCell(cell, color, bold)
   
  def printHeadColumn(self, sheet, row, column, value, color, width):
    self.printCell(sheet, row, column, value, color)
    sheet.column_dimensions[get_column_letter(column + 1)].width = width
    
  def styleCell(self, cell, color, bold=False):
    style = cell.style
    style.font.bold = bold
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    fill = cell.style.fill;
    fill.fill_type = Fill.FILL_SOLID
    fill.start_color = color
    
def loadData(data, params, server):
    orgs = dict()
    porg = server.Get("PotenzialOrg", 'not "userid" is null', "id")
    data.agents = server.Get("Agents", "", "id")
    quests = server.Get("Question", '"idquest" is null or "idquest" is not null', "idquest")
    where = '"created" >= ToDate("{0} 0:0:0") and "created" <= ToDate("{1} 0:0:0")'.format(params.start.strftime('%d.%m.%Y'), (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'))

    pictures = server.Get("PicStoreSrc", where, "id")
    orgs.update(porg)
    data.orgs = orgs
    
    if params.param == 0:
      for item in params.userids:
        server.ChangeUser(item.id)
        aorgs = server.Get("Org", '', 'id')
        server.RestoreUser()
        
        if aorgs != None:
            data.orgs.update(aorgs)

    
    if len(params.userids) > 0:
      arr = ''
      
      for i in params.userids:
        if len(arr) > 0:
          arr += ','
        
        arr += i.id
      
      where = '{0} and "userid" in ({1})'.format(where, arr)
    
    if len(params.quests) > 0:
      arr = ''
      
      for i in params.quests:
        if len(arr) > 0:
          arr += ','
        
        arr += i.id
      
      where = '{0} and "question" in ({1})'.format(where, arr)
    
    
    answers = server.Get("Answer" if params.param == 0 else "MAnswer", where)
    
    if params.param == 1:
      data.agents = server.Get("DivisionManager", "", "login")
      ids = []
      for a in answers:
        if len(a.id) > 0 and not a.id in ids: 
          ids.append(a.id)
          aorgs = server.Get("Org", '"id"=\'{0}\''.format(a.id), 'id')
          
          if aorgs != None:
            data.orgs.update(aorgs)
    
    data.prepare(answers, quests)
    data.orgs = orgs
    
    href = params.hrefBase
    data.loadDocs(answers, pictures, href)
    
    return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLB()
    
    c = sheet.cell(row=0, column=0)
    c.value = "Отчет по анкетам"  
    c.style.font.bold = True
    c.style.font.size = 18

    c = sheet.cell(row=1, column=0)
    c.value = "Интервал: c {0} по {1}".format(params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y'))  
    
    r = 2
    arr = ["Наименование", "Код ТТ", "Координаты", "Адрес", "Дата", "Торговый представитель"]
    r += xlb.printHead(r, arr, sheet, data);
    
    for i in data.items:
      xlb.makeCells(sheet, r, i.values())
      r += 1

    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "quest_rep.xlsx", server)                
    logging.info('end')
    
