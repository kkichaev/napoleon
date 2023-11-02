# -*- coding: cp1251 -*-

import sys;
import tempfile
import io
import time

from openpyxl import Workbook
from datetime import datetime, timedelta, date
from grsoft.xl_base import XLBuilder
from openpyxl.style import *
from openpyxl.cell import get_column_letter

reload(sys);
sys.setdefaultencoding("cp1251")

fullreport = []
shortreport = []

class XLBuilderFull(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value):
        if column > 0:
            if column == 1 or column == 2:
                sheet.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column)
                self.makeBorder(sheet.cell(row=row+1, column=column), Border.BORDER_THIN)
                cell.value = value
                
                if column == 2:
                    cell.style.number_format.format_code = 'd.m'
                    cell.value = value
              
            if  column > 2:
                if  column < 7:
                    cell.value = value[0]
                    
                cc = sheet.cell(row=row+1, column=column)
                self.makeBorder(cc, Border.BORDER_THIN)
                
                if  column < 7:
                    cc.value = value[1]    
                    
                cc.style.alignment. wrap_text = True
        
        if column == 0:
            s = cell.style
            s.alignment.text_rotation = 90
            s.alignment.horizontal = 'center'
            s.alignment.vertical = 'center'
            s.font.bold = True
            s.font.size = 18
            
        self.makeBorder(cell, Border.BORDER_THIN)
        cell.style.alignment. wrap_text = True
        
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment. wrap_text = True
        return column    

class XLBuilderShort(XLBuilder):
    min = None
    
    def __init__(self, h, m):
        self.min = h * 60 + m
        
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment. wrap_text = True
        
        if column > 1:
            c = column * 2 - 2 
            
            sheet.merge_cells(start_row=row, start_column = c, end_row = row, end_column = c+1)
            sheet.cell(row=row, column=c).value = cell.value
            self.paintHeadCell(sheet.cell(row=row, column=c))
            self.paintHeadCell(sheet.cell(row=row, column=c+1))
            
        return column  
    
    def makeCell(self, sheet, row, column, cell, value):
        if column == 0:
            cell.style.number_format.format_code = 'd.m'
            
        if column > 0 and column % 2 == 0:
            if value != None and len(value) > 0:
                tm = value.split(':')
                if len(tm)==2:
                    tmm = int(tm[0]) * 60 + int(tm[1])
                    
                    if tmm > self.min:
                        cell.style.font.color.index = Color.RED
                 
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        
class ItemF:
    fio = None
    vztCnt = 0;
    ordCnt = 0;
    sum = 0;
    items = None
    orgs = None
    
    def __init__(self, server, agent):
        self.fio = agent.name
        st = server.Params[0].start
        fn = server.Params[0].finish
        
        whereStr = '"created"' + " >= ToDate('" + st.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                    '"created"' + " < ToDate('" + fn.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                    '"userid"' + " in ('" + agent.id + "')"
        
        ord = server.Get("Order", whereStr)
        vzt = server.Get("VisitInfo", whereStr)
        rem = server.Get("OrgRemnants", whereStr)
         
        server.ChangeUser("'" + agent.id + "'")
        self.orgs = server.Get("Org", "", "id")
        server.RestoreUser()
        
        docs = [ord, vzt, rem];
        self.ordCnt = len(ord)
        self.vztCnt = calcVzt(docs)
        
        self.sum = 0
        
        for o in ord:
            for i in o.items:
                self.sum += i.cost * i.qty
        
        self.items = []
        
        while st < fn:
            self.items.append(ItemFData(self, server, docs, st))
            st += timedelta(days=1)
     
    def itemTime(self, dt):
        if self.items != None:
            for i in self.items:
                if dt.date() == i.date.date():
                    return (i.begTime, i.endTime)
        
        return (None, None)         
        
def calcVzt(docs):
    ido = []

    for d in docs:
        appendid(ido, d);
    
    return len(ido)
    
def appendid(arr, docs):
    for d in docs:
        if not d.id in arr:
            arr.append(d.id)        
    
class ItemFData:
    date = None
    begDoc = None
    endDoc = None
    begTime = None
    endTime = None
    begOrg = None
    endOrg = None    
    
    def __init__(self, p, server, da, date):
        list = []
        self.date = date
        
        for dc in da:
            for d in dc:
                if date.date() == d.created.date():
                    list.append(d)
                    
        list = sorted(list, key = lambda x: x.created)
        
        sz = len(list)
        
        if sz > 0:
            d = list[0]
            self.begDoc = getDocName(d) 
            self.begOrg = d.id
            self.begTime = d.created
            
            if d.id in p.orgs:
                self.begOrg = p.orgs[d.id].name
            
            d = list[len(list) - 1]
            self.endDoc = getDocName(d)
            self.endOrg = d.id
            self.endTime = d.created 
            
            if d.id in p.orgs:
                self.endOrg = p.orgs[d.id].name
        
def getDocName(d):
    return "зявка" if str(d).startswith('Order') else "посещение"

class Item:
    pass

def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server) 
    print "finish\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')

def doReport(server):
    loadData(server)
    wb = printOut(server)
    XLBuilder().workbookToObject(wb, "visit.xlsx", server)        

def printOut(server):     
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Подробный"

    xlb = XLBuilderFull()
    head = ["ФИО", "", "", "", "Действие", "Время", "Наименование ТТ", "Всего визитов/заказов", "Сумма"]
    xlb.makeHead(sheet, 1, head)
    
    wd = ["пн", "вт", "ср", "чт", "пт", "сб", "вс",]
    
    row = 2
    
    for d in fullreport:
        rc = row+len(d.items)*2 - 1
        sheet.merge_cells(start_row=row, start_column=0, end_row=rc, end_column=0)
        for r in range(row, rc+1):
            xlb.makeBorder(sheet.cell(row=r, column=0), Border.BORDER_THIN)
        
        cell = sheet.cell(row=row, column=0)
        cell.value = d.fio
        cell = sheet.cell(row=row, column=7)
        cell.value = d.vztCnt
        cell = sheet.cell(row=row+1, column=7)
        cell.value = d.ordCnt
        cell = sheet.cell(row=row, column=8)
        cell.value = d.sum
        
        for i in d.items:
            rd = [d.fio, wd[i.date.weekday()], calcExcelDate(i.date), ["Начало", "Завершение"], 
                  [i.begDoc, i.endDoc], [strftime(i.begTime), strftime(i.endTime)],[i.begOrg, i.endOrg], "", ""]
            
            xlb.makeCells(sheet, row, rd)
            row += 2
            
    cc = 2
    for w in [3,11,13,15,9,40,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    sheet = wb.create_sheet()
    sheet.title = "Краткий"
    xlb = XLBuilderShort(server.Params[0].hour, server.Params[0].min)
    
    h = ["",""]
    
    for d in fullreport:
        h.append(d.fio)
        
    xlb.makeHead(sheet, 1, h)
    row = 2  
    
    for d in shortreport:
        rd = [calcExcelDate(d[0]), wd[d[0].weekday()]]
        
        for di in d[1:]: 
            rd.append(strftime(di))
            
        xlb.makeCells(sheet, row, rd)
        row += 1
    
    return wb

def calcExcelDate(dt):
    return (dt.date() - date(1900, 1, 1)).days + 2
     
def strftime(dt):
    return dt.strftime('%H:%M') if dt != None else ""

def collectAgent(server):
    result = []
    a = server.Get("Agents", "", "id")
    ids = server.Params[0].ids
    
    for ai in ids.split(','):
        if ai in a:
            result.append(a[ai])
    
    return result      
      
def loadData(server):
    agents = collectAgent(server)
    
    for a in agents:
        fullreport.append(ItemF(server, a))   
    
    st = server.Params[0].start
    fn = server.Params[0].finish
    
    while st < fn:
        row = []
        row.append(st)
        
        for d in fullreport:
            row.extend(d.itemTime(st))
        
        shortreport.append(row)    
        st += timedelta(days=1)
