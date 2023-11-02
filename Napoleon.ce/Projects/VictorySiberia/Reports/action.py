# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

class Report:
    start = None
    finish = None
    items = None
    agent = None
    division = None
    
    def __init__(self):
        self.start = datetime.now()
        self.finish = datetime.now()
        self.items = list()
        self.agent = ""
        self.division = ""
        
class Item:
    org = None
    vizit = None
    order = None
    
    def __init__(self):
      self.org = ""
      self.vizit = list()
      self.order = 0
      
    def getData(self, row):
      return [self.org, len(self.vizit), self.order , "=IFERROR(C{0}/B{0}, 0)".format(row + 1)]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].userid
    
def loadData(server):
    start, finish, userid = inflateParams(server)
    
    WHERE_STR = '"created" >= ToDate("{0}") and "created" < ToDate("{1}") and "userid"=\'{2}\'';         
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
    
    agent = ""
    server.ChangeUser(userid)
    agent = server.CurrentUser().name
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    server.RestoreUser()

    data = dict()
    
    for name in ['Order', 'VisitInfo', 'OrgRemnants', 'Question', 'Incass', 'ScriptDoc', 'TaskDone']:
        doc = server.Get(name, where)
             
        for d in doc:
            if len(d.id.strip()) == 0:
                continue
            
            if not d.id in data:
                f = Item()
                data[d.id] = f 
                f.org = "{0} / {1}".format(orgs[d.id].name, orgs[d.id].address) if d.id in orgs else d.id
                
            r = data[d.id]
            dt = d.created.date()
            
            if not dt in r.vizit:
                r.vizit.append(dt)
                
            if name == 'Order':
                r.order += 1       
    
    for oid in orgs:
        if not oid in data:   
            f = Item()
            data[oid] = f
            f.org =  "{0} / {1}".format(orgs[oid].name, orgs[oid].address)
            
    div = server.Get("Division","")
    
    dname = ""
    for d in div:
        for da in d.agents:
            if da.id == userid:
                dname = d.name
                break;             
    
    items = sorted(data.values(), cmp=item_cmp)
    res = Report()
    res.start = start
    res.finish = finish - timedelta(days=1)
    res.agent = agent
    res.items = items
    res.division = dname
    
    return res
    
def item_cmp(x, y):
  res = cmp(x.org, y.org)

  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    sh.cell(row=0, column=0).value = '{0} - {1} / {2} / {3}'.format(
        data.start.strftime("%d.%m.%Y"), data.finish.strftime("%d.%m.%Y"), data.agent, data.division)
    
    head = ["Название торговой точки вся база ТП", "Количество посещений", "Количество заявок", "% Заказов"]
    
    r = 1
    xlb.makeHead(sh, r, head)
    
    for d in data.items:
      r += 1
      xlb.makeCells(sh, r, d.getData(r))
    
    setCellWidth(sh, [90,12,12,12])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 3:
      cell.style.number_format._set_format_code('0%')
      
    
def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    