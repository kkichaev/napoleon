# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

import datetime
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import sys
import tempfile

reload(sys);
sys.setdefaultencoding("cp1251")


class ReportDataItem:
    agent = None
    org = None
    device = None
    info = None
    created = None
    links = None
    
    def __init__(self, rep, repdata, pics):
        self.links = list()
        
        if pics:
            for i in rep.items:
                self.links.append(repdata.addPic(i.id))
        
        self.agent = repdata.agents[rep.userid].name if rep.userid in repdata.agents else ""
        self.org = repdata.orgs[rep.id].name if rep.id in repdata.orgs else ""
        self.device = rep.device
        self.info = rep.remark
        self.created = rep.created.strftime("%d/%m/%Y %H:%M:%S")
        
    def __str__(self):
        result = ""
        for property, value in vars(self).iteritems():
            if len(result) > 0:
                result = result + ", "
            result = result + str(property) + ": " + str(value)
        
        return result
    
class PicData:
    name = None
    data = None
    
    def __init__(self):
        name = ""
        data = None
        
class ReportData:
    items = None
    agents = None
    orgs = None
    pics = None
    pcnt = None
    
    def __init__(self, server):
        self.pcnt = 1
        param = server.Params[0];
        self.agents = server.Get("Agents", "", "id")
                
        uid = '"userid" in ' + "(" + param.userid + ")"
        self.orgs = server.Get("Org", uid, "id")
        
        where = '"created" > ToDate("{0}") and "created" <= ToDate("{1}") and {2} order by "created","userid"'.format(
          param.begin.strftime("%d/%m/%Y 0:0:0"), param.end.strftime("%d/%m/%Y 23:59:59"), uid)
        
        rep = server.Get("DefectReport" if param.pics else "DefectReportInfo" , where)
        
        self.items = list()
        self.pics = list()
        
        for r in rep:
            repdata = ReportDataItem(r, self, param.pics)
            self.items.append(repdata)
            
    def addPic(self, pic):
        pd = PicData()
        pd.name = str(self.pcnt)
        pd.data = pic
        self.pics.append(pd)
        self.pcnt = self.pcnt + 1
        
        return pd.name
         
def doReport(data, out):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()

    sheet.cell(row=0, column=0).value = "Дата"
    sheet.cell(row=0, column=1).value = "Оператор"
    sheet.cell(row=0, column=2).value = "Аппарат"
    sheet.cell(row=0, column=3).value = "Сведения"
    sheet.cell(row=0, column=4).value = "Неисправность"
    
    sheet.column_dimensions[get_column_letter(1)].width = 22
    sheet.column_dimensions[get_column_letter(2)].width = 35
    sheet.column_dimensions[get_column_letter(3)].width = 35
    sheet.column_dimensions[get_column_letter(4)].width = 35
    sheet.column_dimensions[get_column_letter(5)].width = 35
    
    row = 1;
    
    for r in data.items:
        sheet.cell(row=row, column=0).value = r.created
        sheet.cell(row=row, column=1).value = r.agent
        sheet.cell(row=row, column=2).value = r.org
        sheet.cell(row=row, column=3).value = r.device
        sheet.cell(row=row, column=4).value = r.info
        
        start_col = 5;
        
        for l in r.links:
            sheet.cell(row=row, column=5).hyperlink = l + ".jpg"
        
        row = row + 1

    repName = "defect_report.xlsx" 
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    print fileName
    
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = out.New()
    obj.name = repName
    obj.file = bytes    
    
def run(server):

    print "start", __name__, datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    type = "Result[name:s,file:b,items[name:s,pic:b]]"
    server.RegisterType(type)
    outObj = server.New("Result")
    
    data = ReportData(server)
    doReport(data, outObj)
    
    for p in data.pics:
        pic = outObj[0].items.New()
        pic.name = p.name
        pic.pic = p.data
        
    server.Put(outObj)
   
    print "done", __name__, datetime.now().strftime('%d/%m/%Y %H:%M:%S')
