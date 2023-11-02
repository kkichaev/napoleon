# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ffC0C0C0"
userOrgs = dict()

class Item:
    data = None
    org = None
    item = None
    remn = None
    order = None
    tara = None
    agent = None
    
    def __init__(self):
        self.remn = 0
        self.order = 0
        self.tara = 0
        
    def getData(self):
        return [self.data.strftime("%d.%m.%Y"), self.agent,  self.org, self.item, self.remn, self.order, self.tara]

WHERE_STR = '"{2}" >= ToDate("{0}") and "{2}" <= ToDate("{1}")'; 

def inflateParams(server):
    return server.Params[0].begin, server.Params[0].end + timedelta(days=1)

def getOrgName(server, id, userid):
    global userOrgs
    if not userid in userOrgs:
        print "get user orgs"
        server.ChangeUser("'" + userid + "'")
        o = server.Get("Org", "", "id")
        server.RestoreUser()
        userOrgs[userid] = o
    
    orgs = userOrgs[userid] 
    return orgs[id].name if id in orgs else "Контрагент с кодом<{0}>".format(id)

def getItemName(id, price):
    return price[id].name if id in price else "Товар с кодом<{0}>".format(id)

def getAgentName(id, agents):
    return agents[id].name if id in agents else "Торговый представитель с кодом<{0}>".format(id)

def traverse(server, docs, data, price, func, agents):        
    for d in docs:
        dd = d.created.date()
        key = str(dd) + d.id + d.userid
        
        if not key in data:
            data[key] = dict();
             
        items = data[key]
        org = getOrgName(server, d.id, d.userid)     
        agent = getAgentName(d.userid, agents)
        
        for di in d.items:
            if not di.id in items:
                i = Item()
                i.org = org
                i.data = d.created
                i.item = getItemName(di.id, price)
                i.agent = agent;
                items[di.id] = i
            
            func(items[di.id], di)
            
def funcRemn(i, di):
    i.remn = i.remn + di.qty
    i.tara = i.tara + di.tara
 
def funcOrd(i, di):
    i.order = i.order + di.qty
                       
def cmpItem(x, y):
    result = cmp(x.data, y.data)
    
    if result == 0:
        result = cmp(x.org, y.org)
        
    return result
         
def collectData(server, orders, remnants, price, agents):
    dt = None
    pos = 0
    
    data = dict()
    traverse(server, remnants, data, price, funcRemn, agents)
    traverse(server, orders, data, price, funcOrd, agents)
            
    result = list()
    for d1 in data.values():
        for d2 in d1.values():
            result.append(d2)
    
    result.sort(cmp=cmpItem)        
                
    return result

def loadData(server):
    start, finish = inflateParams(server)
    
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), "created")
    remnants = server.Get("OrgRemnants", where)
    remnants.sort(cmp= lambda x, y: cmp(x.created, y.created))
    
    orders = server.Get("Order", where)
    orders.sort(cmp= lambda x, y: cmp(x.created, y.created))
    
    sd1 = collectData(server, orders, remnants, server.Get("ManagerPrice", "", "id"), server.Get("Agents", "", "id"))
    
    rd = RptData()
    rd.start = start
    rd.finish = finish - timedelta(days=1)
    
    return rd, sd1 


class XLBuilderEx(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment.wrap_text = True
        self.setBackColor(cell, bkgColor)
        return column    
    
    def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def drawData(sh, xlb, data):
    row = 3
    for d in data:
        rd = d.getData()
        xlb.makeCells(sh, row, rd)
        row += 1
        
def paintHeadCell(xlb, cell):
    cell.style.alignment.wrap_text = True
    xlb.setBackColor(cell, bkgColor)
    xlb.paintHeadCell(cell)
    
def makeHead(xlb, sheet):
    sheet.merge_cells(start_row=1, start_column = 0, end_row = 2, end_column = 0)
    cell = sheet.cell(row=1, column=0)
    paintHeadCell(xlb, cell)
    xlb.makeBorder(sheet.cell(row=2, column=0), XLBuilder.HEAD_BORDER_STYLE)
    cell.value = "Дата"
    
    sheet.merge_cells(start_row=1, start_column = 1, end_row = 2, end_column = 1)
    cell = sheet.cell(row=1, column=1)
    paintHeadCell(xlb, cell)
    xlb.makeBorder(sheet.cell(row=2, column=1), XLBuilder.HEAD_BORDER_STYLE)
    cell.value = "Торговый представитель"
    
    sheet.merge_cells(start_row=1, start_column = 2, end_row = 2, end_column = 2)
    cell = sheet.cell(row=1, column=2)
    paintHeadCell(xlb, cell)
    xlb.makeBorder(sheet.cell(row=2, column=2), XLBuilder.HEAD_BORDER_STYLE)
    cell.value = "Торговая точка"
    
    sheet.merge_cells(start_row=1, start_column = 3, end_row = 2, end_column = 3)
    cell = sheet.cell(row=1, column=3)
    paintHeadCell(xlb, cell)
    xlb.makeBorder(sheet.cell(row=2, column=3), XLBuilder.HEAD_BORDER_STYLE)
    cell.value = "Номенклатура"
    
    sheet.merge_cells(start_row=1, start_column = 4, end_row = 1, end_column = 6)
    cell = sheet.cell(row=1, column=4)
    paintHeadCell(xlb, cell)
    xlb.makeBorder(sheet.cell(row=1, column=5), XLBuilder.HEAD_BORDER_STYLE)
    xlb.makeBorder(sheet.cell(row=1, column=6), XLBuilder.HEAD_BORDER_STYLE)
    cell.value = "Кеги"
    
    cell = sheet.cell(row=2, column=4)
    paintHeadCell(xlb, cell)
    cell.value = "Остаток полные (количество)"
    
    cell = sheet.cell(row=2, column=5)
    paintHeadCell(xlb, cell)
    cell.value = "Заявка"
    
    cell = sheet.cell(row=2, column=6)
    paintHeadCell(xlb, cell)
    cell.value = "Пустые (количество)"
                    
def printData(xlb, sh, rpd, d):
    sh.cell(row=0, column=0).value = "Отчет по возвратным кегам. Период с {0} по {1}".format(rpd.start.strftime("%d/%m/%Y"), rpd.finish.strftime("%d/%m/%Y"))
    makeHead(xlb, sh)
    drawData(sh, xlb, d) 
    setCellWidth(sh, [15,40,40,40,16,16,16])

def printOut(rpd, d1):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    printData(xlb, sh, rpd, d1)
    
    return wb

class RptData:
    agent = None
    start = None
    finish = None
          
def doReport(server):
    rd, data = loadData(server)
    wb = printOut(rd, data)
    XLBuilder().workbookToObject(wb, "visit.xlsx", server)
          
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    