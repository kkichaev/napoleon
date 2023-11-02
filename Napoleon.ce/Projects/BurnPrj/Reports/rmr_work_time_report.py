# -*- coding: cp1251 -*-
from importlib import reload
import logging
from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
#sys.setdefaultencoding("cp1251")

class Report:
  pass

class Data:
  def getData(self):
    res = [self.org, self.format, self.address]
    
    for d in self.days:
      res.append(d.start.strftime("%H:%M") + " " if d.start != None else " ")
      res.append(d.finish.strftime("%H:%M") + " " if d.finish != None else " ")
    return res
      
class DayData():
  def __init__(self):
    self.start = None
    self.finish = None
    
  def __str__(self):
    return 'start: {0}, finish: {1}'.format(self.start.strftime("%H:%M") if self.start != None else " ",
      self.finish.strftime("%H:%M") if self.finish != None else " ")
    
class Item:
  def __init__(self):
    self.data = datetime.now()
    self.finish = datetime.now()
    
  def getData(self):
    return [self.agent]
      
def inflateParams(server):
  return server.Params[0].start, server.Params[0].finish, server.Params[0].userids[0]
    
def loadData(server):
  start, finish, userid = inflateParams(server)
  
  report = Report()
  report.data = {}
  report.start = start
  report.finish = finish
  report.days = []
  
  s = start
  
  while s <= finish:
    d = DayData()
    d.date = s.date()
    report.days.append(d)
    s += timedelta(days=1)
  
  WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
  WHERE_STR = WHERE_STR_ALL + ' and "userid" = \'{2}\''; 

  where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 23:59:00"), userid.id)
  scripts = server.Get("ScriptDoc", where)
  
  division = server.Get("Division", '')
  
  report.division = ""
  
  for d in division:
    for ds in d.agents:
      if ds.id == userid.id:
        report.division = d.name
        break;
        
  agent = ""
  usersOrg = {}
  agents = {}
  
  server.ChangeUser(userid.id)
  agent = server.CurrentUser().name
  agentid = server.CurrentUser().id
  agents[agentid] = agent
  usersOrg[agentid] = server.Get("Org", "", "id")
  porg = server.Get("PotenzialOrg", "", "id")
  usersOrg[agentid].update(porg)
  server.RestoreUser()

  data = [] 
  
  orgmap = {}
  report.agent = agent
  
  for s in scripts:
    if not s.id in report.data:
      report.data[s.id] = Data()
      report.data[s.id].agent = agent
      report.data[s.id].days = []
      report.data[s.id].org = usersOrg[s.userid][s.id].name if s.id in usersOrg[s.userid] else ""
      report.data[s.id].address = usersOrg[s.userid][s.id].address if s.id in usersOrg[s.userid] else ""
      report.data[s.id].format = usersOrg[s.userid][s.id].formatTT if s.id in usersOrg[s.userid] else ""

      for d in range(0,len(report.days)):
        report.data[s.id].days.append(DayData())
        
      report.data[s.id].items = {}
    
    d = Item()
    
    if not s.userid in orgmap:
      orgmap[s.userid] = {}
      
    if not s.created.date() in orgmap[s.userid]:
      orgmap[s.userid][s.created.date()] = 0
    
    orgmap[s.userid][s.created.date()] += 1
    
    d.id = s.id
    d.org = orgmap[s.userid][s.created.date()]
    d.address = usersOrg[s.userid][s.id].address if s.id in usersOrg[s.userid] else ""
    d.data = s.created
    d.userid = s.userid
    d.agent = agents[s.userid] if s.userid in agents else ""
    dt = d.data
    
    if s.items != None:
      for s in s.items:
        if s.state == 1 and s.date > dt and s.date.date() == dt.date():
          dt = s.date
          
    d.finish = dt
    
    data.append(d)
    
  data = sorted(data, key=lambda x: x.data)

  for d in data:
    idx = (d.data - start).days
    
    dd = report.data[d.id].days[idx]
    
    if dd.start == None or d.data < dd.start:
      dd.start = d.data
    
    if dd.finish == None or d.finish > dd.finish:
      dd.finish = d.finish
    
    if not d.org in report.data[d.id].items:
      report.data[d.id].items[d.org] = []
      
      for dx in range(0,len(report.days)):
        report.data[d.id].items[d.org].append(DayData())
    
    orgitem = report.data[d.id].items[d.org][idx]
    
    if orgitem.start == None or d.data < orgitem.start:
      orgitem.start = d.data
    
    if orgitem.finish == None or d.finish > orgitem.finish:
      orgitem.finish = d.finish
  
  return report, agent
    
def item_cmp(x, y):
  res = cmp(x.agent, y.agent)
  
  if res == 0:
    res = cmp(x.data, y.data)

  return res
  
class XLBuilderEx(XLBuilder):
  HEAD_COLOR = "FFD8D8D8"
  
  def adjustHeadCell(self, sheet, cell, row, column):
    self.setBackColor(cell, self.HEAD_COLOR)
    sheet.merge_cells(start_row=row, start_column=column, end_row=row+1, end_column=column)    
    
    return column  
  
def setCellWidth(sh, wa):
  cc = 1
  for w in wa:
      sh.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
def ptintFisrtSheet(xlb, sh, report, agent):
  sh.cell(row=0, column=0).value = report.division
  sh.cell(row=0, column=0).style.font.bold = True
  sh.cell(row=1, column=0).value = report.agent
  sh.cell(row=1, column=0).style.font.bold = True

  head = ["№","Название ТТ", "Формат ТТ", "Адрес ТТ"]
  r = 2
  xlb.makeHead(sh, r, head)
  
  col = len(head)
  
  for d in report.days:
    sh.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
    cell = sh.cell(row=r, column=col)
    cell.value = "{0}".format(d.date.strftime("%d.%m.%Y"))
    XLBuilder().setBackColor(cell, XLBuilderEx.HEAD_COLOR)
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    sh.cell(row=r+1, column=col).value = 'прибыл'
    XLBuilder().setBackColor(sh.cell(row=r+1, column=col), XLBuilderEx.HEAD_COLOR)
    sh.cell(row=r+1, column=col+1).value = 'убыл'
    XLBuilder().setBackColor(sh.cell(row=r+1, column=col+1), XLBuilderEx.HEAD_COLOR)
    col += 2
  
  for c in range(0, len(report.days)*2 + 4):
    for r in range(2,4):
      cel = sh.cell(row=r, column=c)
      cel.style.font.bold = True
      xlb.makeBorder(cel, xlb.HEAD_BORDER_STYLE)
  
  for d in sorted(report.data.values(), key = lambda x: x.agent):
    data = [r - 2]
    data.extend(d.getData())
    r+=1
    xlb.makeCells(sh, r, data)
  
  arr = [10, 30, 30, 30]
  
  for a in range(0, len(report.days) * 2):
    arr.append(12)
    
  setCellWidth(sh, arr)

def printOut(d, a):
  wb = Workbook(False, 'cp1251')
  sh = wb.get_active_sheet()
  
  xlb = XLBuilderEx()
  ptintFisrtSheet(xlb, sh, d, a)
  
  return wb
    
def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data, agent = loadData(server)
  wb = printOut(data, agent)

  XLBuilder().workbookToObject(wb, "rmr_work_time_report.xlsx", server)                
  logging.info('end')
    