# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN

class WeeksItem:
    start = None
    finish = None
    
class Weeks:
    items = None
    
    def __init__(self, start, finish):
        self.items = list()
        item = None;
        idx = 0;
        
        while start <= finish:
            idx = start.weekday()
            
            if idx >= 5:
                item = None
            else:
                if item == None:
                    item = WeeksItem();
                    item.start = start;
                    self.items.append(item)
                    
                item.finish = start;        
                 
            start = start + timedelta(days=1)
            
class ItemCell:
    face = None
    qty = None
    
    def __init__(self):
        self.face = 0.0
        self.qty = 0.0

class ItemData:
    report = None
    data = None
    visit = None
    
    def __init__(self, report):
        self.report = report
        self.data = dict()
        self.visit = 0
        
    def put(self, item):
        cl = None
        if item.id in self.data:
            cl = self.data[item.id]
        else:
            cl = ItemCell()
            self.data[item.id] = cl
            
        cl.face = cl.face + item.face
        cl.qty = cl.qty + item.qty            
        
class PageItem:
    slsnet = None
    city = None
    address = None
    data = None
    report = None
    
    def __init__(self, report):
        self.report = report
        self.data = list()
        
        for w in self.report.weeks.items:
            self.data.append(ItemData(report))
    
    def process(self, c):
        wid = -1
        for i in self.report.weeks.items:
            if i.start <= c.created and i.finish >= c.created:
                wid = self.report.weeks.items.index(i)
                break
        
        if wid != -1:
            id = self.data[wid]
            id.visit = id.visit + 1
            for i in c.items:
                id.put(i)
        
                 
class Page:
    report = None
    user = None
    items = None;
    
    def __init__(self, report, user):
        self.report = report
        self.user = user
        self.items = dict()
        
    def process(self, c):
        item = None
        if c.id in self.items:
            item = self.items[c.id]
        else:
            item = PageItem(self.report)
            self.items[c.id] = item
            
        item.process(c)            
         
class Report:
    pages = None
    weeks = None
    
    """ Список прайс """
    price = None
    """ Позиции в списке по ID """
    pidx = None
    
    slsnet = None
    city = None
    org = None
    
    def __init__(self, server):
        start = server.Params[0].start
        finish = server.Params[0].finish
        cid = server.Params[0].cid
        
        self.weeks = Weeks(start, finish)
        
        conDef = server.Get("ContractDef", '"id"=' + "'" + cid + "'")
        
        pids = ""
        
        for cd in conDef:
            for i in cd.items:
                if len(pids) > 0:
                    pids += ","
                    
                pids += "'" + i.id +"'"
        
        mprice = server.Get("ManagerPrice", '"id" in (' + pids + ')', "id")
        
        self.price = list();
        self.price.extend(mprice.values())
        
        sorted(self.price, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
        
        self.pidx = dict()
        for i in range(0, len(self.price)):
            self.pidx[self.price[i].id] = i    
        
        self.org = server.Get("Org", '"id" is not null', "id")
        agents = server.Get("Agents","","id")
        
        endRange = finish + timedelta(days=1) 
        where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + endRange.strftime('%d/%m/%Y') + '") and "def"='+"'" + cid + "'"
        
        print where
        con = server.Get("Contract", where)
        
        self.pages = dict()
        
        self.slsnet = server.Get("Slsnet","","id")
        self.city = server.Get("City", "", "id")
        
        if con != None:
            for c in con:
                if c.userid in agents and c.id in self.org:
                    self.page_process(server, c, agents[c.userid], self.org[c.id])
        
    def page_process(self, server, c, user, org):
        
        p = None
        if c.userid in self.pages:
            p = self.pages[c.userid]
        else:
            p = Page(self, user)
            self.pages[c.userid] = p;    
            
        p.process(c)
    
    def getCity(self, id):
        result = ""
        
        if id in self.org and self.org[id].cid in self.city:
            result = self.city[self.org[id].cid].name
        
        return result
            
    def getSls(self, id):
        result = ""
        
        if id in self.org and self.org[id].sid in self.slsnet:
            result = self.slsnet[self.org[id].sid].name
        
        return result      
    
    def getAddress(self, id):
        result = ""
        
        if id in self.org:
            result = self.org[id].address 
            
        return result   
              
def setVal(cell, value, vrt = Alignment.VERTICAL_TOP, hrz= Alignment.HORIZONTAL_CENTER, wrap=True, bold=False):
    cell.value = value
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    cell.style.font.bold = bold
                      
def run(server):
    print "contract start"
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    report = Report(server)
    wb = Workbook(False, 'cp1251')
    
    sheet = None
    for page in report.pages.values():
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()
        sheet.title = page.user.name
        
        i = 1
        r = 0
        for w in report.weeks.items:
            sheet.cell(row=r, column=2).value = str(i) + "-я неделя"
            sheet.cell(row=r, column=3).value = "{0:%d}-{1:%d.%m.%Y}".format(w.start, w.finish)
            i = i + 1
            r = r + 1
        
        r = r + 1
        sheet.merge_cells(start_row=r, start_column=0, end_row=r+1, end_column=0)    
        setVal(sheet.cell(row=r, column=0), "№ п/н", bold=True)
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 10
        sheet.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
        setVal(sheet.cell(row=r, column=1), "Cеть", bold=True)
        col_letter = get_column_letter(2) 
        sheet.column_dimensions[col_letter].width = 15
        sheet.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        setVal(sheet.cell(row=r, column=2), "Город", bold=True)
        col_letter = get_column_letter(3) 
        sheet.column_dimensions[col_letter].width = 15
        sheet.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        setVal(sheet.cell(row=r, column=3), "Адрес", bold=True)
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 15
        sheet.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)
        setVal(sheet.cell(row=r, column=4), "неделя", bold=True)
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 15
        sheet.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        setVal(sheet.cell(row=r, column=5), "количество визитов в неделю")
        col_letter = get_column_letter(6) 
        sheet.column_dimensions[col_letter].width = 15   

        c = 6
        for p in report.price:
            setVal(sheet.cell(row=r, column=c), p.name, bold=True)
            sheet.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
            setVal(sheet.cell(row=r + 1, column=c), "Кол-во фейсов")
            setVal(sheet.cell(row=r + 1, column=c + 1), "Наличие по факту (штук)")
            c = c + 2
                 
        r = r + 2
        wc = len(report.weeks.items) - 1
        idx = 1
        for i in page.items:
            sheet.merge_cells(start_row=r, start_column=0, end_row=r+wc, end_column=0)
            setVal(sheet.cell(row=r, column=0), str(idx), vrt = Alignment.VERTICAL_CENTER)
            sheet.merge_cells(start_row=r, start_column=0, end_row=r+wc, end_column=0)
            
            pi = page.items[i]
            sheet.merge_cells(start_row=r, start_column=1, end_row=r+wc, end_column=1)
            setVal(sheet.cell(row=r, column=1), report.getSls(i), vrt = Alignment.VERTICAL_CENTER)
            
            sheet.merge_cells(start_row=r, start_column=2, end_row=r+wc, end_column=2)
            setVal(sheet.cell(row=r, column=2), report.getCity(i), vrt = Alignment.VERTICAL_CENTER)
            
            sheet.merge_cells(start_row=r, start_column=3, end_row=r+wc, end_column=3)
            setVal(sheet.cell(row=r, column=3), report.getAddress(i), vrt = Alignment.VERTICAL_CENTER)
                        
            for pidx in range (0, len(pi.data)):
                setVal(sheet.cell(row=r + pidx, column=4), "{0}-я неделя".format(pidx+1))
                
                idd = pi.data[pidx]
                setVal(sheet.cell(row=r + pidx, column=5), "" if idd.visit == 0 else idd.visit)
                
                c = 6
                for d in idd.data:
                    cidx = report.pidx[d] * 2 + c
                    itc = idd.data[d]
                    setVal(sheet.cell(row=r + pidx, column=cidx), itc.face)
                    setVal(sheet.cell(row=r + pidx, column=cidx + 1), itc.qty)
                
            idx = idx + 1
            r = r + wc + 1
        wc = wc + 1    
        rangeBorders(sheet.range("A7:"+get_column_letter(6 + len(report.price) * 2)+str(len(page.items) + r - 1)) )
            
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    server.Put(outObj)
    
    print "contract finish"