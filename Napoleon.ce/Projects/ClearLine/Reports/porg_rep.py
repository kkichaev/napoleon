# -*- coding: cp1251 -*-
from importlib import reload
import logging
import sys;
import re
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from manager.task import TaskReportData
from rmr_visit_report import Data, AgentSheet
from datetime import datetime
from openpyxl.cell import get_column_letter
from openpyxl.cell import get_column_letter
from orgmap import OrgMap

reload(sys);

class Item:
  def getData(self):
    return [self.name, self.id, self.location, self.address]

class RepData:
  def __init__(self):
    self.items = []
  
def loadData(params, server):
  porg = server.Get("PotenzialOrg", '"userid"=\'{0}\''.format(params.userid), "id")
  print (porg)

  data = RepData()
  
  for p in porg.values():
    item = Item()
    item.name = p.name
    item.id = p.id
    item.location = '{0}, {1}'.format(p.longitude, p.latitude)
    item.address = p.address
    data.items.append(item)
  
  data.items = sorted(data.items, key=lambda lhs: lhs.name)  
  
  return data

class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'
  
  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
    
def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  xlb = XLBuilderEx()

  xlb.makeHead(sheet, 0, ["Наименование", "Код ТТ", "Координаты", "Адрес"], True)

  row = 1

  for i in data.items:
    xlb.makeCells(sheet, row,  i.getData())
    row += 1

  x = 1
  for w in [27,40,27,60]:
    sheet.column_dimensions[get_column_letter(x)].width = w
    x += 1

  return wb
  
def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilder().workbookToObject(wb, "porg_rep.xlsx", server)                
  logging.info('end')