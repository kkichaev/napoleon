# -*- coding: cp1251 -*-

import sys;
import tempfile
import io
import time
from importlib import reload

from openpyxl import Workbook
from datetime import datetime, timedelta, date
from grsoft.xl_base import XLBuilder
from openpyxl.style import *
from openpyxl.cell import get_column_letter

reload(sys)


class XLBuilderEx(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value):
        if column == 1:
            cell.style.number_format.format_code = 'd.m'
        
        XLBuilder.makeCell(self, sheet, row, column, cell, value)


def run(server):
    print ("start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    doReport(server) 
    print ("finish\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

def doReport(server):
    loadData(server)
    wb = printOut(server)
    XLBuilder().workbookToObject(wb, "visit.xlsx", server)        

def printOut(server):     
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    for idx, p in enumerate(report):
        if idx > 0:
            sheet = wb.create_sheet(None, p.agent)
        else:
            sheet.title = p.agent
        
        st = server.Params[0].start
        fn = server.Params[0].finish    
        sheet.cell(row=0, column=0).value = "Период с:{0} по:{1}".format(st.strftime("%d.%m.%Y"), fn.strftime("%d.%m.%Y"))        
    
        xlb = XLBuilderEx()
        head = ["Дата", "Данные"]
        xlb.makeHead(sheet, 1, head)
        
        row = 2
        
        for i in p.items:
            rd = [i.data.strftime('%d/%m/%Y %H:%M'), i.text]
            xlb.makeCells(sheet, row, rd)
            row += 1
            
        cc = 1
        for w in [25,50]:
            sheet.column_dimensions[get_column_letter(cc)].width = w
            cc += 1

    return wb

def collectAgent(server):
    result = []
    a = server.Get("Agents", "", "id")
    ids = server.Params[0].ids
    
    for ai in ids.split(','):
        if ai in a:
            result.append(a[ai])
    
    return result      

def wrapQuotes(ids):
    res = '';
    
    for i in ids.split(','):
        if len(res)>0:
            res += ","
            
        res += "'";
        res += i;
        res += "'";
        
    return res    
        
def loadSyncInfo(server):
    st = server.Params[0].start
    fn = server.Params[0].finish
    ids = server.Params[0].ids
    
    w = '"created"' + " >= ToDate('" + st.strftime("%d/%m/%Y 0:0:0") + "') and " +\
            '"created"' + " < ToDate('" + fn.strftime("%d/%m/%Y 0:0:0") + "') and " +\
            '"userid"' + " in (" + wrapQuotes(ids) + ")"
    
    return server.Get("SyncInfo", w)  

class Page:
    agent = None
    items = None
     
    def __init__(self, a, sy):
        self.agent = a.name
        self.items = []
         
        for s in sy:
            self.items.append(Item(s))
         
class Item:
    data = None
    text = None    
    
    def __init__(self, s):
        self.data = s.created
        self.text = self.parseFlag(s.syncparam)
        
    def parseFlag(self, fl):
        res = '';
        f = int(fl)
        
        if (f & 1) == 1:
            if len(res) > 0:
                res += ", "
            res += 'Очистить базу при приеме'
        
        if (f & 2) == 2:
            if len(res) > 0:
                res += ", "
                
            res += 'Основные данные'
            
        if (f & 4) == 4:
            if len(res) > 0:
                res += ", "
                
            res += 'Документы'
            
        if (f & 8) == 8:
            if len(res) > 0:
                res += ", "
                
            res += 'Фотоотчеты(посещения)'            
            
        if (f & 32) == 32:
            if len(res) > 0:
                res += ", "
                
            res += 'Презентация'      

        if (f & 128) == 128:
            if len(res) > 0:
                res += ", "
                
            res += 'Отгрузки и долги'
            
        if (f & 256) == 256:
            if len(res) > 0:
                res += ", "
                
            res += 'Восстановить заявки'
                             
        return res
    
report = []
     
def loadData(server):
    agents = collectAgent(server)
    syncinfo = loadSyncInfo(server)
    
    mp = dict()
    
    for s in syncinfo:
        if not s.userid in mp:
            mp[s.userid] = list();
            
        mp[s.userid].append(s)    
    
    for a in agents:
        if a.id in mp:
            report.append(Page(a, mp[a.id]))
    
    agents.sort(key=lambda x: x.name.lower())
