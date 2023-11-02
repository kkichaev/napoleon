# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Fill
from openpyxl.style import Color
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

from importlib import reload

import sys;
import tempfile
import io
import time

reload(sys)

class Item:
    date = None
    org = None
    address = None
    vizit = 0
    order = 0
    sum = 0
        
    def __init__(self):
        self.date = datetime.now()
        self.org = ""
        self.address = ""
        self.vizit = 0
        self.order = 0
        self.sum = 0
            
class reversor:
    def __init__(self, obj) -> None:
        self.obj = obj

    def __eq__(self, o) -> bool:
        return self.obj == o

    def __lt__(self, other):
        return other.obj < self.obj

class Report:
    agent = None
    start = None
    finish = None
    ordavg = None
    vztavg = None
    ordcnt = None
    vztcnt = None
    org = None
    items = None
    sum = None
    corg = None
    
    def __init__(self, server):
        userid = server.Params[0].userid
        self.agent = server.Params[0].agent
        self.start = server.Params[0].start
        self.finish = server.Params[0].finish
        
        agent = server.Get("Agent", '', 'id') 
        
        where = "'{0}'".format(userid)
        server.ChangeUser(where)
        self.org = server.Get("Org", "", 'id')
        server.RestoreUser()
        self.corg = server.Get("Org", "", 'id')
        
        where = '"created"' + " >= ToDate('" + self.start.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                    '"created"' + " < ToDate('" + self.finish.strftime("%d/%m/%Y 23:59:59") + "') and " +\
                    '"userid"' + " in ('" + userid + "')"
        
        data = dict() #ключ: дата + id 
        orders = server.Get("Order", where)
        
        for o in orders:
            key = o.created.strftime("%d/%m/%Y")+o.id
            
            item = None
            
            if not key in data:
                item = Item();
                item.date = o.created
                
                org = self.getOrg(o.id)
                
                if org != None:
                    item.org = org.name
                    item.address = org.address
                else:
                    item.org="контрагент код<{0}>".format(o.id)
                    
                sum = 0
                
                for oi in o.items:
                    sum = sum + oi.sum    
                
                item.sum = sum    
                item.order = 1
                item.vizit = 1
                
                data[key] = item
                
        visits = server.Get("Visit", where)
        
        for v in visits:
            self.initVisit(data, v)
                
        remnants = server.Get("OrgRemnants",where)    
        
        for r in remnants:        
            self.initVisit(data, r)
            
        lst = list()
        
        for d in data.values():
            lst.append(d)    
    
        lst.sort(key=lambda x: (x.date, reversor(x.org)))
        summary = dict()
                
        for i in lst:
            item = None
            
            key = i.date.strftime("%d.%m.%y")
            
            if key in summary:
                item = summary[key] 
            else:
                item = Item()
                item.date = i.date
                summary[key] = item

            if item.date < i.date : item.date = i.date                
            item.order = item.order + i.order
            item.vizit = item.vizit + i.vizit
            item.sum = item.sum + i.sum
            
        self.ordcnt = 0
        self.vztcnt = 0     
        self.sum = 0
        
        for s in summary.values():
            lst.append(s)
            self.ordcnt = self.ordcnt + s.order
            self.vztcnt = self.vztcnt + s.vizit
            self.sum = self.sum + s.sum
            
        cnt = len(summary)
        
        if cnt > 0:
            self.ordavg = self.ordcnt / cnt
            self.vztavg = self.vztcnt / cnt
            
        self.items = list()
        lst.sort(key=lambda x: (x.date, reversor(x.org)))
        
        for i in lst:
            self.items.append(i)
       
    def initVisit(self, data, d):
        key = d.created.strftime("%d/%m/%Y")+d.id
            
        item = None
        
        if key in data:
            item = data[key]
        else:
            item = Item()
            item.date = d.created
             
            org = self.getOrg(d.id)

            if org != None:
                item.org = org.name
                item.address = org.address
            else:
                item.org="контрагент код<{0}>".format(d.id)

            data[key] = item
            
        if item.vizit == 0:
            item.vizit = 1        
                            
    def getOrg(self, id):
        result = None
        
        if id in self.org:
            result = self.org[id]
        elif id in self.corg:
            result = self.corg[id]
                
        return result                 
                    
def repName():
    return "res.xlsx"

def weekItemsEquals(w1, w2):
    result = len(w1) == len(w2)
                
    if result:
        for ii in range(0, len(w1)):
            result = w1[ii].name == w2[ii].name
            
            if not result:
                break

    return result
        
def weekEquals(w1, w2):
    result = len(w1) == len(w2)
    
    if result:
        wa = list()
        wb = list()

        wa.extend(w1)
        wb.extend(w2)
        
        wa.sort(key=lambda x: x.name)
        wb.sort(key=lambda x: x.name)
        
        for i in range(0,len(wa)):
            result = wa[i].name == wb[i].name
            
            if result:
                result = weekItemsEquals(wa[i].items, wb[i].items)
                
                if not result:
                    break
                            
    return result

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN
               
def createXML(report):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.cell(row=0, column=0).value = "Отчёт о посещениях и заказах."
    sheet.cell(row=1, column=0).value = "{2} Период {0}-{1}".format(report.start.strftime("%d.%m.%y"),report.finish.strftime("%d.%m.%y"), report.agent) 
    sheet.cell(row=2, column=0).value = "Посещений в среднем: {0} Заказов в среднем: {1}".format(report.vztavg, report.ordavg)
    
    sheet.cell(row=3, column=0).value = "Дата"
    sheet.cell(row=3, column=1).value = "Контрагент"
    sheet.cell(row=3, column=2).value = "Адрес"
    sheet.cell(row=3, column=3).value = "Посещения"
    sheet.cell(row=3, column=4).value = "Время посещения"
    sheet.cell(row=3, column=5).value = "Заказ"
    sheet.cell(row=3, column=6).value = "Сумма заказа"
    
    START_ROW = 4
    r = START_ROW
    for i in report.items:
        sheet.cell(row=r, column=0).value = i.date.strftime("%d.%m.%y") if i.org == "" else ""
        sheet.cell(row=r, column=1).value = i.org
        sheet.cell(row=r, column=2).value = i.address
        sheet.cell(row=r, column=3).value = i.vizit
        sheet.cell(row=r, column=4).value = i.date.strftime("%H:%M") if i.org != "" else ""
        sheet.cell(row=r, column=5).value = i.order
        sheet.cell(row=r, column=6).value = i.sum
        sheet.row_dimensions[r+1].outline_level = 0 if i.org == "" else 1
        r = r + 1
    
    sheet.cell(row=r, column=0).value = "ИТОГО"
    sheet.cell(row=r, column=3).value = report.vztcnt
    sheet.cell(row=r, column=5).value = report.ordcnt
    sheet.cell(row=r, column=6).value = report.sum
    
    sheet.column_dimensions[get_column_letter(1)].width = 11
    sheet.column_dimensions[get_column_letter(2)].width = 45
    sheet.column_dimensions[get_column_letter(3)].width = 35
    sheet.column_dimensions[get_column_letter(4)].width = 12
    sheet.column_dimensions[get_column_letter(5)].width = 18
    sheet.column_dimensions[get_column_letter(6)].width = 12
    sheet.column_dimensions[get_column_letter(7)].width = 16
    
    rangeBorders(sheet.range("A"+str(START_ROW)+":"+get_column_letter(7)+str(r+1)))    

    
    result = tempfile.gettempdir() + '/' + repName()
    wb.save(result)
    
    return result 

def doReport(server):
    report = Report(server)
    xml = createXML(report)
     
    file = io.open(xml, 'rb')
    bytes = file.read(-1)
    file.close()
  
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    obj = outObj.New()
    obj.name = repName()
    obj.file = bytes
      
    server.Put(outObj)
                   
def run(server):
    print ("start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    doReport(server)
    print ("finish\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))