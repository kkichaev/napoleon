# -*- coding: cp1251 -*-
'''
Created on 14 сент. 2016 г.

@author: kkichaev
'''

from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.style import Font, NumberFormat, Color, Fill, Alignment
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter

import sys
import tempfile
import time
from importlib import reload

reload(sys)
sys.setdefaultencoding("cp1251")

class Item:
    name = None
    workDays = None
    visCnt = None
    ordCnt = None
    sum = None
    progress = None
    
    def __init__(self):
        self.name = 0
        self.workDays = 0
        self.visCnt = 0
        self.ordCnt = 0
        self.cum = 0
        self.progress = 0
        self.sum = 0
        
    def getData(self, r):
        return [self.name, self.workDays, self.visCnt, "=IFERROR(C{0}/B{0},0)".format(r+1), self.ordCnt, "=IFERROR(E{0}/B{0},0)".format(r+1), 
                self.sum, "=IFERROR(E{0}/C{0}*100,0)".format(r+1), self.progress * 100]    

class Group(Item):
    items = None
    
    def __init__(self):
        Item.__init__(self)
        self.items = list()

class Report:
    start = None
    finish = None
    items = None
    
    def __init__(self, server):
        self.start = server.Params[0].start
        self.finish = server.Params[0].finish  
        self.items = list()
        
        divids = server.Params[0].ids
        
        where = '"id" in ({0})'.format(divids)
        dvs = server.Get("Division", where)
        ags = server.Get("Agents", '', 'id')
        
        startStr = self.start.strftime("%d/%m/%Y")
        finishStr = self.finish.strftime("%d/%m/%Y")
        
        uids = ""
        
        for d in dvs:
            for a in d.agents:
                if len(uids) > 0:
                    uids += ","
                uids += "'" + a.id + "'"    
                 
        vd = self.collectDocs(server.Get("VisitWithPhoto", startStr + ";" +  finishStr + ";" + uids))
        where = '"created" >= ToDate(\'{0}\') and "created" < ToDate(\'{1}\') and "userid" in ({2})'.format(startStr, finishStr, uids) 
        od = self.collectDocs(server.Get("Order", where))
          
        for d in dvs:
            g = Group()
            g.name = d.name
            
            self.items.append(g)
            
            wdo = list()
            prgso = list()
            
            for a in d.agents:
                i = Item()
                i.name = ags[a.id].name if a.id in ags else "Торговый агент с кодом <{0}>".format(a.id)
                g.items.append(i)
                
                if a.id in vd:
                    vdl = vd[a.id]
                    vdl = sorted(vdl, key=lambda x: x.date)
                    
                    wd = list()
                    vm = dict()
                    
                    # Из визитов с фотографиями строим списки
                    # wd - рабочие дни агента
                    # wdo - рабочие дни подразделения
                    # vm - по дате список id организаций из визитов
                    for v in vdl:
                        d = datetime(v.date.year, v.date.month, v.date.day)
                        
                        if not d in wd:
                            wd.append(d)
                        
                        if not d in wdo:
                            wdo.append(d)
                            
                        if not d in vm:
                            vm[d] = list()
                        
                        vm[d].append(v.id)        
                                
                    i.workDays = len(wd)
                    i.visCnt = len(vdl)
                    
                    # Считаем заявкии сумму
                    if a.id in od:
                        i.ordCnt = self.calcOrderCnt(od[a.id])
                        i.sum = 0
                        for  o in od[a.id]:
                            for oi in o.items:
                                i.sum += oi.qty * oi.cost
                
                    
                    server.ChangeUser("'" + a.id + "'")
                    of = server.Get("OrgFolder", "")
                    cfg = server.Get("ServerConfig", "")
                    server.RestoreUser()
                    
                    days = {"Monday" : "Понедельник", 
                            "Tuesday" : "Вторник",
                            "Wednesday" : "Среда",
                            "Thursday" : "Четверг",
                            "Friday" : "Пятница",
                            "Saturday" : "Суббота",
                            "Sunday" : "Воскресенье" }
                    
                    # список прогрессов по рабочим дням
                    ppd = list();
                    
                    for w in wd:
                        p = self.plannedOrgs(of, cfg, a.id, w, days)
                        vc = 0.0
                        
                        if w in vm:
                            for o in vm[w]:
                                if o in p: 
                                    vc += 1
                        sp = len(p)
                        
                        pr = vc / sp if sp != 0 else 0
                        ppd.append(pr)
                        
                    ps = 0.0
                    
                    for s in ppd:
                        ps += s;    
                    
                    i.progress = ps / len(ppd) if len(ppd) != 0 else 0
                    
                    prgso.append(i.progress)
                                    
                    g.visCnt += i.visCnt
                    g.ordCnt += i.ordCnt
                    g.sum += i.sum
                else:
                    prgso.append(0.0)     
            
            g.workDays = len(wdo)   
            g.items = sorted(g.items, key=lambda x: x.name)
            
            ps = 0.0
                    
            for s in prgso:
                ps += s;
                  
            g.progress =  ps / len(prgso) if len(prgso) != 0 else 0
            
    def calcOrderCnt(self, ords):
        res = list()
        
        for o in ords:
            key = o.created.strftime("%d.%m.%Y") + o.id
            
            if not key in res:
                res.append(key)
        
        return len(res)                  

    def plannedOrgs(self, of, cfg, id, date, days):
        plans = list()
        d = days[date.strftime("%A")]
        widx = self.getWeekIndex(cfg, date, id)
       
        for f in of:
            if f.name == d or f.name == str(widx) + d:
                for i in f.items:            
                    plans.append(i.name)
                
        return plans
    
    def getWeekIndex(self, cfg, data, id):  
        scStart = None
        result = -1 
            
        if cfg != None:
            for c in cfg:
                if c.key == 'SheduleStart' and len(c.value) > 0:
                    scStart = datetime(*(time.strptime(c.value, '%Y-%m-%d')[0:6]))
                    break
                
        if scStart != None:
            d = data - scStart
            result = ((d.days / 7) % 4) + 1;
        
        return result                                
                    
    def collectDocs(self, docs):
        result = dict()
        
        for d in docs:
            if not d.userid in result:
                result[d.userid] = list(); 
                
            result[d.userid].append(d)
        
        return result    

class XLBuilderEx(XLBuilder):
    def makeCells(self, sheet, row, values, color="ffffffff", bold = False, startColumn = 0):
        cc = startColumn
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                self.makeCell(sheet, row, cc, c, value, color, bold)
            cc += 1
            
    def makeCell(self, sheet, row, column, cell, value, color, bold):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        self.setBackColor(cell, color)

        if bold:
            self.makeBorder(cell, XLBuilderEx.HEAD_BORDER_STYLE)
            cell.style.font.bold = True
             
        if column > 0:
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER)
            
def ptintSheet(xlb, sh, data):
    c = sh.cell(row=0,column=0)
    c.value = "Итоговый отчёт подразделения";
    c.style.font.bold = True
    c.style.font.size = 13.5
    
    c = sh.cell(row=2,column=0)
    c.value = "Период {0} - {1}".format(data.start.strftime("%d.%m.%Y"), data.finish.strftime("%d.%m.%Y"))

     
    head = ["Подразделение/агент", "Рабочих дней", "Визиты", "Ср. визит", "Заявки", "Ср. заявки", "Сумма", "Процент заявок", "Прогресс"]
    START_ROW = 4
    r = START_ROW
    xlb.makeHead(sh, r, head, True)
    
    cls = ['ffE5E4E2', 'ffE0FFFF' ]
    gr = []
    
    for x in range(0, len(data.items)):
        d = data.items[x]
        r += 1
        c = cls[x % 2]
        xlb.makeCells(sh, r, d.getData(r), c)
        gr.append(r + 1)
        
        for i in d.items:
            r += 1
            xlb.makeCells(sh, r, i.getData(r), c)
    
    r += 1
    d = ["Итого", "", sumExcelItems(gr, "C"), "", sumExcelItems(gr, "E"), "", sumExcelItems(gr, "G"), "{0}/{1}".format(sumExcelItems(gr, "H"), len(gr)), ""]
    
    xlb.makeCells(sh, r, d, bold = True)
            
    cc = 1
    for w in [45,10,10,10,10,10,15,10,10]:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1    
    
def sumExcelItems(rows, col):
    s = ""
    
    for r in rows:
        if len(s) > 0:
            s += ","
            
        s += col + str(r)
        
    return "=SUM({0})".format(s)     
   
def printOut(report):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    Font.DEFAULT_FONT_NAME = 'Arial'
    Font.DEFAULT_FONT_SIZE = 10 
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, report)
                
    return wb

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
    
def doReport(server):
    wb = printOut(Report(server))
    workbookToObject(wb, "mtxtime.xlsx", server)

def run(server):
    print ("start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    doReport(server)
    print ("finish\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S'))