# -*- coding: cp1251 -*-
import tempfile
# import io

from openpyxl.style import Color, Fill, Border, Alignment

class XLBuilder:
    
    HEAD_BORDER_STYLE = Border.BORDER_MEDIUM
    
    def __init__(self):
        pass

    def makeBorder(self, cell, border):
        borders = cell.style.borders 
        
        borders.top.border_style = border
        borders.bottom.border_style = border
        borders.right.border_style = border  
        borders.left.border_style = border
        
    
    def adjustHeadCell(self, sheet, cell, row, column):
        return column
    
    def setBackColor(self, cell, color):
        fill = cell.style.fill 
        fill.start_color = Color(color)
        fill.end_color = Color(color)
        fill.fill_type = Fill.FILL_SOLID
    
    def paintHeadCell(self, cell):
        self.makeBorder(cell, self.HEAD_BORDER_STYLE)
        
        style = cell.style
        style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        style.alignment.vertical = Alignment.VERTICAL_CENTER
        style.font.bold = True
    
    def makeHead(self, sheet, row, titles, wrap_text = False, startColumn = 0):
        cc = startColumn
        for title in titles:
            if title != None :
                c = sheet.cell(row=row, column=cc)
                c.value = title
                
                self.paintHeadCell(c)
                c.style.alignment.wrap_text = wrap_text
                cc = self.adjustHeadCell(sheet, c, row, cc)            
            cc += 1
    
    def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
        cell.value = value
        self.makeBorder(cell, border)
        cell.style.alignment. wrap_text = True
        
    def makeCells(self, sheet, row, values, startColumn = 0):
        cc = startColumn
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                self.makeCell(sheet, row, cc, c, value)
            cc += 1

    def makeCellsWithBorders(self, sheet, row, values, startColumn = 0):
        cc = startColumn
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                self.makeCell(sheet, row, cc, c, value[0], value[1])
            cc += 1
            
    def workbookToObject(self, wb, repName, server):
        server.RegisterType("Result[name:s,file:b]")
        outObj = server.New("Result")

        tFile = tempfile.TemporaryFile()
        wb.save(tFile)
        tFile.seek(0)
        bytesOut = tFile.read(-1)
        tFile.close()
        
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()
    
        obj = outObj.New()
        obj.name = repName
        obj.file = bytesOut
    
        server.Put(outObj)
               
