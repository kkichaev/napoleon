# -*- coding: cp1251 -*-
from importlib import reload
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute

import datetime

import sys;
reload(sys);
#sys.setdefaultencoding("cp1251")

class Item:
    __slots__ = ['userid', 'username', 'org_in_route', 'visit_in_route', 'visit_out_route', 'order_count', 'order_in_route' 'sum', 'start', 'finish']
    
    def __init__(self):
        self.userid = ''
        self.username = ''
        self.org_in_route = 0.0
        self.visit_in_route = 0.0
        self.visit_out_route = 0.0
        self.order_count = 0.0
        self.order_in_route = 0.0
        self.sum = 0.0
        now = datetime.datetime.now()
        self.start = None
        self.finish = None
        
    def getData(self, row):
      row += 1
      return [self.username, self.org_in_route, self.visit_in_route, self.visit_out_route, "=IFERROR(C{0}/B{0},0)".format(row), 
        self.order_count, self.order_in_route, self.sum, 
        self.order_in_route / (float)(self.visit_in_route) if self.visit_in_route != 0 else 0,
        self.start.strftime("%H:%M") if self.start != None else "", 
        self.finish.strftime("%H:%M") if self.finish != None else "", 
        "=IF((K{0}-J{0})=0,\"\", K{0}-J{0})".format(row)]
      
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()

def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    data = ReportData()
    
    for a in params.agents:
      item = Item()
      
      route = AgentRoute(server, a.id).getDayRoute(params.date)  
      route_ids = []
      
      for i in route:
        route_ids.append(i.id)
      
      item.org_in_route = len(route)
      
      server.ChangeUser("'" + a.id + "'")
      item.userid = a.id
      item.username = server.CurrentUser().name
      server.RestoreUser()
      
      where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
          "'"+a.id+"'",
          params.date.strftime("%d/%m/%Y 0:0:0"),
          params.date.strftime("%d/%m/%Y 23:59:59"))

      orgs_out_route = []
      orgs_order_in_route = []
      
      orders = server.Get("Order", where)
      item.order_count = len(orders)

      for o in orders:
        if not o.id in route_ids and not o.id in orgs_out_route:
          orgs_out_route.append(o.id)
        
        if o.id in route_ids and not o.id in orgs_order_in_route:
          orgs_order_in_route.append(o.id)
          
        for i in o.items:
          item.sum += i.cost * i.qty

      
      scripts = server.Get("ScriptDoc", where)
      
      for s in scripts:
        if not s.id in route_ids and not s.id in orgs_out_route:
          orgs_out_route.append(s.id)
      
      visits = server.Get("VisitInfo", where)
      visit_in_route = []
      
      for v in visits:
        if v.id in route_ids and not v.id in visit_in_route:
          visit_in_route.append(v.id)
        elif not v.id in route_ids and not v.id in orgs_out_route:
          orgs_out_route.append(v.id)
         
        if item.start == None or item.start > v.created:
          item.start = v.created
          
        if item.finish == None or item.finish < v.created:
          item.finish = v.created
          
      remnants = server.Get("OrgStock", where)
      
      for r in remnants:
        if not r.id in route_ids and not r.id in orgs_out_route:
          orgs_out_route.append(r.id)
          
      answer = server.Get("Answer", where)    
      
      for a in answer:
        if not a.id in route_ids and not a.id in orgs_out_route:
          orgs_out_route.append(a.id)
      
      incass = server.Get("Incass", where)
      
      for i in incass:
        if not i.id in route_ids and not i.id in orgs_out_route:
          orgs_out_route.append(i.id)
          
      task = server.Get("TaskDone", where)
      
      for t in task:
        if not t.id in route_ids and not t.id in orgs_out_route:
          orgs_out_route.append(t.id)
          
      #print "orgs_out_route", orgs_out_route
      
      item.visit_out_route = len(orgs_out_route)
      item.order_in_route = len(orgs_order_in_route)
      item.visit_in_route = len(visit_in_route)
      
      data.data.append(item)
      
    data.data = sorted(data.data, cmp=lambda lhs: lhs.username)
    
    return data
    
class XLBuilderEx(XLBuilder):
  RED    = 'FFFF0000'
  ORANGE = 'FFFFC000'
  GREEN  = 'FF00FF00'
  HEAD   = 'FFF2F2F2'
  
  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 4 or column == 8:
      cell.style.number_format._set_format_code('0%')
    
    if column == 9 or column == 10 or column == 11:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME3)
      
    if column == 8:
      if value <= 0.5:
        self.setBackColor(cell,XLBuilderEx.RED)
      elif value <= 0.9:
        self.setBackColor(cell,XLBuilderEx.ORANGE)
      else:   
        self.setBackColor(cell,XLBuilderEx.GREEN)
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=1,column=0)
    c.value = "Ежедневный отчет"
    c.style.font.bold = True
    c.style.font.size = 18
    
    c = sheet.cell(row=3,column=0)
    c.value = "Дата: {0}".format(params.date.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 5
    
    head = ['Агенты', 'Кол-во ТТ по маршруту', 'Визит по маршруту', 'Визит вне маршрута', 'Эффективность визитов', 'Общее кол-во заявок по ТТ', 'Кол-во заявок по маршруту', "Сумма", "Результативность визитов", "Время начало", "Время конец", "Время на маршруте"]
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      values = item.getData(row)
      xlb.makeCells(sheet, row, values)
      row += 1
        
        
    cc = 1
    SZ = 12
    for w in [55,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    