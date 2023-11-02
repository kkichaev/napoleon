# -*- coding: cp1251 -*-
from importlib import reload
import logging

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from grsoft.route import AgentRoute

import tempfile
import sys;

reload(sys);
#sys.setdefaultencoding("cp1251")

class Report:
  __slots__ = {"divname", "start" , "finish", "items", "agents", "date"}
  
  def __init__(self):
      self.date = datetime.now()
      self.items = list()
      self.agents = None
      self.divname = ""

class DivItem:
  __slots__ = {'id', 'name'}
  
  def __init__(self):
    self.id = ''
    self.name = ''
    
  def getData(self):
    return ['\t' + self.name, '', '', '', '', ''];
    
class Item:
  __slots__ = {'id', 'name', 'start', 'finish', 'gpsoff', 'outoftime', 'progress'}
  
  def __init__(self):
    self.id = ""
    self.name = ''
    self.start = ''
    self.finish = ''
    self.gpsoff = 'нет'
    self.outoftime = 'нет'
    self.progress = 0.0
    
  def getData(self):
    return [self.name, self.start, self.finish, self.gpsoff, self.outoftime, self.progress]
      
def loadChildDivision(server, parent, report, params):
  childs = server.Get("Division","\"parent\"={0}".format(parent))
    
  for c in childs:
    di = DivItem()
    di.id = c.id
    di.name = c.name
    report.items.append(di)
    
    for a in c.agents:
      if a.id in report.agents:
        i = Item()
        i.id = a.id
        i.name = report.agents[a.id].name
        report.items.append(i)
        initAgentData(server, params, i)
    
    loadChildDivision(server, c.id, report, params)
        
def initAgentData(server, params, item):
  start = params.date
  finish = params.date
  arr = params.time.split("|")
  
  ar = AgentRoute(server, item.id)
  route = ar.getDayRoute(params.date)  
  route_ids = []
  route_doc = []
  
  for r in route:
    route_ids.append(r.id)
  
  if len(arr) == 2:
    t = arr[0].split(":");
    td1 = timedelta(hours=int(t[0]), minutes=int(t[1]))
    start = start + td1
    t = arr[1].split(":");
    td2 = timedelta(hours=int(t[0]), minutes=int(t[1]))
    finish = finish + td2

  where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+item.id+"'", 
    start.strftime("%d/%m/%Y 00:00:00"), finish.strftime("%d/%m/%Y 23:59:59"))
  
  docNames = ["Order", "VisitInfo", "OrgStock", "Answer", "Incass", "TaskDone", "Sales", "Answer", "Facing"]

  firstDocTime = None
  lastDocTime = None
  
  alldocs = []
  
  for name in docNames:
    docList = server.Get(name, where)
    
    if docList == None:
      continue
      
    for d in docList:
      alldocs.append(d)
      if (not d.id in route_doc) and d.id in route_ids:
        route_doc.append(d.id)

  alldocs = sorted(alldocs, key=lambda lhs: lhs.created)
  
  if len(alldocs) > 0:
    firstDocTime = alldocs[0].created
    lastDocTime = alldocs[-1].created
  
  time = None
  
  for d in alldocs:
    if time == None:
      time = d.created
    else:
      min = (d.created - time).total_seconds() / 60.0
      if min > 60:
        item.outoftime = "*да"
        break
  
  if firstDocTime == None or start < firstDocTime:
    item.start = '*';
    
  if lastDocTime == None or finish > lastDocTime:
    item.finish = '*'
  
  if firstDocTime != None:
    item.start += firstDocTime.strftime('%H:%M')  
    
  if lastDocTime != None:  
    item.finish += lastDocTime.strftime('%H:%M')  
  
  item.progress = float(len(route_doc)) / float(len(route_ids)) if len(route_ids) != 0 else 0.0
  
  where = '"userid"={0} and "date" >= ToDate("{1}") and "date" <= ToDate("{2}")'.format("'"+item.id+"'", 
    start.strftime("%d/%m/%Y %H:%M:00"), finish.strftime("%d/%m/%Y %H:%M:00"))

  ulog = server.Get("UserLog", where)
  
  for log in ulog:
    if log.action == 2:
      item.gpsoff = '*да'
      break
  
def loadData(params, server):
    res = Report()
    res.date = params.date
    res.agents = server.Get("Agents", "", "id")
    div = server.Get("Division","\"id\"={0}".format(params.division))
    
    if div != None and len(div) > 0:
      di = DivItem()
      di.id = div[0].id
      di.name = div[0].name
      res.items.append(di)
      res.divname = div[0].name
      
      for a in div[0].agents:
        if a.id in res.agents:
          i = Item()
          i.id = a.id
          i.name = res.agents[a.id].name 
          res.items.append(i)
          initAgentData(server, params, i)
    
    loadChildDivision(server, params.division, res, params)
    
    return res
    
def item_cmp(x, y):
  res = cmp(x.org, y.org)

  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    cell = sh.cell(row=0, column=0)
    cell.value = 'Аудит работы {0}'.format(data.divname)
    cell.style.font.bold = True

    sh.cell(row=2, column=0).value = 'Дата: {0}'.format(data.date.strftime("%d.%m.%Y"))
    
    head = ["Подразделение/агент", "Время начала работы", "Время окончания работы", "Отключение GPS",
      "Интервал отсутствия более 60 минут", "Процент выполения маршрута"]
    
    r = 4
    xlb.makeHead(sh, r, head)
    
    for d in data.items:
      r += 1
      xlb.makeCells(sh, r, d.getData())
    
    setCellWidth(sh, [90,12,12,12,12,12])
    
class XLBuilderEx(XLBuilder):
  RED = "FF0000"
  GREEN = "00FF00"
  YELLOW = "FFFF00"
  HEADER = "AFAFAF"

  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    self.setBackColor(cell, XLBuilderEx.HEADER)
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    if column == 5 and isinstance(value, float):
      cell.style.number_format._set_format_code('0%')
      color = XLBuilderEx.GREEN
      if value < 0.5:
        color = XLBuilderEx.RED
      elif value < 0.75:
        color = XLBuilderEx.YELLOW
      
      self.setBackColor(cell, color)
      
    if column == 1 or column == 2 or column == 3 or column == 4:
      if "*" in value:
        self.setBackColor(cell, XLBuilderEx.RED)
        value = value[1:]
        
    if column == 0:
      if "\t" in value:
        cell.style.font.bold = True
        
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)    
    
def printOut(d, params):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    XLBuilder().workbookToObject(wb, "auditreport.xlsx", server)         

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "notvisit_report.xlsx", server)                
    logging.info('end')

    