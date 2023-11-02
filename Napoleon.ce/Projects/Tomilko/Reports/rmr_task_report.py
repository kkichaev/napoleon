# -*- coding: cp1251 -*-
import logging
import sys;
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from manager.task import TaskReportData
from rmr_visit_report import Data, AgentSheet
from datetime import datetime
from openpyxl.cell import get_column_letter

reload(sys);
sys.setdefaultencoding("cp1251")

class RepData(Data):
  __slots__ = ["allorgs", "agents"]

  def __init__(self):
    Data.__init__(self)
    self.allorgs = dict()
    self.agents = None
  
  
def loadData(params, server):
  data = RepData()
  data.agents = server.Get('Agents', '', 'id')
    
  allorgs = {}
  
  for ai in params.userids:
    repData = TaskReportData()
    sheet = AgentSheet()
    sheet.id = id
    data.items.append(sheet)
    
    server.ChangeUser("'" + ai.id + "'")
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    sheet.name = server.CurrentUser().name
    server.RestoreUser()
      
    for k in orgs.keys():
      data.allorgs[k] = orgs[k]
  
    where = '"start" <= ToDate("{1}") and "finish" >= ToDate("{0}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 0:0:0"))
      
    where += ' and "userid"="' + ai.id + '"'
    tasks = server.Get('OrgTask', where)
    
    for doc in tasks:
      if doc.orgid in orgs:
        repData.addTask(doc)
    
    where = '"created" >= ToDate("{0}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"))
      
    where += ' and "userid"="' + ai.id + '"'
    doneTask = server.Get('TaskDone', where)
    
    for doc in doneTask:
      if doc.id in orgs:
        repData.addTaskDone(doc)
    
    sheet.items.append(repData)
  
  data.items = sorted(data.items, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))  
  
  return data

class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'
  
  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
    
def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  xlb = XLBuilderEx()
  
  for page in data.items:
    sheet.title = page.name
    cell = sheet.cell(row=0, column=0)
    cell.value = "Отчет по задачам: {0}".format(page.name)
    cell.style.font.bold = True
    
    DATE_FMT = "%d.%m.%Y %H:%M"
    sheet.cell(row=1, column=0).value = "Период: {0} - {1}".format(params.start.strftime(DATE_FMT), params.finish.strftime(DATE_FMT))
    
    row = 3
    
    xlb.makeHead(sheet, row, ["Название магазина", "Адрес", "Период", "Задача", "Выполнена (да /нет)", "Комментарий сотрудника", 'Агент', 'Дата выполнения', 'Постановщик', 'Дата создания'], True)
    
    minDate = datetime(1970,1,1)
    
    for repData in page.items:
      for item in repData.tasks:
          task = item.task
          
          if not task.orgid in data.allorgs:
            continue
            
          value = []
          
          o = data.allorgs[task.orgid]
          value.append(o.name)
          value.append(o.address)
          
          value.append("{0}-{1}".format(task.start.strftime("%d.%m.%Y"), task.finish.strftime("%d.%m.%Y")))
          value.append(task.text)
          value.append( 'Да' if item.isDone else 'Нет')
          value.append(item.comment)

          #task.comment)    
          
          if task.userid in data.agents:
              value.append(data.agents[task.userid].name)
          else:
              value.append(task.userid)
          
          
          value.append(item.responceDate.strftime("%d.%m.%Y") if item.responceDate != None else '')

          value.append(task.manager)
          if task.created < minDate: value.append('')
          else: value.append(task.created.strftime("%d.%m.%Y %H:%M"))

          row += 1
          xlb.makeCells(sheet, row, value)
          
      cc = 1
      for w in [20,20,15,15,15,35,35,15,20,20]:
          sheet.column_dimensions[get_column_letter(cc)].width = w
          cc += 1
        
    sheet = wb.create_sheet()
    
  return wb;
  
def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilder().workbookToObject(wb, "rmr_task_report.xlsx", server)                
  logging.info('end')