# -*- coding: cp1251 -*-

import sys
import locale
import logging
import datetime
import calendar
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat
from calendar import month_name


def AppendChildDivisions(divlist, prnts, pdiv):
    if pdiv.id in prnts:
        for d in prnts[pdiv.id]:
            divlist.append(d)
            AppendChildDivisions(divlist, prnts, d)

def loadAgents(server, divisionId):
    agents = list()
    
    div = server.Get('Division', '', 'id')
    prnts = dict()
    for d in div.values():
        if d.parent in prnts:
            prnts[d.parent].append(d)
        else:
            dlist = list()
            dlist.append(d)
            prnts[d.parent] = dlist
    divlist = list()
    if divisionId in div:
        pdiv = div[divisionId]
        divlist.append(pdiv)
        AppendChildDivisions(divlist, prnts, pdiv)
        
        for d in divlist:
            for ai in d.agents:
                if not ai.id in agents:
                    agents.append(ai.id)
    
    return agents    

def AddMonths(sourcedate,months):
    month = sourcedate.month - 1 + months
    year = int(sourcedate.year + month / 12 )
    month = month % 12 + 1
    day = min(sourcedate.day,calendar.monthrange(year,month)[1])
    
    return datetime.date(year,month,day)


class TotalData:
    __slots__ = ['visitOrgs', 'visits']
    
    def __init__(self):
        self.visitOrgs = 0
        self.visits = 0
        
    def addVisits(self, visits):
        if visits > 0:
            self.visitOrgs += 1
            self.visits += visits 

class PeriodData:
    __slots__ = ['start', 'finish', 'visits', 'docDates']
    
    def __init__(self, start, finish):
        self.start = start
        self.finish = finish + datetime.timedelta(days=-1)
        self.visits = dict()
        self.docDates = dict()
        
    def addDoc(self, doc):
        ret = False
        ddate = doc.date.date()
                
        if ddate >= self.start and ddate <= self.finish:
            ret = True
            dlist = None

            if not doc.id in self.docDates:
                dlist = list()
                self.docDates[doc.id] = dlist
            else:
                dlist = self.docDates[doc.id]
                
            if not ddate in dlist :
                qty = 0
                if doc.id in self.visits:
                    qty = self.visits[doc.id]
                self.visits[doc.id] = qty + 1
                
                dlist.append(ddate)
            
        return ret
    
    def getVisits(self, orgid):
        return self.visits[orgid] if orgid in self.visits else 0

class ReportData:    
    __slots__ = ['periods']

    def addInterval(self, cd, intervalType):
        ed = cd
        if intervalType == 0:
            ed = cd + datetime.timedelta(days=7)
        else:
            ed = AddMonths(cd, 1)
            
        return ed
    
    def __init__(self, params):
        self.periods = list()
        
        cd = params.start.date()
        endDate = params.end.date() + datetime.timedelta(days=1)
        
        while True:
            ed = self.addInterval(cd, params.intervalType)
            pdata = PeriodData(cd, ed)
            self.periods.append(pdata)
            
            if ed >= endDate: break
            cd = ed
        
    def addDoc(self, doc):
        for pdata in self.periods:
            if pdata.addDoc(doc): break
        
def loadData(server, params):
    companies = server.Get('Companies', '')

    orgs = dict()
    company = 'Код <' + params.idCompany + '>'
    if companies != None :
        for c in companies:
            if c.id == params.idCompany:
                company = c.name
                break
    
    orgsQuery = ""
    if len(params.id) != 0:
        orgsQuery = ' and "id" = ' + "'" + params.id + "'"

    sorgs = server.Get('CommonOrgs', '')
    for o in sorgs:
        if o.idCompany == params.idCompany:
            orgs[o.id] = o
    
    agents = loadAgents(server, params.division)
    userids = '"userid" in('
    for uid in agents:
        userids += "'" + uid + "',"        
    userids = userids[:-1] + ')'
    
    q = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.end.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + userids + orgsQuery
    
    data = ReportData(params)
    docs = server.Get('VisitInfo', q)
    if docs != None:
        for d in docs:
            if not d.id in orgs: continue
            data.addDoc(d)
            
    return (company, orgs, data)

def monthName(date):
    months = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
    return str(date.month) if date.month > 12 else months[date.month-1]

class XBuilder (XLBuilder):

    def makeTopHeadCell(self, sheet, row, cc, text, color, width):
        cell = sheet.cell(row=row, column=cc)
        cell.value = text
        sheet.merge_cells(start_row=row, start_column = cc, end_row = row, end_column = cc+2)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER            
        self.setBackColor(cell, color)
        
        row += 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = 'Количество ТТ посещенных минимум 1 раз за период'
        sheet.merge_cells(start_row=row, start_column = cc, end_row = row + 1, end_column = cc)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cell.style.alignment.wrap_text = True
        sheet.column_dimensions[get_column_letter(cc+1)].width = width

        self.makeBorder(cell, Border.BORDER_THIN)            
        cell = sheet.cell(row=row+1, column=cc)
        cell.style.borders.bottom.border_style = Border.BORDER_THIN  
        
        cc += 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = 'Количество посещений за период'
        sheet.merge_cells(start_row=row, start_column = cc, end_row = row + 1, end_column = cc)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER            
        cell.style.alignment.wrap_text = True
        sheet.column_dimensions[get_column_letter(cc+1)].width = width            

        self.makeBorder(cell, Border.BORDER_THIN)            
        cell = sheet.cell(row=row+1, column=cc)
        cell.style.borders.bottom.border_style = Border.BORDER_THIN  
        
        cc += 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = 'Количество визитов в одну ТТ из посещенных за период'
        sheet.merge_cells(start_row=row, start_column = cc, end_row = row + 1, end_column = cc)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER            
        cell.style.alignment.wrap_text = True            
        sheet.column_dimensions[get_column_letter(cc+1)].width = width            

        self.makeBorder(cell, Border.BORDER_THIN)            
        cell = sheet.cell(row=row+1, column=cc)
        cell.style.borders.bottom.border_style = Border.BORDER_THIN
        
        sheet.row_dimensions[row+2].height = 50  
            

    def makeHead(self, sheet, data, row, intervalText, totalData):
        cc = 3
        index = 1
        columnWidth = 20
        cell1Width = 30
        
        colors = [Color.DARKYELLOW, 'FF008080', Color.GREEN, Color.RED, Color.YELLOW]
        
        for pd in data.periods:
            headText = intervalText
            if len(intervalText) == 0:
                headText = monthName(pd.start)
            else:
                headText += ' ' +  str(index)
            self.makeTopHeadCell(sheet, row, cc, headText, colors[index % len(colors) - 1], columnWidth)
            cc += 3
            index += 1
            
            td = TotalData()
            totalData.append(td)

        self.makeTopHeadCell(sheet, row, cc, 'Отклонение (между  последним полным и первым интервалом)', colors[index % len(colors) - 1], columnWidth)

        row += 1
        cc = 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = 'Наименование торговой сети'
        self.makeBorder(cell, Border.BORDER_THIN)            

        cc = 2
        cell = sheet.cell(row=row, column=cc)
        cell.value = 'Количество ТТ в сети, всего'
        sheet.merge_cells(start_row=row, start_column = cc, end_row = row + 1, end_column = cc)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER            
        cell.style.alignment.wrap_text = True
        sheet.column_dimensions[get_column_letter(cc+1)].width = columnWidth
        
        self.makeBorder(cell, Border.BORDER_THIN)                    
        cell = sheet.cell(row=row+1, column=cc)
        cell.style.borders.bottom.border_style = Border.BORDER_THIN  
        
        row += 1
        cc = 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = 'Наименование грузополучателя'
        self.makeBorder(cell, Border.BORDER_THIN)            
        sheet.column_dimensions[get_column_letter(cc+1)].width = cell1Width
        
        return row + 1

def makeCompanyRow(sheet, cr, cc, company, orgsCount, totals):
    firstO = 0
    firstV = 0
    firstDiff = 0
    lastO = 0
    lastV = 0
    lastDiff = 0

    color = 'FFC07070'
    
    bld = XLBuilder()

    cell = sheet.cell(row = cr, column=cc)
    cell.value = company
    bld.setBackColor(cell, color)
    
    cc += 1
    cell = sheet.cell(row = cr, column=cc)
    cell.value = orgsCount
    bld.setBackColor(cell, color)

    isFirst = True
    for v in totals:
        diff = 0 if v.visitOrgs == 0 else float(v.visits) / float(v.visitOrgs)
        if isFirst:
            firstO = v.visitOrgs
            firstV = v.visits
            firstDiff = diff
            isFirst = False
        else:
            lastO = v.visitOrgs
            lastV = v.visits
            lastDiff = diff
            
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = v.visitOrgs
        bld.setBackColor(cell, color)
        
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = v.visits
        bld.setBackColor(cell, color)
        
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = diff
        cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
        bld.setBackColor(cell, color)
        
    cc += 1
    cell = sheet.cell(row = cr, column=cc)
    cell.value = lastO - firstO
    bld.setBackColor(cell, color)
    
    cc += 1
    cell = sheet.cell(row = cr, column=cc)
    cell.value = lastV - firstV
    bld.setBackColor(cell, color)
    
    cc += 1
    cell = sheet.cell(row = cr, column=cc)
    cell.value = lastDiff - firstDiff
    cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
    bld.setBackColor(cell, color)
    

def makeReport(company, orgs, data, params):
    wb = Workbook(False, 'cp1251')

    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    row = 1
    cc = 1
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Отчет по статистике визитов в торговые точки'
    row += 1
    
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Руководитель ___________________________'
    row += 1

    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Период: с {0} по {1}'.format(params.start.strftime("%d/%m/%Y"), params.end.strftime("%d/%m/%Y"))
    row += 1

    intervalText = 'неделя' if params.intervalType == 0 else ''
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Интервал: ' + intervalText
    row += 1

    row += 1
    
    totalData = list()

    bld = XBuilder()
    row = bld.makeHead(sheet, data, row, intervalText, totalData)
    
    orgColor = 'FF70C070'
    
    cr = row + 1
    for org in orgs.values():
        setBackColor = len(params.id) > 0 and params.id == org.id
        
        cc = 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = org.name + "/" + org.address
        if setBackColor: bld.setBackColor(cell, orgColor)
        
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = 1
        if setBackColor: bld.setBackColor(cell, orgColor)
        
        firstVisits = 0
        lastVisits = 0
        firstHaveVisit = 0
        lastHaveVisit = 0
        index = 0
        for pi in data.periods:
            visits = pi.getVisits(org.id)
            totalData[index].addVisits(visits)
            if index == 0: firstVisits = visits
            lastVisits = visits
            
            cc += 1
            cell = sheet.cell(row = cr, column=cc)
            haveVisit = 0 if visits == 0 else 1
            cell.value = haveVisit            
            if setBackColor: bld.setBackColor(cell, orgColor)
            if index == 0: firstHaveVisit = haveVisit
            lastHaveVisit = haveVisit
            
            cc += 1
            cell = sheet.cell(row = cr, column=cc)
            cell.value = visits
            if setBackColor: bld.setBackColor(cell, orgColor)
            
            cc += 1
            cell = sheet.cell(row = cr, column=cc)
            cell.value = visits
            if setBackColor: bld.setBackColor(cell, orgColor)
            index += 1

        diff = lastHaveVisit - firstHaveVisit 
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = 0 if diff == 0 else -1 if diff < 0 else 1
        if setBackColor: bld.setBackColor(cell, orgColor)
        
        diff = lastVisits - firstVisits
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = diff
        if setBackColor: bld.setBackColor(cell, orgColor)
        
        cc += 1
        cell = sheet.cell(row = cr, column=cc)
        cell.value = diff
        index += 1
        if setBackColor: bld.setBackColor(cell, orgColor)
            
        cr += 1

    makeCompanyRow(sheet, row, 1, company, len(orgs), totalData)    
    return wb

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting ")

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))

    (company, orgs, data) = loadData(server, params)
    wb = makeReport(company, orgs, data, params)
    
    XLBuilder().workbookToObject(wb, "visit_stat.xlsx", server)

    logging.info("ended")