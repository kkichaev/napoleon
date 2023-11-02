# -*- coding: cp1251 -*-

from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter

import sys

reload(sys)
sys.setdefaultencoding("cp1251")

def prepareActions(actionData):
    ret = list()
    for a in actionData:
        for item in a.items:
            if item.name in ret: continue
            ret.append(item.name)
        
    return ret

class OrgData:
    __slots__ = ['ho','showcase','corp_block','posm','action','partKI','partPF','items']
    
    def __init__(self):
        self.ho = False
        self.showcase = False
        self.corp_block = False
        self.posm = False
        self.action =''
        self.partKI = 0
        self.partPF = 0
        self.items = list()
        
    def add(self, doc):
        if doc.ho_best > 0 : self.ho = True
        if doc.showcase_best > 0 : self.showcase = True
        if doc.corp_block > 0 : self.corp_block = True
        if doc.posm > 0 : self.posm = True

        self.partKI = doc.share_ki
        self.partPF = doc.share_pf
        
        for item in doc.items:
            if not item.id in self.items:
                self.items.append(item.id)
                
class DateData:
    __slots__ = ['docs']
    
    def __init__(self):
        self.docs = dict()
        
    def add(self, doc):
        if not doc.id in self.docs:
            self.docs[doc.id] = OrgData()
        self.docs[doc.id].add(doc)
        
class AgentData:
    __slots__ = ['dateDocs']
    
    def __init__(self):
        self.dateDocs = dict()
        
    def add(self, doc):
        date = doc.created.date()
        if not date in self.dateDocs:
            self.dateDocs[date] = DateData()
        self.dateDocs[date].add(doc)
    

class ReportData:
    __slots__ = ['agentDocs']
    
    def __init__(self):
        self.agentDocs = dict()
        
    def add(self, doc):
        if not doc.userid in self.agentDocs:
            self.agentDocs[doc.userid] = AgentData()
        self.agentDocs[doc.userid].add(doc)

def loadData(server, param):
    whereUID = '"userid" in('
    for ai in param.agents:
        whereUID += "'" + ai.id + "',"
    whereUID = whereUID[:-1] + ')' 

    orgs = server.Get('CommonOrgs', '', 'id')
    if orgs == None or len(orgs) == 0:
        orgs = server.Get('Org', whereUID, 'id')

#     print len(orgs)
        
    price = server.Get('ManagerPrice','','id')
        
    agents = server.Get('Agents', '', 'id')

    where = '"date" >= (select ifnull(max("date"),0) from "StorcheckGoods" where "date" <= ToDate("{0}")) and "date" < ToDate("{1}")'.format(
        param.start.strftime("%d/%m/%Y 0:0:0"),
        param.finish.strftime("%d/%m/%Y 0:0:0"))
    goods = server.Get('StorcheckGoods', where)
    
    where = '"date" >= (select ifnull(max("date"),0) from "StorcheckActions" where "date" <= ToDate("{0}")) and "date" < ToDate("{1}")'.format(
        param.start.strftime("%d/%m/%Y 0:0:0"),
        param.finish.strftime("%d/%m/%Y 0:0:0"))
    actionData = server.Get('StorcheckActions', where)
#     print where
    actions = prepareActions(actionData)
    
    where = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")'.format(
        param.start.strftime("%d/%m/%Y 0:0:0"),
        param.finish.strftime("%d/%m/%Y 0:0:0"))
    where += ' and ' + whereUID
    docs = server.Get('Storcheck', where)
    
    data = ReportData()
    if docs != None:
        for doc in docs:
            data.add(doc)
            
    return data, orgs, price, agents, goods, actions

def getGoods(date, allGoods):
    curdoc = None
    for d in allGoods:
        ddate = d.date.date()
        if ddate <= date and (curdoc == None or curdoc.date.date() < ddate):
            curdoc = d
            
    newItems = list()
    topItems = list()
    
    if curdoc != None:
        for i in curdoc.items:
            if i.folder == 1:
                if not i.id in newItems: newItems.append(i.id)
            if i.folder == 2:
                if not i.id in topItems: topItems.append(i.id)
    return newItems, topItems

class XLBuilderEx(XLBuilder):
    PART_BG = 'FF99ccff'
    NEW_BG = 'FF339966'
    BASE_STANDART_BG = 'FFffcc00'
    
    def makeBorders(self, sheet, sr, er, sc, ec, border):
        cr = sr
        while cr < er:
            sheet.cell(row=cr, column=sc).style.borders.left.border_style = border
            sheet.cell(row=cr, column=ec).style.borders.left.border_style = border
            cr += 1
            
        cc = sc
        while cc < ec:
            sheet.cell(row=sr, column=cc).style.borders.top.border_style = border
            sheet.cell(row=er-1, column=cc).style.borders.bottom.border_style = border
            cc += 1
            
    
    def drawHead(self, sheet, newItems, topItems, price, row):
        cc = 0

        values = ['№', 'Наименование ТТ']
        for v in values:        
            cell = sheet.cell(row=row, column=cc)
            cell.value = v
            cell.style.font.size = 10
            sheet.merge_cells(start_row=row, start_column = cc, end_row = row+2, end_column = cc)
            self.paintHeadCell(cell)
            
            self.makeBorders(sheet, row, row+3, cc, cc+1, self.HEAD_BORDER_STYLE)
#             sheet.cell(row=row+2, column=cc).style.borders.bottom.border_style = self.HEAD_BORDER_STYLE
                         
            cc += 1
        spanCC = cc
        
        self.drawVCell(sheet, row+1, row+2, cc, 'Категория ТТ')
        
        cc = self.drawSecondHeader(sheet, row+1, cc+1, 'Лучшее место', ['ХО на лучшем месте','Витрина на лучшем месте'])
        
        values = ['Корпоративный блок','POSM','АКЦИЯ']
        for v in values:
            self.drawVCell(sheet, row+1, row+2, cc, v)
            cc+=1

        values = ['Доля полки КИ и Д 50%', 'Доля полки ПФ  40%']
        for v in values:
            self.drawVCell(sheet, row+1, row+2, cc, v, XLBuilderEx.PART_BG)
            cc+=1

        values = []
        for item in newItems:
            name = price[item].name if item in price else item
            values.append(name)            
        cc = self.drawSecondHeader(sheet, row+1, cc, 'НОВИНКИ', values, XLBuilderEx.NEW_BG)
            
        values = []
        for item in topItems:
            name = price[item].name if item in price else item
            values.append(name)            
        cc = self.drawSecondHeader(sheet, row+1, cc, 'Наличие ТОП 30', values)
        sheet.cell(row=row+1, column=cc).style.borders.left.border_style = self.HEAD_BORDER_STYLE
        
        cell = sheet.cell(row=row, column=spanCC)
        cell.value = 'БАЗОВЫЕ СТАНДАРТЫ'
        sheet.merge_cells(start_row=row, start_column = spanCC, end_row = row, end_column = cc-1)
        self.paintHeadCell(cell)
        self.setBackColor(cell, XLBuilderEx.BASE_STANDART_BG)
        self.makeBorders(sheet, row, row+1, spanCC, cc, self.HEAD_BORDER_STYLE)
#         sheet.cell(row=row, column=cc).style.borders.left.border_style = self.HEAD_BORDER_STYLE
        
        ctr = 1
        while True:  
            w = 6 if ctr != 2 else 30           
            sheet.column_dimensions[get_column_letter(ctr)].width = w
            ctr += 1
            if ctr > cc:
                break
        sheet.row_dimensions[row+2].height = 30
        sheet.row_dimensions[row+1].height = 30
            
        
        return row+3
    
    def drawVCell(self, sheet, rowStart, rowEnd, cc, value, bkColor = None):
        cell = sheet.cell(row=rowStart, column=cc)
        if rowStart != rowEnd:
            sheet.merge_cells(start_row=rowStart, start_column = cc, end_row = rowEnd, end_column = cc)
            self.makeBorders(sheet, rowStart, rowEnd+1, cc, cc+1, self.HEAD_BORDER_STYLE)
#             sheet.cell(row=rowEnd, column=cc).style.borders.bottom.border_style = self.HEAD_BORDER_STYLE
        cell.value = value
        cell.style.font.size = 10
        self.paintHeadCell(cell)
        cell.style.alignment.text_rotation = 90
        if bkColor != None:
            self.setBackColor(cell, bkColor)
        
    def drawSecondHeader(self, sheet, rowStart, cc, upHeader, values, bkColor = None):
        startCC = cc
        rc = rowStart + 1
        for v in values:
            self.drawVCell(sheet, rc, rc, cc, v, bkColor)
            cc += 1

        cell = sheet.cell(row=rowStart, column=startCC)
        cell.value = upHeader
        cell.style.font.size = 10
        cell.style.alignment.wrap_text = True 
        sheet.merge_cells(start_row=rowStart, start_column = startCC, end_row = rowStart, end_column = cc-1)
        self.paintHeadCell(cell)
        if bkColor != None:
            self.setBackColor(cell, bkColor)
                    
        return cc            
        
        
    def drawRow(self, sheet, ctr, org, categ, doc, newItems, topItems, row):
#         'ho','showcase','corp_block','posm','action','partKI','partPF','items'
        values = [ctr, org, categ]
        for prop in [doc.ho, doc.showcase, doc.corp_block, doc.posm, (len(doc.action) > 0)]:
            values.append('1' if prop else '0')
        
        values.append(doc.partKI)
        values.append(doc.partPF)
        
        cc = 8
        self.setBackColor(sheet.cell(row=row, column=cc), XLBuilderEx.PART_BG)
        cc += 1
        self.setBackColor(sheet.cell(row=row, column=cc), XLBuilderEx.PART_BG)

        for i in newItems:
            values.append( '1' if i in doc.items else '0')
            cc += 1
            self.setBackColor(sheet.cell(row=row, column=cc), XLBuilderEx.NEW_BG)

        for i in topItems:
            values.append( '1' if i in doc.items else '0')
        
        self.makeCells(sheet, row, values)
   
def drawReportData(row, sheet, xlb, agent, date, datedocs, newItems, topItems, orgs, price):
    cc = 1
    row += 1
    startRow = row
    
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Дата заполнения:'
    cell = sheet.cell(row=row, column=cc + 1)
    cell.value = date.strftime("%d/%m/%Y")
    
    row += 1
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Торговый представитель'
    cell = sheet.cell(row=row, column=cc + 1)
    cell.value = agent
    
    row += 1
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Территория'
    
    row += 1
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Кол-во проанализированных ТТ'
    cell = sheet.cell(row=row, column=cc + 1)
    cell.value = len(datedocs.docs)
    
    row = xlb.drawHead(sheet, newItems, topItems, price, row+2)
    
    ctr = 1
    for id, doc in datedocs.docs.iteritems():
        org = id
        categ = ''
        if id in orgs:
            o = orgs[id]
            org = o.name
            categ = o.category 
        xlb.drawRow(sheet, ctr, org, categ, doc, newItems, topItems, row)
        row += 1
    return row + 1  
    
def makeReport(data, orgs, price, agents, allGoods, actions):
    wb = Workbook(False, 'cp1251')    
    xlb = XLBuilderEx()
    sheet = wb.get_active_sheet()
    sheet.title = "Сторчек"
    row = 1
    cc = 1
    
    cell = sheet.cell(row=row, column=cc)
    cell.value = 'Акции'
    for ac in actions:
        cell = sheet.cell(row=row, column=cc + 1)
        cell.value = ac
        row += 1
    
    for uid, adocs in data.agentDocs.iteritems():
        agent = uid if not uid in agents else agents[uid].name
        for date, ddocs in adocs.dateDocs.iteritems():
            newItems, topItems = getGoods(date, allGoods) 
            
            row = drawReportData(row, sheet, xlb, agent, date, ddocs, newItems, topItems, orgs, price)

    return wb

def run(server):
    param = server.Params[0];
    print "storcheck start"

    data, orgs, price, agents, goods, actions = loadData(server, param)
    
    wb = makeReport(data, orgs, price, agents, goods, actions)
    
    XLBuilder().workbookToObject(wb, "storcheck_report.xlsx", server)
    print "storcheck done"
