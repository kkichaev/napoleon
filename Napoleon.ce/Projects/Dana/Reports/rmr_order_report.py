# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging
# import random

from openpyxl import Workbook, workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Fill, Color, Alignment, NumberFormat, Border
from datetime import datetime, date, timedelta
from rmr_report_style import XLBuilderCommon
from orgmap import OrgMap

reload(sys)
#sys.setdefaultencoding("cp1251")

class Report:
  def __init__(self):
    pass
    
  def getCells(self):
    ret = [''] * len(self.cellidx)
    
    for c in self.cellidx:
      if c in self.orgs:
        o = self.orgs[c]
        name = "{0} ({1}) ".format(o.name, o.address) 
      
      ret[self.cellidx[c]] = name
      
    return ret  

class Val:
  def __init__(self):
    self.qty = 0
    self.pack = 0

class Item:
  def __str__(self):
    return self.id.encode('cp866')
    
  def getData(self):
    res = [self.id, self.name]
    
    q,p = self.putData(res)  
    res.append(q)
    res.append(p)
    
    return res
  
  def putData(self, arr):
    q = 0
    p = 0
    
    for v in self.data:
      arr.append(v.qty)
      r = round(v.pack)
      arr.append(r)
      
      q += v.qty
      p+= r
      
    return q, p
    
  def getFolder(self):
    res = ['', self.fname]
    res.extend([''] * (len(self.data) * 2 + 2))
    
    return res
    
class PriceItem(Item):
  def __init__(self):
    Item.__init__(self)

def filterDocsOrgs(server, docs):
  map = {}
  orgMap = OrgMap(server)
  
  for d in docs:
    o = orgMap.getOrg(d.id, d.userid)
    map[d.id] = o
  
  return map
  
def cellIdxOrgs(map):
  cellidx = dict()   
  idx = 0
  
  for o in sorted(map.values(), key = lambda x: x.name):
    cellidx[o.id] = idx
    
    idx += 1
  
  return cellidx
    
def collectOrgs(server, userids, docs):
  map = filterDocsOrgs(server, docs)
  cellidx = cellIdxOrgs(map)
    
  return map, cellidx  
  
# def compare_item(x,y):
#   r = cmp(x.fname, y.fname)
#   
#   if r == 0:
#     r = cmp(x.name, y.fname)
#     
#   return r

def unpack(list):
  res = ''
  
  for i in list:
    if len(res) > 0:
      res += ','
      
    res += "'{}'".format(i.id)

  return res
  
# def loadData(params, server):
#   r = Report()
  
#   where = '"{0}" >= ToDate("{1} 0:0:0") and "{0}" < ToDate("{2} 0:0:0") and "userid" in ({3})'.format(
#     params.field,
#     params.start.strftime('%d.%m.%Y'), 
#     (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'),
#     unpack(params.userids))
  
#   ord = server.Get('Order', where)
#   price = server.Get('Price', 'setqtyfilter(false)', 'id')
#   folder = server.Get('Folder', '', 'fid')
  
#   logging.info('load data')

#   r.orgs, r.cellidx = collectOrgs(server, params.userids, ord)
  
#   logging.info('load org')

#   items = dict()
  
#   for o in ord:
#     for i in o.items:
#       if not i.id in items:
#         item = Item()
#         item.id = i.id
#         item.name = price[i.id].name if i.id in price  else '{}'.format(i.id)
#         items[i.id] = item
#         item.data = [Val() for x in range(len(r.cellidx))]
#         fid = price[i.id].fid if i.id in price  else ''
#         item.fname = folder[fid].name if fid in folder else '{}'.format(fid)
      
#       item = items[i.id]
      
#       if o.id in r.cellidx:
#         val = item.data[r.cellidx[o.id]]
#         val.qty += i.qty
        
#         if i.id in price:
#           pack = price[i.id].qtyInPack
          
#           if pack != 0:
#             val.pack += i.qty / pack
            
#   r.items = sorted(items.values(),key= lambda x: x.fname + "|" + x.name)
  
#   logging.info('handle data')

#   return r
  
import xlsxwriter as xl
import tempfile
import os

def loadData2(params, server):
  r = Report()
  
  stmt = """
  select "name", name_i, folder, id_i, min("qtyInPack") as qtyInPack, sum("qty") as qty from
  (
  select distinct o."created", o."userid", o."name", oi."name" as name_i, oi.folder as folder, oi."id" as id_i, oi."qtyInPack", oi."qty" from 
   (select o."created", o."userid", o."id", org."name", o."date" from "Order" o 
      left join org on o."id" = org."id") o, 
   (select o."id", o."qty", p."qtyInPack", p."name", p.folder, o."Order$created" as created, o."Order$userid" as userid from "Order$items" o
      left join (select p."id", p.inpack as "qtyInPack", p."name", f."name" as folder from Price p left join "Folder" f on p."folder" = f."fid") p
       on o."id" = p."id") oi 
    where o."userid" = oi."userid" and o."created" = oi."created" 
      and o."{0}" >= ToDate("{1} 0:0:0") and o."{0}" < ToDate("{2}") and o."userid" in ({3})
      ) o
	  group by folder, "name", name_i, id_i
    order by folder, name_i
  """.format(
      params.field,
      params.start.strftime('%d.%m.%Y'), 
      (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'),
      unpack(params.userids))

  ord = server.Query(stmt, "Orders[name@name_i:s,id@id_i:s,folder:s,qtyInPack:n(3),orgs(name)[qty:n(3),name:s]]")
  
  print(len(ord))
  logging.info('load data')

  orgs = list()
  
  for i in ord:
    for oi in i.orgs:
      if not oi.name in orgs:
        orgs.append(oi.name)
  
  logging.info('handle data')

  return (ord, sorted(orgs))

def printOut2(params, data, name, server):
  tFile = os.path.join(tempfile.gettempdir(), name)  
  wb = xl.Workbook(tFile)
  
  items, orgs = data

  sheet = wb.add_worksheet()

  bold = wb.add_format({'bold' : True})
  bold.set_font_size(18)

  sheet.write(0, 0, 'Отчет по заявкам', bold)
  sheet.write(1, 0, '{0}: c {1} по {2}'.format('период', params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y')))
  
  head = ['Артикул', 'Номенклатура']
  for o in orgs: head.append(o)
  head.append('Итого')

  boldHead = wb.add_format({'bold' : True})
  boldHead.set_text_wrap(True)
  boldHead.set_border()
  boldHead.set_bg_color('#F2f2f2')
  boldHead.set_align('center')
  boldHead.set_align('vcenter')

  boldHeadRot = wb.add_format({'bold' : True})
  boldHeadRot.set_rotation(90)
  boldHeadRot.set_text_wrap(True)
  boldHeadRot.set_border()
  boldHeadRot.set_bg_color('#F2f2f2')
  boldHeadRot.set_align('center')
  boldHeadRot.set_align('vcenter')

  cc = 0
  for h in head:
    if cc < 2:
      sheet.merge_range(2, cc, 3, cc, h, boldHead)
      cc += 1
    else:
      sheet.merge_range(2, cc, 2, cc+1, h, boldHeadRot)
      sheet.write(3, cc, "шт", boldHead)
      sheet.write(3, cc+1, "уп", boldHead)
      cc += 2


  sheet.set_row_pixels(2, 180)
  sheet.set_column(0, 0, 13)
  sheet.set_column(1, 1, 55)
  sheet.set_column(2, len(orgs) * 2 + 3, 4)

  cellFmt = wb.add_format({'text_wrap' : True})
  cellFmt.set_border()
  cellFmtBold = wb.add_format({'text_wrap' : True, 'bold' : True})
  cellFmtBold.set_border()

  row = 4
  fname = None
  for i in items:
    if fname == None or fname != i.folder:
      fname = i.folder
      sheet.write(row, 0, fname)
      row += 1

    totQty = 0
    totPack = 0
    values = [i.id, i.name]
    values.extend([None] * 2 * len(orgs))
    for oi in i.orgs:
      ind = 2 * orgs.index(oi.name) + 2
      totQty += oi.qty
      curPack = oi.qty // i.qtyInPack if i.qtyInPack != 0 else 0
      totPack += curPack
      values[ind] = oi.qty
      values[ind + 1] = curPack


    cc = 0  
    for c in values:
      sheet.write(row, cc, c, cellFmt)
      cc += 1

    for c in [totQty, totPack]:
      sheet.write(row, cc, c, cellFmtBold)
      cc += 1
    
    row += 1

  wb.close()
  file = open(tFile, 'rb')
  bytesOut = file.read(-1)
  file.close()

  server.RegisterType("Result[name:s,file:b]")
  outObj = server.New("Result")
  obj = outObj.New()
  obj.name = name
  obj.file = bytesOut

  server.Put(outObj)

class XLBuilderEx(XLBuilderCommon):
  def adjustHeadCell(self, sheet, cell, row, column):
    self.paintHeadCell(sheet.cell(row = row + 1, column=column))
    
    if column > 1:
      s = cell.style
      s.alignment.text_rotation = 90
      s.alignment.wrap_text = True 
      sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
      
      c = sheet.cell(row = row+1, column=column)
      c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
      c.value = 'шт'
      
      column += 1
      c = sheet.cell(row=row, column=column)
      self.paintHeadCell(c)
      
      c = sheet.cell(row = row+1, column=column)
      c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
      c.value = 'уп'
      self.paintHeadCell(c)
      
      SZ = 6
      sheet.column_dimensions[get_column_letter(column)].width = SZ
      sheet.column_dimensions[get_column_letter(column+1)].width = SZ
      
    return column

# def printOut(data, params):
#   wb = Workbook(False, 'cp1251')
#   sheet = wb.get_active_sheet()
#   xlb = XLBuilderEx()
  
#   c = sheet.cell(column=0,row=0)
#   c.value = 'Отчет по заявкам'
#   c.style.font.bold = True
#   c.style.font.size = 18
  
#   c = sheet.cell(column=0,row=1)
#   c.value='{0}: c {1} по {2}'.format('период', params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y'))

#   head = ['Артикул', 'Номенклатура']
#   head.extend(data.getCells())
#   head.append('Итого')
  
#   row = 2
#   xlb.makeHead(sheet, row, head)
#   sheet.row_dimensions[row+1].height = 100
  
#   row = 4
#   fname = None
  
#   for i in data.items:
#     if fname == None or fname != i.fname:
#       fname = i.fname
#       xlb.makeCells(sheet, row, i.getFolder())
#       sheet.cell(row=row, column=1).style.font.bold = True
#       row += 1
      
#     xlb.makeCells(sheet, row, i.getData())

#     sheet.cell(row=row, column=len(data.getCells() * 2) + 2).style.font.bold = True    
#     sheet.cell(row=row, column=len(data.getCells() * 2) + 2 + 1).style.font.bold = True
    
#     row += 1
    
#   x = 1
#   for w in [13,55]:
#     sheet.column_dimensions[get_column_letter(x)].width = w
#     x += 1
  
#   return wb

def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData2(params, server)
  printOut2(params, data, 'rmr_summary_report.xlsx', server)    
  logging.info('end')