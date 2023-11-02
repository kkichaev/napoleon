# -*- coding: cp1251 -*-
from importlib import reload
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys;
reload(sys);
#sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
    __slots__ = ['id', 'name', 'agent', 'date']
    
    def __init__(self):
        self.id = ''
        self.name = ''
        self.agent = ''
        self.date = ''
        
    def getData(self):
      res = [self.agent, self.date.strftime("%d/%m/%Y"), self.name, '']
      return res
        
      
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()

def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    
    data = ReportData() 
    
    for aid in params.agents.split(','):
      agent = ''
      server.ChangeUser("'" + aid + "'")
      agent = server.CurrentUser().name
      orgs = server.Get("Org", "", "id")
      server.RestoreUser()
      
      ar = AgentRoute(server, aid)
      
      date = params.start
      
      while date <= params.finish:
        route = ar.getDayRoute(date)  
        
        route_ids = []
        
        for r in route:
          route_ids.append(r.id)

        where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
            aid,
            date.strftime("%d/%m/%Y 0:0:0"),
            date.strftime("%d/%m/%Y 23:59:59"))
        
        docNames = ["Order", "VisitInfo", "OrgRemnants", "Answer", "Incass", "TaskDone"]

        for name in docNames:
          docList = server.Get(name, where)
          
          if docList == None:
            continue
            
          if docList != None:
            for d in docList:
              if d.id in route_ids:
                route_ids.remove(d.id)
        
        for id in route_ids:
          item = Item()
          item.agent = agent
          item.name = orgs[id].name if id in orgs else id
          item.date = date
          data.data.append(item)
          
        date += timedelta(days=1)
      
     
    data.data = sorted(data.data, cmp=lambda lhs: lhs.name)
    
    return data
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilder()
    row = 1
    
    head = ['№', 'Торговый Агент', 'Дата', 'Контрагент', 'Причина не посещения']
    xlb.makeHead(sheet, row, head, True)
    row += 1
    pos = 1
    for item in data.data:
      d = []
      d.append(pos)
      d.extend(item.getData())
      xlb.makeCells(sheet, row, d)
      row += 1
      pos += 1
        
    cc = 1
    for w in [5,35,15,35,50]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "notvisit_report.xlsx", server)                
    logging.info('end')
    