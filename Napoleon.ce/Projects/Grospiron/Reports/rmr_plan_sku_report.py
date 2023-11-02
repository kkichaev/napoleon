# -*- coding: cp1251 -*-
from importlib import reload
import logging

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from grsoft.route import AgentRoute

import tempfile
import sys;

reload(sys);
#sys.setdefaultencoding("cp1251")

class Report:
  
  def __init__(self):
      self.items = list()
      self.agents = None
      self.divname = ""

class DivItem:
  def __init__(self):
    self.id = ''
    self.name = ''
    self.plan = 0
    self.fact = 0
    
  def getData(self):
    return ['\t' + self.name, self.plan, self.fact, self.fact / self.plan  if  self.plan != 0 else 0];

  def setValue(self, plan , fact):
    if self.parent != None:
      self.parent.setValue(plan, fact)

    self.plan += plan
    self.fact += fact

class Item:
  def __init__(self):
    self.id = ""
    self.name = ''
    self.plan = 0
    self.fact = 0
    
  def getData(self):
    return [self.name, self.plan, self.fact, self.fact / self.plan  if  self.plan != 0 else 0]
      
def loadChildDivision(server, parent, report, params, parentItem, remnants):
  di = DivItem()
  di.id = parent.id
  di.name = parent.name
  di.parent = parentItem
  report.items.append(di)
  
  for a in parent.agents:
    if a.id in report.agents:
      i = Item()
      i.id = a.id
      i.name = report.agents[a.id].name
      i.plan = 0
      i.fact = 0

      report.items.append(i)

      if a.id in remnants:
        for r in remnants[a.id]:
          i.plan += r.plan - 1 if r.plan > 0 else 0
          i.fact += len(r.items)

      di.setValue(i.plan, i.fact)    

  childs = server.Get("Division","\"parent\"={0}".format(parent.id))
    
  for c in childs:
    loadChildDivision(server, c, report, params, di, remnants)
        
def loadData(params, server):
    res = Report()
    res.dataRange = '{0} - {1}'.format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    res.agents = server.Get("Agents", "", "id")
    div = server.Get("Division","\"id\"={0}".format(params.divid))

    where = '"{0}" >= ToDate("{1} 0:0:0") and "{0}" < ToDate("{2} 0:0:0")'.format(
      'created',
      params.start.strftime('%d.%m.%Y'), 
      (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'))

    stmt = '''
      select *, i.id as id_i from OrgRemnants as o
        left join OrgRemnants$items as i on o.created = i.[OrgRemnants$date] and o.userid = i.[OrgRemnants$userid]
        where not i.id in (select id from price where name='_Нет товара_') and {0}
      '''.format(where)
    remnants = server.Query(stmt, "Docs[id:s,userid:s,plan:n,items(userid,created)[id@id_i:s]]")
    rem_dic = {}

    for r in remnants:
      if not r.userid in rem_dic:
        rem_dic[r.userid] = []
      rem_dic[r.userid].append(r)

    loadChildDivision(server, div[0], res, params, None, rem_dic)
    
    return res
    
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    cell = sh.cell(row=0, column=0)
    cell.value = 'План/Факт SKU'
    cell.style.font.bold = True
    sh.cell(row=1, column=0).value = data.dataRange
    
    head = ["Подразделение/агент", "План SKU", "Факт SKU", "Процент"]
    
    r = 2
    xlb.makeHead(sh, r, head)
    
    for d in data.items:
      r += 1
      xlb.makeCells(sh, r, d.getData())
    
    setCellWidth(sh, [50,12,12,12])
    
class XLBuilderEx(XLBuilder):
  RED = "FF0000"
  GREEN = "00FF00"
  YELLOW = "FFFF00"
  HEADER = "AFAFAF"

  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    self.setBackColor(cell, XLBuilderEx.HEADER)
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    if column == 3 and isinstance(value, float):
      cell.style.number_format._set_format_code('0%')
        
    if column == 0:
      if "\t" in value:
        cell.style.font.bold = True
        
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)    
    
def printOut(d, params):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data = loadData(server)
    wb = printOut(data)
    XLBuilder().workbookToObject(wb, "auditreport.xlsx", server)         

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "notvisit_report.xlsx", server)                
    logging.info('end')

    