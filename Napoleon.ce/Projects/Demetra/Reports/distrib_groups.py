# -*- coding: cp1251 -*-

from importlib import reload
import sys
import datetime
import calendar

import time

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat
from datetime import timedelta
from openpyxl.writer.worksheet import row_sort


#reload(sys);
#sys.setdefaultencoding("cp1251")

days = ["Понедельник", 
        "Вторник",
        "Среда",
        "Четверг",
        "Пятница",
        "Суббота",
        "Воскресенье" ]

class DistrGroup:
    __slots__ = ('name','id','items','pos')

    def __init__(self, servObj):
        self.name = servObj.name
        self.id = servObj.id
        self.pos = servObj.pos
        self.items = []
        
        for val in servObj.items:
            if not val.id in self.items:
                self.items.append(val.id)
                
    def has(self, item):
        return item in self.items
    
class GroupData:
    __slots__ = ('exists','ordered','prevOrdered')

    def __init__(self):
        self.exists = False
        self.ordered = False
        self.prevOrdered = False

class OrgGroupData:
    __slots__ = ('groups','distrGroups')
    
    def __init__(self, distrGroups):
        self.groups = {}
        self.distrGroups = distrGroups
        
        for group in distrGroups:
            self.groups[group.id] = GroupData()
            
    def MarkExists(self, distrDoc):
        for item in distrDoc.items:
            if item.id in self.groups and item.exists > 0:
                self.groups[item.id].exists = True
            
            
    def CountItems(self, foo):
        count = 0
        for group in self.groups.values():
            if foo(group):
                count += 1
        return count
        
    def TotalOOS(self):
        return self.CountItems(lambda group: ((not group.exists) and group.prevOrdered))

    def TotalExists(self):
        return self.CountItems(lambda group: group.exists)

    def TotalOrdered(self):
        return self.CountItems(lambda group: (group.exists or group.ordered))

    def FindGroupID(self, item):
        for group in self.distrGroups:
            if group.has(item) :
                return group.id
            
        return None
    

    def MarkOrdered(self, orderDoc, curOrder):
        for item in orderDoc.items:
            group = self.FindGroupID(item.id)
            if group in self.groups:
                if curOrder :
                    self.groups[group].ordered = True
                else :
                    self.groups[group].prevOrdered = True

class DailyOrgKey:
    __slots__ = ('id', 'day')
    
    def __init__(self, doc):
        self.id = doc.id
        self.day = doc.date.date()
        
    def __hash__(self):
        val = hash(self.id)
        val ^= self.day.__hash__()
        
        return val
    
    def __lt__(self, other):
        if self.day < other.day:
            return True 
        elif self.day == other.day:
            return self.id < other.id
        else:
            return False

    def __cmp__(self, other):
        val = cmp(self.day, other.day)
        if val == 0: 
            val = cmp(self.id, other.id)
            
        return val
    

def IsInRoute(doc, route, server):
    curDay = days[doc.date.weekday()]
    widx = getWeekIndex(server, doc.date, doc.userid)
    wday = str(widx) + curDay
    
    if wday in route: 
        curDay = wday
    if curDay in route:
        for v in route[curDay].items:
            if v.name == doc.id :
                return True
    return False

scheduleStart = dict()
def getWeekIndex(server, data, agentid):  
    scStart = None
     
    if agentid in scheduleStart: 
        scStart = scheduleStart[agentid]
    else :
        where = '"userid"' + " in ('" + agentid + "')"
        cfg = server.Get("ServerConfig", where)
        
        if cfg != None:
            for c in cfg:
                if c.key == 'SheduleStart' and len(c.value) > 0:
                    scStart = datetime.datetime(*(time.strptime(c.value, '%Y-%m-%d')[0:6])) 
                    break
                
        scheduleStart[scStart] = scStart;
    
    result = -1
    
    if scStart != None:
        d = data - scStart
        result = ((d.days / 7) % 4) + 1;
    
    return result    

class DailyData:
    __slots__ = ('orgs', 'routeVisits')

    def __init__(self):
        self.orgs = {}
        self.routeVisits = []
#         for org in routeOrgs:
#             self.orgs[org.id] = OrgGroupData(distrGroups)
            
    def PutDistrDocs(self, route, docs, distrGroups, server):
        if docs == None: return
        
        routeDate = {}
        
        for doc in docs:
            curDay = days[doc.date.weekday()]
            
            widx = getWeekIndex(server, doc.date, doc.userid)
            wday = str(widx) + curDay
            
            if wday in route: 
                curDay = wday
            if curDay in route and not (curDay in routeDate):
                dayOrgs = []
                for v in route[curDay].items: dayOrgs.append(v.id)
                routeDate[curDay] = dayOrgs
            
            if not IsInRoute(doc, route, server): continue
            
#             if not doc.id in self.routeVisits: 
            self.routeVisits.append(doc.id)

            key = DailyOrgKey(doc)                    
            if not key in self.orgs:
                self.orgs[key] = OrgGroupData(distrGroups)
            self.orgs[key].MarkExists(doc)
                    

    def PutOrderDocs(self, docs, curDate):
        if docs == None: return

        for doc in docs:
            for key in self.orgs:
                docDate = doc.date.date()
                if (not key.id == doc.id) or key.day < docDate : continue
                self.orgs[key].MarkOrdered(doc, docDate == key.day)


def LoadDailyData(server, param):

    distrGroups = []    
            
    userid = "'" + param.userid + "'"
    
    server.ChangeUser(userid)
    orgs = server.Get("Org", "", "id")
    orgF = server.Get("OrgFolder", "", "name")
    distrObj = server.Get("DistribGroup", "")
    server.RestoreUser()

    if not distrObj == None:
        for do in distrObj:
            distrGroups.append(DistrGroup(do)) 
    
    ddata = DailyData()

    where = '"userid" = {0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
         userid,
         param.start.strftime("%d/%m/%Y 0:0:0"),
         param.finish.strftime("%d/%m/%Y 23:59:59"))
    
    docs = server.Get("DistribGroupDoc", where)
    ddata.PutDistrDocs(orgF, docs, distrGroups, server)
    
    #remove one moonth
    delta = -1
    m, y = (param.dailyReport.month + delta) % 12, param.dailyReport.year + ((param.dailyReport.month)+delta-1) // 12
    if not m: m = 12
    adddays = calendar.monthrange(y, m)[1]
#     adddays = 7
    startDate = param.start + datetime.timedelta(days=-adddays)
    
    where = '"userid" = {0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
         userid,
         startDate.strftime("%d/%m/%Y 0:0:0"),
         param.finish.strftime("%d/%m/%Y 23:59:59"))
    docs = server.Get("Order", where)
    ddata.PutOrderDocs(docs, param.dailyReport.date())
    
    distrGroups.sort(key= lambda x: x.pos)
    
    return (orgs, distrGroups, orgF, ddata)

class XLDailyBuilder(XLBuilder):
    
    def makeHead(self, sheet, row, titles, wrap_text=False):
        cc = 0
        for title in titles:

            if isinstance(title, tuple) :
                c = sheet.cell(row=row-1, column=cc)
                c.value = title[1]                
                self.paintHeadCell(c)
                c.style.alignment.wrap_text = wrap_text
                title = title[0]
            
            c = sheet.cell(row=row, column=cc)
            c.value = title 
            
            self.paintHeadCell(c)
            c.style.alignment.wrap_text = wrap_text

            cc = self.adjustHeadCell(sheet, c, row, cc)
            
            cc += 1
            
    def adjustHeadCell(self, sheet, cell, row, column):
        if column >= 1:
            cell.style.alignment.text_rotation = 90
        return column
    
    def makeCell(self, sheet, row, column, cell, value):
        if isinstance(value, tuple) :
            XLBuilder.makeCell(self, sheet, row, column, cell, value[0])
            if value[1] != Color.WHITE:
                cell.style.fill.end_color = Color(value[1])
                cell.style.fill.start_color = Color(value[1])
                cell.style.fill.fill_type = Fill.FILL_SOLID
        else :
            XLBuilder.makeCell(self, sheet, row, column, cell, value)
        if column > 1 :
            cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER

    def makeTotalCells(self, sheet, row, values, isTopRow, isBottomRow, setFormat = False):
        cc = 1
        for value in values:
            cell = sheet.cell(row=row, column=cc)
            cell.value = value
            
            borders = cell.style.borders 
            
            borders.top.border_style = Border.BORDER_THIN if not isTopRow else Border.BORDER_MEDIUM
            borders.bottom.border_style = Border.BORDER_THIN if not isBottomRow else Border.BORDER_MEDIUM
            borders.right.border_style = Border.BORDER_THIN  if cc != len(values) else Border.BORDER_MEDIUM
            borders.left.border_style = Border.BORDER_THIN if cc != 1 else Border.BORDER_MEDIUM

            cell.style.alignment. wrap_text = True
            if setFormat :
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            cc += 1
        


def MakeBaseData(sheet, orgs, distrGroups, ddata, xlb):
    head = ['', 'Топ Ассортимент']
    keys = list(ddata.orgs.keys())
    keys.sort()
    for key in keys:
        oid = key.id
        orgName = orgs[oid].name if oid in orgs else 'Контрагент с кодом <' + oid + '>'
        date = key.day.strftime("%d/%m/%y")
        tpl = orgName, date
        head.append(tpl)
    
    row = 2
    xlb.makeHead(sheet, row, head, True)
    row += 1
    sheet.row_dimensions[row].height = 150
    cc = 1
    for group in distrGroups:
        data = [cc]
        data.append(group.name)
        for key in keys:
            orgData = ddata.orgs[key]
            color = Color.WHITE
            grpData = orgData.groups[group.id]
            val = 'Н' if grpData.exists else 'П' if grpData.ordered else ''
            if grpData.exists:
                if grpData.ordered:
                    color = Color.GREEN
            elif grpData.prevOrdered:
                color = Color.RED
            data.append((val, color))
        
        xlb.makeCells(sheet, row, data)
        row += 1
        cc += 1
    
    return keys, row

def MakeLegend(sheet, xlb, row):
    cc = 2

    row += 2
    cell = sheet.cell(row=row, column=cc)
    style = cell.style
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell.value = 'Н'
    xlb.makeBorder(cell, Border.BORDER_THIN)
    cell = sheet.cell(row=row, column=cc+1)
    cell.value = 'Товар в наличии'
    
    row += 1
    cell = sheet.cell(row=row, column=cc)
    style = cell.style
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell.value = 'Н'
    xlb.makeBorder(cell, Border.BORDER_THIN)
    xlb.setBackColor(cell, Color.GREEN)
    cell = sheet.cell(row=row, column=cc+1)
    cell.value = 'Товар в наличии, сделана продажа'

    row += 1
    cell = sheet.cell(row=row, column=cc)
    xlb.makeBorder(cell, Border.BORDER_THIN)
    xlb.setBackColor(cell, Color.RED)
    cell = sheet.cell(row=row, column=cc+1)
    cell.value = 'OOS'
    
    row += 1
    cell = sheet.cell(row=row, column=cc)
    style = cell.style
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell.value = 'П'
    xlb.makeBorder(cell, Border.BORDER_THIN)
    xlb.setBackColor(cell, Color.RED)
    cell = sheet.cell(row=row, column=cc+1)
    cell.value = 'OOS, сделана продажа'

    row += 1
    cell = sheet.cell(row=row, column=cc)
    style = cell.style
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell.value = 'П'
    xlb.makeBorder(cell, Border.BORDER_THIN)
    cell = sheet.cell(row=row, column=cc+1)
    cell.value = 'Товар отсутствует, сделана продажа'

    row += 1
    cell = sheet.cell(row=row, column=cc)
    xlb.makeBorder(cell, Border.BORDER_THIN)
    cell = sheet.cell(row=row, column=cc+1)
    cell.value = 'Товар отсутствует'


def MakeDailyReport(sheet, orgs, distrGroups, ddata):
    xlb = XLDailyBuilder()
 
    keys, row = MakeBaseData(sheet, orgs, distrGroups, ddata, xlb)
    
    data1 = ['Наличие до продажи']
    data2 = ['Наличие с продажей']
    data3 = ['OOS']
    data4 = ['База']
    data5 = ['Процент наличия до продажи']
    data6 = ['Процент наличия с продажей']
    data7 = ['Процент OOS']
    
    cc = 3
    clmnStart = None
    clmnEnd = None
    
    for key in keys:
        orgData = ddata.orgs[key]
        data1.append(orgData.TotalExists())        
        data2.append(orgData.TotalOrdered())
        data3.append(orgData.TotalOOS())
        
        clmnName = get_column_letter(cc)
        if clmnStart == None : clmnStart = clmnName
        clmnEnd = clmnName
        
        data4.append('=' + clmnName + str(row+1) + '+' + clmnName + str(row+3))
#         data5.append('=' + clmnName + str(row+1) + '/' + str(row+4) + '*100')
        data5.append('=100-' + clmnName + str(row+7))
        data6.append('=' + clmnName + str(row+2) + '/' + str(len(distrGroups)) + '*100')
        data7.append('=' + clmnName + str(row+3) + '/' + clmnName + str(row+4) + '*100')
        
#         cell = sheet.cell(row=row+7, column=cc-1)
#         cell.value = ('=SUM(' + clmnName + str(row+7) + ':' + clmnName + str(row+6) + ')')
#         cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
        
        cc += 1
        
    xlb.makeTotalCells(sheet, row, data1, True, False)
    row += 1    
    xlb.makeTotalCells(sheet, row, data2, False, False)
    row += 1
    xlb.makeTotalCells(sheet, row, data3, False, False)
    row += 1
    xlb.makeTotalCells(sheet, row, data4, False, True)
    row += 1
    xlb.makeTotalCells(sheet, row, data5, True, False, True)
    row += 1
    xlb.makeTotalCells(sheet, row, data6, False, False, True)
    row += 1
    xlb.makeTotalCells(sheet, row, data7, False, True, True)

    row += 1
    cc = 1
    cell = sheet.cell(row=row, column=cc)
    cell.value = '% OOS за период'
    xlb.makeBorder(cell, Border.BORDER_THIN)
    if clmnStart != None:
        cc += 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = ('=SUM(' + clmnStart + str(row-4) + ':' + clmnEnd + str(row-4) + ') / SUM(' + clmnStart + str(row-3) + ':' + clmnEnd + str(row-3) + ') * 100')
        cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
        xlb.makeBorder(cell, Border.BORDER_THIN)
    
    row += 1
    cc = 1
    cell = sheet.cell(row=row, column=cc)
    cell.value = '% Дистрибьюции за период'
    xlb.makeBorder(cell, Border.BORDER_THIN)
    if clmnStart != None:
        cc += 1
        cell = sheet.cell(row=row, column=cc)
        cell.value = ('=SUM(' + clmnStart + str(row-4) + ':' + clmnEnd + str(row-4) + ")/(" + str(len(distrGroups)) + '*' + str(len(ddata.routeVisits)) + ') * 100')
        cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
        xlb.makeBorder(cell, Border.BORDER_THIN)
   
    MakeLegend(sheet, xlb, row)
    
    cc = 1
    for w in [7,70]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

def GetPrevIntervalStart(dateStart, dateEnd):
    adddays = 0
    if dateStart.day == dateEnd.day :
        if dateStart.month == dateEnd.month :
            adddays = 1
        else :
            delta = -1
            m, y = (dateStart.month + delta) % 12, dateStart.year + ((dateStart.month)+delta-1) // 12
            if not m: m = 12
            adddays = calendar.monthrange(y, m)[1]        
    else :
        adddays = (dateEnd - dateStart).days
        
    if adddays < 7 : adddays = 7 # week is min interval
    return dateStart + timedelta(days=-adddays)

class OrgPeriodReportData :
    __slots__ = ('visits', 'sku', 'oos')
    
    def __init__(self):
        self.visits = 0
        self.sku = 0
        self.oos = 0
     
    def AvgSKU(self) :
        return (0 if self.visits == 0 else self.sku / self.visits)   

    def AvgOOS(self) :
        return (0 if self.visits == 0 else self.oos / self.visits)   

class OrgPeriodDocData :
    __slots_ = ('sales', 'distrib')
    
    def __init__(self):
        self.sales = []
        self.distrib = {}

    def CountExists(self):
        count = 0
        for v in self.distrib.values() :
            if v : count += 1
        return count
    
    def CountOOS(self, saled):
        count = 0
        for k, v in self.distrib.items() :
            if not v and k in saled: count += 1
        return count

class OrgSaledData :
    __slots__ = ('dailyItems')
    
    def __init__(self):
        self.dailyItems = {}
        
    def FindGroupID(self, item, distrGroups):
        for group in distrGroups:
            if group.has(item) :
                return group.id
            
        return None
    def AddSalesDoc(self, doc, distrGroups):
        docData = None
        cd = doc.date.date()
        if not cd in self.dailyItems :
            docData = OrgPeriodDocData()
            self.dailyItems[cd]= docData
        else :
            docData = self.dailyItems[cd]
        
        items = docData.sales    
        for item in doc.items:
            grId = self.FindGroupID(item.id, distrGroups)
            if grId != None and (not grId in items):
                items.append(grId)
                        
    def AddDistribDoc(self, doc):
        docData = None
        cd = doc.date.date()
        if not cd in self.dailyItems :
            docData = OrgPeriodDocData()
            self.dailyItems[cd]= docData
        else :
            docData = self.dailyItems[cd]

        items = docData.distrib

        for item in doc.items:
            if not item.id in items:
                items[item.id] = item.exists > 0
            elif item.exists > 0 :
                items[item.id] = True
    
    def AddSaled(self, saled, items):
        for i in items :
            if not i in saled: 
                saled.append(i)
                        
    def GetPeriodData(self, oosStart, start, finish):
        data = OrgPeriodReportData()
        
        saled = []
        
        dates = list(self.dailyItems.keys())
        dates.sort()
        
        for day in dates :
            dailyData = self.dailyItems[day]
            if day < start :
                if day >= oosStart :
                    self.AddSaled(saled, dailyData.sales)
                continue
                        
            if day > finish :
                break
            
            data.visits += 1
            data.sku += dailyData.CountExists()
            data.oos += dailyData.CountOOS(saled)
       
            self.AddSaled(saled, dailyData.sales)
            
        return data

class SaledData :
    __slots__ = ('orgs', 'route')
    
    def __init__(self, route, start, finish):
        self.orgs = {} 
        self.route = route
        
        ct = start.date()
        while ct <= finish.date():
            curDay = days[ct.weekday()]
            if curDay in route:
                for v in route[curDay].items:
                    if not v.name in self.orgs: self.orgs[v.name] = OrgSaledData()
            
            ct += timedelta(days=1)

    def AddSalesDocs(self, docs, distrGroups):
        for doc in docs:
#             if not IsInRoute(doc, self.route): continue
            if not doc.id in self.orgs: continue
            self.orgs[doc.id].AddSalesDoc(doc, distrGroups) 

    def AddDistribDocs(self, docs, server):
        for doc in docs:
            if not IsInRoute(doc, self.route, server): continue
            if not doc.id in self.orgs: continue
            self.orgs[doc.id].AddDistribDoc(doc) 

class GroupPeriodReportData :
    __slots__ = ('exists', 'oos', 'visits')
    
    def __init__(self):
        self.exists = 0
        self.oos = 0
        self.visits = 0   
        
class GroupSaledData :
    __slots__ = ('saled', 'exists')
    
    def __init__(self):
        self.saled = []
        self.exists = {}

class GroupDailyData :
    __slots__ = ('dailyItems')
    
    def __init__(self):
        self.dailyItems = {}
        
    def AddSaled(self, date, id):
        gsd = None
        if not date in self.dailyItems:
            gsd = GroupSaledData()
            self.dailyItems[date] = gsd
        else :
            gsd = self.dailyItems[date]
        if not id in gsd.saled : gsd.saled.append(id)

    def ExpandSaled(self, saled, src) :
        for val in src:
            if not val in saled : saled.append(val)
    
    def AddExists(self, date, id, exist):
        gsd = None
        if not date in self.dailyItems:
            gsd = GroupSaledData()
            self.dailyItems[date] = gsd
        else :
            gsd = self.dailyItems[date]
        if not id in gsd.exists or exist : 
            gsd.exists[id] = exist
        
    def GetPeriodData(self, oosStart, start, finish, route, debug):
        data = GroupPeriodReportData()
        
        saled = []
        exists = []
        
        dates = list(self.dailyItems.keys())
        dates.sort()
        
        visits = 0
        
        for day in dates :
                            
            dailyData = self.dailyItems[day]
            if day < start :
                if day >= oosStart :
                    self.ExpandSaled(saled, dailyData.saled)
                continue
                        
            if day > finish :
                break            

            if days[day.weekday()] in route:
                visits += len(route[days[day.weekday()]].items)

            for k, v in dailyData.exists.items():
                if not v or k in exists: continue
                exists.append(k)
            self.ExpandSaled(exists, dailyData.saled)
            
            for oid in saled :
                if oid in dailyData.exists and not dailyData.exists[oid] :
                    data.oos += 1
                    
            self.ExpandSaled(saled, dailyData.saled)
        
        data.exists = len(exists)
        data.visits = visits
        
        return data
        
class GroupPeriodData :
    __slots_ = ('items')
    
    def __init__(self, saledData):
        self.items = {}
        for org, v in saledData.orgs.items() :
            for day, data in v.dailyItems.items() :
                for sid in data.sales:
                    gdd = None                     
                    if not sid in self.items :
                        gdd = GroupDailyData()
                        self.items[sid] = gdd;
                    else : 
                        gdd = self.items[sid]
                                            
                    gdd.AddSaled(day, org)
                    
                for eid, dval in data.distrib.items() :
                    gdd = None 
                    if not eid in self.items:
                        gdd = GroupDailyData()
                        self.items[eid] = gdd;
                    else : 
                        gdd = self.items[eid]
                        
                    gdd.AddExists(day, org, dval)
                        
            

def LoadPeriodicData(server, ppStart, prevStart, params, orgs, distrGroups, route):
    
    userid = "'" + params.userid + "'"
    
    where = '"userid" = {0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
         userid,
         prevStart.strftime("%d/%m/%Y 0:0:0"),
         params.finish.strftime("%d/%m/%Y 23:59:59"))
    
    ddocs = server.Get("DistribGroupDoc", where)
    
    where = '"userid" = {0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
         userid,
         ppStart.strftime("%d/%m/%Y 0:0:0"),
         params.finish.strftime("%d/%m/%Y 23:59:59"))
    odocs = server.Get("Order", where)
    
    salesData = SaledData(route, params.start, params.finish)
    salesData.AddDistribDocs(ddocs, server)
    salesData.AddSalesDocs(odocs, distrGroups)

    return salesData

class XLPeriodOrg(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        if column >= 1:
            cell.style.alignment.text_rotation = 90
        return column

class XLPeriodGroup(XLPeriodOrg):
    def makeCell(self, sheet, row, column, cell, value):
        XLPeriodOrg.makeCell(self, sheet, row, column, cell, value)
        if column >= 2:
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)

def MakePeriodOrgReport(sheet, salesData, ppStart, prevStart, start, finish, orgs):
    xlb = XLPeriodOrg()
    
    row = 1
    cell = sheet.cell(row=row, column=0)
    cell.value = 'Период с {0} по {1}'.format(start.strftime("%d.%m.%Y"), finish.strftime("%d.%m.%Y"))
    
    row += 1
    xlb.makeHead(sheet, row, ['', 'Количество визитов за период', 'Среднее SKU', 'Среднее OOS', 'Отклонение SKU', 'Отклонение OOS'], True)
    row += 1
    
    for k, v in salesData.orgs.items() :
        orgName = orgs[k].name if k in orgs else 'Контрагент с кодом <' + k + '>'
        data = [orgName]
        
        curData = v.GetPeriodData(prevStart.date(), start.date(), finish.date())
        prevData = v.GetPeriodData(ppStart.date(), prevStart.date(), (start + timedelta(days=-1)).date())
        
        prevSKU = prevData.AvgSKU()
        prevOOS = prevData.AvgOOS()
        curSKU = curData.AvgSKU()
        curOOS = curData.AvgOOS()
        
        data += [curData.visits, 
                 '0' if curData.visits == 0 else '=' + str(curData.sku) + '/' + str(curData.visits), 
                 '0' if curData.visits == 0 else '=' + str(curData.oos) + '/' + str(curData.visits), 
                 (0 if prevSKU == 0 else curSKU / prevSKU), 
                 (0 if prevOOS == 0 else curOOS / prevOOS) ]
        
        xlb.makeCells(sheet, row, data)
        row += 1
        
    
    cc = 1
    for w in [70]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    

def MakePeriodGroupReport(sheet, salesData, ppStart, prevStart, start, finish, groups, orgCount, route):
    xlb = XLPeriodGroup()
    
    row = 1
    cell = sheet.cell(row=row, column=0)
    cell.value = 'Период с {0} по {1}'.format(start.strftime("%d.%m.%Y"), finish.strftime("%d.%m.%Y"))
    
    row += 1
    xlb.makeHead(sheet, row, ['Топ Ассортимент', 'Охват', 'Процент охвата КБ', 'Изменение процента охвата КБ', '% OOS', 'Изменение % OOS'], True)
    row += 1
    
    for v in groups :
        data = [v.name]
        
        if not v.id in salesData.items : continue
        
        debug = v.id == '112B12DD9F73449786C28B6169DDA6FE'
        v = salesData.items[v.id]

        curData = v.GetPeriodData(prevStart.date(), start.date(), finish.date(), route, debug)
        prevData = v.GetPeriodData(ppStart.date(), prevStart.date(), (start + timedelta(days=-1)).date(), route, False)
        
        curPC = curData.exists / float(orgCount)
        prevPC = prevData.exists / float(orgCount)
        
#         print(v.name + ' oos ' + str(curData.oos))
        
        clmnName = get_column_letter(2)

        curOOS = 0 if curData.visits == 0 else curData.oos / float(curData.visits)
        prevOOS = 0 if prevData.visits == 0 else prevData.oos / float(prevData.visits)
        data += [curData.exists, 
                 '=' + clmnName + str(row+1) + '/' + str(orgCount) + '*100 ', # curPC * 100, 
                 0 if prevPC == 0 else curPC / prevPC * 100 - 100, 
                 '0' if curData.oos == 0 else '=' + str(curData.oos) + '/' + str(curData.visits) + '* 100',
                 (0 if prevOOS == 0 else curOOS / prevOOS * 100 - 100) ]
        
        xlb.makeCells(sheet, row, data)
        row += 1
        
    
    cc = 1
    for w in [70]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

def run(server):
    print( "distrib_rport start")
    params = server.Params[0]
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Дневной"
    
    orgs, distrGroups, route, ddata = LoadDailyData(server, params)
    MakeDailyReport(sheet, orgs, distrGroups, ddata)
    
    if params.periodicReports > 0 :
        prevStart = GetPrevIntervalStart(params.start, params.finish)
        ppStart = GetPrevIntervalStart(prevStart, params.start)
        salesData = LoadPeriodicData(server, ppStart, prevStart, params, orgs, distrGroups, route)
        orgsCount = len(salesData.orgs)
        
        sheet = wb.create_sheet(None, 'Периодический по ТТ')
        MakePeriodOrgReport(sheet, salesData, ppStart, prevStart, params.start, params.finish, orgs)
        
        gpd = GroupPeriodData(salesData)
        sheet = wb.create_sheet(None, 'Периодический по SKU')
        MakePeriodGroupReport(sheet, gpd, ppStart, prevStart, params.start, params.finish, distrGroups, orgsCount, route)

#     wb = printOut(data, params)
# 
    XLBuilder().workbookToObject(wb, "dstr.xlsx", server)                
    print("distrib_rport end")
