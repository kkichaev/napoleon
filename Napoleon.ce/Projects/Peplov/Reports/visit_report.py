# -*- coding: cp1251 -*-
import logging
from importlib import reload

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import time
import datetime

import sys;
reload(sys);
#sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
    __slots__ = ['id', 'name', 'plan', 'fact']
    
    def __init__(self):
        self.id = ''
        self.name = ''
        self.plan = ''
        self.fact = ''
        
    def getData(self, row):
      return [self.name, self.plan, self.fact]
      
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()

def loadData(params, server):
    data = ReportData() 
    agents = server.Get("Agents", "", "id")
    sw = server.Get("StartWork", "")
    
    swhash = {}
    
    for s in sw:
      if not s.userid in swhash:
        swhash[s.userid] = [None] * 7
      
      swhash[s.userid][int(s.day)] = s.time
    
    date = params.date;
    dayidx = date.weekday()
    
    for aid in params.userids.split(','):
      if not aid in swhash or swhash[aid][dayidx] == None or len(swhash[aid][dayidx].strip()) == 0:
        continue
      
      where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
          aid,
          date.strftime("%d/%m/%Y 0:0:0"),
          date.strftime("%d/%m/%Y 23:59:59"))
        
      docNames = ["Order", "VisitInfo", "OrgRemnants", "Answer", "Incass", "TaskDone"]
      userdocs = []
      
      for name in docNames:
        docList = server.Get(name, where)
        
        if docList == None:
          continue
          
        if docList != None:
          for d in docList:
            userdocs.append(d)
      
      userdocs = sorted(userdocs, key=lambda lhs: lhs.created)
      
      t = time.strptime(swhash[aid][dayidx], "%H:%M")
      tm = t.tm_hour * 60 + t.tm_min
      
      if len(userdocs) > 0:
        td = userdocs[0].created.timetuple()
        tdm = td.tm_hour * 60 + td.tm_min
        
        if tdm > tm:
          item = Item()
          item.id = aid
          
          if aid in agents:
            item.name = agents[aid].name
          
          item.plan = swhash[aid][dayidx]
          item.fact = userdocs[0].created.strftime("%H:%M")
          
          data.data.append(item)
      
     
    data.data = sorted(data.data, key=lambda lhs: lhs.name)
    
    return data
    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'
  RED    = 'FFFF0000'
  GREEN  = 'FF00FF00'
  YELLOW = 'FFFFFF00'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Нарушение регламента работы"
    
    c = sheet.cell(row=1,column=0)
    c.value = "Дата: {0}".format(params.date.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 2
    
    head = ['Агент', 'Время план', 'Время факт']
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      xlb.makeCells(sheet, row, item.getData(row))
      row += 1
        
        
    cc = 1
    for w in [50,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)
    
    XLBuilder().workbookToObject(wb, "visit_report.xlsx", server)                
      
    logging.info('end')
    