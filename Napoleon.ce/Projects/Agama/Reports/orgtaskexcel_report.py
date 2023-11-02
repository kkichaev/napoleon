# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
from datetime import datetime
import tempfile
import io

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")


def setBorder(cell):
   cell.style.borders.left.border_style = Border.BORDER_THIN
   cell.style.borders.right.border_style = Border.BORDER_THIN
   cell.style.borders.top.border_style = Border.BORDER_THIN
   cell.style.borders.bottom.border_style = Border.BORDER_THIN

def header(sheet):
    row = 0
    cell = sheet.cell(row=row, column=0) 
    cell.value = "Назначена"
    cell.style.font.bold = True
    setBorder(cell)
    cell =  sheet.cell(row=row, column=1)
    cell.value = "Выполнена"
    cell.style.font.bold = True
    setBorder(cell)
    cell = sheet.cell(row=row, column=2)
    cell.value = "Реалклиент"
    cell.style.font.bold = True
    setBorder(cell)
    cell = sheet.cell(row=row, column=3)
    cell.value = "Адрес ТТ"
    cell.style.font.bold = True
    setBorder(cell)
    cell = sheet.cell(row=row, column=4)
    cell.value = "Формат ТТ"
    cell.style.font.bold = True
    setBorder(cell)
    cell = sheet.cell(row=row, column=5)
    cell.value = "Текст Задачи"
    cell.style.font.bold = True
    setBorder(cell)
    cell = sheet.cell(row=row, column=6)
    cell.value = "Выполнение"
    cell.style.font.bold = True
    setBorder(cell)
    cell = sheet.cell(row=row, column=7)
    cell.value = "Комментарий"
    cell.style.font.bold = True
    setBorder(cell)
    
def scope(val):
    return "'" + val + "'"
    
def agentData(server, sheet, id, agents, param, row, range):
    agent = agents[id]
    sheet.cell(row=row, column=0).value = agent.name
    sheet.cell(row=row, column=0).style.font.bold = True
    sheet.cell(row=row, column=0).style.font.size = 16
    row = row + 1
    
    server.ChangeUser(id)
    agent = server.CurrentUser().name
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    server.RestoreUser()
    
    sheet.cell(row=row, column=0).value = range
    
    header(sheet)
    
    start = param.start.strftime('%d/%m/%Y')
    finish = param.finish.strftime('%d/%m/%Y')
    
    sql = '"userid"=' +\
        scope(id) + ' and "start" >= ToDate(' + scope(start) + ') and "start" <= ToDate(' + scope(finish) +')'
    task = server.Get("OrgTask", sql)   

    taskids = "";
    
    for ot in task:
        if len(taskids) > 0:
            taskids += ","
        taskids += scope(ot.id)    
    
    sql = '"userid"=' + scope(id) + ' and "created" >= ToDate(' + scope(start) + ') and "created" <= ToDate(' + scope(finish) +')'
    taskdone = server.Get("TaskDone", sql)
    
    taskCreated = dict()
    taskText = dict()
    
    taskdonecount = 0
    
    for td in taskdone:
        taskCreated[td.idTask] = td.created
        taskText[td.idTask] = td.text
    
    row = row + 1
    
    for t in task:
        cell = sheet.cell(row=row, column=0)
        cell.value = t.start.strftime('%d.%m.%Y')
        setBorder(cell)  
        
        setBorder(sheet.cell(row=row, column=1))
        setBorder(sheet.cell(row=row, column=6))
        setBorder(sheet.cell(row=row, column=7))
        
        if t.id in taskCreated and t.id in taskText:
            taskdonecount = taskdonecount + 1
            cell = sheet.cell(row=row, column=1)
            cell.value = taskCreated[t.id].strftime('%d.%m.%Y')
            cell = sheet.cell(row=row, column=6)
            cell.value = "1"
            cell = sheet.cell(row=row, column=7)
            cell.value = taskText[t.id]
            
        if t.orgid in orgs:    
            o = orgs[t.orgid]
            cell = sheet.cell(row=row, column=2)
            cell.value = o.realClient
            setBorder(cell)  
            cell = sheet.cell(row=row, column=3)
            cell.value = o.name
            setBorder(cell)
            cell = sheet.cell(row=row, column=4)
            cell.value = o.filter2
            setBorder(cell)    
        
        cell = sheet.cell(row=row, column=5)
        cell.value = t.text
        setBorder(cell)
         
        row = row + 1    
     
    row = row + 1    
    sheet.cell(row=row, column=0).value = "Всего задач"
    sheet.cell(row=row, column=1).value = len(task)
    row = row + 1
    sheet.cell(row=row, column=0).value = "Выполнено"
    sheet.cell(row=row, column=1).value = taskdonecount
    row = row + 1
    sheet.cell(row=row, column=0).value = "% выполнения"
    
    value = 0
    
    if len(task) > 0: 
        value = int(float(taskdonecount) / len(task) * 100)
        
    sheet.cell(row=row, column=1).value = value
                
    row = row + 2

def doReport(server, params, outObj):
    param = server.Params[0];
    range = " Период " + param.start.strftime('%d.%m.%Y') + " - " + param.finish.strftime('%d.%m.%Y');
    agents = server.Get("Agents", "", "id")
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    row = 1  
    for item in param.items:
        agentData(server, sheet, item.id, agents, param, row, range)
        row += 2
            
    sheet.column_dimensions[get_column_letter(1)].width = 15
    sheet.column_dimensions[get_column_letter(2)].width = 15
    sheet.column_dimensions[get_column_letter(3)].width = 15
    sheet.column_dimensions[get_column_letter(4)].width = 15
    sheet.column_dimensions[get_column_letter(5)].width = 15
    sheet.column_dimensions[get_column_letter(6)].width = 30
    sheet.column_dimensions[get_column_letter(7)].width = 15
    sheet.column_dimensions[get_column_letter(8)].width = 30

    repName = "orgtaskexcel.xlsx" 
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)
    
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
    
    obj = outObj.New()
    obj.name = repName
    obj.file = bytes

def run(server):
    print "start orgtask_report "  + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    doReport(server, server.Params[0], outObj)
    server.Put(outObj)
   
    print "done orgtask_report" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
