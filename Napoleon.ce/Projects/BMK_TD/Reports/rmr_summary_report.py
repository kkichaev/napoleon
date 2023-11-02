# -*- coding: cp1251 -*-

from importlib import reload
import sys
import logging

from manager import summary
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from rmr_report_style import XLBuilderCommon
from datetime import timedelta
from openpyxl.style import Border

class Report:
  __slots__ = ['name', 'items']
  
  def __init__(self):
    self.name = ''
    self.items = []

class Item:
  __slots__ = ['id', 'name', 'visit', 'order', 'sum', 'dist', 'progress', 'order_progress', 'order_clients']
  
  def __init__(self):
    self.id = ''
    self.name = ''
    self.visit = 0
    self.order = 0
    self.sum = 0.0
    self.dist = 0.0
    self.progress = 0.0
    self.order_progress = 0.0
    self.order_clients = 0
  
  def __str__(self):
    return 'id: {0}, visit: {1}, order: {2}, sum: {3}, dist: {4}, progress: {5}, order_progress: {6}'.format(self.id, self.visit, self.order, self.sum, self.dist, self.progress, self.order_progress)
    
  def getData(self, row):
    return [self.name, self.visit, self.order, self.order_clients, self.sum, "=IFERROR(D{0}/B{0},0)".format(row), self.progress]

class DivItem(Item):
  __slots__ = ['division', 'childs', 'agents']
  
  def __init__(self, div):
    Item.__init__(self)
    self.division = div
    self.childs = []
    self.agents = []
    self.name = div.name
    
  def updateAgents(self, all):
    for id in self.division.agents:
      if id in all:
        self.agents.append(all[id])
        
    self.agents = sorted(self.agents, key=lambda x: x.name)

  def calcData(self):
    for c in self.childs:
      c.calcData()
      
    for c in self.childs:
      self.incVals(c)
      
    for a in self.agents:
      self.incVals(a)
      
    sz = len(self.childs) + len(self.agents)

    if sz != 0:
      self.order_progress /= sz
      self.progress /= sz
  
  def incVals(self, v):
    self.visit += v.visit
    self.order += v.order
    self.sum += v.sum
    self.dist += v.dist
    self.order_progress += v.order_progress
    self.progress += v.progress
    self.order_clients += v.order_clients

def unpack(list):
  res = ''
  
  for i in list:
    if len(res) > 0:
      res += ','
      
    res += "'{}'".format(i.id)

  return res
    
def loadData(params, server):
  divisions = list()
  rootDivision = server.Get('Division', '"id"=' + str(params.divid))
  agents = server.Get('Agents', '', 'id')
  
  divAgents = summary.loadAgents(server, rootDivision, divisions)

  usersOrder = dict()
  where = '"{0}" >= ToDate("{1} 0:0:0") and "{0}" < ToDate("{2} 0:0:0") and "userid" in ({3})'.format(
    "created",
    params.start.strftime('%d.%m.%Y'), 
    (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'),
    unpack(params.userids))
  
  orders = server.Get('Order', where)

  for o in orders:
    if not o.userid in usersOrder:
      usersOrder[o.userid] = list()

    clients = usersOrder[o.userid]

    if not o.id in clients:
      clients.append(o.id)  

  curDate = params.start
  endDate = params.finish + timedelta(days=1)
  
  map = dict()
  
  while curDate < endDate: 
    #print("curDate", curDate)
    dayRep = summary.dailyReport(server, divAgents, curDate)
    
    for key, value in dayRep.items():
      if not key in map:
        map[key] = Item()
        map[key].id = key
      
      item = map[key]
      item.name = agents[key].name if key in agents else key
      item.visit += len(value.visited)
      item.order += value.orders
      item.sum += value.summa
      item.dist += summary.computeDist(server,curDate,key) / 1000
      
      if len(value.plan) > 0:
        item.progress += value.plannedVisits / float(len(value.plan))
        item.order_progress += item.order / item.visit if item.visit != 0 else 0

    curDate = curDate + timedelta(1)

  for k,v in map.items():
    d = (params.finish - params.start).days

    if d > 0:
      v.progress /= d
      v.order_progress /= d
      
  r = Report()
  dm = dict()
  root = None

  for k in map:
    if k in usersOrder:
      map[k].order_clients = len(usersOrder[k])
  
  for d in divisions:
    if root == None:
      root = DivItem(d)
      root.updateAgents(map)
      dm[d.id] = root
  
    if d.parent in dm:
      p = dm[d.parent]
      dc = DivItem(d)
      dc.updateAgents(map)
      p.childs.append(dc)
      dm[d.id] = dc
      
  if root != None:
    r.name = root.name
    
  root.calcData()  
  putItems(root, r.items)
  
  return r

def putItems(div, out):
  out.append(div)
  
  for a in div.agents:
    out.append(a)
    
  for c in div.childs:
    putItems(c, out)
    
class XLBuilderEx(XLBuilderCommon):
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilderCommon.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 5 or column == 6:
      cell.style.number_format._set_format_code('0%')
    
def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  xlb = XLBuilderEx()
  
  cell = sheet.cell(row=0, column=0)
  cell.value = 'Итоговый отчёт подразделения: {0}'.format(data.name)
  cell.style.font.bold = True
  
  DATE_FMT = "%d.%m.%Y"
  cell = sheet.cell(row=2, column=0)
  cell.value = "Период: {0} - {1}".format(params.start.strftime(DATE_FMT), params.finish.strftime(DATE_FMT))
  
  row = 3
  
  head = ['Подразделение / агент', 'визиты', 'заявки итого', 'Заявки 1 к 1', 'сумма', 'процент заявок', 'прогресс']
  xlb.makeHead(sheet, row, head, True)
  
  row += 1
  
  for i in data.items:
    xlb.makeCells(sheet, row, i.getData(row + 1))
    row += 1

  cc = 1
  for w in [45,20,20,20,20,20]:
    sheet.column_dimensions[get_column_letter(cc)].width = w
    cc += 1
    
  return wb

def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilderCommon().workbookToObject(wb, "rmr_summary_report.xlsx", server)                
  logging.info('end')