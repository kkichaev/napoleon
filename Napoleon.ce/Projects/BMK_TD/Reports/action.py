# -*- coding: cp1251 -*-
from importlib import reload
import logging
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from datetime import datetime
from datetime import timedelta

import sys
reload(sys)
#sys.setdefaultencoding("cp1251")

class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = dict()
        
class AgentPage:
    start = None
    finish = None
    items = None
    agent = None
    division = None
    id = None
    name = None
    
    def __init__(self, agent):
      self.id = agent[0]
      self.name = agent[1]
      self.start = datetime.now()
      self.finish = datetime.now()
      self.items = list()
      self.agent = ""
      self.division = ""
        
class Item:
    org = None
    vizit = None
    order = None
    
    def __init__(self):
      self.org = ""
      self.vizit = list()
      self.order = 0
      
    def getData(self, row):
      return [self.org, len(self.vizit), self.order , "=IFERROR(C{0}/B{0}, 0)".format(row + 1)]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish

def agentName(id, arr):
    name = id
    
    if id in arr:
        name = arr[id].name
        
    return id, name
    
def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    result = ReportData()
    
    for a in params.userids:
      userid = a.id
        
      if not userid in result.data:
          result.data[userid] = AgentPage(agentName(userid, agents))
          
      ap = result.data[userid]
      
      start = params.start
      finish = params.finish + timedelta(days=1)
      
      WHERE_STR = '"created" >= ToDate("{0}") and "created" < ToDate("{1}") and "userid"=\'{2}\'';         
      where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid)
      
      agent = ""
      server.ChangeUser(userid)
      agent = server.CurrentUser().name
      orgs = server.Get("Org", "", "id")
      porg = server.Get("PotenzialOrg", "", "id")
      orgs.update(porg)
      server.RestoreUser()

      data = dict()
      
      for name in ['Order', 'VisitInfo', 'OrgRemnants', 'Question', 'Incass', 'ScriptDoc', 'TaskDone','Layout']:
          doc = server.Get(name, where)
          print(name, where)     
          for d in doc:
              print(name)
              if len(d.id.strip()) == 0:
                  continue
              
              if not d.id in data:
                  f = Item()
                  data[d.id] = f 
                  f.org = "{0} / {1}".format(orgs[d.id].name, orgs[d.id].address) if d.id in orgs else d.id
                  
              r = data[d.id]
              dt = d.created.date()
              
              if not dt in r.vizit:
                  r.vizit.append(dt)
                  
              if name == 'Order':
                  r.order += 1       
      
      for oid in orgs:
          if not oid in data:   
              f = Item()
              data[oid] = f
              f.org =  "{0} / {1}".format(orgs[oid].name, orgs[oid].address)
              
      div = server.Get("Division","")
      
      dname = ""
      for d in div:
          for da in d.agents:
              if da.id == userid:
                  dname = d.name
                  break;             
      
      items = sorted(data.values(), key=lambda lhs: lhs.org)
      
      ap.start = start
      ap.finish = finish - timedelta(days=1)
      ap.agent = agent
      ap.items = items
      ap.division = dname
    
    return result
    
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    sh.cell(row=0, column=0).value = '{0} - {1} / {2} / {3}'.format(
        data.start.strftime("%d.%m.%Y"), data.finish.strftime("%d.%m.%Y"), data.agent, data.division)
    
    head = ["Название торговой точки вся база ТП", "Количество посещений", "Количество заявок", "% Заказов"]
    
    r = 1
    xlb.makeHead(sh, r, head)
    
    for d in data.items:
      r += 1
      xlb.makeCells(sh, r, d.getData(r))
    
    setCellWidth(sh, [90,12,12,12])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 3:
      cell.style.number_format._set_format_code('0%')
      
    
def printOut(d, params):
    wb = Workbook(False, 'cp1251')
    xlb = XLBuilderEx()
    sheet = None
    
    for ap in d.data.values():
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()

        sheet.title = ap.name[:31]
        ptintSheet(xlb, sheet, ap)
                
    return wb

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "rmr_visit_report.xlsx", server)                
    logging.info('end')
    