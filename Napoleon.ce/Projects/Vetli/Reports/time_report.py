import calendar
import logging

from datetime import timedelta
from datetime import datetime
from datetime import date
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from openpyxl.style import NumberFormat


import sys;
from manager import *

bkgColor = "ff8FBC8F"

class Item:
    date = None
    agent = None
    time = None
    dist = None
    cost = None
    km = None
    
    def getData(self):
        return [self.date.strftime('%d.%m.%Y'), self.agent, self.time, (int)(self.dist) / 1000.0, self.cost]

def inflateParams(server):
    return int(server.Params[0].month), int(server.Params[0].year)

def getAgentName(id, agents):
    return agents[id].name if id in agents else "Торговый представитель с кодом<{0}>".format(id)

def loadData(server):
    m, y = inflateParams(server)
    w, d = calendar.monthrange(y,m)
  
    data = dict() # dict by week number
    where = '"month"={0} and "year"={1}'.format(m, y)
    timetr = server.Get("TimeTracking", where)
    agents = server.Get("Agents",'','id')
    cost = 0
    
    if len(timetr) > 0:
        tt = timetr[0]
        cost = tt.cost
        for i in tt.items:
            wn = i.date.isocalendar()[1]
            
            if not wn in data:
                data[wn] = dict()
            ds = i.date.strftime("%d/%m/%Y") 
            k = ds + i.userid
            
            t = Item()
            t.date = i.date
            t.agent = getAgentName(i.userid,agents)
            t.time = '{0}-{1}'.format(i.start, i.finish)
            
            # calc distance for each items
            t.dist = calcDist(server, ds + ' ' + i.start, ds + ' ' + i.finish, i.userid)
            t.cost = tt.cost
            t.km = i.km
            
            if not t.date in data[wn]:
                data[wn][t.date] = list()
                
            data[wn][t.date].append(t)
            
    return data, cost
    
def calcDist(server, start, finish, userid):
  where = '"date" > ToDate("{0}:00") and "date" < ToDate("{1}:00") and "isGSM" = \'0\' and "userid" = "{2}"'.format(start,  finish, userid)
  gpspos = server.Get("GPSPos", where)    
  res = 0
  lp = None
  
  for g in gpspos:
    if lp != None: 
      res += coordutils.distance(lp.latitude, lp.longitude, g.latitude, g.longitude);
      
    lp = g  
  
  return res
  
class XLBuilderEx(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment.wrap_text = True
        self.setBackColor(cell, bkgColor)
        return column    
    
    def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        if column == 1:
            cell.style.alignment.horizontal = Alignment.HORIZONTAL_LEFT
        else:    
            cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
        
        if column == 5:
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER)

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def drawData(sh, xlb, data, row, agrw):
    sh.merge_cells(start_row=row, start_column = 0, end_row =row + len(data), end_column = 0);
    sh.merge_cells(start_row=row, start_column = 4, end_row =row + len(data), end_column = 4);
    
    for d in data:
        dd = d.getData()
        dd.append('={0}{1}/10*{2}{1}'.format(get_column_letter(4),row+1,get_column_letter(5)))
        dd.append(d.km)
        xlb.makeCells(sh, row, dd)
        if not d.agent in agrw:
            agrw[d.agent] = list()
            
        agrw[d.agent].append(row+1)    
        row += 1
        
    ikm = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(4),row-len(data)+1, row)
    ism = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(6),row-len(data)+1, row)
    xlb.makeCells(sh, row, ['','Итого','',ikm,'',ism])
    
    return row
        
def paintHeadCell(xlb, cell):
    cell.style.alignment.wrap_text = True
    xlb.setBackColor(cell, bkgColor)
    xlb.paintHeadCell(cell)
    
def makeHead(xlb, sheet, row):
    cell = sheet.cell(row=row, column=0)
    paintHeadCell(xlb, cell)
    cell.value = "Дата"
    
    cell = sheet.cell(row=row, column=1)
    paintHeadCell(xlb, cell)
    cell.value = "ТП"
    
    cell = sheet.cell(row=row, column=2)
    paintHeadCell(xlb, cell)
    cell.value = "Время"
    
    cell = sheet.cell(row=row, column=3)
    paintHeadCell(xlb, cell)
    cell.value = "КМ"
    
    cell = sheet.cell(row=row, column=4)
    paintHeadCell(xlb, cell)
    cell.value = "Стоимость ГСМ"
    
    cell = sheet.cell(row=row, column=5)
    paintHeadCell(xlb, cell)
    cell.value = "К выдаче руб"
                    
def printData(xlb, sh, data, cost):
    weeks = data.keys()
    weeks = sorted(weeks)
    month = dict()
    
    row = 1
    wkrg = list()
    for w in weeks:
        makeHead(xlb, sh, row)
        row += 1
        dts = data[w].keys()
        dts = sorted(dts)
        
        agrw = dict()
        for d in dts:
            ag = sorted(data[w][d], key = lambda x: x.agent)
            row = drawData(sh, xlb, ag, row, agrw) 
            row += 1
        
        row = drawResultWeek(sh, xlb, row, agrw, month, cost)
        wkrg.append(row)
    
    row = drawResultMonth(sh, xlb, row, month, cost)
    sh.merge_cells(start_row=row, start_column = 0, end_row =row, end_column = 2);
    xlb.makeCells(sh,row,['Итого за месяц','','',funSum(get_column_letter(4), wkrg),'',funSum(get_column_letter(6), wkrg)])    
    
    setCellWidth(sh, [15,26,26,18,18,18,0])
    sh.column_dimensions['G'].visible = False

def drawResultMonth(sh, xlb, row, agrw, cost):
    sh.merge_cells(start_row=row, start_column = 0, end_row =row + len(agrw)-1, end_column = 0);
    sh.merge_cells(start_row=row, start_column = 4, end_row =row + len(agrw)-1, end_column = 4);
    
    ags = sorted(agrw.keys())
    rws = list()
    
    for a in ags:
        xlb.makeCells(sh, row, ['Месяц',a,funSum(get_column_letter(3), agrw[a]), funSum(get_column_letter(4), agrw[a]),
          cost, funcDiv(get_column_letter(4), get_column_letter(5), row+1 , 10)])
          
        rws.append(row+1)
        row += 1
    
    return row

def drawResultWeek(sh, xlb, row, agrw, m, cost):
    sh.merge_cells(start_row=row, start_column = 0, end_row =row + len(agrw)-1, end_column = 0);
    sh.merge_cells(start_row=row, start_column = 4, end_row =row + len(agrw)-1, end_column = 4);
    
    ags = sorted(agrw.keys())
    rws = list()
    
    for a in ags:
        xlb.makeCells(sh, row, ['Неделя',a,funSum(get_column_letter(7), agrw[a]), funSum(get_column_letter(4), agrw[a]) + '+{0}{1}'.format(get_column_letter(3), row + 1),
          cost, funcDiv(get_column_letter(4), get_column_letter(5), row+1 , 10)])
          
        if not a in m:
          m[a] = list()
        
        row += 1
        m[a].append(row)
        rws.append(row)
        
    
    sh.merge_cells(start_row=row, start_column = 0, end_row =row, end_column = 2);
    xlb.makeCells(sh,row,['Итого за неделю','','',funSum(get_column_letter(4), rws),'',funSum(get_column_letter(6), rws)])
    row += 1
    
    return row

def funcDiv(col1, col2, row, coef):
    return '={1}{0} / {3} * {2}{0}'.format(row, col1, col2, coef)
    
def funSum(col, row):
    rs = ''
    for r in row:
        if len(rs) > 0:
            rs += '+'
        rs += '{0}{1}'.format(col,r)
        
    res = '=SUM({0})'.format(rs)

    return res    
    
def printOut(data, cost):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    printData(xlb, sh, data, cost)
    
    return wb
         
def doReport(server):
    data, cost = loadData(server)
    wb = printOut(data, cost)
    XLBuilder().workbookToObject(wb, "time.xlsx", server)
          
def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')
  
  doReport(server)

  logging.info('end')
