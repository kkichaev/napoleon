# -*- coding: cp1251 -*-

import sys;
import locale
import time
import datetime
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat
from time import sleep
from openpyxl.writer.worksheet import row_sort

reload(sys);
sys.setdefaultencoding("cp1251")

class RowData:
    __slots__ = ['date', 'type', 'org', 'distr', 'address', 'agent', 'item', 'inpack', 'qty', 'cost', 'sum', 'remark']
    
    def RowData(self):
        self.inpack = 1
    

def loadDocs(docs, orgs, distrbs, price, data, type, agents):    
    if docs == None: return
    
    for doc in docs:
        distr = 'Код ' + doc.distr if not doc.distr in distrbs else distrbs[doc.distr].name
        org = 'Код ' + doc.id
        addr = ''
        if doc.id in orgs:
            o = orgs[doc.id]
            org = o.name
            addr = o.address
            
        for i in doc.items:
            rd = RowData()
            rd.date = doc.date.strftime("%d.%m.%Y")
            rd.type = 'Заказ' if type == 1 else 'Возврат'
            rd.org = org
            rd.address = addr
            rd.distr = distr
            rd.remark = doc.remark
            rd.agent = agents[doc.userid].name if doc.userid in agents else doc.userid
            
            if not i.id in price:
                rd.item = 'Код ' + i.id
                rd.inpack = 1
            else:
                p = price[i.id]
                rd.item = p.name
                rd.inpack = p.qtyInPack
                
            rd.cost = i.cost
            rd.sum = type * i.cost * i.qty
            rd.qty = type * i.qty

            data.append(rd)

def loadData(server, params):
    
    distribs = ""
    for i in params.items:
        distribs += "'" + i.id + "',"        
    distribs = distribs[:-1]
    
    where = '"created" >= ToDate("{0}") and "created" < ToDate("{1}") and "distr" in ({2})'.format(
        params.start.strftime("%d/%m/%Y 00:00:00"),
        params.end.strftime("%d/%m/%Y 00:00:00"),
        distribs)

    price = server.Get('Price', '', 'id')
    orgs = server.Get('CommonOrgs', '', 'id')
    distr = server.Get('Distributors', '', 'id')

    ords = server.Get("Order", where)
    rets = server.Get("Returns", where)
    agents = server.Get("Agents", "", "id")

    data = list()
    loadDocs(ords, orgs, distr, price, data, 1, agents)
    loadDocs(rets, orgs, distr, price, data, -1, agents)
    
    return data


class XLBuilderEx(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
        XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
        if column == 6 or column == 8 or column == 9:
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)      
            
def printOut(data, params):
    xlb = XLBuilderEx()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    row = 1
    heads = ["Дата Документа", "Тип Документа", "Дистрибьютор", 'Контрагент', 'Адрес доставки', 'Торговый представитель', 'Название номенклатуры', 'Количество (в уп.)', 
             'Количество  (в шт.)', 'Цена с НДС, руб', 'Сумма с НДС', 'Комментарий'] 
    xlb.makeHead(sheet, row, heads, True)
    row += 1
    
    data.sort(key=lambda x: x.date + '|' + x.distr + '|' + x.org + '|' + x.address)
    
    for d in data:
        vals = [d.date, d.type, d.distr, d.org, d.address, d.agent, d.item, d.qty, d.inpack, d.cost, d.sum, d.remark]
        xlb.makeCells(sheet, row, vals)
        row += 1

    cc = 1
    wdh = [11,11,15,25,25,25,40,8,10,8,8,40]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    return wb
    

def run(server):

    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting")
    
    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(server, params)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info("ended")