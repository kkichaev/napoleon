# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN

class ItemCell:
    face = None
    qty = None
    
    def __init__(self):
        self.face = 0.0
        self.qty = 0.0

class PageItem:
    city = None
    user = None
    slsnet = None
    org = None
    data = None
    created = None
    partshelf = 0
    fact = None
    visit = None
    planogram = 0
    date = None
    address = None
    facemy = 0
    faceall = 0
    faceplan = 0
    facefact = 0
    
    def __init__(self):
        self.data = dict()
        self.visit = list()
        
    def process(self, c, report):
        self.city = report.getCity(c.id)
        self.user = report.getUser(c.userid)
        self.slsnet = report.getSls(c.id)
        org = report.getOrg(c.id)
        self.org = org.name
        self.address = org.address
        self.created = c.created.strftime("%d.%m.%Y")
        self.post_process(c, report)
        self.date = datetime(c.created.year, c.created.month, c.created.day)
        
    def post_process(self, c, report):
        ps = report.getPartShelf(c.id)
        if ps == 0: return

        self.partshelf = ps.part
        self.faceplan = report.getFacePlan(c.id);
        self.planogram = report.getPlanogram(c)
        
        other = 0
        my = 0
        all = 0
        
        for i in c.items:
            if i.id in report.mprice:
                self.data[i.id] = i.qty
                
                if report.mprice[i.id].my > 0:
                    my = my + i.qty
                else:
                    other = other + i.qty
                     
        all = my + other
        onepercent  = all / 100
        f = my / onepercent if onepercent != 0 else 0
        self.fact = "{:0.0f}".format(f)
        self.faceall = all
        self.facemy = my
        
        p = self.facemy / self.faceplan * 100 if self.faceplan != 0 else 0
        self.facefact = "{:0.0f}".format(p)
        report.getVisit(c, self.visit)
                   
class MPageItem(PageItem):
    def post_process(self, c, report):
       for i in c.items:
            if i.id in report.mprice:
                self.data[i.id] = i.cost
                     
class Page:
    report = None
    items = None;
    title = None
    width = 0
    
    def __init__(self, report):
        self.report = report
        self.items = list()
        self.title = "Отчет по доли полки"
        self.width = 13
        
    def collumns(self, sheet, r, c):
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Фейсы Сальников", Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Итого фейсы", Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "План фейс", Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Факт фейс", Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "План по доли полки", Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Факт по доли полки", Alignment.VERTICAL_BOTTOM, rotation=90)
        col_letter = get_column_letter(c + 1) 
        sheet.column_dimensions[col_letter].width = 5    
        
        c = c + 1
        sheet.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)
        setVal(sheet.cell(row=r, column=c), "Отметка о выполнении планогаммы", Alignment.VERTICAL_BOTTOM, rotation=90)    
        
    def itemsdata(self, sheet, r, c, i):
        setVal(sheet.cell(row=r, column=c), i.facemy)
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.faceall)
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.faceplan)
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.facefact)
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.partshelf)
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.fact)      
        c = c + 1
        setVal(sheet.cell(row=r, column=c), i.planogram if i.planogram > 0 else "")      
        c = c + 1
        
        for v in i.visit:
            idx = 0
            
            if v.items != None:
                for vi in v.items:
                    cell = sheet.cell(row=r, column=c) 
                    cell.hyperlink = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                    cell.value = "Фото{0}".format(idx + 1)
                    idx = idx + 1
                    c = c + 1
                
    def process(self, report):
        for c in report.contract:
            pi = PageItem()
            pi.process(c, report)
            self.items.append(pi)   
            
        self.items = sorted(self.items, cmp=lambda lhs, rhs: cmp(lhs.created, rhs.created))                  
        
class MonitoringPage(Page):
    def __init__(self, report):
        Page.__init__(self, report)
        self.title = "Ценовой мониторинг"
        self.width = 6
        
    def collumns(self, sheet, r, c):
        pass
    
    def itemsdata(self, sheet, r, c, i):
        pass
    
    def process(self, report):
        for m in report.monitoring:
            pi = MPageItem()
            pi.process(m, report)
            self.items.append(pi)
                         
def price_cmp(lhs, rhs):
        result = (int)(lhs.my - rhs.my)
        
        if result == 0:
            result = cmp(lhs.group, rhs.group)
        
        if result == 0:
            result = cmp(lhs.name, rhs.name)
                
        return result
             
class Report:
    pages = None
    
    """ Список прайс """
    price = None
    """ Позиции в списке по ID """
    pidx = None
    
    slsnet = None
    city = None
    org = None
    agents = None
    partshelf = None
    mprice = None
    visit = None
    planogram = None
    contract = None
    btlplan = None
    
    visitdata = None
    planogramdata = None
    monitoringdata = None
    
    def __init__(self, server):
        start = server.Params[0].start
        finish = server.Params[0].finish
        cid = server.Params[0].cid
        usePhoto = server.Params[0].photo
        
        #self.weeks = Weeks(start, finish)
        
        conDef = server.Get("ContractDef", '"id"=' + "'" + cid + "'")
        
        pids = ""
        
        for cd in conDef:
            for i in cd.items:
                if len(pids) > 0:
                    pids += ","
                    
                pids += "'" + i.id +"'"
        
        self.mprice = server.Get("ManagerPrice", '"id" in (' + pids + ')', "id")
        self.price = list();
        self.price.extend(self.mprice.values())
        self.price = sorted(self.price, cmp=price_cmp)
        
        self.pidx = dict()
        for i in range(0, len(self.price)):
            self.pidx[self.price[i].id] = i    
        
        self.org = server.Get("Org", '"id" is not null', "id")
        self.agents = server.Get("Agents","","id")
        
        endRange = finish + timedelta(days=1) 
        
        where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + endRange.strftime('%d/%m/%Y') + '") and "def"='+"'" + cid + "'"
        self.contract = server.Get("Contract", where)
        
        vobj = "Visit" if usePhoto == 1 else "VisitInfo"
        self.visit = server.Get(vobj, where)
        
        #where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + endRange.strftime('%d/%m/%Y') + '")'
        self.planogram = server.Get("Planogram", where)
        self.monitoring = server.Get("CMonitoring", where)
        self.partshelf = server.Get("PartShelf", '"cid"= ' + "'" + cid + "'", "sid") 
        self.btlplan = server.Get("BtlPlan", "", "id")
        
        self.pages = {MonitoringPage(self), Page(self)}
        
        self.slsnet = server.Get("Slsnet","","id")
        self.city = server.Get("City", "", "id")
        
        for p in self.pages:
            p.process(self)
                    
        
    def getCity(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].cid
        
        return result
    
    def getFacePlan(self, id):
        result = 0.0;
        
        if id in self.btlplan:
            result = self.btlplan[id].face;
        
        #print id, result    
        return result;    
        
    def getOrg(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id]
        
        return result
            
    def getSls(self, id):
        result = id
        
        if id in self.org and self.org[id].sid in self.slsnet:
            result = self.slsnet[self.org[id].sid].name
        
        return result      
    
    def getUser(self, id):
        result = id
        
        if id in self.agents:
            result = self.agents[id].name
            
        return result
      
    def getAddress(self, id):
        result = id
        
        if id in self.org:
            result = self.org[id].address 
            
        return result 
    
    def sameDay(self, d1, d2):
        return d1.day == d2.day and d1.month == d2.month and d1.year == d2.year
     
    def getPartShelf(self, id):
        result = 0
        
        if id in self.org and self.org[id].sid in self.partshelf:
            result = self.partshelf[self.org[id].sid]
            
        return result;
    
    def getVisit(self, c, val):
        if self.visitdata == None and len(self.visit) > 0:
            self.visitdata = dict()
            
            for v in self.visit:
                if not v.id in self.visitdata:
                    self.visitdata[v.id] = dict()
                 
                v1 = self.visitdata[v.id]
                
                if not v.userid in v1:
                    v1[v.userid] = dict()
                
                v2 = v1[v.userid]
                dt = datetime(v.created.year, v.created.month, v.created.day)
                if not dt in v2:
                    v2[dt] = list()
                    
                v2[dt].append(v)         
        
        if self.visitdata != None:
            cdt = datetime(c.created.year, c.created.month, c.created.day)
            
            if c.id in self.visitdata and c.userid in self.visitdata[c.id] and cdt in self.visitdata[c.id][c.userid]:              
                val.extend(self.visitdata[c.id][c.userid][cdt])
    
    def getPlanogram(self, c):
        if self.planogramdata == None:
            self.planogramdata = dict()
            
            for p in self.planogram:
                if not p.id in self.planogramdata:
                    self.planogramdata[p.id] = dict()
                    
                p1 = self.planogramdata[p.id]
                
                if not p.uesrid in p1:
                    p1[p.userid] = dict()
                
                p2 = p1[p.userid]  
                dt = datetime(p.created.year, p.created.month, p.created.day)
                p2[dt] = p
        
        cdt = datetime(c.created.year, c.created.month, c.created.day)
        if c.id in self.planogramdata and c.userid in self.planogramdata[c.id] and cdt in self.planogramdata[c.id][c.userid]:              
            return self.planogramdata[c.id][c.userid][cdt].approved
         
        return 0    
              
def setVal(cell, value, vrt = Alignment.VERTICAL_CENTER, hrz= Alignment.HORIZONTAL_CENTER, wrap=True, bold=False, rotation=0):
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    cell.style.alignment.text_rotation = rotation
    cell.style.font.bold = bold
    cell.value = value
                      
def run(server):
    print "contract start"
    server.RegisterType("Result[name:s,file:b,items[name:s,photo:b]]")
    outObj = server.New("Result")

    report = Report(server)
    wb = Workbook(False, 'cp1251')
    
    sheet = None
    for page in report.pages:
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()
        sheet.title = page.title
        
        i = 1
        r = 0

        sheet.merge_cells(start_row=r, start_column=0, end_row=r+1, end_column=0)    
        setVal(sheet.cell(row=r, column=0), "Город")
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
        setVal(sheet.cell(row=r, column=1), "Мерчендайзер")
        col_letter = get_column_letter(2) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        setVal(sheet.cell(row=r, column=2), "Наименование торговой сети")
        col_letter = get_column_letter(3) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        setVal(sheet.cell(row=r, column=3), "Наименование торговой точки")
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)
        setVal(sheet.cell(row=r, column=4), "Адрес")
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 20
        sheet.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        setVal(sheet.cell(row=r, column=5), "Дата заполнения отчета")
        col_letter = get_column_letter(6) 
        sheet.column_dimensions[col_letter].width = 20

        group = None
        START_COL = 6
        c = START_COL
        pc = c
        isMy = False
        
        for p in report.price:
            if group != None and group != p.group: 
                sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
                setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
                pc = c
                
            group = p.group
            isMy = p.my > 0
            
            setVal(sheet.cell(row=r + 1, column=c), p.name, vrt = Alignment.VERTICAL_BOTTOM, rotation=90)
            col_letter = get_column_letter(c + 1) 
            sheet.column_dimensions[col_letter].width = 5
            
            c = c + 1
            
        sheet.merge_cells(start_row=r, start_column=pc, end_row=r, end_column=c-1)
        setVal(sheet.cell(row=r, column=pc), group, bold=isMy)
        sheet.row_dimensions[2].height = 95
        
        page.collumns(sheet, r, c)
        
        r = r + 2
        dt = None
        sr = r + 1
        dsr = sr
        
        sum_group = dict()
        sum_result = dict()
        last_date = None
        
        for i in page.items:
            if dt == None:
                dt = i.date
                
            if dt != i.date:
                setVal(sheet.cell(row=r, column=0), "Итог дня", hrz= Alignment.HORIZONTAL_LEFT)
                setVal(sheet.cell(row=r, column=5), last_date, hrz= Alignment.HORIZONTAL_RIGHT)
                c = START_COL
                
                for p in report.price:
                    val = sum_group[p.id] if p.id in sum_group else 0
                    setVal(sheet.cell(row=r, column=c), val if val > 0 else "")
                    
                    if not p.id in sum_group:
                        sum_group[p.id] = 0
                    
                    if not p.id in sum_result:
                        sum_result[p.id] = sum_group[p.id]
                    else:
                        sum_result[p.id] = sum_result[p.id] + sum_group[p.id]
                             
                    c = c + 1
                     
                r = r + 1   
                dsr = r
                sum_group = dict()
                dt = i.date
                   
            setVal(sheet.cell(row=r, column=0), i.city, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=1), i.user, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=2), i.slsnet, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=3), i.org, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=4), i.address, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=5), i.created, hrz= Alignment.HORIZONTAL_RIGHT)
            
            c = START_COL;
            last_date = i.created
            
            for p in report.price:
                if p.id in i.data:
                    setVal(sheet.cell(row=r, column=c), i.data[p.id])
                    
                    if not p.id in sum_group:
                        sum_group[p.id] = i.data[p.id]
                    else:     
                        sum_group[p.id] = sum_group[p.id] + i.data[p.id]  
                    
                c = c + 1
                
            page.itemsdata(sheet, r, c, i)
            r = r + 1
        
        setVal(sheet.cell(row=r, column=0), "Итог дня", hrz= Alignment.HORIZONTAL_LEFT)
        setVal(sheet.cell(row=r, column=5), last_date, hrz= Alignment.HORIZONTAL_RIGHT)
        c = START_COL
         
        for p in report.price:
            val = sum_group[p.id] if p.id in sum_group else 0
            setVal(sheet.cell(row=r, column=c), val if val > 0 else "")
            
            if not p.id in sum_group:
                sum_group[p.id] = 0
            
            if not p.id in sum_result:
                sum_result[p.id] = sum_group[p.id]
            else:
                sum_result[p.id] = sum_result[p.id] + sum_group[p.id]
                
            c = c + 1 
         
        setVal(sheet.cell(row=r + 1, column=0), "Итог итого", hrz= Alignment.HORIZONTAL_LEFT)
        c = START_COL
        
        for p in report.price:
            val = sum_result[p.id] if p.id in sum_result else 0
            setVal(sheet.cell(row=r + 1, column=c), val if val > 0 else "")
            c = c + 1 
                   
        rangeBorders(sheet.range("A1:"+get_column_letter(page.width + len(report.price))+str(r + 2)) )
            
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    for v in report.visit:
        idx = 0
        
        if v.items != None:
            for vi in v.items:
                item = obj.items.New()
                item.name = "{0:%d%m%Y%H%M%S}{1}{2}.jpg".format(v.created, v.id, idx)
                item.photo = vi.id
                idx = idx + 1
    
    server.Put(outObj)
    
    print "contract finish"
    