# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN

class PageItem:
    name = None
    data = None
    qty = None
    remark = None
    
    def __init__(self, name, data, qty, remark):
        self.name = name
        self.data = data
        self.qty = qty
        self.remark = remark
                   
class Page:
    report = None
    items = None;
    title = None
    
    def __init__(self, title):
        self.items = list()
        self.title = title
        
    def process(self, report, s):
        for i in s.items:
            self.items.append(PageItem(report.getPrice(i.id), s.created, i.qty, s.remark))                
             
class Report:
    pages = None
    mprice = None
    returns = None
    
    def __init__(self, server):
        start = server.Params[0].start
        finish = server.Params[0].finish
        endRange = finish + timedelta(days=1)
        
        self.mprice = server.Get("ManagerPrice", '', "id")
        self.agents = server.Get("Agents","","id")
        
        where = '"created" >= ToDate("' + start.strftime('%d/%m/%Y') + '") and "created" <= ToDate("' + endRange.strftime('%d/%m/%Y') + '")'
        self.returns = server.Get("Returns", where)
        
        self.pages = dict()
        
        for s in self.returns:
            if s.userid in self.agents:
                if not s.userid in self.pages:
                    self.pages[s.userid] = Page(self.agents[s.userid].name)
                
                self.pages[s.userid].process(self, s)    
    
    def getPrice(self, id):
        result = id
        
        if id in self.mprice:
            result = self.mprice[id].name
        
        return result              
    
def setVal(cell, value, vrt = Alignment.VERTICAL_CENTER, hrz= Alignment.HORIZONTAL_CENTER, wrap=True, bold=False, rotation=0):
    cell.style.alignment.vertical = vrt
    cell.style.alignment.horizontal = hrz
    cell.style.alignment.wrap_text = wrap
    cell.style.alignment.text_rotation = rotation
    cell.style.font.bold = bold
    cell.value = value
                      
def run(server):
    print "return start"
    server.RegisterType("Result[name:s,file:b,items[name:s,photo:b]]")
    outObj = server.New("Result")

    report = Report(server)
    wb = Workbook(False, 'cp1251')
    
    plist = report.pages.values()
    plist = sorted(plist, cmp=lambda lhs, rhs: cmp(lhs.title, rhs.title))

    sheet = None
    idx = 1
    for page in plist:
        if sheet == None:
            sheet = wb.get_active_sheet()
        else:
            sheet = wb.create_sheet()
        
        sheet.title = page.title
          
        r = 0
  
        setVal(sheet.cell(row=r, column=0), "Поз")
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 6
        setVal(sheet.cell(row=r, column=1), "Наименование")
        col_letter = get_column_letter(2) 
        sheet.column_dimensions[col_letter].width = 30
        setVal(sheet.cell(row=r, column=2), "Дата")
        col_letter = get_column_letter(3) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r, column=3), "Кол-во    ")
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 20
        setVal(sheet.cell(row=r, column=4), "Комментарий")
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 50
        
        r = r + 1
        
        for i in page.items:
            setVal(sheet.cell(row=r, column=0), idx)
            setVal(sheet.cell(row=r, column=1), i.name, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=2), i.data, hrz= Alignment.HORIZONTAL_LEFT)
            setVal(sheet.cell(row=r, column=3), i.qty, hrz= Alignment.HORIZONTAL_RIGHT)
            setVal(sheet.cell(row=r, column=4), i.remark, hrz= Alignment.HORIZONTAL_LEFT)
            
            idx = idx + 1
            r = r + 1
            
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    server.Put(outObj)
    
    print "return finish"
    