# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
import tempfile
import io


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def getDivisionAgents(server, division, agents):
   if division == None or len(division) == 0:
      return

   for d in division :
      for a in d.agents :
        agents.append(a.id)

      getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)

def loadAgents(server, division):
   ret = dict()
   agents = server.Get("Agents", "")

   divagents = []
   getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
   divagents = set(divagents)

   for a in agents:
      if a.id in divagents:
         ret[a.id] = a
         agentsDict[a.id] = a

   return ret

def loadAgent(server, agentID):
   ret = dict()
   agents = server.Get("Agents", "")

   for a in agents:
      if a.id == agentID:
         ret[a.id] = a

   return ret

def makeIDStr(server, agents):
   res = '"userid" in ('

   for id in agents.iterkeys():
      res += "'" + id + "',"

   res = res[:-1] + ")"
   return res

agentsDict = dict()

def doAgentReport(server, docs, orgs, repName, outObj):
   wb = Workbook(False, 'cp1251')
   sortedDocs = []
   for d in docs :
      sortedDocs.append(d)
   sortedDocs.sort(key=attrgetter('created'), reverse=True)
     
   sheet = wb.get_active_sheet()
   sheet.title = "Отчет"
   sheet.cell(row=0, column=0).value = "Дата"
   sheet.cell(row=0, column=1).value = "Агент"
   sheet.cell(row=0, column=2).value = "ТТ"
   sheet.cell(row=0, column=3).value = "Товар"
   sheet.cell(row=0, column=4).value = "Фейс план"
   sheet.cell(row=0, column=5).value = "Фейс факт"
   sheet.cell(row=0, column=6).value = "Остаток план"
   sheet.cell(row=0, column=7).value = "Остаток факт"
   
   sheet.column_dimensions[get_column_letter(1)].width = 17
   sheet.column_dimensions[get_column_letter(2)].width = 19
   sheet.column_dimensions[get_column_letter(3)].width = 18
   sheet.column_dimensions[get_column_letter(4)].width = 30
   sheet.column_dimensions[get_column_letter(5)].width = 14
   sheet.column_dimensions[get_column_letter(6)].width = 14
   sheet.column_dimensions[get_column_letter(7)].width = 14
   sheet.column_dimensions[get_column_letter(8)].width = 14
   
   price = server.Get("Price", "", "id")
   
   sr1 = 1
   for d in sortedDocs:
      for i in d.items:
         sheet.cell(row=sr1, column=0).value = d.created.strftime("%d/%m/%Y %H:%M")
         
         if d.userid in agentsDict:
            sheet.cell(row=sr1, column=1).value = agentsDict[d.userid].name
         
         o = None
         
         if d.id in orgs:
            o = orgs[d.id]   
            sheet.cell(row=sr1, column=2).value = o.name
            
         if price != None and i.id in price:
            sheet.cell(row=sr1, column=3).value = price[i.id].name
         
         face = 0;
         qty = 0;
         
         if o != None:
            matrix = o.matrix
            if matrix != None:
               for m in o.matrix:
                  if m.id == i.id:
                     face = m.face
                     qty = m.qty
         
         sheet.cell(row=sr1, column=4).value = face
         sheet.cell(row=sr1, column=6).value = qty
         sheet.cell(row=sr1, column=5).value = i.face
         sheet.cell(row=sr1, column=7).value = i.qty
            
         sr1 += 1

   fileName = tempfile.gettempdir() + '/' + repName
   wb.save(fileName)
 
   file = io.open(fileName, 'rb')
   bytes = file.read(-1)
   file.close()
    
   obj = outObj.New()
   obj.name = repName
   obj.file = bytes


def run(server):

   print "start the report"

   # get list of all params
   params = server.Params
    
   param = params[0]
    
   uidFilter = ""
   
   if param.divisionID != 0 :
       agents = loadAgents(server, param.divisionID)
       uidFilter = makeIDStr(server, agents)
   else:
       agents = loadAgent(server, param.agentID)
       uidFilter = '"userid" in (' + param.agentID + ')'

   orgs = server.Get("Org", uidFilter, "id")
   orgs.update(server.Get("PotenzialOrg", uidFilter, "id"))
   
   dateFilter = '"created" >= ToDate(\'' + param.start.strftime("%d/%m/%Y") + '\') and "created" < ToDate(\'' + param.finish.strftime("%d/%m/%Y") +"')"
   where = uidFilter + " and " + dateFilter;
   
   docs = server.Get("OrgRemnants", where)
   
   server.RegisterType("Result[name:s,file:b]")
   outObj = server.New("Result")
   
   if param.type == 1:
       doAgentReport(server, docs, orgs, "res.xlsx", outObj)
   
# 
#    doReport(orgs, price, docs, agents, param.date, edate, param.isOur, param.isConcurents, "res.xlsx", outObj)
# 
   server.Put(outObj)

   print "done " +  str(len(docs))

   
