# -*- coding: cp1251 -*-

from grsoft.xl_base import XLBuilder
from grsoft import route

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Font
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys)
sys.setdefaultencoding("cp1251")

import get_docs

class WorkTime:
    __slots__ = ('data')
    
    class Item:
        __slots__ = ('id', 'start', 'stop')
        
        def __init__(self, src):
            self.id = src.id
            self.start = src.start
            self.stop = src.stop
    
    def __init__(self, server, param):
        self.data = list()
        
        where = '"userid" in('
        for u in param.users:
            where += "'" + u.id + "',"
        where = where[:-1] + ')'
        
        wt = server.Get('WorkTime', where)
        if wt != None:
            for item in wt:
                self.data.append(WorkTime.Item(item))
                                
    def getWorkTime(self, id, date):
        ret = list()
        for i in self.data:
            if i.id == id and i.start.date() == date :
                ret.append(i)
        return ret
    
class ItemFilter:
    __slots__ = ('ids', 'name')
    
    def __init__(self, price, folders, src):
        self.ids = list()
        self.name = ''
        if src.isFolder == 0 :
            self.ids.append(src.id)
            self.name = price[src.id].name if src.id in price else 'товар с кодом <' + src.id + '>' 
        else:
            self.name = 'папка с кодом <' + src.id + '>'
            started = False
            flist = list()
            level = None
            for f in folders:
                if started: 
                    if f.level <= level: break
                    flist.append(f.id)
                if f.id == src.id:
                    started = True
                    level = f.level
                    flist.append(f.id)
                    self.name = f.name
            for p in price.itervalues():
                if p.fid in flist:
                    self.ids.append(p.id)

    def getQty(self, item):
        return item.qty if item.id in self.ids else 0

class CommonData:
    __slots__ = ('days', 'items', 'workTime')
    
    def __init__(self, server, param):
        self.days = list()
        
        cday = param.start.date()
        while cday <= param.end.date():
            if param.weekday == 0 or param.weekday == cday.isoweekday():
                self.days.append(cday)
            cday = cday + timedelta(days=1)

        price = server.Get('ManagerPrice', '', 'id')
        folders = server.Get('ManagerFolder', '')
        
        self.items = list()
        for item in param.items:
            self.items.append(ItemFilter(price, folders, item))
            
        self.workTime = WorkTime(server, param)
    

class ItemAccum:
    __slots__ = ('filter', 'qty')
    
    def __init__(self, ifilter):
        self.filter = ifilter
        self.qty = 0
        
    def accum(self, item):
        self.qty += self.filter.getQty(item)

class DailyDataItem:
    __slots__ = ('dlvCount', 'dlvSum', 'paySym', 'items', 'orderCount', 'haveDoc')
    
    def __init__(self, items):
        self.items = list()
        self.dlvCount = 0
        self.dlvSum = 0
        self.paySum = 0
        self.orderCount = 0
        self.haveDoc = False
        
        for i in items:
            self.items.append(ItemAccum(i))
            
    def addDlvDoc(self, doc, lastDate):
        if doc.date.date() == lastDate: self.haveDoc = True

        sum = 0
        self.orderCount = 1
        for srci in doc.items:
            if srci.sum != None: sum += srci.sum
            else: sum += srci.qty * srci.cost                 
            
            for i in self.items: 
                i.accum(srci)
                
        self.dlvCount += 1
        self.dlvSum += sum
        if doc.incass != None: self.paySum += doc.incass
            
    def addPayDoc(self, doc, lastDate):
        if doc.date.date() == lastDate: self.haveDoc = True
        self.paySum += doc.sum
        
    def addOtherDoc(self, doc):
        self.haveDoc = True
    
class OrgDailyData:
#
#    day->DailyDataItem
#
    __slots__ = ('data', 'items')
    
    def __init__(self, items):
        self.items = items
        self.data = dict()

    def getKey(self, doc): return doc.date.date() 
             
    def addDlvDoc(self, doc, lastDate):
        key = self.getKey(doc)
        if not(key in self.data):
            self.data[key] = DailyDataItem(self.items)
        self.data[key].addDlvDoc(doc, lastDate)

    def addPayDoc(self, doc, lastDate):
        key = self.getKey(doc)
        if not(key in self.data):
            self.data[key] = DailyDataItem(self.items)
        self.data[key].addPayDoc(doc, lastDate)

    def addOtherDoc(self, doc):
        key = self.getKey(doc)
        if not(key in self.data):
            self.data[key] = DailyDataItem(self.items)
        self.data[key].addOtherDoc(doc)

class OrgData:
#      
#    orgId->OrgDailyData
#
    __slots__ = ('data', 'items', 'orgs', 'balance')
    
    def __init__(self, items, orgs, balance):
        self.items = items
        self.data = dict()
        self.orgs = orgs
        self.balance = balance
        
    def getBalance(self, id):
        debt = 0 if not id in self.orgs else self.orgs[id].debt
        return debt if self.balance == None or (not id in self.balance) else debt - self.balance[id].sum 
        
    def getKey(self, doc): 
        if doc.id in self.orgs:
            return self.orgs[doc.id]
        
        org = OrgData.NoneOrg(doc.id)        
        self.orgs[doc.id] = org
        return org
             
    def addDlvDoc(self, doc, lastDate):
        key = self.getKey(doc)
        if not(key in self.data):
            self.data[key] = OrgDailyData(self.items)
        self.data[key].addDlvDoc(doc, lastDate)

    def addPayDoc(self, doc, lastDate):
        key = self.getKey(doc)
        if not(key in self.data):
            self.data[key] = OrgDailyData(self.items)
        self.data[key].addPayDoc(doc, lastDate)

    def addOtherDoc(self, doc):
        key = self.getKey(doc)
        if not(key in self.data):
            self.data[key] = OrgDailyData(self.items)
        self.data[key].addOtherDoc(doc)
    
    class NoneOrg :
        __slots__ = ('id', 'name')
        def __init__(self, id):
            self.id = id;
            self.name = 'Контрагент с кодом <' + id + '>'
    
class AgentData:
    __slots__ = ('name', 'id', 'route', 'orgData')
    
    def __init__(self, server, agent, items):
        self.name = agent.name
        self.id = agent.id
        self.route = route.AgentRoute(server, agent.id)

        uid = "'" + agent.id + "'"
        server.ChangeUser(uid)
        balance = server.Get('OrgBalance', "", "id")
        server.RestoreUser()
        
        self.orgData = OrgData(items, self.route.orgs, balance)
        
    def addDlvDoc(self, doc, lastDate):
        if doc.userid == self.id:
            self.orgData.addDlvDoc(doc, lastDate)

    def addPayDoc(self, doc, lastDate):
        if doc.userid == self.id:
            self.orgData.addPayDoc(doc, lastDate)

    def addOtherDoc(self, doc):
        if doc.userid == self.id:
            self.orgData.addOtherDoc(doc)

class ReportData:
    __slots__ = ('commonData', 'agents')
    
    def __init__(self, server, param):
        self.commonData = CommonData(server, param)
        self.agents = list()        
        for agent in param.users:
            self.agents.append(AgentData(server, agent, self.commonData.items))
            
    def addDlvDoc(self, doc, lastDate):
        if doc.date.date() in self.commonData.days:
            for agent in self.agents:
                agent.addDlvDoc(doc, lastDate)

    def addPayDoc(self, doc, lastDate):
        if doc.date.date() in self.commonData.days:
            for agent in self.agents:
                agent.addPayDoc(doc, lastDate)

    def addOtherDoc(self, doc):
        if doc.date.date() in self.commonData.days:
            for agent in self.agents:
                agent.addOtherDoc(doc)

def loadData(server, param):
    params = server.Params
    param = params[0]

    class DocsParam:
        __slots__ = ('start', 'end', 'detailed', 'users')
        
        def __init__(self, param):
            self.detailed = 1
            self.start = param.start
            self.end = param.end + timedelta(days=1)
            self.users = param.users

    reportData = ReportData(server, param)
    docs = get_docs.loadData(server, DocsParam(param))
    
    lastDate = param.end.date() 
    if docs != None:
        for doc in docs.dlvList:
            reportData.addDlvDoc(doc, lastDate)
          
        for doc in docs.payList:
            reportData.addPayDoc(doc, lastDate)
            
    agentWhere = ' and "userid" in('
    for ai in param.users:
        agentWhere += "'" + ai.id + "',"
    agentWhere = agentWhere[:-1] + ")"
    
    where = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
        param.end.strftime("%d/%m/%Y 0:0:0"),
        param.end.strftime("%d/%m/%Y 23:59:59"))
    where += agentWhere 
    visits = server.Get('VisitInfo', where)
    if visits != None:
        for doc in visits:
            reportData.addOtherDoc(doc) 
            
    return reportData

def getDayName(wi):
    if wi == 0: return "Все"
    return route.AgentRoute.days[int(wi)-1]

def makeHead(sheet, data, param, row):
    
    xl = XLBuilder()
    values = [ None, 
              ('Период', Border.BORDER_NONE), 
              ('c ' + param.start.strftime("%d/%m/%Y") + ' по ' + param.end.strftime("%d/%m/%Y"), Border.BORDER_NONE), 
              None,
              None, 
              ('Маршрут', Border.BORDER_NONE),
              (getDayName(param.weekday), Border.BORDER_THIN)]
    
    xl.makeCellsWithBorders(sheet, row, values) 

class XLDaily(XLBuilder):
    def drawDailyRow(self, sheet, row, values):
        cc = 0
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                c.value = value
                self.makeBorder(c, Border.BORDER_THIN)
                style = c.style 
                style.alignment.wrap_text = True
                if cc != 2:
                    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                    style.alignment.vertical = Alignment.VERTICAL_CENTER
                
            cc += 1
        
def makeDailyRow(xl, sheet, row, cdate, org, balance, workTime, orgDailyData, index): 
    values = [None, index, org.name, org.category, org.delay, balance]
    st = ''
    delta = 0
    wtimes = workTime.getWorkTime(org.id, cdate)
    for wi in wtimes:
        if st == '' : st = wi.start.strftime('%H:%M')
        delta += (wi.stop - wi.start).seconds
    values.append(st)
    values.append('%d мин' % (delta / 60) if delta > 0 else '')
    
    if orgDailyData != None:
        values.extend(
           ['X' if orgDailyData.haveDoc else '',
            orgDailyData.orderCount if orgDailyData.orderCount > 0 else '',
            orgDailyData.dlvCount if orgDailyData.dlvCount > 0 else '', 
            orgDailyData.dlvSum if orgDailyData.dlvSum > 0 else '', 
            orgDailyData.paySum if orgDailyData.paySum > 0 else ''])
        
        for i in orgDailyData.items:
            values.append(i.qty if i.qty > 0 else '')
    else:
        values.extend(['','','','','','','',''])
                
    xl.drawDailyRow(sheet, row, values)
        
def createDailyReport(sheet, data, param):
    makeHead(sheet, data, param, 2)
    
    xl = XLDaily()
    row = 4    
    for agent in data.agents:
        values = [('', Border.BORDER_NONE), ('Агент', Border.BORDER_NONE),(agent.name, Border.BORDER_NONE)]
        xl.makeCellsWithBorders(sheet, row, values, 0)
        
        titles = ['Название позиции или группы']
        for item in data.commonData.items:
            titles.append(item.name)    
        xl.makeHead(sheet, row, titles, True, 12)        
        row +=1
        
        titles = ['№', 'Клиент', 'Кат.', 'Отср.', 'Ост.лим.', 'Время начала визита', 'Время в т.р.т.', 'Был', 'Кол-во заявок', 'Кол-во накладных', 'Сумма отгрузки', 'Сумма оплат', 'Номенклатура 1', 'Номенклатура 2','Номенклатура 3']
        xl.makeHead(sheet, row, titles, True, 1)
        row += 1
        
        startRow = row
        for cdate in data.commonData.days:
            index = 0
            route = agent.route.getDayRoute(cdate)
            if len(route) == 0: continue            
            for org in route:
                index += 1
                orgDailyData = None 
                if org in agent.orgData.data and cdate in agent.orgData.data[org].data :
                    orgDailyData = agent.orgData.data[org].data[cdate]
                makeDailyRow(xl, sheet, row, cdate, org, agent.orgData.getBalance(org.id), data.commonData.workTime, orgDailyData, index)
                row += 1
            
            # out of route orgs 
            for org,orgdata in agent.orgData.data.iteritems():
                if (not org in route) and (cdate in orgdata.data):
                    index += 1
                    makeDailyRow(xl, sheet, row, cdate, org, agent.orgData.getBalance(org.id), data.commonData.workTime, orgdata.data[cdate], index)
                    row += 1

        values = []
        for ctr in range(7):
            clmn = get_column_letter(ctr+10)                    
            values.append('=SUM(' + clmn + str(startRow+1) + ':' + clmn + str(row) + ')')
        xl.makeHead(sheet, row, values, False, 9)
        
    cc = 1
    for w in [4,7,55,8,8,9,8,8,8,10,10,9,12,16,16,16]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

class PeriodBuilder(XLBuilder):
    __slots__ = ('dayCount')
    
    def __init__(self, dayCount):
        self.dayCount = dayCount
        
    def makeMergedCells(self, sheet, row, values, column):
        cc = column
        for value in values:
            if value != None:
                c = sheet.cell(row=row, column=cc)
                self.makeCell(sheet, row, cc, c, value)
                style = c.style
                style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                style.alignment.vertical = Alignment.VERTICAL_CENTER
                sheet.merge_cells(start_row=row, start_column = cc, end_row = row, end_column = cc + self.dayCount-1)
                
            cc += self.dayCount

    def makeDateCells(self, sheet, row, days, column):
        cc = column
        c = sheet.cell(row=row, column=cc)
        self.makeCell(sheet, row, cc, c, 'Даты')
        style = c.style
        style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        style.alignment.vertical = Alignment.VERTICAL_CENTER
        sheet.merge_cells(start_row=row, start_column = cc, end_row = row, end_column = cc + 5)
        cc += 6
        
        for d in days:
            value = d.strftime('%d/%m')
            
            ccc = cc
            c = sheet.cell(row=row, column=ccc)
            self.makeCell(sheet, row, ccc, c, value)

            ccc += self.dayCount
            c = sheet.cell(row=row, column=ccc )
            self.makeCell(sheet, row, ccc, c, value)
            
            ccc += self.dayCount
            c = sheet.cell(row=row, column=ccc )
            self.makeCell(sheet, row, ccc, c, value)

            ccc += self.dayCount
            c = sheet.cell(row=row, column=ccc )
            self.makeCell(sheet, row, ccc, c, value)

            ccc += self.dayCount
            c = sheet.cell(row=row, column=ccc )
            self.makeCell(sheet, row, ccc, c, value)

            cc += 1

    def makeItemsCells(self, sheet, row, items, column):
        
        self.makeCells(sheet, row, ['№', 'Клиент', 'Кат.', 'Отср.', 'Ост.лим.', 'Был'], column)

        values = ['Сумма отгрузки, грн','Сумма оплат, грн']
        for i in items:
            values.append(i.name)
        self.makeMergedCells(sheet, row, values, column + 6)
        
    def makeThikBorder(self, sheet, startRow, endRow, startColumn):
        border = Border.BORDER_MEDIUM
        
        startColumn += 6
        sheet.cell(row=startRow, column=startColumn + self.dayCount * 2).style.borders.left.border_style = border
        sheet.cell(row=startRow, column=startColumn + self.dayCount * 3).style.borders.left.border_style = border
        sheet.cell(row=startRow, column=startColumn + self.dayCount * 4).style.borders.left.border_style = border
        sheet.cell(row=startRow, column=startColumn + self.dayCount * 5).style.borders.left.border_style = border
        
        for i in range(startRow+1, endRow):
            cc = startColumn
            sheet.cell(row=i, column=cc).style.borders.left.border_style = border
            sheet.cell(row=i, column=cc + self.dayCount).style.borders.left.border_style = border
            sheet.cell(row=i, column=cc + self.dayCount * 2).style.borders.left.border_style = border
            sheet.cell(row=i, column=cc + self.dayCount * 3).style.borders.left.border_style = border
            sheet.cell(row=i, column=cc + self.dayCount * 4).style.borders.left.border_style = border
            sheet.cell(row=i, column=cc + self.dayCount * 5).style.borders.left.border_style = border
            
        for i in range(self.dayCount*5):
            row = startRow + 1 if i < self.dayCount*2 else startRow
            sheet.cell(row=row, column=i + startColumn).style.borders.top.border_style = border
            sheet.cell(row=endRow-1, column=i + startColumn).style.borders.bottom.border_style = border
            
        
def createPeriodReport(sheet, data, param):
    makeHead(sheet, data, param, 2)
    dayCount = len(data.commonData.days)
    
    xl = PeriodBuilder(dayCount)
    startRow = 4
    row = startRow
    startCol = 1
    orderCounts = list()
    
    for cdate in data.commonData.days:
        orderCounts.append(0)
        
    for agent in data.agents:
        values = [('Агент', Border.BORDER_NONE),(agent.name, Border.BORDER_NONE)]
        xl.makeCellsWithBorders(sheet, row, values, startCol)
    
        titles = ['Номенклатура 1', 'Номенклатура 2', 'Номенклатура 3']
        xl.makeMergedCells(sheet, row, titles, startCol + 6 + dayCount * 2)        
        row +=1
        
        xl.makeDateCells(sheet, row, data.commonData.days, startCol)
        row+=1
        
        xl.makeItemsCells(sheet, row, data.commonData.items, startCol)
        row+=1

        route = None
        for cdate in data.commonData.days:
            route = agent.route.getDayRoute(cdate) 
            break
        
        startRow = row
        if route != None:
            index = 0
            for org in route:
                index += 1
                values = [index, org.name, org.category, org.delay, agent.orgData.getBalance(org.id)]
                
                valuedlv = []
                valuepay = []
                valuei1 = []
                valuei2 = []
                valuei3 = []
                
                haveDocs = 'X'
                oci = 0
                for cdate in data.commonData.days:
                    if org in agent.orgData.data and cdate in agent.orgData.data[org].data :
                        orgDailyData = agent.orgData.data[org].data[cdate]
                        
                        if orgDailyData.haveDoc: haveDocs = 'Да'
                        
                        valuedlv.append(orgDailyData.dlvSum if orgDailyData.dlvSum > 0 else '')
                        valuepay.append(orgDailyData.paySum if orgDailyData.paySum > 0 else '')
    
                        i = orgDailyData.items[0]
                        valuei1.append(i.qty if i.qty > 0 else '')
    
                        i = orgDailyData.items[1]
                        valuei2.append(i.qty if i.qty > 0 else '')
    
                        i = orgDailyData.items[2]
                        valuei3.append(i.qty if i.qty > 0 else '')
                        
                        orderCounts[oci] = orderCounts[oci] + orgDailyData.orderCount
                    else:
                        valuedlv.append('')
                        valuepay.append('')
                        valuei1.append('')
                        valuei2.append('')
                        valuei3.append('')
                    oci += 1
    
                values.append(haveDocs)
                values.extend(valuedlv)
                values.extend(valuepay)
                values.extend(valuei1)
                values.extend(valuei2)
                values.extend(valuei3)
                
                xl.makeCells(sheet, row, values, startCol)
                row += 1


    xl.makeThikBorder(sheet, startRow-1, row, startCol)

    values = []
    for ctr in range(dayCount*5):
        clmn = get_column_letter(ctr+8)                    
        values.append('=SUM(' + clmn + str(startRow+1) + ':' + clmn + str(row) + ')')
    xl.makeCells(sheet, row, values, 7)
    
    row += 1
    values = ['Кол-во заказов','','']
    values.extend(orderCounts)
    xl.makeCells(sheet, row, values, 4)
    c = sheet.cell(row=row, column=4)
    c.style.alignment.wrap_text = False
    sheet.merge_cells(start_row=row, start_column = 4, end_row = row, end_column = 6)
    
    
    cc = 2
    for w in [8,50,8,8,9,8]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    for i in range(dayCount):
        w = 8
        sheet.column_dimensions[get_column_letter(cc)].width = w
        sheet.column_dimensions[get_column_letter(cc + dayCount)].width = w 
        sheet.column_dimensions[get_column_letter(cc + dayCount*2)].width = w
        sheet.column_dimensions[get_column_letter(cc + dayCount*3)].width = w
        sheet.column_dimensions[get_column_letter(cc + dayCount*4)].width = w
        cc += 1


def run(server):
    params = server.Params
    param = params[0]

    print "visit_quality start" + str(datetime.now())
    
    data = loadData(server, param)
    
    wb = Workbook(False, 'cp1251')

    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    Font.DEFAULT_FONT_NAME = 'Arial'
    Font.DEFAULT_FONT_SIZE = 9 
    
    sheet.page_margins.left = 0.1
    sheet.page_margins.right = 0.1
    sheet.page_margins.top = 0.2
    sheet.page_margins.bottom = 0.2
    sheet.page_margins.footer = 0.2
    sheet.page_margins.header = 0.2
    
    sheet.page_setup.fitToWidth = True
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.paperSize = 9
    sheet.page_setup.scale=74
     
    if param.weekday == 0 or param.start.date() == param.end.date():
        createDailyReport(sheet, data, param)
    else:
        createPeriodReport(sheet, data, param)
     
    XLBuilder().workbookToObject(wb, "visit_qual.xlsx", server)                
    print 'visit_quality ending ' + str(datetime.now())
    
