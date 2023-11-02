# -*- coding: cp1251 -*-
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
    __slots__ = ['id', 'name', 'visit_in_route', 'visit_out_route', 'visits']
    
    def __init__(self):
        self.id = ''
        self.name = ''
        self.visit_in_route = 0.0
        self.visit_out_route = 0.0
        self.visits = []
        
    def getData(self, row):
      row += 1
      res = ['', self.name, '=D{0}+E{0}'.format(row),self.visit_in_route, self.visit_out_route]
      res.extend(self.visits)
      return res
        
      
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()

def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    
    orgData = {}
    allorgs = {}
    
    for id in params.orgid.split(","):
      orgData[id] = {}
      
      date = params.start
      
      while date <= params.finish:
        orgData[id][date] = 0
        date += timedelta(days=1)
    
    for aid in params.userid.split(','):
      item = Item()
      item.aid = aid
      
      server.ChangeUser("'" + aid + "'")
      orgs = server.Get("Org", "", "id")
      server.RestoreUser()
      
      for k in orgs.keys():
        allorgs[k] = orgs[k]

      ar = AgentRoute(server, aid)
      
      date = params.start
      
      while date <= params.finish:
        route = ar.getDayRoute(date)  
        
        route_ids = []
        
        for i in route:
          route_ids.append(i.id)
          
          if i.id in orgData:
            orgData[i.id][date] = NOT_VISIT_CODE
          
        where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
            aid,
            date.strftime("%d/%m/%Y 0:0:0"),
            date.strftime("%d/%m/%Y 23:59:59"))
        
        docNames = ["Order", "VisitInfo", "OrgRemnants", "Answer", "Incass", "TaskDone", "Sales"]

        for name in docNames:
          docList = server.Get(name, where)
          
          if docList == None:
            continue
            
          if docList != None:
            for d in docList:
              if d.id in orgData:
                if d.id in route_ids:
                  orgData[d.id][date] = IN_ROUT_CODE
                else:
                  if orgData[d.id][date] < IN_ROUT_CODE:
                    orgData[d.id][date] = OUT_ROUT_CODE
                
        date += timedelta(days=1)
      
    data = ReportData()  
    
    for id in orgData:
      d = orgData[id]
      item = Item()
      item.id = id
      item.name = allorgs[id].name if id in allorgs else id
      item.visits = []
      
      date = params.start
      
      while date <= params.finish:
        v = orgData[id][date]
        item.visits.append(v)
        
        if v == IN_ROUT_CODE:
          item.visit_in_route += 1
        elif v == OUT_ROUT_CODE:
          item.visit_out_route += 1
          
        date += timedelta(days=1)
      
      data.data.append(item)
      
    data.data = sorted(data.data, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
    
    return data
    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'
  RED    = 'FFFF0000'
  GREEN  = 'FF00FF00'
  YELLOW = 'FFFFFF00'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    START_ROW = 1
    
    if column == 0:
      cell.value = row - START_ROW
      
    if column > 4:
      cell.style.alignment.wrap_text = False

      val = ''
      if value == NOT_VISIT_CODE:
        val = 'Не посетил'
        self.setBackColor(cell,XLBuilderEx.RED)
      elif value == OUT_ROUT_CODE:
        val = 'Заказ по тел.'
        self.setBackColor(cell,XLBuilderEx.YELLOW)
      elif value == IN_ROUT_CODE:
        val = 'Визит в ТТ'
        self.setBackColor(cell,XLBuilderEx.GREEN)
        
      cell.value = val
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 1
    
    head = ['№', 'Контрагент', 'Кол-во раз в маршруте', 'Кол-во раз в по маршруту', 'Кол-во раз не по маршруту']
    headDays = []
    date = params.start
      
    while date <= params.finish:
      headDays.append(date.strftime('%d.%m.%Y'))
      date += timedelta(days=1)
    
    head.extend(headDays)
      
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      xlb.makeCells(sheet, row, item.getData(row))
      row += 1
        
        
    cc = 1
    for w in [5,50,15,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    for w in headDays:
        sheet.column_dimensions[get_column_letter(cc)].width = 13
        cc += 1
        
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    