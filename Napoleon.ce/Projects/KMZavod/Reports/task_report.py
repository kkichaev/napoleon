# -*- coding: cp1251 -*-
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
  __slots__ = ['id', 'name', 'done', 'exec_date', 'out_date', 'remark', 'created']
  
  def __init__(self):
    self.id = ''
    self.name = ''
    self.done = 0
    self.exec_date = None
    self.out_date = 0
    self.remark = ''
    self.created = None
      
  def getData(self, row):
    row += 1
    res = [self.name, "1", 
      self.done,
      self.exec_date.strftime("%d.%m.%Y") if self.exec_date != None else "", 
      self.done, 
      '=B{0}-C{0}'.format(row),
      self.out_date, 
      self.out_date, 
      self.remark]
    return res
  
  def size(self):
    return 1

class GroupItem(Item):
  __slots__ = ['items']
  
  def __init__(self):
    Item.__init__(self)
    self.items = {}
    self.name = ""
    self.id = ""
    
  def addItem(self, item):
    self.items.append(item)
    
  def toList(self):
    list = self.items.values()
    return sorted(list, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
    
  def getData(self, row):
    row += 1
    res = [self.name, 
      "=SUM({0})".format(self.sumStr('B', row)) if len(self.items) > 0 else 0, 
      "=SUM({0})".format(self.sumStr('C', row)) if len(self.items) > 0 else 0,
      "", 
      "=IFERROR(C{0}/B{0},0)".format(row), 
      "=SUM({0})".format(self.sumStr('F', row)) if len(self.items) > 0 else 0,
      "=SUM({0})".format(self.sumStr('G', row)) if len(self.items) > 0 else 0, 
      "=IFERROR(G{0}/B{0},0)".format(row), 
      ""]
    return res 
  
  def sumStr(self, column, row):
    row += 1
    res = column + str(row)
    arr = self.groupIndex()
    
    for x in range(0,len(arr)-1):
      if len(res) > 0:
        res += ','
      
      row = arr[x] + row
      res += column + str(row)
      
    return res  
      
  def groupIndex(self):
    res = []
    
    for i in self.toList():
      res.append(i.size())
    
    if len(res) == 0:
      res.append(0)
      
    return res
    
  def size(self):
    ret = 1
    
    for i in self.items.values():
      ret += i.size()
      
    return ret  
    
class OrgItem(GroupItem):
  def toList(self):
    list = self.items.values()
    return sorted(list, cmp=lambda lhs, rhs: cmp(lhs.created, rhs.created))
    
class ReportData(GroupItem):
  __slots__ = ["start_data_row"]
  
  def __init__(self):
    GroupItem.__init__(self)
    self.start_data_row = 4
  
  def getData(self, row):
    ret = GroupItem.getData(self,row)
    ret[0] = "Итого"
    
    return ret
    
  def sumStr(self, column, row):
    row = self.start_data_row
    return GroupItem.sumStr(self, column, row)
      
def loadAgentDiv(server):
  divisions = server.Get("Division", "")
  ret = {}
    
  for d in divisions:
    for a in d.agents:
      ret[a.id] = d
      
  return ret    
  
def calcAutoPart(doc):
  ret = 0.0
  auto_check = 0.0
      
  for i in doc.items:
    if i.qty == i.aqty:
      auto_check += 1
      
  auto_part = 0.0

  if len(doc.items) > 0:
    ret = auto_check / len(doc.items)
    
  return ret  
    
def loadData(params, server):
  agents = server.Get("Agents", "", "id")
  
  agent_div = loadAgentDiv(server)
  
  orgids = []
  
  if len(params.orgid.strip()) > 0:
    orgids = params.orgid.strip().split(",")
    
  data = ReportData()
  
  where = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}")'.format(
    params.start.strftime("%d/%m/%Y 0:0:0"),
    params.finish.strftime("%d/%m/%Y 23:59:59"))
    
  taskRemark = server.Get('TaskRemark',where, 'taskid')  
    
  for aid in params.userid.split(','):
    if not aid in agent_div or not aid in agents:
      continue
    
    div = agent_div[aid]
    
    if not div.id in data.items:
      data.items[div.id] = GroupItem()
      data.items[div.id].name = div.name
    
    agentItem = GroupItem()
    
    grp = data.items[div.id]
    grp.items[aid] = agentItem
    
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    agentItem.name = " " * 3 + server.CurrentUser().name
    server.RestoreUser()
    
    where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
      aid,
      params.start.strftime("%d/%m/%Y 0:0:0"),
      params.finish.strftime("%d/%m/%Y 23:59:59"))
      
    tasks = server.Get('OrgTask', where)
    tasksDone = server.Get('TaskDone', where)
    
    taskDoneHash = {}
    
    for t in tasksDone:
      taskDoneHash[t.idTask] = t
      
    taskHash = {}

    for t in tasks:
      if not t.orgid in taskHash:
        taskHash[t.orgid] = []
        
      taskHash[t.orgid].append(t)  
    
    for o in orgs.values():
      if len(orgids) != 0 and not o.id in orgids or not o.id in taskHash:
        continue
        
      if not o.id in agentItem.items:
        agentItem.items[o.id] = OrgItem()
        agentItem.items[o.id].name = " " * 3 * 2 + o.name
        
      orgItem = agentItem.items[o.id]  
    
      if o.id in taskHash:
        for t in taskHash[o.id]:
          taskItem = Item()
          taskItem.name = " " * 3 * 3 + t.text
          taskItem.created = t.created
          
          if t.id in taskDoneHash:
            taskItem.done = 1
            taskItem.exec_date = t.created
              
            if (t.finish + timedelta(days=1)) < taskDoneHash[t.id].created:
              print t.finish, taskDoneHash[t.id].created
              taskItem.out_date = 1
              
          orgItem.items[t.id] = taskItem
          
          if t.id in taskRemark:
            taskItem.remark = taskRemark[t.id].remark
    
  return data
    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 4 or column == 7:
      cell.style.number_format._set_format_code('0%')

def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет по задачам"
    c.style.font.bold = True
    c.style.font.size = 14
    
    c = sheet.cell(row=1,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 2
    
    head = ['Подразделение / агент', 'Назначено', 'Выполнено', 'Дата закрытия задачи', 'Выполнено %', 'Не выполнено',
      'Просрочено', 'Просрочено %', 'Комментарий']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for div in data.toList():
      dt = div.getData(row)
      xlb.makeCells(sheet, row, dt)
      
      for x in range(0,len(dt)):
        sheet.cell(row=row,column=x).style.font.bold = True
      
      row += 1
      
      for agentItem in div.toList():
        xlb.makeCells(sheet, row, agentItem.getData(row))
        
        for x in range(0,len(dt)):
          sheet.cell(row=row,column=x).style.font.bold = True
        
        row += 1
        
        for orgItem in agentItem.toList():
          xlb.makeCells(sheet, row, orgItem.getData(row))
          
          for x in range(0,len(dt)):
            sheet.cell(row=row,column=x).style.font.bold = True
            
          row += 1
          
          for taskItem in orgItem.toList():
            xlb.makeCells(sheet, row, taskItem.getData(row))
            row += 1

    dt = data.getData(row)
    xlb.makeCells(sheet, row, dt)
    for x in range(0,len(dt)):
        sheet.cell(row=row,column=x).style.font.bold = True
    
    cc = 1
    
    for w in [50,15,15,15,15,15,15,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    