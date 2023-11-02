# -*- coding: cp1251 -*-
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
    __slots__ = ['id', 'name', 'order_cnt', 'phone_cnt', 'auto_cnt', 'auto_phone_cnt', 'orgs']
    
    def __init__(self):
        self.id = ''
        self.name = ''
        self.order_cnt = 0
        self.phone_cnt = 0
        self.auto_cnt = 0
        self.auto_phone_cnt = 0
        self.auto_org_cnt = 0
        self.orgs = {}
        
    def getData(self, row):
      row += 1
      res = [self.name, self.order_cnt, self.phone_cnt,'=B{0}-C{0}'.format(row), self.auto_cnt, 
        "=IFERROR(E{0}/B{0},0)".format(row), self.auto_phone_cnt, "=IFERROR(G{0}/B{0},0)".format(row), self.auto_org_cnt, "=IFERROR(I{0}/B{0},0)".format(row)]
      return res
        
    def getOrgList(self):
      return sorted(self.orgs.values(), cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
      
    def size(self):
      ret = 1
      
      for i in self.orgs.values():
        ret += i.size()
        
      return ret   
      
class GroupItem(Item):
    __slots__ = ['items']
    
    def __init__(self, div):
      Item.__init__(self)
      self.items = []
      self.name = div.name
      
    def addItem(self, item):
      self.items.append(item)
      
    def toList(self):
      return sorted(self.items, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
      
    def getData(self, row):
      row += 1
      res = [self.name, 
        "=SUM({0})".format(self.sumStr('B', row)) if len(self.items) > 0 else 0, 
        "=SUM(C{0}:C{1})".format(row + 1, row+len(self.items)) if len(self.items) > 0 else 0,
        '=B{0}-C{0}'.format(row), 
        "=SUM(E{0}:E{1})".format(row + 1, row+len(self.items)) if len(self.items) > 0 else 0, 
        "=IFERROR(E{0}/B{0},0)".format(row), 
        "=SUM(G{0}:G{1})".format(row + 1, row+len(self.items)) if len(self.items) > 0 else 0,
        "=IFERROR(G{0}/B{0},0)".format(row), 
        "=SUM(I{0}:I{1})".format(row + 1, row+len(self.items)) if len(self.items) > 0 else 0,
        "=IFERROR(I{0}/B{0},0)".format(row)]
        
      return res  
    
    def sumStr(self, column, row):
      row += 1
      res = column + str(row)
      arr = self.groupIndex()
      
      for x in range(0,len(arr)-1):
        if len(res) > 0:
          res += ','
        
        row = arr[x] + row
        res += column + str(row)
        
      return res  
        
    def groupIndex(self):
      res = []
      
      for i in self.toList():
        res.append(i.size())
      
      if len(res) == 0:
        res.append(0)
        
      return res
      
    def size(self):
      ret = 1
      
      for i in self.items.values():
        ret += i.size()
        
      return ret 
    
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = {}

    def toList(self):
      list = self.data.values()
      return sorted(list, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
      
def loadAgentDiv(server):
  divisions = server.Get("Division", "")
  ret = {}
    
  for d in divisions:
    for a in d.agents:
      ret[a.id] = d
      
  return ret    
  
def calcAutoPart(doc):
  ret = 0.0
  auto_check = 0.0
      
  for i in doc.items:
    if i.qty == i.aqty:
      auto_check += 1
      
  auto_part = 0.0

  if len(doc.items) > 0:
    ret = auto_check / len(doc.items)
    
  return ret  
    
def loadData(params, server):
  agents = server.Get("Agents", "", "id")
  
  agent_div = loadAgentDiv(server)
  
  orgids = []
  
  if len(params.orgid.strip()) > 0:
    orgids = params.orgid.strip().split(",")
    
  data = ReportData()
  
  for aid in params.userid.split(','):
    if not aid in agent_div:
      continue
      
    div = agent_div[aid]
    
    if not div.id in data.data:
      data.data[div.id] = GroupItem(div)
    
    item = Item()
    item.aid = aid
    
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    item.name = server.CurrentUser().name
    
    server.RestoreUser()
    
    where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
      aid,
      params.start.strftime("%d/%m/%Y 0:0:0"),
      params.finish.strftime("%d/%m/%Y 23:59:59"))
      
    orders = server.Get("Order", where)

    AUTO_ORDER_LIMIT = 0.8
    
    for o in orders:
      if len(orgids) != 0 and not o.id in orgids:
        continue
      
      item.order_cnt += 1
      
      if o.phone != 0:
        item.phone_cnt += 1
      
      auto_part = calcAutoPart(o)
      
      if auto_part >= AUTO_ORDER_LIMIT:  
        item.auto_cnt += 1
        
        if o.phone != 0:
          item.auto_phone_cnt += 1
        else:
          item.auto_org_cnt += 1
      
      if not o.id in item.orgs:
        item.orgs[o.id] = Item()
        item.orgs[o.id].name = orgs[o.id].name if o.id in orgs else 'Контрагент с кодом<{0}>'.format(o.id)
        
      orgitem = item.orgs[o.id]    
      
      orgitem.order_cnt += 1
      
      if o.phone != 0:
        orgitem.phone_cnt += 1
      
      auto_part = calcAutoPart(o)
      
      if auto_part >= AUTO_ORDER_LIMIT:  
        orgitem.auto_cnt += 1
        
        if o.phone != 0:
          orgitem.auto_phone_cnt += 1
        else:
          orgitem.auto_org_cnt += 1
    
    grp = data.data[div.id]
    grp.addItem(item)
    
  return data
    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 5 or column == 7 or column == 9:
      cell.style.number_format._set_format_code('0%')

def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11
    


    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет по автозаказу"
    c.style.font.bold = True

    
    c = sheet.cell(row=1,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 2
    
    head = ['Подразделение / агент', 'Заказов всего', 'Заказов по телефону', 'Заказов из точки', 'Сработало по автозаказу всего', '%',
      'заказов по телефону', '%', 'заказов из точки', '%']
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for grp in data.toList():
      dt = grp.getData(row)
      xlb.makeCells(sheet, row, dt)
      
      for x in range(0,len(dt)):
        sheet.cell(row=row,column=x).style.font.bold = True
      
      row += 1
      
      for item in grp.toList():
        xlb.makeCells(sheet, row, item.getData(row))
        row += 1
        
        for orgitem in item.getOrgList():
          xlb.makeCells(sheet, row, orgitem.getData(row))
          row += 1

    cc = 1
    for w in [50,15,15,15,15,15,15,15,15,15]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    