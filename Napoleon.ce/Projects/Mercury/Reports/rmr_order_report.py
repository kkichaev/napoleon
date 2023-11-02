# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging
# import random

from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Fill, Color, Alignment, NumberFormat, Border
from datetime import datetime, date, timedelta
from rmr_report_style import XLBuilderCommon

reload(sys);
#sys.setdefaultencoding("cp1251")

class Report:
  def __init__(self):
    pass
    
  def getCells(self):
    ret = [''] * len(self.cellidx)
    
    for c in self.cellidx:
      if c in self.orgs:
        o = self.orgs[c]
        name = "{0} ({1}) ".format(o.name, o.address) 
      
      ret[self.cellidx[c]] = name
      
    return ret  

class Val:
  def __init__(self):
    self.qty = 0
    self.pack = 0

class Item:
  def __str__(self):
    return self.id.encode('cp866')
    
  def getData(self):
    res = [self.id, self.name]
    
    q,p = self.putData(res)  
    res.append(q)
    res.append(p)
    
    return res
  
  def putData(self, arr):
    q = 0
    p = 0
    
    for v in self.data:
      arr.append(v.qty)
      r = round(v.pack)
      arr.append(r)
      
      q += v.qty
      p+= r
      
    return q, p
    
  def getFolder(self):
    res = ['', self.fname]
    res.extend([''] * (len(self.data) * 2 + 2))
    
    return res
    
class PriceItem(Item):
  def __init__(self):
    Item.__init__(self)

def getUsersOrgs(server, userids):
  orgs = None
  
  for item in userids:
    server.ChangeUser(item.id)
    aorgs = server.Get("Org", '', 'id')
    server.RestoreUser()  
    
    if orgs == None:
      orgs = aorgs
    else:
      orgs.update(aorgs)
  
  return orgs

def filterDocsOrgs(orgs, docs):
  map = dict()
  for d in docs:
    if not d.id in map and d.id in orgs:
      map[d.id] = orgs[d.id]
  
  return map
  
def cellIdxOrgs(map):
  cellidx = dict()   
  idx = 0
  
  for o in sorted(map.values(), key = lambda x: x.name):
    cellidx[o.id] = idx
    
    idx += 1
  
  return cellidx
    
def collectOrgs(server, userids, docs):
  orgs = getUsersOrgs(server, userids)
  map = filterDocsOrgs(orgs, docs)
  cellidx = cellIdxOrgs(map)
    
  return map, cellidx  
  
# def compare_item(x,y):
#   r = cmp(x.fname, y.fname)
#   
#   if r == 0:
#     r = cmp(x.name, y.fname)
#     
#   return r

def unpack(list):
  res = ''
  
  for i in list:
    if len(res) > 0:
      res += ','
      
    res += "'{}'".format(i.id)

  return res
  
def loadData(params, server):
  r = Report()
  
  where = '"{0}" >= ToDate("{1} 0:0:0") and "{0}" < ToDate("{2} 0:0:0") and "userid" in ({3})'.format(
    params.field,
    params.start.strftime('%d.%m.%Y'), 
    (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'),
    unpack(params.userids))
  
  ord = server.Get('Order', where)
  price = server.Get('Price', '', 'id')
  folder = server.Get('Folder', '', 'fid')
  
  r.orgs, r.cellidx = collectOrgs(server, params.userids, ord)
  
  items = dict()
  
  for o in ord:
    for i in o.items:
      if not i.id in items:
        item = Item()
        item.id = i.id
        item.name = price[i.id].name if i.id in price  else '{}'.format(i.id)
        items[i.id] = item
        item.data = [Val() for x in range(len(r.cellidx))]
        fid = price[i.id].fid if i.id in price  else ''
        item.fname = folder[fid].name if fid in folder else '{}'.format(fid)
      
      item = items[i.id]
      
      if o.id in r.cellidx:
        val = item.data[r.cellidx[o.id]]
        val.qty += i.qty
        
        if i.id in price:
          pack = price[i.id].qtyInPack
          
          if pack != 0:
            val.pack += i.qty / pack
            
  r.items = sorted(items.values(),key= lambda x: x.fname + "|" + x.name)          
  
  return r
  
class XLBuilderEx(XLBuilderCommon):
  def adjustHeadCell(self, sheet, cell, row, column):
    self.paintHeadCell(sheet.cell(row = row + 1, column=column))
    
    if column > 1:
      s = cell.style
      s.alignment.text_rotation = 90
      s.alignment.wrap_text = True 
      sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
      
      c = sheet.cell(row = row+1, column=column)
      c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
      c.value = 'шт'
      
      column += 1
      c = sheet.cell(row=row, column=column)
      self.paintHeadCell(c)
      
      c = sheet.cell(row = row+1, column=column)
      c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
      c.value = 'уп'
      self.paintHeadCell(c)
      
      SZ = 6
      sheet.column_dimensions[get_column_letter(column)].width = SZ
      sheet.column_dimensions[get_column_letter(column+1)].width = SZ
      
    return column
    
def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  xlb = XLBuilderEx()
  
  c = sheet.cell(column=0,row=0)
  c.value = 'Отчет по заявкам'
  c.style.font.bold = True
  c.style.font.size = 18
  
  c = sheet.cell(column=0,row=1)
  c.value='{0}: c {1} по {2}'.format('период', params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y'))

  head = ['Артикул', 'Номенклатура']
  head.extend(data.getCells())
  head.append('Итого')
  
  row = 2
  xlb.makeHead(sheet, row, head)
  sheet.row_dimensions[row+1].height = 100
  
  row = 4
  fname = None
  
  for i in data.items:
    if fname == None or fname != i.fname:
      fname = i.fname
      xlb.makeCells(sheet, row, i.getFolder())
      sheet.cell(row=row, column=1).style.font.bold = True
      row += 1
      
    xlb.makeCells(sheet, row, i.getData())

    sheet.cell(row=row, column=len(data.getCells() * 2) + 2).style.font.bold = True    
    sheet.cell(row=row, column=len(data.getCells() * 2) + 2 + 1).style.font.bold = True
    
    row += 1
    
  x = 1
  for w in [13,55]:
    sheet.column_dimensions[get_column_letter(x)].width = w
    x += 1
  
  return wb

def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilderCommon().workbookToObject(wb, "rmr_summary_report.xlsx", server)                
  logging.info('end')