# -*- coding: cp1251 -*-

from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter

import io
import sys


reload(sys)
sys.setdefaultencoding("cp1251")

def scope(val):
    return "'" + val + "'"

class TaskInfo:
    text = ""
    comment = ""

def getTaskById(server,id):
    result = None
    sql = '"id"=' + scope(id)
    task = server.Get("OrgTask", sql )
    
    if len(task) > 0:
        result = task[0]
        
    return  result   

def getListFromDict(orgid, id, data, check):
    if orgid in data:
        result = data[orgid]
    else:
        result = list()
        data[orgid] = result
         
    check.add(id) 
    
    return result   
    
def init(server):
    taskDone = dict()
    taskMissed = dict()
    taskMissedIds = set()
    taskDoneIds = set()
    
    param = server.Params[0];
    start = param.start.strftime('%d/%m/%Y')
    finish = param.finish.strftime('%d/%m/%Y')
    userid = param.agentID
    sql = '"userid"=' + scope(userid) + ' and "created" >= ToDate(' + scope(start) + ') and "created" < ToDate(' + scope(finish) +')'
    task = server.Get("TaskDone", sql)
    
    sql = '"userid"=' + scope(userid) + ' and "finish" >= ToDate(' + scope(start) + ') and "start" < ToDate(' + scope(finish) +')'
    taskAssign = server.Get("OrgTask", sql, "id")
#    print str(len(taskAssign)) + ' ' + sql
    
    if taskAssign == None:
        taskAssign = []
    
    if task != None:
        for t in task:
#             if len(t.idTask) == 0:
#                 continue
#             idTask = t.idTask
            
            for item in t.items:
                if item.done == 0: continue
                
                ti = TaskInfo()
                
                idTask = item.id
                if not idTask in taskAssign:
                    ct = getTaskById(server, idTask)
                    if ct == None:
                        continue                
                    taskAssign[idTask] = ct
                    
                ct = taskAssign[idTask]
                
                ti.text = ct.text
                ti.comment = t.remark
                
                lst = getListFromDict(t.id, idTask, taskDone, taskDoneIds)
                lst.append(ti)
    
    for id in taskAssign.iterkeys():
        if id in taskDoneIds:
            continue
        
        ct = taskAssign[id]
        ti = TaskInfo()
        ti.text = ct.text
        
        lst = getListFromDict(ct.orgid, ct.id, taskMissed, taskMissedIds)
        lst.append(ti)
        
    return taskDone, taskMissed, taskDoneIds, taskMissedIds
    
#     print task 
#     print start
#     print finish
#     for t in task:
#         for i in t.items:
#             ti = TaskInfo()
#             task = getTaskById(server, i.id)
#             
#             if task != None:
#                 ti.text = task.text
#                 ti.comment = i.text
#                 
#                 if i.done == 1:
#                     print "task done!"
#                     lst = getListFromDict(task.orgid, task.id, taskDone, taskDoneIds)
#                 else:
#                     lst = getListFromDict(task.orgid, task.id, taskMissed, taskMissedIds)
# 
#                 lst.append(ti)
#                  
#     sql = '"userid"=' +\
#         scope(userid) + ' and "start" >= ToDate(' + scope(start) + ') and "start" <= ToDate(' + scope(finish) +')'
#     task = server.Get("OrgTask", sql)   
#     
#     for t in task:
#         if not (t.id in taskMissedIds or t.id in taskDoneIds):
#             ti = TaskInfo()
#             ti.text = t.text 
#             lst = getListFromDict(t.orgid, t.id, taskMissed, taskMissedIds)
#             lst.append(ti)    
#             
#     print taskDone        

def makeGrouHeader(html, title, cnt):  
    html += "<tr bgcolor='#DCDCDC'>"
    html += "<td>"+title+"&nbsp;(" + str(cnt) +")</td>"
    html += "<td>Коментарий</td>"
    html += "</tr>"
    return html

def makeDataRow(html, t):
    html += "<tr>"
    text = "&nbsp;"
    comment = "&nbsp;"
    
    if len(t.text) > 0 :
        text = t.text
        
    if len(t.comment) > 0 :
        comment = t.comment
        
    html += "<td width='50%'>" + text + "</td>"
    html += "<td width='50%'>" + comment + "</td>"  
    html += "</tr>"
    
    return html

def makeHtmlData(html, title, tasklists):
    html = makeGrouHeader(html, title, len(tasklists))
                
    for t in tasklists:
        html = makeDataRow(html, t)
    
    return html
                                              
def doReport(server, param):
    html = "<html><head>" + \
           "<meta http-equiv='content-type' content='text/html; charset=utf-8'></head>" + \
           "<body><FONT FACE='Arial'>"
    html += "<H1>Отчет по заданиям</H1><br>"  
    html += "<H2>Агент:&nbsp;" + param.agentName
    f = param.finish-timedelta(days=1)
    html += "&nbsp;c " + param.start.strftime('%d/%m/%Y') + "&nbsp;по&nbsp;" + f.strftime('%d/%m/%Y') + "</H2>"     
    html += "</FONT></body></html>"       
    uidFilter = '"userid" in ' + "('" + param.agentID + "')"
    orgs = server.Get("Org", uidFilter, "id")
    taskDone, taskMissed, taskDoneIds, taskMissedIds = init(server)
     
    for o in orgs.values() :
        if o.id in taskDone or o.id in taskMissed: 
            html += "<table width='100%' border='1' >"
            html += "<tr  bgcolor='#F5FFFA'>"
            html += "<td colspan='2'><b>" + o.name + "</b></td>"
            html += "</tr>"
            
            if o.id in taskDone:
                html = makeHtmlData(html, "Выполнено", taskDone[o.id])
            
            if o.id in taskMissed:    
                html = makeHtmlData(html, "Не выполнено", taskMissed[o.id])
                
            html += "</table><br>"
    return html

class TaskReportItem:
    __slots__ = ['task', 'comment', 'isDone', 'ended', 'responceDate']
    
    def __init__(self, task):
        self.task = task
        self.comment = ''
        self.isDone = False
        self.ended = ''
        self.responceDate = None

class TaskReportData:
    __slots__ = ['agents', 'orgs', 'tasks', 'taskByID']
    
    def __init__(self):
        self.tasks = list()
        self.taskByID = dict()
        
    def addTask(self, task):
        ti = TaskReportItem(task)
        self.tasks.append(ti)
        self.taskByID[task.id] = ti
        
    def addTaskDone(self, tdone):
        for tdi in tdone.items:
            if not tdi.id in self.taskByID: continue
            
            task = self.taskByID[tdi.id]
            if len(tdi.text) > 0 :
                task.comment = tdi.text
            if tdi.done > 0: 
                task.isDone = True
                task.responceDate = tdone.created
                
        
def taskReportXLS(server, param):
    agents = server.Get('Agents', '', 'id')
    managers = server.Get('DivisionManager','','login')
    
    allorgs = {}

    for ai in param.agents:
      server.ChangeUser("'" + ai.id + "'")
      orgs = server.Get("Org", "", "id")
      server.RestoreUser()
      
      for k in orgs.keys():
        allorgs[k] = orgs[k]
    
      where = '"start" <= ToDate("{1}") and "finish" >= ToDate("{0}")'.format(
          param.start.strftime("%d/%m/%Y 0:0:0"),
          param.finish.strftime("%d/%m/%Y 0:0:0"))
      
      where += ' and "userid"="' + ai.id + '"'
      tasks = server.Get('OrgTask', where)

      where = '"date" >= ToDate("{0}")'.format(
      param.start.strftime("%d/%m/%Y 0:0:0"))
      
      taskRemark = server.Get('TaskRemark',where, 'taskid')  
    
      repData = TaskReportData()
      for doc in tasks:
        if doc.orgid in orgs:
          repData.addTask(doc)
      
      where = '"created" >= ToDate("{0}")'.format(
          param.start.strftime("%d/%m/%Y 0:0:0"))
        
      where += ' and "userid"="' + ai.id + '"'
      doneTask = server.Get('TaskDone', where)
      
      for doc in doneTask:
        if doc.id in orgs:
          repData.addTaskDone(doc)

    wb = Workbook(False, 'cp1251')    
    xlb = XLBuilder()
    sheet = wb.get_active_sheet()
    sheet.title = "отчет по задачам"
    row = 1
    
    xlb.makeHead(sheet, row, ["Название магазина", "Адрес", "Период", "Задача", "Выполнена (да /нет)", "Комментарий сотрудника", 'Агент', 'Дата выполнения', 'Постановщик', 'Дата создания'], True)
    
    minDate = datetime(1970,1,1)
    
    for item in repData.tasks:
        task = item.task
        
        if not task.orgid in allorgs:
          continue
          
        value = []
        
        o = allorgs[task.orgid]
        value.append(o.name)
        value.append(o.address)
        
        value.append("{0}-{1}".format(task.start.strftime("%d.%m.%Y"), task.finish.strftime("%d.%m.%Y")))
        value.append(task.text)
        value.append( 'Да' if item.isDone else 'Нет')
        
        if task.id in taskRemark:
          value.append(taskRemark[task.id].remark)
        else:
          value.append(' ')

        #task.comment)    
        
        if task.userid in agents:
            value.append(agents[task.userid].name)
        else:
            value.append(task.userid)
        
        
        value.append(item.responceDate.strftime("%d.%m.%Y") if item.responceDate != None else '')

        value.append(task.manager)
        if task.created < minDate: value.append('')
        else: value.append(task.created.strftime("%d.%m.%Y %H:%M"))

        row += 1
        xlb.makeCells(sheet, row, value)
        
    cc = 1
    for w in [20,20,15,15,15,35,35,15,20,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    XLBuilder().workbookToObject(wb, "task_report.xlsx", server)
    
def taskReport(server, param):
    server.RegisterType("Result[html:s]")
    objList = server.New("Result")
    obj = objList.New()
    obj.html = doReport(server, param)
    server.Put(objList)

def taskList(server, param):
    server.RegisterType("Result[id:s,name:s,address:s,done:n,missed:n]")
    objList = server.New("Result")
    
    allorgs = dict()
    
    for u in param.agentID.split(','):
      server.ChangeUser("'" + u + "'")
      orgs = server.Get("Org", "", "id")
      server.RestoreUser()
      
      for k in orgs.keys():
        if not k in allorgs:
          allorgs[k] = orgs[k]
    
    taskDone, taskMissed, taskDoneIds, taskMissedIds = init(server)
    
    for o in allorgs.values() :
        obj = objList.New()
        obj.id = o.id
        obj.name = o.name
        obj.address = o.address

        
        if o.id in taskDone:
            obj.done = len(taskDone[o.id])
        
        if o.id in taskMissed:    
            obj.missed = len(taskMissed[o.id])

    
    server.Put(objList)

   
