# -*- coding: cp1251 -*-

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import sys

reload(sys);
sys.setdefaultencoding("cp1251")

class RepData:
    __slots__ = ['id', 'reason']
    
    def __init__(self):
        self.id = ''
        self.reason = ''
    
class OrgData:
    __slots__ = ['id', 'date', 'items']
    
    def __init__(self, id, date):
        self.items = list()
        self.id = id
        self.date = date
        
    def add(self, rejItem, reasons):
        data = RepData()
        data.id = rejItem.id
        
        data.reason = reasons[rejItem.reason].name if rejItem.reason in reasons else rejItem.reason
        
        self.items.append(data)
        
    def get(self, id):
        for ri in self.items:
            if ri.id == id:
                return ri.reason
        return ''

def loadData(server, param):
    uid = "'" + param.userid + "'"
    server.ChangeUser(uid)
    
    price = server.Get('Price', 'SetQtyFilter(false)', 'id')
    orgs = server.Get('Org', '', 'id')
    
    server.RestoreUser()
    
    reasons = server.Get('FocusRejectReason', '', 'id')
    
    where = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}") and "userid" = {2}'.format(
        param.start.strftime("%d/%m/%Y"),
        param.finish.strftime("%d/%m/%Y 23:59:59"),
        uid)
    
    rejected = list()    
    data = list()
    
    docs = server.Get('Order', where)
    if docs != None:  
        for d in docs:
            cur = None
            for ri in d.rejectItems:
                if cur == None:
                    cur = OrgData(d.id, d.date)
                    data.append(cur)
                cur.add(ri, reasons)
                if not ri.id in rejected:
                    rejected.append(ri.id)
                    
    return data, rejected, orgs, price      
            


class XLB(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        if column >= 2:
            cell.style.alignment.text_rotation = 90
        return column
    
def printOut(data, rejected, orgs, price, param):
    xlb = XLB()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"

    cr = 1
    cc = 0
    cell = sheet.cell(row=cr, column=cc)
    cell.value = "Агент: " + param.name

    cr += 1
    cell = sheet.cell(row=cr, column=cc)
    cell.value = 'период с {0} по {1}'.format(
        param.start.strftime("%d/%m/%Y"),
        param.finish.strftime("%d/%m/%Y"))
    
    head = ['Клиенты/Номенклатура', 'Дата']
    wdh = [45]
    for ri in rejected:
        name = price[ri].name if ri in price else 'Товар с кодом ' + ri
        head.append(name)
        wdh.append(15)
        
    cr += 1
    xlb.makeHead(sheet, cr, head, True)
    
    for di in data:
        cr += 1
        name = orgs[di.id].name if di.id in orgs else 'Клиент с кодом ' + di.id
        values = [name, di.date.strftime("%d.%m.%Y")]
        for ri in rejected:
            rsn = di.get(ri)
            values.append(rsn)
        xlb.makeCells(sheet, cr, values)


    cc = 1
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1


    return wb

def run(server):
    param = server.Params[0]
    data, rejected, orgs, price = loadData(server, param)
    wb = printOut(data, rejected, orgs, price, param)
    
    XLBuilder().workbookToObject(wb, "agentorder.xlsx", server) 