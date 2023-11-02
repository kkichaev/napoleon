# -*- coding: cp1251 -*-
from importlib import reload
import sys
import logging
import locale
import time

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta

reload(sys);

class Item:
    def __init__(self):
        self.userid = ''
        self.username = ''
        self.org_in_route = 0.0
        self.visit_in_route = 0.0
        self.visit_out_route = 0.0
        self.order_sum = 0.0
        now = datetime.now()
        self.start = None
        self.finish = None
        self.aid = ''
        self.photocount = 0
        self.sku_base = 0
        self.sku_plan = 0
        self.sku_fact = 0
        self.script_time = 0
        self.out_scritp_time = 0
        
    def getData(self, row):
      row += 1
      return [self.aid, self.username, self.org_in_route, self.visit_in_route, self.visit_out_route, "=G{0}+H{0}".format(row), 
        self.photocount, 
        self.visit_in_route / self.org_in_route if self.org_in_route > 0 else 0,
        (self.visit_in_route) / self.org_in_route if self.org_in_route > 0 else 0,
        # (self.visit_in_route + self.photocount) / self.org_in_route if self.org_in_route > 0 else 0,
        self.sku_base,
        self.sku_plan,
        self.sku_fact,
        self.start.strftime("%H:%M") if self.start != None else "", 
        self.finish.strftime("%H:%M") if self.finish != None else "", 
        "{0:02d}:{1:02d}".format(int(self.script_time / 3600), int(self.script_time / 60 % 60)),
        "{0:02d}:{1:02d}".format(int(self.out_scritp_time / 3600), int(self.out_scritp_time / 60 % 60)),]
      
class ReportData:
    def __init__(self):
      self.data = list()

def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    data = ReportData()
    div = server.Get('Division', '"id"=' + str(params.divid))
    
    if len(div) > 0:
      data.divname = div[0].name

    for aid in params.userids:
      item = Item()
      item.userid = aid.id
      day_start = list()
      day_finish = list()
      item.aid = aid.id
      
      server.ChangeUser("'" + item.aid + "'")
      orgs = server.Get("Org", "", "id") 
      item.username = server.CurrentUser().name
      server.RestoreUser()
      
      for oid in orgs:
        item.sku_base += orgs[oid].sku
      
      ar = AgentRoute(server, item.aid)
      
      date = params.start
      
      while date <= params.finish:
        route = ar.getDayRoute(date)  
        
        route_ids = []
        start = None
        finish = None
        
        for i in route:
          route_ids.append(i.id)
        
        item.org_in_route += len(route)

        server.ChangeUser("'" + item.aid + "'")
        q = date.strftime("%d/%m/%Y 0:0:0") + ';' + date.strftime("%d/%m/%Y 23:59:59")
        photocount = server.Get('PhotoCount', q)
        server.RestoreUser()
      
        server.RestoreUser()
        
        where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
            item.aid,
            date.strftime("%d/%m/%Y 0:0:0"),
            date.strftime("%d/%m/%Y 23:59:59"))
        
        visit_in_route = []
        visit_out_route = []
        sku_plan = 0
        sku_fact = 0

        docNames = ["Order", "VisitInfo", "OrgRemnants", "Answer", "Incass", "TaskDone", "Sales", "ScriptDoc"]
        
        for p in photocount:
          if p.id in route_ids and not p.id in visit_in_route:
            visit_in_route.append(p.id)
            
            if p.id in orgs:
              sku_plan += orgs[p.id].sku
            
          elif not p.id in route_ids and not p.id in visit_out_route:
            visit_out_route.append(p.id)
        
        script_time = 0
        out_scritp_time = 0
        
        for name in docNames:
          docList = server.Get(name, where)
          
          if docList == None:
            continue

          if docList != None:
            sz = len(docList)
            finish_script = None
            
            for d in docList:
              if name == "ScriptDoc":
                ss = d.created
                
                if d.items != None:
                  for i in d.items:
                    if i.state == 1 and i.date > ss:
                      ss = i.date
                
                script_time += (ss - d.created).total_seconds()
                
                if finish_script != None:
                  out_scritp_time += (d.created - finish_script).total_seconds()
                  
                finish_script = ss
                
              if name == "OrgRemnants":
                sku_fact += len(d.items)
                
              if start == None or start > d.created:
                start = d.created
                
              if finish == None or finish < d.created:
                finish = d.created
        
        if start != None:
          day_start.append(start)
          day_finish.append(finish)
        
        item.visit_out_route += len(visit_out_route)
        item.visit_in_route += len(visit_in_route)
        item.photocount += len(photocount)
        item.sku_plan += sku_plan
        item.sku_fact += sku_fact
        item.script_time += script_time
        item.out_scritp_time += out_scritp_time
        
        date += timedelta(days=1)

      if len(day_start) != 0:
         sumH = 0
         for val in day_start:
           sumH += val.hour * 3600 + val.minute * 60 + val.second
         item.start = datetime(2019, 1, 1, int((sumH / 3600) / len(day_start)), int((sumH % 3600 / 60)/ len(day_start)), int((sumH % 60)/len(day_start)))

         sumH = 0
         for val in day_finish:
           sumH += val.hour * 3600 + val.minute * 60 + val.second
         item.finish = datetime(2019, 1, 1, int((sumH / 3600) / len(day_finish)), int((sumH % 3600 / 60)/ len(day_finish)), int((sumH % 60)/len(day_finish)))
      
      data.data.append(item)
      
    data.data = sorted(data.data, key=lambda x: x.username)
    
    return data
    
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 10 or column == 11:
      cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 3
    
    head = ['Дата начала', 'Дата конца', 'филиал', 'ИД', 'ТП', 'Итого точек в маршруте', 'По маршруту посещений', 'Не по маршруту посещений', 'Итого посещений', 
      'Фотоотчеты', '% Посещений', '% Фотоотчетов', 'СКЮ базовая', "СКЮ план", "СКЮ факт", "Начало", "Окончание", "часов отработанных в чистом виде без перемещения", "времени потраченное на перемещение"]
    
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      values = [params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"), data.divname]
      values.extend(item.getData(row))
      
      xlb.makeCells(sheet, row, values)
      row += 1

    cc = 1
    SZ = 12
    
    for w in [SZ,SZ,SZ,SZ,26,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ,SZ]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    