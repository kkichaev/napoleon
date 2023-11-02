# -*- coding: cp1251 -*-
from importlib import reload
import sys
import logging

from grsoft import route
from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

from datetime import timedelta

reload(sys)

class EmptyOrg:
    def __init__(self, id):
        self.id = id
        self.name = 'Контрагент с кодом <' + id + '>'
        self.address = ''

class OrgRouteData :
    def __init__(self, org):
        self.org = org
        # [weekNo=>count]
        self.weekData = dict()
        self.orders = 0
        self.sum = 0

    def addVisit(self, week):
        if not week in self.weekData:
            self.weekData[week] = 1
        else:
            self.weekData[week] += 1

    def addOrder(self, doc):
        self.orders += 1
        for oi in doc.items:
            self.sum += oi.cost * oi.qty

class RouteEx(route.AgentRoute) :
    def orgInRoute(self):
        loaded = dict()
        for ri in self.route.values():
            for rii in ri.items:
                code = rii.name
                corg = self.orgs[code] if code in self.orgs else EmptyOrg(code)
                if not code in loaded:
                    loaded[code] = OrgRouteData(corg) 
        return loaded

class Object:
    pass

class ReportData:
    def __init__(self, server, params) -> None:
        uid = params.userids[0].id
        self.route = RouteEx(server, uid)

        # [id=>OrgRouteData]
        self.orgInRoute = self.route.orgInRoute()

        # [id=>{org,count}]
        self.outRoute = dict()

        agents = server.Get('Agents', '', 'id')
        self.agent = agents[uid]
        self.routeInfo = {}

        userid = "'" + uid + "'" 
        cd = params.start
        endDate = params.finish + timedelta(days=1)
        while cd < endDate :
            ed = cd + timedelta(days=1)
            week = cd.isocalendar()[1]

            whereOrders = '"{0}" >= ToDate("{1}") and "{0}" < ToDate("{2}") and "userid"={3}'.format(
                "created", cd.strftime('%d/%m/%Y'), ed.strftime('%d/%m/%Y'), userid)

            whereVisit = '"{0}" >= ToDate("{1}") and "{0}" < ToDate("{2}") and "userid"={3}'.format(
                "created", cd.strftime('%d/%m/%Y'), ed.strftime('%d/%m/%Y'), userid)

            dailyOrgs = self.route.getDailyOrgList(cd)
            if dailyOrgs == None: dailyOrgs = []
            else:
                ol = []
                for di in dailyOrgs.items:
                    ol.append(di.name)
                    
                    if not di.name in self.routeInfo:
                        self.routeInfo[di.name] = 0

                    self.routeInfo[di.name] = self.routeInfo[di.name] +1    
                dailyOrgs = ol

            docs = server.Get('VisitInfo', whereVisit)
            # docs = server.Get('Order', whereOrders)
            if docs != None:
                for d in docs:
                    code = d.id
                    if not code in self.orgInRoute:
                        if not code in self.outRoute:
                            data = Object()
                            data.org = self.route.orgs[code] if code in self.route.orgs else EmptyOrg(code)
                            data.count = 1
                            self.outRoute[code] = data
                        else:
                            self.outRoute[code].count += 1
                    elif code in dailyOrgs:
                        self.orgInRoute[code].addVisit(week)

            docs = server.Get('Order', whereOrders)
            if docs != None:
                for d in docs:
                    code = d.id
                    if code in self.orgInRoute:
                        self.orgInRoute[code].addOrder(d)
            cd = ed

def loadData(server, params):
    data = ReportData(server, params)
    return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()

    cell = sheet.cell(row=0, column=0)
    cell.value = "Контроль соответсвия маршрутов: {0}".format(data.agent.name)

    cell = sheet.cell(row=1, column=0)
    cell.value = "С {0} по {1}".format(params.start.strftime('%d/%m/%Y'), params.finish.strftime('%d/%m/%Y'))

    xlb = XLBuilder()

    colWidth = [40]
    head = ['Клиент']
    weeks = []
    cd = params.start
    endDate = params.finish + timedelta(days=1)
    while cd < endDate:
        weeks.append(cd.isocalendar()[1])
        head.append(str(len(weeks)) + ' неделя')
        colWidth.append(10)
        cd = cd + timedelta(weeks=1)

    head.extend(['Итого посещений', 'Итого периодичность','Периодичность плановых посещений', 'Кол-во заявок за период','Сумма реализаций за период'])
    colWidth.extend([15,40,15,15])
    row = 2
    xlb.makeHead(sheet, row, head, True)

    orgs = list()
    for d in data.orgInRoute.values():
        orgs.append(d.org)

    row += 1
    notVisited = []

    for o in sorted(orgs, key=lambda x: x.name):
        name = o.name
        if len(o.address) > 0: name += '(' + o.address + ')'
        values = [name]

        rowData = data.orgInRoute[o.id]
        
        wc = 0
        vc = 0
        for w in weeks:
            if w in rowData.weekData:
                visits = rowData.weekData[w]
                values.append(visits)
                wc += 1
                vc += visits
            else:
                values.append('')
        values.extend([vc,wc, data.routeInfo[o.id] if o.id in data.routeInfo else 0, rowData.orders, rowData.sum])

        xlb.makeCells(sheet, row, values)
        row += 1

        if vc == 0 and rowData.orders == 0:
            notVisited.append(name)

    orgInRoute = len(orgs)
    values = ['Итого по маршруту']
    ctr = 2

    if orgInRoute > 0:
        for w in range(0, len(weeks)+4):
            ltr = get_column_letter(ctr)
            sumF = '=SUM({0}{1}:{0}{2})'.format(ltr, 4, orgInRoute + 3)
            values.append(sumF)
            ctr += 1

    xlb.makeCells(sheet, row, values)
    row += 2
    trow = row
    
    head = ['Контрагенты без маршрута','']
    xlb.makeHead(sheet, row, head)
    sheet.merge_cells(start_row = row, start_column = 0, end_row = row, end_column = 1)
    row += 1

    orgs = []
    for o in data.outRoute.values():
        orgs.append(o.org)

    for o in sorted(orgs, key=lambda x: x.name):
        name = o.name
        if len(o.address) > 0: name += '(' + o.address + ')'
        values = [name]

        # print (data.outRoute.keys())
        rowData = data.outRoute[o.id]
        values.append(rowData.count)

        xlb.makeCells(sheet, row, values)
        row += 1

    # values = ['Итого без маршрута', '=SUM(B{0}:B{1})'.format(orgInRoute + 5, row)]
    values = ['Итого без маршрута', str(len(orgs))]
    # print(values[1])
    xlb.makeCells(sheet, row, values)

    row = trow

    head = ['Контрагенты без посещений','']
    xlb.makeHead(sheet, row, head, startColumn=3)
    sheet.merge_cells(start_row = row, start_column = 3, end_row = row, end_column = 4)
    row += 1    

    for name in sorted(notVisited, key=lambda x: x):
        values = [name, 0]
        xlb.makeCells(sheet, row, values, 3)
        row += 1

    x = 1
    for w in colWidth:
      sheet.column_dimensions[get_column_letter(x)].width = w
      x += 1

    return wb

def run(server) :
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(server, params)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "route_control_report.xlsx", server)                
    logging.info('end')
