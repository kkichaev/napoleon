# -*- coding: cp1251 -*-
from importlib import reload
import logging
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from grsoft.route import AgentRoute
from manager.document import docTypes
from datetime import timedelta
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat,Border
from orgmap import OrgMap

import sys;
reload(sys);
#sys.setdefaultencoding("cp1251")

class Data:
  def __init__(self):
    self.items = []

class Item:
  def __init__(self):
    self.orgname = ''
    self.created = ''
    self.remark = ''
    self.agent = ''
    
  def getData(self):
    return [self.agent, self.orgname, self.created.strftime("%d.%m.%Y"), self.remark]
  
def loadData(params, server):
  data = Data()
  orgMap = OrgMap(server)

  for pi in params.userids:
    server.ChangeUser("'" + pi.id + "'")
    agentname = server.CurrentUser().name
    server.RestoreUser()
    
    where = '"userid"="{0}" and "date" >= ToDate("{1}") and "date" <= ToDate("{2}")'.format(
      pi.id,
      params.start.strftime("%d/%m/%Y 00:00:00"),
      params.finish.strftime("%d/%m/%Y 23:59:00"))
    
    
    docs = server.Get("OrgNotes", where)
    
    for d in docs:
      item = Item()
      item.agent = agentname
      item.orgname = orgMap.getOrg(d.id, d.userid).name
      item.created = d.date
      item.remark = d.text
      data.items.append(item)

  data.items = sorted(data.items, key=lambda x: (x.agent, x.created))
  
  return data
  
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'
  
  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
    
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    if column == 6:
      value = "" if value == 0 else value
      
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border) 

    if column == 6:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
    
def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  
  cell = sheet.cell(row=0, column=0)
  cell.value = "Отчет по заметкам"
  cell.style.font.bold = True
  
  DATE_FMT = "%d.%m.%Y"
  sheet.cell(row=1, column=0).value = "Интервал: {0} - {1}".format(params.start.strftime(DATE_FMT), params.finish.strftime(DATE_FMT))
  
  head = ["Агент", "Клиент", "Дата", "Сохраненная заметка"]
  xlb = XLBuilderEx()
  xlb.makeHead(sheet,3,head)
  sheet.row_dimensions[sheet.cell(row=3, column=0).row].height = 34
  
  r = 4
  
  for item in data.items:
    xlb.makeCells(sheet, r,  item.getData())
    r += 1

  x = 1
  for w in [18,40,17,40]:
    sheet.column_dimensions[get_column_letter(x)].width = w
    x += 1

  return wb

def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilder().workbookToObject(wb, "rmr_visit_report.xlsx", server)                
  logging.info('end')
