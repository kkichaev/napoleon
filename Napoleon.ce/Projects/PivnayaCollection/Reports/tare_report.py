# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

class AgentData:
    __slots__ = ['docs']
    
    def __init__(self, server, userid, start, finish):
        where = '"{2}" >= ToDate("{0}") and "{2}" <= ToDate("{1}") and "userid" = "{3}"'.format(
           start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 23:59:59"), "created", userid)
                
        self.docs = server.Get('TareDoc', where)
#         print len(self.docs)

def loadData(server):
    params = server.Params[0]
    
    agents = server.Get('Agents', '', 'id')
    invs = server.Get('Inventory', '', 'id')
    
    data = dict()
    
    orgs = dict()
    for ai in params.items:
        uid = "'" + ai.id + "'"
        server.ChangeUser(uid)
        
        aorgs = server.Get('Org', '', 'id')
        orgs.update(aorgs)
        server.RestoreUser()
        
        data[ai.id] = AgentData(server, ai.id, params.start, params.finish) 
        

    return agents, invs, orgs, data

def printOut(agents, invs, orgs, data):
    wb = Workbook(False, 'cp1251')
    xlb = XLBuilder()
        
    sheet = wb.get_active_sheet()
    sheet.title = "отчет"

    row = 1
    head = ['Дата','ФИО ТП','Наименование клиента','Номенклатура','Задложенность точки','Количество тары в торговой точке','Отклонение']
    xlb.makeHead(sheet, row, head, True)    
    row += 1
    
    for uid, data in data.iteritems():
        agent = agents[uid].name if uid in agents else uid

        for d in data.docs:
            org = d.id if not d.id in orgs else orgs[d.id].name
            for i in d.items:
                inv = i.id if not i.id in invs else invs[i.id].name
                
                diff = i.fact - i.qty 
                values = [d.created.strftime('%d.%m.%Y'), agent, org, inv, i.qty, i.fact, diff]
                xlb.makeCells(sheet, row, values)

                if diff < 0:
                    cell = sheet.cell(row=row, column=6)
                    style = cell.style.font.color = Color(Color.RED)
                    
                row += 1
                
    cc = 1
    wdh = [20,20,35,35,20,20,20]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    return wb

def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    
    agents, invs, orgs, data = loadData(server)

    wb = printOut(agents, invs, orgs, data)
    XLBuilder().workbookToObject(wb, "tare.xlsx", server)

    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    