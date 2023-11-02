# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ffC0C0C0"
userOrgs = dict()

class Item:
    org = None
    item = None
    qty = None
    fact = None
    clear = None
    good = None
    wash = None
    
    def __init__(self):
        self.qty = 0
        self.fact = 0
        self.good = 0
        
    def getData(self):
        return [self.org, self.item,  self.qty, self.fact, self.clear, self.good, self.wash]

WHERE_STR = '"{2}" >= ToDate("{0}") and "{2}" <= ToDate("{1}") and "userid" = "{3}"'; 

def inflateParams(server):
    return server.Params[0].begin, server.Params[0].end + timedelta(days=1), server.Params[0].userid

def getOrgName(server, id, userid):
    global userOrgs
    if not userid in userOrgs:
        server.ChangeUser("'" + userid + "'")
        o = server.Get("Org", "", "id")
        server.RestoreUser()
        userOrgs[userid] = o
    
    orgs = userOrgs[userid] 
    return orgs[id].name if id in orgs else "Контрагент с кодом<{0}>".format(id)

def getItemName(id, price):
    return price[id].name if id in price else "Товар с кодом<{0}>".format(id)

def getAgentName(id, agents):
    return agents[id].name if id in agents else "Торговый представитель с кодом<{0}>".format(id)

         
def collectData(server, inva):
    result = list()
    
    invs = server.Get('Inventory', '', 'id')
    
    for i in inva:
        for ii in i.items:
            item = Item()
            item.org = getOrgName(server, i.id, i.userid)
            item.item = invs[ii.id].name if ii.id in invs else ii.id
            item.qty = ii.qty
            item.fact = ii.fact
            item.clear = '-' if ii.clear == 0 else '+'
            item.good = ii.good
            item.wash = '{0}/{1}'.format(i.penult.strftime("%d.%m.%Y"), i.last.strftime("%d.%m.%Y")) 
            result.append(item)
                
    return result

def loadData(server):
    start, finish, userid = inflateParams(server)
    
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 23:59:59"), "created", userid)
    inva = server.Get("InvAudit", where)
    inva.sort(cmp= lambda x, y: cmp(x.created, y.created))
    
    sd1 = collectData(server, inva)
    
    a = server.Get("Agents", "", "id")
    
    rd = RptData()
    rd.start = start
    rd.finish = finish - timedelta(days=1)
    rd.agent = a[userid].name if userid in a else userid
    
    return rd, sd1 


class XLBuilderEx(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment.wrap_text = True
        self.setBackColor(cell, bkgColor)
        return column    
    
    def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def drawData(sh, xlb, data):
    row = 2
    for d in data:
        rd = d.getData()
        xlb.makeCells(sh, row, rd)
        row += 1
        
def paintHeadCell(xlb, cell):
    cell.style.alignment.wrap_text = True
    xlb.setBackColor(cell, bkgColor)
    xlb.paintHeadCell(cell)
    
def printData(xlb, sh, rpd, d):
    sh.cell(row=0, column=0).value = "Агент: {2}, Период с {0} по {1}".format(rpd.start.strftime("%d/%m/%Y"), rpd.finish.strftime("%d/%m/%Y"), rpd.agent)
    titles = ['Наименование торговой точки', 'Оборудование (пегас, башня, заборная говолка, редуктор)', 'Кол-во по 1С', 'Факт', 'Чистота', 'Исправность', 'Журнал промывок (дата последних двух)'];
    xlb.makeHead(sh, 1, titles, True)
    drawData(sh, xlb, d) 
    setCellWidth(sh, [40,40,16,16,16,16,22])

def printOut(rpd, d1):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    printData(xlb, sh, rpd, d1)
    
    return wb

class RptData:
    agent = None
    start = None
    finish = None
          
def doReport(server):
    rd, data = loadData(server)
    print "data", data
    wb = printOut(rd, data)
    XLBuilder().workbookToObject(wb, "visit.xlsx", server)
          
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    