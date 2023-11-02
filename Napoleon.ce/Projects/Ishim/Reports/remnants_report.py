import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

#import sys;
#reload(sys);
#sys.setdefaultencoding("cp1251")

class PriceItem:
  __slots__ = ['id', 'name', 'plan', 'fact', 'cause']
  
  def __init__(self):
    self.id = ""
    self.plan = 0.0
    self.fact = 0.0
    self.cause = ""
    
  def getData(self, row):
      row += 1
      return ['      ' + self.name, 
        self.plan, 
        self.fact,  
        self.cause,
        "=IFERROR(C{0}/B{0},0)".format(row),]  
      
class OrgItem:
  __slots__ = ['id', 'name', 'data']
  
  def __init__(self):
    self.id = ""
    self.name = ""
    self.data = []
    
  def getData(self, row):
      row += 1
      return ['   ' + self.name, 
        "=SUM(B{0}:B{1})".format(row+1, row+len(self.data)),
        "=SUM(C{0}:C{1})".format(row+1, row+len(self.data)),
        '',
        "=IFERROR(C{0}/B{0},0)".format(row)]       
    
class Item:
    __slots__ = ['aid', 'username', 'data']
    
    def __init__(self):
        self.userid = ''
        self.username = ''
        self.data = []
        
    def getData(self, index, row):
      row += 1
      return [self.username, 
        "=AVERAGE({0})".format(self.getRowItems('B', row)),
        "=AVERAGE({0})".format(self.getRowItems('C', row)),
        '',
        "=IFERROR(C{0}/B{0},0)".format(row)]
      
    def getRowItems(self, column, row):
      row += 1
      res = column + str(row)
      
      for d in range(0, len(self.data)-1):
        res += ','
        row += len(self.data[d].data) + 1
        res += column + str(row)
        
      return res  
      
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
        self.data = list()
        
    def getResult(self, index, row):
      row += 1
      return['Итог', 
        "=AVERAGE({0})".format(self.getRowItems('B', index)),
        "=AVERAGE({0})".format(self.getRowItems('C', index)),
        '',
        "=IFERROR(C{0}/B{0},0)".format(row)]
      
    def getRowItems(self, column, rows):
      res = ''
      
      for r in rows.split(','):
        if len(res) > 0:
          res += ','

        res += column + r
        
      return res    

def loadData(params, server):
    agents = server.Get("Agents", "", "id")
    data = ReportData()
    
    
    for aid in params.userids.split(','):
      item = Item()
      item.aid = aid
      
      where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
            aid,
            params.start.strftime("%d/%m/%Y 0:0:0"),
            params.finish.strftime("%d/%m/%Y 23:59:59"))
      
      date = params.start
      docs = server.Get("OrgRemnants", where)
      strCause = server.Get("StringCause", '"remn"=0 or "remn"!=0',"id")
      
      remn = dict()
      remn_cnt = dict()
      causes = dict()
      
      for d in docs:
        if not d.id in remn:
          remn[d.id] = dict()
        
        if not d.id in causes:
          causes[d.id] = dict()
          
        if not d.id in remn_cnt:
          remn_cnt[d.id] = 1
        else:
          remn_cnt[d.id] += 1
          
        prd = remn[d.id]
        crd = causes[d.id]
        
        for i in d.items:
          if not i.id in prd:
            prd[i.id] = list()
        
          prd[i.id].append(i.qty)   
          
          if len(i.cause) > 0:
            crd[i.id] = i.cause
      
      server.ChangeUser("'" + aid + "'")
      orgs = server.Get("Org", "", "id")
      prop = server.Get("OrgProp", "", "id")
      matrix = server.Get("Matrix", "")
      price = server.Get("Price", "", "id")
      item.username = server.CurrentUser().name
      server.RestoreUser()

      md = dict()
      for m in matrix:
        md[m.name] = m
      
      for oid in orgs:
        if not oid in remn:
          continue
        
        if oid in prop and len(prop[oid].matrix) > 0 and prop[oid].matrix in md:
          o = OrgItem()
          o.id = oid
          o.name = orgs[oid].name
          
          item.data.append(o)
          
          for i in md[prop[oid].matrix].items:
            d = PriceItem()
            d.id = i.id
            d.name = price[i.id].name if i.id in price else i.id
            d.cause = ""
            
            d.plan = remn_cnt[o.id] if o.id in remn_cnt else 0
              
            if oid in remn and i.id in remn[oid]:
              w = remn[oid][i.id]
              d.fact = nonzero(w)
            
            if oid in causes and i.id in causes[oid] and causes[oid][i.id] in strCause:  
              d.cause = strCause[causes[oid][i.id]].text
              
            o.data.append(d)
            
      if len(item.data) > 0:      
        data.data.append(item)
        
    data.data = sorted(data.data, cmp=lambda lhs, rhs: cmp(lhs.username, rhs.username))    
    
    for d in data.data:
      d.data = sorted(d.data, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))    
      
      for y in d.data:
        y.data = sorted(y.data, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))    
      
    return data
    
def nonzero(list):
  res = 0
  
  for i in list:
    if i > 0:
      res += 1
  
  return res
  
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 1 or column == 2:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)  
      
    elif column == 4:
      cell.style.number_format._set_format_code('0%')
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.show_summary_below = False
    
    c = sheet.cell(row=0,column=0)
    c.value = "Отчет с {0} по {1}".format(params.start.strftime("%d.%m.%Y"), params.finish.strftime("%d.%m.%Y"))
    
    xlb = XLBuilderEx()
    row = 1
    
    head = [params.divName, 'План (матрица),  sku', 'Факт (остатки), sku', 'Причина', '%']
    
    xlb.makeHead(sheet, row, head, True)
    
    row += 1
    cnt = 0
    anc = ''
    
    for item in data.data:
      xlb.makeCells(sheet, row, item.getData(cnt, row))
      
      for c in range(0,4):
          sheet.cell(row=row,column = c).style.font.bold = True
          
      row += 1
      cnt += 1
      sheet.row_dimensions[row].outline_level = 0
      
      if len(anc) > 0:
        anc += ','
      
      anc += str(row)
      
      for ii in item.data:
        xlb.makeCells(sheet, row, ii.getData(row))
        
        for c in range(0,4):
          sheet.cell(row=row,column = c).style.font.bold = True
          
        row += 1
        sheet.row_dimensions[row].outline_level = 1
        
        
        
        for xx in ii.data:
          xlb.makeCells(sheet, row, xx.getData(row))
          row += 1
          sheet.row_dimensions[row].outline_level = 2
          
    xlb.makeCells(sheet, row, data.getResult(anc, row))     
    
    for c in range(0,4):
      sheet.cell(row=row,column = c).style.font.bold = True
    
    cc = 1
    SZ = 14
    
    for w in [80,SZ,SZ,SZ]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    return wb
    
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    