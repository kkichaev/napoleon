# -*- coding: cp1251 -*-

import sys;
import locale
import time
import datetime
import logging

from xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat
from time import sleep
from openpyxl.writer.worksheet import row_sort

# reload(sys);
# sys.setdefaultencoding("cp1251")

MAX_COLUMNS = 1500

def putDivisionAgents(server, orgs, d, ret) :
#     print d
    for a in d.agents:
        agent = "'" + a.id + "'"
        if not(agent in ret):
            ret += agent + ","
            refreshOrgs(server, orgs, agent)
    
    return ret
    
def addDivisionAgents(server, orgs, divisions, str, divisionid, parent) :
    ret = str
    for d in divisions:
        if (divisionid != -1 and d.id == divisionid) or (parent != -1 and d.parent == parent):
            agents = putDivisionAgents(server, orgs, d, ret)
            ret = addDivisionAgents(server, orgs, divisions, agents, -1, d.id)

    return ret

def refreshOrgs(server, orgs, userid):
    server.ChangeUser(userid)
    aorgs = server.Get("Org", "", "id")
    server.RestoreUser()
    orgs.update(aorgs)

def loadAgentsData(server, param) :
    ret = ""
    orgs = dict()
    
    if param.division < 0:
        ret = "'" + param.agent + "'"
        refreshOrgs(server, orgs, ret)
    else :
        divisions = server.Get("Division", '')
        ret = addDivisionAgents(server, orgs, divisions, '', param.division, -1)
        ret = ret[:-1]
        
    return ret, orgs

class PriceData:
    __slots__ = ['qty', 'pack']
    
    def __init__(self):
        self.qty = 0
        self.pack = 0
        
    def add(self, orderItem, priceItem):
        self.qty += orderItem.qty
        self.pack += orderItem.qty / priceItem.qtyInPack
        
    def addPriceData(self, priceData):
        if priceData != None:
            self.qty += priceData.qty
            self.pack += priceData.pack
        
class DocData:
    __slots_= ['priceData', 'filter', 'price']
    
    def __init__(self, price, filter):
        self.priceData = dict()
        self.filter = filter
        self.price = price
        
    def add(self, doc, priceTotal):
        added = False
        for item in doc.items:
            if not item.id in self.price: continue

            priceItem = self.price[item.id]
            if len(self.filter) > 0 and priceItem.thermalState != self.filter: continue
            
            pd = None 
            if item.id in self.priceData: pd = self.priceData[item.id]
            else :
                pd = PriceData()
                self.priceData[item.id] = pd
                
            pd.add(item, priceItem)
            
            if item.id in priceTotal: pd = priceTotal[item.id]
            else :
                pd = PriceData()
                priceTotal[item.id] = pd
            pd.add(item, priceItem)
            added = True
            
        return added
 
class OrgKey:
    __slots__ = ['orgs', 'fullName']
    
    def __init__(self, orgs, fullName): 
        self.orgs = orgs
        self.fullName = fullName
    
    def key(self, doc): return doc.id
    
    def name(self, id):
        if not id in self.orgs: return 'Контрагент с кодом <' + id + '>'
        
        org = self.orgs[id]
        ret = org.name
        if self.fullName: ret += ' ' + org.address 
        return ret
    
class AgentKey:
    __slots__ = ['agents']     
    
    def __init__(self, agents):
        self.agents = agents
        
    def key(self, doc): return doc.userid

    def name(self, id):
        return id
#         if not id in self.agents: return 'Торг.агент с кодом <' + id + '>'        
#         return self.agents[id].name

class ReportData:
    __slots__ = ['data', 'keys', 'priceTotal']
    
    def __init__(self, keys):
        self.data = dict()
        self.keys = keys
        self.priceTotal = dict()
        
    def add(self, docs, price, filter):
        rootKey = self.keys[0]
        descKey = self.keys[1]
        
        for doc in docs:
            rk = rootKey.key(doc)
            dk = descKey.key(doc)

            group1Data = None
            if rk in self.data: group1Data = self.data[rk]
            else :
                group1Data = dict()
                self.data[rk] = group1Data
                
            docData = None
            if dk in group1Data: docData = group1Data[dk]
            else :
                docData = DocData(price, filter)
                
            if docData.add(doc, self.priceTotal) :
                group1Data[dk] = docData
            
                
def loadData(server, param):
    uids, orgs = loadAgentsData(server, param)
    
    price = server.Get("ManagerPrice", '', 'id')
    folders = server.Get("ManagerFolder", '', 'id')
    agents = server.Get("Agents", '', 'id')
    
    where = '"userid" in ({0}) and "{3}" >= ToDate("{1}") and "{3}" <= ToDate("{2}")'.format(
        uids,
        param.start.strftime("%d/%m/%Y 0:0:0"),
        param.end.strftime("%d/%m/%Y 23:59:59"),
        'date' if param.docsFromCreated == 0 else 'created')
    
    if param.firm != '' :
        where += ' and "firmCode"=' + "'" + param.firm + "'"
    
    docs = server.Get("Order", where)
    
    agentKey = AgentKey(agents)
    orgKey = OrgKey(orgs, param.fullOrgName == 1)
    
    keys = (agentKey, orgKey) if param.groupBy == 1 else (orgKey, agentKey)
    
    data = ReportData(keys) 
    if docs != None :
        data.add(docs, price, param.thState)
        
    return data, price, folders

class HeadBuilder(XLBuilder):
    
    __slots__ = ['sheet']
    
    def __init__(self):
        self.sheet = None
    
    def makeLabelCell(self, row, col, title):
        cell = self.sheet.cell(row=row, column=col)
        cell.value = title
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    
    def makeHeadRow(self, sheet, data, columnOrder, row, usePack, useQty, onlyTotal, groupByOrg):
        self.sheet = sheet
        self.makeHead(sheet, row, ['Артикул','Базовая номенклатура','Вид упаковки','Терм.состояние'])
        
        cc = 0
        while cc < 4:
            sheet.merge_cells(start_row=row, start_column = cc, end_row = row+2, end_column = cc)
            cell = sheet.cell(row=row+2, column=cc)
            borders = cell.style.borders
            borders.bottom.border_style = XLBuilder.HEAD_BORDER_STYLE 
            cc += 1
            
        startCol = 4
        col = 4;
        
        if onlyTotal == False:
            col = startCol 
            for key in columnOrder :
                v = data.data[key.key]
                fc = col
                cell = sheet.cell(row=row, column=col)
                cell.value = key.value
                if groupByOrg: cell.style.alignment.wrap_text = True
                cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                
                for dkey in key.data: 
                    iv = v[dkey.key]
                    sc = col
                    cell = sheet.cell(row=row+1, column=col)
                    cell.value = dkey.value
                    if groupByOrg: cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                    else: cell.style.alignment.wrap_text = True
                    
                    if useQty :
                        self.makeLabelCell(row+2, col, 'кг')
                        col += 1                        
                    if usePack :
                        self.makeLabelCell(row+2, col, 'ящ')
                        col += 1
                        
                    if useQty and usePack:
                        sheet.merge_cells(start_row=row+1, start_column = sc, end_row = row+1, end_column = col-1)
                if col - fc > 1:
                    sheet.merge_cells(start_row=row, start_column = fc, end_row = row, end_column = col-1)

        cell = sheet.cell(row=row, column=col)
        cell.value = "Итого"
        cell.style.font.bold = True
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
        
        sc = col
        if useQty :
            self.makeLabelCell(row+2, col, 'кг')
            col += 1                        
        if usePack :
            self.makeLabelCell(row+2, col, 'ящ')
            col += 1
        sheet.merge_cells(start_row=row, start_column = sc, end_row = row+1, end_column = col-1)
        
        c=2
        cr = row if groupByOrg else row+1
        while c<col:
            sheet.cell(row=cr, column=c).style.alignment.text_rotation = 90
            c += 1
            
        sheet.row_dimensions[cr+1].height = 120
        
        return row + 3
        

def makeHead(sheet, param, data, columnOrder, groupByOrg):
    row = 0
    c = sheet.cell(row=row, column=1)
    c.value = 'Агент: ' + param.name if param.division < 0 else 'Подразделение: ' + param.name
    row += 1

    filter = 'Период с {0} по {1}'.format(param.start.strftime("%d.%m.%Y"), param.end.strftime("%d.%m.%Y"))
    
    if len(param.firm) > 0: filter += ', фабрика:' + param.firmName
    if len(param.thState) > 0: filter += ', терм.сост.:' + param.thState
    c = sheet.cell(row=row, column=1)
    c.value = filter
    row += 1
    
    hb = HeadBuilder()
    
    return hb.makeHeadRow(sheet, data, columnOrder, row, param.usePack == 1, param.useQty == 1, param.totals == 1, groupByOrg)

class KeyValuePair:
    __slots_ = ['key', 'value', 'data']
    
    def __init__(self, key, value):
        self.key = key
        self.value = value
        self.data = list()

def makeSortOrder(data, price, folders):
    corder = list()
    dataHolder = dict()
    
    priceIds = list()
    
    for k,v in data.data.items():
        kv = KeyValuePair(k, data.keys[0].name(k))
        corder.append(kv)
        for ik, val in v.items():
            kv.data.append(KeyValuePair(ik, data.keys[1].name(ik)))
            
            for pkey, pval in val.priceData.items():
                if pkey in priceIds : continue
                priceIds.append(pkey)
                
                pitem = price[pkey]
                
                pholder = None
                if pitem.fid in dataHolder: pholder = dataHolder[pitem.fid]
                else:
                    pholder = list()
                    dataHolder[pitem.fid] = pholder
                kvp = KeyValuePair(pitem.id, pitem.name)
                kvp.data.append(pitem)
                pholder.append(kvp)            
            
        kv.data.sort(key = lambda x: x.value)
        
    corder.sort(key = lambda x: x.value)
    
    rorder = list()
    for k, v in dataHolder.items():
        fName = folders[k].name if k in folders else 'Папка с кодом <' + k + '>'
        fdata = KeyValuePair(k, fName)
        fdata.data = v
        fdata.data.sort(key = lambda x: x.value)
        rorder.append(fdata)
        
    rorder.sort(key = lambda x: x.value)
    
    return corder, rorder

def stripLeadZeros(id):
    while id[0] == '0' :
        id = id[1:]
    return id

def makeRows(sheet, usePack, useQty, onlyTotals, data, rowOrder, columnOrder, row):
    values = []
    orgTotals = dict()
    
    overflow = False
    for fdata in rowOrder:
        cell = sheet.cell(row=row, column=0)
        cell.value = fdata.value
        cell.style.font.bold = True
        row += 1
        overflow = False

        xlb = XLBuilder()
        for pdata in fdata.data:
            pitem = pdata.data[0]
            values = [stripLeadZeros(pitem.id), pdata.value, pitem.packName, pitem.thermalState]
            
            if not onlyTotals:
                for k1 in columnOrder :
                    rootData = data.data[k1.key]
                    
                    for k2 in k1.data:
                        docData = rootData[k2.key]
                        itemData = docData.priceData[pdata.key] if pdata.key in docData.priceData else None
    
                        if useQty :
                            values.append(itemData.qty if itemData != None else '')
                        if usePack :
                            values.append(itemData.pack if itemData != None else '')
                            
                        ot = orgTotals[k1.key] if k1.key in orgTotals else PriceData()
                        ot.addPriceData(itemData)
                        orgTotals[k1.key] = ot
                        
            itemData = data.priceTotal[pdata.key] if pdata.key in data.priceTotal else None
            if useQty :
                values.append(itemData.qty if itemData != None else '')
            if usePack :
                values.append(itemData.pack if itemData != None else '')
                    
            xlb.makeCells(sheet, row, values, Border.BORDER_NONE)
            row += 1

            if len(values) > MAX_COLUMNS:
                logging.debug("values count %d", len(values))
                c = sheet.cell(row=row, column=0)
                c.value = 'Превышено число допустимых колонок в отчете ' + str(len(values))
                row += 1
                
                overflow = True
                break
        if overflow: break
    
    if not overflow and not onlyTotals:
        values = ['Итого', '', '', '']
        for k1 in columnOrder :
            ot = orgTotals[k1.key] if k1.key in orgTotals else PriceData()
            if useQty: values.append(ot.qty)
            if usePack: values.append(ot.pack)
        cc = 0
        for title in values:
            c = sheet.cell(row=row, column=cc)
            c.value = title
            style = c.style
            style.font.bold = True
            style.alignment.wrap_text = True
            cc += 1
        
    
    cc = 1
    wdh = [11,45,11,11]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    w = 6 if useQty and usePack else 12
    while cc < len(values) + 1:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
    
def printOut(data, param, price, folders):
    xlb = XLBuilder()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"

    
    columnOrder, rowOrder = makeSortOrder(data, price, folders)
    row = makeHead(sheet, param, data, columnOrder, param.groupBy == 0)
    
    logging.info("header done")
    
    makeRows(sheet, param.usePack == 1, param.useQty == 1, param.totals == 1, data, rowOrder, columnOrder, row)
    
    sheet.freeze_panes = "E6"
    return wb

def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting userid")

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))
    
    data, price, folders = loadData(server, params)
    logging.info("data loaded")
    
    wb = printOut(data, params, price, folders)

    XLBuilder().workbookToObject(wb, "agentorder.xlsx", server)                
    logging.info("ended")
