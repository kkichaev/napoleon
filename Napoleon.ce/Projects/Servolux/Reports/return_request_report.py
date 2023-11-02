# -*- coding: cp1251 -*-

import sys;
import locale
import time
import datetime

from xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat
from time import sleep
from openpyxl.writer.worksheet import row_sort

# reload(sys);
# sys.setdefaultencoding("cp1251")

class RepData :
    __slots__ = ['agent', 'org', 'address', 'item', 'qty', 'sum', 'cause', 'svCause', 'mfrDate', 'havePhoto', 'dlvNum', 'dlvDate', 'created', 'svid', 'remark', 'party']
    
    def load(self, doc, item, dlvitem, price, orgs, visit, cause):
        self.agent = doc.userid
        if doc.id in orgs:
            o = orgs[doc.id]
            self.org = o.name
            self.address = o.address
        else:
            self.org = "<Код + '" + doc.id + "'>"
            self.address = ''
        if item.id in price:
            p = price[item.id]
            self.item = p.name + " " + p.thermalState + "/" + p.packName
        else:
            self.item = "<Код + '" + item.id + "'>"
            
        self.qty = dlvitem.qty
        self.sum = dlvitem.qty * dlvitem.cost
        self.cause = cause[item.cause].name if item.cause in cause else ""
        self.svCause = cause[item.svCause].name if item.svCause in cause else ""
        self.mfrDate = item.mfrDate
        self.dlvNum = dlvitem.number
        self.dlvDate = dlvitem.date
        self.created = doc.created
        self.svid = doc.svid
        self.remark = dlvitem.remark
        self.party = dlvitem.party
        
        self.havePhoto = ''
        
        if visit != None and len(visit) > 0:
            for vi in visit[0].items:
                if vi.tag == item.id:
                    self.havePhoto = '+'
                    break
        
        
 
def printOut(params, data):
    xlb = XLBuilder()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"

    row = 0
    c = sheet.cell(row=row, column=0)
    c.value = 'Отчет по заявкам на возвраты с {0} по {1}'.format(
        params.start.strftime("%d.%m.%Y"),
        params.end.strftime("%d.%m.%Y"),)
    
    
    row += 1
    heads = ["№ пп", "Код ТП", "Контрагент", 'адрес доставки', 'наименование продукции', 'кол-во', 'сумма Б/НДС', 'причина возврата', 
             'согласованная причина возврата', 'партия', 'наличие фотоотчета', '№ТТН', 'дата ТТН', 'дата СЗ', 'Время создания', 
             'Статус координатор', 'Комментарий координатора'] 
    xlb.makeHead(sheet, row, heads, True)
    row += 1

    data.sort(key=lambda x: x.agent + '|' + x.org + '|' + x.address + '|' + x.item)

    for d in data:
        vals = [row-1, d.agent, d.org, d.address, d.item, d.qty, d.sum, d.cause, d.svCause, 
                d.party, #d.mfrDate.strftime("%d.%m.%Y"), 
                d.havePhoto, d.dlvNum, 
                d.dlvDate.strftime("%d.%m.%Y"), 
                d.created.strftime("%d.%m.%Y"),
                d.created.strftime("%H:%M"), d.svid, d.remark]
        xlb.makeCells(sheet, row, vals) 
        row += 1

    cc = 1
    wdh = [5,10,25,55,45,10,10, 20, 20,13,5,20,13,13,10,20,30]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    return wb
 
def loadData(server, params):
    
    price = server.Get("ManagerPrice", '', 'id')
    orgs = server.Get("CommonOrgs", '', 'id')
    cause = server.Get("ReturnCause", '', 'id')
    
    where = '"{2}" >= ToDate("{0}") and "{2}" <= ToDate("{1}")'.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.end.strftime("%d/%m/%Y 23:59:59"),
        'created')

    data = list()
    
    docs = server.Get("ReturnRequest", where)
    for d in docs:
        vwhere = '"userid"=' + "'{0}' and created=ToDate('{1}')".format(d.userid, d.visitDoc.strftime("%d/%m/%Y %H:%M:%S"))
        v = server.Get("Visit", vwhere)
        for item in d.items:
            for dlvitem in item.items:
                rdata = RepData()
                rdata.load(d, item, dlvitem, price, orgs, v, cause)
                data.append(rdata)
                
    return data 
    

def run(server):
    repName = 'return_request_report'
    print (repName + " start " + str(datetime.datetime.now()) + " userid: " + server.CurrentUser().id )

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    print (repName + " params " + str(params)) 

    data = loadData(server, params)
    wb = printOut(params, data)

    XLBuilder().workbookToObject(wb, "rr_repoprt.xlsx", server)                
    print (repName + " end "  + str(datetime.datetime.now()))
