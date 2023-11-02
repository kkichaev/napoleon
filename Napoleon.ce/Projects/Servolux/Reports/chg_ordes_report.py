# -*- coding: cp1251 -*-

import sys;
import locale
import time
import datetime

from xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat


class ServerReportData:
    price = None
    orgs = None
    uidwhere = ""
    
    def __init__(self, server, param):
        self.price = server.Get("ManagerPrice", "", "id")
        
        self.orgs = dict()
        
        self.uidwhere = ""
        
        for uid in param.users:
            userid = "'" + uid.id + "'"
            self.uidwhere += userid + ","
            
            server.ChangeUser(userid)
            uorgs = server.Get("Org", "", "id")        
            server.RestoreUser()
            self.orgs.update(uorgs)
        

class ItemData:
    __slots__ = ["name", "qty", "agentQty", "weight", "agentWeight"]
    
    def __init__(self):
        self.name = ""
        self.qty = 0
        self.agentQty = 0
        self.weight = 0
        self.agentWeight = 0

class EmptyOrg :
    __slots__ = ['name', 'address']
    
    def __init__(self, name):
        self.name = name
        self.address = ''

class OrgData:
    org = ""
    address = ""
    items = None
    reportData = None
    
    def __init__(self, order, reportData):
        self.items = dict()

        if not order.id in reportData.orgs:
            o = EmptyOrg("Контрагент с кодом <" + order.id + ">")            
            reportData.orgs[order.id] = o
            
        co = reportData.orgs[order.id]
        self.org = co.name
        self.address = co.address
        self.reportData = reportData

    def updateItems(self, order, prices):
        res = False
        for i in order.items:
            if prices != None and i.id not in prices: continue
                        
            id = None
#             if i.agentQty == 0 or i.agentQty == i.qty: continue
            
            inPack = self.reportData.price[i.id].qtyInPack if i.id in self.reportData.price else 1
            packQty = (i.qty / inPack)
            agentQty = (i.agentQty / inPack) if i.agentQty != 0 else packQty
            
            res = True
            if i.id in self.items:
                id = self.items[i.id]
            else:
                id = ItemData()
                if i.id in self.reportData.price:
                    prc = self.reportData.price[i.id]
                    id.name = prc.name + " " + prc.thermalState + "/" + prc.packName
                else:
                    id.name = i.id
                self.items[i.id] = id 
                
            id.agentQty += agentQty
            id.qty += packQty
            id.weight += i.qty
            id.agentWeight += i.agentQty if i.agentQty != 0 else i.qty
            
        return res

        
    def update(self, order, prices):
        return self.updateItems(order, prices)
                        
#     def updateItems(self, order, isAgentOrder):
#         for i in order.items:            
#             id = None
#             inPack = price[i.id].qtyInPack if i.id in price else 1
#             if i.id in self.items:
#                 id = self.items[i.id]
#             else:
#                 id = ItemData()
#                 if i.id in price:
#                     prc = price[i.id]
#                     id.name = prc.name + " " + prc.thermalState + "/" + prc.packName
#                 else:
#                     id.name = i.id
#                 self.items[i.id] = id 
#                 
#             packQty = (i.qty / inPack)
#             if isAgentOrder:
#                 id.agentQty += packQty
#             else:
#                 id.qty += packQty
# 
#         
#     def update(self, order, chOrder):
#         self.updateItems(order, False)
#         self.updateItems(chOrder if chOrder != None else order, True)
                

def findAgentsOrder(order, chOrders):
    if chOrders != None:
        for co in chOrders:
            if co.id == order.id and co.created == order.created:
                return co
    return None    

def HasItem(order, items):
    for item in order.items:
#         if item.agentQty != 0 and item.agentQty != item.qty :
        if items == None or item.id in items: 
            return True
        
    return False
    
def loadData(server, param):
    reportData = ServerReportData(server, param)
    
#     global price, orgs
#     
#     price = server.Get("ManagerPrice", "", "id")
#     
#     uidwhere = ""
#     
#     for uid in param.users:
#         userid = "'" + uid.id + "'"
#         uidwhere += userid + ","
#         
#         server.ChangeUser(userid)
#         uorgs = server.Get("Org", "", "id")        
#         server.RestoreUser()
#         orgs.update(uorgs)
        
    where = '"userid" in ({0}) and "date" >= ToDate("{1}") and "date" <= ToDate("{2}")'.format(
        reportData.uidwhere[:-1],
        param.begin.strftime("%d/%m/%Y 0:0:0"),
        param.end.strftime("%d/%m/%Y 23:59:59"))
#     where = '"userid" in ({0}) and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
#         reportData.uidwhere[:-1],
#         param.begin.strftime("%d/%m/%Y 0:0:0"),
#         param.end.strftime("%d/%m/%Y 23:59:59"))
    
    orders = server.Get("SVOrderChanges", where) # + ' and "created" <> "modify"')
#     agentsOrders = server.Get("AgentsOrders", where)
    
    firms = set()
    for uid in param.factories:
        firms.add(uid.id)
    prices = None
    if len(param.sku) > 0 :
        prices = set()
        for uid in param.sku:
            prices.add(uid.id)
    
    orgset = None
    if len(param.orgs) > 0:
        orgset = set()
        for uid in param.orgs:
            orgset.add(uid.id)
    
    pdata = dict()
    if orders != None:
        for doc in orders:
            if orgset != None and not (doc.id in orgset): continue
            if doc.firmCode not in firms: continue
            if not HasItem(doc, prices) : continue

            cdata = None
            if doc.id in pdata:
                cdata = pdata[doc.id]
            else:
                cdata = OrgData(doc, reportData)
                
                # remove undef orgs
                if cdata.org != "" :
                    pdata[doc.id] = cdata
                
#             cdata.update(doc, findAgentsOrder(doc, agentsOrders))
            cdata.update(doc, prices)
        
    ret = []
    for doc in pdata.values():
        ret.append(doc)

    locale.setlocale(locale.LC_ALL, 'russian_russia')
    ret.sort(key=lambda x: (x.org.upper(), x.address.upper()))
        # cmp=lambda x, y:
        #     cmp(x.org.upper(), y.org.upper()) if x.org != y.org else
        #     cmp(x.address.upper(), y.address.upper()) if x.address != y.address else 0 )        
    
    return ret

def printOut(data, param):
    xlb = XLBuilder()
    
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"

    row = 0
    end_col = 8 if param.inKG != 0 and param.inBox != 0 else 5
    
    
    sheet.merge_cells(start_row=row, start_column = 0, end_row = row, end_column = end_col)
    cell = sheet.cell(row=row, column=0)
    cell.value = "Период c {0} по {1}".format(param.begin.strftime("%d/%m/%Y"), param.end.strftime("%d/%m/%Y"))
    xlb.makeBorder(cell, Border.BORDER_MEDIUM)
    cell = sheet.cell(row=row, column=6)
    cell.style.borders.left.border_style = Border.BORDER_MEDIUM
    row += 1

    sheet.merge_cells(start_row=row, start_column = 0, end_row = row, end_column = end_col)
    cell = sheet.cell(row=row, column=0)
    cell.value = "Отчет по истории подрезки"
    xlb.makeBorder(cell, Border.BORDER_MEDIUM)
    cell = sheet.cell(row=row, column=6)
    cell.style.borders.left.border_style = Border.BORDER_MEDIUM
    row += 1
    
    heads = ["Наименование контрагента", "АдресТО", "Продукция"] 
    if param.inBox != 0 :
        heads += [ "ТП Заказано,ящ", "ТП Подрезано,ящ", "ТП Факт,ящ"]
    if param.inKG != 0 :
        heads += [ "ТП Заказано,кг", "ТП Подрезано,кг", "ТП Факт,кг"]
    
    xlb.makeHead(sheet, row, heads, True)
    row += 1
    
    totQty = 0
    totAgentQty = 0
    totWeight = 0
    totAgentWeight = 0
    
    for doc in data:
        docQty = 0
        docAgentQty = 0
        docWeight = 0
        docAgentWeight = 0
        
        cc = 0
        
        items = []
        items.extend(doc.items.values())
        items.sort(key=lambda x: x.name.upper())
        itemCount = len(items)
        
        if itemCount > 0 :
            sheet.merge_cells(start_row=row, start_column = cc, end_row = row + itemCount, end_column = cc)
        cell = sheet.cell(row=row, column=cc)
        cell.value = doc.org
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
        cell.style.alignment.wrap_text = True
        xlb.makeBorder(cell, Border.BORDER_THIN)
        cc += 1
        
        if itemCount > 0 :
            sheet.merge_cells(start_row=row, start_column = cc, end_row = row + itemCount - 1, end_column = cc)
        cell = sheet.cell(row=row, column=cc)
        cell.value = doc.address
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
        cell.style.alignment.wrap_text = True        
        xlb.makeBorder(cell, Border.BORDER_THIN)
        xlb.makeBorder(sheet.cell(row=row+itemCount, column=cc), Border.BORDER_THIN)
        cc += 1
    
        startcc = cc
        for item in items:
            cc = startcc
            cell = sheet.cell(row=row, column=cc)
            cell.value = item.name
            cell.style.alignment.wrap_text = True
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1
            
            if param.inBox != 0 :
                cell = sheet.cell(row=row, column=cc)
                cell.value = item.agentQty
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
                xlb.makeBorder(cell, Border.BORDER_THIN)
                cc += 1
    
                cell = sheet.cell(row=row, column=cc)
                cell.value = item.agentQty - item.qty
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
                xlb.makeBorder(cell, Border.BORDER_THIN)
                cc += 1
    
                cell = sheet.cell(row=row, column=cc)
                cell.value = item.qty
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
                xlb.makeBorder(cell, Border.BORDER_THIN)
                cc += 1

            if param.inKG != 0 :
                cell = sheet.cell(row=row, column=cc)
                cell.value = item.agentWeight
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
                xlb.makeBorder(cell, Border.BORDER_THIN)
                cc += 1
    
                cell = sheet.cell(row=row, column=cc)
                cell.value = item.agentWeight - item.weight
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
                xlb.makeBorder(cell, Border.BORDER_THIN)
                cc += 1
    
                cell = sheet.cell(row=row, column=cc)
                cell.value = item.weight
                cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
                xlb.makeBorder(cell, Border.BORDER_THIN)
                cc += 1

            docAgentQty += item.agentQty
            docQty += item.qty
            docAgentWeight += item.agentWeight
            docWeight += item.weight
            row += 1

        cc = startcc
        cell = sheet.cell(row=row, column=cc)
        cell.value = "Итого"
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
        cell.style.font.bold = True
        xlb.makeBorder(cell, Border.BORDER_THIN)
        cc += 1
        
        if param.inBox != 0 :
            cell = sheet.cell(row=row, column=cc)
            cell.value = docAgentQty
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            cell.style.font.bold = True
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1
    
            cell = sheet.cell(row=row, column=cc)
            cell.value = docAgentQty - docQty
            cell.style.font.bold = True
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1
    
            cell = sheet.cell(row=row, column=cc)
            cell.value = docQty
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            cell.style.font.bold = True
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1

        if param.inKG != 0 :
            cell = sheet.cell(row=row, column=cc)
            cell.value = docAgentWeight
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            cell.style.font.bold = True
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1
    
            cell = sheet.cell(row=row, column=cc)
            cell.value = docAgentWeight - docWeight
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            cell.style.font.bold = True
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1
    
            cell = sheet.cell(row=row, column=cc)
            cell.value = docWeight
            cell.style.font.bold = True
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
            xlb.makeBorder(cell, Border.BORDER_THIN)
            cc += 1
        
        row += 1
        
#         sheet.merge_cells(start_row=row, start_column = 0, end_row = row, end_column = 5)
#         cell = sheet.cell(row=row, column=0)
#         cell.value = ""
#         xlb.makeBorder(cell, Border.BORDER_THIN)
#         xlb.setBackColor(cell, Color.YELLOW)
#         sheet.cell(row=row, column=6).style.borders.left.border_style = Border.BORDER_THIN         
#         row += 1

        totQty += docQty
        totAgentQty += docAgentQty
        totWeight += docWeight
        totAgentWeight += docAgentWeight
        
    cc = 0
   
    cellData = ["","", "Итого"]
    if param.inBox != 0 :
        cellData += [totAgentQty, totAgentQty - totQty, totQty]
    if param.inKG != 0 :
        cellData += [totAgentWeight, totAgentWeight - totWeight, totWeight]
        
    for val in cellData:
        cell = sheet.cell(row=row, column=cc)
        cell.value = val
        cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
        xlb.makeBorder(cell, Border.BORDER_THIN)
        xlb.setBackColor(cell, 'FFA500')
        cc += 1
        
    cc = 1
    wdh = [20,35,25,15,15,15]
    if param.inBox !=0 and param.inKG != 0 : wdh += [15,15,15]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    return wb
    
def run(server):
    print( "ch_orders_report start "  + str(datetime.datetime.now()) + " userid: " + server.CurrentUser().id)

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    print ("ch_orders_report params " + str(params))
    
    data = loadData(server, params)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "chgrep.xlsx", server)                
    
    print ("ch_orders_report end "  + str(datetime.datetime.now()))
