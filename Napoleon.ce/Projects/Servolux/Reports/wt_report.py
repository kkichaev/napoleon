# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import time

from xl_base import XLBuilder

# reload(sys);
# sys.setdefaultencoding("cp1251")


class RepData:
    agent = ""
    org = ""
    date = ""
    start = ""
    end = ""
    span = ""
    tdelta = 0
    
    def __init__(self):
        pass
               
def loadData(server, param):
    
    ret = []
    
    userid = "'" + param.userid + "'"
    
    agents = server.Get("Agents", "", "id")
    agent = agents[param.userid].name if param.userid in agents else ""
    
    server.ChangeUser(userid)
    orgs = server.Get("Org", "", "id")
    server.RestoreUser()
    
    where = '"userid" = {0} and "start" >= ToDate("{1}") and "start" <= ToDate("{2}")'.format(
        userid,
        param.begin.strftime("%d/%m/%Y 0:0:0"),
        param.end.strftime("%d/%m/%Y 23:59:59"))
    
    wtObj = server.Get("WorkTime", where)
    if wtObj != None :
        for wt in wtObj:
            if wt.id in orgs:
                rd = RepData()
                rd.agent = agent
                rd.org = orgs[wt.id].address
                rd.date = wt.start.strftime("%d/%m/%Y")
                rd.start = wt.start.strftime("%H:%M")
                rd.end = wt.stop.strftime("%H:%M")
                
                rd.tdelta = wt.stop - wt.start
                sec = rd.tdelta.total_seconds()
                rd.span = "{0:02d}:{1:02d}".format(int(sec / 3600), int((sec % 3600) / 60)) 
                
                ret.append(rd)
     
    return ret

def printOut(data):
    wb = Workbook(False, 'cp1251')
    
    xlb = XLBuilder()

    sheet = wb.get_active_sheet()
    sheet.title = "отчет по визитам"
    
    row = 0
    
    sheet.merge_cells(start_row=row, start_column = 0, end_row = row, end_column = 4) 
    c = sheet.cell(row=row, column=0)
    c.value = "Отчет по визитам"
    xlb.makeBorder(c, Border.BORDER_MEDIUM)
    
    style = c.style
    style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    style.alignment.vertical = Alignment.VERTICAL_CENTER
    style.font.bold = True

    c = sheet.cell(row=row, column=5)
    c.style.borders.left.border_style = Border.BORDER_MEDIUM 
    sheet.row_dimensions[row+1].height = 30

    row+=1
    xlb.makeHead(sheet, row, ["Адрес торгового объекта", "Дата визита", "Время начала визита", "Время окончания визита", "Общее время нахождения на визите"])
    sheet.row_dimensions[row+1].height = 30
 
    totalDelta = 0
    for item in data:
        row += 1
        xlb.makeCells(sheet, row, [item.org, item.date, item.start, item.end, item.span])
        sheet.row_dimensions[row+1].height = 25
        totalDelta += item.tdelta.total_seconds()

    cc = 1
    for w in [80,20,25,25,35]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    row += 1
    sheet.merge_cells(start_row=row, start_column = 0, end_row = row, end_column = 3) 
    c = sheet.cell(row=row, column=0)
    c.value = "ИТОГО: общее время нахождения на визитах"
    xlb.makeBorder(c, Border.BORDER_MEDIUM)
    c.style.font.bold = True

    c = sheet.cell(row=row, column=4)
    c.value = "{0:02d}:{1:02d}".format(int(totalDelta / 3600), int((totalDelta % 3600) / 60))
    c.style.font.bold = True
    xlb.makeBorder(c, Border.BORDER_MEDIUM)
    sheet.row_dimensions[row+1].height = 30
    
    return wb    
               
def run(server):
    repName = 'wt_report'
    print (repName + " start " + str(datetime.now())  + " userid: " + server.CurrentUser().id)
    
    #getcontext().prec = 2 
    #getcontext().rounding = ROUND_05UP
    
    params = server.Params[0]
    print (repName + " params " + str(params) )
    
    data = loadData(server, params)
    wb = printOut(data)

    XLBuilder().workbookToObject(wb, "res.xlsx", server)                
    print (repName + " end "  + str(datetime.now()))
    