# -*- coding: cp1251 -*-

from importlib import reload
import sys;
import logging

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from openpyxl.style import NumberFormat
from rmr_report_style import XLBuilderCommon

reload(sys);

def inflateParams(server):
    return server.Params[0].userids
        
def loadData(server):
    userid = inflateParams(server)
    
    where = '"userid" in({0})'.format(userid)
    fridge = server.Get("InvFrg", where)
    fridge = sorted(fridge, key= lambda x: x.created)
    
    report = Report()
    
    ids = list()
    userorgs = dict()
    rd = list()
    
    for r in fridge:
        if not r.id in ids:
            rd.append(r)
            ids.append(r.id)
    
    for r in rd:
        if not r.userid in userorgs:
            userorgs[r.userid] = dict()
            
            server.ChangeUser("'" + r.userid + "'")
            orgs = server.Get("Org", "", "id")
            
            for id in orgs:
                userorgs[r.userid][id] = orgs[id]
            
            porg = server.Get("PotenzialOrg", "", "id")
            
            for id in porg:
                userorgs[r.userid][id] = porg[id]
            
        
        item = ReportItem()
        item.org = userorgs[r.userid][r.id].name if r.id in userorgs[r.userid] else ""
        item.address = userorgs[r.userid][r.id].address if r.id in userorgs[r.userid] else ""
        item.date = r.created.strftime('%d/%m/%Y')
        
        for n in r.items:
            if n.prez == 1:
                item.numbers.append(' ' + n.number)
        
        item.cnt = len(item.numbers)
        report.items.append(item)        
    
    return report 

def printOut(xlb, rpd):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    n = 0
    for i in rpd.items:
        sz = len(i.numbers)
        if n < sz:
            n = sz
    
    head = ['Контрагент', 'Адрес', 'Кол-во Холодильников Факт', 'Дата']
    
    for s in range(0,n):
        head.append('ИН'+str(s+1))
    
    xlb.makeHead(sheet, 0, head)
    
    row = 1 
    
    for i in rpd.items:
        data = i.getData()
        
        if len(data) < len(head):
            for s in range(len(data), len(head)):
                data.append(" ")
        xlb.makeCells(sheet, row, data)
        row += 1

    cc = 1
    for w in [25,50,25,20]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    return wb

class Report:
    items = None
    
    def __init__(self):
        self.items = list()
        
class ReportItem:
    org = None
    address = None
    cnt = None
    numbers = None
    date = None
    
    def __init__(self):
        self.org = ""
        self.address = ""
        self.cnt = 0
        self.numbers = list()  
        self.date = ""
        
    def getData(self): 
        data = [self.org, self.address, self.cnt, self.date]
        data.extend(self.numbers)
        return data         
          
class XLBuilderEx(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
        XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
        if column > 3:
            cell.style.number_format._set_format_code(NumberFormat.FORMAT_TEXT)         
        
def doReport(server):
    report = loadData(server)
    xlb = XLBuilderEx()
    wb = printOut(xlb, report)
    XLBuilderCommon().workbookToObject(wb, "fridle_prez_rpt.xlsx", server) 
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    doReport(server)
    
    logging.info('end')
    