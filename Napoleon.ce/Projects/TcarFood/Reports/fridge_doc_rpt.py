# -*- coding: cp1251 -*-

from importlib import reload
import sys;
import logging

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from rmr_report_style import XLBuilderCommon

reload(sys);

def inflateParams(server):
    return server.Params[0].userids
        
def loadData(server):
    userid = inflateParams(server)
    
    where = '"userid" in({0})'.format(userid)
    fridge = server.Get("InvFrg", where)
    fridge = sorted(fridge, key= lambda x: x.created)
    
    report = Report()
    
    ids = list()
    userorgs = dict()
    rd = list()
    
    for r in fridge:
        if not r.id in ids:
            rd.append(r)
            ids.append(r.id)
    
    for r in rd:
        if not r.userid in userorgs:
            userorgs[r.userid] = dict()
            
            server.ChangeUser("'" + r.userid + "'")
            orgs = server.Get("Org", "", "id")
            
            for id in orgs:
                userorgs[r.userid][id] = orgs[id]
            
            porg = server.Get("PotenzialOrg", "", "id")
            
            for id in porg:
                userorgs[r.userid][id] = porg[id]
            
        
        item = ReportItem()
        item.org = userorgs[r.userid][r.id].name if r.id in userorgs[r.userid] else ""
        item.address = userorgs[r.userid][r.id].address  if r.id in userorgs[r.userid] else ""
        item.data = r.created
        
        report.items.append(item)        
    
    return report 
    
def printOut(xlb, rpd):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    head = ['Контрагент', 'Адрес', 'Дата последнего документа ХО']
    
    xlb.makeHead(sheet, 0, head)
    
    row = 1 
    
    now = datetime.now()
    mc = now.year*12 + now.month
    
    for i in rpd.items:
        data = i.getData()
        xlb.makeCells(sheet, row, data)
        row += 1
        if abs(mc - (i.data.year * 12 + i.data.month)) >= 6:
            for c in range(0, len(data)):
                cell = sheet.cell(row=row, column=c)
                xlb.setBackColor(cell, "ff0000")
    cc = 1
    for w in [25,55,55]:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    return wb

class Report:
    items = None
    
    def __init__(self):
        self.items = list()
        
class ReportItem:
    org = None
    address = None
    data = None
    
    def __init__(self):
        self.org = ""
        self.address = ""
        self.data = 0
        
    def getData(self): 
        data = [self.org, self.address, self.data.strftime("%d.%m.%Y")]
        return data         
          
def doReport(server):
    report = loadData(server)
    xlb = XLBuilder()
    wb = printOut(xlb, report)
    XLBuilderCommon().workbookToObject(wb, "fridle_doc_rpt.xlsx", server) 
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    doReport(server)
    
    logging.info('end')
    