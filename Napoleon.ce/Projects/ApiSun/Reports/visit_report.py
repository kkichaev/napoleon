# -*- coding: cp1251 -*-
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat, Border, Font, DEFAULTS
from datetime import datetime
from grsoft.route import AgentRoute
from datetime import timedelta
from manager import coordutils

import datetime

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

NOT_VISIT_CODE = 1
OUT_ROUT_CODE = 2
IN_ROUT_CODE = 3

class Item:
    __slots__ = ['id', 'name', 'ntz', 'visits']
    
    def __init__(self):
        self.id = ''
        self.name = ''
        self.ntz = 0
        self.visits = []
        
    def getData(self, row):
      res = [self.name, self.ntz]
      
      for i in reversed(self.visits):
        if i == 0:
          res.append(0)
        else:
          res.append(1)
          
        res.append(i)
        
      return res
        
      
class ReportData:
  __slots__ = ['data', 'planogram', 'orgname', 'lpr', 'fio', 'agentname']
  
  def __init__(self):
      self.data = {}
      self.planogram = []
      self.agentname = ''
      
  def toList(self):
    list = self.data.values()
    return sorted(list, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
    

def loadData(params, server):
  data = ReportData()  
  
  where = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}") and "id"="{2}"'.format(
      params.start.strftime("%d/%m/%Y 0:0:0"),
      params.finish.strftime("%d/%m/%Y 23:59:59"),
      params.id)

  remnants = server.Get("OrgRemnants", where)
  price = server.Get("ManagerPrice", "", "id")
  planogram = server.Get("Planogram", where)
  
  where = '"id" = "{0}"'.format(params.id)
  org = server.Get("Org" , where)
  
  agents = server.Get('Agents', '', 'id')
  
  where  = '"id" = "{0}"'.format(params.id)
  agentorg = server.Get('AgentOrg', where)
  
  if len(agentorg) > 0:
    data.agentname = agents[agentorg[0].userid].name if agentorg[0].userid in agents else agentorg[0].userid
  
  if len(org) > 0:
    data.orgname = org[0].name
    data.lpr = org[0].cheif
    data.fio = org[0].contact
  
  for p in planogram:
    if p.approved == 0 and not p.date in data.planogram:
      data.planogram.append(p.date)
  
  days = (params.finish - params.start).days + 1
  
  for r in remnants:
    for i in r.items:
      if not i.id in data.data:
        data.data[i.id] = Item()
        data.data[i.id].id = i.id
        
        if i.id in price:
          data.data[i.id].name = price[i.id].name
          data.data[i.id].ntz = price[i.id].ntz
        else:
          data.data[i.id].name = "Код товара <{0}>".format(i.id)
          
        data.data[i.id].visits = [0] * days
        
      idx = (r.created - params.start).days
      
      data.data[i.id].visits[idx] = i.qty
      
  return data
    
class XLBuilderEx(XLBuilder):
  RED    = 'FFFF0000'

  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
  
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
      
def printOut(data, params):
    DEFAULTS.font.name="Arial"
    Font.DEFAULT_FONT_NAME = "Arial"
    Font.DEFAULT_FONT_SIZE = 11

    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    
    c = sheet.cell(row=0,column=0)
    c.value = "{0} {1} {2}".format(data.orgname, data.lpr, data.fio)
    
    c = sheet.cell(row=1,column=0)
    c.value = 'Торговый агент: {0}'.format(data.agentname)
    
    xlb = XLBuilderEx()
    row = 2
    
    head = ['Продукт', 'НТЗ']
    date = params.finish
  
    sheet.merge_cells(start_row=row, start_column=0, end_row=row+1, end_column=0)  
    sheet.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)  
    
    c = len(head)
    
    while date >= params.start:
      cell = sheet.cell(row=row,column=c)
      cell.value = date.strftime('%d.%m.%Y')
      
      if date in data.planogram:
        xlb.setBackColor(cell,XLBuilderEx.RED)

        
      sheet.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c+1)  
      sheet.cell(row=row + 1,column=c).value = "Наличие"
      sheet.cell(row=row + 1,column=c + 1).value = "Выкладка"
      
      c += 2
      date += timedelta(days=-1)
    
    xlb.makeHead(sheet, row, head, True)
    
    for c in range(0, len(head) + ((params.finish-params.start).days + 1) * 2):
      for r in range(row, row+2):
        xlb.makeBorder(sheet.cell(row=r, column=c), XLBuilder.HEAD_BORDER_STYLE)

    row += 2

    for item in data.toList():
      xlb.makeCells(sheet, row, item.getData(row))
      row += 1
        
    cc = 1
    for w in [50,8]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    for w in range(0, ((params.finish-params.start).days + 1) * 2):
      sheet.column_dimensions[get_column_letter(cc)].width = 10
      cc += 1
        
    return wb
        
def run(server):
    
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "incass_rep.xlsx", server)                
    logging.info('end')
    