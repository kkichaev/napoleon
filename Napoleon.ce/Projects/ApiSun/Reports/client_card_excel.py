# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
from grsoft.xl_base import XLBuilder
from datetime import timedelta
from datetime import datetime

import tempfile
import io
import client_card


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def setData(sh, xlb, r, t, v, h):
    c = sh.cell(row=r, column=0);
    c.value = t
    c.style.alignment.wrap_text = True
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.vertical = Alignment.VERTICAL_CENTER
 
    sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = sh.cell(row=r, column=1);
    c.value = v
    c.style.alignment.wrap_text = True
     
    for c in range(0,6):
        xlb.makeBorder(sh.cell(row=r, column=c), Border.BORDER_MEDIUM)
         
    sh.row_dimensions[sh.cell(row=r, column=0).row].height = h

def printSheet(xlb, sh, data):
    sh.merge_cells(start_row=1, start_column=0, end_row=1, end_column=5)
    c = sh.cell(row=1, column=0)
    c.value = "Карта клиента."
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.font.bold = True
    
    for c in range(0,6):
        xlb.makeBorder(sh.cell(row=1, column=c), Border.BORDER_MEDIUM)

    setData(sh, xlb, 4, "Название учреждения", data.name, 36)
    setData(sh, xlb, 6, "Адрес", data.address, 25)
    setData(sh, xlb, 8, "Категория", data.categ, 18)
    setData(sh, xlb, 10, "Телефон", data.phone, 28.5)
    setData(sh, xlb, 11, "Эл.почта", data.email, 28.5)
    setData(sh, xlb, 12, "Ф.И.О. директора/врача", data.cheif, 33)
    setData(sh, xlb, 14, "Ф.И.О. сотрудников", data.contact, 28.5)
    setData(sh, xlb, 16, "Закупка у дистрибьютора", data.dealers, 32.35)
    setData(sh, xlb, 18, "До 1 визита(что есть в наличии, на витрине, что назначали Апи-Сан)", data.beforevisit, 63)
    setData(sh, xlb, 20, "Закупки/назначения конкуренты", data.concurent, 47.25)
    setData(sh, xlb, 22, "Дата.Описание визитов.", data.visits, 132.75)
    setData(sh, xlb, 24, "Цель на следующий визит.", data.target, 118.50)
#     
    xlb.setCellWidth(sh, [27,8,8,8,8,27])
    
def printOut(d, xlb):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    printSheet(xlb, sh, d)
    return wb
    
def doReport(server):
    data  = client_card.loadData(server, chr(13) + chr(10))
    xlb = XLBuilder()
    wb = printOut(data, xlb)
    xlb.workbookToObject(wb, "clientcard.xlsx", server)
   
def run(server):
   print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
   doReport(server)
   print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')

   
