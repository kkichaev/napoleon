# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
import tempfile
import io
from datetime import datetime

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def setBorder(cell):
   cell.style.borders.left.border_style = Border.BORDER_THIN
   cell.style.borders.right.border_style = Border.BORDER_THIN
   cell.style.borders.top.border_style = Border.BORDER_THIN
   cell.style.borders.bottom.border_style = Border.BORDER_THIN

def getColor(val1, val2):
    res = val1 - val2
    
    if res == 0:
        return Color.YELLOW
    elif res < 0: 
        return Color.GREEN
    else:
        return Color.RED
    
def doReport(data, outObj):
   wb = Workbook(False, 'cp1251')
   sheet = wb.get_active_sheet()
   
   sheet.cell(row=0, column=0).value = "Изменение ассортимента"
   sheet.cell(row=1, column=0).value = "Период 1 с {0} по {1}: ".format(data.range1start.strftime("%d/%m/%Y"),
      data.range1finish.strftime("%d/%m/%Y"));
   sheet.cell(row=2, column=0).value = "Период 2 с {0} по {1}: ".format(data.range2start.strftime("%d/%m/%Y"),
      data.range2finish.strftime("%d/%m/%Y"));  
   
   cell = sheet.cell(row=4, column=0)
   cell.value = "Торговая точка" 
   cell.style.font.bold = True
   sheet.column_dimensions[get_column_letter(1)].width = 50
   
   STARTPRICECOL = 1 
   col = STARTPRICECOL
   row = 4
   for p in data.price:
     cell = sheet.cell(row=row, column=col)
     cell.value = p.name 
     cell.style.font.bold = True
     cell.style.alignment.wrap_text = True
     sheet.column_dimensions[get_column_letter(col + 1)].width = 10
     col = col + 1
   
   cell = sheet.cell(row=4, column=col)
   cell.value = "Адрес" 
   cell.style.font.bold = True
   sheet.column_dimensions[get_column_letter(col + 1)].width = 40
   cell = sheet.cell(row=4, column=col + 1)
   cell.value = "Телефон" 
   cell.style.font.bold = True 
   cell = sheet.cell(row=4, column=col + 2)
   cell.value = "Вид ТТ" 
   cell.style.font.bold = True
   cell = sheet.cell(row=4, column=col + 3)
   cell.value = "Оптовик" 
   cell.style.font.bold = True 
   cell = sheet.cell(row=4, column=col + 4)
   sheet.column_dimensions[get_column_letter(col + 4)].width = 40
   cell.value = "Лицензия" 
   cell.style.font.bold = True 
   cell = sheet.cell(row=4, column=col + 5)
   cell.value = "Директор" 
   cell.style.font.bold = True
   cell = sheet.cell(row=4, column=col + 6)
   cell.value = "Контактное лицо" 
   cell.style.alignment.wrap_text = True
   cell.style.font.bold = True  
   cell = sheet.cell(row=4, column=col + 7)
   cell.value = "Средняя проходимость" 
   cell.style.font.bold = True 
   cell.style.alignment.wrap_text = True
   cell = sheet.cell(row=4, column=col + 8)
   cell.value = "Агент" 
   cell.style.font.bold = True 
   cell.style.alignment.wrap_text = True
   sheet.column_dimensions[get_column_letter(col + 9)].width = 40
    
   row = row + 1
   
   for key, value in data.items.iteritems():
     sheet.cell(row=row, column=0).value = data.orgs[key].name if key in data.orgs else "Объект с кодом<{0}> не найден".format(key)
     
     col = STARTPRICECOL
     for p in data.price:
        if p.id in value.items:
            item = value.items[p.id]
            cell = sheet.cell(row=row, column=col)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = getColor(item.val1, item.val2)
            setBorder(cell)
        col = col + 1    
     
     if key in data.orgs:
         org = data.orgs[key]
         cell = sheet.cell(row=row, column=col)
         cell.value = org.address
         cell = sheet.cell(row=row, column=col + 1)
         cell.value = org.contactPhone
         cell = sheet.cell(row=row, column=col + 2)
         cell.value = data.orgtypes[org.orgType].name if org.orgType in data.orgtypes else org.orgType 
         cell = sheet.cell(row=row, column=col + 3)
         cell.value = data.dealers[org.dealer].name if org.dealer in data.dealers else org.dealer
         cell = sheet.cell(row=row, column=col + 4)
         cell.value = "да" if org.license == 1 else "нет"
         cell = sheet.cell(row=row, column=col + 5)
         cell.value = org.cheif
         cell = sheet.cell(row=row, column=col + 6)
         cell.value = org.contact
         cell = sheet.cell(row=row, column=col + 7)
         cell.value = org.avgTraff
         cell = sheet.cell(row=row, column=col + 8)
         cell.value = data.agents[value.userid].name if value.userid in data.agents else value.userid
     row = row + 1          

   cell = sheet.cell(row=row, column=0)
   cell.value = "Итого:"
   cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
   cell.style.font.bold = True
   col = STARTPRICECOL
   
   for p in data.price:
       if p.id in data.total and data.total[p.id] != 0:
           cell = sheet.cell(row=row, column=col)
           cell.value = data.total[p.id]
           cell.style.font.bold = True
       col = col + 1    
  
   row = row + 1
   cell = sheet.cell(row=row, column=0)
   cell.value = "Изменение:"
   cell.style.font.bold = True
   cell = sheet.cell(row=row, column=1)
   cell.value = '"{:+0}"'.format(data.totalval)
   cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
   cell.style.font.bold = True
            
   repName = "assort_ch_report.xlsx" 
   fileName = tempfile.gettempdir() + '/' + repName
   wb.save(fileName)

   file = io.open(fileName, 'rb')
   bytes = file.read(-1)
   file.close()

   obj = outObj.New()
   obj.name = repName
   obj.file = bytes

class PriceItem:
    price = None
    val1 = 0
    val2 = 0
    
    def __init__(self, price):
        self.val1 = 0
        self.val2 = 0
        self.price = price
           
class ReportDataItem:
    items = None
    userid = None
    def __init__(self, price, userid):
        self.items = dict()
        self.userid = userid
        
        for p in price:
            self.items[p.id] = PriceItem(p)
            
    def update(self, ri, val):
        if ri in self.items:
            if val == 0:
                self.items[ri].val1 = 1
            else:      
                self.items[ri].val2 = 1
               
class ReportData:
    agents = None
    price = None
    items = None
    range1start = None
    range1finish = None
    range2start = None
    range2finish = None
    orgs = None
    orgtypes = None
    dealers = None
    total = None
    totalval = 0

    def __init__(self, server):
        param = server.Params[0]
        self.range1start = param.range1start
        self.range1finish = param.range1finish
        self.range2start = param.range2start
        self.range2finish = param.range2finish
        self.agents = server.Get("Agents", "", "id")
        self.orgs = server.Get("CommonOrgs", "", "id")
        self.orgtypes = server.Get("OrgType", "", "id")
        self.dealers = server.Get("Dealer", "", "id")
        self.total = dict()
        
        fp = server.Get("Price", "", 'id')
        i =  param.items.split(',')
        self.price = list()
        
        for key, value in fp.iteritems():
            if key in i:
                self.price.append(value)
                
        self.items = dict();
        regions = param.regions.split(',')
        
        fo = dict()
        if len(regions) > 0:
            for o in self.orgs.values():
                if o.parent in regions:
                   fo[o.id] = o
                   
            self.orgs = fo 
        
        where = '"created" > ToDate("{0}") and created <= ToDate("{1}") '.format(
            param.range1start.strftime("%d/%m/%Y 0:0:0"), param.range1finish.strftime("%d/%m/%Y 23:59:59"))
   
        if len(param.agents) > 0:
            where += ' and "userid" in (' + param.agents + ')'

        rem1 = server.Get("OrgRemnants", where);
        self.traverse(rem1, 0)
        
        where = '"created" > ToDate("{0}") and created <= ToDate("{1}") '.format(
            param.range2start.strftime("%d/%m/%Y 0:0:0"), param.range2finish.strftime("%d/%m/%Y 23:59:59"))
   
        if len(param.agents) > 0:
            where += ' and "userid" in (' + param.agents + ')'
            
        rem2 = server.Get("OrgRemnants", where)
        self.traverse(rem2, 1)
        
        for p in self.price:
            for key, value in self.items.iteritems():
                item = value.items[p.id]
                v = item.val2 - item.val1
                
                if p.id in self.total:
                    self.total[p.id] = self.total[p.id] + v
                else:
                    self.total[p.id] = v   
                
                self.totalval = self.totalval + v     
                
        orgs = dict()
         
        for i in self.items.keys():
            orgs[i] = self.orgs[i]
             
        self.orgs = orgs    
            
    def traverse(self, rem, val):
        for r in rem:
            if not r.id in self.orgs:
                continue
            
            rdi = None
            
            if r.id in self.items:
                rdi = self.items[r.id]
            else:
                rdi = ReportDataItem(self.price, r.userid)    
                self.items[r.id] = rdi
                
            for ri in r.items:
                rdi.update(ri.id, val)
           
def run(server):
   print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')

   server.RegisterType("Result[name:s,file:b]")
   outObj = server.New("Result")
   
   reportData = ReportData(server) 
   doReport(reportData, outObj)
      
   server.Put(outObj)

   print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')


   
