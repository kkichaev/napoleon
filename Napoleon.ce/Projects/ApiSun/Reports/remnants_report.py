# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
import tempfile
import io


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def getDivisionAgents(server, division, agents):
   if division == None or len(division) == 0:
      return

   for d in division :
      for a in d.agents :
        agents.append(a.id)

      getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)

def loadAgents(server, division):
   ret = dict()
   agents = server.Get("Agents", "")

   divagents = []
   getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
   divagents = set(divagents)

   for a in agents:
      if a.id in divagents:
         ret[a.id] = a

   return ret

def makeIDStr(server, agents):
   res = '"userid" in ('

   for id in agents.iterkeys():
      res += "'" + id + "',"

   res = res[:-1] + ")"
   return res

def getRemntItemCount(remnants, pid, d):
   result = 0
   
   for r in remnants:
      if r.created.date() == d.date():
         for item in r.items:
            if item.id == pid:
               result += item.qty
               
   return result

def setBorder(cell):
   cell.style.borders.left.border_style = Border.BORDER_THIN
   cell.style.borders.right.border_style = Border.BORDER_THIN
   cell.style.borders.top.border_style = Border.BORDER_THIN
   cell.style.borders.bottom.border_style = Border.BORDER_THIN
   
def dealersStr(server, dealers, orgdeal):
   result = ""
   
   for d in orgdeal:
      if d.id in dealers:
         result += dealers[d.id].name
         result += ", "
         
   if len(result) > 0:
      result = result[:-2]
            
   return result   

def doReport(server, start, finish, orgs, outObj):
   wb = Workbook(False, 'cp1251')
   sheet = wb.get_active_sheet()
      
   dealers = server.Get("Dealer",'',"id")   
   price=server.Get("Price","")
   price = sorted(price, cmp=lambda p1, p2: cmp(p1.name,p2.name))
   
   sr = 1;
   delta = datetime.timedelta(days=1)

   for o in orgs:
      sheet.cell(row=sr, column=0).value = o.name
      
      sheet.cell(row=sr, column=1).value = "адрес: " + o.address + ", руководитель: " + \
         o.cheif + ", телефон руководителя: " + o.cheifPhone + ", контактное лицо: " + \
         o.contact + ", телефон: " + o.contactPhone + ", лицензия: " + ("да" if o.license > 0 else "нет") + \
         ", оптовики: " +dealersStr(server, dealers, o.dealers) + ", средняя проходимость: " + str(o.avgTraff)
         
      d = start
      cl = 2
      
      where = '"id"=\''+o.id+'\' and "created" > ToDate("{0}") and created <= ToDate("{1}")'.format(
         start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 23:59:59"))
      
      remnants = server.Get("OrgRemnants", where)
      
      while d <= finish:
         sheet.cell(row=sr, column=cl).value = d.strftime("%Y-%m-%d")
         d += delta
         cl += 1
      
         
      sr = sr + 1 
      dateqty = dict()
      for p in price:
         cell = sheet.cell(row=sr, column=0)
         cell.value = o.name
#          setBorder(cell)
         cell = sheet.cell(row=sr, column=1)
         cell.value = p.name
#          setBorder(cell)
         
         d = start
         cl = 2
         while d <= finish:
            cell = sheet.cell(row=sr, column=cl)
#            setBorder(cell)
            q = getRemntItemCount(remnants, p.id, d)
    
            if d in dateqty:
               dateqty[d] += q
            else: 
               dateqty[d] = q
                 
#            cell.style.fill.fill_type = Fill.FILL_SOLID
#            cell.style.fill.start_color.index = Color.GREEN if q > 0 else Color.RED

            d += delta
            cl += 1
      
         sr = sr + 1
      sheet.cell(row=sr, column=0).value = o.name   
      sheet.cell(row=sr, column=1).value = "ИТОГО:"
      dd = start
      cll = 2
      
      while dd <= finish:
         sheet.cell(row=sr, column=cll).value = dateqty[dd]
         dd += delta
         cll += 1
         
      sr = sr + 1 
      
      dd = start
      cll = 2
      sheet.cell(row=sr, column=0).value = o.name
      sheet.cell(row=sr, column=1).value = "Посещения:"
      
      while dd <= finish:
         where = '"id"=\''+o.id+'\' and "created" > ToDate("{0}") and created <= ToDate("{1}")'.format(
         dd.strftime("%d/%m/%Y 0:0:0"), dd.strftime("%d/%m/%Y 23:59:59"))
         visit = server.Get("VisitInfo", where)
          
         if len(visit) > 0 :
            sheet.cell(row=sr, column=cll).value = visit[0].remark
            
         dd += delta
         cll += 1
         
      sr = sr + 1  
      sr = sr + 1
   
   sheet.column_dimensions[get_column_letter(1)].width = 35
   sheet.column_dimensions[get_column_letter(2)].width = 73
   
   d = start
   cl = 3
   
   while d <= finish:
      sheet.column_dimensions[get_column_letter(cl)].width = 10
      d += delta
      cl += 1
   
   sheet.auto_filter = sheet.range("A1:"+get_column_letter(2)+str(sr)) 
   repName = "remnants.xlsx" 
   fileName = tempfile.gettempdir() + '/' + repName
   wb.save(fileName)

   file = io.open(fileName, 'rb')
   bytes = file.read(-1)
   file.close()

   obj = outObj.New()
   obj.name = repName
   obj.file = bytes
   
def run(server):

   print "start remnants_report"

   params = server.Params
   param = params[0]
   start = param.start
   finish = param.finish
   orgid = param.orgid
   divid = param.divid
   userid = param.userid
   
   if orgid != "":
      where = '"id"=' + "'" + orgid + "'"
   elif userid != "":
      where = '"id" in (select "id" from AgentOrg where "userid" = ' + "'" + userid + "')"
   elif divid != "":
      agents = loadAgents(server, divid)
      uidFilter = makeIDStr(server, agents)
      where = '"id" in (select "id" from AgentOrg where ' + uidFilter + ')'
     
   orgs = server.Get("CommonOrgs",where)         
   orgs = sorted(orgs, cmp=lambda o1, o2: cmp(o1.name,o2.name))

   server.RegisterType("Result[name:s,file:b]")
   outObj = server.New("Result")
   
   doReport(server, start, finish, orgs, outObj)
   
   server.Put(outObj)

   print "finish remnants_report"

   
