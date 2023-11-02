# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
import tempfile
import io


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def getDivisionAgents(server, division, agents):
   if division == None or len(division) == 0:
      return

   for d in division :
      for a in d.agents :
        agents.append(a.id)

      getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)

def loadAgents(server, division):
   ret = dict()
   agents = server.Get("Agents", "")

   divagents = []
   getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
   divagents = set(divagents)

   for a in agents:
      if a.id in divagents:
         ret[a.id] = a

   return ret

def makeIDStr(server, agents):
   res = '"userid" in ('

   for id in agents.iterkeys():
      res += "'" + id + "',"

   res = res[:-1] + ")"
   return res

def getRemntItemCount(remnants, pid, d):
   result = 0
   
   for r in remnants:
      if r.created.date() == d.date():
         for item in r.items:
            if item.id == pid:
               result += item.qty
               
   return result

def setBorder(cell):
   cell.style.borders.left.border_style = Border.BORDER_THIN
   cell.style.borders.right.border_style = Border.BORDER_THIN
   cell.style.borders.top.border_style = Border.BORDER_THIN
   cell.style.borders.bottom.border_style = Border.BORDER_THIN

def doReport(server, start, finish, filter, uidFilter, outObj):
   wb = Workbook(False, 'cp1251')
   sheet = wb.get_active_sheet()
      
   filterPrice = set()
   price=server.Get("Price","","id")
   filter = unicode(filter,"cp1251").upper();
   
   for p in price.values():
      name = unicode(p.name, "cp1251").upper()
      if name.find(filter) != -1:
         filterPrice.add(p.id)
          
   where = '"created" > ToDate("{0}") and created <= ToDate("{1}")'.format(
         start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 23:59:59"))
   
   if len(uidFilter) > 0:
      where += ' and ' + uidFilter
      
   remnants = server.Get("OrgRemnants", where)   
   
   orgs = server.Get("CommonOrgs", '', "id")
   filterOrgs = set()
   orgItems = dict();
   
   for r in remnants:
      for i in r.items:
         if i.id in filterPrice and r.id in orgs:
            filterOrgs.add(orgs[r.id])
            
            itemSet = None
             
            if r.id in orgItems:
               itemSet = orgItems[r.id]
            else:    
               itemSet = set()
               orgItems[r.id] = itemSet
           
            itemSet.add(i.id)    
            
   filterOrgs = sorted(filterOrgs, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
   
   sr = 1
   cl = 1
   
   for p in filterPrice:
      cell = sheet.cell(row=0, column=cl)
      cell.value = price[p].name
      cell.style.alignment.wrap_text = True
      cl = cl + 1
         
   sheet.cell(row=0, column=cl).value = "адрес"
   sheet.cell(row=0, column=cl).style.font.bold = True
   sheet.cell(row=0, column=cl + 1).value = "телефон"
   sheet.cell(row=0, column=cl + 1).style.font.bold = True
   sheet.cell(row=0, column=cl + 2).value = "вид ТТ"
   sheet.cell(row=0, column=cl + 2).style.font.bold = True
   sheet.cell(row=0, column=cl + 3).value = "оптовик"
   sheet.cell(row=0, column=cl + 3).style.font.bold = True
   sheet.cell(row=0, column=cl + 4).value = "лицензия"
   sheet.cell(row=0, column=cl + 4).style.font.bold = True
   sheet.cell(row=0, column=cl + 5).value = "директор"
   sheet.cell(row=0, column=cl + 5).style.font.bold = True
   sheet.cell(row=0, column=cl + 6).value = "контактное лицо"
   sheet.cell(row=0, column=cl + 6).style.font.bold = True
   sheet.cell(row=0, column=cl + 7).value = "средняя проходимость"
   sheet.cell(row=0, column=cl + 7).style.font.bold = True
      
   for o in filterOrgs: 
      sheet.cell(row=sr, column=0).value = o.name
      
      cl = 1
      for p in filterPrice:
         cell = sheet.cell(row=sr, column=cl)
         cell.style.fill.fill_type = Fill.FILL_SOLID
         cell.style.fill.start_color.index = Color.RED
         setBorder(cell)
         
         if p in orgItems[o.id]:
             cell = sheet.cell(row=sr, column=cl)
             cell.style.fill.fill_type = Fill.FILL_SOLID
             cell.style.fill.start_color.index = Color.GREEN

         cl = cl + 1   
      
      sheet.cell(row=sr, column=cl).value = o.address
      sheet.cell(row=sr, column=cl + 1).value = o.contactPhone
      
      orgType = server.Get("OrgType",'"id" = ' + "'" + o.orgType + "'")
      
      if len(orgType) > 0:
         sheet.cell(row=sr, column=cl + 2).value = orgType[0].name
      else:   
         sheet.cell(row=sr, column=cl + 2).value = "код: " + str(o.orgType)
         
      dealers = "";
      for d in o.dealers:
         dealer = server.Get("Dealer",'"id" = ' + "'" + d.id + "'")
         
         if len(dealers) > 0:
            dealers += ", "
            
         if len(dealer) > 0:
            dealers += dealer[0].name
         else:   
            dealers += "код: " + str(d.id)
         
      sheet.cell(row=sr, column=cl + 3).value = dealers   
      sheet.cell(row=sr, column=cl + 4).value = "да" if o.license > 0 else "нет"
      sheet.cell(row=sr, column=cl + 5).value = o.cheif
      sheet.cell(row=sr, column=cl + 6).value = o.contact
      sheet.cell(row=sr, column=cl + 7).value = "" if o.avgTraff <= 0 else str(o.avgTraff)
      
      sr = sr + 1  

   try:
      sheet.column_dimensions[get_column_letter(1)].width = 35
   except:   
      print "error:", sys.exc_info()[0]
   

   repName = "remnants.xlsx" 
   fileName = tempfile.gettempdir() + '/' + repName
   wb.save(fileName)

   file = io.open(fileName, 'rb')
   bytes = file.read(-1)
   file.close()

   obj = outObj.New()
   obj.name = repName
   obj.file = bytes
   
def run(server):
   print "start present_report"

   params = server.Params
   param = params[0]
   start = param.start
   finish = param.finish
   filter = param.filter
   
   uidFilter = ""
   
   if param.division != -1:
      agents = loadAgents(server, param.division)
      uidFilter = makeIDStr(server, agents)
   elif param.agent != "":   
      uidFilter = '"userid" in (\'' + param.agent +'\')'

   server.RegisterType("Result[name:s,file:b]")
   outObj = server.New("Result")
   
   doReport(server, start, finish, filter, uidFilter, outObj)
      
   server.Put(outObj)

   print "finish present_report"

   
