# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#

from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
import datetime
import tempfile
import io


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def getDivisionAgents(server, division, agents):
   if division == None or len(division) == 0:
      return

   for d in division :
      for a in d.agents :
        agents.append(a.id)

      getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)

def loadAgents(server, division):
   ret = dict()
   agents = server.Get("Agents", "")

   divagents = []
   getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
   divagents = set(divagents)

   for a in agents:
      if a.id in divagents:
         ret[a.id] = a

   return ret

def makeIDStr(server, agents):
   res = '"userid" in ('

   for id in agents.iterkeys():
      res += "'" + id + "',"

   res = res[:-1] + ")"
   return res

def getRemntItemCount(remnants, pid, d):
   result = 0
   
   for r in remnants:
      if r.created.date() == d.date():
         for item in r.items:
            if item.id == pid:
               result += item.qty
               
   return result

def setBorder(cell):
   cell.style.borders.left.border_style = Border.BORDER_THIN
   cell.style.borders.right.border_style = Border.BORDER_THIN
   cell.style.borders.top.border_style = Border.BORDER_THIN
   cell.style.borders.bottom.border_style = Border.BORDER_THIN
   
def setBackColor(cell, color):
  fill = cell.style.fill 
  fill.start_color = Color(color)
  fill.end_color = Color(color)
  fill.fill_type = Fill.FILL_SOLID

def doReport(server, uidFilter, outObj):
   wb = Workbook(False, 'cp1251')
   sheet = wb.get_active_sheet()
   
   where = ""
   
   if len(uidFilter) > 0:
      where = '"id" in (select "id" from agentorg where ' + uidFilter + ')'
   
   orgs = server.Get("CommonOrgs", where)
   orgs = sorted(orgs, cmp=lambda lhs, rhs: cmp(lhs.name, rhs.name))
   
   HEAD   = 'FFF2F2F2'


   sheet.cell(row=0, column=0).value = "наименование"
   sheet.cell(row=0, column=0).style.font.bold = True
   setBorder(sheet.cell(row=0, column=0))
   setBackColor(sheet.cell(row=0, column=0), HEAD)
   sheet.cell(row=0, column=1).value = "адрес"
   sheet.cell(row=0, column=1).style.font.bold = True
   setBorder(sheet.cell(row=0, column=1))
   setBackColor(sheet.cell(row=0, column=1), HEAD)
   sheet.cell(row=0, column=2).value = "телефон"
   sheet.cell(row=0, column=2).style.font.bold = True
   setBorder(sheet.cell(row=0, column=2))
   setBackColor(sheet.cell(row=0, column=2), HEAD)
   sheet.cell(row=0, column=3).value = "emial"
   sheet.cell(row=0, column=3).style.font.bold = True
   setBackColor(sheet.cell(row=0, column=3), HEAD)
   setBorder(sheet.cell(row=0, column=3))
   sheet.cell(row=0, column=4).value = "вид ТТ"
   sheet.cell(row=0, column=4).style.font.bold = True
   setBorder(sheet.cell(row=0, column=4))
   setBackColor(sheet.cell(row=0, column=4), HEAD)
   sheet.cell(row=0, column=5).value = "оптовик"
   sheet.cell(row=0, column=5).style.font.bold = True
   setBorder(sheet.cell(row=0, column=5))
   setBackColor(sheet.cell(row=0, column=5), HEAD)
   sheet.cell(row=0, column=6).value = "лицензия"
   sheet.cell(row=0, column=6).style.font.bold = True
   setBorder(sheet.cell(row=0, column=6))
   setBackColor(sheet.cell(row=0, column=6), HEAD)
   sheet.cell(row=0, column=7).value = "директор"
   sheet.cell(row=0, column=7).style.font.bold = True
   setBorder(sheet.cell(row=0, column=7))
   setBackColor(sheet.cell(row=0, column=7), HEAD)
   sheet.cell(row=0, column=8).value = "контактное лицо"
   sheet.cell(row=0, column=8).style.font.bold = True
   setBorder(sheet.cell(row=0, column=8))
   setBackColor(sheet.cell(row=0, column=8), HEAD)
   sheet.cell(row=0, column=9).value = "средняя проходимость"
   sheet.cell(row=0, column=9).style.font.bold = True
   setBorder(sheet.cell(row=0, column=9))
   setBackColor(sheet.cell(row=0, column=9), HEAD)
   
   sr = 1
         
   for o in orgs: 
      sheet.cell(row=sr, column=0).value = o.name
      sheet.cell(row=sr, column=0).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=0))
      sheet.cell(row=sr, column=1).value = o.address
      sheet.cell(row=sr, column=1).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=1))
      sheet.cell(row=sr, column=2).value = o.contactPhone
      sheet.cell(row=sr, column=2).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=2))
      sheet.cell(row=sr, column=3).value = o.email
      sheet.cell(row=sr, column=3).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=3))
      
      orgType = server.Get("OrgType",'"id" = ' + "'" + o.orgType + "'")
      
      if len(orgType) > 0:
         sheet.cell(row=sr, column=4).value = orgType[0].name
      else:   
         sheet.cell(row=sr, column=4).value = "код: " + str(o.orgType)
         
      sheet.cell(row=sr, column=4).style.alignment.wrap_text = True   
      setBorder(sheet.cell(row=sr, column=4))
         
      dealers = "";
      for d in o.dealers:
         dealer = server.Get("Dealer",'"id" = ' + "'" + d.id + "'")
         
         if len(dealers) > 0:
            dealers += ", "
            
         if len(dealer) > 0:
            dealers += dealer[0].name
         else:   
            dealers += "код: " + str(d.id)
         
      sheet.cell(row=sr, column=5).value = dealers   
      sheet.cell(row=sr, column=5).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=5))
      sheet.cell(row=sr, column=6).value = "да" if o.license > 0 else "нет"
      sheet.cell(row=sr, column=6).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=6))
      sheet.cell(row=sr, column=7).value = o.cheif
      sheet.cell(row=sr, column=7).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=7))
      sheet.cell(row=sr, column=8).value = o.contact
      sheet.cell(row=sr, column=8).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=8))
      sheet.cell(row=sr, column=9).value = "" if o.avgTraff <= 0 else str(o.avgTraff)
      sheet.cell(row=sr, column=9).style.alignment.wrap_text = True
      setBorder(sheet.cell(row=sr, column=9))
      
      sr = sr + 1  

   sheet.column_dimensions[get_column_letter(1)].width = 35
   sheet.column_dimensions[get_column_letter(2)].width = 35
   sheet.column_dimensions[get_column_letter(3)].width = 21
   sheet.column_dimensions[get_column_letter(4)].width = 27
   sheet.column_dimensions[get_column_letter(5)].width = 11
   sheet.column_dimensions[get_column_letter(6)].width = 30
   sheet.column_dimensions[get_column_letter(7)].width = 11
   sheet.column_dimensions[get_column_letter(8)].width = 30
   sheet.column_dimensions[get_column_letter(9)].width = 30
   sheet.column_dimensions[get_column_letter(10)].width = 30


   repName = "remnants.xlsx" 
   fileName = tempfile.gettempdir() + '/' + repName
   wb.save(fileName)

   file = io.open(fileName, 'rb')
   bytes = file.read(-1)
   file.close()

   obj = outObj.New()
   obj.name = repName
   obj.file = bytes
   
def run(server):
   print "start org_report"

   params = server.Params
   param = params[0]

   uidFilter = ""
   print "param.divid", param.divid
   print "param.userid", param.userid
   
   if len(param.divid) > 0:
      agents = loadAgents(server, param.divid)
      uidFilter = makeIDStr(server, agents)
   elif len(param.userid) > 0 :   
      uidFilter = '"userid" in (' + param.userid +')'

   print "uidFilter", uidFilter
   server.RegisterType("Result[name:s,file:b]")
   outObj = server.New("Result")
   
   #try:
   doReport(server, uidFilter, outObj)
   #except:   
   #   print "error:", sys.exc_info()[0]
      
   server.Put(outObj)

   print "finish org_report"

   
