# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class Report:
    items = None
    agent = None
    date = None
    
    def __init__(self):
        self.items = list()
        self.agent = ""
        self.date = datetime.now()

class Item:
    org = None
    item = None
    ost = None
    rekzak = None
    qty = None
    
    def __init__(self):
      self.org = ""
      self.item = ""
      self.ost = 0
      self.rekzak = 0
      self.qty = 0
      
    def getData(self, row):
      return [self.org, self.item, self.ost, self.rekzak if self.rekzak >= 0 else "" , self.qty, "=E{0}-D{0}".format(row+1)]

def inflateParams(server):
    return server.Params[0].date, server.Params[0].userid, server.Params[0].matrix 
    
def loadData(server):
    ret = Report()
    
    date, userid, matrix = inflateParams(server)
    userid = "'" + userid + "'";
    WHERE = '"created" >= ToDate("{0}") and "created" < ToDate("{1}") and "userid" = {2}';         
    where = WHERE.format(date.strftime("%d/%m/%Y 0:0:0"), (date +timedelta(days=1)).strftime("%d/%m/%Y 0:0:0"), userid)
    orders = server.Get("Order", where)
    remnants = server.Get("OrgRemnants", where)
    price = server.Get("Price", "setqtyfilter(false)", "id")
    mtx = server.Get("CommonMatrix", "", "name")
    
    priceFilter = list()
    
    if matrix in mtx:
        for mi in mtx[matrix].items:
            priceFilter.append(mi.id);
    
    server.ChangeUser(userid)
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    agent = server.CurrentUser().name
    server.RestoreUser()
    
    data_dict = dict() # key = order.id + item.id 

    iter_doc(data_dict, orders, orgs, price, ord_item, priceFilter)
    iter_doc(data_dict, remnants, orgs, price, rem_item, priceFilter)
    
    items = list()
    
    for i in data_dict.values():
        items.append(i)
         
    ret.items = sorted(items, cmp=item_cmp)
    ret.date = date
    ret.agent = agent
    
    return ret

def iter_doc(data, docs, orgs, price, func, filter):
    for d in docs:
        for i in d.items:
            if not i.id in filter:
                continue
            
            key = d.id + i.id
             
            if not key in data:
                item = Item()
                item.org = orgs[d.id].name if d.id in orgs else "Контрагент <{0}>".format(d.id)
                item.item = price[i.id].name if i.id in price else "Товар <{0}>".format(i.id)
                data[key] = item
                
            item = data[key]
            func(item , i)
            
def ord_item(item, doc_item):
    item.qty += doc_item.qty
    item.rekzak = doc_item.rekzak
                     
def rem_item(item, doc_item):
    item.ost = doc_item.qty    
        
def item_cmp(x, y):
    res = cmp(x.org, y.org)
  
    if res == 0:
        res = cmp(x.item, y.item)

    return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    sh.cell(row=0, column=0).value = "Агент: {0}".format(data.agent)
    sh.cell(row=1, column=0).value = "Дата: {0}".format(data.date.strftime("%d.%m.%Y"))
    
    head = ["Организация", "Товар", "Остаток", "рекомендованный заказ", "факт.заказ", "отклонение от рекомендованного (шт)"]
    
    r = 2
    xlb.makeHead(sh, r, head)
    
    for d in data.items:
        r += 1
        xlb.makeCells(sh, r, d.getData(r))

    setCellWidth(sh, [30,30])
    
class XLBuilderEx(XLBuilder):
   def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment. wrap_text = True
        return column   
      
    
def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data  = loadData(server)
    wb = printOut(data)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    