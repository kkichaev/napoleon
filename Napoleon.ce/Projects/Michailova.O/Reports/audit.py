# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class Report:
    items = None
    agent = None
    date = None
    titles = None
    
    def __init__(self):
        self.items = list()
        self.agent = ""
        self.date = datetime.now()
        self.titles = [];

class Item:
    pos = None
    org = None
    goods = None
    
    def __init__(self):
      self.pos = 0  
      self.org = ""
      self.goods = []
      
    def getData(self, row):
      res = [self.pos, self.org]
      res.extend(self.goods)
      res.append("=SUM(B{0}:{1}{0})".format(row+1, get_column_letter(len(self.goods) + 1)))  
      return res

def inflateParams(server):
    return server.Params[0].date, server.Params[0].userid, server.Params[0].matrix
    
def loadData(server):
    ret = Report()
    
    date, userid, matrix = inflateParams(server)
    userid = "'" + userid + "'";
    WHERE = '"created" >= ToDate("{0}") and "created" < ToDate("{1}") and "userid" = {2}';         
    where = WHERE.format(date.strftime("%d/%m/%Y 0:0:0"), (date +timedelta(days=1)).strftime("%d/%m/%Y 0:0:0"), userid)
    remnants = server.Get("OrgRemnants", where)
    price = server.Get("Price", "setqtyfilter(false)", "id")
    otm = server.Get("OrgTypeMatrix", "")
    mtx = server.Get("CommonMatrix", "", "name")
    
    server.ChangeUser(userid)
    orgs = server.Get("Org", "", "id")
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    agent = server.CurrentUser().name
    server.RestoreUser()
    
    orgids = list()
    
    for m in otm:
        if m.matrix == matrix:
            orgids.append(m.id)
    
    pids = list()
    
    if matrix in mtx:
        for mi in mtx[matrix].items:
            pids.append(mi.id);
            
    goods = dict() # key = price.id
    
    for i in pids:
        if i in price:
            goods[i] = price[i]        
    
    goods_list = list() # sort goods by name
    for g in goods.values():
        goods_list.append(g)
    
    goods_list = sorted(goods_list, cmp=lambda x,y: cmp(x.name,y.name))    
    goods_idx_list = []
    for g in goods_list:
        goods_idx_list.append(g.id)
    
    data_dict = dict() # key = org.id
    
    visitedOrgs = list()
    
    for r in remnants:
        visitedOrgs.append(r.id)
        
    for oid in orgids:
        if oid in orgs and oid in visitedOrgs:
            item = Item()
            item.org = orgs[oid].name + " " + orgs[oid].address
            item.goods = [0] * len(goods_idx_list)
            data_dict[oid] = item
 
    for r in remnants:
        if r.id in data_dict:
            item = data_dict[r.id]
            
            for i in r.items:
                idx = -1
                if i.id in goods_idx_list:
                    idx = goods_idx_list.index(i.id)
                
                if idx >= 0:
                    item.goods[idx] = 1 if i.qty > 0 else 0
                       
    items = list()
    for i in data_dict.values():
        items.append(i)
        
    ret.items = sorted(items, cmp=item_cmp)
    ret.date = date
    ret.agent = agent
    
    for g in goods_list:
        ret.titles.append(g.name)
        
    p = 1
    
    for i in ret.items:
        i.pos = p
        p += 1    
    
    return ret
        
def item_cmp(x, y):
    res = cmp(x.org, y.org)
  
    if res == 0:
        res = cmd(x.item, y.item)

    return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def rangeBorders(range):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN
                    
def ptintSheet(xlb, sh, data):
    sh.cell(row=0, column=0).value = "Агент: {0}".format(data.agent)
    sh.cell(row=1, column=0).value = "Дата: {0}".format(data.date.strftime("%d.%m.%Y"))
    
    head = ["№ п/п", "Название клиента, адрес"]
    head.extend(data.titles)
    head.append("Всего SKU ассортимента ЯДРА на полке")
     
    r = 2
    xlb.makeHead(sh, r, head)
     
    for d in data.items:
        r += 1
        xlb.makeCells(sh, r, d.getData(r))
 
    er = r
    r+=1
    sh.merge_cells(start_row=r, start_column=0, end_row=r+2, end_column=1)
    sh.cell(row=r, column=0).value="Итоговая оценка за аудит"
    s = sh.cell(row=r, column=0).style
    s.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    s.alignment.vertical = Alignment.VERTICAL_CENTER
    
    sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + len(data.titles) - 1)
    sh.cell(row=r, column=2).value="КОЛ-ВО ТТ В АУДИТЕ"
    sh.cell(row=r, column=2 + len(data.titles)).value="=A{0}".format(r)
    r+=1
    
    sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + len(data.titles) - 1)
    sh.cell(row=r, column=2).value="ОБЩЕЕ КОЛ-ВО СКЮ НА ПОЛКЕ"
    sh.cell(row=r, column=2 + len(data.titles)).value="=SUM({0}4:{0}{1})".format(get_column_letter(2 + len(data.titles) + 1), er+1)
    r+=1
    
    sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + len(data.titles) - 1)
    sh.cell(row=r, column=2).value="СРЕДНЕЕ КОЛ-ВО СКЮ НА ПОЛКЕ"
    sh.cell(row=r, column=2 + len(data.titles)).value="=IFERROR({0}{1}/{0}{2},0)".format(get_column_letter(2 + len(data.titles) + 1), r, r-1)
    
    rangeBorders(sh.range("A{0}:{1}{2}".format(er,get_column_letter(2 + len(data.titles) + 1),r+1)))

    setCellWidth(sh, [5, 45])
    sh.freeze_panes = "C4"
    
class XLBuilderEx(XLBuilder):
   def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment. wrap_text = True
        
        if column > 1:
            cell.style.alignment.text_rotation = 90

        return column   
      
    
def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    data  = loadData(server)
    wb = printOut(data)
    workbookToObject(wb, "mtxtime.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    