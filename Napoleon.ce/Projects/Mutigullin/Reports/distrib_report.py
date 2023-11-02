# -*- coding: cp1251 -*-

import sys;
import locale
import time
import datetime
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment, Border, NumberFormat
from time import sleep
from openpyxl.writer.worksheet import row_sort

reload(sys);
sys.setdefaultencoding("cp1251")

class Report:
    __slots__ = ['data', 'orgs']
    
    def sections(self):
        res = list()
        
        for d in self.data.items:
            if d in self.orgs:
                res.append(self.orgs[d])
                
        return res
    
    def rows(self,s):
        res = list()
        
        if s.id in self.data.items:
            do = self.data.items[s.id]
            
            for i in do.items.values():
                res.append(i)
        
        res = sorted(res, cmp=lambda x,y: cmp(x.name, y.name))
                
        return res
    
    def columns(self, s):
        res = list()
        
        if s.id in self.data.items:
            do = self.data.items[s.id]
            
            for c in do.columns.values():
                res.append(c)
        
        res = sorted(res, cmp=item_cmp)
        return res
    
    def getVal(self, s, i, k):
        res = ''
        
        if s.id in self.data.items:
            do = self.data.items[s.id]
            
            if i.id in do.items:
                di = do.items[i.id]
                
                if k in di.values:
                    res = str(di.values[k])
                    
        return res
        
class Data:
    __slots__ = ['items']
    
    def __init__(self):
        self.items = dict()
        
class DataOrg:
    __slots__ = ['items', 'columns']
    
    def __init__(self):
        self.items = dict()
        self.columns = dict()
        
    @staticmethod    
    def key(d, t):
        return "{0}:{1}".format(d.strftime("%d.%m.%y"), t)
    
    def initColumn(self, k ,d, t):
        if not k in self.columns:
            dd = DataDoc()
            dd.date = d
            dd.type = t
            dd.key = k
            
            self.columns[k] = dd
                    
class DataDoc:
    __slots__ = ['date', 'type', 'key']
    pass

class DataItem:
    __slots__ = ['id', 'name', 'values', 'netorg']
    
    def __init__(self, id, name, netorg):
        self.id = id
        self.name = name
        self.values = dict()
        self.netorg = netorg
        
    def add(self, k, qty):
        if not k in self.values:
            self.values[k] = qty
        else:        
            self.values[k] += qty
            
class NetOrg:
    __slots__ = ['items']
    
    def __init__(self):
        self.items = dict()
        
class NetOrgData:
    __slots__ = ['items']
    
    def __init__(self):
        self.items = dict()
            
def netorgdata(orgs, orgid, priceid, netorg): 
    res = ''
    
    if orgid in orgs:
        o = orgs[orgid]
         
        if o.idNet in netorg.items:
            n = netorg.items[o.idNet]
            
            if priceid in n.items:
                res = n.items[priceid]
                
    return res
 
def loadDocs(data, docs, prices, type, netorg, orgs, params): 
    for d in docs:
        if d.id in orgs:
            netorgid = orgs[d.id].idNet
            
            if d.date >= params.start and d.date <= params.finish and netorgid == params.netorg: 
            
                k = DataOrg.key(d.date, type)
                
                if not d.id in data.items:
                    data.items[d.id] = DataOrg()
                    
                do = data.items[d.id]
                do.initColumn(k, d.date, type) 
                
                for i in d.items:
                    if not i.id in do.items:
                        do.items[i.id] = DataItem(i.id, prices[i.id].name if i.id in prices else i.id, netorgdata(orgs, d.id, i.id, netorg))
                        
                    di = do.items[i.id]
                    di.add(k, i.qty)
                       
def loadData(server, params):
    where = ''

    orgs = dict()
    prices = server.Get("ManagerPrice", "", "id")
    netorg = NetOrg()
    
    for n in server.Get("NetOrg", ""):
        if not n.id in netorg.items:
            netorg.items[n.id] = NetOrgData()
             
        noi = netorg.items[n.id]
         
        for u in n.items:
            noi.items[u.idSrc] = u.idNet
    
    userids = params.userids.split(",")
    data = Data() 
    
    where = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")'.format(params.start.strftime("%d/%m/%Y"), 
        params.finish.strftime("%d/%m/%Y"))
    
    for id in userids:
        server.ChangeUser("'"+id+"'")
        
        o = server.Get("Org", "", "id")
        p = server.Get("PotenzialOrg", "", "id")
        o.update(p)
        
        dlv = server.Get("Delivery", '');
        rmn = server.Get("OrgRemnants", where);
        
        server.RestoreUser()
        
        for k in o.keys():
            if not k in orgs:
                orgs[k] = o[k]
                 
        loadDocs(data, dlv, prices, 1, netorg, orgs, params)
        loadDocs(data, rmn, prices, 2, netorg, orgs, params)
    
    report = Report()
    report.data = data
    report.orgs = orgs
            
    return report

def item_cmp(x, y):
    res = cmp(x.date, y.date)

    if res == 0:
        res = cmp(x.type, y.type)
        
    return res

def rangeBorders(range, bold=False):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = Border.BORDER_MEDIUM if bold else Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_MEDIUM if bold else Border.BORDER_THIN
            cell.style.borders.top.border_style = Border.BORDER_MEDIUM if bold else Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_MEDIUM if bold else Border.BORDER_THIN
       
def printOut(report, params):
    xlb = XLBuilder()
     
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
     
    row = 0
    clmnsz = 0
    bkgColor = "ffC0C0C0"
    
    try:
        for s in report.sections():
            sheet.merge_cells(start_row=row, start_column=0, end_row=row+2, end_column=0)
            c = sheet.cell(row=row, column=0)
            c.value = '№ п/п'
            c.style.alignment.wrap_text = True
            xlb.paintHeadCell(c)
            xlb.setBackColor(c, bkgColor)
            
            sheet.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=1)
            c = sheet.cell(row=row, column=1)
            c.value = 'код сети'
            c.style.alignment.wrap_text = True
            xlb.paintHeadCell(c)
            xlb.setBackColor(c, bkgColor)
            
            sheet.merge_cells(start_row=row, start_column=2, end_row=row+2, end_column=2)
            c = sheet.cell(row=row, column=2)
            c.value = 'продукты'
            c.style.alignment.wrap_text = True
            xlb.paintHeadCell(c)
            xlb.setBackColor(c, bkgColor)
            
            column = 3
            c = sheet.cell(row=row, column=3)
            c.value = s.name
            c.style.alignment.wrap_text = True
            xlb.paintHeadCell(c)
            xlb.setBackColor(c, bkgColor)
            
            docs = report.columns(s)
            sz = len(docs)
            
            if clmnsz < sz:
                clmnsz = sz
            
            for d in docs:
                c = sheet.cell(row=row + 1, column=column)
                c.value = d.date.strftime("%d.%m.%Y")
                xlb.paintHeadCell(c)
                xlb.setBackColor(c, bkgColor)
                
                c = sheet.cell(row=row + 2, column=column)
                c.value = "поставка" if d.type == 1 else "ост"
                c.style.alignment.wrap_text = True
                xlb.paintHeadCell(c)
                xlb.setBackColor(c, bkgColor)
                
                column += 1
            
            sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=column-1)
            
            rangeBorders(sheet.range("{0}{1}:{2}{3}".format(get_column_letter(1), row+1, get_column_letter(sz+3), row+3)), True)
            
            row += 3
            
            dataRow = row
            num = 1
            for r in report.rows(s):
                c = sheet.cell(row=row, column=0)
                c.value = num
                
                c = sheet.cell(row=row, column=1)
                c.value = r.netorg
                
                c = sheet.cell(row=row, column=2)
                c.value = r.name
                
                column = 3
                
                for d in docs:
                    c = sheet.cell(row=row, column=column)
                    c.value = report.getVal(s, r, d.key)
                    
                    column += 1
                    
                row += 1
                num += 1    
                
            rangeBorders(sheet.range("{0}{1}:{2}{3}".format(get_column_letter(1), dataRow+1, get_column_letter(sz+3), row)))
            row += 1
            
        cc = 1
        wdh = [11,11,50]
        
        for i in range(0, clmnsz):
            wdh.append(11)
            
        for w in wdh:
            sheet.column_dimensions[get_column_letter(cc)].width = w
            cc += 1
    except:
        pass
     
    return wb

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting")
    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))
    report = loadData(server, params)
    wb = printOut(report, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info("ended")