# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

routeDataDict = dict()
summary = dict()
scheduleStart = dict()

def getDivisionAgents(server, division, agents):
    if division == None or len(division) == 0:
        return
    
    for d in division :
        for a in d.agents :
            agents.append(a.id)
    
        getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)
      
def loadAgents(server, division):
    ret = dict()
    
    agents = server.Get("Agents", "")
    
    divagents = []
    getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
    divagents = set(divagents)
    
    for a in agents:
        if a.id in divagents:
            ret[a.id] = a
    
    return ret    

def makeIDStr(server, agents):
    res = '"userid" in ('
    
    for id in agents.iterkeys():
        res += "'" + id + "',"
    
    res = res[:-1] + ")"
    return res

def borders(sheet, sr, c1, c2):
    for cl in range(c1, c2) :
       c = sheet.cell(row=sr, column=cl)
       c.style.borders.top.border_style = Border.BORDER_THIN
       c.style.borders.bottom.border_style = Border.BORDER_THIN
       c.style.borders.right.border_style = Border.BORDER_THIN 

def doReportMin(sheet, start, stop, agents):
    sr = 0
    sheet.cell(row=sr, column=0).value = "период, дн.:  " + str((stop - start).days + 1)
    sheet.row_dimensions[sheet.cell(row=sr, column=0).row].height = 30
    col_letter = get_column_letter(1) 
    sheet.column_dimensions[col_letter].width = 16 
    sheet.cell(row=sr, column=1).value = "план"
    sheet.cell(row=sr, column=2).value = "посетил"
    sheet.cell(row=sr, column=3).value = "не посетил"
    col_letter = get_column_letter(4) 
    sheet.column_dimensions[col_letter].width = 16
    sheet.cell(row=sr, column=4).style.alignment.wrap_text = True
    sheet.cell(row=sr, column=4).value = "время работы в ч"
    col_letter = get_column_letter(5) 
    sheet.column_dimensions[col_letter].width = 16
    sheet.cell(row=sr, column=5).style.alignment.wrap_text = True
    sheet.cell(row=sr, column=5).value = "время в точках"
    col_letter = get_column_letter(6) 
    sheet.column_dimensions[col_letter].width = 11
    sheet.cell(row=sr, column=6).style.alignment.wrap_text = True
    sheet.cell(row=sr, column=6).value = "большой перерыв"
    col_letter = get_column_letter(7) 
    sheet.column_dimensions[col_letter].width = 16
    sheet.cell(row=sr, column=7).style.alignment.wrap_text = True
    sheet.cell(row=sr, column=7).value = "время без б перерыва"
    borders(sheet, sr, 0, 8)
          
    plan = 0
    visit = 0
    unvisit = 0
    fullWt = None
    orgTime = None
    bigBreak = None
    nonstop = None
    
    sr += 1          
    for a in agents:
        rd = summary[a]
        
        sheet.cell(row=sr, column=0).value = a.name
        sheet.cell(row=sr, column=1).value = str(rd.plan)
        sheet.cell(row=sr, column=2).value = str(rd.visit)
        sheet.cell(row=sr, column=3).value = str(rd.unvisits)
        
        if rd.fullTime != None:
            sheet.cell(row=sr, column=4).value = "{0:0.1f}".format(float(rd.fullTime.seconds) / float(60 * 60))    
        
        if rd.orgTime != None:
            sheet.cell(row=sr, column=5).value = str(rd.orgTime.seconds / 60)
        
        if rd.bigBreak != None:
            sheet.cell(row=sr, column=6).value = str(rd.bigBreak.seconds / 60)   
            
        if rd.nonstop != None:
            sheet.cell(row=sr, column=7).value = str(rd.nonstop.seconds / 60)
        
        borders(sheet, sr, 0, 8)
        
        plan += rd.plan
        visit += rd.visit
        unvisit += rd.unvisits
        
        if rd.fullTime != None:
            if fullWt == None:
                fullWt = rd.fullTime
            else:
                fullWt += rd.fullTime
                
        if rd.orgTime != None:    
            if orgTime == None:
                orgTime = rd.orgTime
            else:
                orgTime += rd.orgTime
                
        if rd.bigBreak != None:        
            if bigBreak == None:
                bigBreak = rd.bigBreak
            else:
                bigBreak += rd.bigBreak
           
        if rd.nonstop != None:   
            if nonstop == None:
                nonstop = rd.nonstop
            else:
                nonstop += rd.nonstop
        
        sr += 1
            
    sheet.cell(row=sr, column=1).value = str(plan)
    sheet.cell(row=sr, column=2).value = str(visit)
    sheet.cell(row=sr, column=3).value = str(unvisit)
    
    if fullWt != None:
        sheet.cell(row=sr, column=4).value = "{0:0.1f}".format(float(fullWt.seconds) / float(60 * 60))
        
    if orgTime != None:    
        sheet.cell(row=sr, column=5).value = str(orgTime.seconds / 60)
    
    if bigBreak != None:    
        sheet.cell(row=sr, column=6).value = str(bigBreak.seconds / 60)
        
    if nonstop != None:    
        sheet.cell(row=sr, column=7).value = str(nonstop.seconds / 60)
    
    borders(sheet, sr, 0, 8)
        
            
def doReportByDay(sheet, start, stop, agents):
    sr = 0
    while start <= stop:
        plan = 0
        visit = 0
        unvisit = 0
        fullWt = None
        orgTime = None
        bigBreak = None
        nonstop = None
            
        sheet.cell(row=sr, column=0).value = start.strftime("%d/%m/%Y")
        sheet.row_dimensions[sheet.cell(row=sr, column=0).row].height = 30
        col_letter = get_column_letter(1) 
        sheet.column_dimensions[col_letter].width = 11 
        sheet.cell(row=sr, column=1).value = "план"
        sheet.cell(row=sr, column=2).value = "посетил"
        sheet.cell(row=sr, column=3).value = "не посетил"
        col_letter = get_column_letter(4) 
        sheet.column_dimensions[col_letter].width = 11 
        sheet.cell(row=sr, column=4).value = "в первой точке"
        col_letter = get_column_letter(5) 
        sheet.column_dimensions[col_letter].width = 16
        sheet.cell(row=sr, column=5).style.alignment.wrap_text = True
        sheet.cell(row=sr, column=5).value = "в последней точке"
        col_letter = get_column_letter(6) 
        sheet.column_dimensions[col_letter].width = 16
        sheet.cell(row=sr, column=6).style.alignment.wrap_text = True
        sheet.cell(row=sr, column=6).value = "время работы в ч"
        col_letter = get_column_letter(7) 
        sheet.column_dimensions[col_letter].width = 16
        sheet.cell(row=sr, column=7).style.alignment.wrap_text = True
        sheet.cell(row=sr, column=7).value = "время в точках"
        col_letter = get_column_letter(8) 
        sheet.column_dimensions[col_letter].width = 11
        sheet.cell(row=sr, column=8).style.alignment.wrap_text = True
        sheet.cell(row=sr, column=8).value = "большой перерыв"
        col_letter = get_column_letter(9) 
        sheet.column_dimensions[col_letter].width = 16
        sheet.cell(row=sr, column=9).style.alignment.wrap_text = True
        sheet.cell(row=sr, column=9).value = "время без б перерыва"
        col_letter = get_column_letter(10) 
        sheet.column_dimensions[col_letter].width = 16
        sheet.cell(row=sr, column=10).style.alignment.wrap_text = True
        borders(sheet, sr, 0, 10)  
        managers = routeDataDict[start]
        sr += 1
        
        for a in agents:
            rd = managers[a]
            
            sheet.cell(row=sr, column=0).value = a.name
            sheet.cell(row=sr, column=1).value = str(rd.plan)
            sheet.cell(row=sr, column=2).value = str(rd.visit)
            sheet.cell(row=sr, column=3).value = str(rd.unvisits)
            
            if rd.start != None:
                sheet.cell(row=sr, column=4).value = rd.start.strftime("%d/%m/%Y %H:%M")
                
            if rd.stop != None:
                sheet.cell(row=sr, column=5).value = rd.stop.strftime("%d/%m/%Y %H:%M")
            
            if rd.fullTime != None:
                sheet.cell(row=sr, column=6).value = "{0:0.1f}".format(float(rd.fullTime.seconds) / float(60 * 60))    
            
            if rd.orgTime != None:
                sheet.cell(row=sr, column=7).value = str(rd.orgTime.seconds / 60)
            
            if rd.bigBreak != None:
                sheet.cell(row=sr, column=8).value = str(rd.bigBreak.seconds / 60)   
                
            if rd.nonstop != None:
                sheet.cell(row=sr, column=9).value = str(rd.nonstop.seconds / 60)
            
            borders(sheet, sr, 0, 10)
            
            plan += rd.plan
            visit += rd.visit
            unvisit += rd.unvisits
            
            if rd.fullTime != None:
                if fullWt == None:
                    fullWt = rd.fullTime
                else:
                    fullWt += rd.fullTime
                    
            if rd.orgTime != None:    
                if orgTime == None:
                    orgTime = rd.orgTime
                else:
                    orgTime += rd.orgTime
                    
            if rd.bigBreak != None:        
                if bigBreak == None:
                    bigBreak = rd.bigBreak
                else:
                    bigBreak += rd.bigBreak
               
            if rd.nonstop != None:   
                if nonstop == None:
                    nonstop = rd.nonstop
                else:
                    nonstop += rd.nonstop
            
            sr += 1
            
        sheet.cell(row=sr, column=1).value = str(plan)
        sheet.cell(row=sr, column=2).value = str(visit)
        sheet.cell(row=sr, column=3).value = str(unvisit)
        
        if fullWt != None:
            sheet.cell(row=sr, column=6).value = "{0:0.1f}".format(float(fullWt.seconds) / float(60 * 60))
            
        if orgTime != None:    
            sheet.cell(row=sr, column=7).value = str(orgTime.seconds / 60)
        
        if bigBreak != None:    
            sheet.cell(row=sr, column=8).value = str(bigBreak.seconds / 60)
            
        if nonstop != None:    
            sheet.cell(row=sr, column=9).value = str(nonstop.seconds / 60)
            
        borders(sheet, sr, 0, 10)        
        start += timedelta(days=1)
        sr += 2
        
class RouteData:
    plan = 0
    visit = 0
    unvisits = 0
    start = None
    stop = None
    fullTime = None
    orgTime = None
    bigBreak = None
    nonstop = None

def getWeekIndex(server, data, agentid):  
    scStart = None
     
    if agentid in scheduleStart: 
        scStart = scheduleStart[agentid]
    else :
        where = '"userid"' + " in ('" + agentid + "')"
        cfg = server.Get("ServerConfig", where)
        
        if cfg != None:
            for c in cfg:
                if c.key == 'SheduleStart':
                    strptime = lambda date_string, format: datetime(*(time.strptime(date_string, format)[0:6])) 
                    scStart = strptime (c.value, '%Y-%m-%d')
                    break
                
        scheduleStart[scStart] = scStart;
    
    result = -1
    
    if scStart != None:
        d = data - scStart
        result = ((d.days / 7) % 4) + 1;
    
    return result    
               
def run(server):
    print "wt_report start"
    print ""
    
    #getcontext().prec = 2 
    #getcontext().rounding = ROUND_05UP
    
    params = server.Params
    param = params[0]
    days = {"Monday" : "Понедельник", 
            "Tuesday" : "Вторник",
            "Wednesday" : "Среда",
            "Thursday" : "Четверг",
            "Friday" : "Пятница",
            "Saturday" : "Суббота",
            "Sunday" : "Воскресенье",
    };
    
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    
    wb = Workbook(False, 'cp1251')
    sheet = None
    sheet2 = None
    sheet3 = None
    
    sheet = wb.get_active_sheet()
    sheet.title = "развернутый по менеджеру"
    
    sheet2 = wb.create_sheet()
    sheet2.title = "развернутый по дню"
    
    sheet3 = wb.create_sheet()
    sheet3.title = "свернутый"
    
    sr = 0
       
    if param.divisionID != 0 :
        agents = loadAgents(server, param.divisionID)
        
        al = list()
        
        for a in agents.values():
            al.append(a)
        
        al.sort(cmp=lambda x,y: cmp(x.name.lower(), y.name.lower()))
            
        for a in al:
            sheet.cell(row=sr, column=0).value = a.name
            sheet.row_dimensions[sheet.cell(row=sr, column=0).row].height = 30
            col_letter = get_column_letter(1) 
            sheet.column_dimensions[col_letter].width = 21 
            sheet.cell(row=sr, column=1).value = "план"
            sheet.cell(row=sr, column=2).value = "посетил"
            sheet.cell(row=sr, column=3).value = "не посетил"
            col_letter = get_column_letter(4) 
            sheet.column_dimensions[col_letter].width = 11 
            sheet.cell(row=sr, column=4).value = "в первой точке"
            col_letter = get_column_letter(5) 
            sheet.column_dimensions[col_letter].width = 16
            sheet.cell(row=sr, column=5).style.alignment.wrap_text = True
            sheet.cell(row=sr, column=5).value = "в последней точке"
            col_letter = get_column_letter(6) 
            sheet.column_dimensions[col_letter].width = 16
            sheet.cell(row=sr, column=6).style.alignment.wrap_text = True
            sheet.cell(row=sr, column=6).value = "время работы в ч"
            col_letter = get_column_letter(7) 
            sheet.column_dimensions[col_letter].width = 16
            sheet.cell(row=sr, column=7).style.alignment.wrap_text = True
            sheet.cell(row=sr, column=7).value = "время в точках"
            col_letter = get_column_letter(8) 
            sheet.column_dimensions[col_letter].width = 11
            sheet.cell(row=sr, column=8).style.alignment.wrap_text = True
            sheet.cell(row=sr, column=8).value = "большой перерыв"
            col_letter = get_column_letter(9) 
            sheet.column_dimensions[col_letter].width = 16
            sheet.cell(row=sr, column=9).style.alignment.wrap_text = True
            sheet.cell(row=sr, column=9).value = "время без б перерыва"
            col_letter = get_column_letter(10) 
            sheet.column_dimensions[col_letter].width = 16
            sheet.cell(row=sr, column=10).style.alignment.wrap_text = True
            
            for cl in range(0, 10) :
                   c = sheet.cell(row=sr, column=cl)
                   c.style.borders.top.border_style = Border.BORDER_THIN
                   c.style.borders.bottom.border_style = Border.BORDER_THIN
                   c.style.borders.right.border_style = Border.BORDER_THIN  
            
            orgFolder = server.Get("OrgFolder", '"userid"' + " in ('" + a.id + "')")  
           
            b = param.begin
            plan = 0
            visit = 0
            unvisit = 0
            fullWt = None
            orgTime = None
            bigBreak = None
            nonstop = None
            
            while(b <= param.end):
                d = days[b.strftime("%A")]
                widx = getWeekIndex(server, b, a.id)
                
                plans = list()
                visited = list()
                
                if orgFolder != None:
                    for of in orgFolder:
                        if of.name == d or of.name == str(widx) + d:
                            for i in of.items:
                                plans.append(i.name) 
                
                dailyRoute = server.Get("DailyRoute", '"date"' + " = ToDate('" + b.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                                        '"userid"' + " in ('" + a.id + "')")
                
                if dailyRoute != None:
                    for dr in dailyRoute:
                        for i in dr.items:
                            if not i.id in plans:
                                plans.append(i.id)
                
                whereStr = '"created"' + " >= ToDate('" + b.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                    '"created"' + " <= ToDate('" + b.strftime("%d/%m/%Y 23:59:59") + "') and " +\
                    '"userid"' + " in ('" + a.id + "')"
                    
                orders = server.Get("Order", whereStr)
                orders.sort(cmp=lambda x,y: cmp(x.id.lower(), y.id.lower()))
                
                for o in orders:
                    if visited.count(o.id) == 0 :
                       visited.append(o.id)
                
                visits = server.Get("VisitInfo", whereStr)
               
                for v in visits:
                    if visited.count(v.id) == 0 :
                       visited.append(v.id)
                       
                remnants = server.Get("OrgRemnants", whereStr)       
                
                for r in remnants:
                    if visited.count(r.id) == 0 :
                       visited.append(r.id)
            
                monitorings = server.Get("Monitoring", whereStr)       
                if monitorings != None :
                    for m in monitorings:
                        if visited.count(m.id) == 0 :
                           visited.append(m.id)
                       
                visit += len(visited)   
                plan += len(plans)
                
                uv = 0 
                for p in plans:
                    if visited.count(p) == 0 :
                        uv += 1
                
                unvisit += uv        
                wtSql = '"start"' + " >= ToDate('" + b.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                        '"start"' + " <= ToDate('" + b.strftime("%d/%m/%Y 23:59:59") + "') and "+\
                        '"userid"' + " in ('" + a.id + "') order by " + '"start"'
                worktimes = server.Get("WorkTime", wtSql)
                
                start = None
                stop = None
                fwt = None
                tio = None
                bp = None
                dp = None
                
                for worktime in worktimes:   
                    if start == None:
                        start = worktime.start
                        
                    if stop != None:
                        bpt = worktime.start - stop
                        
                        if bpt.total_seconds() > 30 * 60:
                            if bp == None:
                                bp = bpt
                            else:
                                bp += bpt
                                 
                    stop = worktime.stop
                    
                    #print "agent: " +a.name.encode("cp866")
                    #print "start: " + worktime.start.strftime("%d/%m/%Y %H:%M")
                    #print "stop: " + worktime.stop.strftime("%d/%m/%Y %H:%M")
                            
                    if tio == None:
                        tio = worktime.stop - worktime.start
                    else:    
                        tio += worktime.stop - worktime.start
                    
                if stop != None and start != None:    
                    fwt = stop - start
                
                if fwt != None and bp != None:
                   dp = fwt - bp 
                    
                sr += 1  
                sheet.cell(row=sr, column=0).value = b.strftime("%d/%m/%Y %H:%M")
                sheet.cell(row=sr, column=1).value = str(len(plans))
                sheet.cell(row=sr, column=2).value = str(len(visited))
                sheet.cell(row=sr, column=3).value = str(uv)
                
                if start != None:
                    sheet.cell(row=sr, column=4).value = start.strftime("%d/%m/%Y %H:%M")
                    
                if stop != None:
                    sheet.cell(row=sr, column=5).value = stop.strftime("%d/%m/%Y %H:%M")
                
                #if fwt != None:
                #    print "fwt.seconds: " + str(fwt.seconds)
                #    print Decimal(Decimal(fwt.seconds) / (60 * 60))
                    
                if fwt != None:
                    sheet.cell(row=sr, column=6).value ="{0:0.1f}".format(float(fwt.total_seconds()) / float(60 * 60))    
                
                if tio != None:
                    sheet.cell(row=sr, column=7).value = str(int(tio.total_seconds() / 60))
                
                if bp != None:
                    sheet.cell(row=sr, column=8).value = str(int(bp.total_seconds() / 60))   
                    
                if dp != None:
                    sheet.cell(row=sr, column=9).value = str(int(dp.total_seconds() / 60))
                    
                borders(sheet, sr, 0, 10)        
                                                   
                if fwt != None:
                    if fullWt == None:
                        fullWt = fwt
                    else:
                        fullWt += fwt
                    
                if tio != None:    
                    if orgTime == None:
                        orgTime = tio
                    else:
                        orgTime += tio
                        
                if bp != None:        
                    if bigBreak == None:
                        bigBreak = bp
                    else:
                        bigBreak += bp
                   
                if dp != None:   
                    if nonstop == None:
                        nonstop = dp
                    else:
                        nonstop += dp     
               
                routeData = RouteData()
                routeData.plan = len(plans)
                routeData.visit = len(visits)
                routeData.unvisits = uv
                routeData.start = start
                routeData.stop = stop
                routeData.fullTime = fwt
                routeData.orgTime = tio
                routeData.bigBreak = bp
                routeData.nonstop = dp 
                
                if b in routeDataDict:
                    d = routeDataDict[b]
                else:
                    d = dict()
                    routeDataDict[b] = d 
                
                d[a] = routeData
                
                b += timedelta(days=1)
                        
            sr += 1
            sheet.cell(row=sr, column=1).value = str(plan)
            sheet.cell(row=sr, column=2).value = str(visit)
            sheet.cell(row=sr, column=3).value = str(unvisit)
            
            if fullWt != None:
                sheet.cell(row=sr, column=6).value = "{0:0.1f}".format(float(fullWt.total_seconds()) / float(60 * 60))
                
            if orgTime != None:    
                sheet.cell(row=sr, column=7).value = str(int(orgTime.total_seconds() / 60))
            
            if bigBreak != None:    
                sheet.cell(row=sr, column=8).value = str(int(bigBreak.total_seconds() / 60))
                
            if nonstop != None:    
                sheet.cell(row=sr, column=9).value = str(int(nonstop.total_seconds() / 60))
            
            borders(sheet, sr, 0, 10)   
            sr += 2
            
            routeData = RouteData()
            routeData.plan = plan
            routeData.visit = visit
            routeData.unvisits = unvisit
            routeData.fullTime = fullWt
            routeData.orgTime = orgTime
            routeData.bigBreak = bigBreak
            routeData.nonstop = nonstop
            
            summary[a] = routeData
            
        doReportByDay(sheet2, param.begin, param.end, al)
        doReportMin(sheet3, param.begin, param.end, al)
            
    print ""
    print "wt_report done " ;
    
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    server.Put(outObj)