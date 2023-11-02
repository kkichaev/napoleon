# -*- coding: cp1251 -*-
import sys;
import logging
import locale
import time

from grsoft import xl_base
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Fill, Color, Alignment, NumberFormat, Border
from datetime import datetime, date, timedelta
from manager import summary, coordutils
from strings import Resources

reload(sys);
sys.setdefaultencoding("cp1251")

class AgentGps:
    id = None
    items = None
    
    def __init__(self, id):
        self.id = id
        self.items = list()
        
    def add(self, gps):
        self.items.append(gps)
        
    def distance(self, date):
        result = 0
        
        lastpos = None
        
        for g in self.items:
            if g.date.date() == date:
                if lastpos == None:
                    lastpos = g
                    continue

                result += coordutils.distance(lastpos.latitude, lastpos.longitude, g.latitude, g.longitude)
                lastpos = g
                
        return result
        
class Data:
    items = None
    
    def __init__(self):
        self.items = dict()
        
    def add(self, gps):
        if not gps.userid in self.items:
            self.items[gps.userid] = AgentGps(gps.userid)
        
        self.items[gps.userid].add(gps)
        
    def getItems(self, date):
        result = []
        
        for v in self.items.values():
            result.append([v.id, v.distance(date)])
            
        return result    
        
class Report:
    start = None
    finish = None
    users = None
    data = None
    summary = None
    res = None
    
    def __init__(self, params):
        self.start = params.start
        self.finish = params.finish
        self.res = Resources(params.locale)
        self.data = Data()
        self.summary = dict()
    
    def load(self, server):
        gps = server.Get("GPSPos", self.compileWhere()) 
        self.users = server.Get("Agents", "","id")
        
        st = self.start.time()
        ft = self.finish.time()
        
        if ft.hour * ft.minute * ft.second * ft.microsecond == 0:
            ft = ft.replace(23,59,59,999999)
        
        for g in gps:
            gt = g.date.time()
            
            if st <= gt and ft >= gt:
                self.data.add(g)
        
    def compileWhere(self):
        where = '"date" >= ToDate("{0}") and "date" <= ToDate("{1}" and isGSM=0)'.format(
            self.start.strftime("%d/%m/%Y %H:%M:%S"), self.finish.strftime("%d/%m/%Y %H:%M:%S")) 
            
        return where  

    def getItems(self, date):
        result = []
        for i in self.data.getItems(date):
            if not i[0] in self.summary:
                self.summary[i[0]] = 0
                
            self.summary[i[0]] += i[1]    
            result.append([self.getUserName(i[0]), i[1] / 1000])
        
        return result    
        
    def getSummary(self):
        result = []
        
        for k, v in self.summary.items():
            result.append([self.getUserName(k), v / 1000])
        
        return result
        
    def getUserName(self, id):
        return self.users[id].name if id in self.users else "{1} {0}".format(id, Resources.getString("unknown_agent"))
        
    
def createReport(server):
    result = Report(server.Params[0])
    result.load(server)
    return result
    
class XLB(xl_base.XLBuilder):
    header_table_color = Color('FFccccff')
    work_day_color = Color('FFaaff80')
    free_day_color = Color('FFFF0000')
    
    def adjustHeadCell(self, sheet, cell, row, column):
        fill = cell.style.fill;
        fill.fill_type = Fill.FILL_SOLID
        fill.start_color = self.header_table_color
        
        return column
    
    def makeCell(self, sheet, row, column, cell, value):
        xl_base.XLBuilder.makeCell(self, sheet, row, column, cell, value)        
        
        if column == 1:
            cell.style.number_format.format_code = '0.0'
            
    def drawDayCell(self, sheet, row, value, isWork):
        self.makeCell(sheet, row, 0, sheet.cell(row=row, column=0), value)
        self.makeCell(sheet, row, 0, sheet.cell(row=row, column=1), "")
        sheet.merge_cells(start_row=row, start_column=0, end_row=row, end_column=1)
        c = sheet.cell(row=row, column=0)
        fill = c.style.fill;
        fill.fill_type = Fill.FILL_SOLID
        fill.start_color = self.work_day_color if isWork else self.free_day_color
            
def printOut(report):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = report.res.getString("report")
    sheet.merge_cells(start_row=0, start_column=0, end_row=0, end_column=1)
    c = sheet.cell(row = 0, column = 0)
    c.value = report.res.getString("mileage_report")
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.font.bold = True
    c.style.font.size = 14
    sheet.merge_cells(start_row=1, start_column=0, end_row=1, end_column=1)
    c = sheet.cell(row = 1, column = 0)
    c.value = "{4} {0} по {1} {2}-{3}".format(
        report.start.strftime("%d.%m.%Y"), report.finish.strftime("%d.%m.%Y"),
        report.start.strftime("%H:%M"), report.finish.strftime("%H:%M"),
        report.res.getString("report_period_from")
    )
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.font.bold = True
    
    c = sheet.cell(row = 2, column = 0)
    
    xlb = XLB()
    xlb.makeHead(sheet, 2, [report.res.getString("performer"), report.res.getString("distance")])
    day_array = [report.res.getString("sun"), 
                  report.res.getString("mon"), 
                  report.res.getString("tue"),
                  report.res.getString("wed"), 
                  report.res.getString("thu"),
                  report.res.getString("fri"),
                  report.res.getString("sut")]
    
    s = report.start
    row = 3
    
    while s < report.finish:
        dc = int(s.strftime("%w"))
        value = "{0} ({1})".format(day_array[dc], s.strftime("%d.%m.%Y"))
        xlb.drawDayCell(sheet, row, value, dc > 0 and dc < 6)
        row += 1
        
        for item in report.getItems(s.date()):
            xlb.makeCells(sheet, row, item)
            row += 1
            
        s = s + timedelta(days=1)
    
    
    c = sheet.cell(row = row, column = 0)
    xlb.makeHead(sheet, row, [report.res.getString("result_data"), ""])
    sheet.merge_cells(start_row=row, start_column=0, end_row=row, end_column=1)
    row += 1
    xlb.makeHead(sheet, row, [report.res.getString("performer"), report.res.getString("distance")])
    row += 1
    
    for item in report.getSummary():
        xlb.makeCells(sheet, row, item)
        row += 1
    
    cc = 1
    wdh = [45,45]
    for w in wdh:
        sheet.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
    return wb

def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.debug("starting")

    locale.setlocale(locale.LC_ALL, 'american')
    params = server.Params[0]
    logging.info("params " + str(params))
    report = createReport(server)
    wb = printOut(report)
    xl_base.XLBuilder().workbookToObject(wb, "distance", server)                

    logging.info("ended")
