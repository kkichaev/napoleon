# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import NumberFormat
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys
import time
from manager import coordutils

reload(sys);
sys.setdefaultencoding("cp1251")

class Report:
    start = None
    finish = None
    items = None
    
    def __init__(self):
        self.start = datetime.now()
        self.finish = datetime.now()
        self.items = list()

class Item:
    division = None
    agent = None
    data = None
    start = None
    finish = None
    qty_in_route = None
    ord_in_route = None
    ord_out_route = None
    vizit_in_route = None
    vizit_out_route = None
    ptnz_cnt = None
    avg_sum = None
    distance = None
    
    def __init__(self):
      self.division = ""
      self.agent = ""
      self.data = None
      self.start = None
      self.finish = None
      self.qty_in_route = 0
      self.ord_in_route = 0
      self.ord_out_route = 0
      self.vizit_in_route = 0
      self.vizit_out_route = 0
      self.ptnz_cnt = 0
      self.avg_sum = 0
      self.distance = 0
      
    def getData(self, row):
      return [self.division, self.agent, self.data, self.start if self.start != None else "", 
              self.finish if self.finish != None else "", "=E{0}-D{0}".format(row+1), self.qty_in_route, self.ord_in_route, self.ord_out_route,
              self.vizit_in_route, self.vizit_out_route, self.ptnz_cnt, self.avg_sum, self.distance]

def inflateParams(server):
    return server.Params[0].start, server.Params[0].finish, server.Params[0].userids

def getWeekIndex(cfg, data, id):  
    scStart = None
    result = -1 
        
    if cfg != None:
        for c in cfg:
            if c.key == 'SheduleStart' and len(c.value) > 0:
                scStart = datetime(*(time.strptime(c.value, '%Y-%m-%d')[0:6]))
                break
            
    if scStart != None:
        d = data - scStart
        result = ((d.days / 7) % 4) + 1;
    
    return result      
         
def plannedOrgs(of, cfg, id, date, days):
    plans = list()
    d = days[date.strftime("%A")]
    widx = getWeekIndex(cfg, date, id)
    
    for f in of:
        if f.name == d or f.name == str(widx) + d:
            for i in f.items:            
                plans.append(i.name)
                
    return plans

def collectDocs(docdict, docs):
    for d in docs:
        dt = d.created.replace(hour=0, minute=0, second=0, microsecond=0)
        key = d.userid + str(dt)
        
        if not key in docdict:
            docdict[key] = list();
            
        docdict[key].append(d) 
            
def loadData(server):
    start, finish, userids = inflateParams(server)
    
    report = Report()
    report.start = start
    report.finish = finish
    
    finish += timedelta(days=1)
    
    WHERE_STR_ALL = '"created" >= ToDate("{0}") and "created" < ToDate("{1}")';         
    WHERE_STR = WHERE_STR_ALL + ' and "userid" in ({2})'; 

    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userids)
    
    orders = server.Get("Order", where)
    visits = server.Get("VisitInfo", where)
    remnants = server.Get("OrgRemnants", where)
    questions = server.Get("Answer", where)
    incass = server.Get("Incass", where)
    tasks = server.Get("TaskDone", where)
    division = server.Get("Division", '')

    docdict = dict()
    orddict = dict()
    collectDocs(docdict, orders)
    collectDocs(orddict, orders)
    collectDocs(docdict, visits)
    collectDocs(docdict, questions)
    collectDocs(docdict, incass)
    collectDocs(docdict, tasks)
    
    divmap = dict()
    
    for d in division:
        for a in d.agents:
            if not a.id in divmap:
                divmap[a.id] = d.name 
    
    allOrgs = dict()
    datadict = dict(); #key = userid + data 
    agents = dict()
    
    days = {"Monday" : "Понедельник", 
        "Tuesday" : "Вторник",
        "Wednesday" : "Среда",
        "Thursday" : "Четверг",
        "Friday" : "Пятница",
        "Saturday" : "Суббота",
        "Sunday" : "Воскресенье" }
    
    data = list()
    
    for id in userids.split(','):
        server.ChangeUser(id)
        
        agent = server.CurrentUser()
        aid = agent.id
        aname = agent.name
        orgs = server.Get("Org", "", "id")
        porg = server.Get("PotenzialOrg", "", "id")
        orgs.update(porg)
        of = server.Get("OrgFolder", "")
        cfg = server.Get("ServerConfig", "")
        
        server.RestoreUser()
        
        ptzOrgCnt = 0
        
        for p in porg.values():
            if p.created >= start and p.created <= finish:
                ptzOrgCnt += 1
                
        for o in orgs.values():
            if not o.id in allOrgs:
                allOrgs[o.id] = o
        
        
                
        s = start
        
        while s < finish:
            p = plannedOrgs(of, cfg, aid, s, days)
            
            item = Item()
            item.agent = aname
            item.data = s
            item.division = divmap[aid]
            item.qty_in_route = len(p)
            
            where = '"date" > ToDate("{0}") and "date" <= ToDate("{1}") and "isGSM" = \'0\' and "userid"={2}'.format(
            s.strftime("%d/%m/%Y 0:0:0"), s.strftime("%d/%m/%Y 23:59:59"), id)
            
            gpspos = server.Get("GPSPos", where)
            lastpos = None
            gpspos = sorted(gpspos, cmp=lambda l, r: cmp(l.date, r.date))
            distance = 0
            
            for pos in gpspos:
                if lastpos == None:
                    lastpos = pos
                    continue
    
                distance += coordutils.distance(lastpos.latitude, lastpos.longitude, pos.latitude, pos.longitude)
                lastpos = pos
            
            distance /= 1000
            
            key = aid + str(s)
            ordInRouteCnt = 0
            ordNotInRouteCnt = 0
            visitRouteList = list()
            visitNotRouteList = list()
            ordSum = 0
            dcnt = 0
             
            if key in docdict:
                docs = sorted(docdict[key], cmp=lambda x,y: cmp(x.created,y.created)) 
                
                for d in docs:
                    if d.id in p:
                        if not d.id in visitRouteList:
                            visitRouteList.append(d.id)
                    else:
                        if not d.id in visitNotRouteList:
                            visitNotRouteList.append(d.id)
                        
                item.start = docs[0].created
                item.finish = docs[len(docs) - 1].sended
                
                ords = list()
                
                if key in orddict:
                    for o in orddict[key]:
                        dcnt += 1
                        
                        if o.id in p:
                            ordInRouteCnt += 1
                        else:
                            ordNotInRouteCnt += 1 
                        
                        for i in o.items:
                            ordSum += i.qty * i.cost

            item.ord_in_route = ordInRouteCnt
            item.ord_out_route = ordNotInRouteCnt
            item.vizit_in_route = len(visitRouteList)
            item.vizit_out_route = len(visitNotRouteList)
            item.avg_sum = ordSum / dcnt if dcnt > 0 else 0
            item.distance = distance
            item.ptnz_cnt = ptzOrgCnt
            data.append(item)
            
            s += timedelta(days=1)
           
    
    data = sorted(data, cmp=item_cmp ) 
    report.items.extend(data)       

    return report
    
def item_cmp(x, y):
  res = cmp(x.agent, y.agent)
  
  if res == 0:  
    res = cmp(x.data, y.data) * -1

  return res
  
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def ptintSheet(xlb, sh, data):
    sh.cell(row=1, column=0).value = "Период: {0} - {1}".format(data.start.strftime("%d.%m.%Y"),data.finish.strftime("%d.%m.%Y"))
    
    head = ["Подразделение", "Агент", "Дата", "Время начала", "Дата / Время завершения", "Время работы", "Кол-во ТТ в маршруте", "Кол-во заявок по маршруту", "Кол-во заявок не  маршруту",
            "Ко-во посешений по маршруту", "Ко-во посешений не маршруту", "Кол-во новых точек (Потенциальных)", "Средняя сумма заявок в день", "Маршрут в Км."]
    r = 2
    xlb.makeHead(sh, r, head)
    
    for d in data.items:
      r += 1
      xlb.makeCells(sh, r, d.getData(r))
    
    setCellWidth(sh, [24,30,14,10,20,10,10])
    
class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True
    return column
        
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 4:
      cell.style.number_format._set_format_code('yyyy-mm-dd hh:mm:ss')  
    if column == 3 or column == 5:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_DATE_TIME4)
    elif column == 12 or column == 13:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)  
          
      
    
def printOut(d):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheet(xlb, sh, d)
                
    return wb

def doReport(server):
    report = loadData(server)
    wb = printOut(report)
    workbookToObject(wb, "morkreport.xlsx", server)

def workbookToObject(wb, repName, server):
    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    