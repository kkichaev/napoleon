# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment

import datetime
from datetime import timedelta

import sys;
reload(sys);
sys.setdefaultencoding("cp1251")


class QuestHelper:
  LIST_TYPE = 2
  DATASET_TYPE = 5
  PHOTO_TYPE = 7
  ORG_DATASET_TYPE = "Организация"
  BOOL_TYPE = 4

class Quest:
  __slots__ = ["idquest", "quest", "rowidx", "answers"]
  
  def __init__(self, quest):
    self.idquest = quest.idquest
    self.quest = quest
    self.rowidx = dict()
    self.answers = list()
    self.orgs = list()
    
    idx = 0
    quest.items.sort(cmp= lambda x, y: cmp(x.number, y.number))
    
    for i in quest.items:
      if i.type == QuestHelper.LIST_TYPE or i.type == QuestHelper.BOOL_TYPE :
        for v in i.values:
          key = ReportData.ITEM_KEY_FMT.format(quest.idquest, i.iditem, v.value)
          self.rowidx[key] = idx     
          idx += 1          
      else:
        key = ReportData.KEY_FMT.format(quest.idquest, i.iditem)
        self.rowidx[key] = idx      
        idx += 1  

  def makeAnswer(self, d, r, pics, href):
    sz = len(self.rowidx)
    a = [""] * sz
    for n in d.items:
      if n.type ==QuestHelper.LIST_TYPE:
          self.putItemValue(a, ReportData.ITEM_KEY_FMT.format(d.question, n.iditem, n.answer), ["X", n.type2])
      elif n.type == QuestHelper.BOOL_TYPE :
          self.putItemValue(a, ReportData.ITEM_KEY_FMT.format(d.question, n.iditem, n.answer), [n.answer, n.type2])    
      elif n.type == QuestHelper.PHOTO_TYPE:
        val = '=HYPERLINK("{0}{1}", "Фото")'.format(href, pics[n.answer].name) if n.answer in pics else "Фото не найдено!"
        self.putItemValue(a, ReportData.KEY_FMT.format(d.question, n.iditem), [val, n.type2])
          
      elif n.type == QuestHelper.DATASET_TYPE:
        val = ''
        if n.remark == QuestHelper.ORG_DATASET_TYPE:
          val = r.orgs[n.answer].name if n.answer in r.orgs else n.answer
        
        self.putItemValue(a, ReportData.KEY_FMT.format(d.question, n.iditem), [val, n.type2])
      else: 
        self.putItemValue(a, ReportData.KEY_FMT.format(d.question, n.iditem), [n.answer, n.type2])
        
    return a
    
  def addanswer(self, d, r, pics, href):
      a = self.makeAnswer(d,r,pics, href)

      self.answers.append(a)
      self.orgs.append(d.id)
    
  def putItemValue(self, items, key, value):
    if key in self.rowidx:
      idx = self.rowidx[key]
      items[idx] = value

class ManagerDocs:
    __slots__ = ['manager', 'docs']
    
    def __init__(self, manager):
        self.manager = manager
        self.docs = dict()

    def add(self, doc, agents):
        qdocs = None
        if doc.question in self.docs :
            qdocs = self.docs[doc.question]
        else :
            qdocs = dict()
            self.docs[doc.question] = qdocs
            
        docDate = doc.created.date()
        
        agent = agents[doc.agentid].name if doc.agentid in agents else doc.agentid
        doc.agentid = agent  
        
        dateDocs = None
        if not docDate in qdocs:
            dateDocs = list()
            qdocs[docDate] = dateDocs
        else:
            dateDocs = qdocs[docDate]
            
        dateDocs.append(doc)
        
    
class ReportData:
  __slots__ = ['items', 'quests', 'orgs', 'agents', "answers",'managers','pics','href']
  KEY_FMT =  "{0}\t{1}"
  ITEM_KEY_FMT = "{0}\t{1}\t{2}"

  def __init__(self):
    self.items = dict()
    self.quests = dict()
    self.answers = dict()
    
  def prepare(self, docs, quests):
    qu = list()
    
    for d in docs:
      if not d.question in qu:
        qu.append(d.question)
    
    for q in qu:
      if q in quests:
        quest = quests[q]
        self.quests[q] = Quest(quest)

  def loadDocs(self, docs, pics, href):
        self.pics = pics
        self.href = href
        
        for d in docs:
            if not d.userid in self.managers: continue
            mgrDocs = None
            if not d.userid in self.answers: 
                mgr = self.managers[d.userid];
                mgrDocs = ManagerDocs(mgr);
                self.answers[d.userid] = mgrDocs
            else:
                mgrDocs = self.answers[d.userid]
            mgrDocs.add(d, self.agents)
            
def loadData(params, server):
    data = ReportData();
    
    orgs = dict()
    porg = server.Get("PotenzialOrg", "", "id")
    data.agents = server.Get("Agents", "", "id")
    data.managers = server.Get('DivisionManager', '', 'login')

    quests = server.Get("Question", '"idquest" is null or "idquest" is not null', "idquest")
    pictures = server.Get("PicStoreSrc", "", "id")
    orgs.update(porg)
    data.orgs = orgs
    
    if params.param == 0:
      for item in params.userids:
        server.ChangeUser(item.id)
        aorgs = server.Get("Org", '', 'id')
        server.RestoreUser()
        
        if aorgs != None:
            data.orgs.update(aorgs)

    where = '"created" >= ToDate("{0} 0:0:0") and "created" <= ToDate("{1} 0:0:0")'.format(params.start.strftime('%d.%m.%Y'), (params.finish + timedelta(days=1)).strftime('%d.%m.%Y'))
    
    if len(params.userids) > 0:
      arr = ''
      
      for i in params.userids:
        if len(arr) > 0:
          arr += ','
        
        arr += i.id
      
      where = '{0} and "userid" in ({1})'.format(where, arr)
      
    
    if len(params.quests) > 0:
      arr = ''
      
      for i in params.quests:
        if len(arr) > 0:
          arr += ','
        
        arr += i.id
      
      where = '{0} and "question" in ({1})'.format(where, arr)
    
    answers = server.Get("Answer" if params.param == 0 else "MAnswer", where)
    
    if params.param == 1:
      ids = []
      for a in answers:
        if len(a.id) > 0 and not a.id in ids: 
          ids.append(a.id)
          aorgs = server.Get("Org", '"id"=\'{0}\''.format(a.id), 'id')
          
          if aorgs != None:
            data.orgs.update(aorgs)
            
    data.prepare(answers, quests)
    data.orgs = orgs
    
    href = params.hrefBase
    data.loadDocs(answers, pictures, href)
    
    return data
    
class XLBuilderEx(XLBuilder):
  FIXED_CELL_COLOR = "FFB6DDE8"
  HEAD_COLOR = "FFD8D8D8"
  CELL_COLOR = "FFFFFFFG"
  YELLOW = "FFFFFF00"
  
  __slots__ = ['doRotation']
  
  def __init__(self):
      self.doRotation = True
  
  def adjustHeadCell(self, sheet, cell, row, column):
    if column == 0:
      for c in range(0,4):
        self.paintHeadCell(sheet.cell(row=row, column=c))
      column += 3
      sheet.merge_cells(start_row=row, start_column=0, end_row=row, end_column=3)
    else:
      s = cell.style
      if self.doRotation: s.alignment.text_rotation = 90
    
    cell.style.alignment.wrap_text = True    
    self.setBackColor(cell, self.HEAD_COLOR)
    
    return column
    
  def setupCell(self, cell, hor = Alignment.HORIZONTAL_CENTER, ver = Alignment.VERTICAL_CENTER, rot = 0, bold = True):
    style = cell.style
    style.alignment.horizontal = hor
    style.alignment.vertical = ver
    style.alignment.text_rotation = rot
    style.font.bold = True
  
  def printQuest(self, sheet, row, quest):
    start_row = row
    c = sheet.cell(row=row, column=0)
    c.value = quest.quest.name
    self.setupCell(c, Alignment.HORIZONTAL_LEFT, Alignment.VERTICAL_CENTER, 90)
    self.setBackColor(c, self.YELLOW)
    
    idx = 1
    
    for i in quest.quest.items:
      c = sheet.cell(row=row, column=1)
      c.value = idx
      self.setupCell(c)
      self.setBackColor(c, self.FIXED_CELL_COLOR)
      
      idx += 1
      
      c = sheet.cell(row=row, column=2)
      c.value = i.id
      self.setupCell(c, Alignment.HORIZONTAL_LEFT)
      self.setBackColor(c, self.FIXED_CELL_COLOR)
      
      if i.type == QuestHelper.BOOL_TYPE or i.type == QuestHelper.LIST_TYPE:
        sr = row
        
        for v in i.values:
          c = sheet.cell(row=row, column=3)
          c.value = v.value
          self.setupCell(c, Alignment.HORIZONTAL_LEFT)
          self.setBackColor(c, self.FIXED_CELL_COLOR)
          row += 1
          
        sheet.merge_cells(start_row=sr, start_column=1, end_row=row-1, end_column=1)    
        sheet.merge_cells(start_row=sr, start_column=2, end_row=row-1, end_column=2)  
      else:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)  
        row += 1
    
    sheet.merge_cells(start_row=start_row, start_column=0, end_row=row-1, end_column=0)
    
    x = 1
    for w in [3,3,15,15]:
      sheet.column_dimensions[get_column_letter(x)].width = w
      x += 1
    
    return row
    
  def printAnswers(self,sheet, row, col, answers):
      for a in answers:
        a1 = ""
        a2 = ""
        if len(a) >= 2:
          a1 = a[0]
          a2 = a[1]
        
        c = sheet.cell(row=row, column=col)
        c.value = a1
        c.style.alignment.wrap_text = True
        self.makeBorder(c, Border.BORDER_THIN)
        
        c = sheet.cell(row=row, column=col + 1)
        c.value = a2
        c.style.alignment.wrap_text = True
        self.makeBorder(c, Border.BORDER_THIN)
        
        row += 1
      
  def printData(self, sheet, row, data):
   
    maxCol = 0 
    mgrDocs = data.answers.values();
    mgrDocs.sort(cmp = lambda x,y : cmp(x.manager.name.upper(), y.manager.name.upper()))
    for mgr in mgrDocs:
        c = 0
        cell = sheet.cell(row=row, column=c)
        cell.value = "Руководитель:" + mgr.manager.name
        row += 1

        for quid, qdoc in mgr.docs.items():
            quest = data.quests[quid]

            headRow = row      
            head1 = ['Торговый агент']
            head2 = ['Дата']
            head3 = ["Клиент"]
            row += 3
            
            startRow = row
            row = self.printQuest(sheet, row, quest)
            row += 1
            
            c = 4
            dates = qdoc.keys()
            dates.sort()
            for date in dates:
                docs = qdoc[date]
                docs.sort(cmp = lambda x, y : cmp(x.agentid.upper, y.agentid.upper()))
                for doc in docs:
                    org = "{0} ({1})".format(data.orgs[doc.id].name, data.orgs[doc.id].address) if doc.id in data.orgs else doc.id
                    head1.append(doc.agentid)
                    head1.append(doc.agentid)
                    head2.append(date)
                    head2.append(date)
                    head3.append(org) 
                    head3.append(org) 
                    
                    answ = quest.makeAnswer(doc, data, data.pics, data.href)
                    self.printAnswers(sheet, startRow, c, answ)
                    c += 2
    
    
            self.doRotation = False
            self.makeHead(sheet, headRow, head1)
            self.makeHead(sheet, headRow + 1, head2)
            self.doRotation = True
            self.makeHead(sheet, headRow + 2, head3)
            
            print head1, len(head1)
            
            for c in range(4, len(head1) + 3, 2):
              sheet.merge_cells(start_row=headRow, start_column=c, end_row=headRow, end_column=c+1)
              sheet.merge_cells(start_row=headRow+1, start_column=c, end_row=headRow+1, end_column=c+1)
              sheet.merge_cells(start_row=headRow+2, start_column=c, end_row=headRow+2, end_column=c+1)
             
            sheet.row_dimensions[headRow +1].height = 50
            sheet.row_dimensions[headRow +3].height = 100
            
            if maxCol < len(head3) - 1: maxCol = len(head3) - 1
        
    
    for x in range(5, maxCol + 5):
      sheet.column_dimensions[get_column_letter(x)].width = 16
      x += 1
    
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilderEx()
    
    c = sheet.cell(row=0, column=0)
    c.value = "Отчет по анкетам"  
    c.style.font.bold = True
    c.style.font.size = 18

    c = sheet.cell(row=1, column=0)
    c.value = "Интервал: c {0} по {1}".format(params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y'))  
    
    r = 2
    xlb.printData(sheet, r, data)

    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
#     logging.info('start report')

    params = server.Params[0]
#     logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "quest_rep.xlsx", server)                
#     logging.info('end')
    
