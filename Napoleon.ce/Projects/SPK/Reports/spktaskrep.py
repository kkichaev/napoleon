# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color

import datetime


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

class Item:
    __slots__ = ['period', 'status', 'skill', 'strengths', 'razvitie', 'task']
    
    def __init__(self):  
      self.period = ''
      self.status = ''
      self.skill = ''
      self.strengths = ''
      self.razvitie = ''
      self.task = ''
      
    
    def values(self):
      return [self.period, "выполнено" if self.status == 1 else 'не выполнено', self.skill, self.strengths, self.razvitie, self.task]
      
class ReportData:
    __slots__ = ['data']
    
    def __init__(self):
      self.data = list()
        
    def add(self, doc):
      i = Item()
      i.period = "{0} - {1}".format(doc.start.strftime("%d.%m"), doc.finish.strftime("%d.%m"))
      i.status = doc.status
      i.skill = doc.skill
      i.strengths = doc.strengths
      i.razvitie = doc.razvitie
      i.task = doc.task
      
      self.data.append(i)

def item_cmp(x, y):
  res = cmp(x.manager, y.manager)
  
  if res == 0:
      res = cmp(x.data, y.data)

  return res

def loadData(params, server):
    data = ReportData()
    
    q = '"start" <= ToDate("{1}") and "finish" >= ToDate("{0}") and "agentid"=\'{2}\''.format(
        params.start.strftime("%d/%m/%Y 0:0:0"),
        params.finish.strftime("%d/%m/%Y 23:59:59"),
        params.userid)
        
    spktasks = server.Get("SPKTask", q)
    
    for t in spktasks:
      data.add(t)
            
    return data

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
class  XLBuilderEx(XLBuilder):
  HEADER_CELL_COLOR = "FFD8D8D8"
  
  def adjustHeadCell(self, sheet, cell, row, column):
    self.setBackColor(cell, self.HEADER_CELL_COLOR)
    return column
        
def printOut(report, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilderEx()
    
    sheet.cell(row=0, column=0).value = "Отчет по задачам"
    sheet.cell(row=1, column=0).value = "Интервал: c {0} по {1}".format(params.start.strftime('%d.%m.%Y'), params.finish.strftime('%d.%m.%Y'))  
    row = 3
    
    head = ['Интервал', 'Статус', 'Развиваемый навык', 'Сильные стороны', 'Стороны для развития', 'Задания для закрепления и развития']
    
    xlb.makeHead(sheet, row, head, True)
    
    row += 1

    for item in report.data:
      xlb.makeCells(sheet, row, item.values())
      row += 1     
    
    setCellWidth(sheet, [30, 30, 30, 30, 30, 30])

    return wb
    
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    
    logging.info("printOut")
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info('end')
    
