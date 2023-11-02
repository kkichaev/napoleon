# -*- coding: cp1251 -*-
import logging
from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color

import datetime
from datetime import timedelta


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

class DocData:
    __slots__ = ['manager', 'data', 'agent', 'orgname', 'orgaddress', 'photoCount', 'sku', 'start', 'finish', 'photoKey']
    
    def __init__(self):  
      self.start = None
      self.finish = None
      self.photoKey = ''
      pass
    
    def values(self, row):
      result = [self.manager, self.data.strftime('%d.%m.%Y'), self.agent, self.orgname, self.orgaddress, self.photoCount, self.sku,
        self.start.strftime("%H:%M"), self.finish.strftime("%H:%M")]
      td = self.finish - self.start
      ts = td.total_seconds() 
      h = ts / 3600
      ms = ts % 3600
      m = ms / 60
      s = ms % 60
      result.append('%02d:%02d:%02d' % (h,m,s))
      
      return result
      
class GroupData(DocData):
    count = 0
    
    def __init__(self):
      self.count = 0
      
    def values(self, row):
      result = [self.manager, self.data.strftime('%d.%m.%Y'), '', '', '', "=SUM(F{0}:F{1})".format(row + 1, row + self.count), "=SUM(G{0}:G{1})".format(row + 1, row + self.count),
        '', '', '=SUM(J{0}:J{1})'.format(row + 1, row + self.count)]
      
      return result
                    
class ReportData:
    __slots__ = ['data', 'orgs', 'managers', 'agents']
    
    def __init__(self):
      self.data = dict()
        
    def add(self, doc, approve):
      key = self.makeKey(doc.created.strftime("%d.%m.%Y %H:%M:%S"), doc.userid, doc.agentid)

      if not key in self.data:
        data = DocData()
        self.data[key] = data
      
      data = self.data[key]
      
      data.data = doc.aprDate
      data.manager = doc.userid
      data.agent = doc.agentid
            
      if data.manager in self.managers:
        mgr = self.managers[data.manager]
        if len(mgr.name) > 0:
            data.manager = mgr.name
      
      if data.agent in self.agents:
        data.agent = self.agents[data.agent].name
        
      data.orgname = 'Код контрагента <' + doc.id + ">"
      data.orgaddress = ""
      
      if doc.id in self.orgs:
        data.orgname = self.orgs[doc.id].name
        data.orgaddress = self.orgs[doc.id].address
       
      data.photoCount = 0
      data.photoKey = self.makeKey(doc.created.strftime('%d.%m.%Y'), doc.agentid, doc.id)
            
      data.sku = 0
      
      if approve != None :
          data.photoKey= self.makeKey(approve.created.strftime('%d.%m.%Y'), doc.agentid, doc.id)
          data.sku = approve.qty
      
      if data.start == None or data.start > doc.aprDate:
        data.start = doc.aprDate
      
      if data.finish == None  or data.finish < doc.committed:
        data.finish = doc.committed
      if data.finish < data.start:
          data.finish = data.start
    
    def makeKey(self, c, u, i):
      return "{0}\t{1}\t{2}".format(c, u, i)
  
    def setPhoto(self, minDate, maxDate, server):
      if minDate == None or maxDate == None: return

      phdic = dict()

      cdate = minDate
      while cdate <= maxDate:
         edate = cdate + timedelta(days=3)
         if edate > maxDate: edate = maxDate

         q = cdate.strftime("%d/%m/%Y 0:0:0") + ';' + edate.strftime("%d/%m/%Y 23:59:59")
         phd = server.Get('PhotoCount', q)

         for d in phd:
             key = self.makeKey(d.date.strftime('%d.%m.%Y'), d.userid, d.id)
             phdic[key] = d.count
         cdate = edate + timedelta(days=1) 

      for d in self.data.itervalues():
         if d.photoKey in phdic: d.photoCount = phdic[d.photoKey]

      logging.info("Done Photos " + str(len(phdic)))

        
    def getItems(self):  
      res1 = list(self.data.values())
      res1 = sorted(res1, cmp=item_cmp)
      
      groupData = None
      res = list()
      
      for i in res1:
        if groupData == None or groupData.manager != i.manager or groupData.data.strftime('%d.%m.%Y') != i.data.strftime('%d.%m.%Y'):
          groupData = GroupData()
          res.append(groupData)
          groupData.manager = i.manager
          groupData.data = i.data
        
        res.append(i)
        groupData.count += 1
      
      return res

def item_cmp(x, y):
  res = cmp(x.manager, y.manager)
  
  if res == 0:
      res = cmp(x.data, y.data)

  return res


def loadData(params, server):
#     orgs = server.Get('CommonOrgs', '', 'id')
    
    managers = server.Get('DivisionManager', '', 'login')
    agents = server.Get('Agents', '', 'id')

    q = params.start.strftime("%d/%m/%Y") + ";" + params.finish.strftime("%d/%m/%Y 23:59:59")

    orgs = server.Get("ApproveOrgs", q, "id")

    data = ReportData();
    data.managers = managers
    data.agents = agents
    data.orgs = orgs

    cdate = params.start
    minDate = None
    maxDate = None
    while cdate < params.finish:

       q = '"aprDate" >= ToDate("{0}") and "aprDate" <= ToDate("{1}")'.format(
           cdate.strftime("%d/%m/%Y 0:0:0"),
           cdate.strftime("%d/%m/%Y 23:59:59"))

       docs = server.Get('LayoutApproveLog', q)
#       logging.info("LogCount " + str(len(docs)))

       q = cdate.strftime("%d/%m/%Y 0:0:0") + ";" + cdate.strftime("%d/%m/%Y 23:59:59")
       a = server.Get('ApproveReportDocs', q) 
#       logging.info("DocsCount " + str(len(a)))

       adic = dict()
       for ad in a:
          key = ad.id + ad.created.strftime("%d.%m.%Y %H:%M:%S")
          adic[key] = ad
    
       for d in docs:
          if minDate == None or minDate > d.created.date() : minDate = d.created.date()
          if maxDate == None or maxDate < d.created.date() : maxDate = d.created.date()

          key = d.id + d.created.strftime("%d.%m.%Y %H:%M:%S")
          adoc = None
          if key in adic:
             adoc = adic[key]
             d.agentid = adoc.userid

          data.add(d, adoc)
 
       cdate = cdate + timedelta(days=1)

    logging.info("Done docs " + str(len(data.data)))
    data.setPhoto(minDate, maxDate, server)
    return data

class XLBuilderEx(XLBuilder):
  __slots__ = ['useBold']
  
  def __init__(self):
      self.useBold = True
    
  def adjustHeadCell(self, sheet, cell, row, column):
    # cell.style.alignment.text_rotation = 90
    return column
  
  def makeCells(self, sheet, row, values, startColumn, boldFont):
      self.useBold = boldFont
      XLBuilder.makeCells(self, sheet, row, values, startColumn)

  def makeHead(self, sheet, row, titles, wrap_text=False, startColumn=0):
    self.useBold = False
    XLBuilder.makeHead(self, sheet, row, titles, wrap_text, startColumn)
    # sheet.row_dimensions[1].height = 95
    
  def makeCell(self, sheet, row, column, cell, value, border=Border.BORDER_THIN):
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border)
    
    if column == 7 or column == 8:
      cell.style.number_format._set_format_code('hh:mm')
    elif column == 9:
      cell.style.number_format._set_format_code('hh:mm:ss')
    
    if self.useBold:  
        cell.style.font.bold = True
        self.setBackColor(cell, 'FFA500') 

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
 
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilderEx()
    
    sheet.cell(row=0, column=0).value = "ВЫГРУЗКА СТАТИСТИКИ ПО РАБОТЕ ОПЕРАТОРОВ ПО СВЕРКЕ"
    row = 3
    
    head = ['Оператор по сверке', 'Дата', 'Агент (Торговый представитель)', 'Контрагент', 'Адрес ТТ', 'Количество фотографий из ТТ', 'Общее количество SKU В ТТ',
             'Время начала редактирования выкладки в ТТ', 'время окнчания редактирования выкладки в ТТ', 'Итого время работы(час:мин:сек)']
    
    xlb.makeHead(sheet, row, head, True)
    
    row += 1

    for item in data.getItems():
      data = item.values(row + 1)
      xlb.makeCells(sheet, row, data, 0, isinstance(item, GroupData))
      row += 1     
    
    setCellWidth(sheet, [30, 30, 30, 30, 30, 30, 30, 30, 30, 30])

    return wb
    
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    
    logging.info("printOut")
    wb = printOut(data, params)
    logging.info('PrintOut end')

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info('end')
    
