# -*- coding: cp1251 -*-
from openpyxl.workbook import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter

import sys
reload(sys)
sys.setdefaultencoding("cp1251")

class AgentData :
    __slots__ = ['orgs', 'docs', 'agent', 'matrix']
    
    def __init__(self, server, docWhere, agent):
        chuid = "'" + agent.id + "'"

        server.ChangeUser(chuid);
        self.orgs = server.Get('Org', '', 'id')
        src = server.Get('OrgMatrix', '')
        server.RestoreUser()
        
        where = docWhere + ' and "userid" = ' + chuid
        self.docs = server.Get('OrgDistrib', where)
        self.agent = agent

        self.matrix = dict()
        for mi in src:
            if not mi.id in self.matrix:
                self.matrix[mi.id] = list()
            self.matrix[mi.id].append(mi.id_i)
            
    def getMatrixCount(self, id):
        return 0 if not id in self.matrix else len(self.matrix[id])

def loadData(server, param):
    
    agents = server.Get('Agents', '', 'id')
    
    docWhere = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")' . format(
        param.start.strftime('%d/%m/%Y'),
        param.finish.strftime('%d/%m/%Y 23:59:59'))
    
    data = list()
    for ai in param.agents:
        if ai.id in agents:
            ad = AgentData(server, docWhere, agents[ai.id])
            data.append(ad)
            
    return data
   
def docKey(doc):
    return doc.created
    
def printOut(data, server, param):
    
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilder()
    
    cr = 0
    cc = 0
    
    cell = sh.cell(row = cr, column = cc)
    cell.value = 'Отчет по матрице за период {0} - {1} ' . format(param.start.strftime('%d/%m/%Y'), param.finish.strftime('%d/%m/%Y'))    
    cr += 1
    
    titles = ['№ маршрута', 'Дата', 'сеть', 'Адрес', 'Кол-во товара по матрице', 'Кол-во товара по факту', '% исполнения']
    xlb.makeHead(sh, cr, titles)
    cr += 1
    
    for i in data:
        if i.docs == None: continue
        
        for di in sorted(i.docs, key=docKey):
            values = [i.agent.name]
            org = di.id
            adr = '' 
            if di.id in i.orgs :
                o = i.orgs[di.id]
                org = o.name
                adr = o.address
                
            mctr = i.getMatrixCount(di.id)
            haverem = 0
            for item in di.items:
                if len(item.remark) > 0: haverem += 1
            docctr = mctr - haverem
            if docctr < 0: docctr = 0
            prc = "" if mctr == 0 else '{:.2f}%'.format(float(docctr) / mctr * 100.0)
            values.extend([di.created.strftime('%d/%m/%Y'), org, adr, mctr, docctr, prc])
            
            xlb.makeCells(sh, cr, values) 
            cr += 1
    
    setCellWidth(sh, [30,20,20,50,20,20,20])
    return wb

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    
def run(server):
    param = server.Params[0]
    data = loadData(server, param)
    wb = printOut(data, server, param)
    
    XLBuilder().workbookToObject(wb, "matrix_rep.xlsx", server)