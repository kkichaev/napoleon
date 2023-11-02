# -*- coding: cp1251 -*-
from openpyxl.workbook import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter

import sys
reload(sys)
sys.setdefaultencoding("cp1251")

class DataItem:
  __slots__ = ['agent', 'org', 'address',  'item', 'date', 'problem', 'remark', 'created']
  
class AgentData :
    __slots__ = ['agent', 'items']
    
    def __init__(self, server, docWhere, agent, price):
        self.agent = agent.name
        self.items = list()
        
        chuid = "'" + agent.id + "'"

        server.ChangeUser(chuid);
        orgs = server.Get('Org', '', 'id')
        server.RestoreUser()
        
        where = docWhere + ' and "userid" = ' + chuid
        docs = server.Get('OrgDistrib', where)
        visit = server.Get('VisitLite', where)
        visit_hash = self.compileVisitHash(visit)
        keys = []

        
        for d in docs:
          date = d.created.strftime('%d.%m.%Y')
          key = d.id
          remark = visit_hash[date + key] if date + key in visit_hash else ""
          org = orgs[d.id].name if d.id in orgs else d.id
          address = orgs[d.id].address if d.id in orgs else ""
          
          for i in d.items:
            if len(i.remark.strip()) > 0:
              self.items.append(self.compileDataItem(
                agent.name, org, address, price[i.id].name if i.id in price else i.id,
                date, i.remark.strip(), remark, d.created))
            
              k = key + i.id
              
              if not k in keys:
                keys.append(k)
        
        for d in docs:
          where  = "{0};{1}".format(d.created.strftime('%d.%m.%Y %H:%M:%S'),d.userid)       
          pf = server.Get("DistribWithPhoto", where)
          date = d.created.strftime('%d.%m.%Y')
          key = d.id
          remark = visit_hash[date + key] if date + key in visit_hash else ""
          org = orgs[d.id].name if d.id in orgs else d.id
          address = orgs[d.id].address if d.id in orgs else ""
          
          for i in pf:
            k = key + i.itemid
            
            if k in keys:
              self.items.append(self.compileDataItem(
                agent.name, org, address, price[i.itemid].name if i.itemid in price else i.itemid,
                date, '', remark, d.created))
            
        self.items = sorted(self.items, cmp=self.cmpItem)
    
    def cmpItem(self, x, y):
      res = cmp(x.agent, y.agent)
      
      if res == 0:
        res = cmp(x.org, y.org)
        
      if res == 0:
        res = cmp(x.address, y.address)
        
      if res == 0:
        res = cmp(x.item, y.item)
        
      if res == 0:
        res = cmp(x.created, y.created)
        
      return res
      
    def compileDataItem(self, agent, org, address, item, date, problem, remark, created):
      res = DataItem()
      res.agent = agent
      res.org = org
      res.address = address
      res.item = item
      res.date = date
      res.problem = problem
      res.remark = remark
      res.created = created
      
      return res
      
          
    def compileVisitHash(self, docs):
        res = dict()
        
        for d in docs:
          if len(d.remark) > 0:
            key = d.created.strftime('%d.%m.%Y') + d.id
            
            if key in res:
              res[key] += ", " + d.remark  
            else:  
              res[key] = d.remark  
          
        return res

def loadData(server, param):
    
    agents = server.Get('Agents', '', 'id')
    
    docWhere = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")' . format(
        param.start.strftime('%d/%m/%Y'),
        param.finish.strftime('%d/%m/%Y 23:59:59'))
        
    price = server.Get("ManagerPrice", "", "id")
    
    data = list()
    
    for ai in param.agents:
        if ai.id in agents:
            ad = AgentData(server, docWhere, agents[ai.id], price)
            data.append(ad)
            
    return data
    
class XLBuilderEx(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value):
      XLBuilder.makeCell(self, sheet, row, column, cell, value)
      
      if column == 5 and len(value) == 0:
        self.setBackColor(cell,'ffC0C0C0')
        cell.value = "Есть фото"

def printOut(data, server, param):
    
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    
    cr = 0
    cc = 0
    
    cell = sh.cell(row = cr, column = cc)
    cell.value = 'Отчет по решению проблем за период {0} - {1} ' . format(param.start.strftime('%d/%m/%Y'), param.finish.strftime('%d/%m/%Y'))    
    cr += 1
    
    titles = ['Мерчендайзер №', 'Сеть', 'Адрес', 'Товар', 'Дата', 'Причина отсутствия/фото', 'Комментарий ']
    xlb.makeHead(sh, cr, titles)
    cr += 1
    
    for d in data:
      if d.items == None: 
        continue
        
      for i in d.items:
        values = [i.agent, i.org, i.address, i.item, i.date, i.problem, i.remark]
        xlb.makeCells(sh, cr, values) 
        cr += 1
    
    setCellWidth(sh, [20,20,30,50,20,20,20])
    return wb

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    
def run(server):
    param = server.Params[0]
    data = loadData(server, param)
    wb = printOut(data, server, param)
    
    XLBuilder().workbookToObject(wb, "matrix_rep.xlsx", server)