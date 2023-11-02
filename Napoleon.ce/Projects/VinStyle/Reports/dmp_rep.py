# -*- coding: cp1251 -*-
from openpyxl.workbook import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import sys
import tempfile

reload(sys)
sys.setdefaultencoding("cp1251")

class DataItem:
  __slots__ = ['agent', 'org', 'address', 'item', 'date', 'dmp', 'created', 'photo', 'pic']
  
  def getData(self):
    res = []
    res.append(self.agent)
    res.append(self.org)
    res.append(self.address)
    res.append(self.item)
    res.append(self.date)
    res.append(self.dmp)
    
    for p in self.photo:
      res.append(p)
    
    return res
  
class AgentData :
    __slots__ = ['agent', 'items']
    
    def __init__(self, server, docWhere, agent, price, type):
        self.agent = agent.name
        chuid = "'" + agent.id + "'"

        server.ChangeUser(chuid);
        orgs = server.Get('Org', '', 'id')
        server.RestoreUser()
        
        where = docWhere + ' and "userid" = ' + chuid + ' and "dmpId" != ""'
        visit = server.Get('Visit', where)
        
        data = dict()
        
        for v in visit:
          date = v.created.strftime('%d.%m.%Y')
          org = orgs[v.id].name if v.id in orgs else v.id
          address = orgs[v.id].address if v.id in orgs else ""
          
          vc = 1
          
          for i in v.items:
            key = i.itemId + i.dmpId
            
            if not key in data:
              data[key] = self.compileDataItem(
                agent.name, org, address, 
                price[i.itemId].name if i.itemId in price else i.itemId,
                type[i.dmpId].text if i.dmpId in type else i.dmpId,
                date, v.created)
              data[key].photo = []
              data[key].pic = []
        
            s = '=HYPERLINK("{0}.jpg","Фото{1}")'.format(i.key, vc)
            vc += 1
            data[key].photo.append(s);
            data[key].pic.append(i)
            
        self.items = list()
        
        for d in data.values():
          self.items.append(d)
        
        self.items = sorted(self.items, cmp=self.cmpItem)
    
    def cmpItem(self, x, y):
      res = cmp(x.agent, y.agent)
      
      if res == 0:
        res = cmp(x.org, y.org)
        
      if res == 0:
        res = cmp(x.address, y.address)
        
      if res == 0:
        res = cmp(x.item, y.item)
        
      if res == 0:
        res = cmp(x.created, y.created)
        
      return res
      
    def compileDataItem(self, agent, org, address, item, dmp, date, created):
      res = DataItem()
      res.agent = agent
      res.org = org
      res.address = address
      res.item = item
      res.dmp = dmp
      res.date = date
      res.created = created
      
      return res
      
def loadData(server, param):
    agents = server.Get('Agents', '', 'id')
    
    docWhere = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")' . format(
        param.start.strftime('%d/%m/%Y'),
        param.finish.strftime('%d/%m/%Y 23:59:59'))
        
    price = server.Get("ManagerPrice", "", "id")
    dts = server.Get("DMPType", "", "id")
    
    data = list()
    
    for ai in param.agents:
        if ai.id in agents:
            ad = AgentData(server, docWhere, agents[ai.id], price, dts)
            data.append(ad)
            
    return data
    
class XLBuilderEx(XLBuilder):
    def makeCell(self, sheet, row, column, cell, value):
      XLBuilder.makeCell(self, sheet, row, column, cell, value)
        
      if column == 5 and len(value) == 0:
        self.setBackColor(cell,'ffC0C0C0')
        cell.value = "Есть фото"

def printOut(data, server, param):
    
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    
    cr = 0
    cc = 0
    
    cell = sh.cell(row = cr, column = cc)
    cell.value = 'Отчет по ДМП за период {0} - {1} ' . format(param.start.strftime('%d/%m/%Y'), param.finish.strftime('%d/%m/%Y'))    
    cr += 1
    
    titles = ['Мерчендайзер №', 'Сеть', 'Адрес', 'Товар', 'Дата', 'Тип ДМП', 'Ссылка на фото']
    xlb.makeHead(sh, cr, titles)
    cr += 1
    
    sz = 0
    
    for d in data:
      if d.items == None: 
        continue
        
      for i in d.items:
        values = i.getData()
        z = len(values)
        
        if sz < z:
          sz = z
        
        xlb.makeCells(sh, cr, values) 
        cr += 1
    
    if sz > 0:
      rangeBorders(sh.range("A3:{0}{1}".format(get_column_letter(sz),str(cr))), Border.BORDER_THIN)
      rangeBorders(sh.range("A2:{0}2".format(get_column_letter(sz),str(cr))), Border.BORDER_MEDIUM)
      sh.merge_cells(start_row=1, start_column=6, end_row=1, end_column=sz-1)
      
    setCellWidth(sh, [20,20,30,50,20,20,10])
    return wb
    
def rangeBorders(range, border):
    for row in range:
        for cell in row:
            cell.style.borders.left.border_style = border
            cell.style.borders.right.border_style = border
            cell.style.borders.top.border_style = border
            cell.style.borders.bottom.border_style = border
                        
def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1

    
def run(server):
    param = server.Params[0]
    data = loadData(server, param)
    wb = printOut(data, server, param)
    
#    XLBuilder().workbookToObject(wb, "matrix_rep.xlsx", server)
    
    server.RegisterType("Result[name:s,file:b,items[name:s,file:b]]")
    
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = "dmp_rep.xlsx"
    obj.file = bytesOut

    for d in data:
      for i in d.items:
        for p in i.pic:
          print p
          n = obj.items.New()
          n.name = p.key
          print "p.id", p.id
          n.file = p.id

    server.Put(outObj)
