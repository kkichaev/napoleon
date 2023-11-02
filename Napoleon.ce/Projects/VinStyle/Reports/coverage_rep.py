# -*- coding: cp1251 -*-

from openpyxl.workbook import Workbook
from openpyxl.style import Color
from grsoft.xl_base import XLBuilder
from grsoft.route import AgentRoute
from datetime import timedelta
from openpyxl.cell import get_column_letter

import sys
reload(sys)
sys.setdefaultencoding("cp1251")

class AgentData :
    __slots__ = ['orgs', 'route', 'visited', 'agent']
    
    def __init__(self, server, docWhere, agent, start, finish):

        route = AgentRoute(server, agent.id)
        self.orgs = route.orgs;
        self.route = list()
        self.visited = list()
        self.agent = agent
        
        for o in self.orgs.itervalues():
            self.route.append(o.id)

#         cd = start
#         while cd <= finish:
#             cr = route.getDayRoute(cd.date())
#             for oi in cr:
#                 if not oi.id in self.route:
#                     self.route.append(oi.id)
#             
#             cd = cd + timedelta(days=1)
            
        chuid = "'" + agent.id + "'"
        checkDocs = ['Order', 'OrgDistrib', 'VisitInfo']
        
        for curDoc in checkDocs:
            where = docWhere + ' and "userid" = ' + chuid
            docs = server.Get(curDoc, where)
            if docs != None:
                for di in docs:
                    if (di.id in self.route) and (not di.id in self.visited):
                        self.visited.append(di.id)


def loadData(server, param):
    
    agents = server.Get('Agents', '', 'id')
    
    docWhere = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")' . format(
        param.start.strftime('%d/%m/%Y'),
        param.finish.strftime('%d/%m/%Y 23:59:59'))
    
    data = list()
    for ai in param.agents:
        if ai.id in agents:
            ad = AgentData(server, docWhere, agents[ai.id], param.start, param.finish)
            data.append(ad)
            
    return data

class XLBuiderEx (XLBuilder):
    __slots__ = ['backColor']
    
    def __init__(self):
        self.backColor = Color.WHITE
    
    def setBCColor(self, color):
        self.backColor = color
        
    def makeBorder(self, cell, border):
        XLBuilder.makeBorder(self, cell, border)
        self.setBackColor(cell, self.backColor)

def printOut(data, server, param):
    
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilder()
    
    cr = 0
    cc = 0
    
    cell = sh.cell(row = cr, column = cc)
    cell.value = 'Отчет по покрытию территории с {0} по {1} ' . format(param.start.strftime('%d/%m/%Y'), param.finish.strftime('%d/%m/%Y'))    
    cr += 1
    
    titles = ['Мерчендайзер №', 'кол-во точек на территории в маршруте', 'колво посещенных точек', '% покрытия территории']
    xlb.makeHead(sh, cr, titles, True)
    cr += 1
    
    xlbA = XLBuiderEx() 
    for i in data:
        rlen = len(i.route)
        vlen = len(i.visited)
        cvrPrc = "" if rlen == 0 else '{:.2f}%'. format(float(vlen) / rlen * 100.0)
        values = [i.agent.name, rlen, vlen, cvrPrc]

        xlb.makeCells(sh, cr, values) 
        cr += 1
        
        if len(i.route) == 0: continue 
        
        shName = i.agent.name
        if len(shName) > 30: 
           shName=shName[:30]
        shA = wb.create_sheet(None, shName)
        
        ra = 0
        ca = 0
        cell = shA.cell(row = ra, column=ca)
        cell.value = "Покрытие " + cvrPrc
        ra += 1
        
        for ri in i.route:
            org = ri
            adr = ''
            if ri in i.orgs:
                oi = i.orgs[ri]
                org = oi.name
                adr = oi.address
                 
            color = Color.GREEN if ri in i.visited else Color.RED
            xlbA.setBCColor(color)
            
            values = [org, adr]
            xlbA.makeCells(shA, ra, values)
            ra += 1
       
        setCellWidth(shA, [30, 50])    
     
    setCellWidth(sh, [30, 30, 30, 30])  
    return wb

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1


def run(server):
    param = server.Params[0]
    data = loadData(server, param)
    wb = printOut(data, server, param)
    
    XLBuilder().workbookToObject(wb, "covarage_rep.xlsx", server)