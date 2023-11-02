# -*- coding: cp1251 -*-

from openpyxl.workbook import Workbook
from openpyxl.style import Color, NumberFormat
from grsoft.xl_base import XLBuilder
from grsoft.route import AgentRoute
from datetime import timedelta
from openpyxl.cell import get_column_letter

import sys
reload(sys)
sys.setdefaultencoding("cp1251")

class Report:
  __slots__ = ['orgs', 'agents', 'start', 'finish', 'mdata', 'avgdata', 'docs','check']
  KEY_DOC_FMT = "{0}\t{1}\t{2}"
  KEY_CHECK_FMT = "{0}\t{1}\t{2}\{3}"

  def compile(self,server, agents, start, finish):
    self.start = start
    self.finish = finish
    self.agents = agents
    self.mdata = list()
    self.orgs = dict()
    self.avgdata = list()
    self.loadOrgs(server)
    self.loadDocs(server, start, finish)
    self.loadCheck(server, start, finish)
    self.loadPhotoCount(server, start, finish)  
    
  def loadPhotoCount(self, server, start, finish):
    pc = server.Get('PhotoCount', "'{0}';'{1}'".format(start.strftime('%d/%m/%Y'), finish.strftime('%d/%m/%Y 23:59:59')))
    self.photocount = dict()
    
    for p in pc:
      key = self.KEY_DOC_FMT.format(p.date.strftime('%d.%m.%Y %H:%M:%S'), p.id, p.userid)
      self.photocount[key] = p
    
  def loadCheck(self, server, start, finish):
    check = server.Get('DistrCheck','"doccreated" >= ToDate("{0}") and "doccreated" <= ToDate("{1}")' . format(start.strftime('%d/%m/%Y'), finish.strftime('%d/%m/%Y 23:59:59')))  
    self.check= dict()
    
    for c in check:
      key = self.KEY_CHECK_FMT.format(c.doccreated.strftime('%d.%m.%Y %H:%M:%S'), c.id, c.agentid, c.userid)
      
      if not key in self.check:
        self.check[key] = list()
        
      self.check[key].append(c)  
    
  def loadDocs(self, server, start, finish):
    docs = server.Get('OrgDistrib','"created" >= ToDate("{0}") and "created" <= ToDate("{1}")' . format(start.strftime('%d/%m/%Y'), finish.strftime('%d/%m/%Y 23:59:59')))
    self.docs = dict()
    
    for d in docs:
      if not d.userid in self.docs:
        self.docs[d.userid] = list()
        
      self.docs[d.userid].append(d)  
      
  def loadOrgs(self,server):
    porg = server.Get("PotenzialOrg", "", "id")
    self.orgs.update(porg)
    
    for id in self.agents:
      server.ChangeUser(id)
      aorgs = server.Get("Org", '', 'id')
      server.RestoreUser()
      
      if aorgs != None:
          self.orgs.update(aorgs)

class AVGItem:
  __slots__ = ['name', 'checked']
  
  def value(self):
    return [self.name, self.checked]
  
class Item:
  __slots__ = ['data', 'agent', 'org', 'photocount', 'checked']
  
  def value(self):
    return [self.data.strftime('%d.%m.%Y'), self.agent, self.org, self.photocount, self.percent()]
  
  def percent(self):
    return self.checked / self.photocount if self.photocount != 0 else 0
    
def item_cmp(x, y):
  res = cmp(x.data, y.data)

  if res == 0:
      res = cmp(x.org, y.org)

  return res
  
class ManagerData:
  __slots__ = ['id', 'name', 'items']
  
  def __init__(self, m):
    self.id = m.login
    self.name = m.name
  
  def compile(self, server, report, agents):
    self.items = list()
    
    for a in agents:
      if a.id in report.docs:
        for d in report.docs[a.id]:
          i = Item()
          i.data = d.created
          i.agent = report.agents[d.userid].name if d.userid in report.agents else d.userid
          i.org = report.orgs[d.id].name if d.id in report.orgs else d.id
          key = Report.KEY_DOC_FMT.format(d.created.strftime('%d.%m.%Y %H:%M:%S'), d.id, d.userid)
          i.photocount = report.photocount[key].count if key in report.photocount else 0
          key = Report.KEY_CHECK_FMT.format(d.created.strftime('%d.%m.%Y %H:%M:%S'), d.id, d.userid, self.id)
          i.checked = len(report.check[key]) if key in report.check else 0
          self.items.append(i)
    
    self.items = sorted(self.items, cmp=item_cmp)
    
  def avgChecked(self):
    s = 0.0
    
    for i in self.items:
      s += i.percent()
      
    sz = len(self.items)
    res =  s / sz if sz != 0 else 0
    return res

def loadData(server, param):
    managers = server.Get('DivisionManager','')
    divisions = server.Get('Division','','id')
    agents = server.Get('Agents', '', 'id')
    
    docWhere = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")' . format(
        param.start.strftime('%d/%m/%Y'),
        param.finish.strftime('%d/%m/%Y 23:59:59'))
    
    data = Report()
    data.compile(server, agents, param.start, param.finish)
    
    for m in managers:
      if m.division in divisions:
        md = ManagerData(m)
        md.compile(server, data, divisions[m.division].agents)
        data.mdata.append(md)
        
        a = AVGItem()
        a.name = m.name
        a.checked = md.avgChecked()
        
        data.avgdata.append(a)
            
    return data

class XLBuiderEx (XLBuilder):
  def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        
        if column == 4 :
            cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE
            
  def paintRow(self, sheet, row, clmncount, val):
    RED = "FFFF0000"
    GREEN = "FF00FF00"
    
    for c in range(0, clmncount):
      c = sheet.cell(row=row, column=c)
      self.setBackColor(c, GREEN if val == 1 else RED)          
            
class XLBuiderAvgEx (XLBuilder):
  def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        
        if column == 1 :
            cell.style.number_format.format_code = NumberFormat.FORMAT_PERCENTAGE
      
def printOut(data, server, param):
    
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuiderAvgEx()
    
    r = 1
    cell = sh.cell(row = r, column = 0)
    cell.value = 'Отчет по просмотру матриц с {0} по {1} ' . format(param.start.strftime('%d/%m/%Y'), param.finish.strftime('%d/%m/%Y'))    
    r += 1
    
    titles = ['Руководитель', '% просмотра']
    xlb.makeHead(sh, r, titles, True)
    r += 1
    
    for i in data.avgdata:
      xlb.makeCells(sh, r, i.value())
      r+= 1
      
    setCellWidth(sh, [30, 30])    
    
    xlb = XLBuiderEx() 
    
    for m in data.mdata:
      name = m.name[:30]
      sh = wb.create_sheet(None, name)
      r = 1
      titles = ['Дата', 'Агент', 'Клиент', 'Количество фото', '% просмотра']
      xlb.makeHead(sh, r, titles, True)
      r += 1
      
      for i in m.items:
        xlb.makeCells(sh, r, i.value())
        xlb.paintRow(sh, r, len(i.value()), i.percent())
        r+= 1
        
      setCellWidth(sh, [10, 30, 30, 18, 18])    

    return wb

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1


def run(server):
    param = server.Params[0]
    data = loadData(server, param)
    wb = printOut(data, server, param)
    
    XLBuilder().workbookToObject(wb, "covarage_rep.xlsx", server)