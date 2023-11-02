# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

reload(sys);

_month_names = ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

class PTypeItem:
  def __init__(self):
    self.items = {}
    
  def values(self):
    res =  [self.name, self.plan_title, self.month1_title, self.month2_title, self.month3_title, "Факт"]
    return res
    
  def getItems(self):
    return self.items.values()  
  
class AgentItem:
  def __init__(self):
    self.plan = 0
    self.month1 = {}
    self.month2 = {}
    self.month3 = {}
    
  def values(self):
    return [self.name, self.plan, self.monthvals('month1'), self.monthvals('month2'), self.monthvals('month3'), self.fact()]

  def monthvals(self, mf):
    res = 0
    dict = getattr(self, mf)
    cn = len(dict)
    
    for oid in dict:
      org_dict = dict[oid]
      
      for pid in org_dict:
        price_avg = 0
        
        vals = org_dict[pid]
        
        if len(vals) > 0:
          for v in vals:
            price_avg += v
            
          res += price_avg / len(vals)  
          
    return res / cn if cn > 0 else 0
    
  def fact(self):
    return (self.monthvals('month1') + self.monthvals('month2') + self.monthvals('month3')) / 3
    
class ReportData:
  def __init__(self):
    self.items = {}
    self.agents = {}

class XLB(XLBuilder):
  pass

def createPTypeItem(title, agents, month):
  res = PTypeItem()
  res.name = title
  res.plan_title = "План {0} кв".format((int)((month + 2) / 3))
  res.month1_title = _month_names[month - 1]
  res.month2_title = _month_names[month - 1 + 1]
  res.month3_title = _month_names[month - 1 + 2]
  
  for aid in agents:
    ai = AgentItem()
    ai.id = aid
    res.items[ai.id] = ai
    ai.name = agents[aid]
    
  return res  
    
def loadData(data, params, server):
  for u in params.userids:
    server.ChangeUser("'" + u.id + "'")
    name = server.CurrentUser().name
    server.RestoreUser()
    data.agents[u.id] = name

  planItems = ["SKU Лето", "KEY SKU Лето", "17-18 дюйм Лето", "SKU Зима", "KEY SKU Зима", "17-18 дюйм Зима", "ЛГШ"]
  
  for p in range(0, len(planItems)):
    data.items[p] = createPTypeItem(planItems[p], data.agents, params.start.month)

  plans = server.Get('CPlan', '"date"=ToDate("{0}")'.format(params.start.strftime("%d/%m/%Y")))
  
  planField = ["summer", "keySummer", "d17_18Summer", "winter", "keyWinter", "d17_18Winter", "lgsh"]
  
  for p in plans:
    for idx in range(0, len(planField)):
      if p.userid in data.items[idx].items:
        ai = data.items[idx].items[p.userid]
        ai.plan = getattr(p, planField[idx])

  price = server.Get("Price", "", "id")
  
  for u in params.userids:
    aid = u.id
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" < ToDate("{2}")'.format("'"+aid+"'", 
        params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y"))
    remnants = server.Get('OrgRemnants', where)
    quarted = (int)((params.start.month + 2) / 3)
    
    for d in remnants:
      for di in d.items:
        if di.id in price:
          p = price[di.id]
          
          if p.docFilter == 1 and p.season == 2: # Лето
            putRemnant(data, 0, d, di, quarted)
            
          if p.docFilter == 1 and p.season == 2 and p.keySKU == 1: # KEY Лето
            putRemnant(data, 1, d, di, quarted)  
            
          if p.docFilter == 1 and p.season == 2 and (p.diameter == 17 or p.diameter == 18): # 17-18 дюйм Лето 
            putRemnant(data, 2, d, di, quarted)   

          if p.docFilter == 1 and p.season == 1: # Зима
            putRemnant(data, 3, d, di, quarted)
            
          if p.docFilter == 1 and p.season == 1 and p.keySKU == 1: # KEY Зима
            putRemnant(data, 4, d, di, quarted)  
            
          if p.docFilter == 1 and p.season == 1 and (p.diameter == 17 or p.diameter == 18): # 17-18 дюйм Зима
            putRemnant(data, 5, d, di, quarted)
              
          if p.docFilter == 1 and p.autoType == 'ЛГШ': # ЛГШ
            putRemnant(data, 6, d, di, quarted)    
  return data
  
def putRemnant(data, idx, d, di, quarted):
  monthFields = ["month1", "month2", "month2"]
  
  ai = data.items[idx].items[d.userid]
  m = (d.created.month - 1) % quarted
  
  mf = getattr(ai, monthFields[m])
  
  if not d.id in mf:
    mf[d.id] = {}
     
  om = mf[d.id]
  
  if not di.id in om:
    om[di.id] = []
    
  om[di.id].append(di.qty)

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "SKU"
    
    r = 0
    xlb = XLB()
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    sheet.cell(row=0, column=1).value = "{0} Й квартал".format((int)((params.start.month + 2) / 3))
    
    r = 1
    arr = ["",""]
    xlb.makeHead(sheet, r, arr);
    
    for i in data.items:
      item = data.items[i]
      xlb.makeHead(sheet, r, item.values());
      r += 1
      
      for i2 in item.items:
        item2 = item.items[i2]
        xlb.makeCells(sheet, r, item2.values())
        
        r += 1
      
    pos = 3  
    for i in data.items:
      item = data.items[i]
      
      head = [item.name, ""]
      s1 = pos
      s2 = pos+len(item.items)-1
      arr = ["=AVERAGE(C{0}:C{1})".format(s1, s2),
             "=AVERAGE(D{0}:D{1})".format(s1, s2),
             "=AVERAGE(E{0}:E{1})".format(s1, s2),
             "=AVERAGE(F{0}:F{1})".format(s1, s2)]
      
      head.extend(arr)

      xlb.makeHead(sheet, r, head);
      r += 1
      pos += 1 + len(item.items)
      
    cc = 1
    for w in [30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "sku_report.xlsx", server)                
    logging.info('end')
    
