# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

reload(sys);

class Item:
  def values(self):
    pc = self.faceCoordiant / self.faceAll if self.faceAll != 0 else 0
    return [self.id, self.name, self.nameFakt, self.orgReg, self.city, self.address, self.phone, self.web, self.typePTT,
        self.faceAll, self.faceCoordiant, str(round(pc * 100, 2)) + '%',
        self.contactName, self.staffPosition, self.contactPhone]
    # return [self.id, self.name, self.nameFakt, self.orgReg, self.city, self.address, self.phone, self.web, self.typePTT,
    #   self.specPTT, self.avgSell, self.cordiantPart, self.faceAll, self.faceCoordiant, self.contactName, self.staffPosition, self.contactPhone]
    
class ReportData:
  def __init__(self):
    self.items = []
    pass

class XLB(XLBuilder):
  HEAD_COLOR = "FFD8D8D8"

  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.wrap_text = True    
    self.setBackColor(cell, self.HEAD_COLOR)
    
    return column



def loadData(data, params, server):
  if len(params.userids) > 0:
    aid = params.userids[0].id
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    orgReg = server.Get("OrgRegion", "", "id")
    city = server.Get("City", "", "id")
    typePTT = server.Get("TypePTT", "", "id")
    specPTT = server.Get("SpecPTT", "", "id")
    staffPosition = server.Get("StaffPosition", "", "id")
    stock = server.Get("LastStock", "", "id")
    server.RestoreUser()

    for ido in orgs:
      o = orgs[ido]
      item = Item()
      item.id = ido 
      item.name = o.name
      item.nameFakt = o.nameFakt
      item.orgReg = orgReg[o.regionID].name if o.regionID in orgReg else "Регион с кодом <{0}>".format(o.regionID)
      item.city = '' # city[o.cityID].name if o.cityID in city else "Город с кодом <{0}>".format(o.cityID)
      item.address = o.address
      item.phone = o.phone
      item.web = o.web
      item.typePTT = '' # typePTT[o.typepttID].name if o.typepttID in typePTT else "Тип РТТ с кодом <{0}>".format(o.typepttID)
      item.specPTT = specPTT[o.specpttID].name if o.specpttID in specPTT else "Принадлежность с кодом <{0}>".format(o.specpttID)
      item.avgSell = o.avgSell
      item.cordiantPart = o.cordiantPart
      if stock and ido in stock:
        si = stock[ido]
        item.faceAll = si.qty
        item.faceCoordiant = si.aikos
      else:
        item.faceAll = 0
        item.faceCoordiant = 0
      item.contactName = ''
      item.staffPosition = ''
      item.contactPhone = ''
      
      if len(o.contacts) > 0:
        c = o.contacts[0]
        item.contactName = c.name
        item.staffPosition = staffPosition[c.staffPositionID].name if c.staffPositionID in staffPosition else ""
        item.contactPhone = c.phone
    
      data.items.append(item)
  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "КБ new"
    
    xlb = XLB()
    
    r = 0
    arr = ["Код РТТ","Юр. название РТТ", "Факт. название РТТ", "Область / край", "Город / поселок", "Улица, дом", "Телефон РТТ", 
           "Сайт", "Тип РТТ", 
           "Общее кол-во фейсов на основном месте продаж, шт", "Фейсинг Эйкос, шт", "% Эйкос",
           "Контактное лицо ФИО", "Должность", "Телефон"]
    xlb.makeHead(sheet, r, arr);
    
    r += 1
    
    for i in data.items:
      xlb.makeCells(sheet, r, i.values())
      r += 1
    
    cc = 1
    # for w in [12,19,19,19,19,19,14,14,10,10,10,10,10,10,19,19,14]:
    for w in [12,19,19,19,19,19,14,14,10,12,10,19,19,14,15]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "kbnew_report.xlsx", server)                
    logging.info('end')
    
