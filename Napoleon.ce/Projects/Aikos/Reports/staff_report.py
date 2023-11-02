# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging
import calendar
import locale
import datetime

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat
from calendar import monthrange
from datetime import timedelta

reload(sys);

class MonCell:
  pass
  
class ReportData:
  def __init__(self):
    self.items = []
    self.month_names = ["янв.", "фев.", "мар.", "апр.", "май", "июн.", "июл.", "авг.", "сен.", "окт.", "ноя.", "дек." ]
    self.months = {}
    
  def compileMonCells(self, s, f):
    idx = 0
    while s < f:
      mr = monthrange(s.year, s.month)
      mc = MonCell()
      mc.idx = idx
      mc.title = "{0} {1}".format(self.month_names[s.month-1], s.strftime("%y"))
      self.months[s] = mc
      s = s + timedelta(days=mr[1])
      idx += 1
      
  def monthTitles(self):
    res = [None] * len(self.months)
    
    for mc in self.months.values():
      res[mc.idx] = mc.title
      
    return res  

class Item:
  def __init__(self, s, f):
    self.cells = {}
    
    while s < f:
      self.cells[s] = 0
      mr = monthrange(s.year, s.month)
      s = s + timedelta(days=mr[1])
  
  def values(self):
    res = [self.username]
    res.extend(self.cellVales())
    return res
    
  def cellVales(self):
    res = []
    keys = sorted(self.cells.keys())
    
    for k in keys:
      res.append(self.cells[k])
    
    return res
    
class XLB(XLBuilder):
  FIXED_CELL_COLOR = Color("FFB6DDE8")

  def adjustHeadCell(self, sheet, cell, row, column):
    XLBuilder.adjustHeadCell(self, sheet, cell, row, column)
    fill = cell.style.fill;
    fill.fill_type = Fill.FILL_SOLID
    fill.start_color = self.FIXED_CELL_COLOR
        
    return column


def loadData(data, params, server):
  data.compileMonCells(params.start, params.finish)
  
  for u in params.userids:
    item = Item(params.start, params.finish)
    data.items.append(item)
    
    aid = u.id
    server.ChangeUser("'" + aid + "'")
    item.userid = aid
    item.username = server.CurrentUser().name
    server.RestoreUser()
    
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))

    docNames = ["VisitInfo", "OrgRemnants", "CMonitoring"]
    
    for name in docNames:
      docList = server.Get(name, where)
    
      if docList == None:
        continue
      
      for d in docList:
        c = d.created
        s = datetime.datetime(c.year, c.month, 1)
        item.cells[s] = 1

  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Персонал"
    
    xlb = XLB()
    
    r = 0
    arr = ["Территория"]
    arr.extend(data.monthTitles())
      
    xlb.makeHead(sheet, r, arr);
    
    r += 1
    for i in data.items:
      xlb.makeCells(sheet, r, i.values())
      r += 1
    
    cc = 1
    for w in [30]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    arr = [""]
    
    for w in data.months:
      let = get_column_letter(cc)
      sheet.column_dimensions[let].width = 10
      arr.append("=SUM({0}2:{0}{1})".format(let, r))

      cc += 1  
    
    xlb.makeCells(sheet, r, arr);
    
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "staff_report.xlsx", server)                
    logging.info('end')
    
