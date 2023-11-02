# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border, Color, Fill, Alignment, NumberFormat

reload(sys);

class PTypeItem:
  def __init__(self):
    self.items = {}
    
  def values(self):
    res =  [self.name]
    return res
    
  def getItems(self):
    return self.items.values()  
  
class AgentItem:
  def __init__(self):
    self.kb = 0
    self.akb = 0
    self.d17_18 = 0
    self.akb_summer = 0
    self.d17_18_summer = 0
    self.akb_winter = 0
    self.d17_18_winter = 0
    self.lgsh = 0
    
  def values(self):
    return [self.name, self.kb, self.akb, self.d17_18, self.akb_summer, self.d17_18_summer, self.akb_winter, self.d17_18_winter, self.lgsh]
  
class ReportData:
  def __init__(self):
    self.items = {}
    self.agents = {}
    
  def getItems(self):
    return self.items.values()

class XLB(XLBuilder):
  FIXED_CELL_COLOR = Color("FFB6DDE8")
    
  def adjustHeadCell(self, sheet, cell, row, column):
    XLBuilder.adjustHeadCell(self, sheet, cell, row, column)
    fill = cell.style.fill;
    fill.fill_type = Fill.FILL_SOLID
    fill.start_color = self.FIXED_CELL_COLOR

    return column  

def loadData(data, params, server):
  ptypes = server.Get('TypePTT')
  
  for pt in ptypes:
    pi = PTypeItem()
    pi.id = pt.id
    pi.name = pt.name
    
    data.items[pi.id] = pi
    
    for u in params.userids:
      ai = AgentItem()
      ai.id = u.id
      
      pi.items[ai.id] = ai
      
      if ai.id in data.agents:
        ai.name = data.agents[ai.id]
      else:
        server.ChangeUser("'" + ai.id + "'")
        ai.name = server.CurrentUser().name
        server.RestoreUser()
        data.agents[ai.id] = ai.name
        
      pi.items[ai.id] = ai  
      
      
  for u in params.userids:
    aid = u.id
    where = '"userid"={0} and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format("'"+aid+"'", 
      params.start.strftime("%d/%m/%Y 00:00:00"), params.finish.strftime("%d/%m/%Y 23:59:59"))
    remnants = server.Get('OrgRemnants', where)
    server.ChangeUser("'" + aid + "'")
    orgs = server.Get("Org", "", "id")
    price = server.Get("Price", "", "id")
    server.RestoreUser()
    
    akb = {}
    d17_18 = {}
    kb = len(orgs)
    akb_summer = {}
    d17_18_summer = {}
    akb_winter = {}
    d17_18_winter = {}
    lgsh = {}
    
    for r in remnants:
    
      if r.id in orgs:
        org = orgs[r.id]
        
        if not org.typepttID in akb:
          akb[org.typepttID] = []
          
        if not org.typepttID in d17_18:
          d17_18[org.typepttID] = []
          
        if not org.typepttID in akb_summer:
          akb_summer[org.typepttID] = []  
          
        if not org.typepttID in d17_18_summer:
          d17_18_summer[org.typepttID] = []   
          
        if not org.typepttID in d17_18_winter:
          d17_18_winter[org.typepttID] = []   

        if not org.typepttID in akb_winter:
          akb_winter[org.typepttID] = []  
        
        if not org.typepttID in lgsh:
          lgsh[org.typepttID] = []  
          
        for ri in r.items:
          if ri.id in price:
            p = price[ri.id]
            
            if ri.qty > 0 and p.docFilter == 1 and not r.id in akb[org.typepttID]:
              akb[org.typepttID].append(r.id)
              
            if ri.qty > 0 and p.docFilter == 1 and p.diameter in [17,18] and not r.id in d17_18[org.typepttID]:
              d17_18[org.typepttID].append(r.id)
              
            if ri.qty > 0 and p.docFilter == 1 and p.season == 2 and not r.id in akb_summer[org.typepttID]:
              akb_summer[org.typepttID].append(r.id)  
              
            if ri.qty > 0 and p.docFilter == 1 and p.season == 2 and p.diameter in [17,18] and not r.id in d17_18_summer[org.typepttID]:
              d17_18_summer[org.typepttID].append(r.id)  
  
            if ri.qty > 0 and p.docFilter == 1 and p.season == 1 and not r.id in akb_winter[org.typepttID]:
              akb_winter[org.typepttID].append(r.id)  
              
            if ri.qty > 0 and p.docFilter == 1 and p.season == 1 and p.diameter in [17,18] and not r.id in d17_18_winter[org.typepttID]:
              d17_18_winter[org.typepttID].append(r.id)
              
            if ri.qty > 0 and p.docFilter == 1 and p.autoType == 'ЛГШ' and not r.id in lgsh[org.typepttID]:
              lgsh[org.typepttID].append(r.id)  
  
    for pt in data.items:
      pti = data.items[pt]
      
      if aid in pti.items:
        ai = pti.items[aid]
        ai.kb = kb
        ai.akb = 0
        
        if pt in akb:
          ai.akb = len(akb[pt])
        
        ai.d17_18 = 0
        
        if pt in d17_18:
          ai.d17_18 = len(d17_18[pt])
          
        ai.akb_summer = 0
        
        if pt in akb_summer:
          ai.akb_summer = len(akb_summer[pt])  
          
        ai.d17_18_summer = 0
        
        if pt in d17_18_summer:
          ai.d17_18_summer = len(d17_18_summer[pt])   

        ai.akb_winter = 0
        
        if pt in akb_winter:
          ai.akb_winter = len(akb_winter[pt])
          
        ai.d17_18_winter = 0

        if pt in d17_18_winter:
          ai.d17_18_winter = len(d17_18_winter[pt])
          
        ai.lgsh = 0

        if pt in lgsh:
          ai.lgsh = len(lgsh[pt])  
  
  return data

def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "КБ offline"
    
    xlb = XLB()
    
    r = 0
    arr = ["КБ", "АКБ", "17-18 дюйм", "АКБ Лето", "17-18 Лето", "АКБ Зима", "17-18 Зима", "ЛГШ"]
    
    
    for i in data.getItems():
      head = i.values()
      head.extend(arr)
      xlb.makeHead(sheet, r, head);
      r+=1
      
      for i2 in i.getItems():
        xlb.makeCells(sheet, r, i2.values())
        r += 1
    
    pos = 2
    
    for i in data.getItems():
      head = i.values()
      s1 = pos
      s2 = pos+len(i.items)-1
      arr = ["=SUM(B{0}:B{1})".format(s1, s2),
             "=SUM(C{0}:C{1})".format(s1, s2),
             "=SUM(D{0}:D{1})".format(s1, s2),
             "=SUM(E{0}:E{1})".format(s1, s2),
             "=SUM(F{0}:F{1})".format(s1, s2),
             "=SUM(G{0}:G{1})".format(s1, s2),
             "=SUM(H{0}:H{1})".format(s1, s2),
             "=SUM(I{0}:I{1})".format(s1, s2)]
      
      head.extend(arr)
      xlb.makeHead(sheet, r, head);
      r+=1
      pos += 1 + len(i.items)
    
    cc = 1
    
    for w in [20, 20, 20, 20, 20, 20, 20, 20, 20]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
        
    return wb
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
    logging.info('start report')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(ReportData(), params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "kboffline_report.xlsx", server)                
    logging.info('end')
    
