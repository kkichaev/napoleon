# -*- coding: cp1251 -*-
from importlib import reload
import logging
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from grsoft.route import AgentRoute
from manager.document import docTypes
from datetime import timedelta
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat,Border

import sys;
reload(sys);
#sys.setdefaultencoding("cp1251")

class Data:
  __slots__ = ["items"]
  
  def __init__(self):
    self.items = []

class Item:
  __slots__ = ["orgname", "scriptname", "docname", "inroute", "created", "sended", "sum", "remark"]
  
  def __init__(self):
    self.orgname = ''
    self.scriptname = ''
    self.docname = ''
    self.inroute = ''
    self.created = ''
    self.sended = ''
    self.sum = 0
    self.remark = ''
    
  def getData(self, row):
    return [self.created.strftime("%d.%m.%Y %H:%M"), self.orgname, self.docname, 
      self.scriptname, self.inroute, self.sended.strftime("%d.%m.%Y %H:%M"),
      self.sum, self.remark]
  
class AgentSheet:
  __slots__ = ["id", "name", "items"]
  
  def __init__(self):
    self.id = ""
    self.name = ""
    self.items = []

def loadData(params, server):
  data = Data()
  
  for item in params.userids:
    id = item.id
    sheet = AgentSheet()
    sheet.id = id
    data.items.append(sheet)
    
    server.ChangeUser("'" + id + "'")
    orgs = server.Get("Org", "", "id")
    sheet.name = server.CurrentUser().name
    scrDef = server.Get("ScriptDef", "", "id")
    server.RestoreUser()
     
    ar = AgentRoute(server, id)
    date = params.start
    
    while date <= params.finish:
      route = ar.getDayRoute(date)  
      
      route_ids = []
      
      for i in route:
        route_ids.append(i.id)
      
      where = '"userid"="{0}" and "created" >= ToDate("{1}") and "created" <= ToDate("{2}")'.format(
          id,
          date.strftime("%d/%m/%Y 00:00:00"),
          date.strftime("%d/%m/%Y 23:59:00"))
      
      scripts = server.Get("ScriptDoc", where)
      
      scrMap = dict()
      
      for s in scripts:
        for i in s.items:
          if i.state == 1:
            scrMap[i.date] = s.scriptId 
        
      for dt in docTypes:
        docs = dt.docList(server, where)
        
        if docs == None or dt.objectName == "ScriptDoc":
          continue
        
        for d in docs:
          item = Item()
          sheet.items.append(item)
          
          item.orgname = orgs[d.id].name if d.id in orgs else "Контрагент с кодом <{0}>".format(d.id)
          item.docname = dt.title
          item.created = d.created
          item.inroute = "нет" if not d.id in route_ids else ""
          item.sended = d.sended
          
          t = d.created
          if t in scrMap and scrMap[t] in scrDef:
            item.scriptname =  scrDef[scrMap[t]].name
            
          item.sum = d.sum()
              
      date += timedelta(days=1)

  data.items = sorted(data.items, key=lambda lhs: lhs.name)
  
  for i in data.items:
    i.items = sorted(i.items, key=lambda lhs: lhs.created)
  
  return data
  
class XLBuilderEx(XLBuilder):
  HEAD   = 'FFF2F2F2'
  
  def paintHeadCell(self, cell):
    XLBuilder.paintHeadCell(self, cell)
    self.setBackColor(cell,XLBuilderEx.HEAD)
    
  def makeCell(self, sheet, row, column, cell, value, border = Border.BORDER_THIN):
    if column == 6:
      value = "" if value == 0 else value
      
    XLBuilder.makeCell(self, sheet, row, column, cell, value, border) 

    if column == 6:
      cell.style.number_format._set_format_code(NumberFormat.FORMAT_NUMBER_00)
    
def printOut(data, params):
  wb = Workbook(False, 'cp1251')
  sheet = wb.get_active_sheet()
  
  for page in data.items:
    sheet.title = page.name[:31]
    cell = sheet.cell(row=0, column=0)
    cell.value = "Отчет по посещениям: {0}".format(page.name)
    cell.style.font.bold = True
    
    DATE_FMT = "%d.%m.%Y %H:%M"
    sheet.cell(row=1, column=0).value = "Период: {0} - {1}".format(params.start.strftime(DATE_FMT), params.finish.strftime(DATE_FMT))
    
    head = ["Дата", "Контрагент", "Тип посещения", "Сценарий", "План", "Дата передачи", "Сумма", "Комментарий"]
    xlb = XLBuilderEx()
    xlb.makeHead(sheet,3,head)
    sheet.row_dimensions[sheet.cell(row=3, column=0).row].height = 34
    
    r = 4
    for item in page.items:
      xlb.makeCells(sheet, r,  item.getData(r))
      r += 1

    x = 1
    for w in [18,40,17,19,7,16,11,25]:
      sheet.column_dimensions[get_column_letter(x)].width = w
      x += 1

    sheet = wb.create_sheet()
  
  return wb

def run(server):
  logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout, level=logging.DEBUG)    
  logging.info('start report')

  params = server.Params[0]
  logging.info("params " + str(params))

  data = loadData(params, server)
  wb = printOut(data, params)

  XLBuilder().workbookToObject(wb, "rmr_visit_report.xlsx", server)                
  logging.info('end')
