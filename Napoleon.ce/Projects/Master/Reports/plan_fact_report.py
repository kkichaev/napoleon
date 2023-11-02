# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from operator import attrgetter
from openpyxl.cell import get_column_letter
from decimal import *
from openpyxl.style import Color, Fill, Font

import sys;
import tempfile
import io
import time

reload(sys);
sys.setdefaultencoding("cp1251")

scheduleStart = dict()

def getDivisionAgents(server, division, agents):
    if division == None or len(division) == 0:
        return
    
    for d in division :
        for a in d.agents :
            agents.append(a.id)
    
        getDivisionAgents(server, server.Get("Division", '"parent"=' + str(d.id)), agents)
      
def loadAgents(server, division):
    ret = dict()
    
    agents = server.Get("Agents", "")
    
    divagents = []
    getDivisionAgents(server, server.Get("Division", '"id"=' + str(division)), divagents)
    divagents = set(divagents)
    
    for a in agents:
        if a.id in divagents:
            ret[a.id] = a
    
    return ret    

def makeIDStr(server, agents):
    res = '"userid" in ('
    
    for id in agents.iterkeys():
        res += "'" + id + "',"
    
    res = res[:-1] + ")"
    return res

def borders(sheet, sr, c1, c2):
    for cl in range(c1, c2) :
       c = sheet.cell(row=sr, column=cl)
       c.style.borders.top.border_style = Border.BORDER_THIN
       c.style.borders.bottom.border_style = Border.BORDER_THIN
       c.style.borders.right.border_style = Border.BORDER_THIN 

def getWeekIndex(server, data, agentid):  
    scStart = None
     
    if agentid in scheduleStart: 
        scStart = scheduleStart[agentid]
    else :
        where = '"userid"' + " in ('" + agentid + "')"
        cfg = server.Get("ServerConfig", where)
        
        if cfg != None:
            for c in cfg:
                if c.key == 'SheduleStart':
                    strptime = lambda date_string, format: datetime(*(time.strptime(date_string, format)[0:6])) 
                    scStart = strptime (c.value, '%Y-%m-%d')
                    break
                
        scheduleStart[scStart] = scStart;
    
    result = -1
    
    if scStart != None:
        d = data - scStart
        result = ((d.days / 7) % 4) + 1;
    
    return result  

      
def getRoutePerDay(server, orgFolder, a, b1):
    days = {"Monday" : "Понедельник", 
                "Tuesday" : "Вторник",
                "Wednesday" : "Среда",
                "Thursday" : "Четверг",
                "Friday" : "Пятница",
                "Saturday" : "Суббота",
                "Sunday" : "Воскресенье",
        };
    d = days[b1.strftime("%A")]
    widx = getWeekIndex(server, b1, a.id)
    dailyRoute = server.Get("DailyRoute", '"date"' + " = ToDate('" + b1.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                                    '"userid"' + " in ('" + a.id + "')")
    plans = list();
    
    if orgFolder != None:
        for of in orgFolder:
            if of.name == d or of.name == str(widx) + d:
                for i in of.items:
                    if not i.name in plans:
                        plans.append(i.name) 
            
    if dailyRoute != None:
        for dr in dailyRoute:
            for i in dr.items:
                if not i.id in plans:
                    plans.append(i.id)
    
    return plans                

def daysDiff(begin, end):
    a = begin
    result = 0
    weekIdx = begin.isocalendar()[1]
    
    while a <= end:
        if weekIdx != a.isocalendar()[1]:
              weekIdx = a.isocalendar()[1] 
              result += 1
        a += timedelta(days=1)
        result += 1
              
    return result
                                               
def run(server):
    print "wt_report start"
    print ""
    
    #getcontext().prec = 2 
    #getcontext().rounding = ROUND_05UP
    
    params = server.Params
    param = params[0]

    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    
    wb = Workbook(False, 'cp1251')
    sheet = None
    sheet = wb.get_active_sheet()
    sheet.title = "развернутый по менеджеру"
    sheet.freeze_panes = sheet.cell('A4') 
    
    sheet.cell(row=0, column=0).value = "График и выполнение посещений с " +\
        param.begin.strftime("%d/%m/%Y") + " по " + param.end.strftime("%d/%m/%Y")
        
    col_letter = get_column_letter(1) 
    sheet.column_dimensions[col_letter].width = 30     
    
    idx = daysDiff(param.begin, param.end)
    
    sheet.cell(row=1, column=idx+1).value = "итого:"
    sheet.cell(row=2, column=idx+1).value = "план"
    sheet.cell(row=2, column=idx+2).value = "+"
    sheet.cell(row=2, column=idx+3).value = "-"   
    col_letter = get_column_letter(idx + 2) 
    sheet.column_dimensions[col_letter].width = 6
    col_letter = get_column_letter(idx + 3) 
    sheet.column_dimensions[col_letter].width = 6
    col_letter = get_column_letter(idx + 4) 
    sheet.column_dimensions[col_letter].width = 6   
    
    a = param.begin
    weekIdx =  param.begin.isocalendar()[1]
    c = 1
    
    while a <= param.end:
        if weekIdx != a.isocalendar()[1]:
           weekIdx = a.isocalendar()[1] 
           c += 1
           
        sheet.cell(row=2, column=c).value = a.strftime("%d/%m")
        col_letter = get_column_letter(c+1) 
        sheet.column_dimensions[col_letter].width = 7  
        a += timedelta(days=1)
        c += 1
    
    sr = 3
    
    if param.divisionID != 0 :
        agents = loadAgents(server, param.divisionID)
        
        al = list()
        
        for a in agents.values():
            al.append(a)
        
        al.sort(cmp=lambda x,y: cmp(x.name.lower(), y.name.lower()))
            
        for a in al:
            agentIdFilter = '"userid"' + " in ('" + a.id + "')"
            orgFolder = server.Get("OrgFolder", agentIdFilter)  
            sheet.cell(row=sr, column=0).value = a.name   
            sheet.cell(row=sr, column=0).style.font.bold = True
            sheet.cell(row=sr, column=0).style.font.underline = Font.UNDERLINE_SINGLE
            orgs = server.Get("Org", agentIdFilter, "id")
            
            plans = list()
            b1 = param.begin
            
            workDate = datetime.now()

            while(b1 <= param.end):
                pp = getRoutePerDay(server, orgFolder, a, b1)
                
                if len(pp) > 0:
                    for p in pp:
                        if not p in plans:
                            plans.append(p)
                                
                b1 += timedelta(days=1)

            agentSR = sr            
            sr += 1  
            totalPlanAgent = 0
            totalVisitAgent = 0
            totalUnvisitAgent = 0
            weekPlan = dict()
            weekUnvisit = 0
            
            unvisit = dict()
            visitplan = dict()
            totalplan = dict()
            dayPlan = dict()
            dayFact = dict()
            
            weekPlan[weekIdx] = 0
            for pl in plans:
                if orgs.has_key(pl):
                    sheet.cell(row=sr, column=0).value = orgs[pl].name
                    clmn = 1
                    b = param.begin
                    weekIdx = param.begin.isocalendar()[1]
                        
                    while(b <= param.end):
                        visited = False
                        
                        if weekIdx != b.isocalendar()[1]:
                           weekIdx = b.isocalendar()[1] 
                           
                           if weekUnvisit > 0:
                               sheet.cell(row=sr, column=clmn).value = str(weekUnvisit)
                               
                           sheet.cell(row=sr, column=clmn).style.fill.fill_type = Fill.FILL_SOLID
                           sheet.cell(row=sr, column=clmn).style.fill.start_color.index = 'ffffcf60'
                           weekUnvisit = 0
                           clmn += 1
                            
                        whereStr = '"created"' + " >= ToDate('" + b.strftime("%d/%m/%Y 0:0:0") + "') and " +\
                            '"created"' + " <= ToDate('" + b.strftime("%d/%m/%Y 23:59:59") + "') and " +\
                            '"userid"' + " in ('" + a.id + "') and " + '"id" = ' + "'" + orgs[pl].id + "'"
                        
                        orders = server.Get("Order", whereStr)
                        visited = len(orders) > 0
                        
                        if not visited:
                            visits = server.Get("Visit", whereStr)
                            visited = len(visits) > 0
                        
                        if not visited:       
                            remnants = server.Get("OrgRemnants", whereStr)
                            visited = len(remnants) > 0

                        if not visited:                                        
                            monitorings = server.Get("Monitoring", whereStr) 
                            visited = len(monitorings) > 0      
                        
                        dayRoute = getRoutePerDay(server, orgFolder, a, b)
                        
                        if visited and orgs[pl].id in dayRoute:
                            sheet.cell(row=sr, column=clmn).value = "+"
                            if pl in totalplan:
                                totalplan[pl] += 1
                            else:    
                                totalplan[pl] = 1  
                            
                            if pl in visitplan:        
                                visitplan[pl] += 1
                            else:
                                visitplan[pl] = 1
                                    
                            totalPlanAgent += 1
                            totalVisitAgent += 1
                            
                            if weekIdx in weekPlan:
                                weekPlan[weekIdx] += 1
                            else:
                                weekPlan[weekIdx] = 1     
                            
                            if b in dayPlan:
                                dayPlan[b] += 1
                            else:
                                dayPlan[b] = 1
                            
                            if b in dayFact:
                                dayFact[b] += 1
                            else:
                                dayFact[b] = 1
                        else:
                            if orgs[pl].id in dayRoute:
                                if pl in totalplan:
                                    totalplan[pl] += 1
                                else:    
                                    totalplan[pl] = 1   
                                    
                                totalPlanAgent += 1 
                                
                                if weekIdx in weekPlan:
                                    weekPlan[weekIdx] += 1
                                else:
                                    weekPlan[weekIdx] = 1
                                
                                if b in dayPlan:
                                    dayPlan[b] += 1
                                else:
                                    dayPlan[b] = 1
                                        
                                if workDate > b:
                                    sheet.cell(row=sr, column=clmn).value = "-"
                                    sheet.cell(row=sr, column=clmn).style.fill.fill_type = Fill.FILL_SOLID
                                    sheet.cell(row=sr, column=clmn).style.fill.start_color.index = Color.RED
                                    
                                    if pl in unvisit:
                                        unvisit[pl] += 1
                                    else:
                                        unvisit[pl] = 1
                                         
                                    totalUnvisitAgent += 1
                                    weekUnvisit += 1
                                else:         
                                    if orgs[pl].id in dayRoute:
                                        sheet.cell(row=sr, column=clmn).value = "п"
                                        
                        b += timedelta(days=1)
                        clmn += 1
                    
                    if pl in totalplan:    
                        sheet.cell(row=sr, column=clmn).value = str(totalplan[pl])
                        
                    clmn += 1
                    
                    if pl in visitplan:    
                        sheet.cell(row=sr, column=clmn).value = str(visitplan[pl])    
                        
                    clmn += 1
                    
                    if pl in unvisit:    
                        sheet.cell(row=sr, column=clmn).value = str(unvisit[pl])
                        
                    sr += 1  
            
            b1 = param.begin
            c = 1
            weekIdx =  param.begin.isocalendar()[1]
            while b1 <= param.end:
                plan = 0
                fact = 0
                
                if weekIdx != b1.isocalendar()[1]:
                   weekIdx = b1.isocalendar()[1]
                    
                   if weekIdx - 1 in weekPlan:
                        sheet.cell(row=agentSR, column=c).value = str(weekPlan[weekIdx - 1]) 
                        
                   c += 1
                           
                if b1 in dayPlan:
                    plan = dayPlan[b1]
                
                if b1 in dayFact:
                    fact = dayFact[b1]
                        
                sheet.cell(row=agentSR, column=c).value = str(plan) + "/" + str(fact)
                b1 += timedelta(days=1) 
                c += 1
                
            if totalPlanAgent > 0:        
                sheet.cell(row=agentSR, column=idx + 1).value = str(totalPlanAgent)
                
            if totalVisitAgent > 0:    
                sheet.cell(row=agentSR, column=idx + 2).value = str(totalVisitAgent)
            
            if totalUnvisitAgent > 0:    
                sheet.cell(row=agentSR, column=idx + 3).value = str(totalUnvisitAgent)
                
            sr += 1
            
            
    print ""
    print "wt_report done " ;
    
    repName = "res.xlsx"
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)

    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytes
    
    server.Put(outObj)