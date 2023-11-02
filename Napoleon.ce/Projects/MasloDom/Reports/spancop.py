# -*- coding: cp1251 -*-

# types write string without space
# s - string
# n(prec) - double(number), prec == 0  integer
# n - integer
# d - date
# t - time
# dt - datetime
# b - binary
#


from openpyxl import Workbook
from openpyxl.style import Border
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.style import Fill
from operator import attrgetter
from openpyxl.cell import get_column_letter
from datetime import date
import datetime
import tempfile
import io


import sys;
reload(sys);
sys.setdefaultencoding("cp1251")

def get_module_info():
    result = module_info()
    result.name = "distrib"
    result.description = "Процент дистрибуции"
    return result
        
def doReport(server, outObj):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.auto_filter = "a1:w1"
    spancop = server.Get("Spancop", "")
    users = server.Get("Agents", "", "id")
    orgs = server.Get("Org", "", "id")
    categories = server.Get("CategoryProduct", "", "key")
    segments = server.Get("Segment", "", "key")
    chances = server.Get("Chance", "", "key")
    levels = server.Get("ClientLevel", "", "key")
    competitors = server.Get("Competitor", "", "key")
    
    cell = sheet.cell(row=0, column=0)
    cell.value = "Дистрибьютор"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=1)
    cell.value = "Менеджер по продажам"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=2)
    cell.value = "Клиент"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=3)
    cell.value = "Существующий клиент?"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=4)
    cell.value = "Категория продукта"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell = sheet.cell(row=0, column=5)
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell.value = "Сегмент"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=6)
    cell.value = "Предположительная дата 1го заказа"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=7)
    cell.value = "Основные причины Достигнутого успеха или Потери бизнеса (конкретнее)"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=8)
    cell.value = "Вероятность Успеха"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=9)
    cell.value = "Предполагаемый годовой объем (на данный момент)"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=10)
    cell.value = "Текущая оценка прироста до конца " + str(date.today().year)
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=11)
    cell.value = "Фактический прирост объема с начала " + str(date.today().year)
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=12)
    cell.value = "S"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=13)
    cell.value = "P"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=14)
    cell.value = "A"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=15)
    cell.value = "N"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=16)
    cell.value = "C"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=17)
    cell.value = "O"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=18)
    cell.value = "P"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=19)
    cell.value = "Примечания"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=20)
    cell.value = "Холдинговая компания"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=21)
    cell.value = "Уровень клиента"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    cell = sheet.cell(row=0, column=22)
    cell.value = "Конкурент"
    cell.style.font.bold = True
    cell.style.alignment.wrap_text = True
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
    
    sheet.row_dimensions[1].height = 45
    
    rowid = 1
    
    for sp in spancop :
        sheet.cell(row=rowid, column=0).value = "Маслодом"
        
        if sp.userid in users:
            sheet.cell(row=rowid, column=1).value = users[sp.userid].name
        else:
            sheet.cell(row=rowid, column=1).value = sp.userid
        
        if sp.id in orgs:
            sheet.cell(row=rowid, column=2).value = orgs[sp.id].name
        else:
            sheet.cell(row=rowid, column=2).value = sp.org
            
        sheet.cell(row=rowid, column=3).value = "Да" if sp.realclient else "Нет"
        
        if sp.category in categories:  
            sheet.cell(row=rowid, column=4).value = categories[sp.category].value
        else:
            sheet.cell(row=rowid, column=4).value = sp.category       
        
        if sp.segment in segments:
            sheet.cell(row=rowid, column=5).value = segments[sp.segment].value
        else:
            sheet.cell(row=rowid, column=5).value = sp.segment
            
        sheet.cell(row=rowid, column=6).value = sp.first  
        sheet.cell(row=rowid, column=7).value = sp.success
        
        if sp.chance in chances:
            sheet.cell(row=rowid, column=8).value = chances[sp.chance].value
        else:
            sheet.cell(row=rowid, column=8).value = sp.chance     
        
        sheet.cell(row=rowid, column=9).value = sp.cub
        
        
        sheet.cell(row=rowid, column=10).value = str((date.today().replace(month=12, day=31) - sp.first.date()).days * sp.cub/365);
        sheet.cell(row=rowid, column=11).value = str((date.today().replace(month=12, day=31) - sp.s.date()).days * sp.cub/365);
        sheet.cell(row=rowid, column=12).value = sp.s
        sheet.cell(row=rowid, column=13).value = sp.p1
        sheet.cell(row=rowid, column=14).value = sp.a 
        sheet.cell(row=rowid, column=15).value = sp.n
        sheet.cell(row=rowid, column=16).value = sp.c
        sheet.cell(row=rowid, column=17).value = sp.o
        sheet.cell(row=rowid, column=18).value = sp.p2
        sheet.cell(row=rowid, column=19).value = sp.remark
        sheet.cell(row=rowid, column=20).value = sp.holding
        
        if sp.clientLevel in levels:
            sheet.cell(row=rowid, column=21).value = levels[sp.clientLevel].value
        else:
            sheet.cell(row=rowid, column=21).value = sp.clientLevel
         
        if sp.competitor in competitors:
            sheet.cell(row=rowid, column=22).value = competitors[sp.competitor].value
        else:        
            sheet.cell(row=rowid, column=22).value = sp.competitor
            
        rowid = rowid + 1
    
   
    sheet.column_dimensions[get_column_letter(1)].width = 18
    sheet.column_dimensions[get_column_letter(2)].width = 23
    sheet.column_dimensions[get_column_letter(3)].width = 30
    sheet.column_dimensions[get_column_letter(4)].width = 10
    sheet.column_dimensions[get_column_letter(5)].width = 14
    sheet.column_dimensions[get_column_letter(6)].width = 14
    sheet.column_dimensions[get_column_letter(7)].width = 17
    sheet.column_dimensions[get_column_letter(8)].width = 45
    sheet.column_dimensions[get_column_letter(9)].width = 15
    sheet.column_dimensions[get_column_letter(10)].width = 19
    sheet.column_dimensions[get_column_letter(11)].width = 27
    sheet.column_dimensions[get_column_letter(12)].width = 27
    sheet.column_dimensions[get_column_letter(13)].width = 11
    sheet.column_dimensions[get_column_letter(14)].width = 11
    sheet.column_dimensions[get_column_letter(15)].width = 11
    sheet.column_dimensions[get_column_letter(16)].width = 11
    sheet.column_dimensions[get_column_letter(17)].width = 11
    sheet.column_dimensions[get_column_letter(18)].width = 11
    sheet.column_dimensions[get_column_letter(19)].width = 11
    sheet.column_dimensions[get_column_letter(20)].width = 25
    sheet.column_dimensions[get_column_letter(21)].width = 15
    sheet.column_dimensions[get_column_letter(22)].width = 15
    sheet.column_dimensions[get_column_letter(23)].width = 15
    

    repName = "distrib.xlsx" 
    fileName = tempfile.gettempdir() + '/' + repName
    wb.save(fileName)
    
    file = io.open(fileName, 'rb')
    bytes = file.read(-1)
    file.close()
    
    obj = outObj.New()
    obj.name = repName
    obj.file = bytes

def run(server):

    print "start  ", __name__

    server.RegisterType("Result[name:s,file:b]")
    outObj = server.New("Result")
    doReport(server, outObj)
    server.Put(outObj)

    print "finish ", __name__


   
