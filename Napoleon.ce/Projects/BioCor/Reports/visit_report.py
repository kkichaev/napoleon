# -*- coding: cp1251 -*-

from datetime import timedelta
from datetime import datetime
from openpyxl import Workbook
from grsoft.xl_base import XLBuilder
from openpyxl.style import Alignment
from openpyxl.style import Color
from openpyxl.cell import get_column_letter
from openpyxl.style import Border

import tempfile
import sys;

reload(sys);
sys.setdefaultencoding("cp1251")

bkgColor = "ff90ffff"

class RouteItem:
    date = None
    pos = None
    org = None
    addr = None
    contact = None
    phone = None
    modify = None
    comment = None
    
    def __init__(self):
      self.org = ""
      self.addr = ""
      self.contact = " "
      self.phone = " "
      self.comment = ""
      
    def getData(self):
        return [self.date.strftime("%d.%m.%Y"), self.pos, self.org, self.addr, self.contact, self.phone, self.modify.strftime("%d.%m.%Y"),
                self.comment]

class VztItem(RouteItem):
    sklad = None
    polka = None
    time = None
    visits = None
    contact = None
    phone = None
    
    def __init__(self):
        self.sklad = ""
        self.polka = ""
        self.visits = list()
        self.contact = " "
        self.phone = " "
        
    def getData(self):
        result = [self.date.strftime("%d.%m.%Y"), self.pos, self.org, self.addr, self.contact, self.phone, self.sklad, self.polka, self.comment]
        result.append(str(self.time))
        
        for v in self.visits:
          result.append('=HYPERLINK("{0}","{1}")'.format(v.name, v.remark if len(v.remark) > 0 else "Фото"))
        
        return result    

class Photo:
    name = None
    id = None
    remark = None
    
    def __init__(self):
      self.name = ""
      self.remark = ""
      self.id = ""
      
class EmptyRow:
    def draw(self, sh, xlb, st, fn, width):
        # sh.merge_cells(start_row=st, start_column = 0, end_row = fn-1, end_column = 0);
        xlb.drawEmpytRow(sh, fn, width)
        
class PlanEmptyRow(EmptyRow):
    def draw(self, sh, xlb, st, fn, width):
        # sh.merge_cells(start_row=st, start_column = width - 1, end_row = fn - 1, end_column = width - 1);
        EmptyRow.draw(self, sh, xlb, st, fn, width) 
                 
def initOrg(id, i, orgs):
    if id in orgs:
        i.org = orgs[id].name
        i.addr = orgs[id].address
         
        if len(orgs[id].contacts) > 0:
            c = orgs[id].contacts[0]
            i.contact = c.name
            i.phone = c.phone
    else:
        i.org = "Контрагент с кодом<{0}>".format(id)
        i.addr = ""
        i.contact = ""
        i.phone = ""
            
WHERE_STR = '"{3}" >= ToDate("{0}") and "{3}" <= ToDate("{1}")  and "userid"="{2}"'; 

def inflateParams(server):
    return server.Params[0].begin, server.Params[0].end, server.Params[0].userid, server.Params[0].photo
        
def collectWorkTime(server):
    start, finish, userid, dummy = inflateParams(server)
           
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid, "start")
    worktimes = server.Get("WorkTime", where)
    
    wtdata = dict()
    for w in worktimes:
        if not w.start.date() in wtdata:
            wtdata[w.start.date()] = dict()
        
        if w.id in wtdata[w.start.date()]:  
            wtdata[w.start.date()][w.id] = wtdata[w.start.date()][w.id] + w.stop - w.start    
        else:
            wtdata[w.start.date()][w.id] = w.stop - w.start
    
    return wtdata

def collectVizitData(server, wtdata, remnants, orgs, price, visits):
    href = server.Params[0].href

    dt = None
    pos = 0
    
    var = dict()
    
    for v in visits:
      if not v.id in var:
        var[v.id] = dict()
        
      vvar = var[v.id]
      dt = v.date.replace(hour=0, minute=0, second=0, microsecond=0) 
      if not dt in vvar:
        vvar[dt] = list()
       
      vvar[dt].append(v)
    
    maxvizcnt = 0
    
    for r in remnants:
        dt = r.created.replace(hour=0, minute=0, second=0, microsecond=0) 
        
        if r.id in var and dt in var[r.id]:
          for vv in var[r.id][dt]:
            sz = len(vv.items)
            if sz > maxvizcnt:
              maxvizcnt = sz
   
    sd1 = list()
    filenum = 1
    for r in remnants:
        i = VztItem()
        i.date = r.created
        
        if dt == None or dt.date() != r.created.date():
            pos = 1
            
            if dt != None:
              sd1.append(EmptyRow())
              
            dt = r.created
         
        i.data = r.created
        i.pos = pos
        i.comment = r.remark
        initOrg(r.id, i, orgs) 
        
        dt = i.date.replace(hour=0, minute=0, second=0, microsecond=0) 
        
        ic = 1
        if r.id in var and dt in var[r.id]:
          for vv in var[r.id][dt]:
            for vvv in vv.items:
              p = Photo()
              p.name = "{0}{1}".format(href,vvv.name)
              p.remark = vv.remark if len(vv.remark) > 0 else "Фото " + str(ic)
              filenum += 1
              ic += 1
              i.visits.append(p)
        
        vizcnt = len(i.visits)
        
        while vizcnt < maxvizcnt:
          i.visits.append(Photo())
          vizcnt += 1
          
        if (r.created.date() in wtdata) and (r.id in wtdata[r.created.date()]):
            i.time = wtdata[r.created.date()][r.id]
              
        for ri in r.items:
            nm = "Товар с кодом<{0}>".format(ri.id)
             
            if ri.id in price:
                nm = price[ri.id].name
                 
            if ri.qty > 0:
                if len(i.sklad) > 0:
                    i.sklad += ", "
                      
                i.sklad += nm
                  
            if ri.shelf > 0:
                if len(i.polka) > 0:
                    i.polka += ", "
                                   
                i.polka += nm
        pos += 1
        sd1.append(i)   
    
    return sd1, maxvizcnt
    
def collectPlanData(orgs, orgf): 
    sd2 = list()
    for f in orgf:
        pos = 1
        for fi in f.items:
            i = RouteItem()
            i.date = f.date 
            i.pos = pos
            i.modify = f.modify 
            i.comment = fi.comment
            f.items.sort(cmp=lambda x, y: cmp(x.pos, y.pos))
        
            initOrg(fi.name, i, orgs)
            pos += 1
         
            sd2.append(i)   
        
        sd2.append(PlanEmptyRow())    
        
    return sd2
    
def collectReportData(server):
    start ,finish, userid, dummy = inflateParams(server)
    
    where = "\"id\"='{0}'".format(userid)
    agent = server.Get("Agents", where, "id")    
    
    rd = RptData()
    rd.agent = agent[userid].name if userid in agent else userid
    rd.start = start
    rd.finish = finish
    
    return rd
    
def loadData(server):
    start, finish, userid, usephoto = inflateParams(server)
    
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid, "created")
    remnants = server.Get("OrgRemnants", where)
    remnants.sort(cmp= lambda x, y: cmp(x.created, y.created))
    
    visits = []
    if usephoto == 1:
      visits = server.Get("Visit", where)
      
    server.ChangeUser("'" + userid + "'")
    orgs = server.Get("Org", "", "id")
    
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    server.RestoreUser()
    
    where = WHERE_STR.format(start.strftime("%d/%m/%Y 0:0:0"), finish.strftime("%d/%m/%Y 0:0:0"), userid, "date")
    orgf = server.Get("OrgFolder", where)
    orgf.sort(cmp= lambda x, y: cmp(x.date, y.date))
    
    sd1, vsz = collectVizitData(server, collectWorkTime(server), remnants, orgs, server.Get("ManagerPrice", "", "id"), visits)
    sd2 = collectPlanData(orgs, orgf)
    rd = collectReportData(server) 
    
    return rd, sd1, sd2, vsz 


class XLBuilderEx(XLBuilder):
    def adjustHeadCell(self, sheet, cell, row, column):
        cell.style.alignment.wrap_text = True
        self.setBackColor(cell, bkgColor)
        return column    
    
    def makeCell(self, sheet, row, column, cell, value):
        XLBuilder.makeCell(self, sheet, row, column, cell, value)
        cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cell.style.alignment.vertical = Alignment.VERTICAL_CENTER
        
        if column == 9:
            cell.style.number_format.format_code = 'h:mm:ss'
        
        
    def drawEmpytRow(self, sheet, row, cc):
        for c in range(0,cc):
            c = sheet.cell(row=row, column=c)
            self.setBackColor(c, bkgColor)
            self.makeBorder(c, Border.BORDER_THIN)

def makeHead(sh, head, xlb, title, rd):
    xlb.makeHead(sh, 1, head)
    sh.merge_cells(start_row=0, start_column = 0, end_row = 0, end_column = len(head) - 1) 
    cell = sh.cell(row=0, column=0)
    cell.value = "{3}, {0} c {1}г. по {2}г.".format(rd.agent, rd.start.strftime("%d.%m.%Y"), rd.finish.strftime("%d.%m.%Y"), title)
    cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    cell.style.font.bold = True
    cell.style.font.color.index = Color.RED
    xlb.setBackColor(cell, bkgColor)

def setCellWidth(sh, wa):
    cc = 1
    for w in wa:
        sh.column_dimensions[get_column_letter(cc)].width = w
        cc += 1
        
def setPageHeader(sh, xlb, rpd, caption, title, head):
    sh.title = title
    makeHead(sh, head, xlb, caption, rpd)

def drawData(sh, xlb, cc, data):
    row = 2
    er = row
    for d in data:
        if isinstance(d, EmptyRow):
            d.draw(sh, xlb, er, row, cc)
            er = row + 1 
        else: 
            rd = d.getData()
            xlb.makeCells(sh, row, rd)
        row += 1
                    
def ptintSheetVisit(xlb, sh, rpd, d, vsz):
    head = ["Дата", "№", "Контрагент", "Адрес", "Контактное лицо", "Телефон", "Наличие препаратов", "На выкладке", "Комментарии к визиту"]
    head.append("Продолжительность визита")  
    
    setPageHeader(sh, xlb, rpd, "Отчет о визитах", "Отчет", head)
    drawData(sh, xlb, len(head), d) 
    setCellWidth(sh, [10,5,25,25,20,20,25,25,30])

def printSheetPlan(xlb, sh, rpd, d):
    head = ["Дата", "№", "Контрагент", "Адрес", "Контактное лицо", "Телефон", "Последнее изменение", "Комментарий"]
    setPageHeader(sh, xlb, rpd, "План визитов", "План", head)
    drawData(sh, xlb, len(head), d)
    setCellWidth(sh, [15,5,25,25,20,20,15,25])
           
def printOut(rpd, d1, d2, vsz):
    wb = Workbook(False, 'cp1251')
    sh = wb.get_active_sheet()
    
    xlb = XLBuilderEx()
    ptintSheetVisit(xlb, sh, rpd, d1, vsz)
    
    sh = wb.create_sheet()
    printSheetPlan(xlb, sh, rpd, d2)
                
    return wb

class RptData:
    agent = None
    start = None
    finish = None
          
def doReport(server):
    rd, data1, data2, vsz = loadData(server)
    wb = printOut(rd, data1, data2, vsz)
    workbookToObject(wb, "visit.xlsx", server, data1)

def workbookToObject(wb, repName, server, data1):
    server.RegisterType("Result[name:s,file:b,items[name:s,file:b]]")
    outObj = server.New("Result")

    tFile = tempfile.TemporaryFile()
    wb.save(tFile)
    tFile.seek(0)
    bytesOut = tFile.read(-1)
    tFile.close()
    
#         fileName = tempfile.gettempdir() + '/' + repName
#         wb.save(fileName)
#     
#         file = io.open(fileName, 'rb')
#         bytes = file.read(-1)
#         file.close()

    obj = outObj.New()
    obj.name = repName
    obj.file = bytesOut
    
    for d in data1:
      if isinstance(d, VztItem):
        for v in d.visits:
            if v.id != None:
              i = obj.items.New()
              i.name = v.name
              i.file = v.id

    server.Put(outObj)
        
def run(server):
    print "start\t" + __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    doReport(server)
    print "finish\t" +  __name__ + "\t" + datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    