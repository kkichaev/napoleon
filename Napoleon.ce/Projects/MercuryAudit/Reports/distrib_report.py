# -*- coding: cp1251 -*-
from importlib import reload
import sys;
import logging
# import random

from grsoft.route import AgentRoute
from grsoft.orgLocation import OrgLocation, LocationPoint

from grsoft.xl_base import XLBuilder
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import datetime


reload(sys);
#sys.setdefaultencoding("cp1251")

class DocData:
    __slots__ = ['created', 'orgname', 'orgaddress', 'partval', 'orgcateg', 'fc', 'numberho', 'pricevalues', 'remark']
    
    def __init__(self):  
      pass
    
    def values(self):
      result = [self.created.strftime('%d.%m.%Y'), self.orgname, self.orgaddress, self.partval, self.orgcateg, self.fc, self.numberho]
      result.extend(self.pricevalues)
      result.append(self.remark)
      
      return result
                    
class ReportData:
    __slots__ = ['price', 'data', 'pcidx', 'orgs', 'orgType']
    
    def __init__(self):
        self.data = list()
        self.pcidx = dict()
        
    def add(self, doc):
        data = DocData()
        data.created = doc.created
        
        on = '';
        oa = '';
        oc = ''
        
        if doc.id in self.orgs:
          o = self.orgs[doc.id]
          on = o.name
          oa = o.address
          
          if o.typeID in self.orgType:
            for i in self.orgType[o.typeID].items:
              if i.id == o.categID:
                oc = i.name
                break
          
        data.orgname = on
        data.orgaddress = oa
        data.partval = 0
        data.orgcateg = oc
        
        data.numberho = doc.numberho
        data.remark = doc.remark
        data.pricevalues = []
        
        for i in range(0, len(self.price)):
          data.pricevalues.append('')
        
        ownerItems = 0
        concurentItems = 0
        
        for i in doc.items:
          if i.id in self.pcidx:
            data.pricevalues[self.pcidx[i.id]] = i.qty
            ownerItems += i.qty
          else:
            concurentItems += i.qty
        
        
        data.partval = 0 if concurentItems == 0 else ownerItems / concurentItems * 100.0
        
        data.fc = concurentItems
        
        self.data.append(data)
        
    def setPrice(self, pi):
      self.price = []
      
      for p in pi.values():
        self.price.append(p)
      
      self.price = sorted(self.price, key=lambda p: p.name)
      
      for i in range(0, len(self.price)):
        p = self.price[i]
        self.pcidx[p.id] = i
      
    def priceCells(self):
      result = []
      
      for p in self.price:
        result.append(p.name)
      
      return result

def loadData(params, server):
    orgs = server.Get('CommonOrgs', '', 'id')
    porg = server.Get("PotenzialOrg", "", "id")
    orgs.update(porg)
    
    price = server.Get('Price', '"own"=1', 'id')
    orgType = server.Get('OrgType', '', 'id')
    
    agentQuery = '"userid" in('
    for agent in params.agents:
      agentQuery += "'" + agent.id + "',";
        
    agentQuery = agentQuery[:-1] + ")"

    q = '"created" >= ToDate("{0}") and "created" <= ToDate("{1}")'.format(
      params.start.strftime("%d/%m/%Y 0:0:0"),
      params.finish.strftime("%d/%m/%Y 23:59:59"))
    q += ' and ' + agentQuery
        
    data = ReportData();
    data.setPrice(price)
    data.orgs = orgs
    data.orgType = orgType
    
    docs = server.Get('OrgDistrib', q)
    
    if docs != None:
      for d in docs:
        data.add(d)
            
    return data

class XLBuilderEx(XLBuilder):
  def adjustHeadCell(self, sheet, cell, row, column):
    cell.style.alignment.text_rotation = 90
    return column
  
  def makeHead(self, sheet, row, titles, wrap_text = False, startColumn = 0):
    XLBuilder.makeHead(self, sheet, row, titles, wrap_text, startColumn)
    sheet.row_dimensions[1].height = 95
        
def printOut(data, params):
    wb = Workbook(False, 'cp1251')
    sheet = wb.get_active_sheet()
    sheet.title = "Отчет"
    
    xlb = XLBuilderEx()
    
    row = 0
    
    head = ['Дата', 'Организация', 'Адрес', 'Доля полки', 'Категория ТТ', 
             'Фейсы конкурентов', 'Серийный номер']
    st = len(head)         
    pc = data.priceCells()
    
    if len(pc) > 0:
      head.extend(pc)
      
    head.append('комментарий по точке');
    xlb.makeHead(sheet, row, head, True)
    row += 1

    for item in data.data:
      xlb.makeCells(sheet, row, item.values())
      row += 1        
    
    cc = st
    for i in range(0,len(pc)):
      c = sheet.cell(row=row, column=cc)
      cn = get_column_letter(cc+1)
      dsz = len(data.data)
      c.value = "=COUNTIF({0}2:{0}{1},\"<>\"&\"\") / {2}".format(cn, dsz + 1, dsz )
      c.style.number_format._set_format_code('0%')

      cc += 1
    
    cc = 1
    for w in [12,20,20,7,7,7,20]:
      sheet.column_dimensions[get_column_letter(cc)].width = w
      cc += 1
    
    sheet.column_dimensions[get_column_letter(cc + len(pc))].width = 20
    
    return wb
    
        
def run(server):
    logging.basicConfig(format='%(module)s %(asctime)s.%(msecs)03d %(message)s', datefmt='%d.%m.%Y %H:%M:%S', stream=sys.stdout,level=logging.DEBUG)    
    logging.info('start')

    params = server.Params[0]
    logging.info("params " + str(params))

    data = loadData(params, server)
    wb = printOut(data, params)

    XLBuilder().workbookToObject(wb, "distrib_rep.xlsx", server)                
    logging.info('end')
    